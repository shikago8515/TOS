
# -*- coding: utf-8 -*-
"""
Jane 成品表生成模块
从 TMS工具_20260518_2100.pyw 提取的核心逻辑
创建时间: 2026-05-18
"""

import os
import openpyxl
import pandas as pd
from collections import Counter
from datetime import datetime
from openpyxl.cell.cell import MergedCell
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from typing import List, Dict, Any, Optional, Tuple, Set

# 导入工具模块
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from utils.file_utils import ensure_dir


class JaneModule:
    """Jane 成品表生成业务逻辑"""

    DESTINATION_CATEGORY_OVERRIDES = {
        'china': '中国单',
        'south africa': '欧美单',
    }

    REGION_TO_CATEGORY = {
        'EMEA': '欧美单',
        'NAM': '欧美单',
        'LAM': '欧美单',
        'GCA': '亚洲单',
        'APAC': '亚洲单',
    }

    CATEGORY_ORDER = ['欧美单', '亚洲单', '中国单', '其他']
    VALID_CATEGORIES = set(CATEGORY_ORDER)
    CATEGORY_ALIASES = {
        '南非单': '欧美单',
    }
    DEFAULT_CATEGORY = '其他'
    MIN_REQUIRED_FIELD_RATE = 0.9
    MIN_COUNTRY_MATCH_RATE = 0.8
    MAX_TEMPLATE_MISSING_RATE = 0.2
    REQUIRED_TMS_COLUMNS = [
        'Working Number',
        'PO Number',
        'Market PO Number',
        'PO Line Item #',
        'Company Code',
        'Article Number',
        'Customer Request Date (CRD)',
        'PODD',
        'Plant Code',
        'Customer Size Run',
        'Technical Notation',
        'Shipment Mode',
        'Technical Size',
        'Ordered Quantity',
        'Gps Customer Number',
    ]
    COLUMN_CANDIDATES = {
        'Customer Request Date (CRD)': ['Customer Request Date (CRD)', 'Customer Request Date', 'CRD'],
        'Shipment Mode': ['Shipment Mode', 'Shipment Method'],
        'Gps Customer Number': ['Gps Customer Number', 'GPS Customer Number', 'Gps Customer Num', 'GPS Customer Num'],
    }
    STRICT_VALUE_COLUMNS = {
        'Working Number',
        'PO Number',
        'PO Line Item #',
        'Article Number',
        'Technical Size',
        'Ordered Quantity',
        'Gps Customer Number',
    }
    
    def __init__(self):
        pass

    @staticmethod
    def _normalize_text(value: Any) -> str:
        if value is None or pd.isna(value):
            return ''
        text = str(value).strip()
        if text.lower() in {'nan', 'none', 'nat'}:
            return ''
        return text

    @staticmethod
    def _normalize_customer_no(value: Any) -> str:
        text = JaneModule._normalize_text(value)
        if not text:
            return ''
        if text.endswith('.0') and text[:-2].isdigit():
            text = text[:-2]
        if text.isdigit() and len(text) < 6:
            return text.zfill(6)
        return text

    @staticmethod
    def _normalize_size(value: Any) -> str:
        text = JaneModule._normalize_text(value)
        if text.endswith('.0') and text[:-2].isdigit():
            return text[:-2]
        return text

    @staticmethod
    def _format_identifier(value: Any, width: Optional[int] = None) -> str:
        """把 PO/客户号等编号写成文本，避免 Excel 吞掉前导 0。"""

        text = JaneModule._normalize_text(value)
        if not text:
            return ''
        if text.endswith('.0') and text[:-2].isdigit():
            text = text[:-2]
        if width and text.isdigit() and len(text) < width:
            text = text.zfill(width)
        return text

    @staticmethod
    def _customer_size_family(value: Any) -> str:
        text = JaneModule._normalize_size(value).upper()
        if not text:
            return 'BLANK'
        tall_sizes = {'2XST', 'XSTP', 'S/P', 'M/M', 'L/G', 'XLTG', '2XTG'}
        if text in tall_sizes or text.endswith(('T', 'TP', 'TG')):
            return 'TALL'
        if text.startswith('J/'):
            return 'J'
        if text.startswith('A'):
            return 'A'
        if '/' in text:
            return text.split('/', 1)[0]
        return 'STD'

    @staticmethod
    def _to_number(value: Any) -> float:
        if value is None or pd.isna(value):
            return 0
        try:
            return float(value)
        except (TypeError, ValueError):
            text = str(value).strip().replace(',', '')
            try:
                return float(text)
            except (TypeError, ValueError):
                return 0

    @staticmethod
    def _find_column(df: pd.DataFrame, candidates: List[str]) -> Optional[str]:
        columns_by_lower = {str(col).strip().lower(): col for col in df.columns}
        for candidate in candidates:
            if candidate in df.columns:
                return candidate
            found = columns_by_lower.get(candidate.strip().lower())
            if found is not None:
                return found
        return None

    def normalize_tms_columns(self, df: pd.DataFrame, logs: Optional[List[str]] = None) -> pd.DataFrame:
        """Map customer-exported aliases to the internal Jane column names."""

        normalized = df.copy()
        applied_aliases: List[str] = []

        for canonical_column in self.REQUIRED_TMS_COLUMNS:
            candidates = self.COLUMN_CANDIDATES.get(canonical_column, [canonical_column])
            source_column = self._find_column(normalized, candidates)
            if source_column is None or source_column == canonical_column:
                continue

            if canonical_column not in normalized.columns:
                normalized[canonical_column] = normalized[source_column]
            else:
                empty_mask = normalized[canonical_column].apply(
                    lambda value: not bool(self._normalize_text(value))
                )
                normalized.loc[empty_mask, canonical_column] = normalized.loc[empty_mask, source_column]

            applied_aliases.append(f"{source_column} -> {canonical_column}")

        if logs is not None and applied_aliases:
            logs.append("  已兼容 TMS 列名别名：" + "；".join(applied_aliases))

        return normalized

    @classmethod
    def _normalize_int(cls, value: Any) -> Optional[int]:
        text = cls._normalize_text(value)
        if not text:
            return None
        if text.endswith('.0') and text[:-2].isdigit():
            text = text[:-2]
        text = text.lstrip('0') or '0'
        try:
            return int(text)
        except (TypeError, ValueError):
            return None

    @staticmethod
    def _format_key(key: Tuple[int, int, str]) -> str:
        return f"PO={key[0]}, Line={key[1]}, Article={key[2]}"

    @classmethod
    def _sort_size_key(cls, value: Any) -> Tuple[int, int, str]:
        text = cls._normalize_size(value)
        if text.isdigit():
            return (0, int(text), text)
        if len(text) > 1 and text[0].upper() == 'A' and text[1:].isdigit():
            return (1, int(text[1:]), text)
        return (2, 0, text)

    @staticmethod
    def _safe_sheet_name(value: Any, used_names: Optional[Set[str]] = None) -> str:
        used_names = used_names if used_names is not None else set()
        text = str(value).strip() or 'Sheet'
        for char in ['\\', '/', '*', '?', ':', '[', ']']:
            text = text.replace(char, '-')
        text = text[:31] or 'Sheet'
        base = text
        suffix = 1
        while text in used_names:
            suffix_text = f"_{suffix}"
            text = f"{base[:31 - len(suffix_text)]}{suffix_text}"
            suffix += 1
        used_names.add(text)
        return text

    @staticmethod
    def _first_writable_col_right(
        ws: openpyxl.worksheet.worksheet.Worksheet,
        row_idx: int,
        start_col: int
    ) -> int:
        col_idx = start_col
        max_scan_col = max(ws.max_column + 20, start_col)
        while col_idx <= max_scan_col:
            cell = ws.cell(row=row_idx, column=col_idx)
            if not isinstance(cell, MergedCell):
                return col_idx

            moved = False
            for merged_range in ws.merged_cells.ranges:
                if (
                    merged_range.min_row <= row_idx <= merged_range.max_row
                    and merged_range.min_col <= col_idx <= merged_range.max_col
                ):
                    col_idx = merged_range.max_col + 1
                    moved = True
                    break
            if not moved:
                col_idx += 1

        return start_col

    @staticmethod
    def _quote_sheet_name(sheet_name: str) -> str:
        escaped = str(sheet_name).replace("'", "''")
        return f"'{escaped}'"

    @staticmethod
    def _header_text(value: Any) -> str:
        return " ".join(JaneModule._normalize_text(value).replace('\n', ' ').split()).upper()

    def _find_detail_formula_columns(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> Tuple[Optional[int], Optional[int]]:
        article_col = None
        po_qty_col = None
        for row_idx in range(1, min(ws.max_row, 5) + 1):
            for col_idx in range(1, ws.max_column + 1):
                header = self._header_text(ws.cell(row=row_idx, column=col_idx).value)
                if header == 'ARTICLE NUMBER':
                    article_col = col_idx
                elif header == 'PO QTY':
                    po_qty_col = col_idx
            if article_col and po_qty_col:
                break
        return article_col, po_qty_col

    def validate_tms_data(self, df: pd.DataFrame, logs: Optional[List[str]] = None) -> Dict[str, str]:
        """生成前校验 TMS 数据，不只检查列名，也检查关键列内容质量。"""

        if logs is None:
            logs = []

        resolved_columns: Dict[str, str] = {}
        missing_columns: List[str] = []
        for column in self.REQUIRED_TMS_COLUMNS:
            found = self._find_column(df, self.COLUMN_CANDIDATES.get(column, [column]))
            if found is None:
                missing_columns.append(column)
            else:
                resolved_columns[column] = found

        if missing_columns:
            raise ValueError(f"TMS Result Set 缺少必要列：{', '.join(missing_columns)}")

        if len(df) == 0:
            raise ValueError("TMS Result Set 没有数据行")

        hard_errors: List[str] = []
        for logical_name, column in resolved_columns.items():
            if logical_name not in self.STRICT_VALUE_COLUMNS:
                continue
            if logical_name == 'Ordered Quantity':
                continue
            non_empty_count = df[column].apply(lambda value: bool(self._normalize_text(value))).sum()
            fill_rate = non_empty_count / len(df)
            if fill_rate < self.MIN_REQUIRED_FIELD_RATE:
                hard_errors.append(f"{logical_name} 有效值占比 {fill_rate:.1%}")

        qty_col = resolved_columns['Ordered Quantity']
        qty_values = df[qty_col].apply(self._to_number)
        valid_qty_count = (qty_values > 0).sum()
        qty_rate = valid_qty_count / len(df)
        if qty_rate < self.MIN_REQUIRED_FIELD_RATE:
            hard_errors.append(f"Ordered Quantity 正数占比 {qty_rate:.1%}")

        po_col = resolved_columns['PO Number']
        line_col = resolved_columns['PO Line Item #']
        numeric_po_rate = df[po_col].apply(lambda value: self._normalize_int(value) is not None).sum() / len(df)
        numeric_line_rate = df[line_col].apply(lambda value: self._normalize_int(value) is not None).sum() / len(df)
        if numeric_po_rate < self.MIN_REQUIRED_FIELD_RATE:
            hard_errors.append(f"PO Number 数字格式占比 {numeric_po_rate:.1%}")
        if numeric_line_rate < self.MIN_REQUIRED_FIELD_RATE:
            hard_errors.append(f"PO Line Item # 数字格式占比 {numeric_line_rate:.1%}")

        if hard_errors:
            raise ValueError(
                "TMS 数据内容不符合 Jane 成品表生成要求，疑似上传了列名相同但业务数据不匹配的文件："
                + "；".join(hard_errors)
            )

        logs.append(
            "✅ TMS 数据校验通过：必要列存在，PO/Line/Ordered Quantity 等关键字段格式有效"
        )
        return resolved_columns

    def sort_tms_rows_for_generation(
        self,
        df: pd.DataFrame,
        country_lookup: Optional[Dict[str, Dict[str, str]]] = None,
        logs: Optional[List[str]] = None
    ) -> pd.DataFrame:
        """按业务指定顺序生成：Article Number -> Country/Region -> Working Number。"""

        country_lookup = country_lookup or {}
        destination_col = self._find_column(df, ['Country/Region', 'DESTINATION', 'Destination'])

        def normalized_sort_text(value: Any) -> str:
            return self._normalize_text(value).upper()

        def resolve_country(row: pd.Series) -> str:
            destination = self._normalize_text(row.get(destination_col)) if destination_col else ''
            if destination:
                return destination.upper()

            customer_no = self._normalize_customer_no(row.get('Gps Customer Number'))
            country_info = country_lookup.get(customer_no, {})
            return self._normalize_text(country_info.get('destination')).upper()

        source_rows = df.copy().reset_index(drop=True)
        sorted_indices = sorted(
            range(len(source_rows)),
            key=lambda row_idx: (
                normalized_sort_text(source_rows.iloc[row_idx].get('Article Number')),
                resolve_country(source_rows.iloc[row_idx]),
                normalized_sort_text(source_rows.iloc[row_idx].get('Working Number')),
                row_idx,
            ),
        )
        if logs is not None:
            logs.append("  已按 Article Number -> Country/Region -> Working Number 排序生成")
        return source_rows.iloc[sorted_indices].reset_index(drop=True)

    def get_category(self, destination: Any, region: Any) -> str:
        """根据国家/地区与 REGION 返回统计单别。"""

        dest_key = self._normalize_text(destination).lower()
        if dest_key in self.DESTINATION_CATEGORY_OVERRIDES:
            return self.DESTINATION_CATEGORY_OVERRIDES[dest_key]

        region_key = self._normalize_text(region).upper()
        return self.REGION_TO_CATEGORY.get(region_key, self.DEFAULT_CATEGORY)

    def read_country_table(
        self,
        country_path: str,
        logs: Optional[List[str]] = None
    ) -> Dict[str, Dict[str, str]]:
        """读取 country.xlsx，按 CST NO 构建客户号到统计单别的映射。

        分类优先级：
        1. country 表如有 CATEGORY/单别/分类列，直接使用该列；
        2. China 按国家名归入中国单，South Africa 按业务规则归入欧美单；
        3. 其他国家按 REGION 归类。
        """

        if logs is None:
            logs = []

        country_df = pd.read_excel(country_path, dtype=object)
        cst_col = self._find_column(country_df, ['CST NO', 'CSTNO', 'Gps Customer Number'])
        destination_col = self._find_column(country_df, ['DESTINATION', 'Destination'])
        region_col = self._find_column(country_df, ['REGION', 'Region'])
        category_col = self._find_column(country_df, ['CATEGORY', 'Category', '单别', '分类', '类别'])

        missing = [
            name for name, col in [
                ('CST NO', cst_col),
                ('DESTINATION', destination_col),
                ('REGION', region_col),
            ]
            if col is None
        ]
        if missing:
            raise ValueError(f"country.xlsx 缺少必要列：{', '.join(missing)}")

        country_lookup: Dict[str, Dict[str, str]] = {}
        default_rows: List[str] = []
        duplicate_customers: List[str] = []
        invalid_manual_categories: List[str] = []
        for _, row in country_df.iterrows():
            customer_no = self._normalize_customer_no(row[cst_col])
            if not customer_no:
                continue

            destination = self._normalize_text(row[destination_col])
            region = self._normalize_text(row[region_col])
            manual_category = self._normalize_text(row[category_col]) if category_col else ''
            manual_category = self.CATEGORY_ALIASES.get(manual_category, manual_category)
            if manual_category:
                if manual_category in self.VALID_CATEGORIES:
                    category = manual_category
                    source = 'country CATEGORY'
                else:
                    category = self.get_category(destination, region)
                    source = 'fallback'
                    invalid_manual_categories.append(
                        f"{customer_no}={manual_category}"
                    )
            else:
                category = self.get_category(destination, region)
                source = 'destination' if destination.lower() in self.DESTINATION_CATEGORY_OVERRIDES else 'region'

            if customer_no in country_lookup and country_lookup[customer_no]['category'] != category:
                duplicate_customers.append(
                    f"{customer_no}: {country_lookup[customer_no]['category']} -> {category}"
                )

            country_lookup[customer_no] = {
                'category': category,
                'region': region,
                'destination': destination,
                'source': source,
            }
            if category == self.DEFAULT_CATEGORY:
                default_rows.append(f"{customer_no}/{destination}/{region}")

        if invalid_manual_categories:
            logs.append(
                "  ⚠️ country.xlsx 中存在无效单别，已按国家/REGION 兜底："
                + "、".join(invalid_manual_categories[:10])
            )
        if duplicate_customers:
            logs.append(
                "  ⚠️ country.xlsx 中存在重复客户号且分类不同，已使用最后一条："
                + "、".join(duplicate_customers[:10])
            )
        if default_rows:
            logs.append(
                "  ⚠️ country.xlsx 中有客户号未能按规则分类，已归入“其他”："
                + "、".join(default_rows[:10])
            )
        unique_category_counter = Counter(info['category'] for info in country_lookup.values())
        logs.append(f"  Country 分类覆盖（客户号去重后）：{dict(unique_category_counter)}")

        return country_lookup

    def calculate_category_statistics(
        self,
        df: pd.DataFrame,
        country_lookup: Dict[str, Dict[str, str]],
        logs: Optional[List[str]] = None
    ) -> Dict[str, Dict[str, Dict[str, float]]]:
        """按 Working Number + Article Number + 单别统计 Ordered Quantity 总件数。"""

        if logs is None:
            logs = []

        working_col = self._find_column(df, ['Working Number'])
        article_col = self._find_column(df, ['Article Number'])
        qty_col = self._find_column(df, ['Ordered Quantity'])
        customer_col = self._find_column(df, ['Gps Customer Number', 'GPS Customer Number'])
        destination_col = self._find_column(df, ['Country/Region', 'DESTINATION', 'Destination'])
        region_col = self._find_column(df, ['REGION', 'Region', 'region'])

        missing = [
            name for name, col in [
                ('Working Number', working_col),
                ('Article Number', article_col),
                ('Ordered Quantity', qty_col),
                ('Gps Customer Number', customer_col),
            ]
            if col is None
        ]
        if missing:
            raise ValueError(f"TMS Result Set 缺少统计必要列：{', '.join(missing)}")

        stats: Dict[str, Dict[str, Dict[str, float]]] = {}
        matched_rows = 0
        processed_rows = 0
        unmatched_customers: Counter = Counter()
        blank_customer_rows = 0
        used_tms_fallback = False
        for _, row in df.iterrows():
            working_num = self._normalize_text(row[working_col])
            article = self._normalize_text(row[article_col])
            customer_no = self._normalize_customer_no(row[customer_col])
            qty = self._to_number(row[qty_col])
            if not working_num or not article or qty == 0:
                continue

            processed_rows += 1
            country_info = country_lookup.get(customer_no)
            if country_info:
                matched_rows += 1
                category = country_info.get('category', self.DEFAULT_CATEGORY)
            else:
                destination = row[destination_col] if destination_col else ''
                region = row[region_col] if region_col else ''
                if self._normalize_text(destination) or self._normalize_text(region):
                    category = self.get_category(destination, region)
                    used_tms_fallback = True
                else:
                    category = self.DEFAULT_CATEGORY
                if country_lookup and customer_no:
                    unmatched_customers[customer_no] += 1
                elif country_lookup:
                    blank_customer_rows += 1

            stats.setdefault(working_num, {}).setdefault(article, {})
            stats[working_num][article][category] = (
                stats[working_num][article].get(category, 0) + qty
            )

        if processed_rows > 0 and matched_rows == 0 and country_lookup:
            raise ValueError(
                "TMS 的 Gps Customer Number 与 country.xlsx 完全没有匹配，"
                "疑似上传了不对应的 country 文件或客户号格式不一致。"
            )
        if processed_rows > 0 and country_lookup:
            match_rate = matched_rows / processed_rows
            logs.append(f"  Country 匹配率：{matched_rows}/{processed_rows} 行（{match_rate:.1%}）")
            if match_rate < self.MIN_COUNTRY_MATCH_RATE:
                raise ValueError(
                    f"TMS 的 Gps Customer Number 与 country.xlsx 匹配率只有 {match_rate:.1%}，"
                    "低于 80%，疑似上传了不对应的 country 文件或 TMS 数据。"
                )
        if unmatched_customers:
            sample = "、".join(f"{key}({count}行)" for key, count in unmatched_customers.most_common(10))
            logs.append(
                f"  ⚠️ TMS 中有 {len(unmatched_customers)} 个客户号未在 country.xlsx 找到，"
                f"这些行暂归入“其他”：{sample}"
            )
        if blank_customer_rows:
            logs.append(f"  ⚠️ TMS 中有 {blank_customer_rows} 行 Gps Customer Number 为空，暂归入“其他”")
        if used_tms_fallback:
            logs.append("  单别统计兜底：已使用 TMS 的 Country/Region 或 region 字段补充分组")

        return stats

    def find_statistics_area(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        articles: Optional[List[str]] = None
    ) -> Dict[str, Any]:
        """定位成品表里的统计区：分类行、总数行、可选 Article 表头/清单。"""

        articles_set = {self._normalize_text(article) for article in (articles or []) if self._normalize_text(article)}
        category_rows: Dict[str, Tuple[int, int]] = {}
        total_candidates: List[Tuple[int, int]] = []
        article_list_cells: Dict[str, Tuple[int, int, Optional[int]]] = {}

        for row in ws.iter_rows():
            for cell in row:
                value = self._normalize_text(cell.value)
                if not value:
                    continue
                if value in self.CATEGORY_ORDER:
                    category_rows[value] = (cell.row, cell.column)
                elif value == '总数':
                    total_candidates.append((cell.row, cell.column))

        if not category_rows:
            return {
                'category_rows': {},
                'total_cell': total_candidates[0] if total_candidates else None,
                'article_header_cols': {},
                'article_list_cells': {},
            }

        min_category_row = min(row for row, _ in category_rows.values())
        max_category_row = max(row for row, _ in category_rows.values())
        min_category_col = min(col for _, col in category_rows.values())
        total_cell = None
        if total_candidates:
            total_cell = min(
                total_candidates,
                key=lambda row_col: (
                    0 if min_category_row <= row_col[0] <= max_category_row + 12 else 1,
                    abs(row_col[0] - max_category_row),
                    row_col[1],
                )
            )

        list_start_row = max(1, min_category_row - 4)
        list_end_row = min(ws.max_row, max_category_row + 12)
        list_start_col = min(ws.max_column, min_category_col + 2)
        for row_idx in range(list_start_row, list_end_row + 1):
            for col_idx in range(list_start_col, ws.max_column + 1):
                value = self._normalize_text(ws.cell(row=row_idx, column=col_idx).value)
                if value not in articles_set:
                    continue

                value_col = None
                for candidate_col in range(col_idx + 1, min(ws.max_column, col_idx + 10) + 1):
                    candidate_value = ws.cell(row=row_idx, column=candidate_col).value
                    if isinstance(candidate_value, (int, float)):
                        value_col = candidate_col
                        break
                article_list_cells[value] = (row_idx, col_idx, value_col)

        return {
            'category_rows': category_rows,
            'total_cell': total_cell,
            'article_header_cols': {},
            'article_list_cells': article_list_cells,
        }

    def fill_statistics_data(
        self,
        wb: openpyxl.Workbook,
        stats: Dict[str, Dict[str, Dict[str, float]]],
        logs: Optional[List[str]] = None
    ) -> None:
        """把按单别统计的总件数填入每个 Working Number 工作表。"""

        if logs is None:
            logs = []

        def has_formula(cell: openpyxl.cell.cell.Cell) -> bool:
            return isinstance(cell.value, str) and cell.value.startswith('=')

        for working_num, article_stats in stats.items():
            if working_num not in wb.sheetnames:
                logs.append(f"  ⚠️ 统计区跳过：未找到 {working_num} 表")
                continue

            ws = wb[working_num]
            articles = list(article_stats.keys())
            area = self.find_statistics_area(ws, articles)
            category_rows: Dict[str, Tuple[int, int]] = area['category_rows']
            total_cell: Optional[Tuple[int, int]] = area['total_cell']
            article_header_cols: Dict[str, int] = area['article_header_cols']
            article_list_cells: Dict[str, Tuple[int, int, Optional[int]]] = area['article_list_cells']

            if not category_rows:
                logs.append(f"  ⚠️ {working_num} 未找到 欧美单/亚洲单/中国单 统计区，跳过分类统计")
                continue

            category_totals: Dict[str, float] = {category: 0 for category in self.CATEGORY_ORDER}
            article_totals: Dict[str, float] = {}
            for article, category_values in article_stats.items():
                article_total = 0
                for category, qty in category_values.items():
                    category_totals[category] = category_totals.get(category, 0) + qty
                    article_total += qty
                article_totals[article] = article_total

            filled_cells = 0
            if article_header_cols:
                for category, row_col in category_rows.items():
                    row_idx, _ = row_col
                    for article, col_idx in article_header_cols.items():
                        qty = article_stats.get(article, {}).get(category, 0)
                        ws.cell(row=row_idx, column=col_idx, value=qty)
                        filled_cells += 1

                if total_cell:
                    total_row, _ = total_cell
                    for article, col_idx in article_header_cols.items():
                        ws.cell(row=total_row, column=col_idx, value=article_totals.get(article, 0))
                        filled_cells += 1
            else:
                category_value_cols: List[int] = []
                for category, qty in category_totals.items():
                    if category not in category_rows:
                        continue
                    row_idx, label_col = category_rows[category]
                    value_col = self._first_writable_col_right(ws, row_idx, label_col + 1)
                    target_cell = ws.cell(row=row_idx, column=value_col)
                    if not has_formula(target_cell):
                        target_cell.value = qty
                        filled_cells += 1
                    category_value_cols.append(value_col)

                if total_cell:
                    total_row, _ = total_cell
                    total_value_col = max(category_value_cols) if category_value_cols else total_cell[1] + 1
                    total_value_col = self._first_writable_col_right(ws, total_row, total_value_col)
                    target_cell = ws.cell(row=total_row, column=total_value_col)
                    if not has_formula(target_cell):
                        target_cell.value = sum(article_totals.values())
                        filled_cells += 1

            for article, total_qty in article_totals.items():
                list_cell = article_list_cells.get(article)
                if not list_cell:
                    continue
                row_idx, article_col, value_col = list_cell
                if value_col is None:
                    continue
                target_cell = ws.cell(row=row_idx, column=value_col)
                if has_formula(target_cell):
                    continue
                target_cell.value = total_qty
                filled_cells += 1

            logs.append(
                f"  ✅ {working_num} 统计区已填充：{len(articles)} 个 Article，{filled_cells} 个统计单元格"
            )

    def build_data_lookup(
        self,
        df: pd.DataFrame
    ) -> Tuple[Dict[Tuple[int, int, str], pd.Series], Dict[str, Set[Tuple[int, int, str]]]]:
        """构建 TMS 业务键索引：PO Number + PO Line Item # + Article Number。"""

        data_lookup: Dict[Tuple[int, int, str], pd.Series] = {}
        keys_by_working: Dict[str, Set[Tuple[int, int, str]]] = {}

        for _, row in df.iterrows():
            po_num = self._normalize_int(row.get('PO Number'))
            po_line = self._normalize_int(row.get('PO Line Item #'))
            article = self._normalize_text(row.get('Article Number'))
            working_num = self._normalize_text(row.get('Working Number'))
            if po_num is None or po_line is None or not article or not working_num:
                continue

            key = (po_num, po_line, article)
            data_lookup[key] = row
            keys_by_working.setdefault(working_num, set()).add(key)

        return data_lookup, keys_by_working

    def collect_template_keys(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> Set[Tuple[int, int, str]]:
        """从模板工作表读取已有的 PO/Line/Article 业务键，用于防止模板和 TMS 不对应。"""

        keys: Set[Tuple[int, int, str]] = set()
        current_po_num: Optional[int] = None
        current_article_num = ''

        for row_idx in range(1, ws.max_row + 1):
            po_line_item = ws.cell(row=row_idx, column=3).value
            po_line = self._normalize_int(po_line_item)
            if po_line is None:
                continue

            po_num_cell = ws.cell(row=row_idx, column=1).value
            article_num = ws.cell(row=row_idx, column=6).value

            normalized_po = self._normalize_int(po_num_cell)
            if normalized_po is not None:
                current_po_num = normalized_po

            normalized_article = self._normalize_text(article_num)
            if normalized_article:
                current_article_num = normalized_article

            if current_po_num is None or not current_article_num:
                continue

            keys.add((current_po_num, po_line, current_article_num))

        return keys

    def validate_template_alignment(
        self,
        wb: openpyxl.Workbook,
        keys_by_working: Dict[str, Set[Tuple[int, int, str]]],
        working_order: List[Any],
        logs: Optional[List[str]] = None
    ) -> None:
        """校验模板明细行与 TMS 数据是否对应，避免只因列名相同就生成错误文件。"""

        if logs is None:
            logs = []

        matched_sheet_count = 0
        for working_num in working_order:
            wn_str = self._normalize_text(working_num)
            if wn_str not in wb.sheetnames:
                logs.append(f"  ⚠️ 模板中未找到 Working Number 工作表：{wn_str}")
                continue

            tms_keys = keys_by_working.get(wn_str, set())
            template_keys = self.collect_template_keys(wb[wn_str])
            if not tms_keys:
                logs.append(f"  ⚠️ TMS 中 {wn_str} 没有可校验的 PO/Line/Article 业务键")
                continue
            if not template_keys:
                raise ValueError(f"模板工作表 {wn_str} 没有可校验的 PO/Line/Article 明细行")

            matched_keys = tms_keys & template_keys
            if not matched_keys:
                sample = "、".join(self._format_key(key) for key in list(tms_keys)[:5])
                raise ValueError(
                    f"模板工作表 {wn_str} 与 TMS 数据没有任何 PO/Line/Article 匹配，"
                    f"疑似模板和 TMS 不是同一批数据。示例：{sample}"
                )

            matched_sheet_count += 1
            missing_keys = tms_keys - template_keys
            missing_rate = len(missing_keys) / len(tms_keys)
            logs.append(
                f"  {wn_str} 模板匹配：{len(matched_keys)}/{len(tms_keys)} 条 TMS 明细行"
            )
            if missing_rate > self.MAX_TEMPLATE_MISSING_RATE:
                sample = "、".join(self._format_key(key) for key in list(missing_keys)[:5])
                raise ValueError(
                    f"模板工作表 {wn_str} 缺少 {len(missing_keys)} 条 TMS 明细行，"
                    f"缺失比例 {missing_rate:.1%}，已停止生成。示例：{sample}"
                )
            if missing_keys:
                sample = "、".join(self._format_key(key) for key in list(missing_keys)[:5])
                logs.append(
                    f"  ⚠️ {wn_str} 有 {len(missing_keys)} 条 TMS 明细行未在模板找到，将无法填充：{sample}"
                )

        if matched_sheet_count == 0:
            raise ValueError("模板中没有任何 Working Number 工作表能和 TMS 数据匹配")

    def collect_template_sizes(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> Set[str]:
        """读取模板里所有尺码区可填充的 sourcing size。"""

        all_sizes: Set[str] = set()
        for _, size_map in self.build_size_section_maps(ws):
            all_sizes.update(size_map.keys())

        return all_sizes

    def build_size_section_maps(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet
    ) -> List[Tuple[int, Dict[str, int]]]:
        """按每个 sourcing size 行构建尺码到列号的映射。"""

        reserved_headers = {
            'po number', 'market po number', 'po line item #', 'company code',
            'working number', 'article number', 'customer request date (crd)',
            'customer request date', 'crd', 'podd', 'plant code', 'customer size run',
            'technical notation', 'shipment mode', 'gps customer number',
            'sourcing size', 'customer size', 'po qty', 'each po total qty',
            'total'
        }

        sections: List[Tuple[int, Dict[str, int]]] = []
        for row_idx in range(1, ws.max_row + 1):
            title_col = None
            for col_idx in range(1, ws.max_column + 1):
                if self._normalize_text(ws.cell(row=row_idx, column=col_idx).value).lower() == 'sourcing size':
                    title_col = col_idx
                    break
            if title_col is None:
                continue

            size_map: Dict[str, int] = {}
            for col_idx in range(title_col + 1, ws.max_column + 1):
                raw_value = self._normalize_text(ws.cell(row=row_idx, column=col_idx).value)
                size_value = self._normalize_size(raw_value)
                if not size_value:
                    continue
                if raw_value.lower() in reserved_headers or size_value.lower() in reserved_headers:
                    continue
                size_map[size_value] = col_idx

            if size_map:
                sections.append((row_idx, size_map))

        return sections

    @staticmethod
    def _section_for_row(
        sections: List[Tuple[int, Dict[str, int]]],
        row_idx: int
    ) -> Optional[Dict[str, int]]:
        current_section = None
        for section_row, size_map in sections:
            if section_row <= row_idx:
                current_section = size_map
            else:
                break
        return current_section

    def validate_size_alignment(
        self,
        wb: openpyxl.Workbook,
        df: pd.DataFrame,
        working_order: List[Any],
        logs: Optional[List[str]] = None
    ) -> None:
        """校验 TMS Technical Size 是否能落到模板的 sourcing size 列。"""

        if logs is None:
            logs = []

        for working_num in working_order:
            wn_str = self._normalize_text(working_num)
            if wn_str not in wb.sheetnames:
                continue

            available_sizes = self.collect_template_sizes(wb[wn_str])
            if not available_sizes:
                raise ValueError(f"模板工作表 {wn_str} 未找到可识别的 sourcing size 尺码区")

            df_wn = df[df['Working Number'].apply(self._normalize_text) == wn_str]
            tms_sizes = {
                self._normalize_size(value)
                for value in df_wn['Technical Size'].tolist()
                if self._normalize_size(value)
            }
            missing_sizes = tms_sizes - available_sizes
            logs.append(
                f"  {wn_str} 尺码校验：TMS {len(tms_sizes)} 个尺码，模板 {len(available_sizes)} 个尺码"
            )
            if missing_sizes:
                sample = "、".join(sorted(missing_sizes)[:10])
                raise ValueError(
                    f"模板工作表 {wn_str} 缺少 TMS Technical Size 对应的 sourcing size 列：{sample}"
                )

    def create_auto_workbook(
        self,
        df: pd.DataFrame,
        working_order: List[Any],
        country_lookup: Optional[Dict[str, Dict[str, str]]] = None,
        logs: Optional[List[str]] = None
    ) -> openpyxl.Workbook:
        """根据 TMS 数据生成标准成品表。"""

        if logs is None:
            logs = []
        if country_lookup is None:
            country_lookup = {}

        wb = openpyxl.Workbook()
        default_sheet = wb.active
        wb.remove(default_sheet)

        used_sheet_names = {'Summary'}
        thin_black = Side(style='thin', color='000000')
        border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
        header_fill = PatternFill(start_color='5B9BD5', end_color='5B9BD5', fill_type='solid')
        size_fill = PatternFill(start_color='FFF200', end_color='FFF200', fill_type='solid')
        dest_fill = PatternFill(start_color='D9E2F3', end_color='D9E2F3', fill_type='solid')
        subtotal_fill = PatternFill(start_color='00B0F0', end_color='00B0F0', fill_type='solid')
        total_fill = PatternFill(start_color='D9EAF7', end_color='D9EAF7', fill_type='solid')
        category_fills = {
            '欧美单': 'B7DEE8',
            '亚洲单': 'CCC0DA',
            '中国单': 'FCD5B4',
            '其他': 'D9EAD3',
        }

        base_columns = [
            ('PO Number', 'PO Number'),
            ('Market PO Number', 'Market PO Number'),
            ('PO Line Item #', 'PO Line Item #'),
            ('Company Code', 'Company Code'),
            ('Working Number', 'Working Number'),
            ('Article Number', 'Article Number'),
            ('Customer Request Date (CRD)', 'Customer Request Date (CRD)'),
            ('PODD', 'PODD'),
            ('Plant Code', 'Plant Code'),
            ('Customer Size Run', 'Customer Size Run'),
            ('Technical Notation', 'Technical Notation'),
            ('Shipment Mode', 'Shipment Mode'),
            ('Gps Customer Number', 'Gps Customer Number'),
        ]
        identifier_widths = {
            'PO Number': 10,
            'Market PO Number': 10,
            'Gps Customer Number': 6,
        }
        destination_header_col = len(base_columns) + 1
        size_start_col = destination_header_col + 1
        destination_col = self._find_column(df, ['Country/Region', 'DESTINATION', 'Destination'])
        region_col = self._find_column(df, ['REGION', 'Region', 'region'])
        customer_size_col = self._find_column(df, ['Customer Size', 'Customer size'])

        def clean_excel_value(value: Any) -> Any:
            if value is None:
                return None
            try:
                if pd.isna(value):
                    return None
            except (TypeError, ValueError):
                pass
            return value

        def resolve_destination(row: pd.Series) -> str:
            customer_no = self._normalize_customer_no(row.get('Gps Customer Number'))
            tms_destination = self._normalize_text(row.get(destination_col)) if destination_col else ''
            if tms_destination:
                return tms_destination.upper()
            country_info = country_lookup.get(customer_no, {})
            country_destination = self._normalize_text(country_info.get('destination'))
            return country_destination.upper() if country_destination else ''

        def resolve_category(row: pd.Series) -> str:
            customer_no = self._normalize_customer_no(row.get('Gps Customer Number'))
            country_info = country_lookup.get(customer_no)
            if country_info:
                return country_info.get('category', self.DEFAULT_CATEGORY)
            destination = row.get(destination_col) if destination_col else ''
            region = row.get(region_col) if region_col else ''
            if self._normalize_text(destination) or self._normalize_text(region):
                return self.get_category(destination, region)
            return self.DEFAULT_CATEGORY

        def fill_for_category(category: str) -> PatternFill:
            color = category_fills.get(category, category_fills[self.DEFAULT_CATEGORY])
            return PatternFill(start_color=color, end_color=color, fill_type='solid')

        def size_family(row: pd.Series) -> str:
            tech_size = self._normalize_size(row.get('Technical Size')).upper()
            if tech_size.startswith('A') and tech_size[1:].isdigit():
                return 'A'
            if tech_size.isdigit():
                return 'NUM'
            return tech_size[:1] or 'OTHER'

        def build_groups(df_wn: pd.DataFrame) -> List[Dict[str, Any]]:
            """按尺码结构聚类，而不是按国家或客户号拆分。

            只要 Technical Size 对应的 Customer Size 不冲突，就认为属于同一段。
            例如德国/澳洲的 28-52 标准码会合并；Canada Tall 和 Japan A 码会拆开。
            """

            groups: List[Dict[str, Any]] = []
            for _, source_row in df_wn.iterrows():
                tech_size = self._normalize_size(source_row.get('Technical Size'))
                if not tech_size:
                    continue
                customer_size = (
                    self._normalize_size(source_row.get(customer_size_col))
                    if customer_size_col else ''
                ) or tech_size
                family = (
                    size_family(source_row),
                    self._customer_size_family(customer_size),
                )

                target_group = None
                for group in groups:
                    if group['family'] != family:
                        continue
                    existing_customer_size = group['sizes'].get(tech_size)
                    if existing_customer_size is not None and existing_customer_size != customer_size:
                        continue
                    target_group = group
                    break

                if target_group is None:
                    target_group = {
                        'family': family,
                        'first_index': len(groups),
                        'rows': [],
                        'sizes': {},
                    }
                    groups.append(target_group)

                target_group['rows'].append(source_row)
                target_group['sizes'].setdefault(tech_size, customer_size)

            for group in groups:
                group['size_order'] = sorted(group['sizes'].keys(), key=self._sort_size_key)

            family_order = {'NUM': 0, 'OTHER': 1, 'A': 2}
            customer_family_order = {'STD': 0, 'TALL': 1, 'A': 2, 'J': 3}
            return sorted(
                groups,
                key=lambda group: (
                    family_order.get(group['family'][0], 1),
                    customer_family_order.get(group['family'][1], 9),
                    group['first_index']
                )
            )

        def write_base_header(ws: openpyxl.worksheet.worksheet.Worksheet, max_size_slots: int) -> Tuple[int, int]:
            po_qty_col = size_start_col + max_size_slots
            each_po_total_col = po_qty_col + 1
            for col_idx, (header, _) in enumerate(base_columns, 1):
                cell = ws.cell(row=1, column=col_idx, value=header)
                cell.font = Font(bold=True, color='000000')
                cell.fill = header_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

            for row_idx, label in [(1, 'sourcing size'), (2, 'customer size')]:
                cell = ws.cell(row=row_idx, column=destination_header_col, value=label)
                cell.font = Font(bold=True)
                cell.fill = size_fill
                cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                cell.border = border

            for col_idx in range(size_start_col, each_po_total_col + 1):
                for row_idx in [1, 2]:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    cell.fill = size_fill
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.border = border
                    if row_idx == 1:
                        cell.font = Font(bold=True)
            ws.cell(row=1, column=po_qty_col, value='PO QTY')
            ws.cell(row=1, column=each_po_total_col, value='EACH PO\nTOTAL QTY')
            return po_qty_col, each_po_total_col

        def write_size_header(
            ws: openpyxl.worksheet.worksheet.Worksheet,
            row_idx: int,
            group: Dict[str, Any],
            po_qty_col: int,
            each_po_total_col: int
        ) -> None:
            source_row = row_idx
            customer_row = row_idx + 1
            ws.cell(row=source_row, column=destination_header_col, value='sourcing size')
            ws.cell(row=customer_row, column=destination_header_col, value='customer size')
            for target_row in [source_row, customer_row]:
                for col_idx in range(destination_header_col, each_po_total_col + 1):
                    cell = ws.cell(row=target_row, column=col_idx)
                    cell.fill = size_fill
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    cell.font = Font(bold=True)

            for offset, tech_size in enumerate(group['size_order']):
                col_idx = size_start_col + offset
                ws.cell(row=source_row, column=col_idx, value=tech_size)
                ws.cell(row=customer_row, column=col_idx, value=group['sizes'].get(tech_size, tech_size))
            ws.cell(row=source_row, column=po_qty_col, value='PO QTY')

        def write_subtotal_row(
            ws: openpyxl.worksheet.worksheet.Worksheet,
            row_idx: int,
            group: Dict[str, Any],
            po_qty_col: int,
            each_po_total_col: int,
            data_start_row: int,
            data_end_row: int
        ) -> None:
            for col_idx in range(destination_header_col, po_qty_col + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.fill = subtotal_fill
                cell.border = border
                cell.alignment = Alignment(horizontal='center', vertical='center')
            ws.cell(row=row_idx, column=destination_header_col, value='Sub. Total ')
            ws.cell(row=row_idx, column=destination_header_col).font = Font(bold=True)

            for col_idx in range(size_start_col, po_qty_col):
                col_letter = openpyxl.utils.get_column_letter(col_idx)
                formula = (
                    f"=SUM({col_letter}{data_start_row}:{col_letter}{data_end_row})"
                    if data_end_row >= data_start_row else 0
                )
                cell = ws.cell(row=row_idx, column=col_idx, value=formula)
                cell.number_format = '0'

            po_qty_letter = openpyxl.utils.get_column_letter(po_qty_col)
            po_qty_formula = (
                f"=SUM({po_qty_letter}{data_start_row}:{po_qty_letter}{data_end_row})"
                if data_end_row >= data_start_row else 0
            )
            ws.cell(
                row=row_idx,
                column=po_qty_col,
                value=po_qty_formula
            )
            ws.cell(row=row_idx, column=po_qty_col).font = Font(bold=True)
            ws.cell(row=row_idx, column=po_qty_col).number_format = '0'
            ws.cell(row=row_idx, column=each_po_total_col).border = border

        for working_num in working_order:
            wn_str = self._normalize_text(working_num)
            df_wn = df[df['Working Number'].apply(self._normalize_text) == wn_str]
            if len(df_wn) == 0:
                continue

            ws = wb.create_sheet(self._safe_sheet_name(wn_str, used_sheet_names))
            groups = build_groups(df_wn)
            max_size_slots = max((len(group['size_order']) for group in groups), default=1)
            po_qty_col, each_po_total_col = write_base_header(ws, max_size_slots)
            category_helper_col = each_po_total_col + 1
            po_helper_col = each_po_total_col + 2
            po_qty_letter = openpyxl.utils.get_column_letter(po_qty_col)
            each_po_total_letter = openpyxl.utils.get_column_letter(each_po_total_col)
            category_helper_letter = openpyxl.utils.get_column_letter(category_helper_col)
            po_helper_letter = openpyxl.utils.get_column_letter(po_helper_col)
            ws.cell(row=1, column=category_helper_col, value='_Jane_CATEGORY')
            ws.cell(row=1, column=po_helper_col, value='_Jane_PO')
            ws.column_dimensions[category_helper_letter].hidden = True
            ws.column_dimensions[po_helper_letter].hidden = True
            if groups:
                write_size_header(ws, 1, groups[0], po_qty_col, each_po_total_col)

            po_totals: Dict[int, float] = {}
            for _, row in df_wn.iterrows():
                po_num = self._normalize_int(row.get('PO Number'))
                if po_num is None:
                    continue
                po_totals[po_num] = po_totals.get(po_num, 0) + self._to_number(row.get('Ordered Quantity'))

            current_row = 3
            seen_po: Set[int] = set()
            grand_total = 0.0
            for group_index, group in enumerate(groups):
                if group_index > 0:
                    write_size_header(ws, current_row, group, po_qty_col, each_po_total_col)
                    current_row += 2

                size_to_col = {
                    size: size_start_col + offset
                    for offset, size in enumerate(group['size_order'])
                }
                group_data_start_row = current_row
                for source_row in group['rows']:
                    po_num = self._normalize_int(source_row.get('PO Number'))
                    po_seen = po_num is not None and po_num in seen_po
                    destination = resolve_destination(source_row)
                    category = resolve_category(source_row)
                    destination_fill = fill_for_category(category) if destination else dest_fill
                    for col_idx, (_, source_col) in enumerate(base_columns, 1):
                        value = clean_excel_value(source_row.get(source_col))
                        if col_idx in [1, 2] and po_seen:
                            value = None
                        elif source_col in identifier_widths:
                            value = self._format_identifier(value, identifier_widths[source_col])
                        cell = ws.cell(row=current_row, column=col_idx, value=value)
                        cell.border = border
                        cell.alignment = Alignment(vertical='center')
                        if source_col in identifier_widths:
                            cell.number_format = '@'
                        if col_idx in [7, 8]:
                            cell.number_format = 'yyyy/m/d'
                        if col_idx == 13 and destination:
                            cell.fill = destination_fill

                    dest_cell = ws.cell(row=current_row, column=destination_header_col, value=destination)
                    dest_cell.fill = destination_fill
                    dest_cell.border = border
                    dest_cell.alignment = Alignment(vertical='center')

                    for col_idx in range(size_start_col, each_po_total_col + 1):
                        ws.cell(row=current_row, column=col_idx).border = border
                        ws.cell(row=current_row, column=col_idx).alignment = Alignment(horizontal='right', vertical='center')

                    tech_size = self._normalize_size(source_row.get('Technical Size'))
                    qty = self._to_number(source_row.get('Ordered Quantity'))
                    row_total = qty
                    target_col = size_to_col.get(tech_size)
                    if target_col is not None:
                        ws.cell(row=current_row, column=target_col, value=qty)
                        ws.cell(row=current_row, column=target_col).number_format = '0'

                    first_size_letter = openpyxl.utils.get_column_letter(size_start_col)
                    last_size_letter = openpyxl.utils.get_column_letter(po_qty_col - 1)
                    ws.cell(
                        row=current_row,
                        column=po_qty_col,
                        value=f"=SUM({first_size_letter}{current_row}:{last_size_letter}{current_row})"
                    )
                    ws.cell(row=current_row, column=po_qty_col).number_format = '0'
                    ws.cell(row=current_row, column=category_helper_col, value=category)
                    ws.cell(row=current_row, column=po_helper_col, value=po_num)
                    if po_num is not None and po_num not in seen_po:
                        ws.cell(
                            row=current_row,
                            column=each_po_total_col,
                            value=(
                                f"=SUMIFS(${po_qty_letter}:${po_qty_letter},"
                                f"${po_helper_letter}:${po_helper_letter},"
                                f"${po_helper_letter}{current_row})"
                            )
                        )
                        ws.cell(row=current_row, column=each_po_total_col).font = Font(bold=True)
                        ws.cell(row=current_row, column=each_po_total_col).number_format = '0'
                        seen_po.add(po_num)

                    grand_total += row_total
                    current_row += 1

                write_subtotal_row(
                    ws,
                    current_row,
                    group,
                    po_qty_col,
                    each_po_total_col,
                    group_data_start_row,
                    current_row - 1
                )
                current_row += 1

            total_row = current_row
            for col_idx in range(1, each_po_total_col + 1):
                cell = ws.cell(row=total_row, column=col_idx)
                cell.border = border
                cell.fill = total_fill
                cell.alignment = Alignment(vertical='center')
            ws.cell(row=total_row, column=1, value='Total')
            ws.cell(row=total_row, column=1).font = Font(bold=True)
            ws.cell(
                row=total_row,
                column=each_po_total_col,
                value=f"=SUM({each_po_total_letter}3:{each_po_total_letter}{total_row - 1})"
            )
            ws.cell(row=total_row, column=each_po_total_col).font = Font(bold=True)
            ws.cell(row=total_row, column=each_po_total_col).number_format = '0'

            stats_row = total_row + 4
            for index, category in enumerate(self.CATEGORY_ORDER, stats_row):
                ws.cell(row=index, column=5, value=wn_str)
                ws.cell(row=index, column=6, value=category)
                ws.cell(row=index, column=6).fill = PatternFill(
                    start_color=category_fills.get(category, 'FFFFFF'),
                    end_color=category_fills.get(category, 'FFFFFF'),
                    fill_type='solid'
                )
                ws.cell(row=index, column=5).border = border
                ws.cell(row=index, column=6).border = border
                ws.cell(
                    row=index,
                    column=7,
                    value=(
                        f"=SUMIFS(${po_qty_letter}:${po_qty_letter},"
                        f"${category_helper_letter}:${category_helper_letter},F{index})"
                    )
                )
                ws.cell(row=index, column=7).border = border
                ws.cell(row=index, column=7).number_format = '0'
                ws.cell(row=index, column=6).font = Font(bold=True)
            stat_total_row = stats_row + len(self.CATEGORY_ORDER)
            ws.cell(row=stat_total_row, column=5, value='总数')
            ws.cell(
                row=stat_total_row,
                column=7,
                value=f"=SUM(G{stats_row}:G{stat_total_row - 1})"
            )
            for col_idx in range(5, 8):
                ws.cell(row=stat_total_row, column=col_idx).border = border
            ws.cell(row=stat_total_row, column=5).font = Font(bold=True)
            ws.cell(row=stat_total_row, column=7).font = Font(bold=True)
            ws.cell(row=stat_total_row, column=7).number_format = '0'

            article_totals: Dict[str, float] = {}
            for _, row in df_wn.iterrows():
                article = self._normalize_text(row.get('Article Number'))
                if not article:
                    continue
                article_totals[article] = article_totals.get(article, 0) + self._to_number(row.get('Ordered Quantity'))
            article_col = 12
            article_value_col = 13
            detail_article_letter = openpyxl.utils.get_column_letter(6)
            for row_offset, (article, qty) in enumerate(sorted(article_totals.items()), stats_row):
                ws.cell(row=row_offset, column=article_col, value=article)
                ws.cell(
                    row=row_offset,
                    column=article_value_col,
                    value=(
                        f"=SUMIFS(${po_qty_letter}:${po_qty_letter},"
                        f"${detail_article_letter}:${detail_article_letter},"
                        f"{openpyxl.utils.get_column_letter(article_col)}{row_offset})"
                    )
                )
                ws.cell(row=row_offset, column=article_value_col).number_format = '0'

            for row in ws.iter_rows(min_row=1, max_row=total_row, max_col=each_po_total_col):
                for cell in row:
                    cell.border = border

            for row_idx in range(stats_row, stat_total_row + 1):
                for col_idx in range(5, 8):
                    ws.cell(row=row_idx, column=col_idx).border = border
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical='center')
            article_table_end = stats_row + len(article_totals) - 1
            for row_idx in range(stats_row, article_table_end + 1):
                for col_idx in range(article_col, article_value_col + 1):
                    ws.cell(row=row_idx, column=col_idx).border = border
                    ws.cell(row=row_idx, column=col_idx).alignment = Alignment(vertical='center')

            for row in ws.iter_rows(min_row=1, max_row=ws.max_row, max_col=ws.max_column):
                for cell in row:
                    if cell.value is not None:
                        cell.alignment = Alignment(
                            horizontal=cell.alignment.horizontal,
                            vertical='center',
                            wrap_text=cell.alignment.wrap_text,
                        )

            widths = {
                'A': 14, 'B': 16, 'C': 12, 'D': 12, 'E': 18, 'F': 14,
                'G': 16, 'H': 12, 'I': 12, 'J': 14, 'K': 14, 'L': 14,
                'M': 18, 'N': 20,
            }
            for col_letter, width in widths.items():
                ws.column_dimensions[col_letter].width = width
            for col_idx in range(size_start_col, each_po_total_col + 1):
                ws.column_dimensions[openpyxl.utils.get_column_letter(col_idx)].width = 10
            ws.column_dimensions[openpyxl.utils.get_column_letter(each_po_total_col)].width = 13
            ws.freeze_panes = 'A3'
            ws.auto_filter.ref = f"A1:{openpyxl.utils.get_column_letter(each_po_total_col)}{max(total_row, 2)}"

            logs.append(
                f"  ✅ 自动生成 {wn_str}：{sum(len(group['rows']) for group in groups)} 行明细，"
                f"{len(groups)} 个尺码分组，{len(article_totals)} 个 Article"
            )

        self.jane_create_summary(wb, df, working_order)
        return wb
    
    def jane_populate_sheet(self, ws: openpyxl.worksheet.worksheet.Worksheet, 
                           df_wn: pd.DataFrame, working_num: str,
                           data_lookup: Optional[Dict[Tuple[int, int, str], pd.Series]] = None,
                           logs: Optional[List[str]] = None) -> None:
        """填充单个工作表"""
        
        if logs is None:
            logs = []
        
        size_sections = self.build_size_section_maps(ws)
        if len(size_sections) == 0:
            logs.append(f"    ⚠️ 未找到 'sourcing size' 标题行")
            return

        title_row = size_sections[0][0]
        all_template_sizes = sorted({size for _, size_map in size_sections for size in size_map.keys()})
        logs.append(f"    识别到 {len(size_sections)} 个 sourcing size 分组：{all_template_sizes}")

        # 首行通常包含全局的 EACH PO TOTAL QTY 列，后续尺码分组可能只写 PO QTY。
        total_qty_col = None
        for col_idx in range(1, ws.max_column + 1):
            header_val = ws.cell(row=title_row, column=col_idx).value
            if header_val == 'EACH PO TOTAL QTY':
                total_qty_col = col_idx
                break
        
        filled_count = 0
        current_po_num = None
        current_article_num = None
        
        for row_idx in range(1, ws.max_row + 1):
            po_line_item = ws.cell(row=row_idx, column=3).value
            
            if po_line_item is None or po_line_item == 'PO Line Item #':
                continue
            
            po_num_cell = ws.cell(row=row_idx, column=1).value
            article_num = ws.cell(row=row_idx, column=6).value
            
            normalized_po = self._normalize_int(po_num_cell)
            if normalized_po is not None:
                current_po_num = normalized_po
            
            if article_num:
                current_article_num = self._normalize_text(article_num)
            
            po_num = current_po_num
            article = current_article_num
            
            if po_num is None or article is None:
                continue
            
            # 使用数据查找字典，或者原来的遍历方式
            row_data = None
            if data_lookup is not None:
                po_line = self._normalize_int(po_line_item)
                if po_line is not None:
                    key = (po_num, po_line, self._normalize_text(article))
                    row_data = data_lookup.get(key)
            else:
                match_row = df_wn[
                    (df_wn['PO Number'] == po_num) &
                    (df_wn['PO Line Item #'] == int(po_line_item)) &
                    (df_wn['Article Number'] == article)
                ]
                if len(match_row) == 1:
                    row_data = match_row.iloc[0]
            
            if row_data is not None:
                tech_size = self._normalize_size(row_data['Technical Size'])
                qty = self._to_number(row_data['Ordered Quantity'])

                tech_size_to_col = self._section_for_row(size_sections, row_idx)
                if tech_size_to_col is None:
                    continue

                if tech_size in tech_size_to_col:
                    size_col = tech_size_to_col[tech_size]
                    ws.cell(row=row_idx, column=size_col, value=qty)
                    filled_count += 1
                
                if po_num_cell and total_qty_col:
                    ws.cell(row=row_idx, column=total_qty_col, value=self._to_number(row_data.get('Total Qty', qty)))
        
        logs.append(f"    成功填充 {filled_count} 个单元格")
    
    def jane_create_summary(self, wb: openpyxl.Workbook, df: pd.DataFrame, working_order: List[Any]) -> None:
        """创建 Summary 工作表"""
        
        if 'Summary' in wb.sheetnames:
            del wb['Summary']
        
        ws = wb.create_sheet(title='Summary', index=0)

        thin_black = Side(style='thin', color='000000')
        border = Border(left=thin_black, right=thin_black, top=thin_black, bottom=thin_black)
        header_fill = PatternFill(start_color='B7DEE8', end_color='B7DEE8', fill_type='solid')
        header_font = Font(name='Arial', size=10, bold=True)
        body_font = Font(name='Arial', size=10)
        qty_font = Font(name='Arial', size=10, bold=True)

        headers = ['Sketch', 'Working#', 'Article#', 'Color', 'Factory', 'QTY', 'Remark']
        header_row = 2
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=header_row, column=col_idx, value=header)
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)

        row_idx = header_row + 1
        total_qty_all = 0
        
        for wn in working_order:
            wn_str = self._normalize_text(wn)
            df_wn = df[df['Working Number'].apply(self._normalize_text) == wn_str]
            articles = df_wn['Article Number'].unique()
            group_start_row = row_idx
            detail_sheet_name = next(
                (sheet_name for sheet_name in wb.sheetnames if self._normalize_text(sheet_name) == wn_str),
                None
            )
            detail_formula_ref = None
            detail_article_letter = None
            detail_po_qty_letter = None
            if detail_sheet_name:
                detail_article_col, detail_po_qty_col = self._find_detail_formula_columns(wb[detail_sheet_name])
                if detail_article_col and detail_po_qty_col:
                    detail_formula_ref = self._quote_sheet_name(detail_sheet_name)
                    detail_article_letter = openpyxl.utils.get_column_letter(detail_article_col)
                    detail_po_qty_letter = openpyxl.utils.get_column_letter(detail_po_qty_col)

            for article in articles:
                df_article = df_wn[df_wn['Article Number'].apply(self._normalize_text) == self._normalize_text(article)]
                total_qty = df_article['Ordered Quantity'].apply(self._to_number).sum()
                total_qty_all += total_qty
                
                if len(df_article) > 0:
                    factory = df_article['Factory'].iloc[0] if 'Factory' in df_article.columns else ""
                    desc = df_article['Article Description'].iloc[0] if 'Article Description' in df_article.columns else ""
                else:
                    factory = ""
                    desc = ""

                if detail_formula_ref and detail_article_letter and detail_po_qty_letter:
                    qty_value = (
                        f"=SUMIFS({detail_formula_ref}!${detail_po_qty_letter}:${detail_po_qty_letter},"
                        f"{detail_formula_ref}!${detail_article_letter}:${detail_article_letter},"
                        f"C{row_idx})"
                    )
                else:
                    qty_value = total_qty

                row_values = [None, wn_str, article, desc, factory, qty_value, None]
                for col_idx, value in enumerate(row_values, 1):
                    cell = ws.cell(row=row_idx, column=col_idx, value=value)
                    cell.border = border
                    cell.font = qty_font if col_idx == 6 else body_font
                    cell.alignment = Alignment(
                        horizontal='center' if col_idx in [1, 2, 3, 5, 6, 7] else 'left',
                        vertical='center',
                        wrap_text=True,
                    )
                    if col_idx == 6:
                        cell.number_format = '0'

                row_idx += 1

            group_end_row = row_idx - 1
            if group_end_row >= group_start_row:
                if group_end_row > group_start_row:
                    ws.merge_cells(
                        start_row=group_start_row,
                        start_column=2,
                        end_row=group_end_row,
                        end_column=2,
                    )
                    ws.merge_cells(
                        start_row=group_start_row,
                        start_column=7,
                        end_row=group_end_row,
                        end_column=7,
                    )
                ws.cell(
                    row=group_start_row,
                    column=7,
                    value=f"=SUM(F{group_start_row}:F{group_end_row})"
                )
                for col_idx in [2, 7]:
                    cell = ws.cell(row=group_start_row, column=col_idx)
                    cell.border = border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
                    if col_idx == 7:
                        cell.font = qty_font
                        cell.number_format = '0'
        
        for col_idx in range(1, 8):
            cell = ws.cell(row=row_idx, column=col_idx)
            cell.border = border
            cell.font = qty_font
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=True)
        ws.cell(row=row_idx, column=2, value='TOTAL')
        ws.cell(
            row=row_idx,
            column=6,
            value=f"=SUM(F{header_row + 1}:F{row_idx - 1})" if row_idx > header_row + 1 else total_qty_all
        )
        ws.cell(row=row_idx, column=6).number_format = '0'

        widths = {'A': 35, 'B': 18, 'C': 18, 'D': 44, 'E': 18, 'F': 13, 'G': 22}
        for col_letter, width in widths.items():
            ws.column_dimensions[col_letter].width = width
        ws.row_dimensions[header_row].height = 40
        for data_row in range(header_row + 1, row_idx + 1):
            ws.row_dimensions[data_row].height = 40
        ws.freeze_panes = 'A3'
    
    def test_process(self, tms_path: str, country_path: str) -> Dict[str, Any]:
        """Run the same Jane generation path used by /process."""
        return self.process_reports(
            tms_path=tms_path,
            country_path=country_path,
            output_dir=os.path.dirname(tms_path)
        )

    def process_reports(self, tms_path: str, country_path: Optional[str] = None,
                        working_filters: Optional[List[str]] = None, output_dir: str = None) -> Dict[str, Any]:
        """主处理流程"""
        
        result = {
            'success': False,
            'message': '',
            'logs': [],
            'output_path': None,
            'working_count': 0
        }
        
        def log(msg: str):
            result['logs'].append(msg)
        
        try:
            log("=" * 80)
            log("开始生成 Jane 成品表")
            log("=" * 80)
            
            log(f"\n📖 正在读取客户文件...")
            df_result = pd.read_excel(tms_path, sheet_name='Result Set', dtype=object)
            log(f"✅ Result Set读取完成，共 {len(df_result)} 行数据")
            df_result = self.normalize_tms_columns(df_result, result['logs'])
            self.validate_tms_data(df_result, result['logs'])
            
            working_filters_set: set = set()
            if working_filters:
                for part in working_filters:
                    part = part.strip()
                    if part:
                        working_filters_set.add(part)
                log(f"筛选条件：{working_filters_set}")
            
            if working_filters_set:
                df_filtered = df_result[df_result['Working Number'].astype(str).isin(working_filters_set)]
            else:
                df_filtered = df_result

            if len(df_filtered) == 0:
                result['message'] = '没有匹配的数据！请检查筛选条件'
                return result

            if not country_path or not os.path.exists(country_path):
                result['message'] = 'Jane 生成需要上传 country.xlsx'
                log("ERROR: country.xlsx is required for Jane generation")
                return result

            category_stats: Dict[str, Dict[str, Dict[str, float]]] = {}
            country_lookup: Dict[str, Dict[str, str]] = {}
            log(f"\n📖 正在读取 country.xlsx 并计算单别统计...")
            country_lookup = self.read_country_table(country_path, result['logs'])
            df_filtered = self.sort_tms_rows_for_generation(df_filtered, country_lookup, result['logs'])
            working_order = list(df_filtered['Working Number'].unique())
            category_stats = self.calculate_category_statistics(df_filtered, country_lookup, result['logs'])
            log(f"✅ country.xlsx 读取完成：{len(country_lookup)} 个客户号映射")
            log(f"✅ 已计算 {len(category_stats)} 个 Working Number 的单别总件数")
            log(f"筛选后：{len(df_filtered)} 行数据，{len(working_order)} 个 Working Number")

            log(f"\n📐 正在根据客户文件自动生成分段成品表...")
            wb = self.create_auto_workbook(df_filtered, working_order, country_lookup, result['logs'])
            if category_stats:
                log(f"\n📊 正在填充自动生成表的单别统计区...")
                self.fill_statistics_data(wb, category_stats, result['logs'])
             
            log(f"\n💾 正在保存成品表...")
            
            if output_dir is None:
                output_dir = os.path.dirname(tms_path)
            
            ensure_dir(output_dir)
            
            default_name = (
                os.path.splitext(os.path.basename(tms_path))[0] +
                "_成品表_" +
                datetime.now().strftime('%Y%m%d_%H%M%S') +
                ".xlsx"
            )
            
            output_path = os.path.join(output_dir, default_name)
            wb.calculation.calcMode = 'auto'
            wb.calculation.fullCalcOnLoad = True
            wb.calculation.forceFullCalc = True
            wb.save(output_path)
            
            result['output_path'] = output_path
            result['working_count'] = len(working_order)
            
            log(f"\n{'='*80}")
            log(f"✅ 成品表生成完成！")
            log(f"{'='*80}")
            log(f"结果文件：{output_path}")
            log(f"包含 {len(wb.sheetnames)} 个工作表")
            
            result['success'] = True
            result['message'] = f"Jane 成品表生成完成，结果文件：{output_path}"
            
        except Exception as e:
            log(f"\n❌ 错误：{str(e)}")
            import traceback
            log(traceback.format_exc())
            result['message'] = f'处理出错：{str(e)}'
        
        return result

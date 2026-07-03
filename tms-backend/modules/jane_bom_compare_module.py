# -*- coding: utf-8 -*-
"""
Jane PRODUCTION 核对模块。

上传 T1 PRODUCTION 和 Jane-BOM 汇总表后，按 Style ID + Production Lot ID
分组检查生产表一致性，并用 BOM 汇总的 Articles + Factory 材料集合核对多料、缺料。
"""

import copy
import os
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple, Union

import openpyxl
from openpyxl.styles import PatternFill
from openpyxl.utils import get_column_letter

from utils.file_utils import ensure_dir


@dataclass(frozen=True)
class BomMaterial:
    """BOM 或 BOM 汇总中一条可核对的 MAIN COMPONENT 面料记录。"""

    article: str
    factory: str
    material_ref: str
    group_code_supplier: str
    color: str
    working: str
    season: str
    source_file: str
    source_row: int


@dataclass
class ProductionGroup:
    """生产表中同 Style ID + Production Lot ID 的行集合。"""

    key: Tuple[str, str]
    rows: List[int]


class JaneBomCompareModule:
    """Jane PRODUCTION 核对业务逻辑。"""

    MAIN_COMPONENT_SECTION = "MAIN COMPONENT"
    NO_FILL = PatternFill(fill_type=None)
    RED_FILL = PatternFill("solid", fgColor="FFFFC7CE")
    MISSING_ROW_FILL = PatternFill("solid", fgColor="FFFFF2CC")
    RED_FONT_COLOR = "FFFF0000"
    REQUIRED_PRODUCTION_COLUMNS = [
        "Style ID",
        "Production Lot ID",
        "Production Quantity (in Units)",
        "Units",
        "Production Date",
        "Recording Facility ID",
        "Input Material UID/ID",
        "Input Lot Quantity Used (In Units)",
        "Seller Facility ID",
    ]
    REQUIRED_BOM_COLUMNS = [
        "Part Group #",
        "Material Reference #",
        "Group Code Supplier",
        "Color",
    ]
    REQUIRED_BOM_SUMMARY_COLUMNS = [
        "Factory",
        "Articles",
        "Material Reference #",
        "Group Code Supplier",
        "Color",
    ]
    DIAGNOSTIC_HEADERS = [
        ("check result", "Check Result"),
        ("mismatch source", "Mismatch Source"),
        ("check detail", "Check Detail"),
        ("bom source file", "BOM Source File"),
        ("bom source row", "BOM Source Row"),
    ]
    CONSISTENCY_COLUMNS = [
        ("production quantity (in units)", "Production Quantity (in Units)"),
        ("units", "Units"),
        ("production date", "Production Date"),
        ("recording facility id", "Recording Facility ID"),
    ]

    @staticmethod
    def _normalize_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        text = str(value).strip()
        if text.lower() in {"none", "nan", "nat"}:
            return ""
        return text

    @classmethod
    def _normalize_key(cls, value: Any) -> str:
        return cls._normalize_text(value).upper()

    @classmethod
    def _find_header_row(
        cls,
        ws,
        required_columns: Sequence[str],
        max_scan_rows: int = 30,
    ) -> Tuple[int, Dict[str, int]]:
        required = [column.lower() for column in required_columns]
        for row_index in range(1, min(ws.max_row, max_scan_rows) + 1):
            columns: Dict[str, int] = {}
            for column_index in range(1, ws.max_column + 1):
                value = cls._normalize_text(
                    ws.cell(row=row_index, column=column_index).value,
                ).lower()
                if value:
                    columns[value] = column_index
            if all(column in columns for column in required):
                return row_index, columns
        raise ValueError(f"未找到表头：{', '.join(required_columns)}")

    @classmethod
    def _parse_factory_code(cls, value: Any) -> str:
        text = cls._normalize_text(value)
        if "|" in text:
            return cls._normalize_text(text.split("|", 1)[0])
        return text

    @classmethod
    def _parse_article_from_group(cls, value: Any) -> str:
        parts = [part.strip() for part in cls._normalize_text(value).split("|")]
        if len(parts) >= 2:
            return parts[1]
        return ""

    def _read_optional_cell_text(self, ws, row_index: int, columns: Dict[str, int], key: str) -> str:
        column_index = columns.get(key)
        if not column_index:
            return ""
        return self._normalize_text(ws.cell(row=row_index, column=column_index).value)

    def _read_bom_materials(self, bom_path: str) -> List[BomMaterial]:
        wb = openpyxl.load_workbook(bom_path, data_only=False)
        ws = wb.active

        working = self._normalize_text(ws["B1"].value)
        season = self._normalize_text(ws["B3"].value)
        factory = self._parse_factory_code(ws["B5"].value)
        if not factory:
            raise ValueError(f"{os.path.basename(bom_path)} 缺少 Factory")

        header_row, header_columns = self._find_header_row(
            ws,
            self.REQUIRED_BOM_COLUMNS,
        )
        article_columns: List[Tuple[int, str]] = []
        for column_index in range(1, ws.max_column + 1):
            header = self._normalize_text(ws.cell(row=header_row, column=column_index).value)
            if header.lower() != "color":
                continue
            article = self._parse_article_from_group(
                ws.cell(row=header_row - 1, column=column_index).value,
            )
            if article:
                article_columns.append((column_index, article))

        if not article_columns:
            raise ValueError(f"{os.path.basename(bom_path)} 未识别到 Article/Color 列")

        rows: List[BomMaterial] = []
        seen: Set[Tuple[str, str, str]] = set()
        current_section = ""
        main_component_started = False
        for row_index in range(header_row + 1, ws.max_row + 1):
            part_name = self._normalize_text(ws.cell(row=row_index, column=2).value)
            part_group = self._normalize_text(
                ws.cell(row=row_index, column=header_columns["part group #"]).value,
            )
            material_ref = self._normalize_text(
                ws.cell(row=row_index, column=header_columns["material reference #"]).value,
            )

            # BOM 的分段行通常只有 Part Name，没有 Part Group / Material Reference。
            if part_name and not part_group and not material_ref:
                next_section = part_name.upper()
                if main_component_started and next_section != self.MAIN_COMPONENT_SECTION:
                    break
                current_section = next_section
                if current_section == self.MAIN_COMPONENT_SECTION:
                    main_component_started = True
                continue

            if current_section != self.MAIN_COMPONENT_SECTION or not material_ref:
                continue

            group_code_supplier = self._normalize_text(
                ws.cell(row=row_index, column=header_columns["group code supplier"]).value,
            )
            for color_column, article in article_columns:
                color = self._normalize_text(ws.cell(row=row_index, column=color_column).value)
                if not color:
                    continue

                key = (
                    self._normalize_key(article),
                    self._normalize_key(factory),
                    self._normalize_key(material_ref),
                )
                if key in seen:
                    continue
                seen.add(key)
                rows.append(
                    BomMaterial(
                        article=article,
                        factory=factory,
                        material_ref=material_ref,
                        group_code_supplier=group_code_supplier,
                        color=color,
                        working=working,
                        season=season,
                        source_file=os.path.basename(bom_path),
                        source_row=row_index,
                    ),
                )

        return rows

    def _read_bom_summary_materials(self, bom_summary_path: str) -> List[BomMaterial]:
        wb = openpyxl.load_workbook(bom_summary_path, data_only=True)
        ws = wb.active
        header_row, header_columns = self._find_header_row(
            ws,
            self.REQUIRED_BOM_SUMMARY_COLUMNS,
            max_scan_rows=10,
        )

        rows: List[BomMaterial] = []
        seen: Set[Tuple[str, str, str]] = set()
        for row_index in range(header_row + 1, ws.max_row + 1):
            article = self._normalize_text(
                ws.cell(row=row_index, column=header_columns["articles"]).value,
            )
            factory = self._parse_factory_code(
                ws.cell(row=row_index, column=header_columns["factory"]).value,
            )
            material_ref = self._normalize_text(
                ws.cell(row=row_index, column=header_columns["material reference #"]).value,
            )
            color = self._normalize_text(
                ws.cell(row=row_index, column=header_columns["color"]).value,
            )
            if not article or not factory or not material_ref:
                continue

            key = (
                self._normalize_key(article),
                self._normalize_key(factory),
                self._normalize_key(material_ref),
            )
            if key in seen:
                continue
            seen.add(key)
            rows.append(
                BomMaterial(
                    article=article,
                    factory=factory,
                    material_ref=material_ref,
                    group_code_supplier=self._read_optional_cell_text(
                        ws,
                        row_index,
                        header_columns,
                        "group code supplier",
                    ),
                    color=color,
                    working=self._read_optional_cell_text(ws, row_index, header_columns, "working #"),
                    season=self._read_optional_cell_text(ws, row_index, header_columns, "season"),
                    source_file=os.path.basename(bom_summary_path),
                    source_row=row_index,
                ),
            )

        return rows

    def _copy_row(
        self,
        ws,
        source_row: int,
        target_row: int,
        max_column: int,
        copy_values: bool = True,
    ) -> None:
        for column_index in range(1, max_column + 1):
            source = ws.cell(row=source_row, column=column_index)
            target = ws.cell(row=target_row, column=column_index)
            target.value = source.value if copy_values else None
            if source.has_style:
                target._style = copy.copy(source._style)
                if not copy_values:
                    self._clear_fill(target)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy.copy(source.alignment)
            if source.protection:
                target.protection = copy.copy(source.protection)

    def _mark_red(self, cell) -> None:
        cell.fill = copy.copy(self.RED_FILL)

    def _clear_fill(self, cell) -> None:
        cell.fill = copy.copy(self.NO_FILL)

    def _clear_data_fills(self, ws, header_row: int, max_column: int) -> None:
        for row_index in range(header_row + 1, ws.max_row + 1):
            for column_index in range(1, max_column + 1):
                self._clear_fill(ws.cell(row=row_index, column=column_index))

    def _copy_header_style(self, ws, header_row: int, source_col: int, target_col: int) -> None:
        source = ws.cell(row=header_row, column=source_col)
        target = ws.cell(row=header_row, column=target_col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.protection:
            target.protection = copy.copy(source.protection)
        ws.column_dimensions[target.column_letter].width = ws.column_dimensions[source.column_letter].width

    def _insert_column_after(
        self,
        ws,
        header_row: int,
        columns: Dict[str, int],
        source_key: str,
        new_key: str,
        header: str,
    ) -> int:
        source_col = columns[source_key]
        insert_at = source_col + 1
        ws.insert_cols(insert_at)
        for column_name, column_index in list(columns.items()):
            if column_index >= insert_at:
                columns[column_name] = column_index + 1
        columns[new_key] = insert_at
        ws.cell(row=header_row, column=insert_at, value=header)
        self._copy_header_style(ws, header_row, source_col, insert_at)
        return insert_at

    def _append_column(
        self,
        ws,
        header_row: int,
        columns: Dict[str, int],
        new_key: str,
        header: str,
    ) -> int:
        append_at = ws.max_column + 1
        columns[new_key] = append_at
        ws.cell(row=header_row, column=append_at, value=header)
        self._copy_header_style(ws, header_row, max(1, append_at - 1), append_at)
        return append_at

    def _prepare_output_columns(self, ws, header_row: int, columns: Dict[str, int]) -> None:
        self._insert_column_after(
            ws,
            header_row,
            columns,
            "input material uid/id",
            "correct material reference #",
            "Correct Material Reference # (BOM)",
        )
        self._insert_column_after(
            ws,
            header_row,
            columns,
            "seller facility id",
            "correct group code supplier",
            "Correct Group Code Supplier (BOM)",
        )
        self._insert_column_after(
            ws,
            header_row,
            columns,
            "correct group code supplier",
            "料率",
            "料率",
        )
        self._insert_column_after(
            ws,
            header_row,
            columns,
            "料率",
            "颜色",
            "颜色",
        )
        for key, header in self.DIAGNOSTIC_HEADERS:
            self._append_column(ws, header_row, columns, key, header)

    def _set_row_diagnostic(
        self,
        ws,
        row_index: int,
        columns: Dict[str, int],
        result: str,
        source: str,
        detail: str,
        material: Optional[BomMaterial] = None,
    ) -> None:
        result_cell = ws.cell(row=row_index, column=columns["check result"])
        source_cell = ws.cell(row=row_index, column=columns["mismatch source"])
        if not self._normalize_text(result_cell.value):
            result_cell.value = result
        if not self._normalize_text(source_cell.value):
            source_cell.value = source
        detail_cell = ws.cell(row=row_index, column=columns["check detail"])
        existing_detail = self._normalize_text(detail_cell.value)
        detail_cell.value = f"{existing_detail}；{detail}" if existing_detail else detail
        if material:
            file_cell = ws.cell(row=row_index, column=columns["bom source file"])
            row_cell = ws.cell(row=row_index, column=columns["bom source row"])
            if not self._normalize_text(file_cell.value):
                file_cell.value = material.source_file
            if not self._normalize_text(row_cell.value):
                row_cell.value = material.source_row

    def _join_unique(self, values: Sequence[Any]) -> str:
        unique_values: List[str] = []
        seen: Set[str] = set()
        for value in values:
            text = self._normalize_text(value)
            key = self._normalize_key(text)
            if not text or key in seen:
                continue
            seen.add(key)
            unique_values.append(text)
        return " / ".join(unique_values)

    def _set_row_strike(self, ws, row_index: int, max_column: int) -> None:
        for column_index in range(1, max_column + 1):
            cell = ws.cell(row=row_index, column=column_index)
            font = copy.copy(cell.font)
            font.strike = True
            font.color = self.RED_FONT_COLOR
            cell.font = font

    def _set_row_fill(self, ws, row_index: int, max_column: int, fill: PatternFill) -> None:
        for column_index in range(1, max_column + 1):
            ws.cell(row=row_index, column=column_index).fill = copy.copy(fill)

    def _build_expected_material_map(
        self,
        bom_paths: Sequence[str],
        is_summary_input: bool,
        logs: List[str],
    ) -> Tuple[Dict[Tuple[str, str], List[BomMaterial]], Dict[Tuple[str, str], str], int]:
        expected_by_key: Dict[Tuple[str, str], List[BomMaterial]] = {}
        color_by_style_material: Dict[Tuple[str, str], str] = {}
        bom_material_row_count = 0
        for bom_path in bom_paths:
            bom_materials = (
                self._read_bom_summary_materials(bom_path)
                if is_summary_input
                else self._read_bom_materials(bom_path)
            )
            source_row_count = len({material.source_row for material in bom_materials})
            bom_material_row_count += source_row_count
            if is_summary_input:
                logs.append(
                    f"{os.path.basename(bom_path)} 读取 BOM汇总："
                    f"{source_row_count} 个有效行，{len(bom_materials)} 条 Article/Factory 映射",
                )
            else:
                logs.append(
                    f"{os.path.basename(bom_path)} 读取 MAIN COMPONENT："
                    f"{source_row_count} 个有效源行，{len(bom_materials)} 条 Article 映射",
                )
            for material in bom_materials:
                key = (self._normalize_key(material.article), self._normalize_key(material.factory))
                expected_by_key.setdefault(key, []).append(material)
                color_key = (
                    self._normalize_key(material.article),
                    self._normalize_key(material.material_ref),
                )
                if material.color and color_key not in color_by_style_material:
                    color_by_style_material[color_key] = material.color
        return expected_by_key, color_by_style_material, bom_material_row_count

    def _build_production_groups(
        self,
        ws,
        header_row: int,
        columns: Dict[str, int],
    ) -> Dict[Tuple[str, str], ProductionGroup]:
        groups: Dict[Tuple[str, str], ProductionGroup] = {}
        style_col = columns["style id"]
        lot_col = columns["production lot id"]
        material_col = columns["input material uid/id"]
        factory_col = columns["recording facility id"]
        for row_index in range(header_row + 1, ws.max_row + 1):
            style = self._normalize_key(ws.cell(row=row_index, column=style_col).value)
            lot = self._normalize_key(ws.cell(row=row_index, column=lot_col).value)
            material = self._normalize_key(ws.cell(row=row_index, column=material_col).value)
            factory = self._normalize_key(ws.cell(row=row_index, column=factory_col).value)
            if not style and not lot and not material and not factory:
                continue
            key = (style, lot)
            groups.setdefault(key, ProductionGroup(key=key, rows=[])).rows.append(row_index)
        return groups

    def _check_group_consistency(
        self,
        ws,
        group: ProductionGroup,
        columns: Dict[str, int],
    ) -> Tuple[int, int]:
        inconsistent_group = 0
        red_cell_count = 0
        group_has_inconsistency = False
        for column_key, header in self.CONSISTENCY_COLUMNS:
            values: Dict[str, Any] = {}
            for row_index in group.rows:
                cell_value = ws.cell(row=row_index, column=columns[column_key]).value
                normalized = self._normalize_key(cell_value)
                values.setdefault(normalized, cell_value)
            non_empty_values = [value for value in values if value]
            if len(set(non_empty_values)) <= 1:
                continue

            group_has_inconsistency = True
            value_text = self._join_unique(values.values())
            for row_index in group.rows:
                self._mark_red(ws.cell(row=row_index, column=columns[column_key]))
                self._set_row_diagnostic(
                    ws,
                    row_index,
                    columns,
                    "需核对",
                    "Production同组字段不一致",
                    f"同 Style ID + Production Lot ID 下 {header} 不一致：{value_text}",
                )
                red_cell_count += 1

        if group_has_inconsistency:
            inconsistent_group = 1
        return inconsistent_group, red_cell_count

    def _expected_materials_for_group(
        self,
        ws,
        group: ProductionGroup,
        columns: Dict[str, int],
        expected_by_key: Dict[Tuple[str, str], List[BomMaterial]],
    ) -> List[BomMaterial]:
        style_key = self._normalize_key(ws.cell(row=group.rows[0], column=columns["style id"]).value)
        factories: List[str] = []
        seen_factories: Set[str] = set()
        for row_index in group.rows:
            factory = self._normalize_key(
                self._parse_factory_code(
                    ws.cell(row=row_index, column=columns["recording facility id"]).value,
                ),
            )
            if not factory or factory in seen_factories:
                continue
            seen_factories.add(factory)
            factories.append(factory)

        materials: List[BomMaterial] = []
        seen_materials: Set[str] = set()
        for factory in factories:
            for material in expected_by_key.get((style_key, factory), []):
                material_key = self._normalize_key(material.material_ref)
                if material_key in seen_materials:
                    continue
                seen_materials.add(material_key)
                materials.append(material)
        return materials

    def _mark_no_bom_group(
        self,
        ws,
        group: ProductionGroup,
        columns: Dict[str, int],
    ) -> int:
        red_cell_count = 0
        for row_index in group.rows:
            for column_key in ["style id", "recording facility id"]:
                self._mark_red(ws.cell(row=row_index, column=columns[column_key]))
                red_cell_count += 1
            self._set_row_diagnostic(
                ws,
                row_index,
                columns,
                "需核对",
                "Production存在，BOM汇总缺失",
                "Production 有该 Style ID/Recording Facility ID，但 BOM汇总 未找到对应 Articles/Factory。",
            )
        return red_cell_count

    def _find_extra_and_missing_materials(
        self,
        ws,
        group: ProductionGroup,
        columns: Dict[str, int],
        expected_materials: Sequence[BomMaterial],
    ) -> Tuple[Dict[str, List[int]], List[BomMaterial]]:
        expected_by_material = {
            self._normalize_key(material.material_ref): material
            for material in expected_materials
            if self._normalize_key(material.material_ref)
        }
        production_rows_by_material: Dict[str, List[int]] = {}
        for row_index in group.rows:
            material_key = self._normalize_key(
                ws.cell(row=row_index, column=columns["input material uid/id"]).value,
            )
            if material_key:
                production_rows_by_material.setdefault(material_key, []).append(row_index)

        extra_rows = {
            material_key: rows
            for material_key, rows in production_rows_by_material.items()
            if material_key not in expected_by_material
        }
        missing_materials = [
            material
            for material_key, material in expected_by_material.items()
            if material_key not in production_rows_by_material
        ]
        return extra_rows, missing_materials

    def _group_materials_by_ref(
        self,
        expected_materials: Sequence[BomMaterial],
    ) -> Dict[str, List[BomMaterial]]:
        materials_by_ref: Dict[str, List[BomMaterial]] = {}
        for material in expected_materials:
            material_key = self._normalize_key(material.material_ref)
            if material_key:
                materials_by_ref.setdefault(material_key, []).append(material)
        return materials_by_ref

    def _check_group_material_and_supplier_values(
        self,
        ws,
        group: ProductionGroup,
        columns: Dict[str, int],
        expected_materials: Sequence[BomMaterial],
    ) -> int:
        materials_by_ref = self._group_materials_by_ref(expected_materials)
        expected_material_text = self._join_unique(
            [material.material_ref for material in expected_materials],
        )
        mismatch_cell_count = 0

        for row_index in group.rows:
            material_key = self._normalize_key(
                ws.cell(row=row_index, column=columns["input material uid/id"]).value,
            )
            seller_key = self._normalize_key(
                ws.cell(row=row_index, column=columns["seller facility id"]).value,
            )

            supplier_materials = materials_by_ref.get(material_key, [])
            if not supplier_materials:
                supplier_materials = list(expected_materials)
                self._mark_red(ws.cell(row=row_index, column=columns["input material uid/id"]))
                ws.cell(
                    row=row_index,
                    column=columns["correct material reference #"],
                    value=expected_material_text,
                )
                self._set_row_diagnostic(
                    ws,
                    row_index,
                    columns,
                    "需核对",
                    "值不一致：以 BOM 为准",
                    "Input Material UID/ID 不在 BOM MAIN COMPONENT Material Reference # 中。",
                    supplier_materials[0] if supplier_materials else None,
                )
                mismatch_cell_count += 1

            expected_supplier_keys = {
                self._normalize_key(material.group_code_supplier)
                for material in supplier_materials
                if self._normalize_key(material.group_code_supplier)
            }
            if not expected_supplier_keys:
                continue

            if not seller_key or seller_key not in expected_supplier_keys:
                self._mark_red(ws.cell(row=row_index, column=columns["seller facility id"]))
                ws.cell(
                    row=row_index,
                    column=columns["correct group code supplier"],
                    value=self._join_unique(
                        [material.group_code_supplier for material in supplier_materials],
                    ),
                )
                self._set_row_diagnostic(
                    ws,
                    row_index,
                    columns,
                    "需核对",
                    "值不一致：以 BOM 为准",
                    "Seller Facility ID 与 BOM Group Code Supplier 不一致。",
                    supplier_materials[0],
                )
                mismatch_cell_count += 1

        return mismatch_cell_count

    def _insert_missing_material_rows(
        self,
        ws,
        groups_with_missing: Sequence[Tuple[ProductionGroup, Sequence[BomMaterial]]],
        columns: Dict[str, int],
        max_column: int,
    ) -> Set[int]:
        inserted_rows: Set[int] = set()
        for group, materials in sorted(
            groups_with_missing,
            key=lambda item: max(item[0].rows),
            reverse=True,
        ):
            template_row = max(group.rows)
            insert_at = template_row + 1
            for material in sorted(materials, key=lambda item: self._normalize_key(item.material_ref)):
                ws.insert_rows(insert_at)
                self._copy_row(ws, template_row, insert_at, max_column, copy_values=False)

                for column_key in [
                    "style id",
                    "production lot id",
                    "production quantity (in units)",
                    "units",
                    "production date",
                ]:
                    ws.cell(
                        row=insert_at,
                        column=columns[column_key],
                        value=ws.cell(row=template_row, column=columns[column_key]).value,
                    )
                ws.cell(row=insert_at, column=columns["recording facility id"], value=material.factory)
                ws.cell(row=insert_at, column=columns["input material uid/id"], value=material.material_ref)
                ws.cell(row=insert_at, column=columns["input lot quantity used (in units)"], value=None)
                ws.cell(row=insert_at, column=columns["料率"], value=None)
                ws.cell(row=insert_at, column=columns["颜色"], value=material.color or None)
                ws.cell(row=insert_at, column=columns["seller facility id"], value=material.group_code_supplier)

                self._set_row_diagnostic(
                    ws,
                    insert_at,
                    columns,
                    "需补入",
                    "BOM汇总存在，Production缺少材料",
                    f"Production 缺少材料：{material.material_ref}，Factory：{material.factory}",
                    material,
                )
                self._set_row_fill(ws, insert_at, max_column, self.MISSING_ROW_FILL)
                inserted_rows.add(insert_at)
                insert_at += 1
        return inserted_rows

    def _apply_color_values(
        self,
        ws,
        header_row: int,
        columns: Dict[str, int],
        color_by_style_material: Dict[Tuple[str, str], str],
    ) -> None:
        color_col = columns["颜色"]
        for row_index in range(header_row + 1, ws.max_row + 1):
            style_key = self._normalize_key(
                ws.cell(row=row_index, column=columns["style id"]).value,
            )
            material_key = self._normalize_key(
                ws.cell(row=row_index, column=columns["input material uid/id"]).value,
            )
            color = color_by_style_material.get((style_key, material_key))
            ws.cell(row=row_index, column=color_col, value=color or None)

    def _apply_rate_formulas(
        self,
        ws,
        header_row: int,
        columns: Dict[str, int],
        inserted_rows: Set[int],
    ) -> int:
        rate_col = columns["料率"]
        used_col_letter = get_column_letter(columns["input lot quantity used (in units)"])
        style_col_letter = get_column_letter(columns["style id"])
        lot_col_letter = get_column_letter(columns["production lot id"])
        material_col_letter = get_column_letter(columns["input material uid/id"])
        quantity_col_letter = get_column_letter(columns["production quantity (in units)"])
        rate_count = 0
        for row_index in range(header_row + 1, ws.max_row + 1):
            rate_cell = ws.cell(row=row_index, column=rate_col)
            rate_cell.number_format = "0.0000"
            if row_index in inserted_rows:
                rate_cell.value = None
                continue
            check_result = self._normalize_text(
                ws.cell(row=row_index, column=columns["check result"]).value,
            )
            if check_result == "需补入":
                rate_cell.value = None
                continue
            material = self._normalize_text(
                ws.cell(row=row_index, column=columns["input material uid/id"]).value,
            )
            quantity = self._normalize_text(
                ws.cell(row=row_index, column=columns["production quantity (in units)"]).value,
            )
            if not material or not quantity:
                rate_cell.value = None
                continue
            rate_cell.value = (
                f'=IFERROR(SUMIFS(${used_col_letter}:${used_col_letter},'
                f'${style_col_letter}:${style_col_letter},{style_col_letter}{row_index},'
                f'${lot_col_letter}:${lot_col_letter},{lot_col_letter}{row_index},'
                f'${material_col_letter}:${material_col_letter},{material_col_letter}{row_index})/'
                f'{quantity_col_letter}{row_index},"")'
            )
            rate_count += 1
        return rate_count

    def process_reports(
        self,
        production_path: str,
        bom_source: Union[str, os.PathLike, Sequence[str]],
        output_dir: str,
    ) -> Dict[str, Any]:
        logs: List[str] = []
        try:
            is_summary_input = isinstance(bom_source, (str, os.PathLike))
            if is_summary_input:
                source_text = os.fspath(bom_source)
                bom_paths = [source_text] if source_text else []
            else:
                bom_paths = list(bom_source)

            if not bom_paths:
                return {
                    "success": False,
                    "message": "请上传 BOM汇总 文件" if is_summary_input else "请至少上传 1 个 BOM 文件",
                    "logs": logs,
                }

            ensure_dir(output_dir)
            production_wb = openpyxl.load_workbook(production_path)
            production_ws = (
                production_wb["PRODUCTION"]
                if "PRODUCTION" in production_wb.sheetnames
                else production_wb.active
            )
            header_row, columns = self._find_header_row(
                production_ws,
                self.REQUIRED_PRODUCTION_COLUMNS,
                max_scan_rows=10,
            )
            self._prepare_output_columns(production_ws, header_row, columns)
            max_column = production_ws.max_column
            self._clear_data_fills(production_ws, header_row, max_column)

            expected_by_key, color_by_style_material, bom_material_row_count = self._build_expected_material_map(
                bom_paths,
                is_summary_input,
                logs,
            )
            if not expected_by_key:
                return {
                    "success": False,
                    "message": "未从 BOM汇总 中读取到可核对的面料" if is_summary_input else "未从 BOM 中读取到可核对的 MAIN COMPONENT 面料",
                    "logs": logs,
                }

            groups = self._build_production_groups(production_ws, header_row, columns)
            mismatch_cell_count = 0
            inconsistent_group_count = 0
            extra_material_row_count = 0
            no_bom_key_count = 0
            groups_with_missing: List[Tuple[ProductionGroup, Sequence[BomMaterial]]] = []

            for group in groups.values():
                group_inconsistent, red_cells = self._check_group_consistency(
                    production_ws,
                    group,
                    columns,
                )
                inconsistent_group_count += group_inconsistent
                mismatch_cell_count += red_cells

                expected_materials = self._expected_materials_for_group(
                    production_ws,
                    group,
                    columns,
                    expected_by_key,
                )
                if not expected_materials:
                    mismatch_cell_count += self._mark_no_bom_group(production_ws, group, columns)
                    no_bom_key_count += 1
                    continue

                extra_rows, missing_materials = self._find_extra_and_missing_materials(
                    production_ws,
                    group,
                    columns,
                    expected_materials,
                )
                for material_key, rows in extra_rows.items():
                    for row_index in rows:
                        self._set_row_diagnostic(
                            production_ws,
                            row_index,
                            columns,
                            "需删除",
                            "Production多出材料",
                            f"Production 多出材料：{material_key}",
                        )
                        self._set_row_strike(production_ws, row_index, max_column)
                        extra_material_row_count += 1

                mismatch_cell_count += self._check_group_material_and_supplier_values(
                    production_ws,
                    group,
                    columns,
                    expected_materials,
                )

                if missing_materials:
                    groups_with_missing.append((group, missing_materials))

            inserted_rows = self._insert_missing_material_rows(
                production_ws,
                groups_with_missing,
                columns,
                max_column,
            )
            missing_row_count = len(inserted_rows)
            self._apply_color_values(
                production_ws,
                header_row,
                columns,
                color_by_style_material,
            )
            rate_row_count = self._apply_rate_formulas(
                production_ws,
                header_row,
                columns,
                inserted_rows,
            )

            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(output_dir, f"production_compare_{timestamp}.xlsx")
            production_wb.save(output_path)

            return {
                "success": True,
                "message": f"PRODUCTION核对完成，结果文件：{output_path}",
                "logs": logs,
                "bom_count": len(bom_paths),
                "bom_material_row_count": bom_material_row_count,
                "checked_row_count": sum(len(group.rows) for group in groups.values()),
                "mismatch_cell_count": mismatch_cell_count,
                "inconsistent_group_count": inconsistent_group_count,
                "extra_material_row_count": extra_material_row_count,
                "missing_row_count": missing_row_count,
                "rate_row_count": rate_row_count,
                "no_bom_key_count": no_bom_key_count,
                "output_path": output_path,
            }
        except Exception as exc:
            logs.append(f"ERROR: {exc}")
            return {
                "success": False,
                "message": str(exc),
                "logs": logs,
            }

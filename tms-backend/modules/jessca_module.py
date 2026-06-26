
# -*- coding: utf-8 -*-
"""
Jessca 数据核对模块
从 TMS工具_20260518_2100.pyw 提取的核心逻辑
创建时间: 2026-05-18
"""

import os
import re
import openpyxl
import xlrd
import pandas as pd
from collections import defaultdict
from dataclasses import dataclass
from datetime import date, datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
from typing import DefaultDict, List, Dict, Tuple, Any, Optional, Set

# 导入工具模块
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from utils.excel_utils import safe_float, ExcelRowAdapter, XlrdAdapter, OpenpyxlAdapter
from utils.file_utils import ensure_dir, create_thin_border


STATUS_MATCHED = "一致"
STATUS_PRICE_MISMATCH = "价格不一致"
STATUS_STYLE_SUSPECT = "发票款号疑似错误"
STATUS_ARTICLE_SUSPECT = "发票Article疑似错误"
STATUS_MULTI_CANDIDATE = "字段疑似错误-候选多条"
STATUS_REFERENCE_STYLE_SUSPECT = "参考表款号疑似错误"
STATUS_REFERENCE_ARTICLE_SUSPECT = "参考表Article疑似错误"
STATUS_REFERENCE_MULTI_CANDIDATE = "参考表字段疑似错误-候选多条"
STATUS_REFERENCE_MISSING = "参考表未找到"
STATUS_NOT_IN_INVOICE = "未在发票中找到"

DIAGNOSTIC_COLUMNS = [
    "疑似错误字段",
    "诊断说明",
    "建议参考款号",
    "建议参考Article",
    "建议参考价格",
]


@dataclass(frozen=True)
class ReferenceRecord:
    """参考表基准行，row_key 用来回写到原 DataFrame 行。"""

    row_key: Any
    article: str
    style: str
    price: float


@dataclass(frozen=True)
class InvoiceDiagnostic:
    """单条发票明细相对于参考表的诊断结果。"""

    invoice_file: str
    article: str
    style: str
    price: float
    status: str
    suspected_field: str
    message: str
    suggested_style: str
    suggested_article: str
    suggested_price: Optional[float]
    candidate_row_keys: Tuple[Any, ...]


@dataclass(frozen=True)
class InvoiceRecord:
    """FTY 发票明细中可用于参考表价格核对和 TC INV PDF 核对的一行。"""

    invoice_file: str
    invoice_number: str
    invoice_date: str
    po_number: str
    article: str
    style: str
    quantity: Optional[float]
    price: float
    line_amount: Optional[float] = None
    goods_description: str = ""


@dataclass(frozen=True)
class InvoiceSummaryRecord:
    """FTY 发票底部 Total / Final Total 汇总字段。"""

    source_file: str
    total_quantity: Optional[float] = None
    total_amount: Optional[float] = None
    freight_charge: Optional[float] = None
    documentation_charge: Optional[float] = None
    final_total_amount: Optional[float] = None
    currency: str = ""


@dataclass(frozen=True)
class TcInvoiceExtractedPage:
    """TC INV PDF 每页抽取出的文本和表格。"""

    text: str
    tables: List[List[List[Optional[str]]]]


@dataclass(frozen=True)
class TcInvoiceRecord:
    """TC INV PDF 中按 PO/Article/Working No 提取的一行。"""

    source_file: str
    po_number: str
    market_po: str
    working_number: str
    article_number: str
    quantity: Optional[float]
    price: Optional[float] = None
    total_amount: Optional[float] = None
    goods_description: str = ""


@dataclass(frozen=True)
class TcInvoiceSummary:
    """TC INV PDF 底部整单汇总字段。"""

    source_file: str
    total_quantity: Optional[float] = None
    total_carton: Optional[float] = None
    total_gross_weight: Optional[float] = None
    total_net_weight: Optional[float] = None
    total_net_net_weight: Optional[float] = None
    total_po_net_amount: Optional[float] = None
    additional_charge: Optional[float] = None
    documentation_charge: Optional[float] = None
    total_vat: Optional[float] = None
    invoice_total: Optional[float] = None


@dataclass(frozen=True)
class TcComparisonRow:
    """FTY 发票源文件与 TC INV PDF 的核对结果行。"""

    status: str
    issue_detail: str
    po_number: str
    article: str
    style: str
    fty_quantity: Optional[float]
    tc_quantity: Optional[float]
    fty_price: Optional[float]
    tc_price: Optional[float]
    fty_line_amount: Optional[float]
    tc_total_amount: Optional[float]
    fty_goods_description: str
    tc_goods_description: str
    market_po: str
    source_invoice: str
    source_tc_pdf: str


@dataclass(frozen=True)
class TcSummaryComparisonRow:
    """FTY 发票与 TC INV PDF 的整单汇总核对结果。"""

    status: str
    issue_detail: str
    fty_total_quantity: Optional[float]
    tc_total_quantity: Optional[float]
    fty_total_amount: Optional[float]
    tc_total_po_net_amount: Optional[float]
    fty_freight_charge: Optional[float]
    tc_additional_charge: Optional[float]
    fty_documentation_charge: Optional[float]
    tc_documentation_charge: Optional[float]
    fty_final_total_amount: Optional[float]
    tc_invoice_total: Optional[float]
    tc_total_carton: Optional[float]
    tc_total_gross_weight: Optional[float]
    tc_total_net_weight: Optional[float]
    tc_total_net_net_weight: Optional[float]
    tc_total_vat: Optional[float]
    source_invoice: str
    source_tc_pdf: str


@dataclass(frozen=True)
class PackingListRecord:
    """Packing List PDF 中按 PO/Article/Working No 汇总出的一行。"""

    invoice_number: str
    ex_factory_date: str
    po_number: str
    working_number: str
    article_number: str
    customer_number: str
    quantity: Optional[float]
    cartons: Optional[int]


@dataclass(frozen=True)
class PackingComparisonRow:
    """发票源文件与 Packing List PDF 的核对结果行。"""

    status: str
    issue_detail: str
    invoice_number: str
    packing_invoice_number: str
    invoice_date: str
    ex_factory_date: str
    po_number: str
    article: str
    style: str
    invoice_quantity: Optional[float]
    packing_quantity: Optional[float]
    packing_cartons: Optional[int]
    market_po: str
    source_invoice: str


class JesscaModule:
    """Jessca 数据核对业务逻辑"""

    def __init__(self):
        pass

    def read_invoice_records(self, invoice_path: str) -> List[InvoiceRecord]:
        """读取发票完整明细，统一处理 .xls 和 .xlsx 格式。"""

        if invoice_path.lower().endswith('.xls'):
            wb = xlrd.open_workbook(invoice_path)
            ws = wb.sheet_by_index(0)
            adapter = XlrdAdapter(wb, ws)
        else:
            wb = openpyxl.load_workbook(invoice_path, read_only=True)
            ws = wb.active
            adapter = OpenpyxlAdapter(wb, ws)

        try:
            return self._parse_invoice_records(adapter, os.path.basename(invoice_path))
        finally:
            adapter.close()

    def read_invoice_data(self, invoice_path: str) -> List[Dict[str, Any]]:
        """读取发票价格核对数据，保持旧调用方的返回结构不变。"""

        return [
            {
                'article': record.article,
                'style': record.style,
                'price': record.price,
            }
            for record in self.read_invoice_records(invoice_path)
        ]

    def read_invoice_summary(
        self,
        invoice_path: str,
        invoice_records: Optional[List[InvoiceRecord]] = None,
    ) -> InvoiceSummaryRecord:
        """读取 FTY 发票底部 Total / Final Total；读取失败时回退到明细汇总。"""

        invoice_file = os.path.basename(invoice_path)
        try:
            if invoice_path.lower().endswith('.xls'):
                wb = xlrd.open_workbook(invoice_path, formatting_info=False)
                sheet = wb.sheet_by_index(0)
                adapter = XlrdAdapter(wb, sheet)
            else:
                wb = openpyxl.load_workbook(invoice_path, data_only=True)
                ws = wb.active
                adapter = OpenpyxlAdapter(wb, ws)

            try:
                return self._parse_invoice_summary(adapter, invoice_file)
            finally:
                adapter.close()
        except Exception:
            return self._build_invoice_summary_from_records(invoice_file, invoice_records or [])

    def _parse_invoice_data(self, adapter: ExcelRowAdapter) -> List[Dict[str, Any]]:
        """核心解析逻辑，使用适配器访问数据"""
        return [
            {
                'article': record.article,
                'style': record.style,
                'price': record.price,
            }
            for record in self._parse_invoice_records(adapter, "")
        ]

    def _parse_invoice_summary(self, adapter: ExcelRowAdapter, invoice_file: str) -> InvoiceSummaryRecord:
        total_quantity: Optional[float] = None
        total_amount: Optional[float] = None
        freight_charge: Optional[float] = None
        documentation_charge: Optional[float] = None
        final_total_amount: Optional[float] = None
        currency = ""

        for row_idx in range(adapter.get_row_count()):
            row_texts = [
                self._normalize_invoice_text(adapter.get_cell_value(row_idx, col_idx))
                for col_idx in range(adapter.get_col_count(row_idx))
            ]
            labels = [text.strip().upper() for text in row_texts]
            if "TOTAL" in labels:
                total_quantity = self._extract_total_quantity(adapter, row_idx)
                total_amount = self._extract_last_numeric_value(adapter, row_idx)
                currency = currency or self._extract_currency_from_row(adapter, row_idx)
            if "FREIGHT CHARGE" in labels:
                freight_charge = self._extract_last_numeric_value(adapter, row_idx)
                currency = currency or self._extract_currency_from_row(adapter, row_idx)
            if "DOCUMENTATION CHARGE" in labels:
                documentation_charge = self._extract_last_numeric_value(adapter, row_idx)
                currency = currency or self._extract_currency_from_row(adapter, row_idx)
            if "FINAL TOTAL" in labels:
                final_total_amount = self._extract_last_numeric_value(adapter, row_idx)
                currency = currency or self._extract_currency_from_row(adapter, row_idx)

        return InvoiceSummaryRecord(
            source_file=invoice_file,
            total_quantity=total_quantity,
            total_amount=total_amount,
            freight_charge=freight_charge,
            documentation_charge=documentation_charge,
            final_total_amount=final_total_amount,
            currency=currency,
        )

    def _build_invoice_summary_from_records(
        self,
        invoice_file: str,
        records: List[InvoiceRecord],
    ) -> InvoiceSummaryRecord:
        total_quantity = self._sum_optional_numbers(record.quantity for record in records)
        total_amount = self._sum_optional_numbers(record.line_amount for record in records)
        return InvoiceSummaryRecord(
            source_file=invoice_file,
            total_quantity=total_quantity,
            total_amount=total_amount,
            freight_charge=None,
            documentation_charge=None,
            final_total_amount=total_amount,
            currency="",
        )

    def _parse_invoice_records(self, adapter: ExcelRowAdapter, invoice_file: str) -> List[InvoiceRecord]:
        """核心解析逻辑，使用适配器访问数据。"""
        invoice_number, invoice_date = self._extract_invoice_header(adapter)
        data: List[Dict[str, Any]] = []
        pending_items: List[Dict[str, Any]] = []
        current_goods_description = ""

        for row_idx in range(adapter.get_row_count()):
            # 获取第一列值
            col0_val = str(adapter.get_cell_value(row_idx, 0) or "").strip()

            # 识别PO行
            if col0_val.upper().startswith("PO") and "NO" not in col0_val.upper():
                price_val = self._extract_price(adapter, row_idx)
                if price_val is not None:
                    pending_items.append({
                        'po_number': self._normalize_invoice_text(
                            adapter.get_cell_value(row_idx, 1) if adapter.get_col_count(row_idx) > 1 else ""
                        ),
                        'quantity': self._extract_quantity(adapter, row_idx),
                        'article': None,
                        'price': price_val,
                        'line_amount': self._extract_line_amount(adapter, row_idx),
                        'goods_description': current_goods_description,
                    })

            # 识别ARTICLE行
            elif col0_val.upper().startswith("ARTICLE"):
                if adapter.get_col_count(row_idx) > 1:
                    article_val = str(adapter.get_cell_value(row_idx, 1) or "").strip()
                    if pending_items:
                        for item in pending_items:
                            if item.get('article') in (None, ""):
                                item['article'] = article_val

            # 识别STYLE行
            elif col0_val.upper().startswith("STYLE"):
                if adapter.get_col_count(row_idx) > 1:
                    style_val = str(adapter.get_cell_value(row_idx, 1) or "").strip()
                    # 将pending_items转为最终数据
                    for item in pending_items:
                        if item['article'] is not None and item['price'] is not None:
                            data.append({
                                'invoice_file': invoice_file,
                                'invoice_number': invoice_number,
                                'invoice_date': invoice_date,
                                'po_number': item['po_number'],
                                'article': item['article'],
                                'style': style_val,
                                'quantity': item['quantity'],
                                'price': item['price'],
                                'line_amount': item.get('line_amount'),
                                'goods_description': item.get('goods_description', ''),
                            })
                        elif item['price'] is not None:
                            data.append({
                                'invoice_file': invoice_file,
                                'invoice_number': invoice_number,
                                'invoice_date': invoice_date,
                                'po_number': item['po_number'],
                                'article': '',
                                'style': style_val,
                                'quantity': item['quantity'],
                                'price': item['price'],
                                'line_amount': item.get('line_amount'),
                                'goods_description': item.get('goods_description', ''),
                            })
                    pending_items = []
                    current_goods_description = ""

            elif self._looks_like_invoice_goods_description(col0_val):
                current_goods_description = self._normalize_goods_description_text(col0_val)

        return [
            InvoiceRecord(
                invoice_file=str(item.get('invoice_file') or ""),
                invoice_number=str(item.get('invoice_number') or ""),
                invoice_date=str(item.get('invoice_date') or ""),
                po_number=str(item.get('po_number') or ""),
                article=str(item.get('article') or ""),
                style=str(item.get('style') or ""),
                quantity=item.get('quantity'),
                price=float(item.get('price') or 0),
                line_amount=item.get('line_amount'),
                goods_description=str(item.get('goods_description') or ""),
            )
            for item in data
        ]

    @staticmethod
    def _looks_like_invoice_goods_description(value: str) -> bool:
        text = str(value or "").strip()
        if not text:
            return False
        upper_text = text.upper()
        if upper_text.startswith(("PO", "ARTICLE", "STYLE", "MADE IN", "TOTAL")):
            return False
        if "PO NO" in upper_text or "QTY" in upper_text or "UNIT PRICE" in upper_text:
            return False
        if set(text) <= {"*"}:
            return False
        return bool(re.search(r"[A-Z]", upper_text))

    def _extract_invoice_header(self, adapter: ExcelRowAdapter) -> Tuple[str, str]:
        invoice_number = ""
        invoice_date = ""
        for row_idx in range(min(adapter.get_row_count(), 12)):
            for col_idx in range(adapter.get_col_count(row_idx)):
                cell_value = adapter.get_cell_value(row_idx, col_idx)
                if not invoice_number:
                    invoice_number = self._extract_inline_invoice_number(cell_value)
                if not invoice_date:
                    invoice_date = self._extract_inline_invoice_date(cell_value, adapter)

                label = self._normalize_invoice_label(cell_value)
                if not invoice_number and label in {
                    "INV#",
                    "INV",
                    "INVNO",
                    "INVNO.",
                    "INVOICE#",
                    "INVOICENO",
                    "INVOICENO.",
                    "INVOICENUMBER",
                }:
                    invoice_number = self._find_next_non_empty_cell_text(adapter, row_idx, col_idx + 1)
                elif not invoice_date and label == "DATE":
                    invoice_date = self._normalize_invoice_date(
                        adapter.get_cell_value(row_idx, col_idx + 1)
                        if col_idx + 1 < adapter.get_col_count(row_idx)
                        else None,
                        adapter,
                    )
            if invoice_number and invoice_date:
                break
        return invoice_number, invoice_date

    def _extract_inline_invoice_number(self, value: Any) -> str:
        text = self._normalize_invoice_text(value)
        if not text:
            return ""

        invoice_token = r"[A-Z0-9]*\d[A-Z0-9]*(?:[-./][A-Z0-9]*\d[A-Z0-9]*)+"
        match = re.search(
            rf"\b(?:INV(?:OICE)?\s*#?|INV(?:OICE)?\s*NO\.?|NO\.?)\s*[:：]\s*({invoice_token})\b",
            text,
            re.IGNORECASE,
        )
        return match.group(1).strip() if match else ""

    def _extract_inline_invoice_date(self, value: Any, adapter: ExcelRowAdapter) -> str:
        text = self._normalize_invoice_text(value)
        match = re.search(r"\bDATE\s*[:：]\s*(.+)", text, re.IGNORECASE)
        if not match:
            return ""
        return self._normalize_invoice_date(match.group(1).strip(), adapter)

    @staticmethod
    def _normalize_invoice_label(value: Any) -> str:
        text = str(value or "").strip().upper()
        text = re.sub(r"\s+", "", text)
        return text.rstrip(":：")

    @staticmethod
    def _normalize_invoice_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _find_next_non_empty_cell_text(self, adapter: ExcelRowAdapter, row_idx: int, start_col: int) -> str:
        for col_idx in range(start_col, adapter.get_col_count(row_idx)):
            text = self._normalize_invoice_text(adapter.get_cell_value(row_idx, col_idx))
            if text:
                return text
        return ""

    def _normalize_invoice_date(self, value: Any, adapter: ExcelRowAdapter) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        if isinstance(value, (int, float)):
            datemode = getattr(getattr(adapter, "workbook", None), "datemode", None)
            if datemode is not None:
                try:
                    return xlrd.xldate_as_datetime(value, datemode).date().isoformat()
                except Exception:
                    return self._normalize_invoice_text(value)
        text = self._normalize_invoice_text(value)
        text = re.sub(r"^DATE\s*[:：]\s*", "", text, flags=re.IGNORECASE).strip()
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d/%m/%Y", "%Y/%m/%d"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        english_month_date = self._normalize_english_month_date(text)
        if english_month_date:
            return english_month_date
        return text

    @staticmethod
    def _normalize_english_month_date(text: str) -> str:
        month_lookup = {
            "JAN": 1,
            "JANUARY": 1,
            "FEB": 2,
            "FEBRUARY": 2,
            "MAR": 3,
            "MARCH": 3,
            "APR": 4,
            "APRIL": 4,
            "MAY": 5,
            "JUN": 6,
            "JUNE": 6,
            "JUL": 7,
            "JULY": 7,
            "AUG": 8,
            "AUGUST": 8,
            "SEP": 9,
            "SEPT": 9,
            "SEPTEMBER": 9,
            "OCT": 10,
            "OCTOBER": 10,
            "NOV": 11,
            "NOVEMBER": 11,
            "DEC": 12,
            "DECEMBER": 12,
        }
        parts = re.split(r"\s+", text.replace(",", " ").strip())
        if len(parts) != 3:
            return ""

        month_token = parts[0].rstrip(".").upper()
        if month_token not in month_lookup:
            return ""

        try:
            return date(
                int(parts[2]),
                month_lookup[month_token],
                int(parts[1]),
            ).isoformat()
        except ValueError:
            return ""

    def _extract_price(self, adapter: ExcelRowAdapter, row_idx: int) -> Optional[float]:
        """从指定行提取价格"""
        PRICE_MIN, PRICE_MAX = 0.1, 1000  # 价格范围常量
        price_col_candidates = self._find_price_columns(adapter, row_idx)
        for col_idx in price_col_candidates:
            candidate = self._parse_price_value(adapter.get_cell_value(row_idx, col_idx))
            if candidate is not None and PRICE_MIN <= candidate <= PRICE_MAX:
                return candidate

        for col_idx in range(adapter.get_col_count(row_idx)):
            left_text = str(adapter.get_cell_value(row_idx, col_idx - 1) or "").strip().upper() if col_idx > 0 else ""
            right_text = str(adapter.get_cell_value(row_idx, col_idx + 1) or "").strip().upper() if col_idx + 1 < adapter.get_col_count(row_idx) else ""
            if left_text != "USD" or right_text != "USD":
                continue
            candidate = self._parse_price_value(adapter.get_cell_value(row_idx, col_idx))
            if candidate is not None and PRICE_MIN <= candidate <= PRICE_MAX:
                return candidate

        return None

    def _extract_quantity(self, adapter: ExcelRowAdapter, row_idx: int) -> Optional[float]:
        """从 PO 行提取数量，优先按 QTY 表头定位。"""
        for col_idx in self._find_quantity_columns(adapter, row_idx):
            candidate = self._parse_price_value(adapter.get_cell_value(row_idx, col_idx))
            if candidate is not None and candidate >= 0:
                return int(candidate) if float(candidate).is_integer() else candidate
        return None

    def _extract_line_amount(self, adapter: ExcelRowAdapter, row_idx: int) -> Optional[float]:
        """从 PO 行提取 FOB 金额，优先按 FOB 表头定位。"""
        for col_idx in self._find_fob_amount_columns(adapter, row_idx):
            candidate = self._parse_price_value(adapter.get_cell_value(row_idx, col_idx))
            if candidate is not None and candidate >= 0:
                return candidate
        return self._extract_last_numeric_value(adapter, row_idx)

    def _find_fob_amount_columns(self, adapter: ExcelRowAdapter, row_idx: int) -> List[int]:
        amount_cols: List[int] = []
        search_start = max(0, row_idx - 5)
        for header_row in range(search_start, row_idx):
            for col_idx in range(adapter.get_col_count(header_row)):
                header_text = str(adapter.get_cell_value(header_row, col_idx) or "").strip().upper()
                if header_text == "FOB":
                    amount_cols.extend([col_idx, col_idx + 1, col_idx + 2])

        seen = set()
        ordered_cols: List[int] = []
        max_cols = adapter.get_col_count(row_idx)
        for col_idx in amount_cols:
            if 0 <= col_idx < max_cols and col_idx not in seen:
                ordered_cols.append(col_idx)
                seen.add(col_idx)
        return ordered_cols

    def _extract_total_quantity(self, adapter: ExcelRowAdapter, row_idx: int) -> Optional[float]:
        for col_idx in self._find_quantity_columns(adapter, row_idx):
            candidate = self._parse_price_value(adapter.get_cell_value(row_idx, col_idx))
            if candidate is not None and candidate >= 0:
                return int(candidate) if float(candidate).is_integer() else candidate
        for col_idx in range(adapter.get_col_count(row_idx)):
            candidate = self._parse_price_value(adapter.get_cell_value(row_idx, col_idx))
            if candidate is not None and candidate >= 0:
                return int(candidate) if float(candidate).is_integer() else candidate
        return None

    def _extract_last_numeric_value(self, adapter: ExcelRowAdapter, row_idx: int) -> Optional[float]:
        numbers: List[float] = []
        for col_idx in range(adapter.get_col_count(row_idx)):
            candidate = self._parse_price_value(adapter.get_cell_value(row_idx, col_idx))
            if candidate is not None:
                numbers.append(candidate)
        return numbers[-1] if numbers else None

    @staticmethod
    def _extract_currency_from_row(adapter: ExcelRowAdapter, row_idx: int) -> str:
        for col_idx in range(adapter.get_col_count(row_idx)):
            text = str(adapter.get_cell_value(row_idx, col_idx) or "").strip().upper()
            if re.fullmatch(r"[A-Z]{3}", text):
                return text
        return ""

    @staticmethod
    def _sum_optional_numbers(values: Any) -> Optional[float]:
        total = 0.0
        found = False
        for value in values:
            if value is None:
                continue
            total += float(value)
            found = True
        return total if found else None

    def _find_quantity_columns(self, adapter: ExcelRowAdapter, row_idx: int) -> List[int]:
        quantity_cols: List[int] = []
        search_start = max(0, row_idx - 5)
        for header_row in range(search_start, row_idx):
            for col_idx in range(adapter.get_col_count(header_row)):
                header_text = str(adapter.get_cell_value(header_row, col_idx) or "").strip().upper()
                if not header_text:
                    continue
                if ("QTY" in header_text or "QUANTITY" in header_text) and "UNIT PRICE" not in header_text:
                    quantity_cols.append(col_idx)

        seen = set()
        ordered_cols: List[int] = []
        max_cols = adapter.get_col_count(row_idx)
        for col_idx in quantity_cols + [5]:
            if 0 <= col_idx < max_cols and col_idx not in seen:
                ordered_cols.append(col_idx)
                seen.add(col_idx)
        return ordered_cols

    def _find_price_columns(self, adapter: ExcelRowAdapter, row_idx: int) -> List[int]:
        """根据表头定位单价列，避免把 QTY 数量列误当价格。"""
        price_cols: List[int] = []
        search_start = max(0, row_idx - 5)
        for header_row in range(search_start, row_idx):
            for col_idx in range(adapter.get_col_count(header_row)):
                header_text = str(adapter.get_cell_value(header_row, col_idx) or "").strip().upper()
                if not header_text:
                    continue
                if "QTY" in header_text or "QUANTITY" in header_text:
                    continue
                if "UNIT PRICE" in header_text:
                    price_cols.extend([col_idx, col_idx + 1])
                elif header_text == "PRICE":
                    price_cols.append(col_idx)

        seen = set()
        ordered_cols: List[int] = []
        max_cols = adapter.get_col_count(row_idx)
        for col_idx in price_cols + [7]:
            if 0 <= col_idx < max_cols and col_idx not in seen:
                ordered_cols.append(col_idx)
                seen.add(col_idx)
        return ordered_cols

    def _parse_price_value(self, cell_val: Any) -> Optional[float]:
        """解析单元格值为价格候选值"""
        if cell_val is None:
            return None

        if isinstance(cell_val, (int, float)):
            return float(cell_val)

        if isinstance(cell_val, str):
            cell_val_clean = cell_val.strip()
            if not cell_val_clean:
                return None

            # 尝试直接转换
            try:
                return float(cell_val_clean)
            except ValueError:
                pass

            # 使用正则提取数字
            match = re.search(r'(\d+\.?\d*)', cell_val_clean)
            if match:
                try:
                    return float(match.group(1))
                except ValueError:
                    pass

        return None

    @staticmethod
    def _normalize_reference_column_name(column_name: Any) -> str:
        text = str(column_name or "").strip().lower()
        text = re.sub(r"\s+", " ", text)
        return text

    @classmethod
    def _resolve_reference_columns(cls, ref_df: pd.DataFrame) -> Dict[str, Any]:
        # 参考表来自手工维护模板，大小写和空格不稳定，必须先归一化再定位关键列。
        aliases = {
            "article": ("article no.", "article no", "article number", "article", "货号", "款号"),
            "style": ("style no.", "style no", "style number", "style", "款式号", "款式"),
            "price": ("price", "unit price", "reference price", "参考价", "价格"),
        }
        normalized_columns = {
            cls._normalize_reference_column_name(column): column
            for column in ref_df.columns
        }

        resolved: Dict[str, Any] = {}
        for field, field_aliases in aliases.items():
            for alias in field_aliases:
                normalized_alias = cls._normalize_reference_column_name(alias)
                if normalized_alias in normalized_columns:
                    resolved[field] = normalized_columns[normalized_alias]
                    break

        missing_fields = [field for field in ("article", "style", "price") if field not in resolved]
        if missing_fields:
            raise ValueError(
                "参考表缺少必要列："
                + ", ".join(missing_fields)
                + "；当前列："
                + ", ".join(str(column) for column in ref_df.columns)
            )

        return resolved

    @staticmethod
    def _normalize_reference_text(value: Any) -> str:
        if value is None or pd.isna(value):
            return ""
        return str(value).strip()

    def _parse_reference_price(self, value: Any) -> float:
        if value is None or pd.isna(value):
            return 0.0
        parsed = self._parse_price_value(value)
        return parsed if parsed is not None else 0.0

    @staticmethod
    def _price_key(value: Optional[float]) -> int:
        if value is None:
            return 0
        return int(round(float(value) * 100))

    @staticmethod
    def _status_fill(status: str) -> Optional[PatternFill]:
        if status in (STATUS_PRICE_MISMATCH, STATUS_STYLE_SUSPECT,
                      STATUS_ARTICLE_SUSPECT, STATUS_MULTI_CANDIDATE,
                      STATUS_REFERENCE_STYLE_SUSPECT, STATUS_REFERENCE_ARTICLE_SUSPECT,
                      STATUS_REFERENCE_MULTI_CANDIDATE):
            return PatternFill(start_color='FFFFDDDD', end_color='FFFFDDDD', fill_type='solid')
        if status in (STATUS_REFERENCE_MISSING, STATUS_NOT_IN_INVOICE):
            return PatternFill(start_color='FFFFFFCC', end_color='FFFFFFCC', fill_type='solid')
        if status == STATUS_MATCHED:
            return PatternFill(start_color='FFDDFFDD', end_color='FFDDFFDD', fill_type='solid')
        return None

    @staticmethod
    def _status_font(status: str) -> Optional[Font]:
        if status in (STATUS_PRICE_MISMATCH, STATUS_STYLE_SUSPECT,
                      STATUS_ARTICLE_SUSPECT, STATUS_MULTI_CANDIDATE,
                      STATUS_REFERENCE_STYLE_SUSPECT, STATUS_REFERENCE_ARTICLE_SUSPECT,
                      STATUS_REFERENCE_MULTI_CANDIDATE):
            return Font(bold=True, color='FFFF0000')
        if status in (STATUS_REFERENCE_MISSING, STATUS_NOT_IN_INVOICE):
            return Font(bold=True, color='FFCC6600')
        if status == STATUS_MATCHED:
            return Font(bold=True, color='FF00AA00')
        return None

    @staticmethod
    def _is_reference_source_status(status: str) -> bool:
        return status in (
            STATUS_REFERENCE_STYLE_SUSPECT,
            STATUS_REFERENCE_ARTICLE_SUSPECT,
            STATUS_REFERENCE_MULTI_CANDIDATE,
        )

    def _build_reference_records(self, ref_df: pd.DataFrame,
                                 reference_columns: Dict[str, Any]) -> List[ReferenceRecord]:
        records: List[ReferenceRecord] = []
        for row_key, row in ref_df.iterrows():
            records.append(ReferenceRecord(
                row_key=row_key,
                article=self._normalize_reference_text(row.get(reference_columns["article"])),
                style=self._normalize_reference_text(row.get(reference_columns["style"])),
                price=self._parse_reference_price(row.get(reference_columns["price"])),
            ))
        return records

    def _build_reference_index(self, records: List[ReferenceRecord]) -> Dict[str, Dict[Tuple[Any, ...], List[ReferenceRecord]]]:
        indexes: Dict[str, DefaultDict[Tuple[Any, ...], List[ReferenceRecord]]] = {
            "article_style": defaultdict(list),
            "article_price": defaultdict(list),
            "style_price": defaultdict(list),
        }
        for record in records:
            price_key = self._price_key(record.price)
            indexes["article_style"][(record.article, record.style)].append(record)
            indexes["article_price"][(record.article, price_key)].append(record)
            indexes["style_price"][(record.style, price_key)].append(record)
        return {name: dict(index) for name, index in indexes.items()}

    @staticmethod
    def _unique_reference_records(records: List[ReferenceRecord]) -> List[ReferenceRecord]:
        unique_records: List[ReferenceRecord] = []
        seen = set()
        for record in records:
            if record.row_key in seen:
                continue
            unique_records.append(record)
            seen.add(record.row_key)
        return unique_records

    @staticmethod
    def _join_candidate_values(records: List[ReferenceRecord], field_name: str) -> str:
        values: List[str] = []
        for record in records:
            value = str(getattr(record, field_name))
            if value and value not in values:
                values.append(value)
        return " / ".join(values[:5])

    def _make_invoice_diagnostic(self, invoice_file: str, article: str, style: str,
                                 price: float, record: ReferenceRecord, status: str,
                                 suspected_field: str, message: str) -> InvoiceDiagnostic:
        return InvoiceDiagnostic(
            invoice_file=invoice_file,
            article=article,
            style=style,
            price=price,
            status=status,
            suspected_field=suspected_field,
            message=message,
            suggested_style=record.style,
            suggested_article=record.article,
            suggested_price=record.price,
            candidate_row_keys=(record.row_key,),
        )

    def _diagnose_with_article_style(self, invoice_file: str, article: str, style: str,
                                     price: float, records: List[ReferenceRecord]) -> InvoiceDiagnostic:
        matched_records = [
            record for record in records
            if self._price_key(record.price) == self._price_key(price)
        ]
        if matched_records:
            return self._make_invoice_diagnostic(
                invoice_file, article, style, price, matched_records[0],
                STATUS_MATCHED, "", "发票款号、Article、价格均与参考表一致。"
            )

        record = records[0]
        message = f"发票价格 {price:.2f} 与参考价格 {record.price:.2f} 不一致。"
        return self._make_invoice_diagnostic(
            invoice_file, article, style, price, record,
            STATUS_PRICE_MISMATCH, "价格", message
        )

    def _diagnose_without_exact_key(self, invoice_file: str, article: str, style: str,
                                    price: float,
                                    reference_index: Dict[str, Dict[Tuple[Any, ...], List[ReferenceRecord]]],
                                    packing_confirmed: bool = False) -> InvoiceDiagnostic:
        price_key = self._price_key(price)
        article_price_records = reference_index["article_price"].get((article, price_key), [])
        style_price_records = reference_index["style_price"].get((style, price_key), [])
        candidates = self._unique_reference_records(article_price_records + style_price_records)

        if len(candidates) == 1 and article_price_records:
            status = STATUS_REFERENCE_STYLE_SUSPECT if packing_confirmed else STATUS_STYLE_SUSPECT
            message = (
                "FTY 发票与 TC INV PDF 已一致，参考表款号疑似需要维护。"
                if packing_confirmed
                else "发票 Article 和价格可匹配参考表，发票款号疑似抓取错误。"
            )
            return self._make_invoice_diagnostic(
                invoice_file, article, style, price, candidates[0], status,
                "款号", message
            )
        if len(candidates) == 1 and style_price_records:
            status = STATUS_REFERENCE_ARTICLE_SUSPECT if packing_confirmed else STATUS_ARTICLE_SUSPECT
            message = (
                "FTY 发票与 TC INV PDF 已一致，参考表 Article 疑似需要维护。"
                if packing_confirmed
                else "发票款号和价格可匹配参考表，发票 Article 疑似抓取错误。"
            )
            return self._make_invoice_diagnostic(
                invoice_file, article, style, price, candidates[0], status,
                "Article", message
            )
        if candidates:
            return self._make_multi_candidate_diagnostic(
                invoice_file, article, style, price, candidates,
                bool(article_price_records), bool(style_price_records),
                packing_confirmed=packing_confirmed
            )
        return InvoiceDiagnostic(
            invoice_file=invoice_file,
            article=article,
            style=style,
            price=price,
            status=STATUS_REFERENCE_MISSING,
            suspected_field="",
            message="发票款号、Article、价格无法用任意两项匹配参考表。",
            suggested_style="",
            suggested_article="",
            suggested_price=None,
            candidate_row_keys=(),
        )

    def _make_multi_candidate_diagnostic(self, invoice_file: str, article: str, style: str,
                                         price: float, candidates: List[ReferenceRecord],
                                         matched_article_price: bool,
                                         matched_style_price: bool,
                                         packing_confirmed: bool = False) -> InvoiceDiagnostic:
        suspected_field = "款号/Article"
        if matched_article_price and not matched_style_price:
            suspected_field = "款号"
        elif matched_style_price and not matched_article_price:
            suspected_field = "Article"
        suggested_prices = sorted({self._price_key(record.price) for record in candidates})
        suggested_price = suggested_prices[0] / 100 if len(suggested_prices) == 1 else None
        status = STATUS_REFERENCE_MULTI_CANDIDATE if packing_confirmed else STATUS_MULTI_CANDIDATE
        message = (
            f"FTY 发票与 TC INV PDF 已一致，两字段反查命中 {len(candidates)} 条参考记录，需维护参考表。"
            if packing_confirmed
            else f"两字段反查命中 {len(candidates)} 条参考记录，需人工确认。"
        )
        return InvoiceDiagnostic(
            invoice_file=invoice_file,
            article=article,
            style=style,
            price=price,
            status=status,
            suspected_field=suspected_field,
            message=message,
            suggested_style=self._join_candidate_values(candidates, "style"),
            suggested_article=self._join_candidate_values(candidates, "article"),
            suggested_price=suggested_price,
            candidate_row_keys=tuple(record.row_key for record in candidates),
        )

    def _build_invoice_diagnostics(self, all_invoice_data: Dict[Tuple[str, str], Dict[str, float]],
                                   ref_df: pd.DataFrame,
                                   packing_confirmed_invoice_keys: Optional[Set[Tuple[str, str, str, int]]] = None) -> List[InvoiceDiagnostic]:
        reference_columns = self._resolve_reference_columns(ref_df)
        records = self._build_reference_records(ref_df, reference_columns)
        reference_index = self._build_reference_index(records)
        diagnostics: List[InvoiceDiagnostic] = []
        confirmed_keys = packing_confirmed_invoice_keys or set()

        for (article, style), invoice_prices_dict in all_invoice_data.items():
            article_text = self._normalize_reference_text(article)
            style_text = self._normalize_reference_text(style)
            exact_records = reference_index["article_style"].get((article_text, style_text), [])
            for invoice_file, invoice_price in invoice_prices_dict.items():
                packing_confirmed = (
                    self._invoice_diagnostic_key(invoice_file, article_text, style_text, invoice_price)
                    in confirmed_keys
                )
                if exact_records:
                    diagnostics.append(self._diagnose_with_article_style(
                        invoice_file, article_text, style_text, invoice_price, exact_records
                    ))
                else:
                    diagnostics.append(self._diagnose_without_exact_key(
                        invoice_file, article_text, style_text, invoice_price, reference_index,
                        packing_confirmed=packing_confirmed
                    ))
        return diagnostics

    def _invoice_diagnostic_key(self, invoice_file: str, article: Any, style: Any,
                                price: Optional[float]) -> Tuple[str, str, str, int]:
        return (
            str(invoice_file or ""),
            self._normalize_reference_text(article),
            self._normalize_reference_text(style),
            self._price_key(price),
        )

    @staticmethod
    def _diagnostic_severity(status: str) -> int:
        severity_order = {
            STATUS_PRICE_MISMATCH: 0,
            STATUS_STYLE_SUSPECT: 1,
            STATUS_ARTICLE_SUSPECT: 1,
            STATUS_REFERENCE_STYLE_SUSPECT: 1,
            STATUS_REFERENCE_ARTICLE_SUSPECT: 1,
            STATUS_MULTI_CANDIDATE: 2,
            STATUS_REFERENCE_MULTI_CANDIDATE: 2,
            STATUS_MATCHED: 9,
        }
        return severity_order.get(status, 20)

    def _diagnostics_by_reference_row(self, diagnostics: List[InvoiceDiagnostic]) -> Dict[Any, List[InvoiceDiagnostic]]:
        diagnostics_by_row: Dict[Any, List[InvoiceDiagnostic]] = {}
        for diagnostic in diagnostics:
            for row_key in diagnostic.candidate_row_keys:
                diagnostics_by_row.setdefault(row_key, []).append(diagnostic)
        for row_key, row_diagnostics in diagnostics_by_row.items():
            diagnostics_by_row[row_key] = sorted(
                row_diagnostics,
                key=lambda item: self._diagnostic_severity(item.status)
            )
        return diagnostics_by_row

    @staticmethod
    def _count_diagnostics(diagnostics: List[InvoiceDiagnostic]) -> Dict[str, int]:
        counts: Dict[str, int] = {}
        for diagnostic in diagnostics:
            counts[diagnostic.status] = counts.get(diagnostic.status, 0) + 1
        return counts

    @staticmethod
    def _packing_key(po_number: Any, article: Any, style: Any) -> Tuple[str, str, str]:
        return (
            str(po_number or "").strip().upper(),
            str(article or "").strip().upper(),
            str(style or "").strip().upper(),
        )

    def _build_packing_confirmed_invoice_keys(
        self,
        invoice_records: List[InvoiceRecord],
        packing_records: List[PackingListRecord],
    ) -> Set[Tuple[str, str, str, int]]:
        packing_by_key: Dict[Tuple[str, str, str], PackingListRecord] = {
            self._packing_key(record.po_number, record.article_number, record.working_number): record
            for record in packing_records
        }
        confirmed_keys: Set[Tuple[str, str, str, int]] = set()
        for invoice in invoice_records:
            packing = packing_by_key.get(self._packing_key(invoice.po_number, invoice.article, invoice.style))
            if packing is None:
                continue
            # 只有 Packing List 核对完全一致时，才把参考表差异归因到参考表。
            comparison = self._build_packing_comparison_row(invoice, packing)
            if comparison.status == STATUS_MATCHED:
                confirmed_keys.add(
                    self._invoice_diagnostic_key(
                        invoice.invoice_file,
                        invoice.article,
                        invoice.style,
                        invoice.price,
                    )
                )
        return confirmed_keys

    @staticmethod
    def _same_text(left: Any, right: Any) -> bool:
        return str(left or "").strip().upper() == str(right or "").strip().upper()

    @staticmethod
    def _same_number(left: Optional[float], right: Optional[float]) -> bool:
        if left is None or right is None:
            return left is right
        return abs(float(left) - float(right)) < 0.0001

    @staticmethod
    def _normalize_tc_header(value: Any) -> str:
        return re.sub(r"[^A-Z0-9]+", "", str(value or "").upper())

    @staticmethod
    def _parse_optional_number(value: Any) -> Optional[float]:
        if value in (None, ""):
            return None
        text = str(value).strip().replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except (TypeError, ValueError):
            return None

    @classmethod
    def _extract_labeled_number(cls, text: str, label: str) -> Optional[float]:
        escaped_label = re.escape(label)
        pattern = re.compile(
            rf"{escaped_label}\s+([-+]?\d[\d,]*(?:\.\d+)?)",
            re.IGNORECASE,
        )
        matches = list(pattern.finditer(text or ""))
        if not matches:
            return None
        return cls._parse_optional_number(matches[-1].group(1))

    @staticmethod
    def _normalize_goods_description_text(value: Any) -> str:
        text = str(value or "").strip()
        if not text:
            return ""
        text = re.sub(r"\(\s*HTS\s*:[^)]+\)", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\bHTS\s*:\s*[0-9.]+", "", text, flags=re.IGNORECASE)
        text = re.sub(r"\s+", " ", text)
        return text.strip(" ;,")

    @staticmethod
    def _goods_description_key(value: Any) -> str:
        text = JesscaModule._normalize_goods_description_text(value).upper()
        return re.sub(r"[^A-Z0-9]+", "", text)

    @staticmethod
    def _extract_tc_goods_descriptions(text: str) -> List[str]:
        descriptions: List[str] = []
        pattern = re.compile(
            r"Goods Description\s+(?P<description>.*?)(?=\n(?:PO No\b|Customer Size\s+Sourcing Size|HTS\s+Description|Total Quantity\b|Total Carton\b|Total Gross Weight\b|Total Net Weight\b|Total PO Net Amount\b|Net Amount\b|Additional charge\b|Documentation charge\b|Total VAT\b|Invoice Total\b|For and on behalf\b|Page\s+\d+\s+of\s+\d+\b|Goods Description\b)|\Z)",
            re.IGNORECASE | re.DOTALL,
        )
        for match in pattern.finditer(text or ""):
            description = JesscaModule._normalize_goods_description_text(match.group("description"))
            if description:
                descriptions.append(description)
        return descriptions

    def read_tc_invoice_records(self, tc_invoice_path: str) -> List[TcInvoiceRecord]:
        """读取 TC INV PDF 的截图字段：PO、Market PO、Working No、Article、Total Qty、Goods Description。"""

        import pdfplumber

        pages: List[TcInvoiceExtractedPage] = []
        with pdfplumber.open(tc_invoice_path) as pdf:
            for page in pdf.pages:
                pages.append(
                    TcInvoiceExtractedPage(
                        text=page.extract_text(x_tolerance=1, y_tolerance=3) or "",
                        tables=page.extract_tables() or [],
                    )
                )
        return self.parse_tc_invoice_pages(pages, os.path.basename(tc_invoice_path))

    def read_tc_invoice_summary(
        self,
        tc_invoice_path: str,
        tc_records: Optional[List[TcInvoiceRecord]] = None,
    ) -> TcInvoiceSummary:
        """读取 TC INV PDF 底部汇总；读取失败时回退到明细汇总。"""

        import pdfplumber

        source_file = os.path.basename(tc_invoice_path)
        try:
            pages: List[TcInvoiceExtractedPage] = []
            with pdfplumber.open(tc_invoice_path) as pdf:
                for page in pdf.pages:
                    pages.append(
                        TcInvoiceExtractedPage(
                            text=page.extract_text(x_tolerance=1, y_tolerance=3) or "",
                            tables=page.extract_tables() or [],
                        )
                    )
            return self.parse_tc_invoice_summary_pages(pages, source_file)
        except Exception:
            return self._build_tc_summary_from_records(source_file, tc_records or [])

    def parse_tc_invoice_pages(
        self,
        pages: List[TcInvoiceExtractedPage],
        source_file: str,
    ) -> List[TcInvoiceRecord]:
        records: List[TcInvoiceRecord] = []
        money_rows: List[Tuple[Optional[float], Optional[float]]] = []
        full_text = "\n".join(page.text for page in pages)
        descriptions = self._extract_tc_goods_descriptions(full_text)

        for page in pages:
            for table in page.tables:
                records.extend(self._parse_tc_invoice_table(table, source_file))
                money_rows.extend(self._parse_tc_invoice_money_table(table))

        if not records and full_text.strip():
            raise ValueError("未在 TC INV PDF 中识别到 PO 明细表")

        return [
            TcInvoiceRecord(
                source_file=record.source_file,
                po_number=record.po_number,
                market_po=record.market_po,
                working_number=record.working_number,
                article_number=record.article_number,
                quantity=record.quantity,
                price=money_rows[index][0] if index < len(money_rows) else record.price,
                total_amount=money_rows[index][1] if index < len(money_rows) else record.total_amount,
                goods_description=descriptions[index] if index < len(descriptions) else record.goods_description,
            )
            for index, record in enumerate(records)
        ]

    def parse_tc_invoice_summary_pages(
        self,
        pages: List[TcInvoiceExtractedPage],
        source_file: str,
    ) -> TcInvoiceSummary:
        full_text = "\n".join(page.text for page in pages)
        return TcInvoiceSummary(
            source_file=source_file,
            total_quantity=self._extract_labeled_number(full_text, "Total Quantity"),
            total_carton=self._extract_labeled_number(full_text, "Total Carton"),
            total_gross_weight=self._extract_labeled_number(full_text, "Total Gross Weight"),
            total_net_weight=self._extract_labeled_number(full_text, "Total Net Weight"),
            total_net_net_weight=self._extract_labeled_number(full_text, "Total Net Net Weight"),
            total_po_net_amount=self._extract_labeled_number(full_text, "Total PO Net Amount"),
            additional_charge=self._extract_labeled_number(full_text, "Additional charge"),
            documentation_charge=self._extract_labeled_number(full_text, "Documentation charge"),
            total_vat=self._extract_labeled_number(full_text, "Total VAT"),
            invoice_total=self._extract_labeled_number(full_text, "Invoice Total"),
        )

    def _build_tc_summary_from_records(
        self,
        source_file: str,
        records: List[TcInvoiceRecord],
    ) -> TcInvoiceSummary:
        return TcInvoiceSummary(
            source_file=source_file,
            total_quantity=self._sum_optional_numbers(record.quantity for record in records),
            total_po_net_amount=self._sum_optional_numbers(record.total_amount for record in records),
            invoice_total=self._sum_optional_numbers(record.total_amount for record in records),
        )

    def _parse_tc_invoice_table(
        self,
        table: List[List[Optional[str]]],
        source_file: str,
    ) -> List[TcInvoiceRecord]:
        if not table:
            return []

        header_index = -1
        header_lookup: Dict[str, int] = {}
        for row_index, row in enumerate(table):
            lookup = {
                self._normalize_tc_header(cell): col_index
                for col_index, cell in enumerate(row)
                if self._normalize_tc_header(cell)
            }
            if {"PONO", "WORKINGNO", "ARTICLENO", "TOTALQTY"}.issubset(lookup.keys()):
                header_index = row_index
                header_lookup = lookup
                break

        if header_index < 0:
            return []

        records: List[TcInvoiceRecord] = []
        po_index = header_lookup["PONO"]
        market_po_index = header_lookup.get("MARKETPONUMBER", header_lookup.get("MARKETPO"))
        working_index = header_lookup["WORKINGNO"]
        article_index = header_lookup["ARTICLENO"]
        quantity_index = header_lookup["TOTALQTY"]

        for row in table[header_index + 1:]:
            po_number = self._cell_text(row, po_index)
            working_number = self._cell_text(row, working_index)
            article_number = self._cell_text(row, article_index)
            if not po_number or not working_number or not article_number:
                continue
            records.append(
                TcInvoiceRecord(
                    source_file=source_file,
                    po_number=po_number,
                    market_po=self._cell_text(row, market_po_index) if market_po_index is not None else "",
                    working_number=working_number,
                    article_number=article_number,
                    quantity=self._parse_optional_number(self._cell_text(row, quantity_index)),
                    goods_description="",
                )
            )
        return records

    def _parse_tc_invoice_money_table(
        self,
        table: List[List[Optional[str]]],
    ) -> List[Tuple[Optional[float], Optional[float]]]:
        if not table:
            return []

        header_index = -1
        header_lookup: Dict[str, int] = {}
        for row_index, row in enumerate(table):
            lookup = {
                self._normalize_tc_header(cell): col_index
                for col_index, cell in enumerate(row)
                if self._normalize_tc_header(cell)
            }
            if {"TOTALQUANTITY", "PRICE", "TOTALAMOUNT"}.issubset(lookup.keys()):
                header_index = row_index
                header_lookup = lookup
                break

        if header_index < 0:
            return []

        price_index = header_lookup["PRICE"]
        amount_index = header_lookup["TOTALAMOUNT"]
        rows: List[Tuple[Optional[float], Optional[float]]] = []
        for row in table[header_index + 1:]:
            first_cell = self._cell_text(row, 0).upper()
            if first_cell != "QTY":
                continue
            price = self._parse_optional_number(self._cell_text(row, price_index))
            total_amount = self._parse_optional_number(self._cell_text(row, amount_index))
            if price is not None or total_amount is not None:
                rows.append((price, total_amount))
        return rows

    @staticmethod
    def _cell_text(row: List[Optional[str]], index: Optional[int]) -> str:
        if index is None or index >= len(row):
            return ""
        value = row[index]
        return str(value or "").strip()

    def _build_tc_confirmed_invoice_keys(
        self,
        invoice_records: List[InvoiceRecord],
        tc_records: List[TcInvoiceRecord],
    ) -> Set[Tuple[str, str, str, int]]:
        tc_by_key: Dict[Tuple[str, str, str], List[TcInvoiceRecord]] = defaultdict(list)
        for record in tc_records:
            tc_by_key[self._packing_key(record.po_number, record.article_number, record.working_number)].append(record)

        confirmed_keys: Set[Tuple[str, str, str, int]] = set()
        for invoice in invoice_records:
            key = self._packing_key(invoice.po_number, invoice.article, invoice.style)
            candidates = tc_by_key.get(key, [])
            if not candidates:
                continue
            comparison = self._build_tc_comparison_row(invoice, candidates[0])
            if comparison.status == STATUS_MATCHED:
                confirmed_keys.add(
                    self._invoice_diagnostic_key(
                        invoice.invoice_file,
                        invoice.article,
                        invoice.style,
                        invoice.price,
                    )
                )
        return confirmed_keys

    def build_tc_invoice_comparison(
        self,
        invoice_records: List[InvoiceRecord],
        tc_records: List[TcInvoiceRecord],
    ) -> List[TcComparisonRow]:
        """按 PO + Article + Style/Working No 核对 FTY 发票源文件与 TC INV PDF。"""

        tc_by_key: Dict[Tuple[str, str, str], List[TcInvoiceRecord]] = defaultdict(list)
        for record in tc_records:
            tc_by_key[self._packing_key(record.po_number, record.article_number, record.working_number)].append(record)

        rows: List[TcComparisonRow] = []
        for invoice in invoice_records:
            key = self._packing_key(invoice.po_number, invoice.article, invoice.style)
            candidates = tc_by_key.get(key, [])
            tc_record = candidates.pop(0) if candidates else None
            rows.append(self._build_tc_comparison_row(invoice, tc_record))

        for remaining_records in tc_by_key.values():
            for tc_record in remaining_records:
                rows.append(self._build_tc_comparison_row(None, tc_record))

        return rows

    def build_tc_invoice_summary_comparison(
        self,
        invoice_summaries: List[InvoiceSummaryRecord],
        tc_summaries: List[TcInvoiceSummary],
    ) -> List[TcSummaryComparisonRow]:
        """按本次上传批次聚合比较 FTY 与 TC INV 的整单汇总。"""

        if not invoice_summaries and not tc_summaries:
            return []

        fty = self._aggregate_invoice_summaries(invoice_summaries)
        tc = self._aggregate_tc_summaries(tc_summaries)
        issues: List[str] = []
        if not invoice_summaries:
            issues.append("FTY 发票缺少底部汇总")
        if not tc_summaries:
            issues.append("TC INV PDF 缺少底部汇总")
        if invoice_summaries and tc_summaries:
            if not self._same_number(fty.total_quantity, tc.total_quantity):
                issues.append(
                    "Total Quantity 不一致："
                    f"FTY={self._format_optional_number(fty.total_quantity)}；"
                    f"TC={self._format_optional_number(tc.total_quantity)}"
                )
            if not self._same_number(fty.total_amount, tc.total_po_net_amount):
                issues.append(
                    "Total Amount 不一致："
                    f"FTY={self._format_optional_number(fty.total_amount)}；"
                    f"TC Total PO Net Amount={self._format_optional_number(tc.total_po_net_amount)}"
                )
            if not self._same_charge_amount(fty.freight_charge, tc.additional_charge):
                issues.append(
                    "Freight/Additional Charge 不一致："
                    f"FTY={self._format_optional_number(fty.freight_charge)}；"
                    f"TC={self._format_optional_number(tc.additional_charge)}"
                )
            if not self._same_charge_amount(fty.documentation_charge, tc.documentation_charge):
                issues.append(
                    "Documentation Charge 不一致："
                    f"FTY={self._format_optional_number(fty.documentation_charge)}；"
                    f"TC={self._format_optional_number(tc.documentation_charge)}"
                )
            if not self._same_number(fty.final_total_amount, tc.invoice_total):
                issues.append(
                    "Final Total 不一致："
                    f"FTY={self._format_optional_number(fty.final_total_amount)}；"
                    f"TC Invoice Total={self._format_optional_number(tc.invoice_total)}"
                )

        status = STATUS_MATCHED if not issues else ("需核对" if invoice_summaries and tc_summaries else "缺失")
        return [
            TcSummaryComparisonRow(
                status=status,
                issue_detail="；".join(issues),
                fty_total_quantity=fty.total_quantity,
                tc_total_quantity=tc.total_quantity,
                fty_total_amount=fty.total_amount,
                tc_total_po_net_amount=tc.total_po_net_amount,
                fty_freight_charge=fty.freight_charge,
                tc_additional_charge=tc.additional_charge,
                fty_documentation_charge=fty.documentation_charge,
                tc_documentation_charge=tc.documentation_charge,
                fty_final_total_amount=fty.final_total_amount,
                tc_invoice_total=tc.invoice_total,
                tc_total_carton=tc.total_carton,
                tc_total_gross_weight=tc.total_gross_weight,
                tc_total_net_weight=tc.total_net_weight,
                tc_total_net_net_weight=tc.total_net_net_weight,
                tc_total_vat=tc.total_vat,
                source_invoice=fty.source_file,
                source_tc_pdf=tc.source_file,
            )
        ]

    def _aggregate_invoice_summaries(self, summaries: List[InvoiceSummaryRecord]) -> InvoiceSummaryRecord:
        return InvoiceSummaryRecord(
            source_file=", ".join(summary.source_file for summary in summaries if summary.source_file),
            total_quantity=self._sum_optional_numbers(summary.total_quantity for summary in summaries),
            total_amount=self._sum_optional_numbers(summary.total_amount for summary in summaries),
            freight_charge=self._sum_optional_numbers(summary.freight_charge for summary in summaries),
            documentation_charge=self._sum_optional_numbers(summary.documentation_charge for summary in summaries),
            final_total_amount=self._sum_optional_numbers(summary.final_total_amount for summary in summaries),
            currency=next((summary.currency for summary in summaries if summary.currency), ""),
        )

    def _aggregate_tc_summaries(self, summaries: List[TcInvoiceSummary]) -> TcInvoiceSummary:
        return TcInvoiceSummary(
            source_file=", ".join(summary.source_file for summary in summaries if summary.source_file),
            total_quantity=self._sum_optional_numbers(summary.total_quantity for summary in summaries),
            total_carton=self._sum_optional_numbers(summary.total_carton for summary in summaries),
            total_gross_weight=self._sum_optional_numbers(summary.total_gross_weight for summary in summaries),
            total_net_weight=self._sum_optional_numbers(summary.total_net_weight for summary in summaries),
            total_net_net_weight=self._sum_optional_numbers(summary.total_net_net_weight for summary in summaries),
            total_po_net_amount=self._sum_optional_numbers(summary.total_po_net_amount for summary in summaries),
            additional_charge=self._sum_optional_numbers(summary.additional_charge for summary in summaries),
            documentation_charge=self._sum_optional_numbers(summary.documentation_charge for summary in summaries),
            total_vat=self._sum_optional_numbers(summary.total_vat for summary in summaries),
            invoice_total=self._sum_optional_numbers(summary.invoice_total for summary in summaries),
        )

    def _build_tc_comparison_row(
        self,
        invoice: Optional[InvoiceRecord],
        tc_record: Optional[TcInvoiceRecord],
    ) -> TcComparisonRow:
        issues: List[str] = []
        if invoice is None and tc_record is not None:
            issues.append("FTY 发票缺少该 PO/Article/Working No")
        elif tc_record is None and invoice is not None:
            issues.append("TC INV PDF 缺少该 PO/Article/Style")
        elif invoice is not None and tc_record is not None:
            if not self._same_number(invoice.quantity, tc_record.quantity):
                issues.append(f"QTY 不一致：FTY={invoice.quantity or '-'}；TC={tc_record.quantity or '-'}")
            if tc_record.price is not None and not self._same_number(invoice.price, tc_record.price):
                issues.append(
                    "Unit Price 不一致："
                    f"FTY={self._format_optional_number(invoice.price)}；"
                    f"TC={self._format_optional_number(tc_record.price)}"
                )
            if (invoice.line_amount is not None or tc_record.total_amount is not None) and not self._same_number(
                invoice.line_amount,
                tc_record.total_amount,
            ):
                issues.append(
                    "Line Amount 不一致："
                    f"FTY={self._format_optional_number(invoice.line_amount)}；"
                    f"TC={self._format_optional_number(tc_record.total_amount)}"
                )
            if self._goods_description_key(invoice.goods_description) != self._goods_description_key(tc_record.goods_description):
                issues.append(
                    "Goods Description 不一致："
                    f"FTY={invoice.goods_description or '-'}；TC={tc_record.goods_description or '-'}"
                )

        status = "一致" if not issues else ("需核对" if invoice and tc_record else "缺失")
        return TcComparisonRow(
            status=status,
            issue_detail="；".join(issues),
            po_number=(invoice.po_number if invoice else tc_record.po_number if tc_record else ""),
            article=(invoice.article if invoice else tc_record.article_number if tc_record else ""),
            style=(invoice.style if invoice else tc_record.working_number if tc_record else ""),
            fty_quantity=invoice.quantity if invoice else None,
            tc_quantity=tc_record.quantity if tc_record else None,
            fty_price=invoice.price if invoice else None,
            tc_price=tc_record.price if tc_record else None,
            fty_line_amount=invoice.line_amount if invoice else None,
            tc_total_amount=tc_record.total_amount if tc_record else None,
            fty_goods_description=invoice.goods_description if invoice else "",
            tc_goods_description=tc_record.goods_description if tc_record else "",
            market_po=tc_record.market_po if tc_record else "",
            source_invoice=invoice.invoice_file if invoice else "",
            source_tc_pdf=tc_record.source_file if tc_record else "",
        )

    @staticmethod
    def _format_optional_number(value: Optional[float]) -> str:
        if value is None:
            return "-"
        number = float(value)
        if number.is_integer():
            return str(int(number))
        return f"{number:.2f}".rstrip("0").rstrip(".")

    def read_packing_list_records(self, packing_path: str) -> List[PackingListRecord]:
        """读取 Packing List PDF，复用 Draft/Packing 模块的表格解析能力。"""

        from modules.draft_packing_compare_module import DraftPackingCompareModule

        module = DraftPackingCompareModule()
        pages = module._extract_packing_pages(packing_path)
        extracted_records = module.parse_packing_pages(pages)
        full_text = "\n".join(page.text for page in pages)
        invoice_number = self._extract_packing_invoice_number(full_text)
        ex_factory_date = self._extract_packing_ex_factory_date(full_text)

        return [
            PackingListRecord(
                invoice_number=invoice_number,
                ex_factory_date=ex_factory_date,
                po_number=record.po_number,
                working_number=record.working_number,
                article_number=record.article_number,
                customer_number=record.customer_number,
                quantity=record.quantity,
                cartons=record.cartons,
            )
            for record in extracted_records
        ]

    @staticmethod
    def _extract_packing_invoice_number(text: str) -> str:
        match = re.search(r"\b\d{2}-\d{2}-\d{2}-\d{4}\b", text)
        return match.group(0) if match else ""

    @staticmethod
    def _extract_packing_ex_factory_date(text: str) -> str:
        match = re.search(r"Ex-Factory\s+Date[\s\S]{0,120}?(\d{4}-\d{2}-\d{2})", text, re.IGNORECASE)
        return match.group(1) if match else ""

    def build_packing_list_comparison(
        self,
        invoice_records: List[InvoiceRecord],
        packing_records: List[PackingListRecord],
    ) -> List[PackingComparisonRow]:
        """按 PO + Article + Style/Working No 核对发票源文件与 Packing List PDF。"""

        packing_by_key: Dict[Tuple[str, str, str], PackingListRecord] = {
            self._packing_key(record.po_number, record.article_number, record.working_number): record
            for record in packing_records
        }
        rows: List[PackingComparisonRow] = []

        for invoice in invoice_records:
            key = self._packing_key(invoice.po_number, invoice.article, invoice.style)
            packing = packing_by_key.pop(key, None)
            rows.append(self._build_packing_comparison_row(invoice, packing))

        for packing in packing_by_key.values():
            rows.append(self._build_packing_comparison_row(None, packing))

        return rows

    def _build_packing_comparison_row(
        self,
        invoice: Optional[InvoiceRecord],
        packing: Optional[PackingListRecord],
    ) -> PackingComparisonRow:
        issues: List[str] = []
        if invoice is None and packing is not None:
            issues.append("发票源文件缺少该 PO/Article/Working No")
        elif packing is None and invoice is not None:
            issues.append("Packing List PDF 缺少该 PO/Article/Style")
        elif invoice is not None and packing is not None:
            if invoice.invoice_number and packing.invoice_number and not self._same_text(invoice.invoice_number, packing.invoice_number):
                issues.append(f"Invoice Number 不一致：发票={invoice.invoice_number}；Packing={packing.invoice_number}")
            if invoice.invoice_date and packing.ex_factory_date and invoice.invoice_date != packing.ex_factory_date:
                issues.append(f"Date 不一致：发票={invoice.invoice_date}；Packing Ex-Factory={packing.ex_factory_date}")
            if not self._same_number(invoice.quantity, packing.quantity):
                issues.append(f"QTY 不一致：发票={invoice.quantity or '-'}；Packing={packing.quantity or '-'}")

        status = "一致" if not issues else ("需核对" if invoice and packing else "缺失")
        return PackingComparisonRow(
            status=status,
            issue_detail="；".join(issues),
            invoice_number=invoice.invoice_number if invoice else "",
            packing_invoice_number=packing.invoice_number if packing else "",
            invoice_date=invoice.invoice_date if invoice else "",
            ex_factory_date=packing.ex_factory_date if packing else "",
            po_number=(invoice.po_number if invoice else packing.po_number if packing else ""),
            article=(invoice.article if invoice else packing.article_number if packing else ""),
            style=(invoice.style if invoice else packing.working_number if packing else ""),
            invoice_quantity=invoice.quantity if invoice else None,
            packing_quantity=packing.quantity if packing else None,
            packing_cartons=packing.cartons if packing else None,
            market_po=packing.customer_number if packing else "",
            source_invoice=invoice.invoice_file if invoice else "",
        )

    def update_reference_table(
        self,
        ref_df: pd.DataFrame,
        all_invoice_data: Dict[Tuple[str, str], Dict[str, float]],
        packing_confirmed_invoice_keys: Optional[Set[Tuple[str, str, str, int]]] = None,
    ) -> Tuple[pd.DataFrame, Dict[str, int]]:
        result_df = ref_df.copy()
        diagnostics = self._build_invoice_diagnostics(
            all_invoice_data,
            result_df,
            packing_confirmed_invoice_keys=packing_confirmed_invoice_keys,
        )
        diagnostics_by_row = self._diagnostics_by_reference_row(diagnostics)

        statuses: List[str] = []
        invoice_prices: List[Optional[float]] = []
        source_invoices: List[Optional[str]] = []
        suspected_fields: List[str] = []
        diagnostic_messages: List[str] = []
        suggested_styles: List[str] = []
        suggested_articles: List[str] = []
        suggested_prices: List[Optional[float]] = []
        matches = {'一致': 0, '不一致': 0, '未找到': 0}

        for row_key, _ in result_df.iterrows():
            row_diagnostics = diagnostics_by_row.get(row_key, [])
            selected = row_diagnostics[0] if row_diagnostics else None
            if selected is None:
                self._append_missing_reference_status(
                    statuses, invoice_prices, source_invoices,
                    suspected_fields, diagnostic_messages,
                    suggested_styles, suggested_articles, suggested_prices
                )
                matches['未找到'] += 1
                continue

            statuses.append(selected.status)
            invoice_prices.append(selected.price if selected.status != STATUS_MATCHED else None)
            source_invoices.append(selected.invoice_file if selected.status != STATUS_MATCHED else None)
            suspected_fields.append(selected.suspected_field)
            diagnostic_messages.append(selected.message)
            suggested_styles.append(selected.suggested_style)
            suggested_articles.append(selected.suggested_article)
            suggested_prices.append(selected.suggested_price)
            if selected.status == STATUS_MATCHED:
                matches['一致'] += 1
            else:
                matches['不一致'] += 1

        result_df['核对状态'] = statuses
        result_df['发票价格'] = invoice_prices
        result_df['来源发票'] = source_invoices
        result_df['疑似错误字段'] = suspected_fields
        result_df['诊断说明'] = diagnostic_messages
        result_df['建议参考款号'] = suggested_styles
        result_df['建议参考Article'] = suggested_articles
        result_df['建议参考价格'] = suggested_prices

        return result_df, matches

    @staticmethod
    def _append_missing_reference_status(statuses: List[str],
                                         invoice_prices: List[Optional[float]],
                                         source_invoices: List[Optional[str]],
                                         suspected_fields: List[str],
                                         diagnostic_messages: List[str],
                                         suggested_styles: List[str],
                                         suggested_articles: List[str],
                                         suggested_prices: List[Optional[float]]) -> None:
        statuses.append(STATUS_NOT_IN_INVOICE)
        invoice_prices.append(None)
        source_invoices.append(None)
        suspected_fields.append("")
        diagnostic_messages.append("参考表该行未在发票中找到对应记录。")
        suggested_styles.append("")
        suggested_articles.append("")
        suggested_prices.append(None)

    def save_excel_with_summary(self, result_df: pd.DataFrame,
                                all_invoice_data: Dict[Tuple[str, str], Dict[str, float]],
                                invoice_file_names: List[str],
                                invoice_file_paths: List[str],
                                 ref_df: pd.DataFrame,
                                 output_path: str,
                                 tc_comparison_rows: Optional[List[TcComparisonRow]] = None,
                                 tc_summary_rows: Optional[List[TcSummaryComparisonRow]] = None,
                                 tc_confirmed_invoice_keys: Optional[Set[Tuple[str, str, str, int]]] = None) -> Dict[str, Any]:
        """保存 Excel 结果，包含汇总表"""

        wb = openpyxl.Workbook()

        default_sheet = wb.active
        wb.remove(default_sheet)

        thin_border = create_thin_border()

        # 创建核对结果表
        ws_main = wb.create_sheet("核对结果")

        ws_main.row_dimensions[1].height = 30
        headers = list(result_df.columns)
        for col_idx, header in enumerate(headers, 1):
            cell = ws_main.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFFFF', size=11)
            cell.fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row_idx, row in enumerate(result_df.itertuples(index=False), 2):
            ws_main.row_dimensions[row_idx].height = 20
            is_odd_row = ((row_idx - 2) % 2 == 0)
            row_fill = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid') if is_odd_row else None
            row_values = list(row)
            status_value = row_values[headers.index('核对状态')] if '核对状态' in headers else ''

            for col_idx, value in enumerate(row_values, 1):
                cell = ws_main.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='left', vertical='center')
                if row_fill:
                    cell.fill = row_fill
                header_text = str(headers[col_idx - 1])
                if header_text == '核对状态':
                    status_fill = self._status_fill(str(value))
                    status_font = self._status_font(str(value))
                    if status_fill:
                        cell.fill = status_fill
                    if status_font:
                        cell.font = status_font
                elif header_text == '发票价格' and status_value != STATUS_MATCHED and value is not None:
                    cell.font = Font(color='FFFF0000', bold=True)
                elif header_text in DIAGNOSTIC_COLUMNS:
                    cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

        # 创建汇总表：按发票明细纵向展开，避免文件多时横向过宽
        ws_summary = wb.create_sheet("汇总表")
        summary_stats = self._write_compact_summary_sheet(
            ws_summary,
            all_invoice_data,
            invoice_file_names,
            invoice_file_paths,
            ref_df,
            thin_border,
            packing_confirmed_invoice_keys=tc_confirmed_invoice_keys,
        )

        tc_stats = {
            'tc_count': 0,
            'tc_matched_count': 0,
            'tc_issue_count': 0,
            'tc_summary_count': 0,
            'tc_summary_issue_count': 0,
            'tc_total_issue_count': 0,
        }
        if tc_comparison_rows is not None:
            ws_tc = wb.create_sheet("TC INV核对")
            tc_stats = self._write_tc_invoice_comparison_sheet(
                ws_tc,
                tc_comparison_rows,
                tc_summary_rows,
            )

        # 自动调整列宽
        main_column_max_widths = {}
        main_column_min_widths = {}
        for col_idx, header in enumerate(headers, 1):
            header_text = str(header)
            header_lower = header_text.lower()
            if header_text == '来源发票':
                main_column_max_widths[col_idx] = 42
                main_column_min_widths[col_idx] = 18
            elif header_text == '核对状态':
                main_column_max_widths[col_idx] = 18
                main_column_min_widths[col_idx] = 14
            elif header_text == '发票价格' or 'price' in header_lower:
                main_column_max_widths[col_idx] = 16
                main_column_min_widths[col_idx] = 12
            elif 'article' in header_lower or 'style' in header_lower:
                main_column_max_widths[col_idx] = 20
                main_column_min_widths[col_idx] = 14
        self._adjust_column_widths(ws_main, main_column_max_widths, main_column_min_widths)

        summary_column_widths = {
            1: 18,
            2: 15,
            3: 10,
            4: 20,
            5: 10,
            6: 20,
            7: 42,
            8: 16,
            9: 48,
            10: 22,
            11: 22,
            12: 16
        }
        self._adjust_column_widths(ws_summary, summary_column_widths)
        for col_idx, width in summary_column_widths.items():
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            ws_summary.column_dimensions[column_letter].width = width

        wb.save(output_path)

        return {
            'output_path': output_path,
            'sheet_count': len(wb.sheetnames),
            'sheet_names': wb.sheetnames,
            'missing_count': summary_stats['missing_count'],
            'data_count': summary_stats['data_count'],
            'diagnostics': summary_stats['diagnostics'],
            **tc_stats,
        }

    def _write_tc_invoice_comparison_sheet(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        rows: List[TcComparisonRow],
        summary_rows: Optional[List[TcSummaryComparisonRow]] = None,
    ) -> Dict[str, int]:
        summary_rows = summary_rows or []
        thin_border = create_thin_border()
        ws.merge_cells("A1:G1")
        title_cell = ws["A1"]
        title_cell.value = f"TC INV 核对结果 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(size=14, bold=True, color="FF1F2937")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28

        matched_fill = PatternFill(start_color="FFDDFFDD", end_color="FFDDFFDD", fill_type="solid")
        issue_fill = PatternFill(start_color="FFFFDDDD", end_color="FFFFDDDD", fill_type="solid")
        neutral_fill = PatternFill(start_color="FFF3F4F6", end_color="FFF3F4F6", fill_type="solid")
        section_fill = PatternFill(start_color="FFEAF2FF", end_color="FFEAF2FF", fill_type="solid")

        matched_count = sum(1 for row in rows if row.status == STATUS_MATCHED)
        issue_count = len(rows) - matched_count
        summary_count = len(summary_rows)
        summary_issue_count = sum(1 for row in summary_rows if row.status != STATUS_MATCHED)
        summary_display_rows = self._build_tc_summary_display_rows(summary_rows)
        detail_display_rows = self._build_tc_detail_display_rows(rows)

        conclusion, quantity_text, amount_delta = self._build_tc_conclusion(summary_rows, detail_display_rows)
        charge_text = self._build_tc_charge_text(summary_rows[0]) if summary_rows else "未提供汇总"
        top_rows = [
            ("核对结论", conclusion),
            ("数量", quantity_text),
            ("金额差额", amount_delta if amount_delta is not None else "-"),
        ]
        top_rows.insert(2, ("费用", charge_text))
        for row_index, (label, value) in enumerate(top_rows, 2):
            label_cell = ws.cell(row=row_index, column=1, value=label)
            value_cell = ws.cell(row=row_index, column=2, value=value)
            label_cell.font = Font(bold=True, color="FF1F2937")
            label_cell.fill = neutral_fill
            label_cell.border = thin_border
            value_cell.border = thin_border
            value_cell.alignment = Alignment(horizontal="left", vertical="center")
            has_numeric_issue = label == "金额差额" and isinstance(value, (int, float)) and abs(float(value)) > 0.0001
            if "不一致" in str(value) or has_numeric_issue:
                value_cell.fill = issue_fill
                value_cell.font = Font(bold=True, color="FF991B1B")
            elif "一致" in str(value):
                value_cell.fill = matched_fill
                value_cell.font = Font(bold=True, color="FF166534")
            if "不一致" in str(value):
                value_cell.fill = issue_fill
                value_cell.font = Font(bold=True, color="FF991B1B")
            elif "一致" in str(value):
                value_cell.fill = matched_fill
                value_cell.font = Font(bold=True, color="FF166534")

        current_row = 6
        ws.cell(row=current_row, column=1, value="汇总差异")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="FF1F2937")
        ws.cell(row=current_row, column=1).fill = section_fill
        current_row += 1
        self._write_table_header(ws, current_row, ["差异项", "FTY", "TC", "差额", "结果"], thin_border)
        current_row += 1
        if summary_display_rows:
            for display_row in summary_display_rows:
                self._write_simple_tc_display_row(ws, current_row, display_row, thin_border, issue_fill, {1})
                current_row += 1
        else:
            summary_status = "汇总一致" if summary_rows else "未提供汇总"
            fill = matched_fill if summary_rows else neutral_fill
            self._write_simple_tc_display_row(
                ws,
                current_row,
                [summary_status, "-", "-", "-", "一致" if summary_rows else "-"],
                thin_border,
                fill,
                {1},
            )
            current_row += 1

        current_row += 2
        ws.cell(row=current_row, column=1, value="明细差异")
        ws.cell(row=current_row, column=1).font = Font(bold=True, size=12, color="FF1F2937")
        ws.cell(row=current_row, column=1).fill = section_fill
        current_row += 1

        detail_header_row = current_row
        self._write_table_header(
            ws,
            detail_header_row,
            ["PO No", "Article No", "Working/Style No", "差异项", "FTY", "TC", "差额"],
            thin_border,
        )
        current_row += 1
        if detail_display_rows:
            for display_row in detail_display_rows:
                self._write_simple_tc_display_row(ws, current_row, display_row, thin_border, issue_fill, {4, 5, 6})
                current_row += 1
        else:
            self._write_simple_tc_display_row(
                ws,
                current_row,
                ["", "", "", "明细一致", "-", "-", "-"],
                thin_border,
                matched_fill,
                {4},
            )
            current_row += 1

        self._adjust_column_widths(
            ws,
            {
                1: 18,
                2: 18,
                3: 20,
                4: 18,
                5: 28,
                6: 28,
                7: 14,
            },
            {
                1: 12,
                2: 12,
                3: 14,
                4: 12,
                5: 16,
                6: 16,
                7: 10,
            },
        )
        ws.freeze_panes = "A8"
        ws.auto_filter.ref = f"A{detail_header_row}:G{current_row - 1}"
        return {
            "tc_count": len(rows),
            "tc_matched_count": matched_count,
            "tc_issue_count": issue_count,
            "tc_summary_count": summary_count,
            "tc_summary_issue_count": summary_issue_count,
            "tc_total_issue_count": issue_count + summary_issue_count,
        }

    def _build_tc_conclusion(
        self,
        summary_rows: List[TcSummaryComparisonRow],
        detail_display_rows: List[List[Any]],
    ) -> Tuple[str, str, Optional[float]]:
        if not summary_rows:
            return ("明细不一致" if detail_display_rows else "明细一致", "未提供汇总", None)

        summary = summary_rows[0]
        quantity_matched = self._same_number(summary.fty_total_quantity, summary.tc_total_quantity)
        quantity_status = "一致" if quantity_matched else "不一致"
        quantity_text = (
            f"FTY={self._format_optional_number(summary.fty_total_quantity)} / "
            f"TC={self._format_optional_number(summary.tc_total_quantity)} / {quantity_status}"
        )
        amount_delta = self._number_delta(
            self._preferred_summary_amount(summary, "fty"),
            self._preferred_summary_amount(summary, "tc"),
        )
        if amount_delta is None:
            amount_status = "金额未提供"
        else:
            amount_status = "金额一致" if abs(amount_delta) < 0.0001 else "金额不一致"
        return f"{amount_status}；数量{quantity_status}", quantity_text, amount_delta

    def _build_tc_charge_text(self, summary: TcSummaryComparisonRow) -> str:
        freight_status = (
            "一致"
            if self._same_charge_amount(summary.fty_freight_charge, summary.tc_additional_charge)
            else "不一致"
        )
        documentation_status = (
            "一致"
            if self._same_charge_amount(summary.fty_documentation_charge, summary.tc_documentation_charge)
            else "不一致"
        )
        return (
            "Freight/Additional: "
            f"FTY={self._format_optional_number(summary.fty_freight_charge)} / "
            f"TC={self._format_optional_number(summary.tc_additional_charge)} / "
            f"{freight_status}；Documentation: "
            f"FTY={self._format_optional_number(summary.fty_documentation_charge)} / "
            f"TC={self._format_optional_number(summary.tc_documentation_charge)} / "
            f"{documentation_status}"
        )

    def _build_tc_summary_display_rows(self, summary_rows: List[TcSummaryComparisonRow]) -> List[List[Any]]:
        display_rows: List[List[Any]] = []
        for summary in summary_rows:
            if not self._same_number(summary.fty_total_quantity, summary.tc_total_quantity):
                display_rows.append(
                    [
                        "总数量",
                        self._display_optional_number(summary.fty_total_quantity),
                        self._display_optional_number(summary.tc_total_quantity),
                        self._number_delta(summary.fty_total_quantity, summary.tc_total_quantity) or "-",
                        "不一致",
                    ]
                )

            if not self._same_number(summary.fty_total_amount, summary.tc_total_po_net_amount):
                display_rows.append(
                    [
                        "货值合计",
                        self._display_optional_number(summary.fty_total_amount),
                        self._display_optional_number(summary.tc_total_po_net_amount),
                        self._number_delta(summary.fty_total_amount, summary.tc_total_po_net_amount) or "-",
                        "不一致",
                    ]
                )
            if not self._same_number(summary.fty_final_total_amount, summary.tc_invoice_total):
                display_rows.append(
                    [
                        "最终总额",
                        self._display_optional_number(summary.fty_final_total_amount),
                        self._display_optional_number(summary.tc_invoice_total),
                        self._number_delta(summary.fty_final_total_amount, summary.tc_invoice_total) or "-",
                        "不一致",
                    ]
                )
            if not self._same_charge_amount(summary.fty_freight_charge, summary.tc_additional_charge):
                display_rows.append(
                    [
                        "Freight/Additional Charge",
                        self._display_optional_number(summary.fty_freight_charge),
                        self._display_optional_number(summary.tc_additional_charge),
                        self._charge_delta(summary.fty_freight_charge, summary.tc_additional_charge),
                        "不一致",
                    ]
                )
            if not self._same_charge_amount(summary.fty_documentation_charge, summary.tc_documentation_charge):
                display_rows.append(
                    [
                        "Documentation Charge",
                        self._display_optional_number(summary.fty_documentation_charge),
                        self._display_optional_number(summary.tc_documentation_charge),
                        self._charge_delta(summary.fty_documentation_charge, summary.tc_documentation_charge),
                        "不一致",
                    ]
                )
        return display_rows

    def _build_tc_detail_display_rows(self, rows: List[TcComparisonRow]) -> List[List[Any]]:
        display_rows: List[List[Any]] = []
        for row in rows:
            if row.status == STATUS_MATCHED:
                continue

            base = [row.po_number, row.article, row.style]
            if row.status == "缺失":
                fty_value = row.source_invoice or "FTY 发票缺少"
                tc_value = row.source_tc_pdf or "TC INV PDF 缺少"
                display_rows.append([*base, "记录缺失", fty_value, tc_value, "-"])
                continue

            if not self._same_number(row.fty_quantity, row.tc_quantity):
                display_rows.append(
                    [
                        *base,
                        "数量",
                        self._display_optional_number(row.fty_quantity),
                        self._display_optional_number(row.tc_quantity),
                        self._number_delta(row.fty_quantity, row.tc_quantity) or "-",
                    ]
                )
            if row.tc_price is not None and not self._same_number(row.fty_price, row.tc_price):
                display_rows.append(
                    [
                        *base,
                        "Unit Price",
                        self._display_optional_number(row.fty_price),
                        self._display_optional_number(row.tc_price),
                        self._number_delta(row.fty_price, row.tc_price) or "-",
                    ]
                )
            if (row.fty_line_amount is not None or row.tc_total_amount is not None) and not self._same_number(
                row.fty_line_amount,
                row.tc_total_amount,
            ):
                display_rows.append(
                    [
                        *base,
                        "Line Amount",
                        self._display_optional_number(row.fty_line_amount),
                        self._display_optional_number(row.tc_total_amount),
                        self._number_delta(row.fty_line_amount, row.tc_total_amount) or "-",
                    ]
                )
            if self._goods_description_key(row.fty_goods_description) != self._goods_description_key(row.tc_goods_description):
                display_rows.append(
                    [
                        *base,
                        "Goods Description",
                        row.fty_goods_description or "-",
                        row.tc_goods_description or "-",
                        "-",
                    ]
                )
        return display_rows

    @staticmethod
    def _preferred_summary_amount(summary: TcSummaryComparisonRow, side: str) -> Optional[float]:
        if side == "fty":
            return summary.fty_final_total_amount if summary.fty_final_total_amount is not None else summary.fty_total_amount
        return summary.tc_invoice_total if summary.tc_invoice_total is not None else summary.tc_total_po_net_amount

    @staticmethod
    def _number_delta(left: Optional[float], right: Optional[float]) -> Optional[float]:
        if left is None or right is None:
            return None
        return round(float(right) - float(left), 4)

    @staticmethod
    def _same_charge_amount(left: Optional[float], right: Optional[float]) -> bool:
        left_value = 0.0 if left is None else float(left)
        right_value = 0.0 if right is None else float(right)
        return abs(left_value - right_value) < 0.0001

    @staticmethod
    def _charge_delta(left: Optional[float], right: Optional[float]) -> float:
        left_value = 0.0 if left is None else float(left)
        right_value = 0.0 if right is None else float(right)
        return round(right_value - left_value, 4)

    @staticmethod
    def _display_optional_number(value: Optional[float]) -> Any:
        return "-" if value is None else value

    def _write_simple_tc_display_row(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        row_idx: int,
        values: List[Any],
        thin_border: Border,
        fill: PatternFill,
        wrap_columns: Set[int],
    ) -> None:
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.fill = fill
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=col_idx in wrap_columns)

    def _write_table_header(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        row_idx: int,
        headers: List[str],
        thin_border: Border,
    ) -> None:
        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFFFF", size=11)
            cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

    def _write_tc_table_row(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        row_idx: int,
        values: List[Any],
        status: str,
        status_fills: Dict[str, PatternFill],
        thin_border: Border,
        wrap_columns: Set[int],
    ) -> None:
        for col_idx, value in enumerate(values, 1):
            cell = ws.cell(row=row_idx, column=col_idx, value=value)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=col_idx in wrap_columns)
            if col_idx == 1:
                fill = status_fills.get(status)
                if fill:
                    cell.fill = fill
                cell.font = Font(bold=True)

    def _write_packing_list_comparison_sheet(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        rows: List[PackingComparisonRow],
    ) -> Dict[str, int]:
        headers = [
            "Check Status",
            "Issue Detail",
            "Invoice No",
            "Packing Invoice No",
            "Invoice Date",
            "Ex-Factory Date",
            "PO No",
            "Article No",
            "Style / Working No",
            "Invoice QTY",
            "Packing QTY",
            "Packing Cartons",
            "Market PO",
            "Source Invoice",
        ]
        thin_border = create_thin_border()
        title_end_col = openpyxl.utils.get_column_letter(len(headers))
        ws.merge_cells(f"A1:{title_end_col}1")
        title_cell = ws["A1"]
        title_cell.value = f"Packing List 核对结果 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(size=14, bold=True, color="FF1F2937")
        title_cell.alignment = Alignment(horizontal="left", vertical="center")
        ws.row_dimensions[1].height = 28

        for col_idx, header in enumerate(headers, 1):
            cell = ws.cell(row=2, column=col_idx, value=header)
            cell.font = Font(bold=True, color="FFFFFFFF", size=11)
            cell.fill = PatternFill(start_color="FF4472C4", end_color="FF4472C4", fill_type="solid")
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
            cell.border = thin_border

        matched_count = 0
        issue_count = 0
        status_fills = {
            "一致": PatternFill(start_color="FFDDFFDD", end_color="FFDDFFDD", fill_type="solid"),
            "需核对": PatternFill(start_color="FFFFDDDD", end_color="FFFFDDDD", fill_type="solid"),
            "缺失": PatternFill(start_color="FFFFDDDD", end_color="FFFFDDDD", fill_type="solid"),
        }
        for row_idx, row in enumerate(rows, 3):
            values = [
                row.status,
                row.issue_detail,
                row.invoice_number,
                row.packing_invoice_number,
                row.invoice_date,
                row.ex_factory_date,
                row.po_number,
                row.article,
                row.style,
                row.invoice_quantity,
                row.packing_quantity,
                row.packing_cartons,
                row.market_po,
                row.source_invoice,
            ]
            if row.status == "一致":
                matched_count += 1
            else:
                issue_count += 1
            for col_idx, value in enumerate(values, 1):
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="left", vertical="center", wrap_text=(col_idx == 2))
                if col_idx == 1:
                    fill = status_fills.get(row.status)
                    if fill:
                        cell.fill = fill
                    cell.font = Font(bold=True)

        self._adjust_column_widths(
            ws,
            {
                1: 16,
                2: 52,
                3: 18,
                4: 20,
                5: 14,
                6: 16,
                7: 16,
                8: 14,
                9: 20,
                14: 42,
            },
            {
                1: 12,
                2: 24,
                3: 14,
                4: 14,
                5: 12,
                6: 12,
                7: 12,
                8: 12,
                9: 14,
            },
        )
        ws.freeze_panes = "A3"
        return {
            "packing_count": len(rows),
            "packing_matched_count": matched_count,
            "packing_issue_count": issue_count,
        }

    def _write_compact_summary_sheet(self,
                                     ws_summary: openpyxl.worksheet.worksheet.Worksheet,
                                     all_invoice_data: Dict[Tuple[str, str], Dict[str, float]],
                                     invoice_file_names: List[str],
                                     invoice_file_paths: List[str],
                                     ref_df: pd.DataFrame,
                                     thin_border: Border,
                                     packing_confirmed_invoice_keys: Optional[Set[Tuple[str, str, str, int]]] = None) -> Dict[str, Any]:
        """写入窄版汇总表：每张发票命中的价格用纵向明细行展示。"""
        summary_border = Border(
            left=Side(style='thin', color='FFD9E2EC'),
            right=Side(style='thin', color='FFD9E2EC'),
            top=Side(style='thin', color='FFD9E2EC'),
            bottom=Side(style='thin', color='FFD9E2EC')
        )
        no_fill = PatternFill(fill_type=None)
        row_fill_style = PatternFill(start_color='FFF7F9FC', end_color='FFF7F9FC', fill_type='solid')

        header_cells = [
            '款式号',
            '款号',
            '参考价',
            '来源发票',
            '发票价',
            '状态',
            '异常文件',
        ] + DIAGNOSTIC_COLUMNS

        title_end_col = openpyxl.utils.get_column_letter(len(header_cells))
        ws_summary.merge_cells(f'A1:{title_end_col}1')
        title_cell = ws_summary['A1']
        title_cell.value = f"发票价格核对汇总表 - {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
        title_cell.font = Font(size=14, bold=True, color='FF1F2937')
        title_cell.fill = no_fill
        title_cell.alignment = Alignment(horizontal='left', vertical='center')
        ws_summary.row_dimensions[1].height = 28
        for col_idx in range(1, len(header_cells) + 1):
            cell = ws_summary.cell(row=1, column=col_idx)
            cell.border = summary_border
            cell.fill = no_fill

        header_row = 2
        ws_summary.row_dimensions[header_row].height = 24
        for col_idx, header_text in enumerate(header_cells, 1):
            cell = ws_summary.cell(row=header_row, column=col_idx, value=header_text)
            cell.font = Font(bold=True, color='FF1F2937', size=11)
            cell.fill = no_fill
            cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
            cell.border = summary_border

        diagnostics = self._build_invoice_diagnostics(
            all_invoice_data,
            ref_df,
            packing_confirmed_invoice_keys=packing_confirmed_invoice_keys,
        )
        diagnostic_lookup = {
            (diagnostic.article, diagnostic.style, diagnostic.invoice_file): diagnostic
            for diagnostic in diagnostics
        }
        missing_from_ref_count = sum(
            1 for diagnostic in diagnostics
            if diagnostic.status == STATUS_REFERENCE_MISSING
        )
        data_start_row = header_row + 1
        if missing_from_ref_count > 0:
            data_start_row = header_row + 2
            ws_summary.row_dimensions[header_row + 1].height = 24
            ws_summary.merge_cells(f'A{header_row+1}:{title_end_col}{header_row+1}')
            tip_cell = ws_summary.cell(row=header_row+1, column=1, value=f"提示：状态为「参考表未找到」的商品需补入参考表，共 {missing_from_ref_count} 个。")
            tip_cell.font = Font(bold=True, color='FFB45309', size=11)
            tip_cell.fill = PatternFill(start_color='FFFFF7D6', end_color='FFFFF7D6', fill_type='solid')
            tip_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=False)
            for col_idx in range(1, len(header_cells) + 1):
                ws_summary.cell(row=header_row + 1, column=col_idx).border = summary_border

        invoice_path_lookup = dict(zip(invoice_file_names, invoice_file_paths))
        data_row = data_start_row
        data_count = 0

        for (article, style), invoice_prices_dict in all_invoice_data.items():
            rows_to_write = [
                (invoice_name, invoice_prices_dict[invoice_name])
                for invoice_name in invoice_file_names
                if invoice_name in invoice_prices_dict
            ]

            if not rows_to_write:
                rows_to_write = [('-', None)]

            for invoice_name, invoice_price in rows_to_write:
                diagnostic = diagnostic_lookup.get((article, style, invoice_name))
                is_odd_row = (data_count % 2 == 0)
                row_fill = row_fill_style if is_odd_row else None
                ref_price = diagnostic.suggested_price if diagnostic else None
                status = diagnostic.status if diagnostic else STATUS_REFERENCE_MISSING
                abnormal_file = '-'
                if status != STATUS_MATCHED:
                    abnormal_file = (
                        "参考表"
                        if self._is_reference_source_status(status)
                        else invoice_path_lookup.get(invoice_name, invoice_name)
                    )

                ws_summary.row_dimensions[data_row].height = 24
                values = [
                    style,
                    article,
                    ref_price if ref_price is not None else '未找到',
                    invoice_name,
                    invoice_price if invoice_price is not None else '-',
                    status,
                    abnormal_file,
                    diagnostic.suspected_field if diagnostic else '',
                    diagnostic.message if diagnostic else '发票记录未能匹配参考表。',
                    diagnostic.suggested_style if diagnostic else '',
                    diagnostic.suggested_article if diagnostic else '',
                    diagnostic.suggested_price if diagnostic and diagnostic.suggested_price is not None else '',
                ]

                for col_idx, value in enumerate(values, 1):
                    cell = ws_summary.cell(row=data_row, column=col_idx, value=value)
                    cell.border = summary_border
                    cell.alignment = Alignment(horizontal='center', vertical='center', wrap_text=False)
                    if col_idx in (3, 5, 12) and isinstance(value, (int, float)):
                        cell.number_format = '0.00'
                    if row_fill:
                        cell.fill = row_fill
                    if col_idx >= 8:
                        cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)

                invoice_cell = ws_summary.cell(row=data_row, column=4)
                invoice_cell.fill = no_fill
                invoice_cell.font = Font(color='FF1F2937')

                price_cell = ws_summary.cell(row=data_row, column=5)
                status_cell = ws_summary.cell(row=data_row, column=6)
                file_cell = ws_summary.cell(row=data_row, column=7)
                status_fill = self._status_fill(status)
                status_font = self._status_font(status)

                if status_fill:
                    status_cell.fill = status_fill
                if status_font:
                    status_cell.font = status_font
                if status != STATUS_MATCHED:
                    file_cell.value = abnormal_file
                    file_cell.alignment = Alignment(horizontal='left', vertical='center', wrap_text=True)
                    file_cell.font = Font(size=9, color='FF374151')
                    ws_summary.row_dimensions[data_row].height = 38
                else:
                    file_cell.value = '-'

                if status not in (STATUS_MATCHED, STATUS_REFERENCE_MISSING):
                    price_cell.font = Font(color='FFFF0000', bold=True)

                data_count += 1
                data_row += 1

        if data_count > 0:
            ws_summary.auto_filter.ref = f"A{header_row}:{title_end_col}{data_row - 1}"
        ws_summary.freeze_panes = ws_summary.cell(row=data_start_row, column=1)

        return {
            'missing_count': missing_from_ref_count,
            'data_count': data_count,
            'diagnostics': self._count_diagnostics(diagnostics)
        }

    @staticmethod
    def _estimate_text_width(value: Any) -> int:
        if value is None:
            return 0
        return sum(2 if ord(char) > 255 else 1 for char in str(value))

    def _adjust_column_widths(self, ws: openpyxl.worksheet.worksheet.Worksheet,
                             max_widths: Optional[Dict[int, int]] = None,
                             min_widths: Optional[Dict[int, int]] = None) -> None:
        """自动调整Excel列宽"""
        if max_widths is None:
            max_widths = {}
        if min_widths is None:
            min_widths = {}

        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                try:
                    max_length = max(max_length, self._estimate_text_width(cell.value))
                except Exception:
                    pass
            min_width = min_widths.get(col_idx, 9)
            max_width = max_widths.get(col_idx, 38)
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    def process_invoices(
        self,
        invoice_paths: List[str],
        ref_path: str,
        output_dir: str = None,
        packing_path: Optional[str] = None,
        packing_paths: Optional[List[str]] = None,
        tc_invoice_path: Optional[str] = None,
        tc_invoice_paths: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """主处理流程"""

        result = {
            'success': False,
            'message': '',
            'invoice_count': len(invoice_paths),
            'total_items': 0,
            'matches': {'一致': 0, '不一致': 0, '未找到': 0},
            'diagnostics': {},
            'tc_count': 0,
            'tc_matched_count': 0,
            'tc_issue_count': 0,
            'tc_summary_count': 0,
            'tc_summary_issue_count': 0,
            'tc_total_issue_count': 0,
            'output_path': None,
            'logs': []
        }

        def log(msg: str):
            result['logs'].append(msg)

        try:
            log("=" * 80)
            log("开始批量核对")
            log("=" * 80)

            # 读取参考表
            log("\n📖 正在读取参考表...")
            ref_df = pd.read_excel(ref_path)
            log(f"✅ 参考表读取完成，共 {len(ref_df)} 行数据")

            all_invoice_data: Dict[Tuple[str, str], Dict[str, float]] = {}
            all_invoice_records: List[InvoiceRecord] = []
            invoice_summaries: List[InvoiceSummaryRecord] = []
            invoice_file_names: List[str] = []
            invoice_file_paths: List[str] = []

            log(f"\n{'='*80}")
            log(f"开始处理 {len(invoice_paths)} 张发票...")
            log(f"{'='*80}")

            total_items_count = 0

            for idx, invoice_path in enumerate(invoice_paths, 1):
                invoice_filename = os.path.basename(invoice_path)
                invoice_file_names.append(invoice_filename)
                invoice_file_paths.append(invoice_path)

                log(f"\n[{idx}/{len(invoice_paths)}] 处理发票：{invoice_filename}")

                try:
                    invoice_records = self.read_invoice_records(invoice_path)
                    invoice_summaries.append(self.read_invoice_summary(invoice_path, invoice_records))
                    invoice_data = [
                        {
                            'article': record.article,
                            'style': record.style,
                            'price': record.price,
                        }
                        for record in invoice_records
                    ]

                    if len(invoice_data) == 0:
                        log(f"⚠️ 未在发票中识别到有效商品数据，跳过")
                        continue

                    log(f"✅ 成功读取 {len(invoice_data)} 个商品")
                    total_items_count += len(invoice_data)
                    all_invoice_records.extend(invoice_records)

                    for item in invoice_data:
                        key = (item['article'], item['style'])
                        if key not in all_invoice_data:
                            all_invoice_data[key] = {}

                        all_invoice_data[key][invoice_filename] = item['price']

                        if len(invoice_data) <= 10:
                            log(f"  - {item['article']}/{item['style']}: {item['price']:.2f}")

                    if len(invoice_data) > 10:
                        log(f"  ... (还有 {len(invoice_data)-10} 个商品)")

                except Exception as e:
                    log(f"❌ 处理发票 {invoice_filename} 时出错：{str(e)}")
                    import traceback
                    log(traceback.format_exc())

            log(f"\n{'='*80}")
            log(f"开始合并结果到参考表...")
            log(f"{'='*80}")

            result_df, matches = self.update_reference_table(
                ref_df,
                all_invoice_data
            )

            result['matches'] = matches
            result['total_items'] = total_items_count

            log("\n💾 正在保存结果（含汇总表）...")

            if output_dir is None:
                output_dir = os.path.dirname(ref_path)

            ensure_dir(output_dir)
            tc_comparison_rows: Optional[List[TcComparisonRow]] = None
            tc_summary_rows: Optional[List[TcSummaryComparisonRow]] = None
            tc_confirmed_invoice_keys: Set[Tuple[str, str, str, int]] = set()
            effective_tc_paths = tc_invoice_paths if tc_invoice_paths is not None else (
                [tc_invoice_path] if tc_invoice_path else []
            )
            if not effective_tc_paths:
                effective_tc_paths = packing_paths if packing_paths is not None else (
                    [packing_path] if packing_path else []
                )
            if effective_tc_paths:
                log(f"\n📄 正在读取 {len(effective_tc_paths)} 个 TC INV PDF...")
                tc_records: List[TcInvoiceRecord] = []
                tc_summaries: List[TcInvoiceSummary] = []
                for index, current_tc_path in enumerate(effective_tc_paths, start=1):
                    tc_filename = os.path.basename(current_tc_path)
                    log(f"[{index}/{len(effective_tc_paths)}] 读取 TC INV：{tc_filename}")
                    current_tc_records = self.read_tc_invoice_records(current_tc_path)
                    tc_records.extend(current_tc_records)
                    tc_summaries.append(self.read_tc_invoice_summary(current_tc_path, current_tc_records))
                tc_comparison_rows = self.build_tc_invoice_comparison(
                    all_invoice_records,
                    tc_records,
                )
                tc_summary_rows = self.build_tc_invoice_summary_comparison(
                    invoice_summaries,
                    tc_summaries,
                )
                tc_confirmed_invoice_keys = self._build_tc_confirmed_invoice_keys(
                    all_invoice_records,
                    tc_records,
                )
                result['tc_count'] = len(tc_comparison_rows)
                result['tc_matched_count'] = sum(1 for row in tc_comparison_rows if row.status == "一致")
                result['tc_issue_count'] = result['tc_count'] - result['tc_matched_count']
                result['tc_summary_count'] = len(tc_summary_rows)
                result['tc_summary_issue_count'] = sum(1 for row in tc_summary_rows if row.status != "一致")
                result['tc_total_issue_count'] = result['tc_issue_count'] + result['tc_summary_issue_count']
                log(
                    "✅ TC INV 核对完成："
                    f"明细 {result['tc_count']} 行，"
                    f"汇总 {result['tc_summary_count']} 行，"
                    f"总异常 {result['tc_total_issue_count']} 条"
                )

            if tc_confirmed_invoice_keys:
                result_df, matches = self.update_reference_table(
                    ref_df,
                    all_invoice_data,
                    packing_confirmed_invoice_keys=tc_confirmed_invoice_keys,
                )
                result['matches'] = matches

            default_name = (
                os.path.splitext(os.path.basename(ref_path))[0] +
                "_批量核对结果_" +
                datetime.now().strftime('%Y%m%d_%H%M%S') +
                ".xlsx"
            )

            output_path = os.path.join(output_dir, default_name)

            save_result = self.save_excel_with_summary(
                result_df,
                all_invoice_data,
                invoice_file_names,
                invoice_file_paths,
                ref_df,
                output_path,
                tc_comparison_rows=tc_comparison_rows,
                tc_summary_rows=tc_summary_rows,
                tc_confirmed_invoice_keys=tc_confirmed_invoice_keys,
            )

            result['output_path'] = output_path
            result['save_result'] = save_result
            result['diagnostics'] = save_result.get('diagnostics', {})
            result['tc_count'] = save_result.get('tc_count', result['tc_count'])
            result['tc_matched_count'] = save_result.get('tc_matched_count', result['tc_matched_count'])
            result['tc_issue_count'] = save_result.get('tc_issue_count', result['tc_issue_count'])
            result['tc_summary_count'] = save_result.get('tc_summary_count', result['tc_summary_count'])
            result['tc_summary_issue_count'] = save_result.get(
                'tc_summary_issue_count',
                result['tc_summary_issue_count'],
            )
            result['tc_total_issue_count'] = save_result.get(
                'tc_total_issue_count',
                result['tc_total_issue_count'],
            )

            log(f"\n{'='*80}")
            log(f"✅ 批量核对完成！")
            log(f"{'='*80}")
            log(f"处理发票数：{len(invoice_paths)}")
            log(f"商品总数：{total_items_count}")
            log(f"价格一致：{matches['一致']}")
            log(f"价格不一致：{matches['不一致']}")
            log(f"未找到：{matches['未找到']}")
            log(f"{'='*80}")
            log(f"结果文件：{output_path}")

            result['success'] = True
            result['message'] = '批量核对完成'

        except Exception as e:
            log(f"\n❌ 错误：{str(e)}")
            import traceback
            log(traceback.format_exc())
            result['message'] = f'处理出错：{str(e)}'

        return result

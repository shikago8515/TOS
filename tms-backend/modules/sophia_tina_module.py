
# -*- coding: utf-8 -*-
"""
Sophia & Tina 报表生成模块
从 TMS工具_20260518_2100.pyw 提取的核心逻辑
创建时间: 2026-05-18
"""

import os
import openpyxl
import pandas as pd
import tempfile
import zipfile
from collections import defaultdict
from datetime import datetime
from openpyxl.styles import Font, PatternFill, Alignment, Border
from openpyxl.worksheet.table import Table, TableStyleInfo
from openpyxl.utils.datetime import to_excel
from typing import List, Dict, Any, Optional, Tuple, Set
from xml.etree import ElementTree
from xml.sax.saxutils import escape

# 导入工具模块
import sys
sys.path.append(os.path.join(os.path.dirname(__file__), '..'))
from utils.file_utils import ensure_dir, create_thin_border


PIVOT_TEMPLATE_FILENAME = "sophia_tina_pivot_template.xlsx"


def _unique_existing_order(paths: List[str]) -> List[str]:
    seen: Set[str] = set()
    unique_paths: List[str] = []
    for path in paths:
        normalized = os.path.abspath(path)
        if normalized not in seen:
            seen.add(normalized)
            unique_paths.append(normalized)
    return unique_paths


def _candidate_pivot_template_paths() -> List[str]:
    paths: List[str] = []
    override_path = os.environ.get("TOS_SOPHIA_TINA_TEMPLATE_PATH")
    if override_path:
        paths.append(override_path)

    module_dir = os.path.dirname(__file__)
    paths.append(os.path.join(module_dir, "..", "templates", PIVOT_TEMPLATE_FILENAME))

    runtime_root = getattr(sys, "_MEIPASS", None)
    if runtime_root:
        paths.append(os.path.join(runtime_root, "templates", PIVOT_TEMPLATE_FILENAME))

    if getattr(sys, "frozen", False):
        paths.append(os.path.join(os.path.dirname(sys.executable), "_internal", "templates", PIVOT_TEMPLATE_FILENAME))

    return _unique_existing_order(paths)


def _resolve_pivot_template_path() -> str:
    candidates = _candidate_pivot_template_paths()
    for candidate in candidates:
        if os.path.exists(candidate):
            return candidate
    return candidates[0]


class SophiaTinaModule:
    """Sophia & Tina 报表生成业务逻辑"""

    RESULT_HEADERS = [
        'Pack',
        'Season',
        'Factory',
        'Working Number',
        'Article Number',
        'Article Name',
        'CRD',
        'PODD',
        'Gps Customer Number',
        'Country/Region',
        'Shipment Method',
        'Marketing Forecast(M)',
        'Quantity',
        'Factory Price(USD)',
        'Factory Amount(USD)',
        'TMS Price(USD)',
        'TMS Amount(USD)',
    ]

    ARTICLE_PRIORITY = {
        'FINAL': 0,
        'P2': 1,
        'P1': 2,
        'PREC': 3,
    }

    OOXML_MAIN_NS = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    OOXML_REL_NS = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    OOXML_MC_NS = "http://schemas.openxmlformats.org/markup-compatibility/2006"
    OOXML_X14_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"
    OOXML_X14AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2009/9/ac"
    OOXML_X15_NS = "http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"
    OOXML_X15AC_NS = "http://schemas.microsoft.com/office/spreadsheetml/2010/11/ac"
    OOXML_XML_NS = "http://www.w3.org/XML/1998/namespace"
    OOXML_DRAWING_MAIN_NS = "http://schemas.openxmlformats.org/drawingml/2006/main"
    OOXML_CHART_NS = "http://schemas.openxmlformats.org/drawingml/2006/chart"
    OOXML_CHART_2007_NS = "http://schemas.microsoft.com/office/drawing/2007/8/2/chart"
    OOXML_SPREADSHEET_DRAWING_NS = "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"
    OOXML_DRAWING_SLICER_2010_NS = "http://schemas.microsoft.com/office/drawing/2010/slicer"
    OPC_CONTENT_TYPES_NS = "http://schemas.openxmlformats.org/package/2006/content-types"
    OPC_RELATIONSHIPS_NS = "http://schemas.openxmlformats.org/package/2006/relationships"
    COUNTRY_SOURCE_SHEET = "Country Analysis Source"
    SHIP_METHOD_SOURCE_SHEET = "Ship Method Source"
    S2S_SOURCE_SHEET = "S2S Development Source"
    COUNTRY_SUMMARY_BUCKETS = [
        ("CHINA", "China Order"),
        ("UNITED STATES", "USA Order"),
        ("OTHER COUNTRIES", "Other Order"),
    ]
    COUNTRY_SUMMARY_HEADER_STYLE = 25
    COUNTRY_SUMMARY_TEXT_STYLE = 9
    COUNTRY_SUMMARY_PERCENT_STYLE = 10
    COUNTRY_SUMMARY_CURRENCY_STYLE = 11
    HELPER_WORKSHEETS = [
        ("xl/worksheets/sheet13.xml", COUNTRY_SOURCE_SHEET, "rId22", "18"),
        ("xl/worksheets/sheet14.xml", SHIP_METHOD_SOURCE_SHEET, "rId23", "19"),
        ("xl/worksheets/sheet15.xml", S2S_SOURCE_SHEET, "rId24", "20"),
    ]
    COUNTRY_PIVOT_CACHE_ID = "500"
    S2S_PIVOT_CACHE_ID = "501"
    Y2Y_RIGHT_PIVOT_LOCATION_REF = "I20:M34"
    Y2Y_AMOUNT_FIELD_NAMES = {"TMS Amount (USD)", "Y2Y - Amount"}
    SHIP_METHOD_TOTAL_FILL_RGB = "FFD9D9D9"
    PIVOT_TEMPLATE_PATH = os.path.abspath(
        os.path.join(os.path.dirname(__file__), "..", "templates", PIVOT_TEMPLATE_FILENAME)
    )

    def __init__(self):
        pass

    def read_excel_files(self, file_list: List[str]) -> Optional[pd.DataFrame]:
        """读取多个Excel文件并合并数据"""
        dfs: List[pd.DataFrame] = []
        logs: List[str] = []

        for file_path in file_list:
            try:
                df = pd.read_excel(file_path)
                dfs.append(df)
                logs.append(f"  - 读取成功：{os.path.basename(file_path)}，共 {len(df)} 行")
            except Exception as e:
                logs.append(f"  ⚠️ 读取失败：{os.path.basename(file_path)}，错误：{str(e)}")

        if len(dfs) == 0:
            return None, logs

        if len(dfs) == 1:
            return dfs[0], logs

        # 合并多个DataFrame
        result = pd.concat(dfs, axis=0, ignore_index=True)
        logs.append(f"  ✅ 合并成功，总计 {len(result)} 行")

        return result, logs

    def save_st_result(
        self,
        results: List[Dict[str, Any]],
        output_path: str,
        diagnostics: Optional[List[Dict[str, Any]]] = None,
        development_style_rows: Optional[List[Dict[str, Any]]] = None,
    ) -> None:
        """保存 Sophia & Tina 结果报表"""

        template_path = _resolve_pivot_template_path()
        if not os.path.exists(template_path):
            raise FileNotFoundError(f"Sophia/Tina pivot template not found: {template_path}")

        self._save_st_result_from_pivot_template(
            results,
            output_path,
            diagnostics or [],
            development_style_rows or [],
            template_path,
        )
        return

        wb = openpyxl.Workbook()
        thin_border = create_thin_border()
        self._write_summary_sheets(wb, results, development_style_rows or [], thin_border)

        ws = wb.create_sheet("Result")
        header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        odd_row_fill = PatternFill(start_color='FFF2F2F2', end_color='FFF2F2F2', fill_type='solid')
        conflict_fill = PatternFill(start_color='FFFFF2CC', end_color='FFFFF2CC', fill_type='solid')
        missing_fill = PatternFill(start_color='FFFFE5E5', end_color='FFFFE5E5', fill_type='solid')

        ws.row_dimensions[1].height = 30
        for col_idx, header in enumerate(self.RESULT_HEADERS, 1):
            cell = ws.cell(row=1, column=col_idx, value=header)
            cell.font = Font(bold=True, color='FFFFFFFF', size=11)
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal='center', vertical='center')
            cell.border = thin_border

        for row_idx, result in enumerate(results, 2):
            ws.row_dimensions[row_idx].height = 20
            is_odd_row = ((row_idx - 2) % 2 == 0)
            row_fill = odd_row_fill if is_odd_row else None

            for col_idx, header in enumerate(self.RESULT_HEADERS, 1):
                if header == 'Factory Amount(USD)':
                    value = f"=N{row_idx}*M{row_idx}"
                elif header == 'TMS Amount(USD)':
                    value = f"=P{row_idx}*M{row_idx}"
                else:
                    value = result.get(header)
                ws.cell(row=row_idx, column=col_idx, value=value)

            for col_idx in range(1, len(self.RESULT_HEADERS) + 1):
                cell = ws.cell(row=row_idx, column=col_idx)
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if row_fill:
                    cell.fill = row_fill

            if result.get('_factory_price_conflict'):
                ws.cell(row=row_idx, column=14).fill = conflict_fill
                ws.cell(row=row_idx, column=15).fill = conflict_fill
            if result.get('_missing_factory_price'):
                ws.cell(row=row_idx, column=14).fill = missing_fill
                ws.cell(row=row_idx, column=15).fill = missing_fill
            if result.get('_missing_tms_price'):
                ws.cell(row=row_idx, column=16).fill = missing_fill
                ws.cell(row=row_idx, column=17).fill = missing_fill
            if result.get('_missing_article'):
                for col_idx in (2, 12, 16, 17):
                    ws.cell(row=row_idx, column=col_idx).fill = missing_fill
            if result.get('_missing_pack') or result.get('_ambiguous_pack'):
                ws.cell(row=row_idx, column=1).fill = missing_fill if result.get('_missing_pack') else conflict_fill

        ws.freeze_panes = "A2"
        if ws.max_row >= 1:
            table_ref = f"A1:{openpyxl.utils.get_column_letter(len(self.RESULT_HEADERS))}{max(ws.max_row, 1)}"
            table = Table(displayName="Table1", ref=table_ref)
            table.tableStyleInfo = TableStyleInfo(
                name="TableStyleMedium2",
                showFirstColumn=False,
                showLastColumn=False,
                showRowStripes=True,
                showColumnStripes=False,
            )
            ws.add_table(table)

        for row_idx in range(2, ws.max_row + 1):
            for col_idx in (7, 8):
                ws.cell(row=row_idx, column=col_idx).number_format = 'yyyy-mm-dd'
            ws.cell(row=row_idx, column=13).number_format = '#,##0'
            for col_idx in (14, 15, 16, 17):
                ws.cell(row=row_idx, column=col_idx).number_format = '$#,##0.00'

        # 自动调整列宽，避免业务侧打开后需要手动拉宽。
        self._adjust_column_widths(ws, min_widths={
            1: 14,
            2: 12,
            3: 12,
            4: 18,
            5: 14,
            6: 30,
            7: 13,
            8: 13,
            9: 18,
            10: 18,
            11: 18,
            12: 22,
            13: 12,
            14: 18,
            15: 20,
            16: 16,
            17: 18,
        })

        self._write_rules_sheet(wb, thin_border)
        self._write_diagnostics_sheet(wb, diagnostics or [], thin_border)

        wb.save(output_path)

    def _save_st_result_from_pivot_template(
        self,
        results: List[Dict[str, Any]],
        output_path: str,
        diagnostics: List[Dict[str, Any]],
        development_style_rows: List[Dict[str, Any]],
        template_path: str,
    ) -> None:
        """Write editable Result data into the native PivotTable template."""
        table_ref = f"A1:Q{max(len(results) + 1, 1)}"
        replacements = self._build_pivot_template_replacements(
            results,
            diagnostics,
            development_style_rows,
            table_ref,
            template_path,
        )
        skip_entries = {"xl/calcChain.xml"}
        self._replace_xlsx_entries(template_path, output_path, replacements, skip_entries)

    def _build_pivot_template_replacements(
        self,
        results: List[Dict[str, Any]],
        diagnostics: List[Dict[str, Any]],
        development_style_rows: List[Dict[str, Any]],
        table_ref: str,
        template_path: str,
    ) -> Dict[str, bytes]:
        with zipfile.ZipFile(template_path, "r") as archive:
            styles_xml, style_ids = self._updated_styles_xml(archive.read("xl/styles.xml"))
            country_rows = self._country_analysis_source_rows(results)
            ship_method_rows = self._ship_method_source_rows(results)
            s2s_rows = self._s2s_development_source_rows(results, development_style_rows)
            s2s_display_rows = self._s2s_development_analysis_display_rows(s2s_rows)
            country_ref = f"A1:J{max(len(country_rows) + 1, 1)}"
            ship_method_ref = f"A1:H{max(len(ship_method_rows) + 1, 1)}"
            s2s_ref = f"A1:E{max(len(s2s_rows) + 1, 1)}"
            s2s_pivot_ref = f"A1:E{max(len(s2s_display_rows) + 1, 1)}"

            replacements: Dict[str, bytes] = {
                "xl/styles.xml": styles_xml,
                "xl/worksheets/sheet3.xml": self._updated_country_summary_sheet_xml(
                    archive.read("xl/worksheets/sheet3.xml"),
                    results,
                    1,
                    "Season",
                ),
                "xl/worksheets/sheet4.xml": self._updated_country_summary_sheet_xml(
                    archive.read("xl/worksheets/sheet4.xml"),
                    results,
                    1,
                    "Year",
                ),
                "xl/worksheets/sheet7.xml": self._ship_method_analysis_sheet_xml(results, style_ids),
                "xl/worksheets/_rels/sheet3.xml.rels": self._without_pivot_table_relationship(
                    archive.read("xl/worksheets/_rels/sheet3.xml.rels")
                ),
                "xl/worksheets/_rels/sheet4.xml.rels": self._without_pivot_table_relationship(
                    archive.read("xl/worksheets/_rels/sheet4.xml.rels")
                ),
                "xl/worksheets/_rels/sheet7.xml.rels": self._without_pivot_table_relationship(
                    archive.read("xl/worksheets/_rels/sheet7.xml.rels")
                ),
                "xl/worksheets/sheet8.xml": self._development_style_qty_sheet_xml(development_style_rows),
                "xl/worksheets/sheet9.xml": self._with_pivot_table_definition(
                    archive.read("xl/worksheets/sheet9.xml"),
                    "rId1",
                ),
                "xl/worksheets/sheet10.xml": self._result_sheet_xml(results, table_ref, style_ids),
                "xl/worksheets/sheet11.xml": self._rules_sheet_xml(),
                "xl/worksheets/sheet12.xml": self._diagnostics_sheet_xml(diagnostics),
                "xl/worksheets/sheet13.xml": self._simple_sheet_xml(
                    [
                        "Season",
                        "Years",
                        "Factory",
                        "Country Group",
                        "Quantity",
                        "Quantity % (S)",
                        "Quantity % (Y)",
                        "TMS Amount (USD)",
                        "TMS Amount % (S)",
                        "TMS Amount % (Y)",
                    ],
                    country_rows,
                    '<cols><col min="1" max="4" width="20" customWidth="1"/>'
                    '<col min="5" max="10" width="18" customWidth="1"/></cols>',
                ),
                "xl/worksheets/sheet14.xml": self._simple_sheet_xml(
                    [
                        "Factory",
                        "Years",
                        "PODD",
                        "Shipment Method",
                        "Quantity",
                        "Quantity (%)",
                        "TMS Amount (USD)",
                        "TMS Amount (%)",
                    ],
                    ship_method_rows,
                    '<cols><col min="1" max="4" width="18" customWidth="1"/>'
                    '<col min="5" max="8" width="18" customWidth="1"/></cols>',
                ),
                "xl/worksheets/sheet15.xml": self._s2s_development_source_sheet_xml(s2s_rows),
            }
            replacements["xl/tables/table1.xml"] = self._updated_table_xml(
                archive.read("xl/tables/table1.xml"),
                table_ref,
            )
            replacements["xl/workbook.xml"] = self._updated_workbook_xml(archive.read("xl/workbook.xml"))
            replacements["xl/_rels/workbook.xml.rels"] = self._updated_workbook_relationships(
                archive.read("xl/_rels/workbook.xml.rels")
            )
            replacements["[Content_Types].xml"] = self._updated_content_types(
                archive.read("[Content_Types].xml")
            )
            replacements["xl/pivotCache/pivotCacheDefinition3.xml"] = self._pivot_cache_definition_xml(
                self.SHIP_METHOD_SOURCE_SHEET,
                ship_method_ref,
                [
                    ("Factory", "0"),
                    ("Years", "0"),
                    ("PODD", "0"),
                    ("Shipment Method", "0"),
                    ("Quantity", "3"),
                    ("Quantity (%)", "10"),
                    ("TMS Amount (USD)", "166"),
                    ("TMS Amount (%)", "10"),
                ],
            )
            replacements["xl/pivotCache/pivotCacheDefinition5.xml"] = self._pivot_cache_definition_xml(
                self.COUNTRY_SOURCE_SHEET,
                country_ref,
                [
                    ("Season", "0"),
                    ("Years", "0"),
                    ("Factory", "0"),
                    ("Country Group", "0"),
                    ("Quantity", "3"),
                    ("Quantity % (S)", "10"),
                    ("Quantity % (Y)", "10"),
                    ("TMS Amount (USD)", "166"),
                    ("TMS Amount % (S)", "10"),
                    ("TMS Amount % (Y)", "10"),
                ],
            )
            replacements["xl/pivotCache/pivotCacheDefinition6.xml"] = self._pivot_cache_definition_xml(
                self.S2S_SOURCE_SHEET,
                s2s_ref,
                [
                    ("Season", "0"),
                    ("Factory", "0"),
                    ("Development Style Count", "3"),
                    ("Bulk Style Count", "3"),
                    ("Bulk Qty (pcs)", "3"),
                ],
            )
            replacements["xl/pivotCache/pivotCacheRecords5.xml"] = self._empty_pivot_cache_records_xml()
            replacements["xl/pivotCache/pivotCacheRecords6.xml"] = self._empty_pivot_cache_records_xml()
            replacements["xl/pivotCache/_rels/pivotCacheDefinition5.xml.rels"] = self._pivot_cache_relationship_xml(5)
            replacements["xl/pivotCache/_rels/pivotCacheDefinition6.xml.rels"] = self._pivot_cache_relationship_xml(6)
            replacements["xl/pivotTables/pivotTable3.xml"] = self._updated_pivot_table_xml(
                archive.read("xl/pivotTables/pivotTable3.xml"),
                self.COUNTRY_PIVOT_CACHE_ID,
                [0, 2],
                [3, -2],
                [
                    ("Quantity (Y)", 4, None, "3"),
                    ("Quantity Percentage (%)", 5, None, "10"),
                    ("TMS Amount (USD)", 7, None, "166"),
                    ("TMS Amount Percentage (%)", 8, None, "10"),
                ],
                10,
            )
            replacements["xl/pivotTables/pivotTable4.xml"] = self._updated_pivot_table_xml(
                archive.read("xl/pivotTables/pivotTable4.xml"),
                self.COUNTRY_PIVOT_CACHE_ID,
                [1, 2],
                [3, -2],
                [
                    ("Quantity (Y)", 4, None, "3"),
                    ("Quantity (%)", 6, None, "10"),
                    ("TMS Amount (USD)", 7, None, "166"),
                    ("TMS Amount (%)", 9, None, "10"),
                ],
                10,
            )
            replacements["xl/pivotTables/pivotTable8.xml"] = self._updated_pivot_table_xml(
                archive.read("xl/pivotTables/pivotTable8.xml"),
                "267",
                [0, 1, 2, 3],
                [-2],
                [
                    ("Quantity (Y)", 4, None, "3"),
                    ("Quantity (%)", 5, None, "10"),
                    ("TMS Amount (USD)", 6, None, "166"),
                    ("TMS Amount (%)", 7, None, "10"),
                ],
                8,
            )
            replacements["xl/pivotTables/pivotTable9.xml"] = self._updated_pivot_table_xml(
                archive.read("xl/pivotTables/pivotTable9.xml"),
                self.S2S_PIVOT_CACHE_ID,
                [0, 1],
                [-2],
                [
                    ("Development Style Count", 2, None, "3"),
                    ("Bulk Style Count", 3, None, "3"),
                    ("Bulk Qty (pcs)", 4, None, "3"),
                ],
                5,
                location_ref=s2s_pivot_ref,
            )
            replacements["xl/pivotTables/pivotTable5.xml"] = self._y2y_quantity_only_pivot_table_xml(
                archive.read("xl/pivotTables/pivotTable5.xml")
            )
            replacements["xl/charts/chart1.xml"] = self._y2y_quantity_only_chart_xml(
                archive.read("xl/charts/chart1.xml")
            )
            replacements["xl/drawings/drawing1.xml"] = self._normalized_y2y_drawing_xml(
                archive.read("xl/drawings/drawing1.xml")
            )
            replacements["xl/slicerCaches/slicerCache1.xml"] = self._updated_y2y_slicer_cache_xml(
                archive.read("xl/slicerCaches/slicerCache1.xml")
            )

            for name in archive.namelist():
                if name.startswith("xl/pivotCache/pivotCacheDefinition") and name.endswith(".xml"):
                    if name not in replacements:
                        replacements[name] = self._updated_pivot_cache_xml(archive.read(name), table_ref)
                elif name.startswith("xl/pivotCache/pivotCacheRecords") and name.endswith(".xml"):
                    replacements[name] = self._empty_pivot_cache_records_xml()

        return replacements

    @staticmethod
    def _replace_xlsx_entries(
        template_path: str,
        output_path: str,
        replacements: Dict[str, bytes],
        skip_entries: Set[str],
    ) -> None:
        output_dir = os.path.dirname(output_path) or "."
        ensure_dir(output_dir)
        temp_file = tempfile.NamedTemporaryFile(delete=False, suffix=".xlsx", dir=output_dir)
        temp_path = temp_file.name
        temp_file.close()

        try:
            with zipfile.ZipFile(template_path, "r") as source:
                source_names = set(source.namelist())
                with zipfile.ZipFile(temp_path, "w", zipfile.ZIP_DEFLATED) as target:
                    for item in source.infolist():
                        if item.filename in skip_entries:
                            continue
                        data = replacements.get(item.filename)
                        if data is None:
                            data = source.read(item.filename)
                        target.writestr(item, data)

                    for name, data in replacements.items():
                        if name not in source_names and name not in skip_entries:
                            target.writestr(name, data)

            os.replace(temp_path, output_path)
        except Exception:
            if os.path.exists(temp_path):
                os.remove(temp_path)
            raise

    def _result_sheet_xml(
        self,
        results: List[Dict[str, Any]],
        table_ref: str,
        style_ids: Dict[str, int],
    ) -> bytes:
        rows = [self._row_xml(1, [
            self._cell_xml(1, col_idx, header, self._result_header_style(col_idx))
            for col_idx, header in enumerate(self.RESULT_HEADERS, 1)
        ], height="30", custom_height=True)]

        for row_idx, result in enumerate(results, 2):
            cells = []
            for col_idx, header in enumerate(self.RESULT_HEADERS, 1):
                style_id = self._result_data_style(col_idx)
                if header == "Factory Price(USD)" and result.get("_factory_price_conflict"):
                    style_id = style_ids["factory_price_conflict"]
                if header == "Factory Amount(USD)":
                    cells.append(
                        self._cell_xml(
                            row_idx,
                            col_idx,
                            result.get("_factory_amount_value"),
                            style_id,
                            formula=f"N{row_idx}*M{row_idx}",
                        )
                    )
                elif header == "TMS Amount(USD)":
                    cells.append(
                        self._cell_xml(
                            row_idx,
                            col_idx,
                            result.get("_tms_amount_value"),
                            style_id,
                            formula=f"P{row_idx}*M{row_idx}",
                        )
                    )
                elif header in {"CRD", "PODD"}:
                    cells.append(self._date_cell_xml(row_idx, col_idx, result.get(header), style_id))
                else:
                    cells.append(self._cell_xml(row_idx, col_idx, result.get(header), style_id))
            rows.append(self._row_xml(row_idx, cells, height="20.1", custom_height=True))

        dimension = table_ref
        sheet_data = "".join(rows)
        columns_xml = (
            '<cols>'
            '<col min="1" max="1" width="15" customWidth="1"/>'
            '<col min="2" max="3" width="12" customWidth="1"/>'
            '<col min="4" max="4" width="18.42578125" customWidth="1"/>'
            '<col min="5" max="5" width="16.7109375" customWidth="1"/>'
            '<col min="6" max="6" width="31" customWidth="1"/>'
            '<col min="7" max="8" width="21" customWidth="1"/>'
            '<col min="9" max="9" width="23.28515625" customWidth="1"/>'
            '<col min="10" max="11" width="22" customWidth="1"/>'
            '<col min="12" max="12" width="23.42578125" customWidth="1"/>'
            '<col min="13" max="13" width="12" customWidth="1"/>'
            '<col min="14" max="17" width="18" customWidth="1"/>'
            '</cols>'
        )
        return self._worksheet_xml(
            dimension,
            sheet_data,
            columns_xml,
            freeze_panes=True,
            table_rel_id="rId2",
        )

    def _rules_sheet_xml(self) -> bytes:
        rows = [
            ["Field", "Source / Rule"],
            ["Pack", "Factory Price Pack; match by Working Number + Season, then matched Factory Price record Pack, with Working Number fallback when unique."],
            ["Season", "TMS Price Season (M), matched by Working Number (M) + Article Number (A)."],
            ["Factory", "Allocation Factory PO -> Allocation overrides TMS Factory when provided."],
            ["Working Number", "Released / Unreleased Working Number."],
            ["Article Number", "Released / Unreleased Article Number."],
            ["Article Name", "Released / Unreleased Article Description."],
            ["CRD", "Released / Unreleased Customer Request Date (CRD)."],
            ["PODD", "Shipment Method po_delivery_date when split; otherwise Released / Unreleased PODD."],
            ["Gps Customer Number", "Released / Unreleased Gps Customer Number."],
            ["Country/Region", "Released / Unreleased Country/Region."],
            ["Shipment Method", "Shipment Method shippinginstruction_desc when split; otherwise TMS shipment field."],
            ["Marketing Forecast(M)", "TMS Price Marketing Forecast (M)."],
            ["Quantity", "Shipment Method pord_order_qty when single-combo split; otherwise TMS Ordered Quantity."],
            ["Factory Price(USD)", "Factory Price by Season + Working Number + Factory, fallback to Working Number + Factory; when multiple prices exist, keep the first matched value and highlight it."],
            ["Factory Amount(USD)", "Excel formula: Factory Price(USD) * Quantity."],
            ["TMS Price(USD)", "TMS Price Intl. FOB (C), priority Final > P2 > P1 > PREC; Factory Price TMS Price fallback."],
            ["TMS Amount(USD)", "Excel formula: TMS Price(USD) * Quantity."],
        ]
        sheet_rows = []
        for row_idx, row in enumerate(rows, 1):
            style_id = 4 if row_idx == 1 else 5
            cells = [self._cell_xml(row_idx, col_idx, value, style_id) for col_idx, value in enumerate(row, 1)]
            sheet_rows.append(self._row_xml(row_idx, cells))

        return self._worksheet_xml(
            f"A1:B{len(rows)}",
            "".join(sheet_rows),
            '<cols><col min="1" max="1" width="24" customWidth="1"/>'
            '<col min="2" max="2" width="96" customWidth="1"/></cols>',
            freeze_panes=True,
        )

    def _diagnostics_sheet_xml(self, diagnostics: List[Dict[str, Any]]) -> bytes:
        rows: List[List[Any]] = [["Level", "Type", "Key", "Message"]]
        if diagnostics:
            rows.extend(
                [
                    [
                        item.get("level", "WARN"),
                        item.get("type", ""),
                        item.get("key", ""),
                        item.get("message", ""),
                    ]
                    for item in diagnostics
                ]
            )
        else:
            rows.append(["INFO", "OK", "", "No diagnostics."])

        sheet_rows = []
        for row_idx, row in enumerate(rows, 1):
            style_id = 4 if row_idx == 1 else 5
            cells = [self._cell_xml(row_idx, col_idx, value, style_id) for col_idx, value in enumerate(row, 1)]
            sheet_rows.append(self._row_xml(row_idx, cells, height="30" if row_idx > 1 else None))

        return self._worksheet_xml(
            f"A1:D{len(rows)}",
            "".join(sheet_rows),
            '<cols><col min="1" max="1" width="12" customWidth="1"/>'
            '<col min="2" max="2" width="24" customWidth="1"/>'
            '<col min="3" max="3" width="42" customWidth="1"/>'
            '<col min="4" max="4" width="90" customWidth="1"/></cols>',
            freeze_panes=True,
        )

    def _development_style_qty_sheet_xml(self, development_style_rows: List[Dict[str, Any]]) -> bytes:
        return self._simple_sheet_xml(
            ["Season", "Factory", "Development Style Count"],
            self._development_style_qty_rows(development_style_rows),
            '<cols><col min="1" max="2" width="18" customWidth="1"/>'
            '<col min="3" max="3" width="26" customWidth="1"/></cols>',
        )

    def _s2s_development_source_sheet_xml(self, source_rows: List[List[Any]]) -> bytes:
        headers = ["Season", "Factory", "Development Style Count", "Bulk Style Count", "Bulk Qty (pcs)"]
        sheet_rows = [
            self._row_xml(
                1,
                [self._cell_xml(1, col_idx, header, 4) for col_idx, header in enumerate(headers, 1)],
            )
        ]
        for row_idx, row in enumerate(source_rows, 2):
            cells = [
                self._cell_xml(row_idx, 1, row[0], 5),
                self._cell_xml(row_idx, 2, row[1], 5),
                self._cell_xml(row_idx, 3, row[2], 5, formula=self._s2s_development_count_formula(row_idx)),
                self._cell_xml(row_idx, 4, row[3], 5),
                self._cell_xml(row_idx, 5, row[4], 5),
            ]
            sheet_rows.append(self._row_xml(row_idx, cells))

        end_row = max(len(source_rows) + 1, 1)
        return self._worksheet_xml(
            f"A1:E{end_row}",
            "".join(sheet_rows),
            '<cols><col min="1" max="2" width="18" customWidth="1"/>'
            '<col min="3" max="5" width="24" customWidth="1"/></cols>',
            freeze_panes=True,
        )

    @staticmethod
    def _s2s_development_count_formula(row_idx: int) -> str:
        return (
            "SUMIFS('Development Style Qty'!$C:$C,"
            f"'Development Style Qty'!$A:$A,A{row_idx},"
            f"'Development Style Qty'!$B:$B,B{row_idx})"
        )

    def _simple_sheet_xml(
        self,
        headers: List[str],
        rows: List[List[Any]],
        columns_xml: str,
    ) -> bytes:
        sheet_rows = [
            self._row_xml(
                1,
                [self._cell_xml(1, col_idx, header, 4) for col_idx, header in enumerate(headers, 1)],
            )
        ]
        for row_idx, row in enumerate(rows, 2):
            cells = [
                self._cell_xml(row_idx, col_idx, value, 5)
                for col_idx, value in enumerate(row, 1)
            ]
            sheet_rows.append(self._row_xml(row_idx, cells))

        end_column = openpyxl.utils.get_column_letter(len(headers))
        end_row = max(len(rows) + 1, 1)
        return self._worksheet_xml(
            f"A1:{end_column}{end_row}",
            "".join(sheet_rows),
            columns_xml,
            freeze_panes=True,
        )

    def _country_analysis_source_rows(self, results: List[Dict[str, Any]]) -> List[List[Any]]:
        season_totals: Dict[Tuple[str, str], Dict[str, float]] = defaultdict(self._empty_metrics)
        year_totals: Dict[Tuple[int, str], Dict[str, float]] = defaultdict(self._empty_metrics)
        normalized_rows: List[Dict[str, Any]] = []

        for row in results:
            podd = self._to_datetime(row.get("PODD"))
            season = self._clean_text(row.get("Season"))
            factory = self._clean_text(row.get("Factory"))
            normalized = {
                "season": season,
                "year": podd.year if podd else "",
                "factory": factory,
                "country_group": self._country_bucket(row.get("Country/Region")),
                "quantity": self._optional_float(row.get("Quantity")) or 0.0,
                "tms_amount": self._optional_float(row.get("_tms_amount_value")) or 0.0,
            }
            normalized_rows.append(normalized)
            season_totals[(season, factory)]["quantity"] += normalized["quantity"]
            season_totals[(season, factory)]["tms_amount"] += normalized["tms_amount"]
            if podd:
                year_totals[(podd.year, factory)]["quantity"] += normalized["quantity"]
                year_totals[(podd.year, factory)]["tms_amount"] += normalized["tms_amount"]

        rows: List[List[Any]] = []
        for row in sorted(
            normalized_rows,
            key=lambda item: (
                *self._season_sort_key(item["season"]),
                self._year_sort_key(item["year"]),
                item["factory"],
                item["country_group"],
            ),
        ):
            season_total = season_totals[(row["season"], row["factory"])]
            year_total = (
                year_totals[(row["year"], row["factory"])]
                if row["year"] != ""
                else self._empty_metrics()
            )
            rows.append([
                row["season"],
                row["year"],
                row["factory"],
                row["country_group"],
                row["quantity"],
                row["quantity"] / season_total["quantity"] if season_total["quantity"] else 0,
                row["quantity"] / year_total["quantity"] if year_total["quantity"] else 0,
                row["tms_amount"],
                row["tms_amount"] / season_total["tms_amount"] if season_total["tms_amount"] else 0,
                row["tms_amount"] / year_total["tms_amount"] if year_total["tms_amount"] else 0,
            ])
        return rows

    def _ship_method_source_rows(self, results: List[Dict[str, Any]]) -> List[List[Any]]:
        metrics: Dict[Tuple[str, int, int, str], Dict[str, float]] = defaultdict(self._empty_metrics)
        monthly_totals: Dict[Tuple[str, int, int], Dict[str, float]] = defaultdict(self._empty_metrics)

        for row in results:
            podd = self._to_datetime(row.get("PODD"))
            if podd is None:
                continue
            factory = self._clean_text(row.get("Factory"))
            shipment_method = self._clean_text(row.get("Shipment Method"))
            key = (factory, podd.year, podd.month, shipment_method)
            total_key = (factory, podd.year, podd.month)
            self._add_metrics(metrics[key], row)
            self._add_metrics(monthly_totals[total_key], row)

        rows: List[List[Any]] = []
        for (factory, year, month, shipment_method), values in sorted(
            metrics.items(),
            key=lambda item: self._ship_method_sort_key(item[0]),
        ):
            total = monthly_totals[(factory, year, month)]
            quantity_total = total["quantity"]
            tms_amount_total = total["tms_amount"]
            rows.append([
                factory,
                year,
                datetime(2000, month, 1).strftime("%b"),
                shipment_method,
                values["quantity"],
                values["quantity"] / quantity_total if quantity_total else 0,
                values["tms_amount"],
                values["tms_amount"] / tms_amount_total if tms_amount_total else 0,
            ])
        return rows

    def _ship_method_analysis_sheet_xml(
        self,
        results: List[Dict[str, Any]],
        style_ids: Optional[Dict[str, int]] = None,
    ) -> bytes:
        headers = [
            "Factory",
            "Years",
            "PODD",
            "Shipment Method",
            "Quantity (Y)",
            "Quantity (%)",
            "TMS Amount (USD)",
            "TMS Amount (%)",
        ]
        summary_rows = self._ship_method_analysis_rows(results, style_ids)
        sheet_rows = [
            self._row_xml(
                1,
                [self._cell_xml(1, col_idx, header, 4) for col_idx, header in enumerate(headers, 1)],
            )
        ]
        for row_idx, row in enumerate(summary_rows, 2):
            values = row["values"]
            formulas = row["formulas"]
            style_ids = row["style_ids"]
            cells = [
                self._cell_xml(
                    row_idx,
                    col_idx,
                    value,
                    style_ids[col_idx - 1],
                    formula=formulas[col_idx - 1],
                )
                for col_idx, value in enumerate(values, 1)
            ]
            sheet_rows.append(self._row_xml(row_idx, cells))

        end_row = max(len(summary_rows) + 1, 1)
        dimension = f"A1:H{end_row}"
        return self._worksheet_xml(
            dimension,
            "".join(sheet_rows),
            '<cols><col min="1" max="1" width="18" customWidth="1"/>'
            '<col min="2" max="2" width="12" customWidth="1"/>'
            '<col min="3" max="4" width="18" customWidth="1"/>'
            '<col min="5" max="8" width="18" customWidth="1"/></cols>',
            freeze_panes=True,
            auto_filter_ref=dimension,
        )

    def _ship_method_analysis_rows(
        self,
        results: List[Dict[str, Any]],
        style_ids: Optional[Dict[str, int]] = None,
    ) -> List[Dict[str, Any]]:
        detail_metrics: Dict[Tuple[str, int, int, str], Dict[str, float]] = defaultdict(self._empty_metrics)
        monthly_totals: Dict[Tuple[str, int, int], Dict[str, float]] = defaultdict(self._empty_metrics)
        yearly_totals: Dict[Tuple[str, int], Dict[str, float]] = defaultdict(self._empty_metrics)

        for row in results:
            podd = self._to_datetime(row.get("PODD"))
            if podd is None:
                continue
            factory = self._clean_text(row.get("Factory"))
            shipment_method = self._clean_text(row.get("Shipment Method"))
            detail_key = (factory, podd.year, podd.month, shipment_method)
            month_key = (factory, podd.year, podd.month)
            year_key = (factory, podd.year)
            self._add_metrics(detail_metrics[detail_key], row)
            self._add_metrics(monthly_totals[month_key], row)
            self._add_metrics(yearly_totals[year_key], row)

        rows: List[Dict[str, Any]] = []
        previous_month_key: Optional[Tuple[str, int, int]] = None
        previous_year: Optional[int] = None
        for factory, year, month, shipment_method in sorted(
            detail_metrics,
            key=self._ship_method_sort_key,
        ):
            month_key = (factory, year, month)
            if previous_year is not None and year != previous_year:
                if previous_month_key is not None:
                    rows.append(self._ship_method_total_row(
                        previous_month_key,
                        monthly_totals,
                        yearly_totals,
                        len(rows) + 2,
                        style_ids,
                    ))
                    previous_month_key = None
                rows.extend(self._ship_method_year_total_rows(previous_year, yearly_totals, len(rows) + 2, style_ids))

            if previous_month_key is not None and previous_month_key != month_key:
                rows.append(self._ship_method_total_row(
                    previous_month_key,
                    monthly_totals,
                    yearly_totals,
                    len(rows) + 2,
                    style_ids,
                ))
            detail_row_index = len(rows) + 2
            metrics = detail_metrics[(factory, year, month, shipment_method)]
            month_total = monthly_totals[month_key]
            rows.append(self._ship_method_detail_row(
                factory,
                year,
                month,
                shipment_method,
                metrics,
                month_total,
                detail_row_index,
            ))
            previous_month_key = month_key
            previous_year = year

        if previous_month_key is not None:
            rows.append(self._ship_method_total_row(
                previous_month_key,
                monthly_totals,
                yearly_totals,
                len(rows) + 2,
                style_ids,
            ))
        if previous_year is not None:
            rows.extend(self._ship_method_year_total_rows(previous_year, yearly_totals, len(rows) + 2, style_ids))
        return rows

    def _ship_method_detail_row(
        self,
        factory: str,
        year: int,
        month: int,
        shipment_method: str,
        metrics: Dict[str, float],
        month_total: Dict[str, float],
        row_index: int,
    ) -> Dict[str, Any]:
        quantity = metrics["quantity"]
        tms_amount = metrics["tms_amount"]
        return {
            "values": [
                factory,
                year,
                datetime(2000, month, 1).strftime("%b"),
                shipment_method,
                quantity,
                quantity / month_total["quantity"] if month_total["quantity"] else 0,
                tms_amount,
                tms_amount / month_total["tms_amount"] if month_total["tms_amount"] else 0,
            ],
            "formulas": [
                None,
                None,
                None,
                None,
                self._ship_method_month_sumifs_formula("M", row_index, include_method=True),
                self._ship_method_month_percentage_formula("M", "E", row_index),
                self._ship_method_month_sumifs_formula("Q", row_index, include_method=True),
                self._ship_method_month_percentage_formula("Q", "G", row_index),
            ],
            "style_ids": self._ship_method_style_ids(),
        }

    def _ship_method_total_row(
        self,
        month_key: Tuple[str, int, int],
        monthly_totals: Dict[Tuple[str, int, int], Dict[str, float]],
        yearly_totals: Dict[Tuple[str, int], Dict[str, float]],
        row_index: int,
        style_ids: Optional[Dict[str, int]] = None,
    ) -> Dict[str, Any]:
        factory, year, month = month_key
        month_total = monthly_totals[month_key]
        year_total = yearly_totals[(factory, year)]
        quantity = month_total["quantity"]
        tms_amount = month_total["tms_amount"]
        return {
            "values": [
                factory,
                year,
                f"{datetime(2000, month, 1).strftime('%b')} Total",
                None,
                quantity,
                quantity / year_total["quantity"] if year_total["quantity"] else 0,
                tms_amount,
                tms_amount / year_total["tms_amount"] if year_total["tms_amount"] else 0,
            ],
            "formulas": [
                None,
                None,
                None,
                None,
                self._ship_method_month_sumifs_formula("M", row_index, include_method=False),
                self._ship_method_year_percentage_formula("M", "E", row_index),
                self._ship_method_month_sumifs_formula("Q", row_index, include_method=False),
                self._ship_method_year_percentage_formula("Q", "G", row_index),
            ],
            "style_ids": self._ship_method_total_style_ids(style_ids),
        }

    def _ship_method_year_total_rows(
        self,
        year: int,
        yearly_totals: Dict[Tuple[str, int], Dict[str, float]],
        start_row_index: int,
        style_ids: Optional[Dict[str, int]] = None,
    ) -> List[Dict[str, Any]]:
        rows: List[Dict[str, Any]] = []
        year_keys = sorted(key for key in yearly_totals if key[1] == year)
        for offset, year_key in enumerate(year_keys):
            row_index = start_row_index + offset
            rows.append(self._ship_method_year_total_row(year_key, yearly_totals[year_key], row_index, style_ids))
        return rows

    def _ship_method_year_total_row(
        self,
        year_key: Tuple[str, int],
        year_total: Dict[str, float],
        row_index: int,
        style_ids: Optional[Dict[str, int]] = None,
    ) -> Dict[str, Any]:
        factory, year = year_key
        quantity = year_total["quantity"]
        tms_amount = year_total["tms_amount"]
        return {
            "values": [
                factory,
                f"{year} Total",
                None,
                None,
                quantity,
                1 if quantity else 0,
                tms_amount,
                1 if tms_amount else 0,
            ],
            "formulas": [
                None,
                None,
                None,
                None,
                self._ship_method_year_sumifs_formula("M", row_index, year),
                self._ship_method_year_percentage_formula("M", "E", row_index, year),
                self._ship_method_year_sumifs_formula("Q", row_index, year),
                self._ship_method_year_percentage_formula("Q", "G", row_index, year),
            ],
            "style_ids": self._ship_method_total_style_ids(style_ids),
        }

    def _ship_method_style_ids(self) -> List[int]:
        return [
            self.COUNTRY_SUMMARY_TEXT_STYLE,
            self.COUNTRY_SUMMARY_TEXT_STYLE,
            self.COUNTRY_SUMMARY_TEXT_STYLE,
            self.COUNTRY_SUMMARY_TEXT_STYLE,
            self.COUNTRY_SUMMARY_TEXT_STYLE,
            self.COUNTRY_SUMMARY_PERCENT_STYLE,
            self.COUNTRY_SUMMARY_CURRENCY_STYLE,
            self.COUNTRY_SUMMARY_PERCENT_STYLE,
        ]

    def _ship_method_total_style_ids(self, style_ids: Optional[Dict[str, int]]) -> List[int]:
        if not style_ids:
            return self._ship_method_style_ids()
        return [
            style_ids["ship_method_total_text"],
            style_ids["ship_method_total_text"],
            style_ids["ship_method_total_text"],
            style_ids["ship_method_total_text"],
            style_ids["ship_method_total_text"],
            style_ids["ship_method_total_percent"],
            style_ids["ship_method_total_currency"],
            style_ids["ship_method_total_percent"],
        ]

    @staticmethod
    def _ship_method_month_start_formula(row_index: int) -> str:
        return (
            f'DATE($B{row_index},'
            f'MATCH(LEFT($C{row_index},3),{{"Jan","Feb","Mar","Apr","May","Jun",'
            f'"Jul","Aug","Sep","Oct","Nov","Dec"}},0),1)'
        )

    def _ship_method_month_sumifs_formula(self, metric_column: str, row_index: int, include_method: bool) -> str:
        month_start = self._ship_method_month_start_formula(row_index)
        criteria = [
            f"Result!$C:$C,$A{row_index}",
            f'Result!$H:$H,">="&{month_start}',
            f'Result!$H:$H,"<"&EDATE({month_start},1)',
        ]
        if include_method:
            criteria.append(f"Result!$K:$K,$D{row_index}")
        return f"SUMIFS(Result!${metric_column}:${metric_column},{','.join(criteria)})"

    def _ship_method_year_sumifs_formula(self, metric_column: str, row_index: int, year: Optional[int] = None) -> str:
        start_date = f"DATE({year},1,1)" if year is not None else f"DATE($B{row_index},1,1)"
        end_date = f"DATE({year + 1},1,1)" if year is not None else f"DATE($B{row_index}+1,1,1)"
        return (
            f'SUMIFS(Result!${metric_column}:${metric_column},Result!$C:$C,$A{row_index},'
            f'Result!$H:$H,">="&{start_date},'
            f'Result!$H:$H,"<"&{end_date})'
        )

    def _ship_method_month_percentage_formula(self, metric_column: str, value_column: str, row_index: int) -> str:
        denominator = self._ship_method_month_sumifs_formula(metric_column, row_index, include_method=False)
        return f"IFERROR({value_column}{row_index}/{denominator},0)"

    def _ship_method_year_percentage_formula(
        self,
        metric_column: str,
        value_column: str,
        row_index: int,
        year: Optional[int] = None,
    ) -> str:
        denominator = self._ship_method_year_sumifs_formula(metric_column, row_index, year)
        return f"IFERROR({value_column}{row_index}/{denominator},0)"

    def _s2s_development_source_rows(
        self,
        results: List[Dict[str, Any]],
        development_style_rows: List[Dict[str, Any]],
    ) -> List[List[Any]]:
        development_totals = self._development_style_totals(development_style_rows)
        bulk_styles: Dict[Tuple[str, str], Set[str]] = defaultdict(set)
        quantities: Dict[Tuple[str, str], float] = defaultdict(float)
        for row in results:
            key = (self._clean_text(row.get("Season")), self._clean_text(row.get("Factory")))
            working = self._clean_text(row.get("Working Number"))
            if working:
                bulk_styles[key].add(working)
            quantities[key] += self._optional_float(row.get("Quantity")) or 0.0

        rows: List[List[Any]] = []
        source_keys = set(development_totals) | set(bulk_styles) | set(quantities)
        for season, factory in sorted(source_keys, key=self._season_factory_sort_key):
            rows.append([
                season,
                factory,
                development_totals.get((season, factory), 0),
                len(bulk_styles[(season, factory)]),
                quantities[(season, factory)],
            ])
        return rows

    def _development_style_qty_rows(self, development_style_rows: List[Dict[str, Any]]) -> List[List[Any]]:
        return [
            [season, factory, count]
            for (season, factory), count in sorted(
                self._development_style_totals(development_style_rows).items(),
                key=lambda item: self._season_factory_sort_key(item[0]),
            )
        ]

    def _development_style_totals(self, development_style_rows: List[Dict[str, Any]]) -> Dict[Tuple[str, str], int]:
        totals: Dict[Tuple[str, str], int] = defaultdict(int)
        for row in development_style_rows:
            season = self._clean_text(row.get("Season"))
            factory = self._clean_text(row.get("Factory"))
            if not factory:
                continue
            totals[(season, factory)] += int(self._optional_float(row.get("Development Style Count")) or 0)
        return totals

    def _s2s_development_analysis_sheet_xml(self, source_rows: List[List[Any]]) -> bytes:
        return self._simple_sheet_xml(
            ["Season", "Factory", "Development Style Count", "Bulk Style Count", "Bulk Qty (pcs)"],
            self._s2s_development_analysis_display_rows(source_rows),
            '<cols><col min="1" max="2" width="18" customWidth="1"/>'
            '<col min="3" max="5" width="24" customWidth="1"/></cols>',
        )

    def _s2s_development_analysis_display_rows(self, source_rows: List[List[Any]]) -> List[List[Any]]:
        grouped: Dict[str, List[List[Any]]] = defaultdict(list)
        for row in source_rows:
            if len(row) < 5:
                continue
            grouped[self._clean_text(row[0])].append(row)

        display_rows: List[List[Any]] = []
        grand_development_count = 0
        grand_bulk_count = 0
        grand_bulk_quantity = 0.0
        for season in sorted(grouped, key=self._season_sort_key):
            season_label = season or "(blank)"
            season_development_count = 0
            season_bulk_count = 0
            season_bulk_quantity = 0.0
            first_factory_for_season = True
            for row in sorted(grouped[season], key=lambda item: self._clean_text(item[1])):
                factory = self._clean_text(row[1])
                development_count = int(self._optional_float(row[2]) or 0)
                bulk_count = int(self._optional_float(row[3]) or 0)
                bulk_quantity = self._optional_float(row[4]) or 0.0
                display_rows.append([
                    season_label if first_factory_for_season else "",
                    factory,
                    development_count,
                    bulk_count,
                    bulk_quantity,
                ])
                first_factory_for_season = False
                season_development_count += development_count
                season_bulk_count += bulk_count
                season_bulk_quantity += bulk_quantity

            display_rows.append([
                f"{season_label} Total",
                "",
                season_development_count,
                season_bulk_count,
                season_bulk_quantity,
            ])
            grand_development_count += season_development_count
            grand_bulk_count += season_bulk_count
            grand_bulk_quantity += season_bulk_quantity

        if display_rows:
            display_rows.append([
                "Grand Total",
                "",
                grand_development_count,
                grand_bulk_count,
                grand_bulk_quantity,
            ])
        return display_rows

    def _updated_country_summary_sheet_xml(
        self,
        xml_bytes: bytes,
        results: List[Dict[str, Any]],
        start_row: int,
        period_kind: str,
    ) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        sheet_data = root.find(f"{{{self.OOXML_MAIN_NS}}}sheetData")
        if sheet_data is None:
            return xml_bytes

        for row in list(sheet_data):
            sheet_data.remove(row)

        summary_rows = self._country_summary_block_rows(results, period_kind, start_row)
        for offset, (row_kind, values, formulas) in enumerate(summary_rows):
            row_index = start_row + offset
            style_ids = self._country_summary_style_ids(row_kind)
            cells = [
                self._cell_xml(
                    row_index,
                    column_index,
                    value,
                    style_ids[column_index - 1],
                    formula=formulas[column_index - 1],
                )
                for column_index, value in enumerate(values, 1)
            ]
            sheet_data.append(self._row_element(self._row_xml(row_index, cells)))

        dimension = root.find(f"{{{self.OOXML_MAIN_NS}}}dimension")
        if dimension is not None:
            max_row = max(
                [
                    self._worksheet_row_number(row) or 0
                    for row in sheet_data.findall(f"{{{self.OOXML_MAIN_NS}}}row")
                ]
                or [start_row + len(summary_rows) - 1]
            )
            dimension.set("ref", f"A1:F{max_row}")

        return self._serialize_spreadsheet_xml(root)

    def _country_summary_block_rows(
        self,
        results: List[Dict[str, Any]],
        period_kind: str,
        start_row: int,
    ) -> List[Tuple[str, List[Any], List[Optional[str]]]]:
        grouped_metrics: Dict[Tuple[Any, str], Dict[str, Dict[str, float]]] = defaultdict(
            lambda: defaultdict(self._empty_metrics)
        )
        period_labels: Dict[Any, str] = {}

        for row in results:
            period_key, period_label = self._country_summary_period(row, period_kind)
            if period_key is None:
                continue
            factory = self._clean_text(row.get("Factory")) or "(blank)"
            bucket = self._country_bucket(row.get("Country/Region"))
            metrics = grouped_metrics[(period_key, factory)][bucket]
            metrics["quantity"] += self._optional_float(row.get("Quantity")) or 0.0
            metrics["tms_amount"] += self._country_summary_tms_amount(row)
            period_labels[period_key] = period_label

        empty_formulas: List[Optional[str]] = [None] * 6
        if not grouped_metrics:
            header_label = "Year" if period_kind == "Year" else "Season"
            return [(
                "header",
                [header_label, "Factory", *[label for _, label in self.COUNTRY_SUMMARY_BUCKETS], "Total Order"],
                empty_formulas,
            )]

        summary_rows: List[Tuple[str, List[Any], List[Optional[str]]]] = []
        previous_period_key: Any = None
        for period_key, factory in sorted(
            grouped_metrics,
            key=lambda key: self._country_summary_sort_key(key, period_kind),
        ):
            if period_key != previous_period_key:
                if previous_period_key is not None:
                    summary_rows.append(("spacer", [None] * 6, empty_formulas))
                summary_rows.append((
                    "header",
                    [
                        period_labels[period_key],
                        "Factory",
                        *[label for _, label in self.COUNTRY_SUMMARY_BUCKETS],
                        "Total Order",
                    ],
                    empty_formulas,
                ))
                previous_period_key = period_key

            bucket_metrics = grouped_metrics[(period_key, factory)]
            quantities = [bucket_metrics[bucket]["quantity"] for bucket, _ in self.COUNTRY_SUMMARY_BUCKETS]
            values = [bucket_metrics[bucket]["tms_amount"] for bucket, _ in self.COUNTRY_SUMMARY_BUCKETS]
            total_quantity = sum(quantities)
            total_value = sum(values)
            quantity_percentages = [
                quantity / total_quantity if total_quantity else 0
                for quantity in quantities
            ]
            value_percentages = [
                value / total_value if total_value else 0
                for value in values
            ]
            qty_row = start_row + len(summary_rows)
            value_row = qty_row + 2
            summary_rows.extend([
                (
                    "qty",
                    ["Qty(pcs)", factory, *quantities, total_quantity],
                    self._country_summary_metric_formulas("M", period_kind, period_key, qty_row),
                ),
                (
                    "qty_pct",
                    ["Qty percentage", factory, *quantity_percentages, None],
                    self._country_summary_percentage_formulas(qty_row),
                ),
                (
                    "value",
                    ["Value(usd)", factory, *values, total_value],
                    self._country_summary_metric_formulas("Q", period_kind, period_key, value_row),
                ),
                (
                    "value_pct",
                    ["Value percentage", factory, *value_percentages, None],
                    self._country_summary_percentage_formulas(value_row),
                ),
            ])
        return summary_rows

    def _country_summary_metric_formulas(
        self,
        metric_column: str,
        period_kind: str,
        period_key: Any,
        row_index: int,
    ) -> List[Optional[str]]:
        formulas: List[Optional[str]] = [None, None]
        formulas.extend([
            self._country_summary_sumifs_formula(metric_column, period_kind, period_key, row_index, "CHINA"),
            self._country_summary_sumifs_formula(metric_column, period_kind, period_key, row_index, "UNITED STATES"),
            self._country_summary_sumifs_formula(metric_column, period_kind, period_key, row_index, "OTHER COUNTRIES"),
            f"SUM(C{row_index}:E{row_index})",
        ])
        return formulas

    @staticmethod
    def _country_summary_percentage_formulas(source_row: int) -> List[Optional[str]]:
        return [
            None,
            None,
            f"IF($F{source_row}=0,0,C{source_row}/$F{source_row})",
            f"IF($F{source_row}=0,0,D{source_row}/$F{source_row})",
            f"IF($F{source_row}=0,0,E{source_row}/$F{source_row})",
            None,
        ]

    def _country_summary_sumifs_formula(
        self,
        metric_column: str,
        period_kind: str,
        period_key: Any,
        row_index: int,
        bucket: str,
    ) -> str:
        criteria = [
            f"Result!$C:$C,$B{row_index}",
        ]
        if bucket == "OTHER COUNTRIES":
            criteria.extend([
                'Result!$J:$J,"<>CHINA"',
                'Result!$J:$J,"<>UNITED STATES"',
            ])
        else:
            criteria.append(f"Result!$J:$J,{self._excel_formula_string(bucket)}")

        if period_kind == "Year":
            year = int(period_key)
            criteria.extend([
                f'Result!$H:$H,">="&DATE({year},1,1)',
                f'Result!$H:$H,"<"&DATE({year + 1},1,1)',
            ])
        else:
            criteria.append(f"Result!$B:$B,{self._excel_formula_string(str(period_key))}")

        return f"SUMIFS(Result!${metric_column}:${metric_column},{','.join(criteria)})"

    @staticmethod
    def _excel_formula_string(value: str) -> str:
        return '"' + value.replace('"', '""') + '"'

    def _country_summary_period(
        self,
        row: Dict[str, Any],
        period_kind: str,
    ) -> Tuple[Optional[Any], str]:
        if period_kind == "Year":
            podd = self._to_datetime(row.get("PODD"))
            if podd is None:
                return None, ""
            return podd.year, f"Year {podd.year}"

        season = self._clean_text(row.get("Season"))
        if season:
            return season, f"Season {season}"
        return "", "Season (blank)"

    @classmethod
    def _country_summary_sort_key(cls, key: Tuple[Any, str], period_kind: str) -> Tuple[Any, ...]:
        period_key, factory = key
        if period_kind == "Year" and isinstance(period_key, int):
            return (0, period_key, factory)
        return (*cls._season_sort_key(period_key), factory)

    @classmethod
    def _season_factory_sort_key(cls, key: Tuple[str, str]) -> Tuple[Any, ...]:
        season, factory = key
        return (*cls._season_sort_key(season), cls._clean_text(factory))

    @classmethod
    def _season_sort_key(cls, value: Any) -> Tuple[int, int, int, str]:
        season = cls._clean_text(value).upper()
        if not season:
            return (2, 9999, 99, "")

        compact = "".join(char for char in season if char.isalnum())
        prefix = compact[:2]
        year_digits = "".join(char for char in compact[2:] if char.isdigit())
        if prefix in {"SS", "FW"} and year_digits:
            if len(year_digits) >= 4:
                year = int(year_digits[:4])
            else:
                year = 2000 + int(year_digits[:2])
            season_order = 0 if prefix == "SS" else 1
            return (0, year, season_order, season)

        return (1, 9999, 99, season)

    @staticmethod
    def _year_sort_key(value: Any) -> Tuple[int, Any]:
        if isinstance(value, int):
            return (0, value)
        return (1, str(value))

    @staticmethod
    def _ship_method_sort_key(key: Tuple[str, int, int, str]) -> Tuple[int, int, str, str]:
        factory, year, month, shipment_method = key
        return (year, month, factory, shipment_method)

    def _country_summary_tms_amount(self, row: Dict[str, Any]) -> float:
        tms_amount = self._optional_float(row.get("_tms_amount_value"))
        if tms_amount is None:
            tms_amount = self._optional_float(row.get("TMS Amount(USD)"))
        return tms_amount or 0.0

    def _country_summary_style_ids(self, row_kind: str) -> List[Optional[int]]:
        if row_kind == "spacer":
            return [None] * 6
        if row_kind == "header":
            return [self.COUNTRY_SUMMARY_HEADER_STYLE] * 6
        if row_kind == "qty_pct":
            return [
                self.COUNTRY_SUMMARY_TEXT_STYLE,
                self.COUNTRY_SUMMARY_TEXT_STYLE,
                self.COUNTRY_SUMMARY_PERCENT_STYLE,
                self.COUNTRY_SUMMARY_PERCENT_STYLE,
                self.COUNTRY_SUMMARY_PERCENT_STYLE,
                self.COUNTRY_SUMMARY_PERCENT_STYLE,
            ]
        if row_kind == "value":
            return [
                self.COUNTRY_SUMMARY_TEXT_STYLE,
                self.COUNTRY_SUMMARY_TEXT_STYLE,
                self.COUNTRY_SUMMARY_CURRENCY_STYLE,
                self.COUNTRY_SUMMARY_CURRENCY_STYLE,
                self.COUNTRY_SUMMARY_CURRENCY_STYLE,
                self.COUNTRY_SUMMARY_CURRENCY_STYLE,
            ]
        if row_kind == "value_pct":
            return [
                self.COUNTRY_SUMMARY_TEXT_STYLE,
                self.COUNTRY_SUMMARY_TEXT_STYLE,
                self.COUNTRY_SUMMARY_PERCENT_STYLE,
                self.COUNTRY_SUMMARY_PERCENT_STYLE,
                self.COUNTRY_SUMMARY_PERCENT_STYLE,
                self.COUNTRY_SUMMARY_TEXT_STYLE,
            ]
        return [self.COUNTRY_SUMMARY_TEXT_STYLE] * 6

    def _row_element(self, row_xml: str) -> ElementTree.Element:
        wrapper = ElementTree.fromstring(
            f'<root xmlns="{self.OOXML_MAIN_NS}" xmlns:x14ac="{self.OOXML_X14AC_NS}">{row_xml}</root>'
        )
        return list(wrapper)[0]

    @staticmethod
    def _worksheet_row_number(row: ElementTree.Element) -> Optional[int]:
        value = row.attrib.get("r")
        if value is None:
            return None
        try:
            return int(value)
        except ValueError:
            return None

    def _worksheet_xml(
        self,
        dimension: str,
        sheet_data: str,
        columns_xml: str,
        freeze_panes: bool = False,
        table_rel_id: Optional[str] = None,
        auto_filter_ref: Optional[str] = None,
    ) -> bytes:
        if freeze_panes:
            sheet_views = (
                '<sheetViews><sheetView workbookViewId="0">'
                '<pane ySplit="1" topLeftCell="A2" activePane="bottomLeft" state="frozen"/>'
                '<selection pane="bottomLeft"/></sheetView></sheetViews>'
            )
        else:
            sheet_views = '<sheetViews><sheetView workbookViewId="0"/></sheetViews>'

        table_parts = (
            f'<tableParts count="1"><tablePart r:id="{table_rel_id}"/></tableParts>'
            if table_rel_id
            else ""
        )
        auto_filter = f'<autoFilter ref="{auto_filter_ref}"/>' if auto_filter_ref else ""
        xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            f'<worksheet xmlns="{self.OOXML_MAIN_NS}" xmlns:r="{self.OOXML_REL_NS}" '
            f'xmlns:mc="{self.OOXML_MC_NS}" mc:Ignorable="x14ac" '
            f'xmlns:x14ac="{self.OOXML_X14AC_NS}">'
            f'<dimension ref="{dimension}"/>'
            f'{sheet_views}'
            '<sheetFormatPr defaultRowHeight="15" x14ac:dyDescent="0.25"/>'
            f'{columns_xml}'
            f'<sheetData>{sheet_data}</sheetData>'
            f'{auto_filter}'
            '<pageMargins left="0.7" right="0.7" top="0.75" bottom="0.75" header="0.3" footer="0.3"/>'
            f'{table_parts}'
            '</worksheet>'
        )
        return xml.encode("utf-8")

    def _row_xml(
        self,
        row_idx: int,
        cells: List[str],
        height: Optional[str] = None,
        custom_height: bool = False,
    ) -> str:
        attrs = [f'r="{row_idx}"', f'spans="1:{len(cells)}"', 'x14ac:dyDescent="0.25"']
        if height:
            attrs.insert(2, f'ht="{height}"')
        if custom_height:
            attrs.insert(3 if height else 2, 'customHeight="1"')
        return f"<row {' '.join(attrs)}>{''.join(cells)}</row>"

    def _date_cell_xml(self, row_idx: int, col_idx: int, value: Any, style_id: int) -> str:
        parsed = self._to_datetime(value)
        if parsed is None:
            return self._cell_xml(row_idx, col_idx, value, style_id)
        return self._cell_xml(row_idx, col_idx, to_excel(parsed), style_id)

    def _cell_xml(
        self,
        row_idx: int,
        col_idx: int,
        value: Any,
        style_id: Optional[int],
        formula: Optional[str] = None,
    ) -> str:
        ref = f"{openpyxl.utils.get_column_letter(col_idx)}{row_idx}"
        style_attr = f' s="{style_id}"' if style_id is not None else ""
        if formula:
            cached = "" if self._is_blank(value) else f"<v>{self._format_excel_number(value)}</v>"
            return f'<c r="{ref}"{style_attr}><f>{escape(formula)}</f>{cached}</c>'

        if self._is_blank(value):
            return f'<c r="{ref}"{style_attr}/>'

        if isinstance(value, bool):
            return f'<c r="{ref}"{style_attr} t="b"><v>{1 if value else 0}</v></c>'

        if self._is_number(value):
            return f'<c r="{ref}"{style_attr}><v>{self._format_excel_number(value)}</v></c>'

        text = escape(str(value))
        return (
            f'<c r="{ref}"{style_attr} t="inlineStr">'
            f'<is><t xml:space="preserve">{text}</t></is></c>'
        )

    @staticmethod
    def _is_number(value: Any) -> bool:
        if isinstance(value, bool):
            return False
        return isinstance(value, (int, float)) and not pd.isna(value)

    @staticmethod
    def _format_excel_number(value: Any) -> str:
        return format(float(value), ".15g")

    @staticmethod
    def _result_header_style(col_idx: int) -> int:
        if col_idx == 1:
            return 32
        if col_idx == 17:
            return 34
        return 33

    @staticmethod
    def _result_data_style(col_idx: int) -> int:
        if col_idx == 1:
            return 35
        if col_idx in {7, 8}:
            return 37
        if col_idx in {14, 15, 16}:
            return 39
        if col_idx == 17:
            return 40
        return 36

    def _updated_styles_xml(self, xml_bytes: bytes) -> Tuple[bytes, Dict[str, int]]:
        root = ElementTree.fromstring(xml_bytes)
        fills = root.find(f"{{{self.OOXML_MAIN_NS}}}fills")
        cell_xfs = root.find(f"{{{self.OOXML_MAIN_NS}}}cellXfs")
        if fills is None or cell_xfs is None:
            return xml_bytes, {"factory_price_conflict": self._result_data_style(14)}

        conflict_fill_id = self._find_or_append_solid_fill(fills, "FFFFF2CC")
        normal_price_style = cell_xfs[self._result_data_style(14)]
        conflict_style_id = self._find_or_append_style_with_fill(
            cell_xfs,
            normal_price_style,
            conflict_fill_id,
        )
        ship_total_fill_id = self._find_or_append_solid_fill(fills, self.SHIP_METHOD_TOTAL_FILL_RGB)
        ship_total_text_style_id = self._find_or_append_style_with_fill(
            cell_xfs,
            cell_xfs[self.COUNTRY_SUMMARY_TEXT_STYLE],
            ship_total_fill_id,
        )
        ship_total_percent_style_id = self._find_or_append_style_with_fill(
            cell_xfs,
            cell_xfs[self.COUNTRY_SUMMARY_PERCENT_STYLE],
            ship_total_fill_id,
        )
        ship_total_currency_style_id = self._find_or_append_style_with_fill(
            cell_xfs,
            cell_xfs[self.COUNTRY_SUMMARY_CURRENCY_STYLE],
            ship_total_fill_id,
        )
        fills.set("count", str(len(list(fills))))
        cell_xfs.set("count", str(len(list(cell_xfs))))
        return self._serialize_spreadsheet_xml(root), {
            "factory_price_conflict": conflict_style_id,
            "ship_method_total_text": ship_total_text_style_id,
            "ship_method_total_percent": ship_total_percent_style_id,
            "ship_method_total_currency": ship_total_currency_style_id,
        }

    def _find_or_append_solid_fill(self, fills: ElementTree.Element, rgb: str) -> int:
        for index, fill in enumerate(list(fills)):
            fg_color = fill.find(f".//{{{self.OOXML_MAIN_NS}}}fgColor")
            if fg_color is not None and fg_color.attrib.get("rgb") == rgb:
                return index

        fill = ElementTree.Element(f"{{{self.OOXML_MAIN_NS}}}fill")
        pattern = ElementTree.SubElement(fill, f"{{{self.OOXML_MAIN_NS}}}patternFill", {"patternType": "solid"})
        ElementTree.SubElement(pattern, f"{{{self.OOXML_MAIN_NS}}}fgColor", {"rgb": rgb})
        ElementTree.SubElement(pattern, f"{{{self.OOXML_MAIN_NS}}}bgColor", {"rgb": rgb})
        fills.append(fill)
        return len(list(fills)) - 1

    def _find_or_append_style_with_fill(
        self,
        cell_xfs: ElementTree.Element,
        base_style: ElementTree.Element,
        fill_id: int,
    ) -> int:
        base_attrib = dict(base_style.attrib)
        base_attrib["fillId"] = str(fill_id)
        base_attrib["applyFill"] = "1"
        for index, style in enumerate(list(cell_xfs)):
            if dict(style.attrib) == base_attrib:
                return index

        style = ElementTree.Element(f"{{{self.OOXML_MAIN_NS}}}xf", base_attrib)
        for child in list(base_style):
            style.append(ElementTree.fromstring(ElementTree.tostring(child)))
        cell_xfs.append(style)
        return len(list(cell_xfs)) - 1

    def _updated_table_xml(self, xml_bytes: bytes, table_ref: str) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        root.set("ref", table_ref)
        auto_filter = root.find(f"{{{self.OOXML_MAIN_NS}}}autoFilter")
        if auto_filter is not None:
            auto_filter.set("ref", table_ref)
        return self._serialize_spreadsheet_xml(root)

    def _updated_pivot_cache_xml(self, xml_bytes: bytes, table_ref: str) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        root.set("refreshOnLoad", "1")
        root.set("enableRefresh", "1")
        root.set("recordCount", "0")
        source = root.find(f".//{{{self.OOXML_MAIN_NS}}}worksheetSource")
        if source is not None:
            source.set("sheet", "Result")
            source.set("ref", table_ref)
        return self._serialize_spreadsheet_xml(root)

    def _with_pivot_table_definition(self, xml_bytes: bytes, relationship_id: str) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        rel_attr = f"{{{self.OOXML_REL_NS}}}id"
        for child in root.findall(f"{{{self.OOXML_MAIN_NS}}}pivotTableDefinition"):
            if child.attrib.get(rel_attr) == relationship_id:
                return self._serialize_spreadsheet_xml(root)

        pivot_definition = ElementTree.Element(
            f"{{{self.OOXML_MAIN_NS}}}pivotTableDefinition",
            {rel_attr: relationship_id},
        )
        ext_list_tag = f"{{{self.OOXML_MAIN_NS}}}extLst"
        for index, child in enumerate(list(root)):
            if child.tag == ext_list_tag:
                root.insert(index, pivot_definition)
                break
        else:
            root.append(pivot_definition)
        return self._serialize_spreadsheet_xml(root)

    def _pivot_cache_definition_xml(
        self,
        sheet_name: str,
        source_ref: str,
        fields: List[Tuple[str, str]],
    ) -> bytes:
        cache_fields = []
        for name, num_fmt_id in fields:
            if num_fmt_id in {"3", "10", "166"}:
                shared_items = '<sharedItems containsBlank="1" containsNumber="1"/>'
            else:
                shared_items = '<sharedItems containsBlank="1"/>'
            cache_fields.append(
                f'<cacheField name="{escape(name)}" numFmtId="{num_fmt_id}">{shared_items}</cacheField>'
            )

        xml = (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            f'<pivotCacheDefinition xmlns="{self.OOXML_MAIN_NS}" xmlns:r="{self.OOXML_REL_NS}" '
            'r:id="rId1" refreshOnLoad="1" enableRefresh="1" recordCount="0" '
            'createdVersion="5" refreshedVersion="5" minRefreshableVersion="3">'
            '<cacheSource type="worksheet">'
            f'<worksheetSource ref="{source_ref}" sheet="{escape(sheet_name)}"/>'
            '</cacheSource>'
            f'<cacheFields count="{len(fields)}">{"".join(cache_fields)}</cacheFields>'
            '</pivotCacheDefinition>'
        )
        return xml.encode("utf-8")

    def _empty_pivot_cache_records_xml(self) -> bytes:
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            f'<pivotCacheRecords xmlns="{self.OOXML_MAIN_NS}" xmlns:r="{self.OOXML_REL_NS}" count="0"/>'
        ).encode("utf-8")

    def _pivot_cache_relationship_xml(self, cache_index: int) -> bytes:
        return (
            '<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            f'<Relationships xmlns="{self.OPC_RELATIONSHIPS_NS}">'
            '<Relationship Id="rId1" '
            'Type="http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheRecords" '
            f'Target="pivotCacheRecords{cache_index}.xml"/>'
            '</Relationships>'
        ).encode("utf-8")

    def _updated_pivot_table_xml(
        self,
        xml_bytes: bytes,
        cache_id: str,
        row_fields: List[int],
        column_fields: List[int],
        data_fields: List[Tuple[str, int, Optional[str], str]],
        field_count: int,
        location_ref: Optional[str] = None,
    ) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        root.set("cacheId", cache_id)
        if location_ref:
            location = root.find(f"{{{self.OOXML_MAIN_NS}}}location")
            if location is not None:
                location.set("ref", location_ref)
        self._replace_pivot_child(root, "pivotFields", self._pivot_fields_element(
            field_count,
            row_fields,
            [field for field in column_fields if field >= 0],
            [field_index for _, field_index, _, _ in data_fields],
        ))
        self._replace_pivot_child(root, "rowFields", self._pivot_axis_fields_element("rowFields", row_fields))
        self._replace_pivot_child(root, "colFields", self._pivot_axis_fields_element("colFields", column_fields))
        self._replace_pivot_child(root, "dataFields", self._pivot_data_fields_element(data_fields))
        return self._serialize_spreadsheet_xml(root)

    def _replace_pivot_child(
        self,
        root: ElementTree.Element,
        child_name: str,
        replacement: ElementTree.Element,
    ) -> None:
        tag = f"{{{self.OOXML_MAIN_NS}}}{child_name}"
        for index, child in enumerate(list(root)):
            if child.tag == tag:
                root.remove(child)
                root.insert(index, replacement)
                return
        root.append(replacement)

    def _pivot_fields_element(
        self,
        field_count: int,
        row_fields: List[int],
        column_fields: List[int],
        data_field_indexes: List[int],
    ) -> ElementTree.Element:
        root = ElementTree.Element(f"{{{self.OOXML_MAIN_NS}}}pivotFields", {"count": str(field_count)})
        for index in range(field_count):
            attrs = {"compact": "0", "outline": "0", "showAll": "0"}
            if index in row_fields:
                attrs["axis"] = "axisRow"
            elif index in column_fields:
                attrs["axis"] = "axisCol"
            if index in data_field_indexes:
                attrs["dataField"] = "1"
            ElementTree.SubElement(root, f"{{{self.OOXML_MAIN_NS}}}pivotField", attrs)
        return root

    def _pivot_axis_fields_element(self, name: str, fields: List[int]) -> ElementTree.Element:
        root = ElementTree.Element(f"{{{self.OOXML_MAIN_NS}}}{name}", {"count": str(len(fields))})
        for field in fields:
            ElementTree.SubElement(root, f"{{{self.OOXML_MAIN_NS}}}field", {"x": str(field)})
        return root

    def _pivot_data_fields_element(
        self,
        data_fields: List[Tuple[str, int, Optional[str], str]],
    ) -> ElementTree.Element:
        root = ElementTree.Element(f"{{{self.OOXML_MAIN_NS}}}dataFields", {"count": str(len(data_fields))})
        for name, field_index, subtotal, num_fmt_id in data_fields:
            attrs = {"name": name, "fld": str(field_index), "numFmtId": num_fmt_id}
            if subtotal:
                attrs["subtotal"] = subtotal
            ElementTree.SubElement(root, f"{{{self.OOXML_MAIN_NS}}}dataField", attrs)
        return root

    def _y2y_quantity_only_pivot_table_xml(self, xml_bytes: bytes) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        location = root.find(f"{{{self.OOXML_MAIN_NS}}}location")
        if location is not None:
            location.set("ref", self.Y2Y_RIGHT_PIVOT_LOCATION_REF)

        data_fields = root.find(f"{{{self.OOXML_MAIN_NS}}}dataFields")
        retained_field_indexes: Set[str] = set()
        if data_fields is not None:
            for data_field in list(data_fields):
                field_name = (data_field.attrib.get("name") or "").strip()
                if field_name in self.Y2Y_AMOUNT_FIELD_NAMES:
                    data_fields.remove(data_field)
                    continue
                field_index = data_field.attrib.get("fld")
                if field_index is not None:
                    retained_field_indexes.add(field_index)
            data_fields.set("count", str(len(list(data_fields))))

        self._sync_pivot_data_field_flags(root, retained_field_indexes)
        self._trim_pivot_column_items_for_data_field_count(root, 2)
        return self._serialize_spreadsheet_xml(root)

    def _sync_pivot_data_field_flags(
        self,
        root: ElementTree.Element,
        retained_field_indexes: Set[str],
    ) -> None:
        pivot_fields = root.find(f"{{{self.OOXML_MAIN_NS}}}pivotFields")
        if pivot_fields is None:
            return
        for index, pivot_field in enumerate(pivot_fields.findall(f"{{{self.OOXML_MAIN_NS}}}pivotField")):
            if str(index) in retained_field_indexes:
                pivot_field.set("dataField", "1")
            else:
                pivot_field.attrib.pop("dataField", None)

    def _trim_pivot_column_items_for_data_field_count(
        self,
        root: ElementTree.Element,
        retained_data_field_count: int,
    ) -> None:
        col_items = root.find(f"{{{self.OOXML_MAIN_NS}}}colItems")
        if col_items is None:
            return

        trimmed_items: List[ElementTree.Element] = []
        for item in list(col_items):
            value_nodes = item.findall(f"{{{self.OOXML_MAIN_NS}}}x")
            if len(value_nodes) >= 2:
                trimmed_items.append(item)
                continue
            data_field_value = value_nodes[0].attrib.get("v") if value_nodes else None
            if data_field_value is None:
                trimmed_items.append(item)
                continue
            try:
                data_field_index = int(data_field_value)
            except ValueError:
                trimmed_items.append(item)
                continue
            if data_field_index < retained_data_field_count:
                trimmed_items.append(item)

        for item in list(col_items):
            col_items.remove(item)
        for item in trimmed_items:
            col_items.append(item)
        col_items.set("count", str(len(trimmed_items)))

    def _y2y_quantity_only_chart_xml(self, xml_bytes: bytes) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        retained_series: List[ElementTree.Element] = []
        series_tag = f"{{{self.OOXML_CHART_NS}}}ser"

        for parent in root.iter():
            for child in list(parent):
                if child.tag != series_tag:
                    continue
                series_name = (self._chart_series_name(child) or "").strip()
                if series_name in self.Y2Y_AMOUNT_FIELD_NAMES:
                    parent.remove(child)
                    continue
                retained_series.append(child)

        for index, series in enumerate(retained_series):
            for tag_name in ("idx", "order"):
                node = series.find(f"{{{self.OOXML_CHART_NS}}}{tag_name}")
                if node is not None:
                    node.set("val", str(index))
        return self._serialize_chart_xml(root)

    def _chart_series_name(self, series: ElementTree.Element) -> Optional[str]:
        name_node = series.find(f".//{{{self.OOXML_CHART_NS}}}tx//{{{self.OOXML_CHART_NS}}}v")
        return name_node.text if name_node is not None else None

    def _normalized_y2y_drawing_xml(self, xml_bytes: bytes) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        return self._serialize_drawing_xml(root)

    def _updated_y2y_slicer_cache_xml(self, xml_bytes: bytes) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        pivot_tables = root.find(f"{{{self.OOXML_X14_NS}}}pivotTables")
        if pivot_tables is None:
            pivot_tables = ElementTree.SubElement(root, f"{{{self.OOXML_X14_NS}}}pivotTables")

        existing_names = {
            pivot_table.attrib.get("name")
            for pivot_table in pivot_tables.findall(f"{{{self.OOXML_X14_NS}}}pivotTable")
        }
        for name in ("PivotTable15", "PivotTable16"):
            if name not in existing_names:
                ElementTree.SubElement(
                    pivot_tables,
                    f"{{{self.OOXML_X14_NS}}}pivotTable",
                    {"tabId": "8", "name": name},
                )
        xml = self._serialize_spreadsheet_xml(root)
        return self._ensure_xmlns_prefix(xml, "x", self.OOXML_MAIN_NS)

    @staticmethod
    def _ensure_xmlns_prefix(xml_bytes: bytes, prefix: str, namespace: str) -> bytes:
        declaration = f'xmlns:{prefix}="{namespace}"'
        xml = xml_bytes.decode("utf-8")
        if declaration in xml:
            return xml_bytes

        first_tag_start = xml.find("<", xml.find("?>") + 2 if "?>" in xml[:80] else 0)
        if first_tag_start < 0:
            return xml_bytes
        first_tag_end = xml.find(">", first_tag_start)
        if first_tag_end < 0:
            return xml_bytes
        tag_close_offset = -1 if xml[first_tag_end - 1] == "/" else 0
        insert_at = first_tag_end + tag_close_offset
        xml = f"{xml[:insert_at]} {declaration}{xml[insert_at:]}"
        return xml.encode("utf-8")

    def _updated_workbook_xml(self, xml_bytes: bytes) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        sheets = root.find(f"{{{self.OOXML_MAIN_NS}}}sheets")
        if sheets is not None:
            existing_sheet_names = {
                sheet.attrib.get("name")
                for sheet in sheets.findall(f"{{{self.OOXML_MAIN_NS}}}sheet")
            }
            for _, sheet_name, relationship_id, sheet_id in self.HELPER_WORKSHEETS:
                if sheet_name not in existing_sheet_names:
                    ElementTree.SubElement(
                        sheets,
                        f"{{{self.OOXML_MAIN_NS}}}sheet",
                        {
                            "name": sheet_name,
                            "sheetId": sheet_id,
                            f"{{{self.OOXML_REL_NS}}}id": relationship_id,
                            "state": "hidden",
                        },
                    )

        pivot_caches = root.find(f"{{{self.OOXML_MAIN_NS}}}pivotCaches")
        if pivot_caches is None:
            pivot_caches = ElementTree.SubElement(root, f"{{{self.OOXML_MAIN_NS}}}pivotCaches")
        existing_cache_ids = {
            cache.attrib.get("cacheId")
            for cache in pivot_caches.findall(f"{{{self.OOXML_MAIN_NS}}}pivotCache")
        }
        for cache_id, relationship_id in (
            (self.COUNTRY_PIVOT_CACHE_ID, "rId25"),
            (self.S2S_PIVOT_CACHE_ID, "rId26"),
        ):
            if cache_id not in existing_cache_ids:
                ElementTree.SubElement(
                    pivot_caches,
                    f"{{{self.OOXML_MAIN_NS}}}pivotCache",
                    {"cacheId": cache_id, f"{{{self.OOXML_REL_NS}}}id": relationship_id},
                )

        calc_pr = root.find(f"{{{self.OOXML_MAIN_NS}}}calcPr")
        if calc_pr is None:
            calc_pr = ElementTree.SubElement(root, f"{{{self.OOXML_MAIN_NS}}}calcPr")
        calc_pr.set("fullCalcOnLoad", "1")
        calc_pr.set("forceFullCalc", "1")
        return self._serialize_workbook_xml(root)

    def _updated_workbook_relationships(self, xml_bytes: bytes) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        for relationship in list(root):
            if relationship.attrib.get("Type", "").endswith("/calcChain"):
                root.remove(relationship)

        existing_ids = {relationship.attrib.get("Id") for relationship in root}
        for sheet_path, _, relationship_id, _ in self.HELPER_WORKSHEETS:
            if relationship_id not in existing_ids:
                ElementTree.SubElement(
                    root,
                    f"{{{self.OPC_RELATIONSHIPS_NS}}}Relationship",
                    {
                        "Id": relationship_id,
                        "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/worksheet",
                        "Target": sheet_path.replace("xl/", ""),
                    },
                )
        for cache_index, relationship_id in (("5", "rId25"), ("6", "rId26")):
            if relationship_id not in existing_ids:
                ElementTree.SubElement(
                    root,
                    f"{{{self.OPC_RELATIONSHIPS_NS}}}Relationship",
                    {
                        "Id": relationship_id,
                        "Type": "http://schemas.openxmlformats.org/officeDocument/2006/relationships/pivotCacheDefinition",
                        "Target": f"pivotCache/pivotCacheDefinition{cache_index}.xml",
                    },
                )
        return self._serialize_package_relationships_xml(root)

    def _updated_content_types(self, xml_bytes: bytes) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        for child in list(root):
            if child.attrib.get("PartName") == "/xl/calcChain.xml":
                root.remove(child)

        for sheet_path, _, _, _ in self.HELPER_WORKSHEETS:
            self._ensure_content_type_override(
                root,
                f"/{sheet_path}",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.worksheet+xml",
            )
        for cache_index in ("5", "6"):
            self._ensure_content_type_override(
                root,
                f"/xl/pivotCache/pivotCacheDefinition{cache_index}.xml",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheDefinition+xml",
            )
            self._ensure_content_type_override(
                root,
                f"/xl/pivotCache/pivotCacheRecords{cache_index}.xml",
                "application/vnd.openxmlformats-officedocument.spreadsheetml.pivotCacheRecords+xml",
            )
        return self._serialize_content_types_xml(root)

    def _ensure_content_type_override(
        self,
        root: ElementTree.Element,
        part_name: str,
        content_type: str,
    ) -> None:
        for child in root:
            if child.attrib.get("PartName") == part_name:
                child.set("ContentType", content_type)
                return
        ElementTree.SubElement(
            root,
            f"{{{self.OPC_CONTENT_TYPES_NS}}}Override",
            {"PartName": part_name, "ContentType": content_type},
        )

    @classmethod
    def _without_calc_chain_relationship(cls, xml_bytes: bytes) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        for relationship in list(root):
            if relationship.attrib.get("Type", "").endswith("/calcChain"):
                root.remove(relationship)
        return cls._serialize_package_relationships_xml(root)

    @classmethod
    def _without_pivot_table_relationship(cls, xml_bytes: bytes) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        for relationship in list(root):
            if relationship.attrib.get("Type", "").endswith("/pivotTable"):
                root.remove(relationship)
        return cls._serialize_package_relationships_xml(root)

    @classmethod
    def _without_calc_chain_content_type(cls, xml_bytes: bytes) -> bytes:
        root = ElementTree.fromstring(xml_bytes)
        for child in list(root):
            if child.attrib.get("PartName") == "/xl/calcChain.xml":
                root.remove(child)
        return cls._serialize_content_types_xml(root)

    @classmethod
    def _register_spreadsheet_namespaces(cls) -> None:
        ElementTree.register_namespace("", cls.OOXML_MAIN_NS)
        ElementTree.register_namespace("r", cls.OOXML_REL_NS)
        ElementTree.register_namespace("mc", cls.OOXML_MC_NS)
        ElementTree.register_namespace("x14", cls.OOXML_X14_NS)
        ElementTree.register_namespace("x14ac", cls.OOXML_X14AC_NS)
        ElementTree.register_namespace("x15", cls.OOXML_X15_NS)
        ElementTree.register_namespace("x15ac", cls.OOXML_X15AC_NS)

    @classmethod
    def _serialize_spreadsheet_xml(cls, root: ElementTree.Element) -> bytes:
        cls._register_spreadsheet_namespaces()
        return cls._serialize_xml(root)

    @classmethod
    def _serialize_chart_xml(cls, root: ElementTree.Element) -> bytes:
        ElementTree.register_namespace("c", cls.OOXML_CHART_NS)
        ElementTree.register_namespace("a", cls.OOXML_DRAWING_MAIN_NS)
        ElementTree.register_namespace("r", cls.OOXML_REL_NS)
        ElementTree.register_namespace("mc", cls.OOXML_MC_NS)
        ElementTree.register_namespace("c14", cls.OOXML_CHART_2007_NS)
        return cls._serialize_xml(root)

    @classmethod
    def _serialize_drawing_xml(cls, root: ElementTree.Element) -> bytes:
        ElementTree.register_namespace("", cls.OOXML_SPREADSHEET_DRAWING_NS)
        ElementTree.register_namespace("a", cls.OOXML_DRAWING_MAIN_NS)
        ElementTree.register_namespace("c", cls.OOXML_CHART_NS)
        ElementTree.register_namespace("r", cls.OOXML_REL_NS)
        ElementTree.register_namespace("mc", cls.OOXML_MC_NS)
        ElementTree.register_namespace("a14", cls.OOXML_DRAWING_SLICER_2010_NS)
        return cls._serialize_xml(root)

    @classmethod
    def _serialize_workbook_xml(cls, root: ElementTree.Element) -> bytes:
        cls._register_spreadsheet_namespaces()
        cls._ensure_markup_compat_namespace_declarations(root)
        return cls._serialize_xml(root)

    @classmethod
    def _serialize_content_types_xml(cls, root: ElementTree.Element) -> bytes:
        ElementTree.register_namespace("", cls.OPC_CONTENT_TYPES_NS)
        return cls._serialize_xml(root)

    @classmethod
    def _serialize_package_relationships_xml(cls, root: ElementTree.Element) -> bytes:
        ElementTree.register_namespace("", cls.OPC_RELATIONSHIPS_NS)
        return cls._serialize_xml(root)

    @classmethod
    def _ensure_markup_compat_namespace_declarations(cls, root: ElementTree.Element) -> None:
        prefix_namespaces = {
            "x14": cls.OOXML_X14_NS,
            "x14ac": cls.OOXML_X14AC_NS,
            "x15": cls.OOXML_X15_NS,
            "x15ac": cls.OOXML_X15AC_NS,
        }
        compatibility_attrs = {"Ignorable", "Requires"}

        required_prefixes: Set[str] = set()
        for element in root.iter():
            for attr_name, attr_value in element.attrib.items():
                local_name = attr_name.rsplit("}", 1)[-1] if attr_name.startswith("{") else attr_name
                if local_name not in compatibility_attrs:
                    continue
                for token in attr_value.split():
                    required_prefixes.add(token.split(":", 1)[0])

        for prefix in sorted(required_prefixes):
            namespace_uri = prefix_namespaces.get(prefix)
            if namespace_uri:
                root.set(f"xmlns:{prefix}", namespace_uri)

    @staticmethod
    def _serialize_xml(root: ElementTree.Element) -> bytes:
        return (
            b'<?xml version="1.0" encoding="UTF-8" standalone="yes"?>\n'
            + ElementTree.tostring(root, encoding="utf-8", short_empty_elements=True)
        )

    def _write_rules_sheet(self, wb: openpyxl.Workbook, thin_border: Border) -> None:
        """写入字段来源和取值规则，避免说明行污染 Result 数据区。"""
        ws = wb.create_sheet("Rules")
        rows = [
            ["Field", "Source / Rule"],
            ["Pack", "Factory Price 表 Pack 列：按 Working Number + Season 匹配；命中 Factory Price 记录时使用同一记录 Pack；唯一 Working Number 时允许兜底。"],
            ["Season", "TMS Price 表 Season (M)，按 Working Number (M) + Article Number (A) 匹配。"],
            ["Factory", "优先按 Allocation Factory 表 PO -> Allocation 覆盖；未上传或未匹配时使用 Released / Unreleased 表 Factory。"],
            ["Working Number", "Released / Unreleased 表 P 列 Working Number。"],
            ["Article Number", "Released / Unreleased 表 S 列 Article Number。"],
            ["Article Name", "Released / Unreleased 表 T 列 Article Description。"],
            ["CRD", "Released / Unreleased 表 U 列 Customer Request Date (CRD)。"],
            ["PODD", "Released / Unreleased 表 W 列 PODD。"],
            ["Gps Customer Number", "Released / Unreleased 表 AB 列 Gps Customer Number。"],
            ["Country/Region", "Released / Unreleased 表 AC 列 Country/Region。"],
            ["Shipment Method", "未上传 Shipment Method 文件时使用 Released / Unreleased 表 Shipment Method/Shipment Mode；上传后按 PO 覆盖。"],
            ["Marketing Forecast(M)", "TMS Price 表 Marketing Forecast (M)。"],
            ["Quantity", "默认按 Result 明细维度汇总 Released / Unreleased 表 Ordered Quantity；上传 Shipment Method 后，单一业务组合 PO 按 pord_order_qty 拆分。"],
            ["Factory Price(USD)", "Factory Price 表 G 列，优先按 Season + Working Number + Factory 匹配，失败后按 Working Number + Factory 兜底；多价格时保留首个匹配值并高亮。"],
            ["Factory Amount(USD)", "Excel 公式：Factory Price(USD) * Quantity。"],
            ["TMS Price(USD)", "TMS Price 表 Intl. FOB (C)，优先级 Final > P2 > P1 > PREC；缺失时使用 Factory Price 表 TMS Price 后备。"],
            ["TMS Amount(USD)", "Excel 公式：TMS Price(USD) * Quantity。"],
        ]
        for row in rows:
            ws.append(row)

        header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.freeze_panes = "A2"
        ws.column_dimensions['A'].width = 24
        ws.column_dimensions['B'].width = 96

    def _write_diagnostics_sheet(self, wb: openpyxl.Workbook,
                                 diagnostics: List[Dict[str, Any]],
                                 thin_border: Border) -> None:
        """写入可追踪的异常信息，方便业务核对缺失和高亮原因。"""
        ws = wb.create_sheet("Diagnostics")
        ws.append(["Level", "Type", "Key", "Message"])
        if diagnostics:
            for item in diagnostics:
                ws.append([
                    item.get("level", "WARN"),
                    item.get("type", ""),
                    item.get("key", ""),
                    item.get("message", ""),
                ])
        else:
            ws.append(["INFO", "OK", "", "No diagnostics."])

        header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')
        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(vertical='top', wrap_text=True)
        ws.freeze_panes = "A2"
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 24
        ws.column_dimensions['C'].width = 42
        ws.column_dimensions['D'].width = 90

    @staticmethod
    def _estimate_text_width(value: Any) -> int:
        if value is None:
            return 0
        return sum(2 if ord(char) > 255 else 1 for char in str(value))

    def _adjust_column_widths(self, ws: openpyxl.worksheet.worksheet.Worksheet,
                             max_widths: Optional[Dict[int, int]] = None,
                             min_widths: Optional[Dict[int, int]] = None) -> None:
        """自动调整Excel列宽"""
        if max_widths is None:
            max_widths = {}
        if min_widths is None:
            min_widths = {}

        for col_idx in range(1, ws.max_column + 1):
            max_length = 0
            column_letter = openpyxl.utils.get_column_letter(col_idx)
            for row in range(1, ws.max_row + 1):
                cell = ws.cell(row=row, column=col_idx)
                try:
                    max_length = max(max_length, self._estimate_text_width(cell.value))
                except Exception:
                    pass
            min_width = min_widths.get(col_idx, 9)
            max_width = max_widths.get(col_idx, 34)
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    @classmethod
    def _normalize_header(cls, value: Any) -> str:
        text = cls._clean_text(value).lower()
        return ''.join(char for char in text if char.isalnum())

    @classmethod
    def _find_column(cls, df: pd.DataFrame, aliases: List[str]) -> Optional[str]:
        normalized_aliases = {cls._normalize_header(alias) for alias in aliases}
        for column in df.columns:
            if cls._normalize_header(column) in normalized_aliases:
                return column
        return None

    @classmethod
    def _require_column(cls, df: pd.DataFrame, aliases: List[str], source_name: str) -> str:
        column = cls._find_column(df, aliases)
        if column is None:
            raise ValueError(f"{source_name} 缺少: {' / '.join(aliases)}")
        return column

    def read_excel_files_with_detected_header(
        self,
        file_list: List[str],
        required_aliases: List[List[str]],
        scan_all_sheets: bool = False,
    ) -> Tuple[Optional[pd.DataFrame], List[str]]:
        """读取标题行位置不固定的 Excel 文件，按字段名而不是固定行列定位。"""
        dfs: List[pd.DataFrame] = []
        logs: List[str] = []

        for file_path in file_list:
            file_name = os.path.basename(file_path)
            try:
                if scan_all_sheets:
                    # Factory Price 文件可能把辅助 sheet 放在最前面，需按必需字段选择第一个有效 sheet。
                    raw_sheets = pd.read_excel(file_path, sheet_name=None, header=None)
                    selected_df: Optional[pd.DataFrame] = None
                    for sheet_name, raw_df in raw_sheets.items():
                        header_row_index = self._detect_header_row(raw_df, required_aliases)
                        if header_row_index is None:
                            continue
                        selected_df = self._build_detected_header_dataframe(raw_df, header_row_index)
                        dfs.append(selected_df)
                        logs.append(
                            f"  - 读取成功：{file_name} / {sheet_name}，"
                            f"标题行 {header_row_index + 1}，共 {len(selected_df)} 行"
                        )
                        break
                    if selected_df is None:
                        logs.append(f"  ⚠️ 未找到标题行：{file_name}")
                    continue

                raw_df = pd.read_excel(file_path, header=None)
                header_row_index = self._detect_header_row(raw_df, required_aliases)
                if header_row_index is None:
                    logs.append(f"  ⚠️ 未找到标题行：{file_name}")
                    continue
                df = self._build_detected_header_dataframe(raw_df, header_row_index)
                dfs.append(df)
                logs.append(
                    f"  - 读取成功：{file_name}，标题行 {header_row_index + 1}，共 {len(df)} 行"
                )
            except Exception as e:
                logs.append(f"  ⚠️ 读取失败：{file_name}，错误：{str(e)}")

        if not dfs:
            return None, logs
        if len(dfs) == 1:
            return dfs[0], logs

        result = pd.concat(dfs, axis=0, ignore_index=True)
        logs.append(f"  ✅ 合并成功，总计 {len(result)} 行")
        return result, logs

    @staticmethod
    def _build_detected_header_dataframe(raw_df: pd.DataFrame, header_row_index: int) -> pd.DataFrame:
        headers = raw_df.iloc[header_row_index].tolist()
        df = raw_df.iloc[header_row_index + 1:].copy()
        df.columns = headers
        return df.dropna(how='all')

    def _detect_header_row(self, raw_df: pd.DataFrame, required_aliases: List[List[str]]) -> Optional[int]:
        for row_index in range(min(len(raw_df), 12)):
            normalized_headers = {
                self._normalize_header(value)
                for value in raw_df.iloc[row_index].tolist()
                if not self._is_blank(value)
            }
            if all(
                any(self._normalize_header(alias) in normalized_headers for alias in aliases)
                for aliases in required_aliases
            ):
                return row_index
        return None

    @classmethod
    def _to_datetime(cls, value: Any) -> Optional[datetime]:
        if cls._is_blank(value):
            return None
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime().replace(tzinfo=None)
        if isinstance(value, datetime):
            return value.replace(tzinfo=None)
        try:
            parsed = pd.to_datetime(value)
        except (TypeError, ValueError):
            return None
        if pd.isna(parsed):
            return None
        return parsed.to_pydatetime().replace(tzinfo=None)

    @classmethod
    def _country_bucket(cls, value: Any) -> str:
        country = cls._clean_text(value).upper()
        if country == "CHINA":
            return "CHINA"
        if country in {"UNITED STATES", "USA", "US"}:
            return "UNITED STATES"
        return "OTHER COUNTRIES"

    @staticmethod
    def _empty_metrics() -> Dict[str, float]:
        return {"quantity": 0.0, "factory_amount": 0.0, "tms_amount": 0.0}

    @classmethod
    def _add_metrics(cls, target: Dict[str, float], row: Dict[str, Any]) -> None:
        quantity = cls._optional_float(row.get("Quantity")) or 0.0
        factory_amount = cls._optional_float(row.get("_factory_amount_value")) or 0.0
        tms_amount = cls._optional_float(row.get("_tms_amount_value")) or 0.0
        target["quantity"] += quantity
        target["factory_amount"] += factory_amount
        target["tms_amount"] += tms_amount

    def _append_summary_table(
        self,
        ws: openpyxl.worksheet.worksheet.Worksheet,
        headers: List[str],
        rows: List[List[Any]],
        thin_border: Border,
    ) -> None:
        header_fill = PatternFill(start_color='FF4472C4', end_color='FF4472C4', fill_type='solid')
        ws.append(headers)
        for row in rows:
            ws.append(row)

        for cell in ws[1]:
            cell.font = Font(bold=True, color='FFFFFFFF')
            cell.fill = header_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal='center', vertical='center')

        for row in ws.iter_rows(min_row=2):
            for cell in row:
                cell.border = thin_border
                cell.alignment = Alignment(horizontal='center', vertical='center')
                if isinstance(cell.value, datetime):
                    cell.number_format = 'yyyy-mm-dd'
                elif isinstance(cell.value, float):
                    cell.number_format = '#,##0.00'

        ws.freeze_panes = "A2"
        self._adjust_column_widths(ws, min_widths={idx: 13 for idx in range(1, len(headers) + 1)})

    def _write_summary_sheets(
        self,
        wb: openpyxl.Workbook,
        results: List[Dict[str, Any]],
        development_style_rows: List[Dict[str, Any]],
        thin_border: Border,
    ) -> None:
        default_sheet = wb.active
        wb.remove(default_sheet)
        self._write_monthly_summary(wb, results, thin_border)
        self._write_seasonal_summary(wb, results, thin_border)
        self._write_country_analysis(wb, results, "Country Analysis (S)", "Season", thin_border)
        self._write_country_analysis(wb, results, "Country Analysis (Y)", "Year", thin_border)
        self._write_y2y_comparison(wb, results, thin_border)
        self._write_factory_order_analysis(wb, results, thin_border)
        self._write_ship_method_analysis(wb, results, thin_border)
        self._write_development_style_qty(wb, development_style_rows, thin_border)
        self._write_s2s_development_analysis(wb, results, development_style_rows, thin_border)

    def _write_monthly_summary(
        self,
        wb: openpyxl.Workbook,
        results: List[Dict[str, Any]],
        thin_border: Border,
    ) -> None:
        groups: Dict[Tuple[int, int, str], Dict[str, float]] = defaultdict(self._empty_metrics)
        for row in results:
            podd = self._to_datetime(row.get("PODD"))
            if podd is None:
                continue
            key = (podd.year, podd.month, self._clean_text(row.get("Factory")))
            self._add_metrics(groups[key], row)

        rows = [
            [year, datetime(2000, month, 1).strftime("%b"), factory, metrics["quantity"], metrics["factory_amount"], metrics["tms_amount"]]
            for (year, month, factory), metrics in sorted(groups.items())
        ]
        ws = wb.create_sheet("Monthly Summary (By Fty)")
        self._append_summary_table(
            ws,
            ["Years", "PODD", "Factory", "Quantity (Y)", "Factory Amount (USD)", "TMS Amount (USD)"],
            rows,
            thin_border,
        )

    def _write_seasonal_summary(
        self,
        wb: openpyxl.Workbook,
        results: List[Dict[str, Any]],
        thin_border: Border,
    ) -> None:
        groups: Dict[Tuple[str, str], Dict[str, float]] = defaultdict(self._empty_metrics)
        for row in results:
            key = (self._clean_text(row.get("Season")), self._clean_text(row.get("Factory")))
            self._add_metrics(groups[key], row)

        rows = [
            [season, factory, metrics["quantity"], metrics["factory_amount"], metrics["tms_amount"]]
            for (season, factory), metrics in sorted(
                groups.items(),
                key=lambda item: self._season_factory_sort_key(item[0]),
            )
        ]
        ws = wb.create_sheet("Seasonal Summary (By Fty)")
        self._append_summary_table(
            ws,
            ["Season", "Factory", "Quantity (Y)", "Factory Amount (USD)", "TMS Amount (USD)"],
            rows,
            thin_border,
        )

    def _write_country_analysis(
        self,
        wb: openpyxl.Workbook,
        results: List[Dict[str, Any]],
        title: str,
        period_kind: str,
        thin_border: Border,
    ) -> None:
        groups: Dict[Tuple[Any, str, str], Dict[str, float]] = defaultdict(self._empty_metrics)
        totals: Dict[Tuple[Any, str], Dict[str, float]] = defaultdict(self._empty_metrics)
        for row in results:
            podd = self._to_datetime(row.get("PODD"))
            if period_kind == "Year":
                if podd is None:
                    continue
                period = podd.year
            else:
                period = self._clean_text(row.get("Season"))
            factory = self._clean_text(row.get("Factory"))
            bucket = self._country_bucket(row.get("Country/Region"))
            self._add_metrics(groups[(period, factory, bucket)], row)
            self._add_metrics(totals[(period, factory)], row)

        rows: List[List[Any]] = []
        for period, factory in sorted(
            totals,
            key=lambda key: self._country_summary_sort_key(key, period_kind),
        ):
            total = totals[(period, factory)]
            row_values: List[Any] = [period, factory]
            for bucket in ("CHINA", "UNITED STATES", "OTHER COUNTRIES"):
                metrics = groups[(period, factory, bucket)]
                quantity_pct = metrics["quantity"] / total["quantity"] if total["quantity"] else 0
                amount_pct = metrics["tms_amount"] / total["tms_amount"] if total["tms_amount"] else 0
                row_values.extend([metrics["quantity"], quantity_pct, metrics["tms_amount"], amount_pct])
            row_values.extend([total["quantity"], 1 if total["quantity"] else 0, total["tms_amount"], 1 if total["tms_amount"] else 0])
            rows.append(row_values)

        ws = wb.create_sheet(title)
        self._append_summary_table(
            ws,
            [
                period_kind,
                "Factory",
                "CHINA Quantity (Y)",
                "CHINA Quantity (%)",
                "CHINA TMS Amount (USD)",
                "CHINA TMS Amount (%)",
                "UNITED STATES Quantity (Y)",
                "UNITED STATES Quantity (%)",
                "UNITED STATES TMS Amount (USD)",
                "UNITED STATES TMS Amount (%)",
                "OTHER COUNTRIES Quantity (Y)",
                "OTHER COUNTRIES Quantity (%)",
                "OTHER COUNTRIES TMS Amount (USD)",
                "OTHER COUNTRIES TMS Amount (%)",
                "Total Quantity (Y)",
                "Total Quantity (%)",
                "Total TMS Amount (USD)",
                "Total TMS Amount (%)",
            ],
            rows,
            thin_border,
        )

    def _write_y2y_comparison(
        self,
        wb: openpyxl.Workbook,
        results: List[Dict[str, Any]],
        thin_border: Border,
    ) -> None:
        groups: Dict[Tuple[int, int], Dict[str, float]] = defaultdict(self._empty_metrics)
        for row in results:
            podd = self._to_datetime(row.get("PODD"))
            if podd is None:
                continue
            self._add_metrics(groups[(podd.year, podd.month)], row)

        rows: List[List[Any]] = []
        for year, month in sorted(groups):
            metrics = groups[(year, month)]
            previous = groups.get((year - 1, month))
            qty_y2y = (
                (metrics["quantity"] - previous["quantity"]) / previous["quantity"]
                if previous and previous["quantity"]
                else ""
            )
            amount_y2y = (
                (metrics["tms_amount"] - previous["tms_amount"]) / previous["tms_amount"]
                if previous and previous["tms_amount"]
                else ""
            )
            rows.append([year, datetime(2000, month, 1).strftime("%b"), metrics["quantity"], qty_y2y, metrics["tms_amount"], amount_y2y])

        ws = wb.create_sheet("Y2Y Comparison(Qtty & Value)")
        self._append_summary_table(
            ws,
            ["Years", "PODD", "Quantity (Y)", "Y2Y - Qty", "TMS Amount (USD)", "Y2Y - Amount"],
            rows,
            thin_border,
        )

    def _write_factory_order_analysis(
        self,
        wb: openpyxl.Workbook,
        results: List[Dict[str, Any]],
        thin_border: Border,
    ) -> None:
        groups: Dict[str, Dict[str, float]] = defaultdict(self._empty_metrics)
        for row in results:
            self._add_metrics(groups[self._clean_text(row.get("Factory"))], row)

        rows = []
        for factory, metrics in sorted(groups.items()):
            quantity = metrics["quantity"]
            avg_factory = metrics["factory_amount"] / quantity if quantity else 0
            avg_tms = metrics["tms_amount"] / quantity if quantity else 0
            rows.append([factory, quantity, avg_factory, metrics["factory_amount"], avg_tms, metrics["tms_amount"]])

        ws = wb.create_sheet("Fty Order Analysis")
        self._append_summary_table(
            ws,
            ["Factory", "Quantity (Y)", "Average Fty px (USD)", "Factory Amount (USD)", "Average TMS px (USD)", "TMS Amount (USD)"],
            rows,
            thin_border,
        )

    def _write_ship_method_analysis(
        self,
        wb: openpyxl.Workbook,
        results: List[Dict[str, Any]],
        thin_border: Border,
    ) -> None:
        rows = [row["values"] for row in self._ship_method_analysis_rows(results)]

        ws = wb.create_sheet("Ship Method Analysis")
        self._append_summary_table(
            ws,
            ["Factory", "Years", "PODD", "Shipment Method", "Quantity (Y)", "Quantity (%)", "TMS Amount (USD)", "TMS Amount (%)"],
            rows,
            thin_border,
        )

    def _write_development_style_qty(
        self,
        wb: openpyxl.Workbook,
        development_style_rows: List[Dict[str, Any]],
        thin_border: Border,
    ) -> None:
        ws = wb.create_sheet("Development Style Qty")
        self._append_summary_table(
            ws,
            ["Season", "Factory", "Development Style Count"],
            self._development_style_qty_rows(development_style_rows),
            thin_border,
        )

    def _write_s2s_development_analysis(
        self,
        wb: openpyxl.Workbook,
        results: List[Dict[str, Any]],
        development_style_rows: List[Dict[str, Any]],
        thin_border: Border,
    ) -> None:
        rows = self._s2s_development_source_rows(results, development_style_rows)
        ws = wb.create_sheet("S2S Development Analysis")
        self._append_summary_table(
            ws,
            ["Season", "Factory", "Development Style Count", "Bulk Style Count", "Bulk Qty (pcs)"],
            rows,
            thin_border,
        )

    @staticmethod
    def _clean_key(value: Any) -> str:
        """Normalize Excel key values and avoid treating blank/NaN as real keys."""
        if pd.isna(value):
            return ""
        text = str(value).strip()
        if text.lower() in {"nan", "none"}:
            return ""
        return text

    @classmethod
    def _make_article_key(cls, working_num: Any, article_num: Any) -> Tuple[str, str]:
        return (cls._clean_key(working_num).upper(), cls._clean_key(article_num).upper())

    @staticmethod
    def _is_blank(value: Any) -> bool:
        if value is None:
            return True
        try:
            if pd.isna(value):
                return True
        except (TypeError, ValueError):
            pass
        return str(value).strip().lower() in {"", "nan", "none", "nat"}

    @classmethod
    def _clean_text(cls, value: Any) -> str:
        if cls._is_blank(value):
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, int) and not isinstance(value, bool):
            return str(value)
        return str(value).strip()

    @classmethod
    def _clean_output_value(cls, value: Any) -> Any:
        if cls._is_blank(value):
            return ""
        if isinstance(value, pd.Timestamp):
            return value.to_pydatetime().replace(tzinfo=None)
        if isinstance(value, datetime):
            return value.replace(tzinfo=None)
        return value

    @classmethod
    def _group_key_value(cls, value: Any) -> str:
        if cls._is_blank(value):
            return ""
        if isinstance(value, pd.Timestamp):
            return value.strftime('%Y-%m-%d')
        if isinstance(value, datetime):
            return value.strftime('%Y-%m-%d')
        return str(value).strip().upper()

    @classmethod
    def _optional_float(cls, value: Any) -> Optional[float]:
        if cls._is_blank(value):
            return None
        try:
            if isinstance(value, str):
                return float(value.strip().replace(",", ""))
            return float(value)
        except (TypeError, ValueError):
            return None

    @classmethod
    def _article_priority_rank(cls, milestone: Any) -> int:
        key = cls._clean_key(milestone).upper().replace(" ", "")
        return cls.ARTICLE_PRIORITY.get(key, 99)

    @staticmethod
    def _select_article_candidate(candidates: List[Dict[str, Any]]) -> Dict[str, Any]:
        """严格按 Milestone 顺序取值：Final > P2 > P1 > PREC。"""
        for milestone in ("FINAL", "P2", "P1", "PREC"):
            for item in candidates:
                if item['milestone'].upper().replace(" ", "") == milestone:
                    return item

        return candidates[0]


    @classmethod
    def _make_factory_group_key(cls, season: Any, working_num: Any, factory: Any) -> Tuple[str, str, str]:
        return (
            cls._clean_key(season).upper(),
            cls._clean_key(working_num).upper(),
            cls._clean_key(factory).upper(),
        )

    @classmethod
    def _make_factory_article_key(cls, season: Any, working_num: Any,
                                  factory: Any, article_num: Any) -> Tuple[str, str, str, str]:
        return cls._make_factory_group_key(season, working_num, factory) + (
            cls._clean_key(article_num).upper(),
        )

    def process_reports(
        self,
        tms_paths: List[str],
        tms_price_paths: List[str],
        price_paths: List[str],
        pack_paths: Optional[List[str]] = None,
        output_dir: Optional[str] = None,
        allocation_paths: Optional[List[str]] = None,
        shipment_method_paths: Optional[List[str]] = None,
    ) -> Dict[str, Any]:
        """主处理流程"""

        result = {
            'success': False,
            'message': '',
            'logs': [],
            'output_path': None,
            'working_count': 0,
            'result_count': 0,
            'diagnostics_count': 0
        }

        def log(msg: str):
            result['logs'].append(msg)

        try:
            log("=" * 80)
            log("开始生成 Sophia & Tina 报表")
            log("=" * 80)

            # 读取各类文件
            log("\n📖 正在读取 TMS 文件...")
            tms_df, tms_logs = self.read_excel_files(tms_paths)
            result['logs'].extend(tms_logs)

            if tms_df is None:
                result['message'] = 'TMS 文件读取失败'
                return result

            log(f"✅ TMS 文件读取完成，共 {len(tms_df)} 行数据")

            log("\n📖 正在读取 TMS Price 文件...")
            tms_price_df, tms_price_logs = self.read_excel_files(tms_price_paths)
            result['logs'].extend(tms_price_logs)

            if tms_price_df is None:
                result['message'] = 'TMS Price 文件读取失败'
                return result

            log(f"✅ TMS Price 文件读取完成，共 {len(tms_price_df)} 行数据")

            log("\n📖 正在读取 Factory Price 文件...")
            price_df, price_logs = self.read_excel_files_with_detected_header(
                price_paths,
                [["Season"], ["Working Number"], ["Article Number"], ["Factory"], ["Factory Price"]],
                scan_all_sheets=True,
            )
            result['logs'].extend(price_logs)

            if price_df is None:
                result['message'] = 'Factory Price 文件读取失败'
                return result

            log(f"✅ Factory Price 文件读取完成，共 {len(price_df)} 行数据")

            pack_df: Optional[pd.DataFrame] = None
            if pack_paths:
                log("\n📖 正在读取兼容 Pack 文件...")
                pack_df, pack_logs = self.read_excel_files(pack_paths)
                result['logs'].extend(pack_logs)
                if pack_df is not None:
                    log(f"✅ Pack 文件读取完成，共 {len(pack_df)} 行数据")

            allocation_df: Optional[pd.DataFrame] = None
            if allocation_paths:
                log("\n📖 正在读取 Allocation Factory 文件...")
                allocation_df, allocation_logs = self.read_excel_files_with_detected_header(
                    allocation_paths,
                    [["Allocation"], ["PO"]],
                )
                result['logs'].extend(allocation_logs)
                if allocation_df is not None:
                    log(f"✅ Allocation Factory 文件读取完成，共 {len(allocation_df)} 行数据")

            shipment_df: Optional[pd.DataFrame] = None
            if shipment_method_paths:
                log("\n📖 正在读取 Shipment Method 文件...")
                shipment_df, shipment_logs = self.read_excel_files_with_detected_header(
                    shipment_method_paths,
                    [["purchasingdocument"], ["shippinginstruction_desc"], ["po_delivery_date"], ["pord_order_qty", "Sum of pord_order_qty"]],
                )
                result['logs'].extend(shipment_logs)
                if shipment_df is not None:
                    log(f"✅ Shipment Method 文件读取完成，共 {len(shipment_df)} 行数据")

            # 检查必要列
            log(f"\n🔍 正在检查必要的列...")
            try:
                tms_columns = {
                    "factory": self._require_column(tms_df, ["Factory"], "TMS"),
                    "po": self._find_column(tms_df, ["PO Number"]),
                    "working": self._require_column(tms_df, ["Working Number"], "TMS"),
                    "article": self._require_column(tms_df, ["Article Number"], "TMS"),
                    "article_name": self._require_column(tms_df, ["Article Description"], "TMS"),
                    "crd": self._require_column(tms_df, ["Customer Request Date (CRD)"], "TMS"),
                    "podd": self._require_column(tms_df, ["PODD"], "TMS"),
                    "shipment": self._find_column(tms_df, ["Shipment Method", "Shipment Mode", "Shipment Method/Shipment Mode"]),
                    "gps": self._require_column(tms_df, ["Gps Customer Number"], "TMS"),
                    "country": self._require_column(tms_df, ["Country/Region"], "TMS"),
                    "quantity": self._require_column(tms_df, ["Ordered Quantity"], "TMS"),
                }
                tms_price_columns = {
                    "working": self._require_column(tms_price_df, ["Working Number (M)"], "TMS Price"),
                    "article": self._require_column(tms_price_df, ["Article Number (A)"], "TMS Price"),
                    "season": self._require_column(tms_price_df, ["Season (M)"], "TMS Price"),
                    "marketing_forecast": self._require_column(tms_price_df, ["Marketing Forecast (M)"], "TMS Price"),
                    "milestone": self._require_column(tms_price_df, ["Milestone (C)"], "TMS Price"),
                    "tms_price": self._require_column(tms_price_df, ["Intl. FOB (C)"], "TMS Price"),
                    "factory_group": self._find_column(tms_price_df, ["Factory Group Code (MF)"]),
                }
                price_columns = {
                    "pack": self._find_column(price_df, ["Pack"]),
                    "season": self._require_column(price_df, ["Season"], "Factory Price"),
                    "working": self._require_column(price_df, ["Working Number"], "Factory Price"),
                    "article": self._require_column(price_df, ["Article Number"], "Factory Price"),
                    "factory": self._require_column(price_df, ["Factory"], "Factory Price"),
                    "factory_price": self._require_column(price_df, ["Factory Price"], "Factory Price"),
                    "tms_price": self._find_column(price_df, ["TMS Price"]),
                }
            except ValueError as exc:
                msg = "缺少必要的列：\n" + str(exc)
                log(f"❌ {msg}")
                result['message'] = msg
                return result

            log("✅ 所有必要列检查通过")

            log(f"\n🔄 正在按 TMS 明细维度汇总 Ordered Quantity...")
            log(f"📊 TMS 文件共 {len(tms_df)} 行数据")

            diagnostics: List[Dict[str, Any]] = []
            seen_diagnostics: Set[Tuple[str, str, str]] = set()

            def add_diagnostic(level: str, diag_type: str, key: str, message: str) -> None:
                # 同一个问题只写一次，避免 Diagnostics 被明细行重复撑大。
                dedupe_key = (diag_type, key, message)
                if dedupe_key in seen_diagnostics:
                    return
                seen_diagnostics.add(dedupe_key)
                diagnostics.append({
                    "level": level,
                    "type": diag_type,
                    "key": key,
                    "message": message,
                })

            working_numbers: Set[str] = set()
            used_article_keys: Set[Tuple[str, str]] = set()
            tms_groups: Dict[Tuple[str, ...], Dict[str, Any]] = {}
            po_combo_keys: Dict[str, Set[Tuple[str, ...]]] = defaultdict(set)

            for _, row in tms_df.iterrows():
                working_num = self._clean_text(row[tms_columns["working"]])
                article_num = self._clean_text(row[tms_columns["article"]])
                if not working_num or not article_num:
                    continue

                po_number = self._clean_text(row[tms_columns["po"]]) if tms_columns["po"] else ""
                source_factory = self._clean_text(row[tms_columns["factory"]])
                shipment_method = self._clean_text(row[tms_columns["shipment"]]) if tms_columns["shipment"] else ""
                working_numbers.add(working_num.upper())
                used_article_keys.add((working_num.upper(), article_num.upper()))
                group_values = {
                    '_po_number': po_number,
                    '_source_factory': source_factory,
                    'Factory': source_factory,
                    'Working Number': working_num,
                    'Article Number': article_num,
                    'Article Name': self._clean_text(row[tms_columns["article_name"]]),
                    'CRD': self._clean_output_value(row[tms_columns["crd"]]),
                    'PODD': self._clean_output_value(row[tms_columns["podd"]]),
                    'Shipment Method': shipment_method,
                    'Gps Customer Number': self._clean_text(row[tms_columns["gps"]]),
                    'Country/Region': self._clean_text(row[tms_columns["country"]]),
                }
                key = tuple([
                    self._group_key_value(po_number),
                    self._group_key_value(source_factory),
                    self._group_key_value(working_num),
                    self._group_key_value(article_num),
                    self._group_key_value(row[tms_columns["article_name"]]),
                    self._group_key_value(row[tms_columns["crd"]]),
                    self._group_key_value(row[tms_columns["podd"]]),
                    self._group_key_value(shipment_method),
                    self._group_key_value(row[tms_columns["gps"]]),
                    self._group_key_value(row[tms_columns["country"]]),
                ])
                if key not in tms_groups:
                    tms_groups[key] = {
                        **group_values,
                        'Quantity': 0.0,
                    }
                combo_key = (
                    self._group_key_value(source_factory),
                    self._group_key_value(working_num),
                    self._group_key_value(article_num),
                    self._group_key_value(row[tms_columns["gps"]]),
                    self._group_key_value(row[tms_columns["country"]]),
                )
                if po_number:
                    po_combo_keys[po_number.upper()].add(combo_key)
                qty = self._optional_float(row[tms_columns["quantity"]]) or 0.0
                tms_groups[key]['Quantity'] += qty

            log(f"✅ TMS 明细汇总完成：{len(tms_groups)} 行，{len(working_numbers)} 个 Working Number")

            allocation_lookup: Dict[str, str] = {}
            if allocation_df is not None:
                allocation_po_col = self._require_column(allocation_df, ["PO"], "Allocation Factory")
                allocation_col = self._require_column(allocation_df, ["Allocation"], "Allocation Factory")
                allocation_conflicts: Dict[str, Set[str]] = defaultdict(set)
                for _, row in allocation_df.iterrows():
                    po_key = self._clean_key(row[allocation_po_col]).upper()
                    allocation_value = self._clean_text(row[allocation_col])
                    if not po_key or not allocation_value:
                        continue
                    allocation_conflicts[po_key].add(allocation_value)
                    if po_key not in allocation_lookup:
                        allocation_lookup[po_key] = allocation_value
                for po_key, values in allocation_conflicts.items():
                    if len(values) > 1:
                        add_diagnostic(
                            "WARN",
                            "ALLOCATION_PO_CONFLICT",
                            po_key,
                            f"Allocation Factory 同一个 PO 存在多个 Allocation：{sorted(values)}；Result 使用首个值。",
                        )

            shipment_lookup: Dict[str, List[Dict[str, Any]]] = defaultdict(list)
            if shipment_df is not None:
                shipment_po_col = self._require_column(shipment_df, ["purchasingdocument"], "Shipment Method")
                shipment_method_col = self._require_column(shipment_df, ["shippinginstruction_desc"], "Shipment Method")
                shipment_podd_col = self._require_column(shipment_df, ["po_delivery_date"], "Shipment Method")
                shipment_qty_col = self._require_column(
                    shipment_df,
                    ["pord_order_qty", "Sum of pord_order_qty"],
                    "Shipment Method",
                )
                for _, row in shipment_df.iterrows():
                    po_key = self._clean_key(row[shipment_po_col]).upper()
                    if not po_key:
                        continue
                    shipment_lookup[po_key].append({
                        "Shipment Method": self._clean_text(row[shipment_method_col]),
                        "PODD": self._clean_output_value(row[shipment_podd_col]),
                        "Quantity": self._optional_float(row[shipment_qty_col]),
                    })

            expanded_groups: List[Dict[str, Any]] = []
            for group in tms_groups.values():
                po_key = self._clean_key(group.get("_po_number")).upper()
                shipment_records = shipment_lookup.get(po_key, [])
                if shipment_records and len(po_combo_keys.get(po_key, set())) <= 1:
                    for shipment_record in shipment_records:
                        next_group = dict(group)
                        next_group["Shipment Method"] = shipment_record["Shipment Method"]
                        next_group["PODD"] = shipment_record["PODD"]
                        if shipment_record["Quantity"] is not None:
                            next_group["Quantity"] = shipment_record["Quantity"]
                        expanded_groups.append(next_group)
                else:
                    if shipment_records and len(po_combo_keys.get(po_key, set())) > 1:
                        add_diagnostic(
                            "WARN",
                            "SHIPMENT_PO_MULTI_COMBO",
                            po_key,
                            "Shipment Method 表有该 PO，但 TMS 中同一 PO 对应多个 Article/客户组合；已保留 TMS 原数量，未按 Shipment 拆分。",
                        )
                    expanded_groups.append(group)

            log("📊 正在构建 TMS Price / Factory Price / Pack 查找字典...")

            article_candidates: Dict[Tuple[str, str], List[Dict[str, Any]]] = {}
            development_working_by_season_factory: Dict[Tuple[str, str], Set[str]] = defaultdict(set)
            for _, row in tms_price_df.iterrows():
                key = self._make_article_key(row[tms_price_columns["working"]], row[tms_price_columns["article"]])
                if not key[0] or not key[1]:
                    continue
                article_candidates.setdefault(key, []).append({
                    'season': self._clean_text(row[tms_price_columns["season"]]),
                    'marketing_forecast': self._clean_output_value(row[tms_price_columns["marketing_forecast"]]),
                    'tms_price': self._optional_float(row[tms_price_columns["tms_price"]]),
                    'milestone': self._clean_text(row[tms_price_columns["milestone"]]),
                    'rank': self._article_priority_rank(row[tms_price_columns["milestone"]]),
                })
                factory_group_col = tms_price_columns.get("factory_group")
                if factory_group_col:
                    season = self._clean_text(row[tms_price_columns["season"]])
                    factory_group = self._clean_text(row[factory_group_col])
                    working = self._clean_text(row[tms_price_columns["working"]])
                    if season and factory_group and working:
                        development_working_by_season_factory[(season, factory_group)].add(working)

            article_lookup: Dict[Tuple[str, str], Dict[str, Any]] = {}
            for key, candidates in article_candidates.items():
                selected = self._select_article_candidate(candidates)
                article_lookup[key] = selected
                if len(candidates) > 1 and key in used_article_keys:
                    milestones = sorted({item['milestone'] for item in candidates if item['milestone']})
                    add_diagnostic(
                        "INFO",
                        "ARTICLE_PRIORITY",
                        " / ".join(key),
                        f"Article 存在 {len(candidates)} 条成本记录，已按 Final > P2 > P1 > PREC 选择 {selected['milestone']}；候选 Milestone：{milestones}。",
                    )

            pack_lookup: Dict[Tuple[str, str], Any] = {}
            pack_by_working: Dict[str, Set[Any]] = {}
            price_group_records: Dict[Tuple[str, str, str], List[Dict[str, Any]]] = defaultdict(list)
            price_factory_records: Dict[Tuple[str, str], List[Dict[str, Any]]] = defaultdict(list)
            for _, row in price_df.iterrows():
                working_key = self._clean_key(row[price_columns["working"]]).upper()
                season_key = self._clean_key(row[price_columns["season"]]).upper()
                factory_key = self._clean_key(row[price_columns["factory"]]).upper()
                article_key = self._clean_key(row[price_columns["article"]]).upper()
                factory_price = self._optional_float(row[price_columns["factory_price"]])
                tms_price_fallback = (
                    self._optional_float(row[price_columns["tms_price"]])
                    if price_columns.get("tms_price")
                    else None
                )
                pack_value = (
                    self._clean_output_value(row[price_columns["pack"]])
                    if price_columns.get("pack")
                    else ""
                )
                if not working_key:
                    continue
                pack_by_working.setdefault(working_key, set()).add(pack_value)
                if season_key and (working_key, season_key) not in pack_lookup:
                    pack_lookup[(working_key, season_key)] = pack_value
                if factory_price is None:
                    continue
                price_record = {
                    "pack": pack_value,
                    "factory_price": factory_price,
                    "tms_price": tms_price_fallback,
                    "factory": factory_key,
                    "article": article_key,
                }
                if factory_key:
                    price_factory_records[(working_key, factory_key)].append(price_record)
                group_key = self._make_factory_group_key(season_key, working_key, factory_key)
                if not all(group_key):
                    continue
                price_group_records[group_key].append(price_record)

            if pack_df is not None:
                pack_working_col = self._find_column(pack_df, ["Working Number"])
                pack_season_col = self._find_column(pack_df, ["Season"])
                pack_col = self._find_column(pack_df, ["Pack"])
                if pack_working_col and pack_season_col and pack_col:
                    for _, row in pack_df.iterrows():
                        working_key = self._clean_key(row[pack_working_col]).upper()
                        season_key = self._clean_key(row[pack_season_col]).upper()
                        pack_value = self._clean_output_value(row[pack_col])
                        if not working_key:
                            continue
                        pack_by_working.setdefault(working_key, set()).add(pack_value)
                        if season_key and (working_key, season_key) not in pack_lookup:
                            pack_lookup[(working_key, season_key)] = pack_value

            price_group_lookup = {
                key: records[0]
                for key, records in price_group_records.items()
            }
            price_factory_lookup = {
                key: records[0]
                for key, records in price_factory_records.items()
            }
            factory_conflict_groups = {
                key: {record["factory_price"] for record in records}
                for key, records in price_group_records.items()
                if len({record["factory_price"] for record in records}) > 1
            }
            factory_fallback_conflict_groups = {
                key: {record["factory_price"] for record in records}
                for key, records in price_factory_records.items()
                if len({record["factory_price"] for record in records}) > 1
            }
            development_style_rows = [
                {
                    "Season": season,
                    "Factory": factory,
                    "Development Style Count": len(workings),
                }
                for (season, factory), workings in sorted(development_working_by_season_factory.items())
            ]

            log(f"✅ TMS Price 匹配键：{len(article_lookup)}；Factory Price 匹配键：{len(price_group_lookup)}；Pack 匹配键：{len(pack_lookup)}")

            log(f"\n📊 正在生成 17 列 Result 数据...")
            results: List[Dict[str, Any]] = []
            missing_article_count = 0
            missing_pack_count = 0
            ambiguous_pack_count = 0
            missing_factory_prices = 0
            missing_tms_prices = 0
            highlighted_factory_conflicts = 0

            for group in expanded_groups:
                working_key = self._clean_key(group['Working Number']).upper()
                article_key = self._clean_key(group['Article Number']).upper()
                source_factory_key = self._clean_key(group['_source_factory']).upper()
                po_key = self._clean_key(group.get('_po_number')).upper()
                output_factory = allocation_lookup.get(po_key, group['Factory'])
                factory_key = self._clean_key(output_factory).upper()
                article_data = article_lookup.get((working_key, article_key))

                season_value = article_data.get('season') if article_data else ""
                marketing_forecast = article_data.get('marketing_forecast') if article_data else ""
                tms_price = article_data.get('tms_price') if article_data else None
                missing_article = article_data is None
                if missing_article:
                    missing_article_count += 1
                    add_diagnostic(
                        "WARN",
                        "ARTICLE_MISSING",
                        f"{working_key} / {article_key}",
                        "TMS 明细未在 Article 表匹配到 Working Number (M) + Article Number (A)。",
                    )

                season_key = self._clean_key(season_value).upper()
                pack_value = pack_lookup.get((working_key, season_key))
                missing_pack = False
                ambiguous_pack = False

                factory_lookup_key = self._make_factory_group_key(season_value, working_key, factory_key)
                price_record = price_group_lookup.get(factory_lookup_key)
                matched_price_key: Tuple[str, ...] = factory_lookup_key
                matched_price_rule = "Season + Working Number + Factory"
                conflict_values = factory_conflict_groups.get(factory_lookup_key)
                if price_record is None and factory_key != source_factory_key:
                    source_factory_lookup_key = self._make_factory_group_key(
                        season_value,
                        working_key,
                        source_factory_key,
                    )
                    price_record = price_group_lookup.get(source_factory_lookup_key)
                    if price_record is not None:
                        matched_price_key = source_factory_lookup_key
                        matched_price_rule = "Season + Working Number + Factory"
                        conflict_values = factory_conflict_groups.get(source_factory_lookup_key)
                        add_diagnostic(
                            "INFO",
                            "FACTORY_PRICE_SOURCE_FACTORY_FALLBACK",
                            f"{po_key or 'NO_PO'} / {working_key} / {article_key}",
                            "Allocation 覆盖后的 Factory 未匹配到价格，已按 TMS 原 Factory 匹配 Factory Price。",
                        )
                if price_record is None:
                    factory_fallback_key = (working_key, factory_key)
                    price_record = price_factory_lookup.get(factory_fallback_key)
                    if price_record is not None:
                        matched_price_key = factory_fallback_key
                        matched_price_rule = "Working Number + Factory"
                        conflict_values = factory_fallback_conflict_groups.get(factory_fallback_key)
                if price_record is None and factory_key != source_factory_key:
                    source_factory_fallback_key = (working_key, source_factory_key)
                    price_record = price_factory_lookup.get(source_factory_fallback_key)
                    if price_record is not None:
                        matched_price_key = source_factory_fallback_key
                        matched_price_rule = "Working Number + Factory"
                        conflict_values = factory_fallback_conflict_groups.get(source_factory_fallback_key)
                        add_diagnostic(
                            "INFO",
                            "FACTORY_PRICE_SOURCE_FACTORY_FALLBACK",
                            f"{po_key or 'NO_PO'} / {working_key} / {article_key}",
                            "Allocation 覆盖后的 Factory 未匹配到价格，已按 TMS 原 Factory + Working Number 匹配 Factory Price。",
                        )
                factory_price_conflict = bool(conflict_values)
                if factory_price_conflict:
                    highlighted_factory_conflicts += 1
                    add_diagnostic(
                        "WARN",
                        "FACTORY_PRICE_CONFLICT",
                        " / ".join(matched_price_key),
                        f"Factory Price 表同一个 {matched_price_rule} 存在多个 Factory Price：{sorted(conflict_values)}；Result 已填写第一个匹配价格并高亮对应价格单元格。",
                    )
                factory_price = price_record.get("factory_price") if price_record else None
                if price_record and price_record.get("pack"):
                    pack_value = price_record["pack"]
                if not pack_value:
                    fallback_packs = {pack for pack in pack_by_working.get(working_key, set()) if pack}
                    if len(fallback_packs) == 1:
                        pack_value = next(iter(fallback_packs))
                    elif len(fallback_packs) > 1:
                        ambiguous_pack = True
                        ambiguous_pack_count += 1
                        add_diagnostic(
                            "WARN",
                            "PACK_AMBIGUOUS",
                            f"{working_key} / {season_key or 'NO_SEASON'}",
                            f"Pack 表同一个 Working Number 有多个 Pack，无法在缺少 Season 精确匹配时兜底：{sorted(fallback_packs)}。",
                        )
                    else:
                        missing_pack = True
                        missing_pack_count += 1
                        add_diagnostic(
                            "WARN",
                            "PACK_MISSING",
                            f"{working_key} / {season_key or 'NO_SEASON'}",
                            "Pack 表未匹配到 Working Number + Season、命中的 Factory Price 记录 Pack，或唯一 Working Number 兜底值。",
                        )
                missing_factory_price = factory_price is None
                if missing_factory_price:
                    missing_factory_prices += 1
                    add_diagnostic(
                        "WARN",
                        "FACTORY_PRICE_MISSING",
                        " / ".join(factory_lookup_key),
                        "Factory Price 表未按 Season + Working Number + Factory 或 Working Number + Factory 匹配到价格。",
                    )

                if tms_price is None:
                    if price_record and price_record.get("tms_price") is not None:
                        tms_price = price_record["tms_price"]
                        add_diagnostic(
                            "INFO",
                            "TMS_PRICE_FACTORY_PRICE_FALLBACK",
                            f"{working_key} / {article_key}",
                            "TMS Price 文件未匹配到 Intl. FOB (C)，已使用 Factory Price 表中的 TMS Price。",
                        )
                if tms_price is None:
                    missing_tms_prices += 1
                    add_diagnostic(
                        "WARN",
                        "TMS_PRICE_MISSING",
                        f"{working_key} / {article_key}",
                        "TMS Price 文件未匹配到可用 Intl. FOB (C)，TMS Price 与 TMS Amount 留空。",
                    )

                quantity = group['Quantity']
                factory_amount = factory_price * quantity if factory_price is not None else None
                tms_amount = tms_price * quantity if tms_price is not None else None

                results.append({
                    'Pack': pack_value or "",
                    'Season': season_value or "",
                    'Factory': output_factory,
                    'Working Number': group['Working Number'],
                    'Article Number': group['Article Number'],
                    'Article Name': group['Article Name'],
                    'CRD': group['CRD'],
                    'PODD': group['PODD'],
                    'Gps Customer Number': group['Gps Customer Number'],
                    'Country/Region': group['Country/Region'],
                    'Shipment Method': group['Shipment Method'],
                    'Marketing Forecast(M)': marketing_forecast,
                    'Quantity': quantity,
                    'Factory Price(USD)': factory_price,
                    'Factory Amount(USD)': factory_amount,
                    'TMS Price(USD)': tms_price,
                    'TMS Amount(USD)': tms_amount,
                    '_factory_amount_value': factory_amount,
                    '_tms_amount_value': tms_amount,
                    '_factory_price_conflict': factory_price_conflict,
                    '_missing_factory_price': missing_factory_price,
                    '_missing_tms_price': tms_price is None,
                    '_missing_article': missing_article,
                    '_missing_pack': missing_pack,
                    '_ambiguous_pack': ambiguous_pack,
                })

            log(f"\n✅ 共生成 {len(results)} 条 Result 明细记录")
            if missing_article_count:
                log(f"  ⚠️ {missing_article_count} 条 Result 明细未匹配到 TMS Price 记录")
            if missing_pack_count:
                log(f"  ⚠️ {missing_pack_count} 条 Result 明细未匹配到 Pack")
            if ambiguous_pack_count:
                log(f"  ⚠️ {ambiguous_pack_count} 条 Result 明细 Pack 存在多值歧义")
            if missing_factory_prices:
                log(f"  ⚠️ {missing_factory_prices} 条 Result 明细未匹配到 Factory Price")
            if missing_tms_prices:
                log(f"  ⚠️ {missing_tms_prices} 条 Result 明细未匹配到 TMS Price Intl. FOB (C)")
            if highlighted_factory_conflicts:
                log(f"  ⚠️ {highlighted_factory_conflicts} 条 Result 明细因同 Season/Working/Factory 多 Factory Price 已高亮")

            result['working_count'] = len(working_numbers)
            result['result_count'] = len(results)
            result['diagnostics_count'] = len(diagnostics)

            log(f"\n💾 正在保存结果报表...")

            if output_dir is None:
                output_dir = os.path.dirname(tms_paths[0])

            ensure_dir(output_dir)

            default_name = (
                "Sophia_Tina_Result_" +
                datetime.now().strftime('%Y%m%d_%H%M%S') +
                ".xlsx"
            )

            output_path = os.path.join(output_dir, default_name)

            self.save_st_result(results, output_path, diagnostics, development_style_rows)

            result['output_path'] = output_path

            log(f"\n{'='*80}")
            log(f"✅ 报表生成完成！")
            log(f"{'='*80}")
            log(f"结果文件：{output_path}")

            result['success'] = True
            result['message'] = f"Sophia & Tina 报表生成完成，共生成 {len(results)} 条 Result 明细记录"

        except Exception as e:
            log(f"\n❌ 错误：{str(e)}")
            import traceback
            log(traceback.format_exc())
            result['message'] = f'处理出错：{str(e)}'

        return result

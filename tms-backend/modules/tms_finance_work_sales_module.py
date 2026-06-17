# -*- coding: utf-8 -*-
"""
TMS 财务 - Work Sales 数据写入模块。
"""

from __future__ import annotations

import os
import re
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Callable, Dict, Iterable, List, Optional, Tuple
from uuid import uuid4

import openpyxl
import xlrd
from openpyxl.formula.translate import Translator
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.worksheet import Worksheet


RowFingerprint = Tuple[str, ...]


@dataclass(frozen=True)
class BulkSalesColumns:
    invoice: int
    style: int
    buyer_unit_price: int
    factory_unit_price: int
    quantity: int
    sales_amount_net: int
    purchase_amount: int
    handover_date: int


@dataclass(frozen=True)
class BulkSalesRow:
    source_sheet: str
    source_row: int
    invoice: str
    style: str
    buyer_unit_price: Any
    factory_unit_price: Any
    quantity: Any
    sales_amount_net: Any
    purchase_amount: Any
    handover_date: Any


@dataclass(frozen=True)
class TurnoverColumns:
    style: int
    unit_price_exclude_vat: int
    quantity: int
    total_system_include_vat: int
    merchandiser: int
    handover_date: int
    invoice: int


@dataclass(frozen=True)
class TurnoverSection:
    name: str
    header_row: int
    data_start_row: int
    subtotal_row: int
    columns: TurnoverColumns


class TmsFinanceWorkSalesModule:
    """从 iPlex BULK Sales 导出表重建 TURNOVER 的 Turnover Details 明细。"""

    DETAIL_SHEET_NAME = "Turnover Details"
    SALES_SECTION = "sales"
    PURCHASE_SECTION = "purchase"
    OUTPUT_PREFIX = "tms_finance_work_sales"
    DEFAULT_MERCHANDISER = "Caroline"
    SALES_VAS_RATE = "0.483581"
    VAT_RATE = "0.13"
    SALES_STATUS = "Issued VAT inv."
    PURCHASE_STATUS = "received VAT inv."
    SALES_TRAILING_ROWS = 4

    def process_files(
        self,
        bulk_sales_path: Optional[str] = None,
        turnover_path: Optional[str] = None,
        output_dir: Optional[str] = None,
        **legacy_kwargs: Any,
    ) -> Dict[str, Any]:
        bulk_sales_path = bulk_sales_path or legacy_kwargs.get("iplix_path")
        if not bulk_sales_path:
            raise ValueError("请上传 BULK Sales 导出表")
        if not turnover_path:
            raise ValueError("请上传 TURNOVER 目标表")

        output_root = output_dir or os.path.dirname(os.path.abspath(turnover_path))
        os.makedirs(output_root, exist_ok=True)

        source_rows = self._extract_bulk_sales_rows(bulk_sales_path)
        if not source_rows:
            raise ValueError("BULK Sales 导出表未读取到有效明细")

        workbook = openpyxl.load_workbook(turnover_path, data_only=False)
        try:
            if self.DETAIL_SHEET_NAME not in workbook.sheetnames:
                raise ValueError("TURNOVER 文件缺少 Turnover Details Sheet")
            ws = workbook[self.DETAIL_SHEET_NAME]
            cleared_sales_count, cleared_purchase_count = self._rebuild_turnover_details(
                ws,
                source_rows,
            )

            output_filename = f"{self.OUTPUT_PREFIX}_{uuid4().hex}.xlsx"
            output_path = os.path.join(output_root, output_filename)
            workbook.save(output_path)
        finally:
            workbook.close()

        sales_written_count = len(source_rows)
        purchase_written_count = len(source_rows)
        duplicate_count = 0
        return {
            "success": True,
            "message": (
                "Work Sales 数据写入完成："
                f"Sales 写入 {sales_written_count} 行，"
                f"Purchase 写入 {purchase_written_count} 行。"
            ),
            "output_path": output_path,
            "output_file": output_filename,
            "source_row_count": len(source_rows),
            "extracted_count": len(source_rows),
            "sales_written_count": sales_written_count,
            "purchase_written_count": purchase_written_count,
            "cleared_sales_count": cleared_sales_count,
            "cleared_purchase_count": cleared_purchase_count,
            "sales_appended_count": sales_written_count,
            "purchase_appended_count": purchase_written_count,
            "duplicate_count": duplicate_count,
            "diagnostic_count": 0,
            "diagnostics": [],
            "totals": {
                "sales_written_count": sales_written_count,
                "purchase_written_count": purchase_written_count,
                "cleared_sales_count": cleared_sales_count,
                "cleared_purchase_count": cleared_purchase_count,
                "sales_appended_count": sales_written_count,
                "purchase_appended_count": purchase_written_count,
            },
            "source_summary": {
                "source_rows": len(source_rows),
                "sales_rows": sales_written_count,
                "purchase_rows": purchase_written_count,
                "sales_written_rows": sales_written_count,
                "purchase_written_rows": purchase_written_count,
                "cleared_sales_rows": cleared_sales_count,
                "cleared_purchase_rows": cleared_purchase_count,
                "duplicate_rows": duplicate_count,
            },
            "logs": [
                f"BULK Sales 读取 {len(source_rows)} 行",
                f"清空旧 Sales 明细 {cleared_sales_count} 行",
                f"清空旧 Purchase 明细 {cleared_purchase_count} 行",
                f"Sales 明细写入 {sales_written_count} 行",
                f"Purchase 明细写入 {purchase_written_count} 行",
            ],
        }

    def _extract_bulk_sales_rows(self, workbook_path: str) -> List[BulkSalesRow]:
        extension = os.path.splitext(workbook_path)[1].lower()
        if extension == ".xls":
            return self._extract_xls_bulk_sales_rows(workbook_path)
        if extension in {".xlsx", ".xlsm"}:
            return self._extract_openpyxl_bulk_sales_rows(workbook_path)
        raise ValueError("BULK Sales 导出表仅支持 .xls / .xlsx / .xlsm")

    def _extract_openpyxl_bulk_sales_rows(self, workbook_path: str) -> List[BulkSalesRow]:
        workbook = openpyxl.load_workbook(workbook_path, data_only=True, read_only=True)
        try:
            ws = workbook[workbook.sheetnames[0]]
            header_map = self._build_header_map_from_reader(
                ws.max_column,
                lambda column: ws.cell(1, column).value,
            )
            columns = self._build_bulk_sales_columns(header_map)
            rows: List[BulkSalesRow] = []
            for row_index in range(2, ws.max_row + 1):
                row = self._build_bulk_sales_row(
                    ws.title,
                    row_index,
                    columns,
                    lambda column, current_row=row_index: ws.cell(current_row, column).value,
                )
                if row is not None:
                    rows.append(row)
            return rows
        finally:
            workbook.close()

    def _extract_xls_bulk_sales_rows(self, workbook_path: str) -> List[BulkSalesRow]:
        book = xlrd.open_workbook(workbook_path, on_demand=True)
        try:
            sheet = book.sheet_by_index(0)
            header_map = self._build_header_map_from_reader(
                sheet.ncols,
                lambda column: sheet.cell_value(0, column - 1),
            )
            columns = self._build_bulk_sales_columns(header_map)
            rows: List[BulkSalesRow] = []
            for row_index in range(2, sheet.nrows + 1):
                row = self._build_bulk_sales_row(
                    sheet.name,
                    row_index,
                    columns,
                    lambda column, current_row=row_index: self._read_xls_cell(
                        sheet,
                        current_row,
                        column,
                        book.datemode,
                    ),
                )
                if row is not None:
                    rows.append(row)
            return rows
        finally:
            book.release_resources()

    def _build_bulk_sales_columns(self, header_map: Dict[str, int]) -> BulkSalesColumns:
        return BulkSalesColumns(
            invoice=self._required_column(header_map, ["SALES INVOICE NUMBER", "INVOICE NUMBER"]),
            style=self._required_column(header_map, ["STYLE NUMBER"]),
            buyer_unit_price=self._required_column(header_map, ["BUYER UNIT PX", "BUYER  UNIT PX"]),
            factory_unit_price=self._required_column(
                header_map,
                ["FACTORY UNIT PX", "FACTORY  UNIT PX"],
            ),
            quantity=self._required_column(header_map, ["SHIP QTY", "SHIP QUANTITY"]),
            sales_amount_net=self._required_column(
                header_map,
                ["*SALES INVOICE AMT (NET)", "SALES INVOICE AMT (NET)"],
            ),
            purchase_amount=self._required_column(
                header_map,
                ["PUR INVOICE AMOUNT", "PURCHASE INVOICE AMOUNT"],
            ),
            handover_date=self._required_column(header_map, ["HANDOVER DATE"]),
        )

    def _build_bulk_sales_row(
        self,
        sheet_name: str,
        row_index: int,
        columns: BulkSalesColumns,
        read_cell: Callable[[int], Any],
    ) -> Optional[BulkSalesRow]:
        invoice = self._clean_text(read_cell(columns.invoice))
        style = self._clean_text(read_cell(columns.style))
        quantity = read_cell(columns.quantity)
        if not (invoice or style or self._clean_text(quantity)):
            return None
        if not invoice or not style:
            return None
        return BulkSalesRow(
            source_sheet=sheet_name,
            source_row=row_index,
            invoice=invoice,
            style=style,
            buyer_unit_price=self._money_or_blank(read_cell(columns.buyer_unit_price)),
            factory_unit_price=self._money_or_blank(read_cell(columns.factory_unit_price)),
            quantity=self._number_for_cell(quantity),
            sales_amount_net=self._money_or_blank(read_cell(columns.sales_amount_net)),
            purchase_amount=self._money_or_blank(read_cell(columns.purchase_amount)),
            handover_date=read_cell(columns.handover_date),
        )

    def _find_turnover_section(self, ws: Worksheet, section_name: str) -> TurnoverSection:
        header_row = (
            self._find_sales_header_row(ws)
            if section_name == self.SALES_SECTION
            else self._find_purchase_header_row(ws)
        )
        columns = self._build_turnover_columns(ws, header_row, section_name)
        data_start_row = header_row + 1
        subtotal_row = self._find_subtotal_row(ws, data_start_row, columns)
        return TurnoverSection(
            name=section_name,
            header_row=header_row,
            data_start_row=data_start_row,
            subtotal_row=subtotal_row,
            columns=columns,
        )

    def _find_sales_header_row(self, ws: Worksheet) -> int:
        for row in range(1, ws.max_row + 1):
            headers = self._header_values(ws, row)
            if (
                "STYLENUMBER" in headers
                and "SALESINVOICENUMBER" in headers
                and "PURCHASEAMOUNT" not in headers
            ):
                return row
        raise ValueError("Turnover Details 缺少 Sales 明细表头")

    def _find_purchase_header_row(self, ws: Worksheet) -> int:
        for row in range(1, ws.max_row + 1):
            headers = self._header_values(ws, row)
            if "STYLENUMBER" in headers and "PURCHASEAMOUNT" in headers:
                return row
        raise ValueError("Turnover Details 缺少 Purchase 明细表头")

    def _build_turnover_columns(
        self,
        ws: Worksheet,
        header_row: int,
        section_name: str,
    ) -> TurnoverColumns:
        header_map = self._build_header_map_from_reader(
            ws.max_column,
            lambda column: ws.cell(header_row, column).value,
        )
        amount_aliases = (
            ["TOTAL AMOUNT IN IPLEX SYSTEM (INCLUDE VAT)"]
            if section_name == self.SALES_SECTION
            else ["TOTAL AMOUNT IN IPLEX SYSTEM (EXCLUDE VAT)"]
        )
        return TurnoverColumns(
            style=self._required_column(header_map, ["STYLE NUMBER"]),
            unit_price_exclude_vat=self._required_column(
                header_map,
                ["UNIT PRICE(EXCLUDE VAT)", "UNIT PRICE EXCLUDE VAT"],
            ),
            quantity=self._required_column(header_map, ["SHIP QUANTITY", "SHIP QTY"]),
            total_system_include_vat=self._required_column(
                header_map,
                amount_aliases,
            ),
            merchandiser=self._required_column(header_map, ["MERCH", "MERCHANDISER"]),
            handover_date=self._required_column(header_map, ["HANDOVER DATE"]),
            invoice=self._required_column(header_map, ["SALES INVOICE NUMBER", "INVOICE NUMBER"]),
        )

    def _find_subtotal_row(
        self,
        ws: Worksheet,
        data_start_row: int,
        columns: TurnoverColumns,
    ) -> int:
        subtotal_row = self._find_subtotal_row_or_none(ws, data_start_row, columns)
        if subtotal_row is None:
            raise ValueError("Turnover Details 缺少明细小计行")
        return subtotal_row

    def _find_subtotal_row_or_none(
        self,
        ws: Worksheet,
        data_start_row: int,
        columns: TurnoverColumns,
        stop_row: Optional[int] = None,
    ) -> Optional[int]:
        end_row = min(stop_row or ws.max_row, ws.max_row)
        for row in range(data_start_row, end_row + 1):
            quantity_value = ws.cell(row, columns.quantity).value
            style_value = self._clean_text(ws.cell(row, columns.style).value)
            if isinstance(quantity_value, str) and quantity_value.upper().startswith("=SUM("):
                return row
            if not style_value and self._row_has_sum_formula(ws, row):
                return row
        return None

    def _append_section_rows(
        self,
        ws: Worksheet,
        section: TurnoverSection,
        rows: List[BulkSalesRow],
        section_name: str,
    ) -> None:
        insert_at = section.subtotal_row
        count = len(rows)
        template_row = max(section.data_start_row, insert_at - 1)
        self._insert_rows_with_formula_translation(ws, insert_at, count)
        for offset, item in enumerate(rows):
            target_row = insert_at + offset
            self._copy_template_row(ws, template_row, target_row)
            if section_name == self.SALES_SECTION:
                self._write_sales_row(ws, target_row, item, section.columns)
            else:
                self._write_purchase_row(ws, target_row, item, section.columns)

    def _write_sales_row(
        self,
        ws: Worksheet,
        row_index: int,
        item: BulkSalesRow,
        columns: TurnoverColumns,
    ) -> None:
        ws.cell(row_index, columns.style).value = item.style
        ws.cell(row_index, columns.unit_price_exclude_vat).value = item.buyer_unit_price
        ws.cell(row_index, columns.quantity).value = item.quantity
        ws.cell(row_index, columns.total_system_include_vat).value = item.sales_amount_net
        ws.cell(row_index, columns.merchandiser).value = self.DEFAULT_MERCHANDISER
        ws.cell(row_index, columns.handover_date).value = item.handover_date
        ws.cell(row_index, columns.invoice).value = item.invoice

    def _write_purchase_row(
        self,
        ws: Worksheet,
        row_index: int,
        item: BulkSalesRow,
        columns: TurnoverColumns,
    ) -> None:
        ws.cell(row_index, columns.style).value = item.style
        ws.cell(row_index, columns.unit_price_exclude_vat).value = item.factory_unit_price
        ws.cell(row_index, columns.quantity).value = item.quantity
        ws.cell(row_index, columns.total_system_include_vat).value = item.purchase_amount
        ws.cell(row_index, columns.merchandiser).value = self.DEFAULT_MERCHANDISER
        ws.cell(row_index, columns.handover_date).value = item.handover_date
        ws.cell(row_index, columns.invoice).value = item.invoice

    def _rebuild_turnover_details(
        self,
        ws: Worksheet,
        source_rows: List[BulkSalesRow],
    ) -> Tuple[int, int]:
        sales_header_row = self._find_sales_header_row(ws)
        purchase_header_row = self._find_purchase_header_row(ws)
        sales_columns = self._build_turnover_columns(ws, sales_header_row, self.SALES_SECTION)
        purchase_columns = self._build_turnover_columns(
            ws,
            purchase_header_row,
            self.PURCHASE_SECTION,
        )
        purchase_title_row = self._find_purchase_title_row(ws, purchase_header_row)

        sales_start_row = sales_header_row + 1
        existing_sales_subtotal = self._find_subtotal_row_or_none(
            ws,
            sales_start_row,
            sales_columns,
            stop_row=purchase_title_row - 1,
        )
        sales_data_end_row = (existing_sales_subtotal or purchase_title_row) - 1
        cleared_sales_count = self._count_existing_detail_rows(
            ws,
            sales_start_row,
            sales_data_end_row,
            sales_columns,
        )

        purchase_start_row = purchase_header_row + 1
        existing_purchase_subtotal = self._find_subtotal_row_or_none(
            ws,
            purchase_start_row,
            purchase_columns,
            stop_row=ws.max_row,
        )
        purchase_data_end_row = (existing_purchase_subtotal or (ws.max_row + 1)) - 1
        cleared_purchase_count = self._count_existing_detail_rows(
            ws,
            purchase_start_row,
            purchase_data_end_row,
            purchase_columns,
        )

        required_sales_rows = len(source_rows) + self.SALES_TRAILING_ROWS
        current_sales_rows = purchase_title_row - sales_start_row
        self._resize_row_region(ws, sales_start_row, current_sales_rows, required_sales_rows)

        sales_header_row = self._find_sales_header_row(ws)
        purchase_header_row = self._find_purchase_header_row(ws)
        sales_columns = self._build_turnover_columns(ws, sales_header_row, self.SALES_SECTION)
        sales_start_row = sales_header_row + 1
        self._write_sales_block(ws, sales_start_row, source_rows, sales_columns)

        purchase_header_row = self._find_purchase_header_row(ws)
        purchase_columns = self._build_turnover_columns(
            ws,
            purchase_header_row,
            self.PURCHASE_SECTION,
        )
        purchase_start_row = purchase_header_row + 1
        existing_purchase_subtotal = self._find_subtotal_row_or_none(
            ws,
            purchase_start_row,
            purchase_columns,
            stop_row=ws.max_row,
        )
        current_purchase_rows = (
            existing_purchase_subtotal - purchase_start_row + 1
            if existing_purchase_subtotal is not None
            else max(0, ws.max_row - purchase_start_row + 1)
        )
        required_purchase_rows = len(source_rows) + 1
        self._resize_row_region(
            ws,
            purchase_start_row,
            current_purchase_rows,
            required_purchase_rows,
        )

        purchase_header_row = self._find_purchase_header_row(ws)
        purchase_columns = self._build_turnover_columns(
            ws,
            purchase_header_row,
            self.PURCHASE_SECTION,
        )
        self._write_purchase_block(ws, purchase_header_row + 1, source_rows, purchase_columns)
        return cleared_sales_count, cleared_purchase_count

    def _find_purchase_title_row(self, ws: Worksheet, purchase_header_row: int) -> int:
        for row in range(purchase_header_row - 1, 0, -1):
            title = self._clean_text(ws.cell(row, 1).value).upper()
            if title.startswith("PURCHASE DETAILS"):
                return row
        raise ValueError("Turnover Details 缺少 Purchase Details 标题")

    def _count_existing_detail_rows(
        self,
        ws: Worksheet,
        start_row: int,
        end_row: int,
        columns: TurnoverColumns,
    ) -> int:
        if end_row < start_row:
            return 0
        count = 0
        for row in range(start_row, end_row + 1):
            style = self._clean_text(ws.cell(row, columns.style).value)
            invoice = self._clean_text(ws.cell(row, columns.invoice).value)
            if style or invoice:
                count += 1
        return count

    def _resize_row_region(
        self,
        ws: Worksheet,
        start_row: int,
        current_count: int,
        required_count: int,
    ) -> None:
        if current_count < required_count:
            ws.insert_rows(start_row + current_count, required_count - current_count)
        elif current_count > required_count:
            ws.delete_rows(start_row + required_count, current_count - required_count)

    def _write_sales_block(
        self,
        ws: Worksheet,
        start_row: int,
        rows: List[BulkSalesRow],
        columns: TurnoverColumns,
    ) -> None:
        template_row = start_row
        for offset, item in enumerate(rows):
            target_row = start_row + offset
            self._prepare_row_for_rebuild(ws, template_row, target_row)
            self._write_sales_row(ws, target_row, item, columns)
            self._write_sales_formulas(ws, target_row)

        subtotal_row = start_row + len(rows)
        self._prepare_row_for_rebuild(ws, template_row, subtotal_row)
        self._write_subtotal_formulas(ws, subtotal_row, start_row, subtotal_row - 1)

        tax_carry_row = subtotal_row + 1
        self._prepare_row_for_rebuild(ws, template_row, tax_carry_row)
        ws.cell(tax_carry_row, 9).value = f"=I{subtotal_row}"

        for row in range(tax_carry_row + 1, tax_carry_row + 3):
            self._prepare_row_for_rebuild(ws, template_row, row)

    def _write_purchase_block(
        self,
        ws: Worksheet,
        start_row: int,
        rows: List[BulkSalesRow],
        columns: TurnoverColumns,
    ) -> None:
        template_row = start_row
        for offset, item in enumerate(rows):
            target_row = start_row + offset
            self._prepare_row_for_rebuild(ws, template_row, target_row)
            self._write_purchase_row(ws, target_row, item, columns)
            self._write_purchase_formulas(ws, target_row)

        subtotal_row = start_row + len(rows)
        self._prepare_row_for_rebuild(ws, template_row, subtotal_row)
        self._write_subtotal_formulas(ws, subtotal_row, start_row, subtotal_row - 1)

    def _write_sales_formulas(self, ws: Worksheet, row_index: int) -> None:
        ws.cell(row_index, 5).value = f"=ROUND(C{row_index}*D{row_index},2)"
        ws.cell(row_index, 6).value = f"=ROUND({self.SALES_VAS_RATE}*D{row_index},2)"
        ws.cell(row_index, 8).value = f"=ROUND(E{row_index}+F{row_index}+G{row_index},2)"
        ws.cell(row_index, 9).value = f"=ROUND(H{row_index}*{self.VAT_RATE},2)"
        ws.cell(row_index, 10).value = f"=H{row_index}+I{row_index}"
        ws.cell(row_index, 11).value = f"=J{row_index}"
        ws.cell(row_index, 14).value = f"=J{row_index}-L{row_index}"
        ws.cell(row_index, 15).value = self.SALES_STATUS

    def _write_purchase_formulas(self, ws: Worksheet, row_index: int) -> None:
        ws.cell(row_index, 5).value = f"=ROUND(C{row_index}*D{row_index},2)"
        ws.cell(row_index, 7).value = f"=ROUND(E{row_index}*{self.VAT_RATE},2)"
        ws.cell(row_index, 9).value = f"=E{row_index}+G{row_index}+F{row_index}"
        ws.cell(row_index, 10).value = f"=L{row_index}+G{row_index}"
        ws.cell(row_index, 13).value = f"=E{row_index}-L{row_index}"
        ws.cell(row_index, 14).value = self.PURCHASE_STATUS

    def _write_subtotal_formulas(
        self,
        ws: Worksheet,
        subtotal_row: int,
        start_row: int,
        end_row: int,
    ) -> None:
        for column in range(4, 15):
            letter = get_column_letter(column)
            ws.cell(subtotal_row, column).value = f"=SUM({letter}{start_row}:{letter}{end_row})"

    def _prepare_row_for_rebuild(
        self,
        ws: Worksheet,
        template_row: int,
        target_row: int,
    ) -> None:
        if 1 <= template_row <= ws.max_row and template_row != target_row:
            self._copy_row_style(ws, template_row, target_row)
        self._clear_row_values(ws, target_row)

    def _copy_row_style(self, ws: Worksheet, template_row: int, target_row: int) -> None:
        source_dimension = ws.row_dimensions[template_row]
        target_dimension = ws.row_dimensions[target_row]
        target_dimension.height = source_dimension.height
        target_dimension.hidden = source_dimension.hidden
        target_dimension.outlineLevel = source_dimension.outlineLevel
        target_dimension.collapsed = source_dimension.collapsed
        for column in range(1, ws.max_column + 1):
            source_cell = ws.cell(template_row, column)
            target_cell = ws.cell(target_row, column)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)

    def _clear_row_values(self, ws: Worksheet, row_index: int) -> None:
        for column in range(1, ws.max_column + 1):
            ws.cell(row_index, column).value = None

    def _build_existing_fingerprints(
        self,
        ws: Worksheet,
        section: TurnoverSection,
    ) -> set[RowFingerprint]:
        fingerprints: set[RowFingerprint] = set()
        for row in range(section.data_start_row, section.subtotal_row):
            if not self._clean_text(ws.cell(row, section.columns.style).value):
                continue
            fingerprints.add(self._build_target_fingerprint(ws, row, section))
        return fingerprints

    def _build_target_fingerprint(
        self,
        ws: Worksheet,
        row_index: int,
        section: TurnoverSection,
    ) -> RowFingerprint:
        columns = section.columns
        return (
            section.name,
            self._clean_text(ws.cell(row_index, columns.style).value),
            self._normalize_money(ws.cell(row_index, columns.unit_price_exclude_vat).value),
            self._normalize_quantity(ws.cell(row_index, columns.quantity).value),
            self._normalize_money(ws.cell(row_index, columns.total_system_include_vat).value),
            self._normalize_date_for_key(ws.cell(row_index, columns.handover_date).value),
            self._clean_text(ws.cell(row_index, columns.invoice).value),
        )

    def _build_source_fingerprint(
        self,
        item: BulkSalesRow,
        section_name: str,
    ) -> RowFingerprint:
        amount = (
            item.sales_amount_net
            if section_name == self.SALES_SECTION
            else item.purchase_amount
        )
        unit_price = (
            item.buyer_unit_price
            if section_name == self.SALES_SECTION
            else item.factory_unit_price
        )
        return (
            section_name,
            self._clean_text(item.style),
            self._normalize_money(unit_price),
            self._normalize_quantity(item.quantity),
            self._normalize_money(amount),
            self._normalize_date_for_key(item.handover_date),
            self._clean_text(item.invoice),
        )

    def _insert_rows_with_formula_translation(
        self,
        ws: Worksheet,
        insert_at: int,
        amount: int,
    ) -> None:
        if amount <= 0:
            return
        formulas: List[Tuple[int, int, str]] = []
        for row in range(insert_at, ws.max_row + 1):
            for column in range(1, ws.max_column + 1):
                value = ws.cell(row, column).value
                if isinstance(value, str) and value.startswith("="):
                    formulas.append((row, column, value))
        ws.insert_rows(insert_at, amount)
        for old_row, column, formula in formulas:
            new_row = old_row + amount
            origin = f"{get_column_letter(column)}{old_row}"
            target = f"{get_column_letter(column)}{new_row}"
            ws.cell(new_row, column).value = self._translate_formula(formula, origin, target)

    def _copy_template_row(self, ws: Worksheet, template_row: int, target_row: int) -> None:
        source_dimension = ws.row_dimensions[template_row]
        target_dimension = ws.row_dimensions[target_row]
        target_dimension.height = source_dimension.height
        target_dimension.hidden = source_dimension.hidden
        target_dimension.outlineLevel = source_dimension.outlineLevel
        target_dimension.collapsed = source_dimension.collapsed

        for column in range(1, ws.max_column + 1):
            source_cell = ws.cell(template_row, column)
            target_cell = ws.cell(target_row, column)
            if source_cell.has_style:
                target_cell.font = copy(source_cell.font)
                target_cell.fill = copy(source_cell.fill)
                target_cell.border = copy(source_cell.border)
                target_cell.alignment = copy(source_cell.alignment)
                target_cell.number_format = source_cell.number_format
                target_cell.protection = copy(source_cell.protection)
            value = source_cell.value
            if isinstance(value, str) and value.startswith("="):
                origin = f"{get_column_letter(column)}{template_row}"
                target = f"{get_column_letter(column)}{target_row}"
                value = self._translate_formula(value, origin, target)
            target_cell.value = value

    def _rewrite_subtotal_formulas(self, ws: Worksheet, section: TurnoverSection) -> None:
        data_end_row = max(section.data_start_row, section.subtotal_row - 1)
        for column in range(4, 15):
            letter = get_column_letter(column)
            ws.cell(section.subtotal_row, column).value = (
                f"=SUM({letter}{section.data_start_row}:{letter}{data_end_row})"
            )

    def _translate_formula(self, formula: str, origin: str, target: str) -> str:
        try:
            return Translator(formula, origin=origin).translate_formula(target)
        except Exception:
            return formula

    def _row_has_sum_formula(self, ws: Worksheet, row: int) -> bool:
        for column in range(1, ws.max_column + 1):
            value = ws.cell(row, column).value
            if isinstance(value, str) and value.upper().startswith("=SUM("):
                return True
        return False

    def _header_values(self, ws: Worksheet, row: int) -> set[str]:
        return {
            self._normalize_header(ws.cell(row, column).value)
            for column in range(1, ws.max_column + 1)
            if self._normalize_header(ws.cell(row, column).value)
        }

    def _build_header_map_from_reader(
        self,
        max_column: int,
        read_cell: Callable[[int], Any],
    ) -> Dict[str, int]:
        return {
            self._normalize_header(read_cell(column)): column
            for column in range(1, max_column + 1)
            if self._normalize_header(read_cell(column))
        }

    def _required_column(self, header_map: Dict[str, int], aliases: List[str]) -> int:
        for alias in aliases:
            normalized = self._normalize_header(alias)
            if normalized in header_map:
                return header_map[normalized]
        raise ValueError(f"缺少必需字段：{' / '.join(aliases)}")

    def _normalize_header(self, value: Any) -> str:
        return re.sub(r"[\s*]+", "", self._clean_text(value)).upper()

    def _read_xls_cell(
        self,
        sheet: xlrd.sheet.Sheet,
        row: int,
        column: int,
        datemode: int,
    ) -> Any:
        cell = sheet.cell(row - 1, column - 1)
        if cell.ctype == xlrd.XL_CELL_DATE:
            return xlrd.xldate.xldate_as_datetime(cell.value, datemode)
        return cell.value

    def _money_or_blank(self, value: Any) -> Any:
        amount = self._parse_decimal(value)
        return "" if amount is None else self._money_for_response(amount)

    def _number_for_cell(self, value: Any) -> Any:
        number = self._parse_decimal(value)
        if number is None:
            return value
        if number == number.to_integral_value():
            return int(number)
        return float(number)

    def _normalize_money(self, value: Any) -> str:
        amount = self._parse_decimal(value)
        return "" if amount is None else format(self._round_money(amount), ".2f")

    def _normalize_quantity(self, value: Any) -> str:
        number = self._parse_decimal(value)
        if number is None:
            return self._clean_text(value)
        if number == number.to_integral_value():
            return str(int(number))
        normalized = format(number.normalize(), "f")
        return normalized.rstrip("0").rstrip(".")

    def _normalize_date_for_key(self, value: Any) -> str:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        return self._clean_text(value)

    def _parse_decimal(self, value: Any) -> Optional[Decimal]:
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value).replace(",", "").strip())
        except (InvalidOperation, AttributeError):
            return None

    def _round_money(self, value: Decimal) -> Decimal:
        return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def _money_for_response(self, value: Decimal) -> float:
        return float(self._round_money(value))

    def _clean_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).replace("\u200b", "").replace("\ufeff", "").strip()

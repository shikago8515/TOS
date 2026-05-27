# -*- coding: utf-8 -*-
"""
Jane BOM 汇总模块。

根据 BOM 文件里的 Working # 和 Season 匹配 Pack.xlsx，再把 MAIN COMPONENT
物料按 Article/Color 展开成汇总表。
"""

import os
from datetime import datetime
from typing import Any, Dict, List, Optional, Sequence, Tuple

import openpyxl
from openpyxl.styles import Font, PatternFill

from utils.file_utils import ensure_dir


class JaneBomSummaryModule:
    """Jane-BOM 汇总业务逻辑。"""

    OUTPUT_HEADERS = [
        "Pack",
        "Working #",
        "Season",
        "Factory",
        "Articles",
        "Part Group #",
        "Material Reference #",
        "Material Name",
        "Group Code Supplier",
        "Supplier Name",
        "Material Description",
        "Color",
    ]
    REQUIRED_PACK_COLUMNS = {
        "pack": "Pack",
        "season": "Season",
        "working number": "Working Number",
    }
    MAIN_COMPONENT_SECTION = "MAIN COMPONENT"

    @staticmethod
    def _normalize_text(value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        if text.lower() in {"none", "nan", "nat"}:
            return ""
        return text

    @classmethod
    def _normalize_key(cls, value: Any) -> str:
        return cls._normalize_text(value).upper()

    @classmethod
    def _find_header_row(cls, ws, required_columns: Sequence[str], max_scan_rows: int = 30) -> Tuple[int, Dict[str, int]]:
        required = {column.lower(): column for column in required_columns}
        for row_index in range(1, min(ws.max_row, max_scan_rows) + 1):
            columns: Dict[str, int] = {}
            for column_index in range(1, ws.max_column + 1):
                value = cls._normalize_text(ws.cell(row=row_index, column=column_index).value).lower()
                if value:
                    columns[value] = column_index
            if all(column in columns for column in required):
                return row_index, columns
        raise ValueError(f"未找到表头：{', '.join(required_columns)}")

    @classmethod
    def _parse_article_from_group(cls, value: Any) -> str:
        parts = [part.strip() for part in cls._normalize_text(value).split("|")]
        if len(parts) >= 2:
            return parts[1]
        return ""

    def _read_pack_lookup(self, pack_path: str, logs: List[str]) -> Dict[Tuple[str, str], str]:
        wb = openpyxl.load_workbook(pack_path, data_only=True)
        ws = wb.active
        header_row, header_columns = self._find_header_row(ws, self.REQUIRED_PACK_COLUMNS.values(), max_scan_rows=10)

        pack_col = header_columns["pack"]
        season_col = header_columns["season"]
        working_col = header_columns["working number"]
        lookup: Dict[Tuple[str, str], str] = {}

        for row_index in range(header_row + 1, ws.max_row + 1):
            pack = self._normalize_text(ws.cell(row=row_index, column=pack_col).value)
            season = self._normalize_key(ws.cell(row=row_index, column=season_col).value)
            working = self._normalize_key(ws.cell(row=row_index, column=working_col).value)
            if not working or not season:
                continue

            key = (working, season)
            existing = lookup.get(key)
            if existing and existing != pack:
                raise ValueError(f"Pack 表存在重复映射：{working} + {season} -> {existing} / {pack}")
            lookup[key] = pack

        logs.append(f"Pack 映射读取完成：{len(lookup)} 个 Working Number + Season")
        return lookup

    def _parse_bom_rows(
        self,
        bom_path: str,
        pack_lookup: Dict[Tuple[str, str], str],
        logs: List[str],
    ) -> List[Dict[str, Any]]:
        wb = openpyxl.load_workbook(bom_path, data_only=False)
        ws = wb.active

        working = self._normalize_text(ws["B1"].value)
        season = self._normalize_text(ws["B3"].value)
        factory = self._normalize_text(ws["B5"].value)
        if not working or not season or not factory:
            raise ValueError(f"{os.path.basename(bom_path)} 缺少 Working #、Season 或 Factory")

        header_row, header_columns = self._find_header_row(
            ws,
            [
                "Part Group #",
                "Material Reference #",
                "Material Name",
                "Group Code Supplier",
                "Supplier Name",
                "Material Description",
                "Color",
            ],
        )

        article_columns: List[Tuple[int, str]] = []
        for column_index in range(1, ws.max_column + 1):
            header = self._normalize_text(ws.cell(row=header_row, column=column_index).value)
            if header.lower() != "color":
                continue
            article = self._parse_article_from_group(ws.cell(row=header_row - 1, column=column_index).value)
            if article:
                article_columns.append((column_index, article))

        if not article_columns:
            raise ValueError(f"{os.path.basename(bom_path)} 未识别到 Article/Color 列")

        pack = pack_lookup.get((self._normalize_key(working), self._normalize_key(season)), "")
        if not pack:
            logs.append(f"WARN: {os.path.basename(bom_path)} 未匹配到 Pack：{working} + {season}")

        rows: List[Dict[str, Any]] = []
        current_section = ""
        for row_index in range(header_row + 1, ws.max_row + 1):
            part_name = self._normalize_text(ws.cell(row=row_index, column=2).value)
            part_group = self._normalize_text(ws.cell(row=row_index, column=header_columns["part group #"]).value)
            material_ref = self._normalize_text(ws.cell(row=row_index, column=header_columns["material reference #"]).value)

            # BOM 的 section 行只有 Part Name，没有 Part Group / Material Reference。
            if part_name and not part_group and not material_ref:
                current_section = part_name.upper()
                continue

            if current_section != self.MAIN_COMPONENT_SECTION or not material_ref:
                continue

            base_row = {
                "Pack": pack,
                "Working #": working,
                "Season": season,
                "Factory": factory,
                "Part Group #": part_group,
                "Material Reference #": material_ref,
                "Material Name": self._normalize_text(ws.cell(row=row_index, column=header_columns["material name"]).value),
                "Group Code Supplier": self._normalize_text(ws.cell(row=row_index, column=header_columns["group code supplier"]).value),
                "Supplier Name": self._normalize_text(ws.cell(row=row_index, column=header_columns["supplier name"]).value),
                "Material Description": self._normalize_text(ws.cell(row=row_index, column=header_columns["material description"]).value),
            }

            for color_column, article in article_columns:
                color = self._normalize_text(ws.cell(row=row_index, column=color_column).value)
                if not color:
                    continue
                rows.append({
                    **base_row,
                    "Articles": article,
                    "Color": color,
                })

        logs.append(f"{os.path.basename(bom_path)} 解析完成：{len(rows)} 行")
        return rows

    def _create_output_workbook(self, rows: List[Dict[str, Any]]) -> openpyxl.Workbook:
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(self.OUTPUT_HEADERS)

        for row in rows:
            ws.append([row.get(header, "") for header in self.OUTPUT_HEADERS])

        # 表头样式保持克制，便于用户继续筛选和编辑。
        header_fill = PatternFill("solid", fgColor="D9EAF7")
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = header_fill
        ws.freeze_panes = "A2"
        ws.auto_filter.ref = ws.dimensions

        for column_cells in ws.columns:
            max_length = max(len(self._normalize_text(cell.value)) for cell in column_cells)
            ws.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 10), 48)

        return wb

    def process_reports(
        self,
        bom_paths: Sequence[str],
        pack_path: str,
        output_dir: str,
    ) -> Dict[str, Any]:
        logs: List[str] = []
        try:
            if not bom_paths:
                return {
                    "success": False,
                    "message": "请至少上传 1 个 BOM 文件",
                    "logs": logs,
                }

            ensure_dir(output_dir)
            logs.append("开始生成 Jane-BOM 汇总")
            pack_lookup = self._read_pack_lookup(pack_path, logs)

            rows: List[Dict[str, Any]] = []
            for bom_path in bom_paths:
                rows.extend(self._parse_bom_rows(bom_path, pack_lookup, logs))

            wb = self._create_output_workbook(rows)
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(output_dir, f"jane_bom_summary_{timestamp}.xlsx")
            wb.save(output_path)

            return {
                "success": True,
                "message": f"Jane-BOM 汇总生成完成，结果文件：{output_path}",
                "logs": logs,
                "row_count": len(rows),
                "bom_count": len(bom_paths),
                "output_path": output_path,
            }
        except Exception as exc:
            logs.append(f"ERROR: {exc}")
            return {
                "success": False,
                "message": str(exc),
                "logs": logs,
            }

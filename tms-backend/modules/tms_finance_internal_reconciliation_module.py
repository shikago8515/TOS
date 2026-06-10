# -*- coding: utf-8 -*-
"""
TMS 财务 - 内销对账表数据提取模块。
"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from typing import Any, Dict, Iterable, List, Optional, Tuple
from uuid import uuid4

import openpyxl
from openpyxl.utils.cell import column_index_from_string
from openpyxl.worksheet.worksheet import Worksheet


BusinessKey = Tuple[str, str, str, str]


@dataclass(frozen=True)
class SourceColumns:
    style: int
    article: int
    description: int
    order: int
    customer_order: int
    customer_no: int
    warehouse: int
    delivery_date: int
    quantity: int
    purchase_amount: int
    tms_unit_price: int
    promo_fee: int
    sales_amount: int
    vendor_code: int
    vendor_name: int
    mr: int
    commercial_invoice: Optional[int]


@dataclass
class ExtractedRow:
    source_file: str
    source_sheet: str
    source_row: int
    key: BusinessKey
    values: Dict[str, Any]
    commercial_invoice: str


@dataclass(frozen=True)
class ExclusionConfig:
    rows: set[int]
    columns: set[int]
    diagnostics: List[Dict[str, Any]]


class TmsFinanceInternalReconciliationModule:
    """按来源行顺序回填内销对账单 `未清账` 尾部已有行。"""

    TARGET_SHEET_NAME = "未清账"
    EXCLUSION_SHEET_NAME = "_Exclude_Rows_Cols"
    TARGET_HEADERS = [
        "Date1",
        "Sales invoice",
        "Date2",
        "Purchase invoice",
        "REMARK",
        "VENDOR",
        "CUSTOMER",
        "QTY",
        "Purchase amount",
        "Sales amount  with Tax 13%",
        "货描",
        "WORKING  NO.(款号)",
        "PO ORDER NO.",
        "ARTICLE NO.(货号)",
        "CUSTOMER NO.(客户编码)",
        "CUSTOMER ORDER NO.(客户订单号)",
        "交期",
        "MR",
        "Commercial Invoice",
        "工厂Promo附加费",
        "采购",
        "销售",
    ]
    TARGET_FIELD_HEADERS = {
        "date1": "Date1",
        "sales_invoice": "Sales invoice",
        "date2": "Date2",
        "purchase_invoice": "Purchase invoice",
        "remark": "REMARK",
        "vendor": "VENDOR",
        "customer": "CUSTOMER",
        "quantity": "QTY",
        "purchase_amount": "Purchase amount",
        "sales_amount": "Sales amount  with Tax 13%",
        "description": "货描",
        "style": "WORKING  NO.(款号)",
        "order": "PO ORDER NO.",
        "article": "ARTICLE NO.(货号)",
        "customer_no": "CUSTOMER NO.(客户编码)",
        "customer_order": "CUSTOMER ORDER NO.(客户订单号)",
        "delivery_date": "交期",
        "mr": "MR",
        "commercial_invoice": "Commercial Invoice",
        "promo_fee": "工厂Promo附加费",
        "purchase_display": "采购",
        "sales_display": "销售",
    }
    WRITABLE_FIELDS = [
        "remark",
        "vendor",
        "customer",
        "quantity",
        "purchase_amount",
        "sales_amount",
        "description",
        "style",
        "order",
        "article",
        "customer_no",
        "customer_order",
        "delivery_date",
        "mr",
        "commercial_invoice",
        "promo_fee",
        "purchase_display",
        "sales_display",
    ]
    MONEY_FIELDS = {"purchase_amount", "sales_amount", "promo_fee"}
    VENDOR_CODE_MAP = {
        "ELP008": "新龙泰",
        "1L8006": "新龙泰",
        "ELP012": "万代",
        "1L8012": "万代",
    }
    VENDOR_NAME_ALIASES = {
        "丹东新龙太": "新龙泰",
        "丹东新龙泰": "新龙泰",
        "新龙太": "新龙泰",
        "新龙泰": "新龙泰",
    }

    def process_files(
        self,
        source_paths: Iterable[str] | str | None = None,
        target_path: Optional[str] = None,
        output_dir: Optional[str] = None,
        legacy_output_dir: Optional[str] = None,
        *,
        sample_path: Optional[str] = None,
        bulk_path: Optional[str] = None,
    ) -> Dict[str, Any]:
        source_path_list, normalized_target_path, normalized_output_dir = self._normalize_process_args(
            source_paths,
            target_path,
            output_dir,
            legacy_output_dir,
            sample_path,
            bulk_path,
        )
        if not source_path_list:
            raise ValueError("请至少上传一个 Sample/Bulk 来源文件")
        if not normalized_target_path:
            raise ValueError("请上传内销对账大表")

        output_root = normalized_output_dir or os.path.dirname(os.path.abspath(normalized_target_path))
        os.makedirs(output_root, exist_ok=True)

        target_wb = self._load_workbook(normalized_target_path)
        if self.TARGET_SHEET_NAME not in target_wb.sheetnames:
            raise ValueError("内销对账单缺少 未清账 Sheet")

        target_ws = target_wb[self.TARGET_SHEET_NAME]
        target_columns = self._resolve_target_columns(target_ws)
        exclusion_config = self._parse_exclusion_config(target_wb, target_ws)
        source_rows: List[ExtractedRow] = []
        for source_path in source_path_list:
            source_rows.extend(self._extract_source_rows(source_path))

        first_blank_row = self._find_first_empty_row(target_ws)
        target_row_count = max(first_blank_row - 2, 0)
        candidate_rows = [
            row
            for row in range(2, first_blank_row)
            if row not in exclusion_config.rows
        ]
        if len(candidate_rows) < len(source_rows):
            raise ValueError(
                f"目标表可回填行不足：来源 {len(source_rows)} 行，可回填 {len(candidate_rows)} 行"
            )

        write_rows = candidate_rows[-len(source_rows):] if source_rows else []
        totals = {
            "quantity": Decimal("0"),
            "purchase_amount": Decimal("0"),
            "sales_amount_with_tax": Decimal("0"),
        }
        sample_count = 0
        book_count = 0

        for item, target_row in zip(source_rows, write_rows):
            self._write_target_row(
                target_ws,
                target_row,
                item,
                target_columns,
                exclusion_config.columns,
            )
            totals["quantity"] += self._decimal_or_zero(item.values.get("quantity"))
            totals["purchase_amount"] += self._decimal_or_zero(item.values.get("purchase_amount"))
            totals["sales_amount_with_tax"] += self._decimal_or_zero(item.values.get("sales_amount"))
            if item.values.get("remark") == "Sample":
                sample_count += 1
            elif item.values.get("remark") == "Book":
                book_count += 1

        output_filename = f"tms_finance_internal_reconciliation_{uuid4().hex}.xlsx"
        output_path = os.path.join(output_root, output_filename)
        target_wb.save(output_path)

        updated_count = len(write_rows)
        diagnostics = exclusion_config.diagnostics
        return {
            "success": True,
            "message": f"内销对账表数据提取完成：回填 {updated_count} 行。",
            "output_path": output_path,
            "output_file": output_filename,
            "updated_count": updated_count,
            "source_row_count": len(source_rows),
            "target_row_count": target_row_count,
            "excluded_rows": sorted(exclusion_config.rows),
            "excluded_columns": sorted(exclusion_config.columns),
            "appended_count": 0,
            "skipped_count": 0,
            "duplicate_count": 0,
            "diagnostic_count": len(diagnostics),
            "diagnostics": diagnostics,
            "totals": {
                "quantity": self._number_for_response(totals["quantity"]),
                "purchase_amount": self._money_for_response(totals["purchase_amount"]),
                "sales_amount_with_tax": self._money_for_response(totals["sales_amount_with_tax"]),
            },
            "source_summary": {
                "sample_rows": sample_count,
                "book_rows": book_count,
                "source_rows": len(source_rows),
                "source_files": len(source_path_list),
            },
            "logs": [
                f"来源文件 {len(source_path_list)} 个，提取 {len(source_rows)} 行",
                f"目标首个全空行：{first_blank_row}",
                f"回填 {updated_count} 行，未追加新行",
            ],
        }

    def _normalize_process_args(
        self,
        source_paths: Iterable[str] | str | None,
        target_path: Optional[str],
        output_dir: Optional[str],
        legacy_output_dir: Optional[str],
        sample_path: Optional[str],
        bulk_path: Optional[str],
    ) -> Tuple[List[str], Optional[str], Optional[str]]:
        if legacy_output_dir is not None:
            legacy_sample_path = self._clean_path(source_paths)
            legacy_bulk_path = self._clean_path(target_path)
            legacy_target_path = self._clean_path(output_dir)
            return (
                [path for path in [legacy_sample_path, legacy_bulk_path] if path],
                legacy_target_path,
                legacy_output_dir,
            )

        if sample_path or bulk_path:
            return (
                [path for path in [sample_path, bulk_path] if path],
                target_path,
                output_dir,
            )

        if source_paths is None:
            return [], target_path, output_dir
        if isinstance(source_paths, (str, os.PathLike)):
            return [os.fspath(source_paths)], target_path, output_dir
        return [os.fspath(path) for path in source_paths], target_path, output_dir

    def _clean_path(self, value: Any) -> Optional[str]:
        if value is None:
            return None
        return os.fspath(value)

    def _load_workbook(self, workbook_path: str, *, data_only: bool = False) -> openpyxl.Workbook:
        return openpyxl.load_workbook(
            workbook_path,
            data_only=data_only,
            read_only=False,
            keep_vba=workbook_path.lower().endswith(".xlsm"),
        )

    def _extract_source_rows(self, workbook_path: str) -> List[ExtractedRow]:
        values_wb = self._load_workbook(workbook_path, data_only=True)
        formula_wb = self._load_workbook(workbook_path, data_only=False)
        rows: List[ExtractedRow] = []

        for values_ws in values_wb.worksheets:
            if values_ws.sheet_state != "visible":
                continue
            formula_ws = formula_wb[values_ws.title]
            columns = self._find_source_columns(values_ws)
            if columns is None:
                continue
            header_row = self._find_header_row(values_ws)
            if header_row is None:
                continue

            # 来源表为两层表头；遇到首个整行全空立即停止，避免处理 Total 或历史残留。
            row_index = header_row + 2
            while row_index <= values_ws.max_row:
                if self._is_row_empty(values_ws, row_index):
                    break
                if self._is_business_row(values_ws, row_index, columns):
                    rows.append(
                        self._build_extracted_row(
                            workbook_path,
                            values_ws,
                            formula_ws,
                            row_index,
                            columns,
                        ),
                    )
                row_index += 1

        return rows

    def _find_source_columns(self, ws: Worksheet) -> Optional[SourceColumns]:
        header_row = self._find_header_row(ws)
        if header_row is None:
            return None

        header_map = self._build_header_map(ws, header_row)
        required = [
            "款号",
            "颜色",
            "描述",
            "订单号",
            "客户订单号",
            "客户编号",
            "客户仓库",
            "大货交期",
            "数量",
            "TMS业务",
        ]
        if any(self._normalize_header(name) not in header_map for name in required):
            return None

        return SourceColumns(
            style=header_map[self._normalize_header("款号")],
            article=header_map[self._normalize_header("颜色")],
            description=header_map[self._normalize_header("描述")],
            order=header_map[self._normalize_header("订单号")],
            customer_order=header_map[self._normalize_header("客户订单号")],
            customer_no=header_map[self._normalize_header("客户编号")],
            warehouse=header_map[self._normalize_header("客户仓库")],
            delivery_date=header_map[self._normalize_header("大货交期")],
            quantity=header_map[self._normalize_header("数量")],
            purchase_amount=self._find_nested_column(
                ws,
                header_row,
                parent_names=["工厂价格"],
                child_keywords=["PO总金额", "总金额"],
                fallback=14,
            ),
            tms_unit_price=self._find_nested_column(
                ws,
                header_row,
                parent_names=["TMS价格"],
                child_keywords=["单价"],
                fallback=15,
            ),
            promo_fee=self._find_nested_column(
                ws,
                header_row,
                parent_names=["TMS价格"],
                child_keywords=["VAS/SHAS附加费", "附加费", "Promo"],
                fallback=17,
            ),
            sales_amount=self._find_nested_column(
                ws,
                header_row,
                parent_names=["TMS价格"],
                child_keywords=["PO总金额", "总金额"],
                fallback=18,
            ),
            vendor_code=self._find_nested_column(
                ws,
                header_row,
                parent_names=["工厂"],
                child_keywords=["工厂code", "Code"],
                fallback=19,
            ),
            vendor_name=self._find_nested_column(
                ws,
                header_row,
                parent_names=["工厂"],
                child_keywords=["中文", "工厂"],
                fallback=20,
                exclude_keywords=["code"],
            ),
            mr=header_map[self._normalize_header("TMS业务")],
            commercial_invoice=header_map.get(self._normalize_header("TMS发票#")),
        )

    def _find_header_row(self, ws: Worksheet) -> Optional[int]:
        required = {
            self._normalize_header("款号"),
            self._normalize_header("订单号"),
            self._normalize_header("数量"),
        }
        for row_index in range(1, min(ws.max_row, 10) + 1):
            values = {
                self._normalize_header(ws.cell(row_index, column).value)
                for column in range(1, ws.max_column + 1)
            }
            if required.issubset(values):
                return row_index
        return None

    def _build_header_map(self, ws: Worksheet, header_row: int) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for column in range(1, ws.max_column + 1):
            header = self._normalize_header(ws.cell(header_row, column).value)
            if header:
                mapping[header] = column
        return mapping

    def _find_nested_column(
        self,
        ws: Worksheet,
        header_row: int,
        *,
        parent_names: List[str],
        child_keywords: List[str],
        fallback: int,
        exclude_keywords: Optional[List[str]] = None,
    ) -> int:
        parent_set = {self._normalize_header(name) for name in parent_names}
        child_set = [self._normalize_header(name) for name in child_keywords]
        exclude_set = [self._normalize_header(name) for name in (exclude_keywords or [])]
        current_parent = ""
        for column in range(1, ws.max_column + 1):
            parent = self._normalize_header(ws.cell(header_row, column).value)
            if parent:
                current_parent = parent
            child = self._normalize_header(ws.cell(header_row + 1, column).value)
            if current_parent not in parent_set or not child:
                continue
            if exclude_set and any(keyword in child for keyword in exclude_set):
                continue
            if any(keyword in child for keyword in child_set):
                return column
        return fallback

    def _is_business_row(self, ws: Worksheet, row: int, columns: SourceColumns) -> bool:
        style = self._clean_text(ws.cell(row, columns.style).value)
        order = self._normalize_order(ws.cell(row, columns.order).value)
        quantity = ws.cell(row, columns.quantity).value
        return bool(style and order and self._parse_decimal(quantity) is not None)

    def _build_extracted_row(
        self,
        workbook_path: str,
        values_ws: Worksheet,
        formula_ws: Worksheet,
        row: int,
        columns: SourceColumns,
    ) -> ExtractedRow:
        warehouse = self._clean_text(values_ws.cell(row, columns.warehouse).value)
        remark = "Sample" if warehouse.upper() == "MSO" else "Book"
        vendor = self._map_vendor(
            values_ws.cell(row, columns.vendor_code).value,
            values_ws.cell(row, columns.vendor_name).value,
        )
        style = self._normalize_style(values_ws.cell(row, columns.style).value)
        order = self._normalize_order(values_ws.cell(row, columns.order).value)
        article = self._clean_text(values_ws.cell(row, columns.article).value)
        commercial_invoice = (
            self._clean_text(values_ws.cell(row, columns.commercial_invoice).value)
            if columns.commercial_invoice
            else ""
        )
        purchase_amount = self._read_purchase_amount(values_ws, row, columns)
        sales_amount = self._read_sales_amount(values_ws, formula_ws, row, columns)
        promo_fee = self._money_or_blank(values_ws.cell(row, columns.promo_fee).value)
        quantity = self._number_for_cell(values_ws.cell(row, columns.quantity).value)
        values = {
            "remark": remark,
            "vendor": vendor,
            "customer": "adidas",
            "quantity": quantity,
            "purchase_amount": purchase_amount,
            "sales_amount": sales_amount,
            "description": self._clean_text(values_ws.cell(row, columns.description).value),
            "style": style,
            "order": order,
            "article": article,
            "customer_no": self._clean_text(values_ws.cell(row, columns.customer_no).value),
            "customer_order": self._clean_placeholder(values_ws.cell(row, columns.customer_order).value),
            "delivery_date": values_ws.cell(row, columns.delivery_date).value,
            "mr": self._clean_text(values_ws.cell(row, columns.mr).value),
            "commercial_invoice": commercial_invoice,
            "promo_fee": promo_fee,
        }
        return ExtractedRow(
            source_file=os.path.basename(workbook_path),
            source_sheet=values_ws.title,
            source_row=row,
            key=(remark, style, order, article),
            values=values,
            commercial_invoice=commercial_invoice,
        )

    def _read_purchase_amount(self, ws: Worksheet, row: int, columns: SourceColumns) -> Optional[float]:
        amount = self._parse_decimal(ws.cell(row, columns.purchase_amount).value)
        if amount is not None:
            return self._money_for_response(amount)

        quantity = self._decimal_or_zero(ws.cell(row, columns.quantity).value)
        unit_price = self._decimal_or_zero(ws.cell(row, 12).value)
        tax = self._round_money(quantity * unit_price * Decimal("0.13"))
        return self._money_for_response(quantity * unit_price + tax)

    def _read_sales_amount(
        self,
        values_ws: Worksheet,
        formula_ws: Worksheet,
        row: int,
        columns: SourceColumns,
    ) -> Optional[float]:
        amount = self._parse_decimal(values_ws.cell(row, columns.sales_amount).value)
        if amount is not None:
            return self._money_for_response(amount)

        quantity = self._decimal_or_zero(values_ws.cell(row, columns.quantity).value)
        unit_price = self._decimal_or_zero(values_ws.cell(row, columns.tms_unit_price).value)
        promo_fee = self._decimal_or_zero(values_ws.cell(row, columns.promo_fee).value)
        tax = self._round_money((quantity * unit_price + promo_fee) * Decimal("0.13"))
        formula_value = self._clean_text(formula_ws.cell(row, columns.sales_amount).value)
        if formula_value.startswith("="):
            return self._money_for_response(quantity * unit_price + tax + promo_fee)
        return None

    def _resolve_target_columns(self, ws: Worksheet) -> Dict[str, int]:
        header_map = self._build_target_header_map(ws)
        missing_headers: List[str] = []
        columns: Dict[str, int] = {}
        for field, header in self.TARGET_FIELD_HEADERS.items():
            normalized_header = self._normalize_header(header)
            column = header_map.get(normalized_header)
            if column is None:
                missing_headers.append(header)
            else:
                columns[field] = column
        if missing_headers:
            raise ValueError("内销对账单缺少必要列：" + "、".join(missing_headers))
        return columns

    def _build_target_header_map(self, ws: Worksheet) -> Dict[str, int]:
        mapping: Dict[str, int] = {}
        for column in range(1, ws.max_column + 1):
            header = self._normalize_header(ws.cell(1, column).value)
            if header:
                mapping[header] = column
        return mapping

    def _parse_exclusion_config(
        self,
        workbook: openpyxl.Workbook,
        target_ws: Worksheet,
    ) -> ExclusionConfig:
        if self.EXCLUSION_SHEET_NAME not in workbook.sheetnames:
            return ExclusionConfig(
                rows=set(),
                columns=set(),
                diagnostics=[{"reason": "未提供排除行列配置"}],
            )

        rules_ws = workbook[self.EXCLUSION_SHEET_NAME]
        target_header_map = self._build_target_header_map(target_ws)
        rows: set[int] = set()
        columns: set[int] = set()
        errors: List[str] = []

        for row_index in range(2, rules_ws.max_row + 1):
            for token in self._split_rule_tokens(rules_ws.cell(row_index, 1).value):
                parsed_row = self._parse_excluded_row(token)
                if parsed_row is None:
                    errors.append(f"A{row_index} 行号无效：{token}")
                else:
                    rows.add(parsed_row)

            for token in self._split_rule_tokens(rules_ws.cell(row_index, 2).value):
                parsed_column = self._parse_excluded_column(token, target_ws, target_header_map)
                if parsed_column is None:
                    errors.append(f"B{row_index} 列无效：{token}")
                else:
                    columns.add(parsed_column)

        if errors:
            raise ValueError("排除配置无效：" + "；".join(errors))
        return ExclusionConfig(rows=rows, columns=columns, diagnostics=[])

    def _split_rule_tokens(self, value: Any) -> List[str]:
        text = self._clean_text(value)
        if not text:
            return []
        return [token.strip() for token in re.split(r"[,，;；\r\n\t]+", text) if token.strip()]

    def _parse_excluded_row(self, token: str) -> Optional[int]:
        if not token.isdigit():
            return None
        row = int(token)
        return row if row > 0 else None

    def _parse_excluded_column(
        self,
        token: str,
        target_ws: Worksheet,
        target_header_map: Dict[str, int],
    ) -> Optional[int]:
        if re.fullmatch(r"[A-Za-z]{1,3}", token):
            try:
                column = column_index_from_string(token.upper())
            except ValueError:
                return None
            return column if 1 <= column <= target_ws.max_column else None

        column = target_header_map.get(self._normalize_header(token))
        if column is None:
            return None
        return column

    def _find_first_empty_row(self, ws: Worksheet) -> int:
        max_column = max(ws.max_column, len(self.TARGET_HEADERS))
        for row in range(2, ws.max_row + 2):
            if all(ws.cell(row, column).value in (None, "") for column in range(1, max_column + 1)):
                return row
        return ws.max_row + 1

    def _write_target_row(
        self,
        ws: Worksheet,
        row: int,
        item: ExtractedRow,
        target_columns: Dict[str, int],
        excluded_columns: set[int],
    ) -> None:
        values = dict(item.values)
        values["purchase_display"] = self._build_invoice_display(
            item.commercial_invoice,
            ws.cell(row, target_columns["purchase_invoice"]).value,
        )
        values["sales_display"] = self._build_invoice_display(
            item.commercial_invoice,
            ws.cell(row, target_columns["sales_invoice"]).value,
        )

        for field in self.WRITABLE_FIELDS:
            column = target_columns[field]
            if column in excluded_columns:
                continue
            cell = ws.cell(row, column)
            cell.value = values.get(field, "")
            if field in self.MONEY_FIELDS:
                cell.number_format = "0.00"

    def _build_invoice_display(self, commercial_invoice: str, invoice_value: Any) -> str:
        invoice = self._clean_text(invoice_value)
        if not commercial_invoice or not invoice:
            return ""
        return f"#{commercial_invoice}({invoice})"

    def _map_vendor(self, vendor_code: Any, vendor_name: Any) -> str:
        code = self._normalize_vendor_key(self._clean_text(vendor_code))
        name = self._clean_text(vendor_name)
        if code in self.VENDOR_CODE_MAP:
            return self.VENDOR_CODE_MAP[code]
        normalized_name = self._normalize_vendor_key(name)
        return self.VENDOR_NAME_ALIASES.get(normalized_name, name or code)

    def _is_row_empty(self, ws: Worksheet, row: int) -> bool:
        return all(
            self._clean_text(ws.cell(row, column).value) == ""
            for column in range(1, ws.max_column + 1)
        )

    def _normalize_vendor_key(self, value: str) -> str:
        return value.strip().upper() if value.isascii() else value.strip()

    def _clean_placeholder(self, value: Any) -> str:
        text = self._clean_text(value)
        return "" if text.lower() in {"(blank)", "(空白)", "none"} else text

    def _clean_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).replace("\u200b", "").replace("\ufeff", "").strip()

    def _normalize_header(self, value: Any) -> str:
        return re.sub(r"\s+", "", self._clean_text(value)).lower()

    def _normalize_style(self, value: Any) -> str:
        text = self._clean_text(value).upper()
        while text.endswith("."):
            text = text[:-1]
        return text

    def _normalize_order(self, value: Any) -> str:
        text = self._clean_text(value)
        if text.endswith(".0") and text[:-2].isdigit():
            return text[:-2]
        return text

    def _money_or_blank(self, value: Any) -> Any:
        amount = self._parse_decimal(value)
        return "" if amount is None else self._money_for_response(amount)

    def _number_for_cell(self, value: Any) -> Any:
        number = self._parse_decimal(value)
        if number is None:
            return value
        return self._number_for_response(number)

    def _parse_decimal(self, value: Any) -> Optional[Decimal]:
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value).replace(",", "").strip())
        except (InvalidOperation, AttributeError):
            return None

    def _decimal_or_zero(self, value: Any) -> Decimal:
        return self._parse_decimal(value) or Decimal("0")

    def _round_money(self, value: Decimal) -> Decimal:
        return value.quantize(Decimal("0.01"), rounding=ROUND_HALF_UP)

    def _money_for_response(self, value: Decimal) -> float:
        return float(self._round_money(value))

    def _number_for_response(self, value: Decimal) -> int | float:
        if value == value.to_integral_value():
            return int(value)
        return float(value)

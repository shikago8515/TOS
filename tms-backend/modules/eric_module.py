# -*- coding: utf-8 -*-
"""
Eric Excel data processing integration module.

Flow:
1. Insert PO Number1 / Article Number1 helper columns.
2. Split repeated PO header blocks into separate sheets.
3. Unpivot size columns into a normalized Final_Data sheet.
"""

import copy
import os
import re
from contextlib import redirect_stdout
from datetime import date, datetime, timedelta
from io import StringIO
from typing import Any, Callable, Dict, List, Optional, Sequence, Tuple

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Font, PatternFill
from openpyxl.styles import numbers
from openpyxl.utils import get_column_letter

try:
    import xlrd
except ImportError:  # pragma: no cover - backend requirements include xlrd
    xlrd = None


DATE_FORMAT = "MM/DD/YYYY"
EXCEL_DATE_FORMAT = numbers.FORMAT_DATE_XLSX14
DATE_INPUT_FORMATS = ("%Y-%m-%d", "%Y/%m/%d", "%m/%d/%y", "%m/%d/%Y", "%d/%m/%y", "%d/%m/%Y")
DATE_SERIAL_MIN = 20000
DATE_SERIAL_MAX = 80000
SIZE_COL_START = 10
SIZE_NAME_ALIASES = {
    "A2XL": "A/2XL",
}
FINAL_DATA_MIN_WIDTHS = {
    1: 13,
    2: 18,
    3: 16,
    4: 12,
    5: 15,
    6: 21,
    7: 24,
    8: 13,
    9: 17,
    10: 10,
    11: 11,
}
FINAL_DATA_MAX_WIDTH = 36
MAX_AUTOFIT_SCAN_ROWS = 1000
MAX_AUTOFIT_SCAN_COLS = 80
CHECK_OK = "OK"
CHECK_FINAL_ONLY = "FINAL_ONLY"
CHECK_YTIC_ONLY = "YTIC_ONLY"
CHECK_QTY_DIFF = "QTY_DIFF"
YTIC_DESTINATION_HEADERS = [
    "CUSTOMER PO NUMBER",
    "*STYLE NUMBER",
    "CUSTOMER SEASON",
    "CUSTOMER SEASON YEAR",
    "TMS OFFICE",
    "DISTRIBUTOR DIVISION",
    "*CUSTOMER DELIVERY DATE",
    "DELIVERY DATE",
    "DATE MARGIN",
    "*SHIP MODE",
    "SHIP MODE",
    "SHIP MODE NOTE",
    "*DESTINATION",
    "DESTINATION",
    "DESTINATION COUNT",
    "PO QTY",
    "SIZE RANGE",
    "UOM",
]
YTIC_SP_HEADERS = [
    "CUSTOMER PO NUMBER",
    "*STYLE NUMBER",
    "CUSTOMER SEASON",
    "CUSTOMER SEASON YEAR",
    "BILL TO",
    "CONSIGNEE",
    "SHIP TO",
    "DATE",
    "*CUSTOMER DELIVERY DATE",
    "",
    "*SHIP MODE",
    "*DESTINATION",
    "PO QTY",
    "SIZE RANGE",
    "UOM",
]
# Eric 要求这些国家在 SP_Extract 中整行标浅绿色，便于人工核对。
ERIC_SP_COUNTRY_HIGHLIGHT_COLOR = "FFEBF1DE"
ERIC_SP_COUNTRY_HIGHLIGHT_DESTINATIONS = {
    "GERMANY",
    "UNITED STATES",
    "ITALY",
    "UNITED KINGDOM",
    "JAPAN",
    "HONG KONG",
    "MONACO",
}

SheetRows = List[List[Any]]
WorkbookRows = Dict[str, SheetRows]
QuantityKey = Tuple[str, str, str]
RowFillResolver = Callable[[Sequence[Any]], Optional[PatternFill]]


def ensure_dir(dir_path: str):
    os.makedirs(dir_path, exist_ok=True)


class EricModule:
    """Excel\u6570\u636e\u5904\u7406\u6574\u5408\u5de5\u5177-Eric."""

    @staticmethod
    def get_actual_data_rows(ws):
        actual_rows = 0
        for row in range(2, ws.max_row + 1):
            b_value = ws.cell(row=row, column=2).value
            if b_value is not None and b_value != "":
                actual_rows = row
        if actual_rows == 0:
            for col in range(1, ws.max_column + 1):
                for row in range(2, ws.max_row + 1):
                    cell_value = ws.cell(row=row, column=col).value
                    if cell_value is not None and cell_value != "":
                        actual_rows = max(row, actual_rows)
        return max(actual_rows, 2)

    @staticmethod
    def is_header_row(ws, row: int) -> bool:
        val = ws.cell(row=row, column=1).value
        return bool(val and 'PO Number' in str(val))

    @staticmethod
    def is_empty_row(ws, row: int, max_col: Optional[int] = None) -> bool:
        max_col = max_col or ws.max_column
        return all(ws.cell(row=row, column=col).value is None for col in range(1, max_col + 1))

    @staticmethod
    def copy_cell_style(source, target):
        if not source.has_style:
            return
        try:
            target.font = copy.copy(source.font)
            target.border = copy.copy(source.border)
            target.fill = copy.copy(source.fill)
            target.number_format = source.number_format
            target.alignment = copy.copy(source.alignment)
        except Exception:
            pass

    @staticmethod
    def copy_column_widths(source_ws, target_ws, col_count: int):
        for col in range(1, col_count + 1):
            col_letter = get_column_letter(col)
            if source_ws.column_dimensions[col_letter].width:
                target_ws.column_dimensions[col_letter].width = source_ws.column_dimensions[col_letter].width

    @staticmethod
    def estimate_text_width(value) -> int:
        if value is None:
            return 0
        return sum(2 if ord(char) > 255 else 1 for char in str(value))

    @staticmethod
    def autofit_columns(ws, min_width=8, max_width=50):
        max_row = min(ws.max_row, MAX_AUTOFIT_SCAN_ROWS)
        max_column = min(ws.max_column, MAX_AUTOFIT_SCAN_COLS)
        for col_idx in range(1, max_column + 1):
            max_length = 0
            column_letter = get_column_letter(col_idx)
            for row_idx in range(1, max_row + 1):
                try:
                    cell = ws.cell(row=row_idx, column=col_idx)
                    max_length = max(max_length, EricModule.estimate_text_width(cell.value))
                except Exception:
                    pass
            adjusted_width = min(max(max_length + 2, min_width), max_width)
            ws.column_dimensions[column_letter].width = adjusted_width

    @staticmethod
    def apply_final_data_column_widths(ws):
        EricModule.autofit_columns(ws, min_width=8, max_width=FINAL_DATA_MAX_WIDTH)
        for col_index, min_width in FINAL_DATA_MIN_WIDTHS.items():
            col_letter = get_column_letter(col_index)
            current_width = ws.column_dimensions[col_letter].width or 0
            ws.column_dimensions[col_letter].width = max(current_width, min_width)

    @staticmethod
    def add_auto_filter(ws):
        if ws.max_column == 0:
            return
        max_col_letter = get_column_letter(ws.max_column)
        ws.auto_filter.ref = f"A1:{max_col_letter}{ws.max_row}"

    @staticmethod
    def format_date_value(value):
        if value is None or value == "":
            return ""
        return EricModule.normalize_date_output_value(value)

    @staticmethod
    def is_output_date_header(header: Any) -> bool:
        text = re.sub(r"\s+", " ", str(header or "").strip().upper())
        if text == "PODD":
            return True
        return "DATE" in text and "MARGIN" not in text

    @staticmethod
    def normalize_date_output_value(value: Any) -> Any:
        if value is None or value == "":
            return value
        if isinstance(value, datetime):
            return value.date()
        if isinstance(value, date):
            return value
        if isinstance(value, (int, float)) and not isinstance(value, bool):
            if DATE_SERIAL_MIN <= float(value) <= DATE_SERIAL_MAX:
                try:
                    excel_epoch = datetime(1899, 12, 30)
                    return (excel_epoch + timedelta(days=float(value))).date()
                except Exception:
                    return value
            return value
        if isinstance(value, str):
            text = value.strip()
            if not text:
                return ""
            if re.fullmatch(r"\d+(?:\.\d+)?", text):
                numeric_value = float(text)
                if DATE_SERIAL_MIN <= numeric_value <= DATE_SERIAL_MAX:
                    try:
                        excel_epoch = datetime(1899, 12, 30)
                        return (excel_epoch + timedelta(days=numeric_value)).date()
                    except Exception:
                        return value
            for fmt in DATE_INPUT_FORMATS:
                try:
                    return datetime.strptime(text, fmt).date()
                except ValueError:
                    continue
            return value
        return value

    @staticmethod
    def apply_date_column_formats(ws, headers: Sequence[str]):
        date_columns = [
            index
            for index, header in enumerate(headers, start=1)
            if EricModule.is_output_date_header(header)
        ]
        for column in date_columns:
            for row in range(2, ws.max_row + 1):
                cell = ws.cell(row=row, column=column)
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = EXCEL_DATE_FORMAT

    @staticmethod
    def format_ytic_date(value: Any) -> Any:
        return EricModule.normalize_date_output_value(value)

    @staticmethod
    def convert_text_to_number(value):
        if value is None or value == "":
            return value
        try:
            return int(value)
        except (ValueError, TypeError):
            try:
                return float(value)
            except (ValueError, TypeError):
                return value

    @staticmethod
    def normalize_size_name(value):
        if value is None:
            return ""
        size_name = str(value).strip()
        return SIZE_NAME_ALIASES.get(size_name, size_name)

    @staticmethod
    def normalize_output_value(value):
        if value is None:
            return None
        if isinstance(value, str) and not value.strip():
            return None
        return value

    def add_and_fill_columns(self, input_file: str, output_file: str) -> str:
        print("[1/3] \u6dfb\u52a0 PO Number1 / Article Number1\uff0c\u5e76\u5411\u4e0b\u586b\u5145 PO Number1")
        wb = load_workbook(input_file)

        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            actual_rows = self.get_actual_data_rows(ws)
            print(f"  \u5904\u7406 Sheet: {sheet_name}\uff0c\u5b9e\u9645\u6570\u636e\u884c: {actual_rows}")

            ws.insert_cols(8)
            ws.cell(row=1, column=8).value = "PO Number1"
            for row in range(2, actual_rows + 1):
                ws.cell(row=row, column=8).value = ws.cell(row=row, column=1).value

            ws.insert_cols(9)
            ws.cell(row=1, column=9).value = "Article Number1"
            for row in range(2, actual_rows + 1):
                ws.cell(row=row, column=9).value = ws.cell(row=row, column=3).value

            last_value = None
            filled_count = 0
            for row in range(2, actual_rows + 1):
                cell = ws.cell(row=row, column=8)
                current_val = cell.value
                if current_val is None or current_val == "":
                    if last_value is not None:
                        cell.value = last_value
                        filled_count += 1
                else:
                    last_value = current_val
            print(f"    PO Number1 \u7a7a\u503c\u586b\u5145: {filled_count}")

        wb.save(output_file)
        print(f"  \u8f93\u51fa step1: {output_file}")
        return output_file

    def process_split(self, input_file: str, output_file: str) -> str:
        print("[2/3] \u6309 PO \u6807\u9898\u884c\u62c6\u5206 Sheet")
        wb = load_workbook(input_file, data_only=True)
        new_wb = Workbook()
        new_wb.remove(new_wb.active)
        sheet_counter = 1

        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            header_rows = [r for r in range(1, ws.max_row + 1) if self.is_header_row(ws, r)]
            print(f"  \u539f Sheet: {ws_name}\uff0c\u627e\u5230\u6570\u636e\u5757: {len(header_rows)}")

            for idx, header_row in enumerate(header_rows):
                data_end = header_rows[idx + 1] - 1 if idx + 1 < len(header_rows) else ws.max_row
                if idx + 1 == len(header_rows):
                    data_end = max(
                        (r for r in range(ws.max_row, header_row, -1) if not self.is_empty_row(ws, r)),
                        default=ws.max_row
                    )

                new_ws = new_wb.create_sheet(title=str(sheet_counter))

                for source_row in range(header_row, data_end + 1):
                    if self.is_empty_row(ws, source_row):
                        continue
                    # 拆分文件只作为下一步数据源，跳过样式复制能明显降低大文件耗时。
                    new_ws.append([
                        ws.cell(row=source_row, column=col).value
                        for col in range(1, ws.max_column + 1)
                    ])

                first_po = ws.cell(row=header_row + 1, column=1).value if header_row + 1 <= data_end else "N/A"
                print(f"    \u62c6\u5206 Sheet {sheet_counter}\uff0cPO: {first_po}")
                sheet_counter += 1

        if not new_wb.sheetnames:
            raise ValueError("\u672a\u627e\u5230\u5305\u542b 'PO Number' \u7684\u6807\u9898\u884c\uff0c\u65e0\u6cd5\u62c6\u5206\u6570\u636e\u5757")

        new_wb.save(output_file)
        print(f"  \u8f93\u51fa split: {output_file}")
        return output_file

    def process_unpivot(self, input_file: str, output_file: str) -> str:
        print("[3/3] \u9006\u900f\u89c6\u5c3a\u7801\u5217\u5e76\u751f\u6210 Final_Data")
        wb = load_workbook(input_file, data_only=True)
        new_wb = Workbook()
        new_wb.remove(new_wb.active)

        new_headers = [
            "PO Number", "Working Number", "Article Number", "PODD", "Shipment Mode",
            "Gps Customer Number", "Country", "PO Number1", "Article Number1", "Size", "Quantity"
        ]

        ws_out = new_wb.create_sheet("Final_Data")
        for c, header in enumerate(new_headers, 1):
            ws_out.cell(row=1, column=c, value=header)

        row_count = 0
        emitted_po_keys = set()

        def append_final_row(fixed_values, size_name, quantity):
            nonlocal row_count
            row_values = list(fixed_values)
            po_key = row_values[7] or row_values[0]
            if po_key:
                if po_key in emitted_po_keys:
                    row_values[0] = None
                else:
                    emitted_po_keys.add(po_key)
            output_row = [
                self.normalize_output_value(value)
                for value in row_values + [self.normalize_size_name(size_name), quantity]
            ]
            ws_out.append(output_row)
            row_count += 1

            current_row = ws_out.max_row
            ws_out.cell(row=current_row, column=4).number_format = EXCEL_DATE_FORMAT
            ws_out.cell(row=current_row, column=8).number_format = numbers.FORMAT_GENERAL
            ws_out.cell(row=current_row, column=6).number_format = "@"
            ws_out.cell(row=current_row, column=10).number_format = "@"

        for ws_name in wb.sheetnames:
            ws = wb[ws_name]
            headers = [ws.cell(row=1, column=col).value for col in range(1, ws.max_column + 1)]
            size_cols = [(c, headers[c - 1]) for c in range(SIZE_COL_START, ws.max_column + 1) if headers[c - 1]]
            print(f"  \u5904\u7406 Sheet: {ws_name}\uff0c\u5c3a\u7801\u5217: {len(size_cols)}")

            for row in range(2, ws.max_row + 1):
                if self.is_empty_row(ws, row, 9):
                    continue
                fixed = [ws.cell(row=row, column=c).value or "" for c in range(1, 10)]
                fixed[3] = self.format_date_value(fixed[3])
                fixed[7] = self.convert_text_to_number(fixed[7])

                has_qty = False
                for col, size_name in size_cols:
                    val = ws.cell(row=row, column=col).value
                    if val is None or not str(val).strip():
                        continue
                    try:
                        qty = float(val)
                    except (TypeError, ValueError):
                        continue
                    if qty > 0:
                        append_final_row(fixed, size_name, int(qty))
                        has_qty = True
                if not has_qty:
                    append_final_row(fixed, "", "")

        self.apply_final_data_column_widths(ws_out)
        self.add_auto_filter(ws_out)
        new_wb.save(output_file)
        print(f"  \u8f93\u51fa final: {output_file}")
        print(f"  Final_Data \u884c\u6570: {row_count}")
        return output_file

    @staticmethod
    def normalize_text_key(value) -> str:
        if value is None:
            return ""
        return str(value).strip().upper()

    @staticmethod
    def normalize_po_key(value) -> str:
        if value is None or value == "":
            return ""
        text = str(value).strip()
        if text.endswith(".0"):
            text = text[:-2]
        digits = re.sub(r"\D", "", text)
        if not digits:
            return ""
        return digits.zfill(10)

    @staticmethod
    def normalize_quantity(value) -> Optional[int]:
        if value is None or value == "":
            return None
        try:
            return int(round(float(value)))
        except (TypeError, ValueError):
            return None

    @staticmethod
    def is_meaningful_row(row: Sequence[Any]) -> bool:
        return any(value is not None and str(value).strip() != "" for value in row)

    @staticmethod
    def clean_header(value: Any, fallback_index: int) -> str:
        if value is None or str(value).strip() == "":
            return f"Column {fallback_index + 1}"
        return str(value).strip()

    @staticmethod
    def read_sheet_table(rows: SheetRows) -> Tuple[List[str], List[List[Any]]]:
        if not rows:
            return [], []

        header_index = 0
        for index, row in enumerate(rows):
            if EricModule.is_meaningful_row(row):
                header_index = index
                break

        headers = [
            EricModule.clean_header(value, index)
            for index, value in enumerate(rows[header_index])
        ]
        data_rows = [
            list(row)
            for row in rows[header_index + 1:]
            if EricModule.is_meaningful_row(row)
        ]
        return headers, data_rows

    @staticmethod
    def index_headers(headers: Sequence[str]) -> Dict[str, int]:
        return {str(header).strip().upper(): index for index, header in enumerate(headers)}

    @staticmethod
    def row_value(row: Sequence[Any], index: int) -> Any:
        if 0 <= index < len(row):
            return row[index]
        return None

    @staticmethod
    def normalize_ship_mode(value: Any, destination: Any = None) -> str:
        text = EricModule.normalize_text_key(value)
        if text in {"BY SEA", "SEA", "OCEAN"}:
            return "Ocean"
        if text in {"BY AIR", "AIR"}:
            return "Air"
        if text in {"BY COURIER", "COURIER", "AIR EXPRESS"}:
            if EricModule.normalize_text_key(destination) == "GERMANY":
                return "Ocean"
            return "Air Express"
        return "" if not text else str(value).strip()

    @staticmethod
    def standardize_ytic_headers(headers: Sequence[str], standard_headers: Sequence[str]) -> List[str]:
        if len(headers) == len(standard_headers):
            return list(standard_headers)
        return list(headers)

    def read_workbook_rows(self, file_path: str) -> WorkbookRows:
        extension = os.path.splitext(file_path)[1].lower()
        if extension == ".xls":
            if xlrd is None:
                raise ValueError("后端缺少 xlrd，无法读取 .xls 文件")
            book = xlrd.open_workbook(file_path, on_demand=True)
            try:
                return {
                    sheet_name: [
                        [sheet.cell_value(row, col) for col in range(sheet.ncols)]
                        for row in range(sheet.nrows)
                    ]
                    for sheet_name in book.sheet_names()
                    for sheet in [book.sheet_by_name(sheet_name)]
                }
            finally:
                book.release_resources()

        wb = load_workbook(file_path, read_only=True, data_only=True)
        try:
            return {
                ws.title: [list(row) for row in ws.iter_rows(values_only=True)]
                for ws in wb.worksheets
            }
        finally:
            wb.close()

    def read_final_data(self, final_file: str) -> Dict[str, Any]:
        wb = load_workbook(final_file, read_only=True, data_only=True)
        try:
            if "Final_Data" not in wb.sheetnames:
                raise ValueError("结果文件缺少 Final_Data 工作表")

            ws = wb["Final_Data"]
            row_iter = ws.iter_rows(values_only=True)
            header_row = next(row_iter, None)
            if not header_row:
                raise ValueError("Final_Data 工作表为空")

            headers = [str(value).strip() if value is not None else "" for value in header_row]
            records = [
                dict(zip(headers, row))
                for row in row_iter
                if self.is_meaningful_row(row)
            ]
        finally:
            wb.close()

        quantity_map: Dict[QuantityKey, int] = {}
        quantity_order: List[QuantityKey] = []
        po_set = set()
        po_order: List[str] = []
        for record in records:
            po_key = self.normalize_po_key(record.get("PO Number1") or record.get("PO Number"))
            article_key = self.normalize_text_key(record.get("Article Number1") or record.get("Article Number"))
            size_key = self.normalize_text_key(self.normalize_size_name(record.get("Size")))
            quantity = self.normalize_quantity(record.get("Quantity"))

            if po_key:
                if po_key not in po_set:
                    po_order.append(po_key)
                    po_set.add(po_key)
            if po_key and article_key and size_key and quantity is not None:
                key = (po_key, article_key, size_key)
                if key not in quantity_map:
                    quantity_order.append(key)
                quantity_map[key] = quantity_map.get(key, 0) + quantity

        return {
            "headers": headers,
            "records": records,
            "quantity_map": quantity_map,
            "quantity_order": quantity_order,
            "po_set": po_set,
            "po_order": po_order,
        }

    def find_ytic_blocks(self, rows: SheetRows) -> List[Tuple[int, int]]:
        starts = [
            index
            for index, row in enumerate(rows[1:], 1)
            if self.normalize_po_key(self.row_value(row, 0))
        ]

        blocks = []
        for index, start in enumerate(starts):
            end = starts[index + 1] if index + 1 < len(starts) else len(rows)
            blocks.append((start, end))
        return blocks

    def find_row_containing(self, rows: SheetRows, start: int, end: int, text: str) -> Optional[int]:
        needle = text.upper()
        for row_index in range(start, min(end, len(rows))):
            row_text = " | ".join(
                self.normalize_text_key(value)
                for value in rows[row_index]
                if value is not None
            )
            if needle in row_text:
                return row_index
        return None

    def find_exact_section_label(self, rows: SheetRows, start: int, end: int, text: str) -> Optional[int]:
        needle = self.normalize_text_key(text)
        for row_index in range(start, min(end, len(rows))):
            if any(self.normalize_text_key(value) == needle for value in rows[row_index]):
                return row_index
        return None

    def find_section_header(
        self,
        rows: SheetRows,
        start: int,
        end: int,
        required: Sequence[str],
    ) -> Optional[Tuple[int, Dict[str, int]]]:
        required_set = {item.upper() for item in required}
        for row_index in range(start, min(end, len(rows))):
            header_map = {
                self.normalize_text_key(value): col_index
                for col_index, value in enumerate(rows[row_index])
                if self.normalize_text_key(value)
            }
            if required_set.issubset(header_map.keys()):
                return row_index, header_map
        return None

    def find_first_section_data_row(
        self,
        rows: SheetRows,
        section_row: int,
        block_end: int,
        required: Sequence[str],
    ) -> Optional[List[Any]]:
        header_result = self.find_section_header(rows, section_row + 1, block_end, required)
        if header_result is None:
            return None
        header_row, _header_map = header_result
        return self.first_meaningful_row_after(rows, header_row + 1, block_end)

    @staticmethod
    def calculate_date_margin(left: Any, right: Any) -> Any:
        if left in (None, "") or right in (None, ""):
            return ""
        if isinstance(left, (int, float)) and isinstance(right, (int, float)):
            margin = float(left) - float(right)
            return int(margin) if margin.is_integer() else margin

        left_date = EricModule.normalize_date_output_value(left)
        right_date = EricModule.normalize_date_output_value(right)
        if isinstance(left_date, (datetime, date)) and isinstance(right_date, (datetime, date)):
            return (left_date - right_date).days
        return ""

    @staticmethod
    def normalize_destination_count_flag(value: Any) -> bool:
        if isinstance(value, bool):
            return value
        if isinstance(value, (int, float)):
            return value != 0
        text = str(value).strip().lower()
        # 兼容 Excel 里可能被读成文本的 0/false。
        return text not in ("", "0", "false")

    def extract_ytic_size_rows_from_source(
        self,
        rows: SheetRows,
        final_quantity_map: Dict[QuantityKey, int],
    ) -> List[Dict[str, Any]]:
        size_rows: List[Dict[str, Any]] = []
        for block_start, block_end in self.find_ytic_blocks(rows):
            po_key = self.normalize_po_key(self.row_value(rows[block_start], 0))
            sub_detail_row = self.find_row_containing(rows, block_start, block_end, "PO Sub Details")
            if sub_detail_row is None:
                continue

            header_result = self.find_section_header(
                rows,
                sub_detail_row + 1,
                block_end,
                ("COLOR", "SIZE", "QTY"),
            )
            if header_result is None:
                continue

            header_row, header_map = header_result
            color_col = header_map["COLOR"]
            size_col = header_map["SIZE"]
            qty_col = header_map["QTY"]

            for row in rows[header_row + 1:block_end]:
                color = self.normalize_text_key(self.row_value(row, color_col))
                size = self.normalize_text_key(self.normalize_size_name(self.row_value(row, size_col)))
                qty = self.normalize_quantity(self.row_value(row, qty_col))
                if not (po_key and color and size and qty is not None):
                    if self.is_meaningful_row(row):
                        continue
                    break

                key = (po_key, color, size)
                actual_qty = final_quantity_map.get(key)
                size_rows.append({
                    "po": po_key,
                    "article": color,
                    "size": size,
                    "quantity": qty,
                    "actual_quantity": actual_qty,
                    "margin": None if actual_qty is None else actual_qty - qty,
                    "status": self.status_for_quantities(actual_qty, qty),
                })

        return size_rows

    def extract_ytic_size_rows_from_sheet(
        self,
        rows: SheetRows,
        final_quantity_map: Dict[QuantityKey, int],
    ) -> List[Dict[str, Any]]:
        headers, data_rows = self.read_sheet_table(rows)
        header_index = self.index_headers(headers)
        required = ("PO NO", "COLOR", "SIZE", "QTY")
        if not all(name in header_index for name in required):
            return []

        size_rows: List[Dict[str, Any]] = []
        for row in data_rows:
            po_key = self.normalize_po_key(self.row_value(row, header_index["PO NO"]))
            article = self.normalize_text_key(self.row_value(row, header_index["COLOR"]))
            size = self.normalize_text_key(self.normalize_size_name(self.row_value(row, header_index["SIZE"])))
            qty = self.normalize_quantity(self.row_value(row, header_index["QTY"]))
            if not (po_key and article and size and qty is not None):
                continue

            key = (po_key, article, size)
            actual_qty = final_quantity_map.get(key)
            size_rows.append({
                "po": po_key,
                "article": article,
                "size": size,
                "quantity": qty,
                "actual_quantity": actual_qty,
                "margin": None if actual_qty is None else actual_qty - qty,
                "status": self.status_for_quantities(actual_qty, qty),
            })

        return size_rows

    def extract_destination_rows_from_source(self, rows: SheetRows) -> List[List[Any]]:
        output_rows: List[List[Any]] = []
        for block_start, _block_end in self.find_ytic_blocks(rows):
            row = rows[block_start + 5] if block_start + 5 < len(rows) else []
            po_key = self.normalize_po_key(self.row_value(rows[block_start], 0))
            destination = self.row_value(row, 12)
            output_rows.append([
                po_key,
                self.row_value(row, 2),
                self.row_value(row, 4),
                self.row_value(row, 5),
                self.row_value(row, 7),
                self.row_value(row, 9),
                self.row_value(row, 10),
                self.format_ytic_date(self.row_value(row, 10)),
                0,
                self.row_value(row, 11),
                self.normalize_ship_mode(self.row_value(row, 11), destination),
                "",
                destination,
                self.normalize_destination_name(destination),
                1,
                self.row_value(row, 13),
                self.row_value(row, 14),
                self.row_value(row, 15),
            ])
        return output_rows

    @staticmethod
    def normalize_destination_name(value: Any) -> str:
        text = "" if value is None else str(value).strip()
        # YTIC 参考表里 Peru 大小写不完全一致；源表无额外标记时保留原值。
        if text.upper() == "PERU":
            return text
        return EricModule.normalize_text_key(value)

    @staticmethod
    def to_processed_ytic_row(row: Sequence[Any]) -> List[Any]:
        return [
            value
            for index, value in enumerate(row)
            if index not in (1, 3)
        ]

    def first_meaningful_row_after(self, rows: SheetRows, start: int, end: int) -> Optional[List[Any]]:
        for index in range(start, min(end, len(rows))):
            if self.is_meaningful_row(rows[index]):
                return rows[index]
        return None

    def extract_sp_rows_from_source(self, rows: SheetRows) -> List[List[Any]]:
        output_rows: List[List[Any]] = []
        processed_rows = [self.to_processed_ytic_row(row) for row in rows]

        for block_start, block_end in self.find_ytic_blocks(rows):
            po_key = self.normalize_po_key(self.row_value(rows[block_start], 0))
            line_row = processed_rows[block_start + 5] if block_start + 5 < len(processed_rows) else []
            output_rows.append([
                po_key,
                *[self.row_value(line_row, index) for index in range(1, 8)],
                self.row_value(line_row, 8),
                "",
                *[self.row_value(line_row, index) for index in range(9, 14)],
            ])

            sales_data = None
            sales_row = self.find_exact_section_label(rows, block_start, block_end, "Sales Confirmation")
            if sales_row is not None:
                sales_data = self.find_first_section_data_row(
                    rows,
                    sales_row,
                    block_end,
                    ("CUSTOMER", "*SELLER/SUPPLIER", "*BILL TO", "*CONSIGNEE"),
                )

            purchase_data = None
            purchase_row = self.find_exact_section_label(rows, block_start, block_end, "Purchase Confirmation")
            if purchase_row is not None:
                purchase_data = self.find_first_section_data_row(
                    rows,
                    purchase_row,
                    block_end,
                    ("PC NUMBER", "FACTORY", "*BILL TO", "*CONSIGNEE", "SHIP TO"),
                )

            if sales_data:
                output_rows.append([
                    "",
                    "",
                    self.row_value(sales_data, 4),
                    self.row_value(sales_data, 5),
                    self.row_value(sales_data, 6),
                    self.row_value(sales_data, 7),
                    self.row_value(sales_data, 8),
                    self.row_value(sales_data, 9),
                    self.row_value(sales_data, 10),
                    self.calculate_date_margin(
                        self.row_value(sales_data, 10),
                        self.row_value(purchase_data or [], 10),
                    ),
                    self.row_value(sales_data, 11),
                    self.row_value(sales_data, 12),
                    "",
                    "",
                    "",
                ])

            if purchase_data:
                output_rows.append([
                    "",
                    "",
                    self.row_value(purchase_data, 4),
                    self.row_value(purchase_data, 5),
                    self.row_value(purchase_data, 6),
                    self.row_value(purchase_data, 7),
                    self.row_value(purchase_data, 8),
                    self.row_value(purchase_data, 9),
                    self.row_value(purchase_data, 10),
                    "",
                    self.row_value(purchase_data, 11),
                    self.row_value(purchase_data, 12),
                    "",
                    "",
                    "",
                ])

        return output_rows

    def extract_ytic_tables(
        self,
        ytic_file: str,
        final_quantity_map: Dict[QuantityKey, int],
    ) -> Dict[str, Any]:
        workbook_rows = self.read_workbook_rows(ytic_file)
        lower_name_map = {name.lower(): name for name in workbook_rows}
        source_sheet_name = lower_name_map.get("源表") or next(iter(workbook_rows))
        source_rows = workbook_rows[source_sheet_name]

        size_rows = []
        if "提取尺寸表" in workbook_rows:
            size_rows = self.extract_ytic_size_rows_from_sheet(
                workbook_rows["提取尺寸表"],
                final_quantity_map,
            )
        if not size_rows:
            size_rows = self.extract_ytic_size_rows_from_source(source_rows, final_quantity_map)

        if "提取目的地表" in workbook_rows:
            destination_headers, destination_rows = self.read_sheet_table(workbook_rows["提取目的地表"])
        else:
            destination_headers = list(YTIC_DESTINATION_HEADERS)
            destination_rows = self.extract_destination_rows_from_source(source_rows)
        destination_headers = self.standardize_ytic_headers(destination_headers, YTIC_DESTINATION_HEADERS)

        if "提取SP表" in workbook_rows:
            sp_headers, sp_rows = self.read_sheet_table(workbook_rows["提取SP表"])
        else:
            sp_headers = list(YTIC_SP_HEADERS)
            sp_rows = self.extract_sp_rows_from_source(source_rows)
        sp_headers = self.standardize_ytic_headers(sp_headers, YTIC_SP_HEADERS)

        quantity_map: Dict[QuantityKey, int] = {}
        quantity_order: List[QuantityKey] = []
        po_set = set()
        po_order: List[str] = []
        for row in size_rows:
            po_key = row["po"]
            article_key = row["article"]
            size_key = row["size"]
            quantity = row["quantity"]
            if po_key not in po_set:
                po_order.append(po_key)
                po_set.add(po_key)
            key = (po_key, article_key, size_key)
            if key not in quantity_map:
                quantity_order.append(key)
            quantity_map[key] = quantity_map.get(key, 0) + quantity

        return {
            "size_rows": size_rows,
            "quantity_map": quantity_map,
            "quantity_order": quantity_order,
            "po_set": po_set,
            "po_order": po_order,
            "destination_headers": destination_headers,
            "destination_rows": destination_rows,
            "sp_headers": sp_headers,
            "sp_rows": sp_rows,
            "source_sheet": source_sheet_name,
        }

    @staticmethod
    def status_for_quantities(final_quantity: Optional[int], ytic_quantity: Optional[int]) -> str:
        if final_quantity is None:
            return CHECK_YTIC_ONLY
        if ytic_quantity is None:
            return CHECK_FINAL_ONLY
        if final_quantity == ytic_quantity:
            return CHECK_OK
        return CHECK_QTY_DIFF

    def build_size_check_rows(
        self,
        final_quantity_map: Dict[QuantityKey, int],
        ytic_quantity_map: Dict[QuantityKey, int],
        preferred_order: Optional[Sequence[QuantityKey]] = None,
    ) -> List[List[Any]]:
        rows: List[List[Any]] = []
        all_keys = set(final_quantity_map) | set(ytic_quantity_map)
        ordered_keys: List[QuantityKey] = []
        seen_keys = set()
        if preferred_order is not None:
            # reconciliation 主表以 Check Excel 的尺寸行作为范围，Final_Data 独有项留给 PO 诊断表。
            for key in preferred_order:
                if key in all_keys and key not in seen_keys:
                    ordered_keys.append(key)
                    seen_keys.add(key)
        else:
            ordered_keys.extend(sorted(all_keys))

        for key in ordered_keys:
            final_qty = final_quantity_map.get(key)
            ytic_qty = ytic_quantity_map.get(key)
            status = self.status_for_quantities(final_qty, ytic_qty)
            difference = None
            if final_qty is not None and ytic_qty is not None:
                difference = final_qty - ytic_qty
            rows.append([
                key[0],
                key[1],
                key[2],
                final_qty,
                ytic_qty,
                difference,
                status,
            ])
        return rows

    def build_po_check_rows(self, final_po_set, ytic_po_set, preferred_order: Optional[Sequence[str]] = None) -> List[List[Any]]:
        rows: List[List[Any]] = []
        all_po = final_po_set | ytic_po_set
        ordered_po: List[str] = []
        seen_po = set()
        for po_key in preferred_order or []:
            if po_key in all_po and po_key not in seen_po:
                ordered_po.append(po_key)
                seen_po.add(po_key)
        ordered_po.extend(sorted(all_po - seen_po))

        for po_key in ordered_po:
            in_final = po_key in final_po_set
            in_ytic = po_key in ytic_po_set
            if in_final and in_ytic:
                status = CHECK_OK
            elif in_final:
                status = CHECK_FINAL_ONLY
            else:
                status = CHECK_YTIC_ONLY
            rows.append([po_key, in_final, in_ytic, status])
        return rows

    @staticmethod
    def normalize_compare_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, bool):
            return str(value)
        if isinstance(value, int):
            return str(value)
        if isinstance(value, float):
            return str(int(value)) if value.is_integer() else str(value).strip()
        text = str(value).strip()
        if re.fullmatch(r"\d+\.0", text):
            return text[:-2]
        return text

    @staticmethod
    def ordered_unique_text(values: Sequence[Any]) -> List[str]:
        ordered: List[str] = []
        seen = set()
        for value in values:
            text = EricModule.normalize_compare_text(value)
            if text and text not in seen:
                ordered.append(text)
                seen.add(text)
        return ordered

    def build_po_text_compare_rows(
        self,
        final_data: Dict[str, Any],
        ytic_data: Dict[str, Any],
    ) -> List[List[Any]]:
        final_po_values = [
            record.get("PO Number")
            for record in final_data.get("records", [])
        ]
        destination_headers = ytic_data.get("destination_headers", [])
        destination_header_index = self.index_headers(destination_headers)
        destination_po_index = destination_header_index.get("CUSTOMER PO NUMBER")
        ytic_po_values = []
        if destination_po_index is not None:
            ytic_po_values = [
                self.row_value(row, destination_po_index)
                for row in ytic_data.get("destination_rows", [])
            ]

        final_po_order = self.ordered_unique_text(final_po_values)
        ytic_po_order = self.ordered_unique_text(ytic_po_values)
        final_po_set = set(final_po_order)
        ytic_po_set = set(ytic_po_order)

        ordered_po: List[str] = []
        seen_po = set()
        # 按 YTIC 目的地表顺序优先输出，方便和来源表人工核对。
        for po_text in [*ytic_po_order, *final_po_order]:
            if po_text and po_text not in seen_po:
                ordered_po.append(po_text)
                seen_po.add(po_text)

        rows: List[List[Any]] = []
        for po_text in ordered_po:
            in_final = po_text in final_po_set
            in_ytic = po_text in ytic_po_set
            if in_final and in_ytic:
                status = CHECK_OK
            elif in_final:
                status = CHECK_FINAL_ONLY
            else:
                status = CHECK_YTIC_ONLY
            rows.append([po_text, in_final, in_ytic, status])
        return rows

    def write_table_sheet(
        self,
        wb: Workbook,
        title: str,
        headers: Sequence[str],
        rows: Sequence[Sequence[Any]],
        row_fill_resolver: Optional[RowFillResolver] = None,
    ):
        ws = wb.create_sheet(title[:31])
        ws.append(list(headers))
        date_column_indexes = {
            index
            for index, header in enumerate(headers, start=1)
            if self.is_output_date_header(header)
        }
        for row in rows:
            normalized_row = [
                self.normalize_date_output_value(value) if index in date_column_indexes else value
                for index, value in enumerate(row, start=1)
            ]
            ws.append(normalized_row)
            current_row = ws.max_row
            for column in date_column_indexes:
                cell = ws.cell(row=current_row, column=column)
                if isinstance(cell.value, (datetime, date)):
                    cell.number_format = EXCEL_DATE_FORMAT
            if row_fill_resolver:
                row_fill = row_fill_resolver(normalized_row)
                if row_fill:
                    for column in range(1, len(headers) + 1):
                        ws.cell(row=current_row, column=column).fill = copy.copy(row_fill)

        if headers:
            max_col_letter = get_column_letter(len(headers))
            ws.auto_filter.ref = f"A1:{max_col_letter}{max(ws.max_row, 1)}"
            ws.freeze_panes = "A2"
            header_fill = PatternFill("solid", fgColor="1F4E78")
            header_font = Font(bold=True, color="FFFFFF")
            for cell in ws[1]:
                cell.fill = header_fill
                cell.font = header_font

        self.autofit_columns(ws, min_width=8, max_width=42)

    @staticmethod
    def build_ytic_sp_row_fill_resolver(headers: Sequence[str]) -> RowFillResolver:
        header_index = EricModule.index_headers(headers)
        destination_index = header_index.get("*DESTINATION")
        if destination_index is None:
            destination_index = header_index.get("DESTINATION")
        highlight_fill = PatternFill("solid", fgColor=ERIC_SP_COUNTRY_HIGHLIGHT_COLOR)

        def resolve(row: Sequence[Any]) -> Optional[PatternFill]:
            if destination_index is None:
                return None
            destination = EricModule.row_value(row, destination_index)
            destination_key = EricModule.normalize_text_key(destination)
            if destination_key in ERIC_SP_COUNTRY_HIGHLIGHT_DESTINATIONS:
                return highlight_fill
            return None

        return resolve

    def write_reconciliation_workbook(
        self,
        output_file: str,
        final_data: Dict[str, Any],
        ytic_data: Dict[str, Any],
        size_check_rows: List[List[Any]],
        po_check_rows: List[List[Any]],
        summary_rows: List[List[Any]],
    ) -> str:
        wb = Workbook()
        ws_summary = wb.active
        ws_summary.title = "Summary"
        ws_summary.append(["项目", "值"])
        for row in summary_rows:
            ws_summary.append(row)

        for cell in ws_summary[1]:
            cell.fill = PatternFill("solid", fgColor="1F4E78")
            cell.font = Font(bold=True, color="FFFFFF")
        ws_summary.freeze_panes = "A2"
        self.autofit_columns(ws_summary, min_width=12, max_width=46)

        size_check_export_rows = [
            [
                row[0],
                row[1],
                row[2],
                row[4],
                row[3],
                f"=E{row_index}-D{row_index}",
            ]
            for row_index, row in enumerate(size_check_rows, start=2)
        ]
        self.write_table_sheet(
            wb,
            "Size_Check",
            ["PO Number", "Article Number", "Size", "PO Quantity", "Final Quantity", "margin"],
            size_check_export_rows,
        )
        self.write_table_sheet(
            wb,
            "PO_Check",
            ["PO Number", "In Final_Data", "In YTIC", "Status"],
            po_check_rows,
        )

        final_headers = final_data["headers"]
        final_rows = [
            [record.get(header) for header in final_headers]
            for record in final_data["records"]
        ]
        self.write_table_sheet(wb, "Final_Data", final_headers, final_rows)

        ytic_size_rows = [
            [
                row["po"],
                row["article"],
                row["size"],
                row["quantity"],
                row["actual_quantity"],
                row["margin"],
                row["status"],
            ]
            for row in ytic_data["size_rows"]
        ]
        self.write_table_sheet(
            wb,
            "YTIC_Size_Extract",
            ["PO NO", "COLOR", "SIZE", "QTY", "Actual QTY", "QTY Margin", "Status"],
            ytic_size_rows,
        )
        ytic_destination_rows = []
        for row_index, row in enumerate(ytic_data["destination_rows"], start=2):
            export_row = list(row)
            if len(export_row) > 8 and export_row[8] not in (None, ""):
                # I 列日期差值由后面的 DELIVERY DATE 减前面的 *CUSTOMER DELIVERY DATE。
                export_row[8] = f"=H{row_index}-G{row_index}"
            if len(export_row) > 14 and export_row[14] not in (None, ""):
                # O 列用 Excel 布尔值表达目的地是否匹配。
                export_row[14] = self.normalize_destination_count_flag(export_row[14])
            ytic_destination_rows.append(export_row)
        self.write_table_sheet(
            wb,
            "Destination1Extract",
            ytic_data["destination_headers"],
            ytic_destination_rows,
        )
        ytic_sp_rows = []
        for row_index, row in enumerate(ytic_data["sp_rows"], start=2):
            export_row = list(row)
            if len(export_row) > 9 and export_row[9] not in (None, ""):
                # J 列日期差值由当前行和下一行的 *CUSTOMER DELIVERY DATE 公式计算。
                export_row[9] = f"=I{row_index}-I{row_index + 1}"
            ytic_sp_rows.append(export_row)
        self.write_table_sheet(
            wb,
            "SP_Extract",
            ytic_data["sp_headers"],
            ytic_sp_rows,
            row_fill_resolver=self.build_ytic_sp_row_fill_resolver(ytic_data["sp_headers"]),
        )
        for sheet_name in ("Summary", "PO_Check", "YTIC_Size_Extract"):
            if sheet_name in wb.sheetnames:
                del wb[sheet_name]
        self.write_table_sheet(
            wb,
            "PO_Text_Compare",
            ["PO Text", "In Final_Data PO Number", "In YTIC Destination CUSTOMER PO NUMBER", "Status"],
            self.build_po_text_compare_rows(final_data, ytic_data),
        )

        wb.save(output_file)
        return output_file

    def process_reconciliation(self, pack_file: str, ytic_file: str, output_dir: str) -> Dict[str, Any]:
        result: Dict[str, Any] = {
            "success": False,
            "message": "",
            "logs": [],
            "output_path": None,
            "row_count": 0,
            "difference_count": 0,
            "po_difference_count": 0,
            "size_check_count": 0,
        }

        ensure_dir(output_dir)
        buffer = StringIO()
        try:
            with redirect_stdout(buffer):
                final_result = self.process_file(pack_file, output_dir)
                if not final_result["success"] or not final_result["output_path"]:
                    raise ValueError(final_result["message"] or "Final_Data 生成失败")

                final_data = self.read_final_data(final_result["output_path"])
                ytic_data = self.extract_ytic_tables(ytic_file, final_data["quantity_map"])

                size_check_rows = self.build_size_check_rows(
                    final_data["quantity_map"],
                    ytic_data["quantity_map"],
                    ytic_data.get("quantity_order", []),
                )
                po_check_rows = self.build_po_text_compare_rows(final_data, ytic_data)

                difference_count = sum(1 for row in size_check_rows if row[-1] != CHECK_OK)
                po_difference_count = sum(1 for row in po_check_rows if row[-1] != CHECK_OK)
                final_only_count = sum(1 for row in size_check_rows if row[-1] == CHECK_FINAL_ONLY)
                ytic_only_count = sum(1 for row in size_check_rows if row[-1] == CHECK_YTIC_ONLY)
                qty_diff_count = sum(1 for row in size_check_rows if row[-1] == CHECK_QTY_DIFF)

                summary_rows = [
                    ["Final_Data 行数", len(final_data["records"])],
                    ["Final_Data PO 数", len(final_data["po_set"])],
                    ["YTIC 尺寸行数", len(ytic_data["size_rows"])],
                    ["YTIC PO 数", len(ytic_data["po_set"])],
                    ["Size_Check 总项", len(size_check_rows)],
                    ["数量完全一致项", len(size_check_rows) - difference_count],
                    ["数量差异项", qty_diff_count],
                    ["Final_Data 独有项", final_only_count],
                    ["YTIC 独有项", ytic_only_count],
                    ["PO 差异项", po_difference_count],
                    ["YTIC 来源 Sheet", ytic_data["source_sheet"]],
                    ["核对状态", "通过" if difference_count == 0 and po_difference_count == 0 else "存在差异"],
                ]

                base_name = os.path.splitext(os.path.basename(pack_file))[0]
                timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
                output_file = os.path.join(output_dir, f"{base_name}_eric_reconcile_{timestamp}.xlsx")
                self.write_reconciliation_workbook(
                    output_file,
                    final_data,
                    ytic_data,
                    size_check_rows,
                    po_check_rows,
                    summary_rows,
                )

                print(f"[4/4] 生成 Eric 核对诊断包: {output_file}")
                print(f"  Size_Check 总项: {len(size_check_rows)}")
                print(f"  数量差异项: {difference_count}")
                print(f"  PO 差异项: {po_difference_count}")

            result.update({
                "success": True,
                "message": "Eric 最终核对诊断包生成完成",
                "logs": [line for line in buffer.getvalue().splitlines() if line.strip()],
                "output_path": output_file,
                "row_count": len(final_data["records"]),
                "difference_count": difference_count,
                "po_difference_count": po_difference_count,
                "size_check_count": len(size_check_rows),
            })
        except Exception as exc:
            result.update({
                "success": False,
                "message": f"Eric 最终核对失败：{exc}",
                "logs": [line for line in buffer.getvalue().splitlines() if line.strip()] + [str(exc)],
            })

        return result

    def process_file(self, input_file: str, output_dir: str) -> Dict[str, Any]:
        result: Dict[str, Any] = {
            "success": False,
            "message": "",
            "logs": [],
            "output_path": None,
            "row_count": 0,
        }

        ensure_dir(output_dir)
        base_name = os.path.splitext(os.path.basename(input_file))[0]
        timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
        output_prefix = os.path.join(output_dir, f"{base_name}_eric_{timestamp}")
        step1 = f"{output_prefix}_step1.xlsx"
        step2 = f"{output_prefix}_split.xlsx"
        final = f"{output_prefix}_final.xlsx"

        buffer = StringIO()
        try:
            with redirect_stdout(buffer):
                self.add_and_fill_columns(input_file, step1)
                self.process_split(step1, step2)
                self.process_unpivot(step2, final)

            row_count = 0
            wb = load_workbook(final, read_only=True, data_only=True)
            if "Final_Data" in wb.sheetnames:
                row_count = max(wb["Final_Data"].max_row - 1, 0)
            wb.close()

            result.update({
                "success": True,
                "message": "Eric Excel \u6570\u636e\u5904\u7406\u5b8c\u6210",
                "logs": [line for line in buffer.getvalue().splitlines() if line.strip()],
                "output_path": final,
                "row_count": row_count,
            })
        except Exception as exc:
            result.update({
                "success": False,
                "message": f"Eric Excel \u6570\u636e\u5904\u7406\u5931\u8d25\uff1a{exc}",
                "logs": [line for line in buffer.getvalue().splitlines() if line.strip()] + [str(exc)],
            })

        return result

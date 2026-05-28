# -*- coding: utf-8 -*-
"""
Jane OUTBOUND 核对模块。

以 T1 OUTBOUND 为输出底稿，对照 TMS Released Order 的 Result Set，
按 Style/PO/Line/Factory 匹配并标红数量、PODD 等差异。
"""

import copy
import os
from dataclasses import dataclass
from datetime import date, datetime
from typing import Any, Dict, List, Optional, Sequence, Set, Tuple

import openpyxl
from openpyxl.styles import PatternFill

from utils.file_utils import ensure_dir


@dataclass(frozen=True)
class TmsOutboundRow:
    """TMS Result Set 中可用于 OUTBOUND 核对的一行。"""

    row_index: int
    style: str
    po_number: str
    line_number: str
    factory: str
    working_number: str
    podd: Any
    ordered_quantity: Any


class JaneOutboundCompareModule:
    """Jane-OUTBOUND 核对业务逻辑。"""

    NO_FILL = PatternFill(fill_type=None)
    RED_FILL = PatternFill("solid", fgColor="FFFFC7CE")
    REQUIRED_OUTBOUND_COLUMNS = [
        "Style Number",
        "Invoice Date/Delivery Date",
        "Outbound Quantity (in Units)",
        "Units",
        "PO Number",
        "Line Number",
        "Recording Facility ID",
        "Other Reason",
    ]
    REQUIRED_TMS_COLUMNS = [
        "Factory",
        "PO Number",
        "PO Line Item #",
        "Working Number",
        "Article Number",
        "PODD",
        "Ordered Quantity",
    ]
    DIAGNOSTIC_HEADERS = [
        ("check result", "Check Result"),
        ("mismatch source", "Mismatch Source"),
        ("check detail", "Check Detail"),
        ("copy of tms source row", "Copy of TMS Source Row"),
    ]

    @staticmethod
    def _normalize_text(value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        if isinstance(value, datetime):
            return value.strftime("%Y-%m-%d")
        if isinstance(value, date):
            return value.strftime("%Y-%m-%d")
        text = str(value).strip()
        if text.lower() in {"none", "nan", "nat"}:
            return ""
        return text

    @classmethod
    def _normalize_key(cls, value: Any) -> str:
        return cls._normalize_text(value).upper()

    @classmethod
    def _normalize_number(cls, value: Any) -> Optional[float]:
        text = cls._normalize_text(value).replace(",", "")
        if not text:
            return None
        try:
            return float(text)
        except ValueError:
            return None

    @classmethod
    def _normalize_date(cls, value: Any) -> str:
        if isinstance(value, datetime):
            return value.date().isoformat()
        if isinstance(value, date):
            return value.isoformat()
        text = cls._normalize_text(value)
        if not text:
            return ""
        for fmt in ("%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%m/%d/%Y", "%d.%m.%Y"):
            try:
                return datetime.strptime(text, fmt).date().isoformat()
            except ValueError:
                continue
        return text

    @classmethod
    def _find_header_row(
        cls,
        ws,
        required_columns: Sequence[str],
        max_scan_rows: int = 20,
    ) -> Tuple[int, Dict[str, int]]:
        required = [column.lower() for column in required_columns]
        for row_index in range(1, min(ws.max_row, max_scan_rows) + 1):
            columns: Dict[str, int] = {}
            for column_index in range(1, ws.max_column + 1):
                value = cls._normalize_text(
                    ws.cell(row=row_index, column=column_index).value,
                ).lower()
                if value and value not in columns:
                    columns[value] = column_index
            if all(column in columns for column in required):
                return row_index, columns
        raise ValueError(f"未找到表头：{', '.join(required_columns)}")

    def _mark_red(self, cell) -> None:
        cell.fill = copy.copy(self.RED_FILL)

    def _clear_fill(self, cell) -> None:
        cell.fill = copy.copy(self.NO_FILL)

    def _clear_data_fills(self, ws, header_row: int, max_column: int) -> None:
        for row_index in range(header_row + 1, ws.max_row + 1):
            for column_index in range(1, max_column + 1):
                self._clear_fill(ws.cell(row=row_index, column=column_index))

    def _copy_row(
        self,
        ws,
        source_row: int,
        target_row: int,
        max_column: int,
        copy_values: bool = True,
    ) -> None:
        for column_index in range(1, max_column + 1):
            source = ws.cell(row=source_row, column=column_index)
            target = ws.cell(row=target_row, column=column_index)
            target.value = source.value if copy_values else None
            if source.has_style:
                target._style = copy.copy(source._style)
                if not copy_values:
                    self._clear_fill(target)
            if source.number_format:
                target.number_format = source.number_format
            if source.alignment:
                target.alignment = copy.copy(source.alignment)
            if source.protection:
                target.protection = copy.copy(source.protection)

    def _copy_header_style(self, ws, header_row: int, source_col: int, target_col: int) -> None:
        source = ws.cell(row=header_row, column=source_col)
        target = ws.cell(row=header_row, column=target_col)
        if source.has_style:
            target._style = copy.copy(source._style)
        if source.number_format:
            target.number_format = source.number_format
        if source.alignment:
            target.alignment = copy.copy(source.alignment)
        if source.protection:
            target.protection = copy.copy(source.protection)
        ws.column_dimensions[target.column_letter].width = ws.column_dimensions[source.column_letter].width

    def _insert_column_after(
        self,
        ws,
        header_row: int,
        columns: Dict[str, int],
        source_key: str,
        new_key: str,
        header: str,
    ) -> int:
        source_col = columns[source_key]
        insert_at = source_col + 1
        ws.insert_cols(insert_at)
        for column_name, column_index in list(columns.items()):
            if column_index >= insert_at:
                columns[column_name] = column_index + 1
        columns[new_key] = insert_at
        ws.cell(row=header_row, column=insert_at, value=header)
        self._copy_header_style(ws, header_row, source_col, insert_at)
        return insert_at

    def _append_column(
        self,
        ws,
        header_row: int,
        columns: Dict[str, int],
        new_key: str,
        header: str,
    ) -> int:
        append_at = ws.max_column + 1
        columns[new_key] = append_at
        ws.cell(row=header_row, column=append_at, value=header)
        self._copy_header_style(ws, header_row, max(1, append_at - 1), append_at)
        return append_at

    def _prepare_output_columns(self, ws, header_row: int, columns: Dict[str, int]) -> None:
        self._insert_column_after(
            ws,
            header_row,
            columns,
            "invoice date/delivery date",
            "correct podd",
            "Correct PODD",
        )
        self._insert_column_after(
            ws,
            header_row,
            columns,
            "outbound quantity (in units)",
            "correct ordered quantity",
            "Correct Ordered Quantity",
        )
        self._insert_column_after(
            ws,
            header_row,
            columns,
            "units",
            "correct units",
            "Correct Units",
        )
        self._insert_column_after(
            ws,
            header_row,
            columns,
            "other reason",
            "correct working number",
            "Correct Working Number",
        )
        for key, header in self.DIAGNOSTIC_HEADERS:
            self._append_column(ws, header_row, columns, key, header)

    def _set_row_diagnostic(
        self,
        ws,
        row_index: int,
        columns: Dict[str, int],
        result: str,
        source: str,
        detail: str,
        tms_row_index: Optional[int] = None,
    ) -> None:
        ws.cell(row=row_index, column=columns["check result"], value=result)
        ws.cell(row=row_index, column=columns["mismatch source"], value=source)
        detail_cell = ws.cell(row=row_index, column=columns["check detail"])
        existing_detail = self._normalize_text(detail_cell.value)
        detail_cell.value = f"{existing_detail}；{detail}" if existing_detail else detail
        if tms_row_index:
            ws.cell(row=row_index, column=columns["copy of tms source row"], value=tms_row_index)

    def _read_tms_rows(self, tms_path: str) -> List[TmsOutboundRow]:
        wb = openpyxl.load_workbook(tms_path, data_only=True)
        ws = wb["Result Set"] if "Result Set" in wb.sheetnames else wb.active
        header_row, columns = self._find_header_row(ws, self.REQUIRED_TMS_COLUMNS)

        rows: List[TmsOutboundRow] = []
        for row_index in range(header_row + 1, ws.max_row + 1):
            style = self._normalize_text(ws.cell(row=row_index, column=columns["article number"]).value)
            po_number = self._normalize_text(ws.cell(row=row_index, column=columns["po number"]).value)
            line_number = self._normalize_text(ws.cell(row=row_index, column=columns["po line item #"]).value)
            factory = self._normalize_text(ws.cell(row=row_index, column=columns["factory"]).value)
            working_number = self._normalize_text(ws.cell(row=row_index, column=columns["working number"]).value)
            if not any([style, po_number, line_number, factory, working_number]):
                continue
            rows.append(
                TmsOutboundRow(
                    row_index=row_index,
                    style=style,
                    po_number=po_number,
                    line_number=line_number,
                    factory=factory,
                    working_number=working_number,
                    podd=ws.cell(row=row_index, column=columns["podd"]).value,
                    ordered_quantity=ws.cell(row=row_index, column=columns["ordered quantity"]).value,
                ),
            )
        return rows

    def _row_key(self, style: Any, po_number: Any, line_number: Any, factory: Any) -> Tuple[str, str, str, str]:
        return (
            self._normalize_key(style),
            self._normalize_key(po_number),
            self._normalize_key(line_number),
            self._normalize_key(factory),
        )

    def _append_issue(
        self,
        issues: List[List[Any]],
        issue_type: str,
        field: str,
        key: Tuple[str, str, str, str],
        working_number: str,
        t1_value: Any,
        tms_value: Any,
        t1_row: Optional[int],
        tms_row: Optional[int],
        detail: str,
    ) -> None:
        issues.append([
            issue_type,
            field,
            key[0],
            key[1],
            working_number,
            key[2],
            key[3],
            self._normalize_text(t1_value),
            self._normalize_text(tms_value),
            t1_row or "",
            tms_row or "",
            detail,
        ])

    def _append_missing_outbound_row(
        self,
        ws,
        tms_row: TmsOutboundRow,
        template_row: int,
        columns: Dict[str, int],
        max_column: int,
    ) -> int:
        target_row = ws.max_row + 1
        self._copy_row(ws, template_row, target_row, max_column, copy_values=False)

        # 新增行只填 TMS 能提供且 T1 OUTBOUND 需要核对的字段。
        ws.cell(target_row, columns["style number"], tms_row.style)
        ws.cell(target_row, columns["invoice date/delivery date"], self._normalize_date(tms_row.podd))
        ws.cell(target_row, columns["outbound quantity (in units)"], tms_row.ordered_quantity)
        ws.cell(target_row, columns["units"], "Each")
        ws.cell(target_row, columns["po number"], tms_row.po_number)
        ws.cell(target_row, columns["line number"], tms_row.line_number)
        ws.cell(target_row, columns["recording facility id"], tms_row.factory)
        ws.cell(target_row, columns["other reason"], tms_row.working_number)

        for column_name in [
            "style number",
            "po number",
            "line number",
            "recording facility id",
            "other reason",
        ]:
            self._mark_red(ws.cell(row=target_row, column=columns[column_name]))
        self._set_row_diagnostic(
            ws,
            target_row,
            columns,
            "需补入",
            "Copy of TMS存在，T1 OUTBOUND缺失",
            "Copy of TMS 有该 Style/PO/Line/Factory，T1 OUTBOUND 缺少对应行。",
            tms_row.row_index,
        )

        return target_row

    def process_reports(
        self,
        outbound_path: str,
        tms_path: str,
        output_dir: str,
    ) -> Dict[str, Any]:
        logs: List[str] = []
        try:
            ensure_dir(output_dir)
            wb = openpyxl.load_workbook(outbound_path)
            ws = wb["OUTBOUND"] if "OUTBOUND" in wb.sheetnames else wb.active
            header_row, columns = self._find_header_row(ws, self.REQUIRED_OUTBOUND_COLUMNS)
            self._prepare_output_columns(ws, header_row, columns)
            max_column = ws.max_column
            self._clear_data_fills(ws, header_row, max_column)

            tms_rows = self._read_tms_rows(tms_path)
            tms_by_key = {
                self._row_key(row.style, row.po_number, row.line_number, row.factory): row
                for row in tms_rows
            }

            style_col = columns["style number"]
            date_col = columns["invoice date/delivery date"]
            quantity_col = columns["outbound quantity (in units)"]
            units_col = columns["units"]
            po_col = columns["po number"]
            line_col = columns["line number"]
            factory_col = columns["recording facility id"]
            working_col = columns["other reason"]

            outbound_keys: Set[Tuple[str, str, str, str]] = set()
            working_numbers: Set[str] = set()
            issues: List[List[Any]] = []
            checked_row_count = 0
            matched_row_count = 0
            missing_tms_row_count = 0
            difference_cell_count = 0
            first_data_row = header_row + 1

            for row_index in range(first_data_row, ws.max_row + 1):
                key = self._row_key(
                    ws.cell(row=row_index, column=style_col).value,
                    ws.cell(row=row_index, column=po_col).value,
                    ws.cell(row=row_index, column=line_col).value,
                    ws.cell(row=row_index, column=factory_col).value,
                )
                working_number = self._normalize_key(ws.cell(row=row_index, column=working_col).value)
                if not any(key) and not working_number:
                    continue

                checked_row_count += 1
                outbound_keys.add(key)
                if working_number:
                    working_numbers.add(working_number)

                tms_row = tms_by_key.get(key)
                if not tms_row:
                    for column_index in [style_col, po_col, line_col, factory_col]:
                        self._mark_red(ws.cell(row=row_index, column=column_index))
                    self._set_row_diagnostic(
                        ws,
                        row_index,
                        columns,
                        "需核对",
                        "T1 OUTBOUND存在，Copy of TMS缺失",
                        "T1 OUTBOUND 存在，但 Copy of TMS 找不到相同 Style/PO/Line/Factory。",
                    )
                    missing_tms_row_count += 1
                    self._append_issue(
                        issues,
                        "TMS 缺失",
                        "匹配键",
                        key,
                        working_number,
                        "",
                        "",
                        row_index,
                        None,
                        "T1 OUTBOUND 存在，但 TMS Result Set 找不到相同 Style/PO/Line/Factory。",
                    )
                    continue

                matched_row_count += 1
                expected_date = self._normalize_date(tms_row.podd)
                actual_date = self._normalize_date(ws.cell(row=row_index, column=date_col).value)
                if actual_date != expected_date:
                    self._mark_red(ws.cell(row=row_index, column=date_col))
                    ws.cell(row=row_index, column=columns["correct podd"], value=expected_date)
                    self._set_row_diagnostic(
                        ws,
                        row_index,
                        columns,
                        "需核对",
                        "值不一致：以 Copy of TMS 为准",
                        "Invoice Date/Delivery Date 与 Copy of TMS PODD 不一致。",
                        tms_row.row_index,
                    )
                    difference_cell_count += 1
                    self._append_issue(
                        issues,
                        "值不一致",
                        "Invoice Date/Delivery Date",
                        key,
                        working_number or self._normalize_key(tms_row.working_number),
                        actual_date,
                        expected_date,
                        row_index,
                        tms_row.row_index,
                        "T1 日期与 TMS PODD 不一致。",
                    )

                expected_quantity = self._normalize_number(tms_row.ordered_quantity)
                actual_quantity = self._normalize_number(ws.cell(row=row_index, column=quantity_col).value)
                if actual_quantity != expected_quantity:
                    self._mark_red(ws.cell(row=row_index, column=quantity_col))
                    ws.cell(
                        row=row_index,
                        column=columns["correct ordered quantity"],
                        value=tms_row.ordered_quantity,
                    )
                    self._set_row_diagnostic(
                        ws,
                        row_index,
                        columns,
                        "需核对",
                        "值不一致：以 Copy of TMS 为准",
                        "Outbound Quantity 与 Copy of TMS Ordered Quantity 不一致。",
                        tms_row.row_index,
                    )
                    difference_cell_count += 1
                    self._append_issue(
                        issues,
                        "值不一致",
                        "Outbound Quantity (in Units)",
                        key,
                        working_number or self._normalize_key(tms_row.working_number),
                        ws.cell(row=row_index, column=quantity_col).value,
                        tms_row.ordered_quantity,
                        row_index,
                        tms_row.row_index,
                        "T1 出库数量与 TMS Ordered Quantity 不一致。",
                    )

                if self._normalize_key(ws.cell(row=row_index, column=units_col).value) != "EACH":
                    self._mark_red(ws.cell(row=row_index, column=units_col))
                    ws.cell(row=row_index, column=columns["correct units"], value="Each")
                    self._set_row_diagnostic(
                        ws,
                        row_index,
                        columns,
                        "需核对",
                        "值不一致：以规则为准",
                        "T1 OUTBOUND Units 应为 Each。",
                        tms_row.row_index,
                    )
                    difference_cell_count += 1
                    self._append_issue(
                        issues,
                        "值不一致",
                        "Units",
                        key,
                        working_number or self._normalize_key(tms_row.working_number),
                        ws.cell(row=row_index, column=units_col).value,
                        "Each",
                        row_index,
                        tms_row.row_index,
                        "T1 Units 应为 Each。",
                    )

                expected_working = self._normalize_key(tms_row.working_number)
                if working_number and expected_working and working_number != expected_working:
                    self._mark_red(ws.cell(row=row_index, column=working_col))
                    ws.cell(
                        row=row_index,
                        column=columns["correct working number"],
                        value=tms_row.working_number,
                    )
                    self._set_row_diagnostic(
                        ws,
                        row_index,
                        columns,
                        "需核对",
                        "值不一致：以 Copy of TMS 为准",
                        "Other Reason 与 Copy of TMS Working Number 不一致。",
                        tms_row.row_index,
                    )
                    difference_cell_count += 1
                    self._append_issue(
                        issues,
                        "值不一致",
                        "Other Reason",
                        key,
                        working_number,
                        ws.cell(row=row_index, column=working_col).value,
                        tms_row.working_number,
                        row_index,
                        tms_row.row_index,
                        "T1 Other Reason 与 TMS Working Number 不一致。",
                    )

            missing_outbound_row_count = 0
            template_row = first_data_row if ws.max_row >= first_data_row else header_row
            for tms_row in tms_rows:
                tms_working = self._normalize_key(tms_row.working_number)
                if tms_working not in working_numbers:
                    continue
                key = self._row_key(tms_row.style, tms_row.po_number, tms_row.line_number, tms_row.factory)
                if key in outbound_keys:
                    continue
                appended_row = self._append_missing_outbound_row(
                    ws,
                    tms_row,
                    template_row,
                    columns,
                    max_column,
                )
                outbound_keys.add(key)
                missing_outbound_row_count += 1
                self._append_issue(
                    issues,
                    "T1 OUTBOUND 缺失",
                    "整行",
                    key,
                    tms_working,
                    "",
                    "TMS 行存在",
                    appended_row,
                    tms_row.row_index,
                    "TMS Result Set 存在相关 Working Number 行，但 T1 OUTBOUND 缺少该 Style/PO/Line/Factory。",
                )

            if "OUTBOUND_Check" in wb.sheetnames:
                del wb["OUTBOUND_Check"]
            timestamp = datetime.now().strftime("%Y%m%d_%H%M%S")
            output_path = os.path.join(output_dir, f"jane_outbound_compare_{timestamp}.xlsx")
            wb.save(output_path)

            logs.append(f"TMS Result Set 读取完成：{len(tms_rows)} 行")
            logs.append(f"T1 OUTBOUND 核对完成：{checked_row_count} 行，差异 {len(issues)} 条")
            return {
                "success": True,
                "message": f"Jane-OUTBOUND 核对完成，结果文件：{output_path}",
                "logs": logs,
                "checked_row_count": checked_row_count,
                "matched_row_count": matched_row_count,
                "missing_tms_row_count": missing_tms_row_count,
                "missing_outbound_row_count": missing_outbound_row_count,
                "difference_cell_count": difference_cell_count,
                "issue_count": len(issues),
                "output_path": output_path,
            }
        except Exception as exc:
            logs.append(f"ERROR: {exc}")
            return {
                "success": False,
                "message": str(exc),
                "logs": logs,
            }

# -*- coding: utf-8 -*-
"""Generic Excel template mapper module."""

from __future__ import annotations

import os
from copy import copy
from dataclasses import dataclass
from datetime import date, datetime
from pathlib import Path
from typing import Any, Sequence
from uuid import uuid4

import openpyxl
import xlrd
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class ExcelTemplateFieldMapping:
    target_column: int
    source_column: int
    target_header: str = ""
    source_header: str = ""
    required: bool = True


@dataclass(frozen=True)
class ExcelTemplateMapperConfig:
    source_sheet_name: str
    template_sheet_name: str
    source_header_row: int
    source_data_start_row: int
    template_header_row: int
    template_data_start_row: int
    mappings: list[ExcelTemplateFieldMapping]


@dataclass(frozen=True)
class SheetData:
    name: str
    rows: list[list[Any]]
    max_row: int
    max_column: int


class ExcelTemplateMapperModule:
    """Map rows from one uploaded workbook into a user-provided template workbook."""

    SAMPLE_ROW_LIMIT = 20

    def inspect_workbook(
        self,
        workbook_path: str | os.PathLike[str],
        *,
        sheet_name: str | None = None,
        header_row: int = 1,
    ) -> dict[str, Any]:
        sheets = self._read_workbook(workbook_path)
        if not sheets:
            raise ValueError("工作簿中没有可读取的 sheet")

        selected = self._select_sheet(sheets, sheet_name)
        self._validate_row_index(header_row, selected.max_row, "表头行")
        header_values = selected.rows[header_row - 1] if header_row <= len(selected.rows) else []
        sample_values = selected.rows[header_row] if header_row < len(selected.rows) else []

        return {
            "sheets": [
                {
                    "name": sheet.name,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                }
                for sheet in sheets
            ],
            "selected_sheet": {
                "name": selected.name,
                "header_row": header_row,
                "max_row": selected.max_row,
                "max_column": selected.max_column,
                "data_row_count": max(selected.max_row - header_row, 0),
                "headers": [
                    {
                        "index": index,
                        "letter": get_column_letter(index),
                        "label": self._display_text(self._read_cell(header_values, index)),
                        "sample_value": self._display_text(self._read_cell(sample_values, index)),
                    }
                    for index in range(1, selected.max_column + 1)
                ],
                "sample_rows": self._build_sample_rows(selected, header_row + 1),
            },
        }

    def process_files(
        self,
        *,
        source_path: str | os.PathLike[str],
        template_path: str | os.PathLike[str],
        config: ExcelTemplateMapperConfig,
        output_dir: str | os.PathLike[str],
    ) -> dict[str, Any]:
        self._validate_config(config)
        source_sheet = self._select_sheet(self._read_workbook(source_path), config.source_sheet_name)
        self._validate_row_index(config.source_header_row, source_sheet.max_row, "源文件表头行")
        self._validate_row_index(config.source_data_start_row, source_sheet.max_row, "源文件数据起始行")

        unmapped_required_fields = self._unmapped_required_fields(config.mappings)
        if unmapped_required_fields:
            raise ValueError(f"必填模板字段未匹配：{', '.join(unmapped_required_fields)}")

        template_workbook = openpyxl.load_workbook(template_path)
        try:
            if config.template_sheet_name not in template_workbook.sheetnames:
                raise ValueError(f"未找到模板 sheet：{config.template_sheet_name}")
            template_sheet = template_workbook[config.template_sheet_name]
            self._validate_row_index(config.template_header_row, template_sheet.max_row, "模板表头行")
            self._validate_template_start_row(config.template_data_start_row)

            source_rows = self._source_data_rows(source_sheet, config.source_data_start_row)
            self._write_mapped_rows(template_sheet, source_rows, config)

            output_path = Path(output_dir) / f"excel_template_mapper_{uuid4().hex}.xlsx"
            template_workbook.save(output_path)
        finally:
            template_workbook.close()

        return {
            "success": True,
            "message": f"通用 Excel 映射完成，结果文件：{output_path}",
            "output_path": str(output_path),
            "source_row_count": len(source_rows),
            "written_row_count": len(source_rows),
            "mapped_field_count": len([mapping for mapping in config.mappings if mapping.source_column > 0]),
            "unmapped_required_fields": [],
        }

    def _read_workbook(self, workbook_path: str | os.PathLike[str]) -> list[SheetData]:
        path = Path(workbook_path)
        suffix = path.suffix.lower()
        if suffix == ".xls":
            return self._read_xls_workbook(path)
        if suffix in {".xlsx", ".xlsm"}:
            return self._read_openpyxl_workbook(path)
        raise ValueError("仅支持 .xls / .xlsx / .xlsm 文件")

    def _read_openpyxl_workbook(self, path: Path) -> list[SheetData]:
        workbook = openpyxl.load_workbook(path, data_only=True, read_only=True)
        try:
            sheets: list[SheetData] = []
            for worksheet in workbook.worksheets:
                max_row = worksheet.max_row or 0
                max_column = worksheet.max_column or 0
                rows = [
                    list(row)
                    for row in worksheet.iter_rows(
                        min_row=1,
                        max_row=max_row,
                        max_col=max_column,
                        values_only=True,
                    )
                ]
                sheets.append(SheetData(worksheet.title, rows, max_row, max_column))
            return sheets
        finally:
            workbook.close()

    def _read_xls_workbook(self, path: Path) -> list[SheetData]:
        book = xlrd.open_workbook(str(path), on_demand=True)
        try:
            sheets: list[SheetData] = []
            for sheet in book.sheets():
                rows = [
                    [
                        self._read_xls_cell(sheet, row_index, column_index, book.datemode)
                        for column_index in range(sheet.ncols)
                    ]
                    for row_index in range(sheet.nrows)
                ]
                sheets.append(SheetData(sheet.name, rows, sheet.nrows, sheet.ncols))
            return sheets
        finally:
            book.release_resources()

    def _read_xls_cell(self, sheet: xlrd.sheet.Sheet, row_index: int, column_index: int, datemode: int) -> Any:
        value = sheet.cell_value(row_index, column_index)
        if sheet.cell_type(row_index, column_index) == xlrd.XL_CELL_DATE:
            return xlrd.xldate.xldate_as_datetime(value, datemode)
        return value

    def _select_sheet(self, sheets: Sequence[SheetData], sheet_name: str | None) -> SheetData:
        if sheet_name:
            for sheet in sheets:
                if sheet.name == sheet_name:
                    return sheet
            raise ValueError(f"未找到 sheet：{sheet_name}")
        return sheets[0]

    def _validate_row_index(self, row_index: int, max_row: int, label: str) -> None:
        if row_index < 1 or row_index > max(max_row, 1):
            raise ValueError(f"{label}无效")

    def _validate_template_start_row(self, row_index: int) -> None:
        if row_index < 1:
            raise ValueError("模板写入起始行无效")

    def _validate_config(self, config: ExcelTemplateMapperConfig) -> None:
        if not config.source_sheet_name.strip():
            raise ValueError("缺少源文件 sheet")
        if not config.template_sheet_name.strip():
            raise ValueError("缺少模板 sheet")
        if config.source_header_row < 1:
            raise ValueError("源文件表头行无效")
        if config.source_data_start_row < 1:
            raise ValueError("源文件数据起始行无效")
        if config.template_header_row < 1:
            raise ValueError("模板表头行无效")
        if config.template_data_start_row < 1:
            raise ValueError("模板写入起始行无效")
        if not config.mappings:
            raise ValueError("缺少字段映射")
        for mapping in config.mappings:
            if mapping.target_column < 1:
                raise ValueError("模板目标列无效")

    def _unmapped_required_fields(self, mappings: Sequence[ExcelTemplateFieldMapping]) -> list[str]:
        fields: list[str] = []
        for mapping in mappings:
            if mapping.required and mapping.source_column < 1:
                fields.append(mapping.target_header or get_column_letter(mapping.target_column))
        return fields

    def _source_data_rows(self, sheet: SheetData, data_start_row: int) -> list[list[Any]]:
        rows: list[list[Any]] = []
        for row in sheet.rows[data_start_row - 1:]:
            if any(value not in (None, "") for value in row):
                rows.append(row)
        return rows

    def _write_mapped_rows(
        self,
        template_sheet: Any,
        source_rows: Sequence[Sequence[Any]],
        config: ExcelTemplateMapperConfig,
    ) -> None:
        for row_offset, source_row in enumerate(source_rows):
            target_row = config.template_data_start_row + row_offset
            self._copy_template_row_style(template_sheet, config.template_data_start_row, target_row)
            for mapping in config.mappings:
                if mapping.source_column < 1:
                    continue
                template_sheet.cell(row=target_row, column=mapping.target_column).value = self._read_cell(
                    source_row,
                    mapping.source_column,
                )

    def _copy_template_row_style(self, worksheet: Any, template_row: int, target_row: int) -> None:
        if target_row == template_row:
            return
        for column_index in range(1, worksheet.max_column + 1):
            source_cell = worksheet.cell(row=template_row, column=column_index)
            target_cell = worksheet.cell(row=target_row, column=column_index)
            if source_cell.has_style:
                target_cell._style = copy(source_cell._style)
            if source_cell.number_format:
                target_cell.number_format = source_cell.number_format

    def _build_sample_rows(self, sheet: SheetData, start_row: int) -> list[dict[str, Any]]:
        sample_rows: list[dict[str, Any]] = []
        end_row = min(sheet.max_row, start_row + self.SAMPLE_ROW_LIMIT - 1)
        for row_number in range(start_row, end_row + 1):
            row = sheet.rows[row_number - 1] if row_number <= len(sheet.rows) else []
            sample_rows.append({
                "row_number": row_number,
                "values": [
                    self._display_text(self._read_cell(row, column_index))
                    for column_index in range(1, sheet.max_column + 1)
                ],
            })
        return sample_rows

    def _read_cell(self, row_values: Sequence[Any], column_index: int) -> Any:
        if column_index < 1 or column_index > len(row_values):
            return None
        return row_values[column_index - 1]

    def _display_text(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, datetime):
            return value.isoformat(sep=" ")
        if isinstance(value, date):
            return value.isoformat()
        return str(value)

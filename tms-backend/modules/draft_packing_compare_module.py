# -*- coding: utf-8 -*-
"""产地证与 Packing List PDF 字段提取核对模块。"""

from __future__ import annotations

import os
import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any, Iterable, Sequence
from uuid import uuid4

import openpyxl
import pdfplumber
from openpyxl.styles import Alignment, Font, PatternFill
from openpyxl.utils import get_column_letter

from modules.origin_certificate_ocr import (
    OriginCertificateOcrUnavailableError,
    OriginCertificateOcrParser,
    OriginCertificateRecord,
    RapidOriginCertificateOcr,
)


@dataclass(frozen=True)
class PackingExtractedPage:
    text: str
    tables: list[list[list[str | None]]]


@dataclass
class ExtractedRecord:
    source: str
    po_number: str
    invoice_number: str = ""
    working_number: str = ""
    article_number: str = ""
    customer_number: str = ""
    quantity: int | None = None
    cartons: int | None = None
    cartons_in_words: str = ""
    goods_description: str = ""
    hs_code: str = ""
    issues: list[str] = field(default_factory=list)

    def match_key(self) -> tuple[str, str, str]:
        return (
            self.po_number,
            self.working_number.upper(),
            self.article_number.upper(),
        )


@dataclass(frozen=True)
class ComparisonIssue:
    po_number: str
    key: tuple[str, str, str]
    source: str
    field_name: str
    issue_type: str
    detail: str


@dataclass
class ComparisonGroup:
    key: tuple[str, str, str]
    draft: ExtractedRecord | None
    packing: ExtractedRecord | None
    status: str
    issues: list[ComparisonIssue]

    @property
    def po_number(self) -> str:
        return self.key[0]

    @property
    def issue_detail(self) -> str:
        return "；".join(issue.detail for issue in self.issues)


@dataclass
class ComparisonResult:
    groups: list[ComparisonGroup]
    issues: list[ComparisonIssue]
    group_count: int
    issue_count: int
    mismatch_count: int
    missing_field_count: int


@dataclass(frozen=True)
class InvoiceComparisonSheet:
    invoice_number: str
    sheet_name: str
    comparison: ComparisonResult


class DraftPackingCompareModule:
    """解析两份 PDF 并生成产地证/Packing 上下对比 Excel。"""

    ORIGIN_CERTIFICATE_SOURCE = "产地证"
    OUTPUT_SHEET = "产地证 vs Packing"
    ISSUES_SHEET = "Issues"
    UNMATCHED_SHEET = "未匹配"
    HEADERS = [
        "PO Number",
        "Source",
        "INV number",
        "Working Number / Style Number",
        "Article Number",
        "Cust Number / Market PO Number",
        "Quantity",
        "Cartons",
        "Cartons In Words",
        "Goods Description",
        "HS Code / HTS Code",
        "Check Status",
        "Issue Detail",
    ]
    ISSUE_HEADERS = [
        "PO Number",
        "Match Key",
        "Source",
        "Field",
        "Issue Type",
        "Detail",
    ]
    BATCH_ISSUE_HEADERS = ["INV number", "Sheet Name", *ISSUE_HEADERS]
    FIELD_COLUMNS = {
        "Working Number / Style Number": 4,
        "Article Number": 5,
        "Cust Number / Market PO Number": 6,
        "Quantity": 7,
        "Cartons": 8,
        "Goods Description": 10,
        "HS Code / HTS Code": 11,
        "Record": 1,
    }
    NO_CONTENT_FILL_COLUMNS = {FIELD_COLUMNS["HS Code / HTS Code"]}
    ITEM_HEADER_RE = re.compile(
        r"(?m)^(?P<item>\d+)\s+(?:CART\s+NO\.\s*)?(?P<cartons_words>[A-Z][A-Z\s-]+?)\s*"
        r"\((?P<cartons>\d+)\)\s+CARTONS\s+OF\s+ADIDAS\s+BRAND\s+GARMENT\b.*$",
        re.IGNORECASE,
    )
    PO_RE = re.compile(r"PO#\s*(?P<po>\d{10})", re.IGNORECASE)
    PACKING_BLOCK_RE = re.compile(
        r"(?ms)^(?P<po>\d{10})\s+1\b.*?Goods Description\s*\n(?P<description>.*?)(?=\nPO No PO Line|\nTotal Quantity|\Z)"
    )
    DETAIL_ROW_RE = re.compile(
        r"^(?P<range>\d{4})\s+\d+\s+\d+\s+\d+\s+(?P<po>\d{10})\s+\d+\s+"
        r"(?P<article>[A-Z0-9]+)\s+.*?\s+(?P<qty>\d+)\s+(?P<ctn>\d+)\s+"
        r"(?:R\s+)?\d",
        re.IGNORECASE,
    )
    PACKING_DESCRIPTION_FOOTER_RE = re.compile(
        r"\b(This document is a summary|The complete document may be accessed on the system|Page\s+\d+\s+of\s+\d+)\b",
        re.IGNORECASE,
    )
    PACKING_DESCRIPTION_HTS_RE = re.compile(r"\bHTS\s*:\s*(?P<hts>[0-9][0-9.]{3,})", re.IGNORECASE)
    RED_FILL = PatternFill("solid", fgColor="FFFFC7CE")
    YELLOW_FILL = PatternFill("solid", fgColor="FFFFF2CC")
    HEADER_FILL = PatternFill("solid", fgColor="FFD9EAF7")
    SEPARATOR_FILL = PatternFill("solid", fgColor="FFEFF6FF")

    def __init__(self, origin_certificate_ocr: RapidOriginCertificateOcr | None = None) -> None:
        self.origin_certificate_ocr = origin_certificate_ocr or RapidOriginCertificateOcr()
        self.origin_certificate_text_parser = OriginCertificateOcrParser()

    def parse_origin_certificate_pdf(self, pdf_path: str | os.PathLike[str]) -> list[ExtractedRecord]:
        text = self._extract_pdf_text(pdf_path)
        records = self.parse_draft_text(text)
        text_records = [
            self._record_from_origin_certificate(record)
            for record in self.origin_certificate_text_parser.parse_text(text)
        ]
        if text_records and self._records_completeness_score(text_records) > self._records_completeness_score(records):
            return text_records
        if records:
            return records

        try:
            origin_records = self.origin_certificate_ocr.extract_records(pdf_path)
        except OriginCertificateOcrUnavailableError as exc:
            raise ValueError(str(exc)) from exc
        return [self._record_from_origin_certificate(record) for record in origin_records]

    def parse_draft_pdf(self, pdf_path: str | os.PathLike[str]) -> list[ExtractedRecord]:
        return self.parse_origin_certificate_pdf(pdf_path)

    def parse_packing_pdf(self, pdf_path: str | os.PathLike[str]) -> list[ExtractedRecord]:
        return self.parse_packing_pages(self._extract_packing_pages(pdf_path))

    def _extract_pdf_text(self, pdf_path: str | os.PathLike[str]) -> str:
        with pdfplumber.open(pdf_path) as pdf:
            return "\n".join(page.extract_text(x_tolerance=1, y_tolerance=3) or "" for page in pdf.pages)

    def _record_from_origin_certificate(self, record: OriginCertificateRecord) -> ExtractedRecord:
        extracted = ExtractedRecord(
            source=self.ORIGIN_CERTIFICATE_SOURCE,
            po_number=record.po_number,
            working_number=record.working_number,
            article_number=record.article_number,
            customer_number=record.customer_number,
            quantity=record.quantity,
            cartons=record.cartons,
            cartons_in_words=record.cartons_in_words,
            goods_description=record.goods_description,
            hs_code=record.hs_code,
        )
        if not extracted.po_number:
            extracted.issues.append(f"{self.ORIGIN_CERTIFICATE_SOURCE} PO Number 字段缺失或定位困难")
        self._add_missing_record_issues(extracted, self.ORIGIN_CERTIFICATE_SOURCE)
        return extracted

    def _records_completeness_score(self, records: Sequence[ExtractedRecord]) -> int:
        score = 0
        for record in records:
            fields = (
                record.po_number,
                record.working_number,
                record.article_number,
                record.customer_number,
                record.quantity,
                record.cartons,
                record.goods_description,
                record.hs_code,
            )
            score += sum(1 for value in fields if value not in (None, ""))
        return score

    def parse_draft_text(self, text: str) -> list[ExtractedRecord]:
        item_matches = list(self.ITEM_HEADER_RE.finditer(text))
        po_matches = list(self.PO_RE.finditer(text))
        records: list[ExtractedRecord] = []

        for index, po_match in enumerate(po_matches):
            item_match = self._find_previous_item_header(item_matches, po_match.start())
            next_item_start = self._find_next_item_start(item_matches, po_match.start())
            next_po_start = po_matches[index + 1].start() if index + 1 < len(po_matches) else len(text)
            block_end = min(next_item_start, next_po_start)
            po_block = text[po_match.start() : block_end]
            context_text = text[item_match.end() : po_match.start()] if item_match else ""

            cartons = self._parse_int(item_match.group("cartons")) if item_match else None
            record = ExtractedRecord(
                source=self.ORIGIN_CERTIFICATE_SOURCE,
                po_number=po_match.group("po"),
                working_number=self._find_labeled_value(po_block, r"STYLE\s+NO\.?"),
                article_number=self._find_labeled_value(po_block, r"ARTICLE\s+NO\.?"),
                customer_number=self._find_labeled_value(po_block, r"CUST\s+NO\.?"),
                quantity=self._parse_int(self._find_labeled_value(po_block, r"QUANTITY")),
                cartons=cartons,
                cartons_in_words=self._normalize_text(item_match.group("cartons_words")) if item_match else "",
                goods_description=self._extract_draft_description(context_text),
                hs_code=self._find_labeled_value(po_block, r"HS\s+Code"),
            )
            self._add_missing_record_issues(record, self.ORIGIN_CERTIFICATE_SOURCE)
            records.append(record)

        return records

    def parse_packing_pages(self, pages: Sequence[PackingExtractedPage]) -> list[ExtractedRecord]:
        records: list[ExtractedRecord] = []
        by_key: dict[tuple[str, str, str], ExtractedRecord] = {}
        page_texts = [page.text for page in pages]
        detail_totals = self._parse_packing_detail_totals(page_texts)
        invoice_number = self._extract_packing_invoice_number(page_texts)

        for page in pages:
            last_record: ExtractedRecord | None = None
            for table in page.tables:
                if not table:
                    continue

                header = table[0]
                normalized_header = [self._normalize_header(cell) for cell in header]
                if self._is_packing_summary_header(normalized_header):
                    for row in table[1:]:
                        record = self._record_from_packing_summary_row(normalized_header, row)
                        if not record:
                            continue
                        record.invoice_number = invoice_number
                        detail_total = detail_totals.get(record.po_number)
                        if detail_total:
                            if record.quantity is None:
                                record.quantity = detail_total["quantity"]
                            if record.cartons is None:
                                record.cartons = detail_total["cartons"]
                        records.append(record)
                        by_key[record.match_key()] = record
                        last_record = record
                    continue

                if last_record and "hts" in normalized_header:
                    self._attach_hts_row(last_record, normalized_header, table[1:])

            self._attach_packing_descriptions(page.text, records)

        for record in records:
            self._add_missing_record_issues(record, "Packing List")

        return list(by_key.values()) if by_key else records

    def build_comparison_result(
        self,
        draft_records: Sequence[ExtractedRecord],
        packing_records: Sequence[ExtractedRecord],
    ) -> ComparisonResult:
        draft_keys = {
            id(record): self._comparison_key(record, packing_records)
            for record in draft_records
        }
        packing_keys = {
            id(record): self._comparison_key(record, draft_records)
            for record in packing_records
        }
        draft_record_ids = {id(record) for record in draft_records}
        draft_by_key = {draft_keys[id(record)]: record for record in draft_records}
        packing_by_key = {packing_keys[id(record)]: record for record in packing_records}
        ordered_keys: list[tuple[str, str, str]] = []
        for record in list(draft_records) + list(packing_records):
            key = draft_keys[id(record)] if id(record) in draft_record_ids else packing_keys[id(record)]
            if key not in ordered_keys:
                ordered_keys.append(key)

        groups: list[ComparisonGroup] = []
        all_issues: list[ComparisonIssue] = []
        for key in ordered_keys:
            draft = draft_by_key.get(key)
            packing = packing_by_key.get(key)
            issues: list[ComparisonIssue] = []

            if draft is None:
                issues.append(
                    ComparisonIssue(
                        po_number=key[0],
                        key=key,
                        source=self.ORIGIN_CERTIFICATE_SOURCE,
                        field_name="Record",
                        issue_type="字段缺失或定位困难",
                        detail="产地证 未找到对应记录",
                    )
                )
            if packing is None:
                issues.append(
                    ComparisonIssue(
                        po_number=key[0],
                        key=key,
                        source="Packing List",
                        field_name="Record",
                        issue_type="字段缺失或定位困难",
                        detail="Packing List 未找到对应记录",
                    )
                )

            for record in (draft, packing):
                if record:
                    issues.extend(self._issues_from_record(record, key))

            if draft and packing:
                issues.extend(self._compare_records(draft, packing, key))

            status = self._status_from_issues(issues)
            group = ComparisonGroup(
                key=key,
                draft=draft,
                packing=packing,
                status=status,
                issues=issues,
            )
            groups.append(group)
            all_issues.extend(issues)

        mismatch_count = sum(1 for issue in all_issues if issue.issue_type == "值不一致")
        missing_field_count = sum(1 for issue in all_issues if issue.issue_type == "字段缺失或定位困难")
        return ComparisonResult(
            groups=groups,
            issues=all_issues,
            group_count=len(groups),
            issue_count=len(all_issues),
            mismatch_count=mismatch_count,
            missing_field_count=missing_field_count,
        )

    def _comparison_key(
        self,
        record: ExtractedRecord,
        counterpart_records: Sequence[ExtractedRecord],
    ) -> tuple[str, str, str]:
        strict_key = record.match_key()
        if strict_key[1] or not strict_key[0] or not strict_key[2]:
            return strict_key

        partial_matches = {
            counterpart.match_key()
            for counterpart in counterpart_records
            if counterpart.po_number == record.po_number
            and counterpart.article_number.upper() == record.article_number.upper()
        }
        if len(partial_matches) == 1:
            return next(iter(partial_matches))
        return strict_key

    def process_files(
        self,
        draft_pdf_path: str | os.PathLike[str],
        packing_pdf_path: str | os.PathLike[str],
        output_dir: str | os.PathLike[str],
    ) -> dict[str, Any]:
        draft_records = self.parse_origin_certificate_pdf(draft_pdf_path)
        packing_records = self.parse_packing_pdf(packing_pdf_path)
        return self.process_extracted_data(draft_records, packing_records, output_dir)

    def process_file_batches(
        self,
        draft_pdf_paths: Sequence[str | os.PathLike[str]],
        packing_pdf_paths: Sequence[str | os.PathLike[str]],
        output_dir: str | os.PathLike[str],
    ) -> dict[str, Any]:
        draft_records: list[ExtractedRecord] = []
        packing_records: list[ExtractedRecord] = []

        for packing_pdf_path in packing_pdf_paths:
            packing_records.extend(self.parse_packing_pdf(packing_pdf_path))
        for draft_pdf_path in draft_pdf_paths:
            draft_records.extend(self.parse_origin_certificate_pdf(draft_pdf_path))

        self._assign_origin_invoice_numbers(draft_records, packing_records)
        invoice_groups = self._group_records_by_invoice(draft_records, packing_records)

        output_root = Path(output_dir)
        output_root.mkdir(parents=True, exist_ok=True)
        output_path = output_root / f"draft_packing_compare_{uuid4().hex}.xlsx"

        used_sheet_names: set[str] = set()
        sheets: list[InvoiceComparisonSheet] = []
        for invoice_number, grouped_records in invoice_groups:
            sheet_name = self._make_unique_sheet_title(invoice_number, used_sheet_names)
            sheets.append(
                InvoiceComparisonSheet(
                    invoice_number=invoice_number,
                    sheet_name=sheet_name,
                    comparison=self.build_comparison_result(
                        grouped_records["draft"],
                        grouped_records["packing"],
                    ),
                )
            )

        self._write_batch_workbook(sheets, output_path)

        group_count = sum(sheet.comparison.group_count for sheet in sheets)
        issue_count = sum(sheet.comparison.issue_count for sheet in sheets)
        mismatch_count = sum(sheet.comparison.mismatch_count for sheet in sheets)
        missing_field_count = sum(
            sheet.comparison.missing_field_count for sheet in sheets
        )

        return {
            "success": True,
            "message": f"产地证与 Packing List 核对完成，生成 {group_count} 组对比记录，{len(sheets)} 个结果 Sheet",
            "output_path": str(output_path),
            "group_count": group_count,
            "issue_count": issue_count,
            "mismatch_count": mismatch_count,
            "missing_field_count": missing_field_count,
            "draft_count": len(draft_records),
            "packing_count": len(packing_records),
            "sheet_count": len(sheets),
            "draft_file_count": len(draft_pdf_paths),
            "packing_file_count": len(packing_pdf_paths),
            "logs": [
                f"产地证文件数：{len(draft_pdf_paths)}",
                f"Packing List 文件数：{len(packing_pdf_paths)}",
                f"产地证识别记录数：{len(draft_records)}",
                f"Packing List 识别记录数：{len(packing_records)}",
                f"结果 Sheet 数：{len(sheets)}",
                f"核对分组数：{group_count}",
                f"问题数：{issue_count}",
                f"输出文件：{output_path}",
            ],
        }

    def process_extracted_data(
        self,
        draft_records: Sequence[ExtractedRecord],
        packing_records: Sequence[ExtractedRecord],
        output_dir: str | os.PathLike[str],
    ) -> dict[str, Any]:
        comparison = self.build_comparison_result(draft_records, packing_records)
        output_root = Path(output_dir)
        output_root.mkdir(parents=True, exist_ok=True)
        output_path = output_root / f"draft_packing_compare_{uuid4().hex}.xlsx"
        self._write_workbook(comparison, output_path)

        return {
            "success": True,
            "message": f"产地证与 Packing List 核对完成，生成 {comparison.group_count} 组对比记录",
            "output_path": str(output_path),
            "group_count": comparison.group_count,
            "issue_count": comparison.issue_count,
            "mismatch_count": comparison.mismatch_count,
            "missing_field_count": comparison.missing_field_count,
            "draft_count": len(draft_records),
            "packing_count": len(packing_records),
            "sheet_count": 1,
            "draft_file_count": 1,
            "packing_file_count": 1,
            "logs": [
                f"产地证识别记录数：{len(draft_records)}",
                f"Packing List 识别记录数：{len(packing_records)}",
                f"核对分组数：{comparison.group_count}",
                f"问题数：{comparison.issue_count}",
                f"输出文件：{output_path}",
            ],
        }

    def _assign_origin_invoice_numbers(
        self,
        draft_records: Sequence[ExtractedRecord],
        packing_records: Sequence[ExtractedRecord],
    ) -> None:
        for draft in draft_records:
            if draft.invoice_number:
                continue

            candidate_invoices = {
                packing.invoice_number
                for packing in packing_records
                if packing.invoice_number
                and self._records_can_share_invoice_group(draft, packing)
            }
            if len(candidate_invoices) == 1:
                draft.invoice_number = next(iter(candidate_invoices))
                continue

            detail = f"{self.ORIGIN_CERTIFICATE_SOURCE} INV number 缺失，无法唯一匹配 Packing 发票号"
            if detail not in draft.issues:
                draft.issues.append(detail)

    def _records_can_share_invoice_group(
        self,
        draft: ExtractedRecord,
        packing: ExtractedRecord,
    ) -> bool:
        if not draft.po_number or draft.po_number != packing.po_number:
            return False

        draft_article = draft.article_number.upper()
        packing_article = packing.article_number.upper()
        draft_working = draft.working_number.upper()
        packing_working = packing.working_number.upper()

        if draft_article and packing_article and draft_article != packing_article:
            return False
        if draft_working and packing_working and draft_working != packing_working:
            return False

        return bool(
            (draft_article and packing_article)
            or (draft_working and packing_working)
        )

    def _group_records_by_invoice(
        self,
        draft_records: Sequence[ExtractedRecord],
        packing_records: Sequence[ExtractedRecord],
    ) -> list[tuple[str, dict[str, list[ExtractedRecord]]]]:
        grouped: dict[str, dict[str, list[ExtractedRecord]]] = {}
        order: list[str] = []

        def group_for(invoice_number: str) -> dict[str, list[ExtractedRecord]]:
            key = invoice_number or self.UNMATCHED_SHEET
            if key not in grouped:
                grouped[key] = {"draft": [], "packing": []}
                order.append(key)
            return grouped[key]

        for packing in packing_records:
            if not packing.invoice_number:
                detail = "Packing List INV number 字段缺失或定位困难"
                if detail not in packing.issues:
                    packing.issues.append(detail)
            group_for(packing.invoice_number)["packing"].append(packing)

        for draft in draft_records:
            group_for(draft.invoice_number)["draft"].append(draft)

        return [(invoice_number, grouped[invoice_number]) for invoice_number in order]

    def _make_unique_sheet_title(self, raw_title: str, used_titles: set[str]) -> str:
        base_title = self._sanitize_sheet_title(raw_title)
        title = base_title
        suffix = 2
        while title.lower() in used_titles:
            suffix_text = f"_{suffix}"
            title = f"{base_title[: 31 - len(suffix_text)]}{suffix_text}"
            suffix += 1

        used_titles.add(title.lower())
        return title

    def _sanitize_sheet_title(self, raw_title: str) -> str:
        title = re.sub(r"[\[\]:*?/\\]", "_", self._normalize_text(raw_title))
        title = title.strip().strip("'") or self.UNMATCHED_SHEET
        return title[:31] or self.UNMATCHED_SHEET

    def _extract_packing_pages(self, pdf_path: str | os.PathLike[str]) -> list[PackingExtractedPage]:
        pages: list[PackingExtractedPage] = []
        with pdfplumber.open(pdf_path) as pdf:
            for page in pdf.pages:
                pages.append(
                    PackingExtractedPage(
                        text=page.extract_text(x_tolerance=1, y_tolerance=3) or "",
                        tables=page.extract_tables() or [],
                    )
                )
        return pages

    def _find_previous_item_header(
        self,
        item_matches: Sequence[re.Match[str]],
        position: int,
    ) -> re.Match[str] | None:
        previous = [match for match in item_matches if match.start() < position]
        return previous[-1] if previous else None

    def _find_next_item_start(self, item_matches: Sequence[re.Match[str]], position: int) -> int:
        for match in item_matches:
            if match.start() > position:
                return match.start()
        return 10**12

    def _find_labeled_value(self, text: str, label_pattern: str) -> str:
        match = re.search(label_pattern + r"\s*:?\s*([A-Z0-9\-./]+)", text, re.IGNORECASE)
        return self._normalize_text(match.group(1)) if match else ""

    def _extract_draft_description(self, context_text: str) -> str:
        description_lines: list[str] = []
        for raw_line in context_text.splitlines():
            line = self._normalize_text(raw_line)
            if not line:
                continue
            line = self._strip_draft_marks_prefix(line)
            if re.match(r"^(PO#|ARTICLE\s+NO|STYLE\s+NO|CUST\s+NO|QUANTITY|HS\s+Code|PSR|\(SEE ATTACHMENT\))", line, re.I):
                break
            if re.match(r"^(Page\s+\d+|Attachment|Serial No\.|Reference No\.|ASEAN-|FORM E|Issued in|See Overleaf)", line, re.I):
                continue
            if "CARTONS OF ADIDAS BRAND GARMENT" in line.upper():
                continue
            description_lines.append(line)

        return self._normalize_description(" ".join(description_lines))

    def _strip_draft_marks_prefix(self, line: str) -> str:
        prefixes = [
            "CART NO.",
            "CUST O/N",
            "PO NO",
            "ART NO",
            "SIZE ARTICLE NO.",
            "QTY STYLE NO.",
            "MADE IN CHINA",
        ]
        for prefix in prefixes:
            if line.upper().startswith(prefix):
                return line[len(prefix) :].strip()
        return line

    def _parse_packing_detail_totals(self, page_texts: Iterable[str]) -> dict[str, dict[str, int]]:
        totals: dict[str, dict[str, int]] = {}
        for text in page_texts:
            for raw_line in text.splitlines():
                detail = self._parse_packing_detail_line(self._normalize_text(raw_line))
                if not detail:
                    continue
                po_number = detail["po"]
                totals.setdefault(po_number, {"quantity": 0, "cartons": 0})
                totals[po_number]["quantity"] += detail["quantity"]
                totals[po_number]["cartons"] += detail["cartons"]
        return totals

    def _extract_packing_invoice_number(self, page_texts: Iterable[str]) -> str:
        invoice_token_re = re.compile(r"\b[A-Z0-9]*\d[A-Z0-9]*(?:[-./][A-Z0-9]*\d[A-Z0-9]*)+\b", re.IGNORECASE)
        lines = [self._normalize_text(line) for text in page_texts for line in text.splitlines()]
        for index, line in enumerate(lines):
            if not re.search(r"\bInvoice\s+Number\b", line, re.IGNORECASE):
                continue

            candidate_lines: list[str] = []
            suffix = re.split(r"\bInvoice\s+Number\b", line, maxsplit=1, flags=re.IGNORECASE)[-1]
            if suffix.strip():
                candidate_lines.append(suffix)
            candidate_lines.extend(lines[index + 1 : index + 6])

            for candidate_line in candidate_lines:
                match = invoice_token_re.search(candidate_line)
                if match:
                    return self._normalize_text(match.group(0))
        return ""

    def _parse_packing_detail_line(self, line: str) -> dict[str, Any] | None:
        tokens = line.split()
        if len(tokens) < 12 or not re.fullmatch(r"\d{4}", tokens[0]):
            return None
        if not re.fullmatch(r"\d{10}", tokens[4]):
            return None

        first_decimal_index = next(
            (
                index
                for index, token in enumerate(tokens[7:], start=7)
                if re.fullmatch(r"\d+\.\d+", token)
            ),
            None,
        )
        if first_decimal_index is None:
            return None

        numeric_prefix = [
            int(token)
            for token in tokens[7:first_decimal_index]
            if re.fullmatch(r"\d+", token)
        ]
        if len(numeric_prefix) < 3:
            return None

        return {
            "po": tokens[4],
            "article": tokens[6],
            "quantity": numeric_prefix[-2],
            "cartons": numeric_prefix[-1],
        }

    def _is_packing_summary_header(self, normalized_header: Sequence[str]) -> bool:
        required = {"pono", "workingno", "articleno"}
        return required.issubset(set(normalized_header)) and (
            "marketponumber" in normalized_header or "marketpo" in normalized_header
        )

    def _record_from_packing_summary_row(
        self,
        normalized_header: Sequence[str],
        row: Sequence[str | None],
    ) -> ExtractedRecord | None:
        po_number = self._value_by_header(normalized_header, row, "pono")
        if not re.fullmatch(r"\d{10}", po_number):
            return None

        return ExtractedRecord(
            source="Packing List",
            po_number=po_number,
            working_number=self._value_by_header(normalized_header, row, "workingno"),
            article_number=self._value_by_header(normalized_header, row, "articleno"),
            customer_number=self._value_by_header(normalized_header, row, "marketponumber")
            or self._value_by_header(normalized_header, row, "marketpo"),
            quantity=self._parse_int(self._value_by_header(normalized_header, row, "qty")),
            cartons=self._parse_int(self._value_by_header(normalized_header, row, "ctncount")),
        )

    def _attach_hts_row(
        self,
        record: ExtractedRecord,
        normalized_header: Sequence[str],
        rows: Sequence[Sequence[str | None]],
    ) -> None:
        for row in rows:
            hts = self._value_by_header(normalized_header, row, "hts")
            if hts:
                record.hs_code = hts
                return

    def _attach_packing_descriptions(self, text: str, records: Sequence[ExtractedRecord]) -> None:
        descriptions: dict[str, str] = {}
        for match in self.PACKING_BLOCK_RE.finditer(text):
            description = self._normalize_packing_description(match.group("description"))
            if description:
                descriptions[match.group("po")] = description

        for record in records:
            if record.po_number in descriptions and not record.goods_description:
                record.goods_description = descriptions[record.po_number]
            if record.goods_description and not record.hs_code:
                hts_match = self.PACKING_DESCRIPTION_HTS_RE.search(record.goods_description)
                if hts_match:
                    record.hs_code = hts_match.group("hts")

    def _add_missing_record_issues(self, record: ExtractedRecord, source: str) -> None:
        checks = [
            ("Working Number / Style Number", record.working_number),
            ("Article Number", record.article_number),
            ("Cust Number / Market PO Number", record.customer_number),
            ("Quantity", record.quantity),
            ("Cartons", record.cartons),
            ("Goods Description", record.goods_description),
            ("HS Code / HTS Code", record.hs_code),
        ]
        for field_name, value in checks:
            if value is None or value == "":
                record.issues.append(f"{source} {field_name} 字段缺失或定位困难")

    def _issues_from_record(
        self,
        record: ExtractedRecord,
        key: tuple[str, str, str],
    ) -> list[ComparisonIssue]:
        issues: list[ComparisonIssue] = []
        for issue in record.issues:
            field_name = self._field_name_from_issue(issue)
            self._append_unique_issue(
                issues,
                ComparisonIssue(
                    po_number=record.po_number,
                    key=key,
                    source=record.source,
                    field_name=field_name,
                    issue_type="字段缺失或定位困难",
                    detail=issue,
                ),
            )
        return issues

    def _compare_records(
        self,
        draft: ExtractedRecord,
        packing: ExtractedRecord,
        key: tuple[str, str, str],
    ) -> list[ComparisonIssue]:
        comparisons = [
            (
                "Working Number / Style Number",
                draft.working_number,
                packing.working_number,
                self._normalize_compare_text,
            ),
            (
                "Article Number",
                draft.article_number,
                packing.article_number,
                self._normalize_compare_text,
            ),
            (
                "Cust Number / Market PO Number",
                draft.customer_number,
                packing.customer_number,
                self._normalize_compare_text,
            ),
            ("Quantity", draft.quantity, packing.quantity, self._normalize_compare_number),
            ("Cartons", draft.cartons, packing.cartons, self._normalize_compare_number),
            (
                "Goods Description",
                draft.goods_description,
                packing.goods_description,
                self._normalize_compare_description,
            ),
            ("HS Code / HTS Code", draft.hs_code, packing.hs_code, self._normalize_compare_text),
        ]
        issues: list[ComparisonIssue] = []
        for field_name, draft_value, packing_value, normalizer in comparisons:
            if draft_value in (None, "") or packing_value in (None, ""):
                continue
            if normalizer(draft_value) == normalizer(packing_value):
                continue
            issues.append(
                ComparisonIssue(
                    po_number=draft.po_number,
                    key=key,
                    source="Both",
                    field_name=field_name,
                    issue_type="值不一致",
                    detail=f"{field_name} 不一致：产地证={draft_value}；Packing List={packing_value}",
                )
            )
        return issues

    def _status_from_issues(self, issues: Sequence[ComparisonIssue]) -> str:
        if not issues:
            return "一致"
        if any(issue.issue_type == "字段缺失或定位困难" for issue in issues):
            return "需反馈"
        return "需核对"

    def _write_workbook(self, comparison: ComparisonResult, output_path: Path) -> None:
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = self.OUTPUT_SHEET
        issues_ws = workbook.create_sheet(self.ISSUES_SHEET)

        self._write_comparison_sheet(ws, comparison)
        issues_ws.append(self.ISSUE_HEADERS)
        self._style_header(issues_ws)

        for issue in comparison.issues:
            issues_ws.append(
                [
                    issue.po_number,
                    " | ".join(issue.key),
                    issue.source,
                    issue.field_name,
                    issue.issue_type,
                    issue.detail,
                ]
            )

        self._autosize(issues_ws)
        workbook.save(output_path)

    def _write_batch_workbook(
        self,
        sheets: Sequence[InvoiceComparisonSheet],
        output_path: Path,
    ) -> None:
        workbook = openpyxl.Workbook()
        workbook.remove(workbook.active)

        for sheet in sheets:
            ws = workbook.create_sheet(sheet.sheet_name)
            self._write_comparison_sheet(ws, sheet.comparison)

        issues_ws = workbook.create_sheet(self.ISSUES_SHEET)
        issues_ws.append(self.BATCH_ISSUE_HEADERS)
        self._style_header(issues_ws)
        for sheet in sheets:
            for issue in sheet.comparison.issues:
                issues_ws.append(
                    [
                        sheet.invoice_number,
                        sheet.sheet_name,
                        issue.po_number,
                        " | ".join(issue.key),
                        issue.source,
                        issue.field_name,
                        issue.issue_type,
                        issue.detail,
                    ]
                )

        self._autosize(issues_ws)
        workbook.save(output_path)

    def _write_comparison_sheet(self, ws, comparison: ComparisonResult) -> None:
        ws.append(self.HEADERS)
        self._style_header(ws)

        for index, group in enumerate(comparison.groups):
            draft_row = ws.max_row + 1
            ws.append(self._row_values(group.draft, self.ORIGIN_CERTIFICATE_SOURCE, group))
            packing_row = ws.max_row + 1
            ws.append(self._row_values(group.packing, "Packing List", group))
            self._apply_issue_fills(ws, group, draft_row, packing_row)
            if index < len(comparison.groups) - 1:
                self._append_separator_row(ws)

        self._autosize(ws)

    def _row_values(
        self,
        record: ExtractedRecord | None,
        source: str,
        group: ComparisonGroup,
    ) -> list[Any]:
        invoice_number = self._invoice_number_for_group(group)
        if record is None:
            return [
                group.po_number,
                source,
                invoice_number,
                group.key[1],
                group.key[2],
                "",
                "",
                "",
                "",
                "",
                "",
                group.status,
                group.issue_detail,
            ]

        return [
            record.po_number,
            source,
            record.invoice_number or invoice_number,
            record.working_number,
            record.article_number,
            record.customer_number,
            record.quantity,
            record.cartons,
            record.cartons_in_words,
            record.goods_description,
            record.hs_code,
            group.status,
            group.issue_detail,
        ]

    def _invoice_number_for_group(self, group: ComparisonGroup) -> str:
        if group.packing and group.packing.invoice_number:
            return group.packing.invoice_number
        if group.draft and group.draft.invoice_number:
            return group.draft.invoice_number
        return ""

    def _apply_issue_fills(
        self,
        ws,
        group: ComparisonGroup,
        draft_row: int,
        packing_row: int,
    ) -> None:
        for issue in group.issues:
            column = self.FIELD_COLUMNS.get(issue.field_name)
            if not column:
                continue
            if column in self.NO_CONTENT_FILL_COLUMNS:
                continue
            fill = self.YELLOW_FILL if issue.issue_type == "字段缺失或定位困难" else self.RED_FILL
            if issue.source == self.ORIGIN_CERTIFICATE_SOURCE:
                ws.cell(row=draft_row, column=column).fill = fill
            elif issue.source == "Packing List":
                ws.cell(row=packing_row, column=column).fill = fill
            elif issue.source == "Both":
                ws.cell(row=draft_row, column=column).fill = fill
                ws.cell(row=packing_row, column=column).fill = fill
            else:
                ws.cell(row=draft_row, column=column).fill = fill
                ws.cell(row=packing_row, column=column).fill = fill

    def _append_separator_row(self, ws) -> None:
        separator_row = ws.max_row + 1
        ws.append([None] * len(self.HEADERS))
        for column_index in range(1, len(self.HEADERS) + 1):
            ws.cell(row=separator_row, column=column_index).fill = self.SEPARATOR_FILL

    def _style_header(self, ws) -> None:
        for cell in ws[1]:
            cell.font = Font(bold=True)
            cell.fill = self.HEADER_FILL
            cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)

    def _autosize(self, ws) -> None:
        for column_cells in ws.columns:
            column_letter = get_column_letter(column_cells[0].column)
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            ws.column_dimensions[column_letter].width = min(max(max_length + 2, 12), 60)
        for row in ws.iter_rows():
            for cell in row:
                cell.alignment = Alignment(vertical="top", wrap_text=True)

    def _field_name_from_issue(self, issue: str) -> str:
        for field_name in self.FIELD_COLUMNS:
            if field_name in issue:
                return field_name
        if "HTS" in issue or "HS" in issue:
            return "HS Code / HTS Code"
        if "明细 Quantity" in issue:
            return "Quantity"
        if "明细 Cartons" in issue:
            return "Cartons"
        return "Record"

    def _append_unique_issue(self, issues: list[ComparisonIssue], issue: ComparisonIssue) -> None:
        signature = (issue.source, issue.field_name, issue.issue_type, issue.detail)
        if signature not in {
            (entry.source, entry.field_name, entry.issue_type, entry.detail)
            for entry in issues
        }:
            issues.append(issue)

    def _value_by_header(
        self,
        normalized_header: Sequence[str],
        row: Sequence[str | None],
        header_key: str,
    ) -> str:
        try:
            index = normalized_header.index(header_key)
        except ValueError:
            return ""
        if index >= len(row):
            return ""
        return self._normalize_text(row[index])

    def _normalize_header(self, value: str | None) -> str:
        return re.sub(r"[^a-z0-9]+", "", self._normalize_text(value).lower())

    def _normalize_text(self, value: Any) -> str:
        if value is None:
            return ""
        text = str(value).replace("\r", "\n").replace("\xa0", " ").replace("©", "'")
        return re.sub(r"\s+", " ", text).strip()

    def _normalize_description(self, value: Any) -> str:
        text = self._normalize_text(value)
        text = re.sub(r"\b[A-Z]{3}\.\d{2},\d{4}\b", "", text)
        text = re.sub(r"\s+,", ",", text)
        text = re.sub(r",\s+", ",", text)
        return self._normalize_text(text)

    def _normalize_packing_description(self, value: Any) -> str:
        description_lines: list[str] = []
        for raw_line in str(value).replace("\r", "\n").splitlines():
            line = self._normalize_text(raw_line)
            if not line:
                continue
            footer_match = self.PACKING_DESCRIPTION_FOOTER_RE.search(line)
            if footer_match:
                line = line[: footer_match.start()].strip()
                if line:
                    description_lines.append(line)
                break
            description_lines.append(line)

        return self._normalize_description(" ".join(description_lines))

    def _normalize_compare_text(self, value: Any) -> str:
        return self._normalize_text(value).upper()

    def _normalize_compare_description(self, value: Any) -> str:
        return self._normalize_description(value).upper()

    def _normalize_compare_number(self, value: Any) -> str:
        parsed = self._parse_int(value)
        return "" if parsed is None else str(parsed)

    def _parse_int(self, value: Any) -> int | None:
        text = self._normalize_text(value).replace(",", "")
        if not text:
            return None
        match = re.search(r"-?\d+", text)
        return int(match.group(0)) if match else None

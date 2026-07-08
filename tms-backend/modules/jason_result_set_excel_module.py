from __future__ import annotations

import re
from copy import copy
from dataclasses import dataclass
from datetime import datetime
from pathlib import Path
from typing import Any, Iterable
from uuid import uuid4

import openpyxl
from openpyxl.styles import Font, PatternFill


@dataclass(frozen=True)
class ResultSetRow:
    working_number: str
    article_number: str
    po_number: str
    market_po_number: str
    gps_customer_number: str
    assigned_factory: str
    podd: datetime | None
    price_per_unit: float
    total_adjustments: float
    ordered_quantity: float
    shipped_quantity: float
    shipped_status: str
    order_type: str
    pack: Any = None
    season: Any = None
    description: Any = None
    tms_merchandiser: Any = None
    factory_merchandiser: Any = None


@dataclass
class ResultSetGroup:
    working_number: str
    article_number: str
    po_number: str
    market_po_number: str
    gps_customer_number: str
    assigned_factory: str
    podd: datetime | None
    price_per_unit: float
    total_adjustments: float
    pack: Any = None
    season: Any = None
    description: Any = None
    tms_merchandiser: Any = None
    factory_merchandiser: Any = None
    ordered_quantity: float = 0
    shipped_quantity: float = 0
    statuses: set[str] | None = None

    def __post_init__(self) -> None:
        if self.statuses is None:
            self.statuses = set()


@dataclass(frozen=True)
class TemplateLookupRow:
    bulk_handover_date: Any


@dataclass(frozen=True)
class TargetRow:
    pack: Any
    season: Any
    working_number: str
    article_number: str
    description: Any
    po_number: str
    market_po_number: str
    gps_customer_number: str
    customer_warehouse: str
    bulk_handover_date: Any
    podd: datetime | None
    ordered_quantity: float
    tms_price: float
    total_adjustments: float
    factory_code: str
    factory_name: str
    tms_merchandiser: Any
    factory_merchandiser: Any


class JasonResultSetExcelModule:
    """Generate Jason target Excel rows from an adidas Result Set export."""

    REQUIRED_COLUMNS = [
        "Working Number",
        "Article Number",
        "PO Number",
        "Market PO Number",
        "Gps Customer Number",
        "Assigned Factory",
        "PODD",
        "Price Per Unit",
        "Total Adjustments",
        "Ordered Quantity",
        "Shipped Qty",
        "Shipped Status",
        "Order Type",
    ]
    ALLOWED_ORDER_TYPES = {"ZGPS", "ZREG"}
    TARGET_HEADER_MERGES = (
        "A1:A2",
        "B1:B2",
        "C1:C2",
        "D1:D2",
        "E1:E2",
        "F1:F2",
        "G1:G2",
        "H1:H2",
        "I1:I2",
        "J1:J2",
        "K1:K2",
        "L1:N1",
        "O1:R1",
        "S1:T1",
        "U1:U2",
        "V1:V2",
        "W1:W2",
    )
    TARGET_PRICE_FILL_COLUMNS = frozenset(range(12, 19))
    TARGET_SHEET_ORDER = ("目标表", "Sheet1", "Sheet2", "Sheet3")
    SHEET1_HEADERS = (
        "Working Number",
        "Article Number",
        "PO Number",
        "Market PO Number",
        "Gps Customer Number",
        "Assigned Factory",
        "PODD",
        "Price Per Unit",
        "Total Adjustments",
        "Sum of Ordered Quantity",
    )

    def __init__(self, template_path: str | Path | None = None) -> None:
        self.template_path = Path(template_path) if template_path is not None else self._default_template_path()

    def process_result_set(
        self,
        *,
        result_set_path: str | Path,
        target_month: str,
        output_dir: str | Path,
    ) -> dict[str, Any]:
        year, month = self._parse_target_month(target_month)
        template_context = self._load_template_context()
        rows = self._read_result_set_rows(result_set_path)
        target_rows, warnings, counts = self._build_target_rows(
            rows,
            target_year=year,
            target_month=month,
            template_context=template_context,
        )

        output_path = Path(output_dir) / f"jason_result_set_excel_{uuid4().hex}.xlsx"
        self._write_output_workbook(template_context["workbook"], target_rows, output_path)

        return {
            "success": True,
            "message": f"Jason Result Set Excel 生成完成，共写入 {len(target_rows)} 行",
            "output_path": str(output_path),
            "written_row_count": len(target_rows),
            "not_shipped_row_count": counts["not_shipped"],
            "partial_row_count": counts["partial"],
            "unknown_lookup_count": counts["unknown_lookup"],
            "warnings": warnings,
        }

    @classmethod
    def _default_template_path(cls) -> Path:
        return Path(__file__).resolve().parent.parent / "templates" / "jason_result_set_excel_template.xlsx"

    def _load_template_context(self) -> dict[str, Any]:
        if not self.template_path.exists():
            raise ValueError("Jason Result Set Excel 模板不存在")

        workbook = openpyxl.load_workbook(self.template_path, data_only=False, keep_links=False)
        if "目标表" not in workbook.sheetnames:
            workbook.close()
            raise ValueError("Jason Result Set Excel 模板缺少目标表")

        target_sheet = workbook["目标表"]
        warehouse_lookup = self._read_two_column_lookup(workbook, "Sheet2")
        factory_lookup = self._read_two_column_lookup(workbook, "Sheet3")
        row_lookup, article_lookup = self._read_target_lookup(target_sheet)
        style_templates = [
            self._capture_cell_style(target_sheet.cell(row=3, column=column_index))
            for column_index in range(1, 24)
        ]

        return {
            "workbook": workbook,
            "target_sheet": target_sheet,
            "warehouse_lookup": warehouse_lookup,
            "factory_lookup": factory_lookup,
            "row_lookup": row_lookup,
            "article_lookup": article_lookup,
            "style_templates": style_templates,
        }

    def _read_two_column_lookup(
        self,
        workbook: openpyxl.Workbook,
        sheet_name: str,
    ) -> dict[str, str]:
        if sheet_name not in workbook.sheetnames:
            return {}

        sheet = workbook[sheet_name]
        lookup: dict[str, str] = {}
        for row in sheet.iter_rows(min_row=2, values_only=True):
            key = self._normalize_identifier(row[0] if len(row) > 0 else None)
            value = self._normalize_text(row[1] if len(row) > 1 else None)
            if key:
                lookup[key] = value
        return lookup

    def _read_target_lookup(
        self,
        target_sheet: openpyxl.worksheet.worksheet.Worksheet,
    ) -> tuple[dict[tuple[str, str], TemplateLookupRow], dict[str, TemplateLookupRow]]:
        row_lookup: dict[tuple[str, str], TemplateLookupRow] = {}
        article_lookup: dict[str, TemplateLookupRow] = {}

        for row_index in range(3, target_sheet.max_row + 1):
            working_number = self._normalize_identifier(target_sheet.cell(row=row_index, column=3).value)
            article_number = self._normalize_identifier(target_sheet.cell(row=row_index, column=4).value)
            po_number = self._normalize_identifier(target_sheet.cell(row=row_index, column=6).value)
            if not working_number and not article_number and not po_number:
                continue
            if not working_number or not article_number:
                continue
            if str(po_number).startswith("="):
                continue

            lookup_row = TemplateLookupRow(
                bulk_handover_date=target_sheet.cell(row=row_index, column=10).value,
            )
            row_lookup[(working_number, article_number)] = lookup_row
            article_lookup.setdefault(article_number, lookup_row)

        return row_lookup, article_lookup

    def _read_result_set_rows(self, result_set_path: str | Path) -> list[ResultSetRow]:
        workbook = openpyxl.load_workbook(result_set_path, data_only=True, read_only=True, keep_links=False)
        try:
            sheet = workbook["Result Set"] if "Result Set" in workbook.sheetnames else workbook.worksheets[0]
            header_values = next(sheet.iter_rows(min_row=1, max_row=1, values_only=True), None)
            if not header_values:
                raise ValueError("Result Set 表头为空")
            headers = {
                self._normalize_header(value): index
                for index, value in enumerate(header_values)
                if self._normalize_header(value)
            }
            missing_columns = [
                column_name
                for column_name in self.REQUIRED_COLUMNS
                if self._normalize_header(column_name) not in headers
            ]
            if missing_columns:
                raise ValueError(f"Result Set 缺少必需列：{', '.join(missing_columns)}")

            rows: list[ResultSetRow] = []
            for values in sheet.iter_rows(min_row=2, values_only=True):
                if not any(value not in (None, "") for value in values):
                    continue
                rows.append(self._parse_result_set_row(values, headers))
            return rows
        finally:
            workbook.close()

    def _parse_result_set_row(
        self,
        values: tuple[Any, ...],
        headers: dict[str, int],
    ) -> ResultSetRow:
        def read(column_name: str) -> Any:
            index = headers[self._normalize_header(column_name)]
            return values[index] if index < len(values) else None

        def read_optional(column_name: str) -> Any:
            index = headers.get(self._normalize_header(column_name))
            return values[index] if index is not None and index < len(values) else None

        return ResultSetRow(
            working_number=self._normalize_identifier(read("Working Number")),
            article_number=self._normalize_identifier(read("Article Number")),
            po_number=self._normalize_identifier(read("PO Number")),
            market_po_number=self._normalize_identifier(read("Market PO Number")),
            gps_customer_number=self._normalize_identifier(read("Gps Customer Number")),
            assigned_factory=self._normalize_identifier(read("Assigned Factory")),
            podd=self._coerce_datetime(read("PODD")),
            price_per_unit=self._coerce_float(read("Price Per Unit")),
            total_adjustments=self._coerce_float(read("Total Adjustments")),
            ordered_quantity=self._coerce_float(read("Ordered Quantity")),
            shipped_quantity=self._coerce_float(read("Shipped Qty")),
            shipped_status=self._normalize_text(read("Shipped Status")),
            order_type=self._normalize_identifier(read("Order Type")),
            pack=self._normalize_optional_output_value(read_optional("Pack")),
            season=self._normalize_optional_output_value(read_optional("Season")),
            description=self._normalize_optional_output_value(read_optional("Description")),
            tms_merchandiser=self._normalize_optional_output_value(read_optional("TMS Merchandiser")),
            factory_merchandiser=self._normalize_optional_output_value(read_optional("Factory Merchandiser")),
        )

    def _build_target_rows(
        self,
        rows: Iterable[ResultSetRow],
        *,
        target_year: int,
        target_month: int,
        template_context: dict[str, Any],
    ) -> tuple[list[TargetRow], list[str], dict[str, int]]:
        warehouse_lookup: dict[str, str] = template_context["warehouse_lookup"]
        factory_lookup: dict[str, str] = template_context["factory_lookup"]
        row_lookup: dict[tuple[str, str], TemplateLookupRow] = template_context["row_lookup"]
        article_lookup: dict[str, TemplateLookupRow] = template_context["article_lookup"]
        allowed_gps = set(warehouse_lookup)
        groups: dict[tuple[Any, ...], ResultSetGroup] = {}

        for row in rows:
            if row.order_type not in self.ALLOWED_ORDER_TYPES:
                continue
            if allowed_gps and row.gps_customer_number not in allowed_gps:
                continue
            if row.shipped_status == "Fully Shipped":
                continue
            if row.shipped_status != "Partially Shipped" and not self._is_target_month(row.podd, target_year, target_month):
                continue

            group_key = (
                row.working_number,
                row.article_number,
                row.po_number,
                row.market_po_number,
                row.gps_customer_number,
                row.assigned_factory,
                row.podd,
                row.price_per_unit,
                row.total_adjustments,
            )
            group = groups.get(group_key)
            if group is None:
                group = ResultSetGroup(
                    working_number=row.working_number,
                    article_number=row.article_number,
                    po_number=row.po_number,
                    market_po_number=row.market_po_number,
                    gps_customer_number=row.gps_customer_number,
                    assigned_factory=row.assigned_factory,
                    podd=row.podd,
                    price_per_unit=row.price_per_unit,
                    total_adjustments=row.total_adjustments,
                    pack=row.pack,
                    season=row.season,
                    description=row.description,
                    tms_merchandiser=row.tms_merchandiser,
                    factory_merchandiser=row.factory_merchandiser,
                )
                groups[group_key] = group

            group.ordered_quantity += row.ordered_quantity
            group.shipped_quantity += row.shipped_quantity
            group.statuses.add(row.shipped_status)
            self._fill_group_optional_fields(group, row)

        warnings: list[str] = []
        target_rows: list[TargetRow] = []
        counts = {"not_shipped": 0, "partial": 0, "unknown_lookup": 0}

        for group in sorted(groups.values(), key=self._group_sort_key):
            is_partial = "Partially Shipped" in group.statuses
            quantity = group.ordered_quantity - group.shipped_quantity if is_partial else group.ordered_quantity
            if quantity <= 0:
                continue

            lookup = row_lookup.get((group.working_number, group.article_number)) or article_lookup.get(group.article_number)
            if lookup is None:
                lookup = TemplateLookupRow(group.podd)
                counts["unknown_lookup"] += 1
                warnings.append(
                    f"未找到模板 lookup：Working Number={group.working_number}, Article={group.article_number}"
                )

            adjustment = (
                round(group.total_adjustments * group.shipped_quantity / group.ordered_quantity, 2)
                if is_partial and group.ordered_quantity
                else group.total_adjustments
            )
            target_rows.append(
                TargetRow(
                    pack=group.pack,
                    season=group.season,
                    working_number=group.working_number,
                    article_number=group.article_number,
                    description=group.description,
                    po_number=group.po_number,
                    market_po_number=group.market_po_number,
                    gps_customer_number=group.gps_customer_number,
                    customer_warehouse=warehouse_lookup.get(group.gps_customer_number, ""),
                    bulk_handover_date=lookup.bulk_handover_date or group.podd,
                    podd=group.podd,
                    ordered_quantity=self._format_quantity(quantity),
                    tms_price=group.price_per_unit,
                    total_adjustments=adjustment,
                    factory_code=group.assigned_factory,
                    factory_name=factory_lookup.get(group.assigned_factory, ""),
                    tms_merchandiser=group.tms_merchandiser,
                    factory_merchandiser=group.factory_merchandiser,
                )
            )
            counts["partial" if is_partial else "not_shipped"] += 1

        return target_rows, warnings, counts

    def _write_output_workbook(
        self,
        workbook: openpyxl.Workbook,
        target_rows: list[TargetRow],
        output_path: Path,
    ) -> None:
        self._write_sheet1_summary(workbook, target_rows)
        target_sheet = workbook["目标表"]
        style_templates = [
            self._capture_cell_style(target_sheet.cell(row=3, column=column_index))
            for column_index in range(1, 24)
        ]

        if target_sheet.max_row >= 3:
            target_sheet.delete_rows(3, target_sheet.max_row - 2)
        self._ensure_target_header_merges(target_sheet)
        self._sync_factory_subheader_fill(target_sheet)

        for row_offset, target_row in enumerate(target_rows, start=3):
            self._apply_row_style(target_sheet, row_offset, style_templates)
            self._clear_non_price_data_row_fill(target_sheet, row_offset)
            values = [
                target_row.pack,
                target_row.season,
                target_row.working_number,
                target_row.article_number,
                target_row.description,
                target_row.po_number,
                target_row.market_po_number,
                target_row.gps_customer_number,
                target_row.customer_warehouse,
                target_row.bulk_handover_date,
                target_row.ordered_quantity,
                None,
                f"=ROUND(0.13*L{row_offset}*K{row_offset},2)",
                f"=ROUND(L{row_offset}*K{row_offset}+M{row_offset},2)",
                target_row.tms_price,
                f"=ROUND(0.13*(K{row_offset}*O{row_offset}+Q{row_offset}),2)",
                target_row.total_adjustments,
                f"=ROUND(O{row_offset}*K{row_offset}+P{row_offset}+Q{row_offset},2)",
                target_row.factory_code,
                target_row.factory_name,
                None,
                target_row.tms_merchandiser,
                target_row.factory_merchandiser,
            ]
            for column_index, value in enumerate(values, start=1):
                target_sheet.cell(row=row_offset, column=column_index).value = value

        total_row = 3 + len(target_rows)
        self._apply_row_style(target_sheet, total_row, style_templates)
        self._clear_row_fill(target_sheet, total_row)
        last_data_row = total_row - 1
        if target_rows:
            target_sheet.cell(row=total_row, column=6).value = f"=COUNTA(F3:F{last_data_row})"
            target_sheet.cell(row=total_row, column=11).value = f"=SUM(K3:K{last_data_row})"
            target_sheet.cell(row=total_row, column=13).value = f"=SUM(M3:M{last_data_row})"
            target_sheet.cell(row=total_row, column=14).value = f"=SUM(N3:N{last_data_row})"
            target_sheet.cell(row=total_row, column=16).value = f"=SUM(P3:P{last_data_row})"
            target_sheet.cell(row=total_row, column=17).value = f"=SUM(Q3:Q{last_data_row})"
            target_sheet.cell(row=total_row, column=18).value = f"=SUM(R3:R{last_data_row})"

        output_path.parent.mkdir(parents=True, exist_ok=True)
        self._ensure_workbook_sheet_order(workbook)
        workbook.save(output_path)
        workbook.close()

    def _write_sheet1_summary(self, workbook: openpyxl.Workbook, target_rows: list[TargetRow]) -> None:
        if "Sheet1" in workbook.sheetnames:
            del workbook["Sheet1"]
        sheet = workbook.create_sheet("Sheet1", 0)

        sheet.cell(row=1, column=1).value = "Order Type"
        sheet.cell(row=1, column=2).value = "(Multiple Items)"
        for column_index, header in enumerate(self.SHEET1_HEADERS, start=1):
            cell = sheet.cell(row=3, column=column_index)
            cell.value = header
            cell.font = Font(bold=True)

        for row_index, target_row in enumerate(target_rows, start=4):
            values = [
                target_row.working_number,
                target_row.article_number,
                target_row.po_number,
                target_row.market_po_number,
                target_row.gps_customer_number,
                target_row.factory_code,
                target_row.podd,
                target_row.tms_price,
                target_row.total_adjustments,
                target_row.ordered_quantity,
            ]
            for column_index, value in enumerate(values, start=1):
                sheet.cell(row=row_index, column=column_index).value = value

        last_detail_row = 3 + len(target_rows)
        total_row = last_detail_row + 1
        sheet.cell(row=total_row, column=1).value = "Grand Total"
        sheet.cell(row=total_row, column=1).font = Font(bold=True)
        sheet.cell(row=total_row, column=10).value = sum(row.ordered_quantity or 0 for row in target_rows)
        sheet.cell(row=total_row, column=10).font = Font(bold=True)

        sheet.auto_filter.ref = f"A3:J{last_detail_row}"
        for row_index in range(4, total_row):
            sheet.cell(row=row_index, column=7).number_format = "yyyy-mm-dd"
            sheet.cell(row=row_index, column=8).number_format = "#,##0.00"
            sheet.cell(row=row_index, column=9).number_format = "#,##0.00"
            sheet.cell(row=row_index, column=10).number_format = "#,##0"
        sheet.cell(row=total_row, column=10).number_format = "#,##0"

    def _ensure_target_header_merges(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
    ) -> None:
        for merged_range in list(sheet.merged_cells.ranges):
            if (
                merged_range.min_row <= 2
                and merged_range.max_row >= 1
                and merged_range.min_col <= 23
                and merged_range.max_col >= 1
            ):
                sheet.unmerge_cells(str(merged_range))

        for range_ref in self.TARGET_HEADER_MERGES:
            sheet.merge_cells(range_ref)

        sheet.auto_filter.ref = None

    @staticmethod
    def _sync_factory_subheader_fill(
        sheet: openpyxl.worksheet.worksheet.Worksheet,
    ) -> None:
        factory_header_fill = copy(sheet["S1"].fill)
        sheet["S2"].fill = copy(factory_header_fill)
        sheet["T2"].fill = copy(factory_header_fill)

    def _ensure_workbook_sheet_order(self, workbook: openpyxl.Workbook) -> None:
        ordered_sheet_index = 0
        for sheet_name in self.TARGET_SHEET_ORDER:
            if sheet_name not in workbook.sheetnames:
                continue
            current_index = workbook.sheetnames.index(sheet_name)
            workbook.move_sheet(workbook[sheet_name], offset=ordered_sheet_index - current_index)
            ordered_sheet_index += 1

    def _capture_cell_style(self, cell: openpyxl.cell.cell.Cell) -> dict[str, Any]:
        return {
            "style": copy(cell._style),
            "number_format": cell.number_format,
            "font": copy(cell.font),
            "fill": copy(cell.fill),
            "border": copy(cell.border),
            "alignment": copy(cell.alignment),
            "protection": copy(cell.protection),
        }

    def _apply_row_style(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        row_index: int,
        style_templates: list[dict[str, Any]],
    ) -> None:
        for column_index, style in enumerate(style_templates, start=1):
            cell = sheet.cell(row=row_index, column=column_index)
            cell._style = copy(style["style"])
            cell.number_format = style["number_format"]
            cell.font = copy(style["font"])
            cell.fill = copy(style["fill"])
            cell.border = copy(style["border"])
            cell.alignment = copy(style["alignment"])
            cell.protection = copy(style["protection"])

    @staticmethod
    def _clear_row_fill(sheet: openpyxl.worksheet.worksheet.Worksheet, row_index: int) -> None:
        for column_index in range(1, 24):
            sheet.cell(row=row_index, column=column_index).fill = PatternFill()

    def _clear_non_price_data_row_fill(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        row_index: int,
    ) -> None:
        for column_index in range(1, 24):
            if column_index not in self.TARGET_PRICE_FILL_COLUMNS:
                sheet.cell(row=row_index, column=column_index).fill = PatternFill()

    @staticmethod
    def _fill_group_optional_fields(group: ResultSetGroup, row: ResultSetRow) -> None:
        if group.pack is None and row.pack is not None:
            group.pack = row.pack
        if group.season is None and row.season is not None:
            group.season = row.season
        if group.description is None and row.description is not None:
            group.description = row.description
        if group.tms_merchandiser is None and row.tms_merchandiser is not None:
            group.tms_merchandiser = row.tms_merchandiser
        if group.factory_merchandiser is None and row.factory_merchandiser is not None:
            group.factory_merchandiser = row.factory_merchandiser

    def _group_sort_key(self, group: ResultSetGroup) -> tuple[str, str, str, str, str]:
        return (
            group.working_number,
            group.po_number,
            group.article_number,
            group.market_po_number,
            group.gps_customer_number,
        )

    @staticmethod
    def _parse_target_month(target_month: str) -> tuple[int, int]:
        if not re.fullmatch(r"\d{4}-\d{2}", str(target_month or "")):
            raise ValueError("target_month 必须是 YYYY-MM")
        year_text, month_text = target_month.split("-", 1)
        year = int(year_text)
        month = int(month_text)
        if month < 1 or month > 12:
            raise ValueError("target_month 月份无效")
        return year, month

    @staticmethod
    def _is_target_month(value: datetime | None, year: int, month: int) -> bool:
        return value is not None and value.year == year and value.month == month

    @staticmethod
    def _coerce_datetime(value: Any) -> datetime | None:
        if isinstance(value, datetime):
            return value
        return None

    @staticmethod
    def _coerce_float(value: Any) -> float:
        if value in (None, ""):
            return 0.0
        try:
            return float(str(value).replace(",", ""))
        except (TypeError, ValueError):
            return 0.0

    @staticmethod
    def _format_quantity(value: float) -> int | float:
        return int(value) if float(value).is_integer() else value

    @staticmethod
    def _normalize_header(value: Any) -> str:
        return str(value or "").strip().lower()

    @staticmethod
    def _normalize_text(value: Any) -> str:
        if value is None:
            return ""
        text = str(value).strip()
        return "" if text.lower() in {"none", "nan"} else text

    @classmethod
    def _normalize_optional_output_value(cls, value: Any) -> Any:
        if value is None:
            return None
        if isinstance(value, str):
            text = cls._normalize_text(value)
            return text or None
        return value

    @classmethod
    def _normalize_identifier(cls, value: Any) -> str:
        text = cls._normalize_text(value)
        if text.endswith(".0") and text[:-2].isdigit():
            return text[:-2]
        return text

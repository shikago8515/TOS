# -*- coding: utf-8 -*-
"""iPlex 双表数据核对模块。"""

from __future__ import annotations

import os
from dataclasses import dataclass
from decimal import Decimal, InvalidOperation, ROUND_HALF_UP
from pathlib import Path
from typing import Any, Sequence
from uuid import uuid4

import openpyxl
import xlrd
from openpyxl.styles import Font, PatternFill
from openpyxl.utils import get_column_letter


@dataclass(frozen=True)
class IplexDualTableCompareConfig:
    main_sheet_name: str
    lookup_sheet_name: str
    main_header_row: int
    lookup_header_row: int
    main_key_column: int
    lookup_key_column: int
    four_digit_main_column: int
    four_digit_lookup_column: int
    two_digit_main_column: int
    two_digit_lookup_column: int
    four_digit_result_header: str = "4位小数差值"
    two_digit_result_header: str = "2位小数差值"


@dataclass(frozen=True)
class SheetData:
    name: str
    rows: list[list[Any]]
    max_row: int
    max_column: int


class IplexDualTableCompareModule:
    """按用户指定列生成 VLOOKUP 差值核对 Excel。"""

    HELPER_SHEET_NAME = "_iplex_lookup"
    HEADER_FILL = PatternFill("solid", fgColor="FFD9EAF7")
    MISMATCH_FILL = PatternFill("solid", fgColor="FFFFC7CE")

    def inspect_workbook(
        self,
        workbook_path: str | os.PathLike[str],
        *,
        sheet_name: str | None = None,
        header_row: int = 1,
    ) -> dict[str, Any]:
        sheets = self._read_workbook(workbook_path)
        if not sheets:
            raise ValueError("工作簿中没有可读取的 sheet")

        selected = self._select_sheet(sheets, sheet_name)
        self._validate_row_index(header_row, selected.max_row, "表头行")
        header_values = selected.rows[header_row - 1] if header_row <= len(selected.rows) else []
        sample_values = (
            selected.rows[header_row]
            if header_row < len(selected.rows)
            else []
        )

        return {
            "sheets": [
                {
                    "name": sheet.name,
                    "max_row": sheet.max_row,
                    "max_column": sheet.max_column,
                }
                for sheet in sheets
            ],
            "selected_sheet": {
                "name": selected.name,
                "header_row": header_row,
                "max_row": selected.max_row,
                "max_column": selected.max_column,
                "data_row_count": max(selected.max_row - header_row, 0),
                "headers": [
                    {
                        "index": index,
                        "letter": get_column_letter(index),
                        "label": self._display_text(self._read_cell(header_values, index)),
                        "sample_value": self._display_text(self._read_cell(sample_values, index)),
                    }
                    for index in range(1, selected.max_column + 1)
                ],
            },
        }

    def process_files(
        self,
        *,
        main_path: str | os.PathLike[str],
        lookup_path: str | os.PathLike[str],
        config: IplexDualTableCompareConfig,
        output_dir: str | os.PathLike[str],
    ) -> dict[str, Any]:
        main_sheet = self._select_sheet(self._read_workbook(main_path), config.main_sheet_name)
        lookup_sheet = self._select_sheet(self._read_workbook(lookup_path), config.lookup_sheet_name)
        self._validate_config(config, main_sheet, lookup_sheet)

        output_workbook = openpyxl.Workbook()
        result_sheet = output_workbook.active
        result_sheet.title = self._safe_sheet_title(main_sheet.name)
        helper_sheet = output_workbook.create_sheet(self.HELPER_SHEET_NAME)
        helper_sheet.sheet_state = "hidden"

        main_max_column = main_sheet.max_column
        four_result_column = main_max_column + 1
        two_result_column = main_max_column + 2
        self._write_main_values(result_sheet, main_sheet, main_max_column)
        self._write_helper_sheet(helper_sheet, lookup_sheet, config)
        lookup_last_row = max(lookup_sheet.max_row - config.lookup_header_row + 1, 1)
        helper_last_row = max(lookup_last_row, 2)
        helper_range = f"'{self.HELPER_SHEET_NAME}'!$A$2:$C${helper_last_row}"

        result_sheet.cell(config.main_header_row, four_result_column).value = config.four_digit_result_header
        result_sheet.cell(config.main_header_row, two_result_column).value = config.two_digit_result_header

        preview_rows = self._build_preview_rows(main_sheet, lookup_sheet, config)
        mismatch_row_numbers = {row["row_number"] for row in preview_rows}

        for row_index in range(config.main_header_row + 1, main_sheet.max_row + 1):
            row_values = main_sheet.rows[row_index - 1] if row_index - 1 < len(main_sheet.rows) else []
            key = self._normalize_key(self._read_cell(row_values, config.main_key_column))
            if not key:
                continue

            key_ref = f"{get_column_letter(config.main_key_column)}{row_index}"
            four_main_ref = f"{get_column_letter(config.four_digit_main_column)}{row_index}"
            two_main_ref = f"{get_column_letter(config.two_digit_main_column)}{row_index}"
            result_sheet.cell(row_index, four_result_column).value = (
                f"=ROUND({four_main_ref}-VLOOKUP({key_ref},{helper_range},2,FALSE),4)"
            )
            result_sheet.cell(row_index, two_result_column).value = (
                f"=ROUND({two_main_ref}-VLOOKUP({key_ref},{helper_range},3,FALSE),2)"
            )
            result_sheet.cell(row_index, four_result_column).number_format = "0.0000"
            result_sheet.cell(row_index, two_result_column).number_format = "0.00"
            if row_index in mismatch_row_numbers:
                for column_index in range(1, two_result_column + 1):
                    result_sheet.cell(row_index, column_index).fill = self.MISMATCH_FILL

        self._style_result_sheet(
            result_sheet,
            header_row=config.main_header_row,
            max_row=main_sheet.max_row,
            max_column=two_result_column,
        )

        output_path = Path(output_dir) / f"iplex_dual_table_compare_{uuid4().hex}.xlsx"
        output_workbook.save(output_path)

        summary = self._build_summary(main_sheet, lookup_sheet, config)
        return {
            "success": True,
            "message": f"iPlex 双表核对完成，结果文件：{output_path}",
            "logs": [],
            "output_path": str(output_path),
            "preview_rows": preview_rows,
            **summary,
        }

    def _read_workbook(self, workbook_path: str | os.PathLike[str]) -> list[SheetData]:
        path = Path(workbook_path)
        suffix = path.suffix.lower()
        if suffix == ".xls":
            return self._read_xls_workbook(path)
        if suffix in {".xlsx", ".xlsm"}:
            return self._read_openpyxl_workbook(path)
        raise ValueError("仅支持 .xls / .xlsx / .xlsm 文件")

    def _read_openpyxl_workbook(self, path: Path) -> list[SheetData]:
        workbook = openpyxl.load_workbook(path, data_only=False, read_only=True)
        try:
            sheets: list[SheetData] = []
            for worksheet in workbook.worksheets:
                max_row = worksheet.max_row or 0
                max_column = worksheet.max_column or 0
                rows = [
                    list(row)
                    for row in worksheet.iter_rows(
                        min_row=1,
                        max_row=max_row,
                        max_col=max_column,
                        values_only=True,
                    )
                ]
                sheets.append(
                    SheetData(
                        name=worksheet.title,
                        rows=rows,
                        max_row=max_row,
                        max_column=max_column,
                    )
                )
            return sheets
        finally:
            workbook.close()

    def _read_xls_workbook(self, path: Path) -> list[SheetData]:
        book = xlrd.open_workbook(str(path), on_demand=True)
        try:
            sheets: list[SheetData] = []
            for sheet in book.sheets():
                rows = [list(sheet.row_values(row_index)) for row_index in range(sheet.nrows)]
                sheets.append(
                    SheetData(
                        name=sheet.name,
                        rows=rows,
                        max_row=sheet.nrows,
                        max_column=sheet.ncols,
                    )
                )
            return sheets
        finally:
            book.release_resources()

    def _select_sheet(self, sheets: Sequence[SheetData], sheet_name: str | None) -> SheetData:
        if sheet_name:
            for sheet in sheets:
                if sheet.name == sheet_name:
                    return sheet
            raise ValueError(f"未找到 sheet：{sheet_name}")
        return sheets[0]

    def _validate_config(
        self,
        config: IplexDualTableCompareConfig,
        main_sheet: SheetData,
        lookup_sheet: SheetData,
    ) -> None:
        self._validate_row_index(config.main_header_row, main_sheet.max_row, "主表表头行")
        self._validate_row_index(config.lookup_header_row, lookup_sheet.max_row, "查找表表头行")
        for label, column, max_column in (
            ("主表关键列", config.main_key_column, main_sheet.max_column),
            ("查找表关键列", config.lookup_key_column, lookup_sheet.max_column),
            ("主表 4 位数值列", config.four_digit_main_column, main_sheet.max_column),
            ("查找表 4 位数值列", config.four_digit_lookup_column, lookup_sheet.max_column),
            ("主表 2 位数值列", config.two_digit_main_column, main_sheet.max_column),
            ("查找表 2 位数值列", config.two_digit_lookup_column, lookup_sheet.max_column),
        ):
            if column < 1 or column > max_column:
                raise ValueError(f"{label} 超出可用列范围")

    def _validate_row_index(self, row_index: int, max_row: int, label: str) -> None:
        if row_index < 1 or row_index > max(max_row, 1):
            raise ValueError(f"{label} 超出可用行范围")

    def _write_main_values(
        self,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        sheet: SheetData,
        max_column: int,
    ) -> None:
        for row_index, row_values in enumerate(sheet.rows, start=1):
            for column_index in range(1, max_column + 1):
                worksheet.cell(row_index, column_index).value = self._read_cell(row_values, column_index)

    def _write_helper_sheet(
        self,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        lookup_sheet: SheetData,
        config: IplexDualTableCompareConfig,
    ) -> None:
        worksheet.append(["Lookup Key", "4 Digit Lookup Value", "2 Digit Lookup Value"])
        for source_row in lookup_sheet.rows[config.lookup_header_row:]:
            worksheet.append(
                [
                    self._read_cell(source_row, config.lookup_key_column),
                    self._read_cell(source_row, config.four_digit_lookup_column),
                    self._read_cell(source_row, config.two_digit_lookup_column),
                ]
            )

    def _style_result_sheet(
        self,
        worksheet: openpyxl.worksheet.worksheet.Worksheet,
        *,
        header_row: int,
        max_row: int,
        max_column: int,
    ) -> None:
        for cell in worksheet[header_row]:
            if cell.column <= max_column:
                cell.font = Font(bold=True)
                cell.fill = self.HEADER_FILL

        worksheet.auto_filter.ref = (
            f"A{header_row}:{get_column_letter(max_column)}{max(max_row, header_row)}"
        )
        worksheet.freeze_panes = f"A{header_row + 1}"
        for column_index in range(1, max_column + 1):
            column_letter = get_column_letter(column_index)
            width = min(
                max(
                    len(self._display_text(worksheet.cell(row=row_index, column=column_index).value))
                    for row_index in range(1, min(max_row, 50) + 1)
                )
                + 2,
                36,
            )
            worksheet.column_dimensions[column_letter].width = max(width, 10)

    def _build_summary(
        self,
        main_sheet: SheetData,
        lookup_sheet: SheetData,
        config: IplexDualTableCompareConfig,
    ) -> dict[str, int]:
        lookup_map: dict[str, list[Any]] = {}
        for row in lookup_sheet.rows[config.lookup_header_row:]:
            key = self._normalize_key(self._read_cell(row, config.lookup_key_column))
            if key and key not in lookup_map:
                lookup_map[key] = row

        matched_count = 0
        main_row_count = 0
        four_digit_mismatch_count = 0
        two_digit_mismatch_count = 0
        for row in main_sheet.rows[config.main_header_row:]:
            key = self._normalize_key(self._read_cell(row, config.main_key_column))
            if not key:
                continue

            main_row_count += 1
            lookup_row = lookup_map.get(key)
            if lookup_row is None:
                continue
            matched_count += 1
            if self._rounded_difference(
                self._read_cell(row, config.four_digit_main_column),
                self._read_cell(lookup_row, config.four_digit_lookup_column),
                4,
            ) not in (None, Decimal("0.0000")):
                four_digit_mismatch_count += 1
            if self._rounded_difference(
                self._read_cell(row, config.two_digit_main_column),
                self._read_cell(lookup_row, config.two_digit_lookup_column),
                2,
            ) not in (None, Decimal("0.00")):
                two_digit_mismatch_count += 1

        lookup_row_count = max(lookup_sheet.max_row - config.lookup_header_row, 0)
        return {
            "main_row_count": main_row_count,
            "lookup_row_count": lookup_row_count,
            "matched_count": matched_count,
            "unmatched_count": max(main_row_count - matched_count, 0),
            "four_digit_mismatch_count": four_digit_mismatch_count,
            "two_digit_mismatch_count": two_digit_mismatch_count,
        }

    def _build_preview_rows(
        self,
        main_sheet: SheetData,
        lookup_sheet: SheetData,
        config: IplexDualTableCompareConfig,
    ) -> list[dict[str, Any]]:
        lookup_map: dict[str, list[Any]] = {}
        for row in lookup_sheet.rows[config.lookup_header_row:]:
            key = self._normalize_key(self._read_cell(row, config.lookup_key_column))
            if key and key not in lookup_map:
                lookup_map[key] = row

        preview_rows: list[dict[str, Any]] = []
        for row_index, row in enumerate(main_sheet.rows[config.main_header_row:], start=config.main_header_row + 1):
            key = self._normalize_key(self._read_cell(row, config.main_key_column))
            if not key:
                continue

            lookup_row = lookup_map.get(key)
            four_main_value = self._read_cell(row, config.four_digit_main_column)
            two_main_value = self._read_cell(row, config.two_digit_main_column)

            if lookup_row is None:
                preview_rows.append(
                    {
                        "row_number": row_index,
                        "key": key,
                        "status": "未匹配",
                        "four_digit": {
                            "main_value": self._format_decimal_value(four_main_value, 4),
                            "lookup_value": "#N/A",
                            "difference": "#N/A",
                        },
                        "two_digit": {
                            "main_value": self._format_decimal_value(two_main_value, 2),
                            "lookup_value": "#N/A",
                            "difference": "#N/A",
                        },
                    }
                )
                continue

            four_lookup_value = self._read_cell(lookup_row, config.four_digit_lookup_column)
            two_lookup_value = self._read_cell(lookup_row, config.two_digit_lookup_column)
            four_difference = self._rounded_difference(four_main_value, four_lookup_value, 4)
            two_difference = self._rounded_difference(two_main_value, two_lookup_value, 2)
            has_four_mismatch = four_difference not in (None, Decimal("0.0000"))
            has_two_mismatch = two_difference not in (None, Decimal("0.00"))
            if not has_four_mismatch and not has_two_mismatch:
                continue

            preview_rows.append(
                {
                    "row_number": row_index,
                    "key": key,
                    "status": "不一致",
                    "four_digit": {
                        "main_value": self._format_decimal_value(four_main_value, 4),
                        "lookup_value": self._format_decimal_value(four_lookup_value, 4),
                        "difference": self._format_decimal(four_difference, 4),
                    },
                    "two_digit": {
                        "main_value": self._format_decimal_value(two_main_value, 2),
                        "lookup_value": self._format_decimal_value(two_lookup_value, 2),
                        "difference": self._format_decimal(two_difference, 2),
                    },
                }
            )

        return preview_rows

    def _rounded_difference(self, left: Any, right: Any, digits: int) -> Decimal | None:
        left_decimal = self._to_decimal(left)
        right_decimal = self._to_decimal(right)
        if left_decimal is None or right_decimal is None:
            return None

        quant = Decimal("1").scaleb(-digits)
        return (left_decimal - right_decimal).quantize(quant, rounding=ROUND_HALF_UP)

    def _format_decimal_value(self, value: Any, digits: int) -> str:
        decimal_value = self._to_decimal(value)
        if decimal_value is None:
            return ""
        return self._format_decimal(decimal_value, digits)

    def _format_decimal(self, value: Decimal | None, digits: int) -> str:
        if value is None:
            return ""
        quant = Decimal("1").scaleb(-digits)
        return format(value.quantize(quant, rounding=ROUND_HALF_UP), f".{digits}f")

    def _to_decimal(self, value: Any) -> Decimal | None:
        if value in (None, ""):
            return None
        try:
            return Decimal(str(value))
        except (InvalidOperation, ValueError):
            return None

    def _read_cell(self, row_values: Sequence[Any], one_based_column: int) -> Any:
        index = one_based_column - 1
        return row_values[index] if 0 <= index < len(row_values) else None

    def _display_text(self, value: Any) -> str:
        if value is None:
            return ""
        return str(value).replace("\n", " ").strip()

    def _normalize_key(self, value: Any) -> str:
        if value is None:
            return ""
        if isinstance(value, float) and value.is_integer():
            return str(int(value))
        return str(value).strip()

    def _safe_sheet_title(self, title: str) -> str:
        return title[:31] or "Result"

from __future__ import annotations

import base64
import binascii
from io import BytesIO
import json
import logging
from datetime import datetime, time, timedelta, timezone
from typing import Any
from urllib.parse import quote
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side
from openpyxl.utils import get_column_letter
from openpyxl.worksheet.table import Table, TableStyleInfo
from fastapi import APIRouter, File, Form, HTTPException, Query, UploadFile
from fastapi.responses import StreamingResponse
from pydantic import BaseModel, Field

from utils.credential_crypto import decrypt_secret, encrypt_secret
from utils.minio_storage import (
    build_object_key,
    download_url_bytes,
    get_minio_bucket,
    get_object_response,
    put_object_bytes,
    sanitize_object_segment,
    sha256_bytes,
)
from utils.mysql_store import (
    count_automation_runs,
    create_automation_batch,
    create_automation_batch_attempt,
    create_automation_run,
    delete_automation_batch,
    delete_automation_credentials,
    delete_automation_run,
    delete_excel_template,
    get_automation_batch,
    get_automation_batch_attempt,
    get_automation_run_file,
    get_automation_credentials,
    get_automation_run,
    get_excel_template,
    insert_automation_run_file,
    list_automation_batches,
    list_automation_credentials,
    list_automation_run_files,
    list_automation_runs,
    list_excel_templates,
    update_automation_batch_attempt,
    update_automation_batch_checkpoint,
    update_automation_run,
    update_excel_template,
    upsert_automation_credentials,
    upsert_excel_template,
)


router = APIRouter(prefix="/automation", tags=["Automation Storage"])
logger = logging.getLogger(__name__)
BEIJING_TZ = timezone(timedelta(hours=8))

INFOR_NEXUS_SHARED_CREDENTIAL_IDS = (
    "shipping-automation",
    "shipping-automation-demo",
    "shipping-automation-2",
    "xinlongtai-shipping-automation",
    "po-auto-download",
    "packing-list-auto-download",
    "released-bulk-automation",
)

MICROSOFT_SHARED_CREDENTIAL_IDS = (
    "microsoft-login-n8n",
    "ticket-owner-statistics",
)

TICKET_OWNER_AUTOMATION_ID = "ticket-owner-statistics"
TICKET_OWNER_BUCKET_KEY = "ticket_owner_statistics"
TICKET_OWNER_OBJECT_ROOT = "ticket-owner-statistics"
TICKET_OWNER_COLUMNS = (
    "Case Number",
    "Task Type",
    "Request",
    "PO Number",
    "Working Number",
    "Factory",
    "Merch",
)
TICKET_OWNER_EXCEL_NAME = "Ticket ownership.xlsx"
EXCEL_CONTENT_TYPE = "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet"
PACKING_LIST_AUTOMATION_ID = "packing-list-auto-download"
PACKING_LIST_BUCKET_KEY = "packing_list_auto_download"
PACKING_LIST_OBJECT_ROOT = "packing-list-auto-download"


class CredentialPayload(BaseModel):
    username: str = Field(min_length=1)
    password: str = Field(min_length=1)
    accountKey: str = "default"


class RunFilePayload(BaseModel):
    url: str
    fileName: str = ""
    fileRole: str = "result_file"
    contentType: str = ""
    contentBase64: str = ""


class RunUpdatePayload(BaseModel):
    status: str
    message: str = ""
    result: Any = None
    resultFiles: list[RunFilePayload] = Field(default_factory=list)


class PackingListAttemptPayload(BaseModel):
    runId: str = ""
    mode: str = "new"


class PackingListCheckpointFilePayload(BaseModel):
    fileName: str
    fileRole: str = "packing_list_pdf"
    contentType: str = "application/pdf"
    contentBase64: str


class PackingListCheckpointPayload(BaseModel):
    runId: str = ""
    attemptId: str = ""
    mode: str = "continue"
    status: str = "running"
    message: str = ""
    checkpoint: dict[str, Any] = Field(default_factory=dict)
    groupResult: dict[str, Any] | None = None
    result: dict[str, Any] | None = None
    files: list[PackingListCheckpointFilePayload] = Field(default_factory=list)


class PackingListInterruptPayload(BaseModel):
    runId: str = ""
    attemptId: str = ""
    message: str = "用户已停止本机执行器，当前批次已中断，可从断点继续。"


class TicketOwnerAttemptPayload(BaseModel):
    runId: str = ""
    mode: str = "new"


class TicketOwnerCheckpointPayload(BaseModel):
    runId: str = ""
    attemptId: str = ""
    mode: str = "continue"
    status: str = "running"
    message: str = ""
    checkpoint: dict[str, Any] = Field(default_factory=dict)
    itemResult: dict[str, Any] | None = None
    result: dict[str, Any] | None = None


class TicketOwnerInterruptPayload(BaseModel):
    runId: str = ""
    attemptId: str = ""
    message: str = "用户已停止本机执行器，当前 Ticket ownership 批次已中断，可从断点继续。"


class TemplateUpdatePayload(BaseModel):
    moduleId: str | None = None
    templateKey: str | None = None
    displayName: str | None = None
    isActive: bool | None = None


@router.get("/credentials/{automation_id}/accounts")
def list_credentials(automation_id: str) -> dict[str, Any]:
    accounts: list[dict[str, Any]] = []
    seen_account_keys: set[str] = set()
    for source_automation_id in _credential_lookup_ids(automation_id):
        for row in list_automation_credentials(source_automation_id):
            account_key = _normalize_account_key(row.get("account_key"))
            if account_key in seen_account_keys:
                continue
            seen_account_keys.add(account_key)
            accounts.append(_credential_account_payload(automation_id, row, source_automation_id))

    return {
        "ok": True,
        "automationId": automation_id,
        "accounts": accounts,
    }


@router.get("/credentials/{automation_id}")
def read_credentials(automation_id: str, accountKey: str = Query("default")) -> dict[str, Any]:
    account_key = _normalize_account_key(accountKey)
    row, source_automation_id = _get_credentials_with_alias(automation_id, account_key)
    return _credential_public_payload(automation_id, account_key, row, source_automation_id)


@router.put("/credentials/{automation_id}")
def save_credentials(automation_id: str, payload: CredentialPayload) -> dict[str, Any]:
    username = payload.username.strip()
    password = payload.password
    account_key = _normalize_account_key(payload.accountKey)
    if not username or not password:
        raise HTTPException(status_code=400, detail="请填写当前网站登录账号密码。")

    row = upsert_automation_credentials(
        automation_id,
        account_key,
        username,
        encrypt_secret(password),
    )
    return _credential_public_payload(automation_id, account_key, row, automation_id)


@router.delete("/credentials/{automation_id}")
def clear_credentials(automation_id: str, accountKey: str = Query("default")) -> dict[str, Any]:
    account_key = _normalize_account_key(accountKey)
    row, source_automation_id = _get_credentials_with_alias(automation_id, account_key)
    delete_automation_credentials(source_automation_id if row else automation_id, account_key)
    return _credential_public_payload(automation_id, account_key, None, source_automation_id)


@router.post("/credentials/{automation_id}/resolve")
def resolve_credentials(automation_id: str, accountKey: str = Query("default")) -> dict[str, Any]:
    account_key = _normalize_account_key(accountKey)
    resolved_row, source_automation_id, password, saw_stored_credentials = _resolve_credentials_with_aliases(
        automation_id,
        account_key,
    )
    if resolved_row:
        return {
            "ok": True,
            "automationId": automation_id,
            "sourceAutomationId": source_automation_id,
            "accountKey": account_key,
            "username": resolved_row["username"],
            "password": password,
        }
    if not saw_stored_credentials:
        raise HTTPException(status_code=404, detail="未保存当前网站登录账号密码。请先填写并保存。")
    raise HTTPException(
        status_code=409,
        detail=(
            "已找到该 Infor Nexus 账号记录，但本机后端无法解密密码。"
            "请确认本机 tms-backend/config/settings.yaml 中的 security.credential_key "
            "与保存该密码时使用的后端一致；如果当前连接的是远程数据库，请使用服务器同一套凭据密钥。"
        ),
    )


@router.get("/templates")
def read_templates(
    moduleId: str | None = Query(None),
    includeInactive: bool = Query(False),
    limit: int = Query(500, ge=1, le=1000),
) -> dict[str, Any]:
    templates = [
        _template_payload(row)
        for row in list_excel_templates(moduleId, include_inactive=includeInactive, limit=limit)
    ]
    return {
        "ok": True,
        "moduleId": moduleId or "",
        "templates": templates,
    }


@router.post("/templates")
async def upload_template(
    module_id: str = Form(...),
    template_key: str = Form("default"),
    display_name: str = Form("Excel 模板"),
    file: UploadFile = File(...),
) -> dict[str, Any]:
    safe_filename = sanitize_object_segment(file.filename or "template.xlsx")
    if not safe_filename.lower().endswith((".xlsx", ".xls", ".xlsm")):
        raise HTTPException(status_code=400, detail="Template file must be an Excel workbook.")

    content = await file.read()
    if not content:
        raise HTTPException(status_code=400, detail="Template file is empty.")

    bucket = get_minio_bucket("templates")
    object_key = build_object_key("templates", module_id, template_key, safe_filename)
    storage_record = put_object_bytes(
        bucket=bucket,
        object_key=object_key,
        content=content,
        content_type=file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
    )
    row = upsert_excel_template({
        "module_id": module_id,
        "template_key": template_key or "default",
        "display_name": display_name or safe_filename,
        "bucket": storage_record["bucket"],
        "object_key": storage_record["object_key"],
        "original_filename": safe_filename,
        "content_type": file.content_type or "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        "file_size": storage_record["file_size"],
        "sha256": storage_record["sha256"],
    })
    return {
        "ok": True,
        "template": _template_payload(row),
    }


@router.patch("/templates/{template_id}")
def update_template(template_id: int, payload: TemplateUpdatePayload) -> dict[str, Any]:
    if not get_excel_template(template_id, include_inactive=True):
        raise HTTPException(status_code=404, detail="模板记录不存在。")

    updates: dict[str, Any] = {}
    if payload.moduleId is not None:
        updates["module_id"] = payload.moduleId.strip()
    if payload.templateKey is not None:
        updates["template_key"] = payload.templateKey.strip() or "default"
    if payload.displayName is not None:
        updates["display_name"] = payload.displayName.strip() or "Excel 模板"
    if payload.isActive is not None:
        updates["is_active"] = 1 if payload.isActive else 0

    row = update_excel_template(template_id, updates)
    if not row:
        raise HTTPException(status_code=404, detail="模板记录不存在。")
    return {
        "ok": True,
        "template": _template_payload(row),
    }


@router.delete("/templates/{template_id}")
def delete_template(template_id: int) -> dict[str, Any]:
    if not delete_excel_template(template_id):
        raise HTTPException(status_code=404, detail="模板记录不存在或已经删除。")
    return {
        "ok": True,
        "templateId": template_id,
    }


@router.get("/templates/{template_id}/download")
def download_template(template_id: int):
    row = get_excel_template(template_id)
    if not row:
        raise HTTPException(status_code=404, detail="当前模块的 Excel 模板未配置，请联系管理员上传模板。")

    try:
        response = get_object_response(row["bucket"], row["object_key"])
    except Exception as exc:
        _raise_template_storage_error(template_id, row, exc)
    filename = row.get("original_filename") or f"template-{template_id}.xlsx"
    headers = {
        "Content-Disposition": _attachment_content_disposition(filename),
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        response.stream(32 * 1024),
        media_type=row.get("content_type") or "application/octet-stream",
        headers=headers,
        background=None,
    )


@router.post("/runs")
async def create_run(
    automation_id: str = Form(...),
    module_id: str = Form(""),
    run_name: str = Form(""),
    message: str = Form(""),
    source_file: UploadFile | None = File(None),
) -> dict[str, Any]:
    run_id = f"{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:10]}"
    row = create_automation_run({
        "run_id": run_id,
        "automation_id": automation_id,
        "module_id": module_id or automation_id,
        "run_name": run_name or automation_id,
        "status": "running",
        "message": message,
        "started_at": datetime.utcnow(),
    })

    files: list[dict[str, Any]] = []
    warnings: list[str] = []
    if source_file is not None and source_file.filename:
        stored_file = await _try_store_upload_file(
            run_id=run_id,
            automation_id=automation_id,
            upload=source_file,
            file_role="source_excel",
            bucket_key="run_files",
            warnings=warnings,
        )
        if stored_file:
            files.append(stored_file)

    return {
        "ok": True,
        "run": _run_payload(row),
        "files": [_run_file_payload(file_row) for file_row in files],
        "warnings": warnings,
    }


@router.post("/packing-list-auto-download/batches")
async def create_packing_list_batch(
    run_id: str = Form(""),
    batch_name: str = Form(""),
    source_file: UploadFile = File(...),
) -> dict[str, Any]:
    safe_filename = sanitize_object_segment(source_file.filename or "packing-list.xlsx")
    if not safe_filename.lower().endswith((".xlsx", ".xls")):
        raise HTTPException(status_code=400, detail="自动下载箱单批次只支持 Excel 文件。")

    content = await source_file.read()
    if not content:
        raise HTTPException(status_code=400, detail="自动下载箱单批次源 Excel 为空。")

    now = datetime.utcnow()
    batch_id = f"plad-{now.strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}"
    date_folder = _packing_list_date_folder(now)
    bucket = get_minio_bucket(PACKING_LIST_BUCKET_KEY)
    object_prefix = f"{PACKING_LIST_OBJECT_ROOT}/{date_folder}/{batch_id}"
    object_key = f"{object_prefix}/source/{safe_filename}"
    storage_record = put_object_bytes(
        bucket=bucket,
        object_key=object_key,
        content=content,
        content_type=source_file.content_type or EXCEL_CONTENT_TYPE,
    )
    source_file_payload = {
        "bucket": bucket,
        "objectKey": object_key,
        "originalFilename": safe_filename,
        "contentType": source_file.content_type or EXCEL_CONTENT_TYPE,
        "fileSize": storage_record["file_size"],
        "sha256": storage_record["sha256"],
        "downloadPath": f"/api/automation/packing-list-auto-download/batches/{batch_id}/source/download",
    }
    row = create_automation_batch({
        "batch_id": batch_id,
        "automation_id": PACKING_LIST_AUTOMATION_ID,
        "module_id": PACKING_LIST_AUTOMATION_ID,
        "run_id": run_id,
        "batch_name": batch_name or safe_filename,
        "source_file_name": safe_filename,
        "source_file_sha256": storage_record["sha256"],
        "source_file": source_file_payload,
        "status": "pending",
        "message": "批次已创建，等待执行。",
        "total_count": 0,
        "completed_count": 0,
        "failed_count": 0,
        "pending_count": 0,
        "checkpoint": {
            "batchId": batch_id,
            "automationId": PACKING_LIST_AUTOMATION_ID,
            "status": "pending",
            "items": [],
            "groupResults": [],
            "createdAt": now.isoformat(),
            "updatedAt": now.isoformat(),
        },
        "bucket": bucket,
        "object_prefix": object_prefix,
    })
    file_row = None
    if run_id:
        file_row = insert_automation_run_file({
            "run_id": run_id,
            "file_role": "source_excel",
            "bucket": bucket,
            "object_key": object_key,
            "original_filename": safe_filename,
            "content_type": source_file.content_type or EXCEL_CONTENT_TYPE,
            "file_size": storage_record["file_size"],
            "sha256": storage_record["sha256"],
        })
    return {
        "ok": True,
        "batch": _packing_list_batch_payload(row),
        "sourceFile": source_file_payload,
        "file": _run_file_payload(file_row) if file_row else None,
    }


@router.get("/packing-list-auto-download/batches")
def list_packing_list_batches(limit: int = Query(10, ge=1, le=50)) -> dict[str, Any]:
    rows = list_automation_batches(PACKING_LIST_AUTOMATION_ID, limit=limit)
    return {
        "ok": True,
        "batches": [_packing_list_batch_payload(row) for row in rows],
    }


@router.get("/packing-list-auto-download/batches/latest")
def read_latest_packing_list_batch() -> dict[str, Any]:
    rows = list_automation_batches(PACKING_LIST_AUTOMATION_ID, limit=20, include_payloads=False)
    resumable_row = next((row for row in rows if _is_resumable_batch(row)), None)
    if resumable_row:
        resumable_row = get_automation_batch(resumable_row["batch_id"]) or resumable_row
    return {
        "ok": True,
        "batch": _packing_list_batch_payload(resumable_row) if resumable_row else None,
    }


@router.get("/packing-list-auto-download/batches/{batch_id}")
def read_packing_list_batch(batch_id: str) -> dict[str, Any]:
    row = get_automation_batch(batch_id)
    if not row or row.get("automation_id") != PACKING_LIST_AUTOMATION_ID:
        raise HTTPException(status_code=404, detail="自动下载箱单批次不存在。")
    return {
        "ok": True,
        "batch": _packing_list_batch_payload(row),
    }


@router.get("/packing-list-auto-download/batches/{batch_id}/source/download")
def download_packing_list_batch_source(batch_id: str):
    row = get_automation_batch(batch_id)
    if not row or row.get("automation_id") != PACKING_LIST_AUTOMATION_ID:
        raise HTTPException(status_code=404, detail="自动下载箱单批次不存在。")
    source_file = _safe_json_loads(row.get("source_file_json")) or {}
    bucket = source_file.get("bucket") or row.get("bucket")
    object_key = source_file.get("objectKey") or source_file.get("object_key")
    if not bucket or not object_key:
        raise HTTPException(status_code=404, detail="该批次没有可下载的源 Excel。")
    try:
        response = get_object_response(bucket, object_key)
    except Exception as exc:
        raise HTTPException(status_code=503, detail="批次源 Excel 暂时无法从 MinIO 下载。") from exc
    filename = source_file.get("originalFilename") or row.get("source_file_name") or "packing-list.xlsx"
    headers = {
        "Content-Disposition": _attachment_content_disposition(filename),
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        response.stream(32 * 1024),
        media_type=source_file.get("contentType") or EXCEL_CONTENT_TYPE,
        headers=headers,
        background=None,
    )


@router.post("/packing-list-auto-download/batches/{batch_id}/attempts")
def create_packing_list_batch_attempt(batch_id: str, payload: PackingListAttemptPayload) -> dict[str, Any]:
    row = get_automation_batch(batch_id)
    if not row or row.get("automation_id") != PACKING_LIST_AUTOMATION_ID:
        raise HTTPException(status_code=404, detail="自动下载箱单批次不存在。")
    attempt_id = f"pla-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}"
    attempt = create_automation_batch_attempt({
        "attempt_id": attempt_id,
        "batch_id": batch_id,
        "run_id": payload.runId,
        "mode": _normalize_packing_list_resume_mode(payload.mode),
        "status": "running",
        "message": "attempt started",
        "started_at": datetime.utcnow(),
    })
    return {
        "ok": True,
        "attempt": _packing_list_attempt_payload(attempt),
        "batch": _packing_list_batch_payload(row),
    }


@router.post("/packing-list-auto-download/batches/{batch_id}/checkpoint")
def write_packing_list_checkpoint(batch_id: str, payload: PackingListCheckpointPayload) -> dict[str, Any]:
    row = get_automation_batch(batch_id)
    if not row or row.get("automation_id") != PACKING_LIST_AUTOMATION_ID:
        raise HTTPException(status_code=404, detail="自动下载箱单批次不存在。")

    attempt_id = payload.attemptId.strip()
    if attempt_id and not get_automation_batch_attempt(attempt_id):
        create_automation_batch_attempt({
            "attempt_id": attempt_id,
            "batch_id": batch_id,
            "run_id": payload.runId,
            "mode": _normalize_packing_list_resume_mode(payload.mode),
            "status": "running",
            "message": "attempt started from checkpoint write",
            "started_at": datetime.utcnow(),
        })

    stored_files = [
        _store_packing_list_checkpoint_file(row, payload, file_payload)
        for file_payload in payload.files
    ]
    checkpoint = _merge_packing_list_checkpoint(
        existing=_safe_json_loads(row.get("checkpoint_json")) or {},
        payload=payload,
        stored_files=stored_files,
    )
    counts = _packing_list_checkpoint_counts(checkpoint)
    status = _normalize_packing_list_batch_status(payload.status, counts)
    updated_row = update_automation_batch_checkpoint(batch_id, {
        "run_id": payload.runId,
        "status": status,
        "message": payload.message or checkpoint.get("message", ""),
        "total_count": counts["total"],
        "completed_count": counts["completed"],
        "failed_count": counts["failed"],
        "pending_count": counts["pending"],
        "checkpoint": checkpoint,
    })
    if attempt_id:
        update_automation_batch_attempt(attempt_id, {
            "status": status if status in {"success", "failed", "partial", "interrupted"} else "running",
            "message": payload.message or checkpoint.get("message", ""),
            "result": payload.result or {"checkpoint": checkpoint},
            "finished_at": datetime.utcnow() if status in {"success", "failed", "partial", "interrupted"} else None,
        })
    return {
        "ok": True,
        "batch": _packing_list_batch_payload(updated_row or row),
        "storedFiles": [_run_file_payload(file_row) for file_row in stored_files],
    }


@router.post("/packing-list-auto-download/batches/{batch_id}/interrupt")
def interrupt_packing_list_batch(batch_id: str, payload: PackingListInterruptPayload) -> dict[str, Any]:
    row = get_automation_batch(batch_id)
    if not row or row.get("automation_id") != PACKING_LIST_AUTOMATION_ID:
        raise HTTPException(status_code=404, detail="自动下载箱单批次不存在。")

    current_status = str(row.get("status") or "").lower()
    if current_status in {"success", "failed", "partial", "interrupted"}:
        return {
            "ok": True,
            "batch": _packing_list_batch_payload(row),
            "changed": False,
        }

    existing_checkpoint = _safe_json_loads(row.get("checkpoint_json")) or {}
    now = datetime.utcnow().isoformat()
    message = payload.message or "用户已停止本机执行器，当前批次已中断，可从断点继续。"
    checkpoint = {
        **existing_checkpoint,
        "status": "interrupted",
        "message": message,
        "interruptedAt": now,
        "updatedAt": now,
    }
    counts = _packing_list_checkpoint_counts(checkpoint)
    updated_row = update_automation_batch_checkpoint(batch_id, {
        "run_id": payload.runId or row.get("run_id", ""),
        "status": "interrupted",
        "message": message,
        "total_count": counts["total"] or int(row.get("total_count") or 0),
        "completed_count": counts["completed"] or int(row.get("completed_count") or 0),
        "failed_count": counts["failed"] or int(row.get("failed_count") or 0),
        "pending_count": counts["pending"] or int(row.get("pending_count") or 0),
        "checkpoint": checkpoint,
    })
    attempt_id = payload.attemptId.strip() or str(checkpoint.get("attemptId") or "").strip()
    if attempt_id:
        update_automation_batch_attempt(attempt_id, {
            "status": "interrupted",
            "message": message,
            "result": {"checkpoint": checkpoint},
            "finished_at": datetime.utcnow(),
        })
    return {
        "ok": True,
        "batch": _packing_list_batch_payload(updated_row or row),
        "changed": True,
    }


@router.post("/ticket-owner-statistics/batches/{batch_id}/attempts")
def create_ticket_owner_batch_attempt(batch_id: str, payload: TicketOwnerAttemptPayload) -> dict[str, Any]:
    row = get_automation_batch(batch_id)
    if not row or row.get("automation_id") != TICKET_OWNER_AUTOMATION_ID:
        raise HTTPException(status_code=404, detail="Ticket ownership 批次不存在。")
    attempt_id = f"tos-attempt-{datetime.utcnow().strftime('%Y%m%d%H%M%S')}-{uuid4().hex[:8]}"
    attempt = create_automation_batch_attempt({
        "attempt_id": attempt_id,
        "batch_id": batch_id,
        "run_id": payload.runId,
        "mode": _normalize_packing_list_resume_mode(payload.mode),
        "status": "running",
        "message": "attempt started",
        "started_at": datetime.utcnow(),
    })
    return {
        "ok": True,
        "attempt": _packing_list_attempt_payload(attempt),
        "batch": _packing_list_batch_payload(row),
    }


@router.post("/ticket-owner-statistics/batches/{batch_id}/checkpoint")
def write_ticket_owner_checkpoint(batch_id: str, payload: TicketOwnerCheckpointPayload) -> dict[str, Any]:
    row = get_automation_batch(batch_id)
    if not row or row.get("automation_id") != TICKET_OWNER_AUTOMATION_ID:
        raise HTTPException(status_code=404, detail="Ticket ownership 批次不存在。")

    attempt_id = payload.attemptId.strip()
    if attempt_id and not get_automation_batch_attempt(attempt_id):
        create_automation_batch_attempt({
            "attempt_id": attempt_id,
            "batch_id": batch_id,
            "run_id": payload.runId,
            "mode": _normalize_packing_list_resume_mode(payload.mode),
            "status": "running",
            "message": "attempt started from checkpoint write",
            "started_at": datetime.utcnow(),
        })

    checkpoint = _merge_ticket_owner_checkpoint(
        batch_id=batch_id,
        existing=_safe_json_loads(row.get("checkpoint_json")) or {},
        payload=payload,
        stored_files=[],
    )
    stored_files = _store_ticket_owner_checkpoint_result_files(payload.runId, checkpoint, payload.result)
    if stored_files:
        checkpoint = _merge_ticket_owner_checkpoint(
            batch_id=batch_id,
            existing=checkpoint,
            payload=payload,
            stored_files=stored_files,
        )
    counts = _ticket_owner_checkpoint_counts(checkpoint)
    status = _normalize_ticket_owner_batch_status(payload.status, counts)
    checkpoint["status"] = status
    updated_row = update_automation_batch_checkpoint(batch_id, {
        "run_id": payload.runId,
        "status": status,
        "message": payload.message or checkpoint.get("message", ""),
        "total_count": counts["total"],
        "completed_count": counts["completed"],
        "failed_count": counts["failed"],
        "pending_count": counts["pending"],
        "checkpoint": checkpoint,
    })
    if attempt_id:
        update_automation_batch_attempt(attempt_id, {
            "status": status if status in {"success", "failed", "partial", "interrupted"} else "running",
            "message": payload.message or checkpoint.get("message", ""),
            "result": payload.result or {"checkpoint": checkpoint},
            "finished_at": datetime.utcnow() if status in {"success", "failed", "partial", "interrupted"} else None,
        })
    return {
        "ok": True,
        "batch": _packing_list_batch_payload(updated_row or row),
        "storedFiles": [_run_file_payload(file_row) for file_row in stored_files],
    }


@router.post("/ticket-owner-statistics/batches/{batch_id}/interrupt")
def interrupt_ticket_owner_batch(batch_id: str, payload: TicketOwnerInterruptPayload) -> dict[str, Any]:
    row = get_automation_batch(batch_id)
    if not row or row.get("automation_id") != TICKET_OWNER_AUTOMATION_ID:
        raise HTTPException(status_code=404, detail="Ticket ownership 批次不存在。")

    current_status = str(row.get("status") or "").lower()
    if current_status in {"success", "failed", "partial", "interrupted"}:
        return {
            "ok": True,
            "batch": _packing_list_batch_payload(row),
            "changed": False,
        }

    existing_checkpoint = _safe_json_loads(row.get("checkpoint_json")) or {}
    now = datetime.utcnow().isoformat()
    message = payload.message or "用户已停止本机执行器，当前 Ticket ownership 批次已中断，可从断点继续。"
    checkpoint = {
        **existing_checkpoint,
        "status": "interrupted",
        "message": message,
        "interruptedAt": now,
        "updatedAt": now,
    }
    counts = _ticket_owner_checkpoint_counts(checkpoint)
    updated_row = update_automation_batch_checkpoint(batch_id, {
        "run_id": payload.runId or row.get("run_id", ""),
        "status": "interrupted",
        "message": message,
        "total_count": counts["total"] or int(row.get("total_count") or 0),
        "completed_count": counts["completed"] or int(row.get("completed_count") or 0),
        "failed_count": counts["failed"] or int(row.get("failed_count") or 0),
        "pending_count": counts["pending"] or int(row.get("pending_count") or 0),
        "checkpoint": checkpoint,
    })
    attempt_id = payload.attemptId.strip() or str(checkpoint.get("attemptId") or "").strip()
    if attempt_id:
        update_automation_batch_attempt(attempt_id, {
            "status": "interrupted",
            "message": message,
            "result": {"checkpoint": checkpoint},
            "finished_at": datetime.utcnow(),
        })
    return {
        "ok": True,
        "batch": _packing_list_batch_payload(updated_row or row),
        "changed": True,
    }


@router.delete("/ticket-owner-statistics/batches/{batch_id}")
def delete_ticket_owner_batch(batch_id: str) -> dict[str, Any]:
    row = get_automation_batch(batch_id)
    if not row or row.get("automation_id") != TICKET_OWNER_AUTOMATION_ID:
        raise HTTPException(status_code=404, detail="Ticket ownership 批次不存在。")
    deleted = delete_automation_batch(batch_id)
    return {
        "ok": True,
        "batchId": batch_id,
        "deleted": deleted,
    }


@router.patch("/runs/{run_id}")
async def finish_run(run_id: str, payload: RunUpdatePayload) -> dict[str, Any]:
    if not get_automation_run(run_id):
        raise HTTPException(status_code=404, detail="Automation run was not found.")

    row = update_automation_run(
        run_id,
        payload.status,
        payload.message,
        payload.result,
        datetime.utcnow(),
    )
    files: list[dict[str, Any]] = []
    warnings: list[str] = []
    server_generated_roles: set[str] = set()
    if payload.result is not None:
        result_file = _try_store_result_json(run_id, row["automation_id"], payload.result, warnings)
        if result_file:
            files.append(result_file)
        ticket_owner_file = _try_store_ticket_owner_excel(
            run_id,
            row["automation_id"],
            payload.result,
            warnings,
        )
        if ticket_owner_file:
            files.append(ticket_owner_file)
            server_generated_roles.add("result_excel")

    for file_payload in payload.resultFiles:
        if file_payload.fileRole in server_generated_roles:
            continue
        result_file = _try_store_remote_result_file(run_id, row["automation_id"], file_payload, warnings)
        if result_file:
            files.append(result_file)

    return {
        "ok": True,
        "run": _run_payload(row),
        "files": [_run_file_payload(file_row) for file_row in files],
        "warnings": warnings,
    }


@router.delete("/runs/{run_id}")
def delete_run(run_id: str) -> dict[str, Any]:
    if not get_automation_run(run_id):
        raise HTTPException(status_code=404, detail="Automation run was not found.")
    deleted = delete_automation_run(run_id)
    return {
        "ok": True,
        **deleted,
    }


@router.get("/runs")
def read_runs(
    automationId: str | None = Query(None),
    moduleId: str | None = Query(None),
    status: str | None = Query(None),
    keyword: str | None = Query(None),
    dateFrom: str | None = Query(None),
    dateTo: str | None = Query(None),
    page: int = Query(1, ge=1),
    pageSize: int = Query(30, ge=1, le=100),
    limit: int | None = Query(None, ge=1, le=100),
) -> dict[str, Any]:
    try:
        started_from, started_to = _parse_beijing_date_range(dateFrom, dateTo)
        effective_page_size = limit or pageSize
        offset = (page - 1) * effective_page_size
        total = count_automation_runs(
            automationId,
            module_id=moduleId,
            status=status,
            keyword=keyword,
            started_from=started_from,
            started_to=started_to,
        )
        runs = [
            _run_payload(row)
            for row in list_automation_runs(
                automationId,
                effective_page_size,
                module_id=moduleId,
                status=status,
                keyword=keyword,
                started_from=started_from,
                started_to=started_to,
                offset=offset,
            )
        ]
    except HTTPException:
        raise
    except Exception as exc:
        logger.exception("Failed to read automation run records.")
        raise HTTPException(
            status_code=503,
            detail="无法读取自动化执行记录：远程 MySQL 数据库暂时不可用或连接被服务器断开。请确认 172.16.48.208 的 MySQL 服务正常后重试。",
        ) from exc
    return {
        "ok": True,
        "runs": runs,
        "pagination": {
            "page": page,
            "pageSize": effective_page_size,
            "total": total,
        },
    }


@router.get("/runs/{run_id}")
def read_run_detail(run_id: str) -> dict[str, Any]:
    row = get_automation_run(run_id)
    if not row:
        raise HTTPException(status_code=404, detail="Automation run was not found.")
    return {
        "ok": True,
        "run": _run_payload(row),
        "files": [_run_file_payload(file_row) for file_row in list_automation_run_files(run_id)],
    }


@router.get("/runs/{run_id}/files")
def read_run_files(run_id: str) -> dict[str, Any]:
    if not get_automation_run(run_id):
        raise HTTPException(status_code=404, detail="Automation run was not found.")
    return {
        "ok": True,
        "runId": run_id,
        "files": [_run_file_payload(row) for row in list_automation_run_files(run_id)],
    }


@router.get("/run-files/{file_id}/download")
def download_run_file(file_id: int):
    row = get_automation_run_file(file_id)
    if not row:
        raise HTTPException(status_code=404, detail="执行文件不存在。")

    try:
        response = get_object_response(row["bucket"], row["object_key"])
    except Exception as exc:
        logger.exception(
            "Failed to download automation run file: file_id=%s bucket=%s object_key=%s",
            file_id,
            row.get("bucket"),
            row.get("object_key"),
        )
        raise HTTPException(
            status_code=503,
            detail="执行文件暂时无法下载，请联系管理员检查 MinIO 存储连接。",
        ) from exc

    filename = row.get("original_filename") or f"automation-file-{file_id}"
    headers = {
        "Content-Disposition": _attachment_content_disposition(filename),
        "Cache-Control": "no-store",
    }
    return StreamingResponse(
        response.stream(32 * 1024),
        media_type=row.get("content_type") or "application/octet-stream",
        headers=headers,
        background=None,
    )


def _credential_public_payload(
    automation_id: str,
    account_key: str,
    row: dict[str, Any] | None,
    source_automation_id: str | None = None,
) -> dict[str, Any]:
    return {
        "ok": True,
        "automationId": automation_id,
        "sourceAutomationId": source_automation_id or automation_id,
        "accountKey": account_key,
        "hasStoredCredentials": bool(row),
        "username": row["username"] if row else "",
    }


def _credential_account_payload(
    automation_id: str,
    row: dict[str, Any],
    source_automation_id: str,
) -> dict[str, Any]:
    return {
        "automationId": automation_id,
        "sourceAutomationId": source_automation_id,
        "accountKey": _normalize_account_key(row.get("account_key")),
        "hasStoredCredentials": True,
        "username": row.get("username", ""),
        "createdAt": _format_datetime(row.get("created_at")),
        "updatedAt": _format_datetime(row.get("updated_at")),
    }


def _get_credentials_with_alias(automation_id: str, account_key: str) -> tuple[dict[str, Any] | None, str]:
    account_key = _normalize_account_key(account_key)
    for candidate_id in _credential_lookup_ids(automation_id):
        row = get_automation_credentials(candidate_id, account_key)
        if row:
            return row, candidate_id
    return None, automation_id


def _resolve_credentials_with_aliases(
    automation_id: str,
    account_key: str,
) -> tuple[dict[str, Any] | None, str, str, bool]:
    account_key = _normalize_account_key(account_key)
    saw_stored_credentials = False
    for candidate_id in _credential_lookup_ids(automation_id):
        row = get_automation_credentials(candidate_id, account_key)
        if not row:
            continue
        saw_stored_credentials = True
        try:
            password = decrypt_secret(row["password_ciphertext"])
        except ValueError:
            logger.warning(
                "Automation credential decrypt failed: automation_id=%s source_automation_id=%s account_key=%s",
                automation_id,
                candidate_id,
                account_key,
            )
            continue
        return row, candidate_id, password, saw_stored_credentials
    return None, automation_id, "", saw_stored_credentials


def _normalize_account_key(value: Any) -> str:
    key = str(value or "default").strip()
    return (key or "default")[:120]


def _credential_lookup_ids(automation_id: str) -> list[str]:
    return [automation_id]


def _template_payload(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": row["id"],
        "moduleId": row["module_id"],
        "templateKey": row["template_key"],
        "displayName": row["display_name"],
        "originalFilename": row.get("original_filename", ""),
        "contentType": row.get("content_type", ""),
        "fileSize": row.get("file_size", 0),
        "sha256": row.get("sha256", ""),
        "isActive": bool(row.get("is_active", 1)),
        "downloadPath": f"/api/automation/templates/{row['id']}/download",
        "createdAt": _format_datetime(row.get("created_at")),
        "updatedAt": _format_datetime(row.get("updated_at")),
    }


def _attachment_content_disposition(filename: str) -> str:
    ascii_filename = filename.encode("ascii", errors="ignore").decode("ascii").strip()
    fallback_filename = ascii_filename or "template.xlsx"
    quoted_filename = quote(filename, safe="")
    return f'attachment; filename="{fallback_filename}"; filename*=UTF-8\'\'{quoted_filename}'


def _run_payload(row: dict[str, Any]) -> dict[str, Any]:
    started_at = row.get("started_at")
    finished_at = row.get("finished_at")
    created_at = row.get("created_at")
    updated_at = row.get("updated_at")
    return {
        "runId": row["run_id"],
        "automationId": row["automation_id"],
        "moduleId": row["module_id"],
        "runName": row.get("run_name", ""),
        "status": row.get("status", ""),
        "message": row.get("message", ""),
        "result": _safe_json_loads(row.get("result_json")),
        "startedAt": _format_datetime(started_at),
        "finishedAt": _format_datetime(finished_at),
        "createdAt": _format_datetime(created_at),
        "updatedAt": _format_datetime(updated_at),
        "startedAtBeijing": _format_beijing_datetime(started_at),
        "finishedAtBeijing": _format_beijing_datetime(finished_at),
        "createdAtBeijing": _format_beijing_datetime(created_at),
        "updatedAtBeijing": _format_beijing_datetime(updated_at),
        "durationSeconds": _duration_seconds(started_at, finished_at),
    }


def _run_file_payload(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": row["id"],
        "runId": row["run_id"],
        "fileRole": row["file_role"],
        "bucket": row["bucket"],
        "objectKey": row["object_key"],
        "originalFilename": row.get("original_filename", ""),
        "contentType": row.get("content_type", ""),
        "fileSize": row.get("file_size", 0),
        "sha256": row.get("sha256", ""),
        "createdAt": _format_datetime(row.get("created_at")),
        "downloadPath": f"/api/automation/run-files/{row['id']}/download",
    }


def _packing_list_batch_payload(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if not row:
        return None
    checkpoint = _safe_json_loads(row.get("checkpoint_json")) or {}
    source_file = _safe_json_loads(row.get("source_file_json")) or {}
    return {
        "batchId": row["batch_id"],
        "automationId": row.get("automation_id", PACKING_LIST_AUTOMATION_ID),
        "moduleId": row.get("module_id", PACKING_LIST_AUTOMATION_ID),
        "runId": row.get("run_id", ""),
        "batchName": row.get("batch_name", ""),
        "status": row.get("status", ""),
        "message": row.get("message", ""),
        "sourceFileName": row.get("source_file_name", ""),
        "sourceFileSha256": row.get("source_file_sha256", ""),
        "sourceFile": source_file,
        "totalCount": int(row.get("total_count") or 0),
        "completedCount": int(row.get("completed_count") or 0),
        "failedCount": int(row.get("failed_count") or 0),
        "pendingCount": int(row.get("pending_count") or 0),
        "checkpoint": checkpoint,
        "bucket": row.get("bucket", ""),
        "objectPrefix": row.get("object_prefix", ""),
        "resumable": _is_resumable_batch(row),
        "createdAt": _format_datetime(row.get("created_at")),
        "updatedAt": _format_datetime(row.get("updated_at")),
    }


def _packing_list_attempt_payload(row: dict[str, Any] | None) -> dict[str, Any] | None:
    if not row:
        return None
    return {
        "attemptId": row["attempt_id"],
        "batchId": row["batch_id"],
        "runId": row.get("run_id", ""),
        "mode": row.get("mode", ""),
        "status": row.get("status", ""),
        "message": row.get("message", ""),
        "result": _safe_json_loads(row.get("result_json")) or {},
        "startedAt": _format_datetime(row.get("started_at")),
        "finishedAt": _format_datetime(row.get("finished_at")),
        "createdAt": _format_datetime(row.get("created_at")),
        "updatedAt": _format_datetime(row.get("updated_at")),
    }


def _is_resumable_batch(row: dict[str, Any]) -> bool:
    status = str(row.get("status") or "").lower()
    if status in {"success", "completed"}:
        return False
    total_count = int(row.get("total_count") or 0)
    completed_count = int(row.get("completed_count") or 0)
    return total_count == 0 or completed_count < total_count


def _packing_list_date_folder(value: datetime | None = None) -> str:
    source = value or datetime.utcnow()
    if source.tzinfo is None:
        source = source.replace(tzinfo=timezone.utc)
    return source.astimezone(BEIJING_TZ).strftime("%Y-%m-%d")


def _normalize_packing_list_resume_mode(value: str) -> str:
    mode = str(value or "").strip().lower()
    if mode in {"continue", "retry-failed", "restart", "new"}:
        return mode
    return "continue"


def _normalize_packing_list_batch_status(status: str, counts: dict[str, int]) -> str:
    raw = str(status or "").strip().lower()
    if raw in {"interrupted", "failed"}:
        return raw
    if counts["total"] > 0 and counts["completed"] >= counts["total"]:
        return "success"
    if raw in {"success", "partial"}:
        return "partial" if counts["failed"] or counts["pending"] else "success"
    if counts["failed"] and counts["pending"] == 0:
        return "partial"
    return "running"


def _merge_packing_list_checkpoint(
    *,
    existing: dict[str, Any],
    payload: PackingListCheckpointPayload,
    stored_files: list[dict[str, Any]],
) -> dict[str, Any]:
    now = datetime.utcnow().isoformat()
    incoming = payload.checkpoint if isinstance(payload.checkpoint, dict) else {}
    checkpoint = {
        **existing,
        **incoming,
        "updatedAt": now,
        "attemptId": payload.attemptId or incoming.get("attemptId") or existing.get("attemptId", ""),
        "runId": payload.runId or incoming.get("runId") or existing.get("runId", ""),
        "mode": _normalize_packing_list_resume_mode(payload.mode or incoming.get("mode", "")),
        "message": payload.message or incoming.get("message") or existing.get("message", ""),
    }

    items_by_no: dict[str, dict[str, Any]] = {}
    for item in existing.get("items") or []:
        if isinstance(item, dict):
            key = str(item.get("no") or item.get("groupKey") or "").strip()
            if key:
                items_by_no[key] = dict(item)
    for item in incoming.get("items") or []:
        if isinstance(item, dict):
            key = str(item.get("no") or item.get("groupKey") or "").strip()
            if key:
                items_by_no[key] = {**items_by_no.get(key, {}), **item}

    group_results_by_no: dict[str, dict[str, Any]] = {}
    for group in existing.get("groupResults") or []:
        if isinstance(group, dict):
            key = str(group.get("no") or "").strip()
            if key:
                group_results_by_no[key] = dict(group)
    for group in incoming.get("groupResults") or []:
        if isinstance(group, dict):
            key = str(group.get("no") or "").strip()
            if key:
                group_results_by_no[key] = {**group_results_by_no.get(key, {}), **group}

    group_result = payload.groupResult or None
    if isinstance(group_result, dict):
        no = str(group_result.get("no") or "").strip()
        if no:
            group_files = [_stored_file_checkpoint_payload(row) for row in stored_files]
            merged_group = {**group_results_by_no.get(no, {}), **group_result}
            if group_files:
                merged_group["storedFiles"] = group_files
            group_results_by_no[no] = merged_group
            item = {
                **items_by_no.get(no, {}),
                "no": no,
                "status": "success" if group_result.get("ok") else "failed",
                "message": group_result.get("error") or payload.message or "",
                "poNumbers": group_result.get("poNumbers") or [],
                "successfulPoNumber": group_result.get("successfulPoNumber") or "",
                "downloadedFilePath": group_result.get("filePath") or "",
                "updatedAt": now,
            }
            if group_files:
                item["storedFiles"] = group_files
            items_by_no[no] = item

    result = payload.result if isinstance(payload.result, dict) else None
    if result:
        checkpoint["result"] = result
        for group in result.get("groupResults") or []:
            if isinstance(group, dict):
                no = str(group.get("no") or "").strip()
                if no:
                    group_results_by_no[no] = {**group_results_by_no.get(no, {}), **group}
                    items_by_no.setdefault(no, {
                        "no": no,
                        "status": "success" if group.get("ok") else "failed",
                        "message": group.get("error") or "",
                        "poNumbers": group.get("poNumbers") or [],
                        "updatedAt": now,
                    })

    checkpoint["items"] = list(items_by_no.values())
    checkpoint["groupResults"] = list(group_results_by_no.values())
    counts = _packing_list_checkpoint_counts(checkpoint)
    checkpoint.update({
        "totalCount": counts["total"],
        "completedCount": counts["completed"],
        "failedCount": counts["failed"],
        "pendingCount": counts["pending"],
        "status": _normalize_packing_list_batch_status(payload.status, counts),
    })
    return checkpoint


def _merge_ticket_owner_checkpoint(
    *,
    batch_id: str,
    existing: dict[str, Any],
    payload: TicketOwnerCheckpointPayload,
    stored_files: list[dict[str, Any]],
) -> dict[str, Any]:
    now = datetime.utcnow().isoformat()
    incoming = payload.checkpoint if isinstance(payload.checkpoint, dict) else {}
    checkpoint = {
        **existing,
        **incoming,
        "batchId": batch_id,
        "attemptId": payload.attemptId or incoming.get("attemptId") or existing.get("attemptId", ""),
        "runId": payload.runId or incoming.get("runId") or existing.get("runId", ""),
        "mode": _normalize_packing_list_resume_mode(payload.mode or incoming.get("mode", "")),
        "message": payload.message or incoming.get("message") or existing.get("message", ""),
        "updatedAt": now,
    }

    items_by_key: dict[str, dict[str, Any]] = {}
    for item in existing.get("items") or []:
        if isinstance(item, dict):
            key = _ticket_owner_item_key(item)
            if key:
                items_by_key[key] = dict(item)
    for item in incoming.get("items") or []:
        if isinstance(item, dict):
            key = _ticket_owner_item_key(item)
            if key:
                items_by_key[key] = {**items_by_key.get(key, {}), **item}

    if isinstance(payload.itemResult, dict):
        key = _ticket_owner_item_key(payload.itemResult)
        if key:
            item_status = str(payload.itemResult.get("status") or "").strip().lower()
            if not item_status:
                item_status = "success" if payload.itemResult.get("ok") is not False else "failed"
            items_by_key[key] = {
                **items_by_key.get(key, {}),
                **payload.itemResult,
                "itemKey": key,
                "status": item_status,
                "updatedAt": now,
            }

    if isinstance(payload.result, dict):
        checkpoint["result"] = payload.result

    if stored_files:
        checkpoint["storedFiles"] = [_stored_file_checkpoint_payload(row) for row in stored_files]

    checkpoint["items"] = list(items_by_key.values())
    counts = _ticket_owner_checkpoint_counts(checkpoint)
    checkpoint.update({
        "totalCount": counts["total"],
        "completedCount": counts["completed"],
        "failedCount": counts["failed"],
        "pendingCount": counts["pending"],
        "status": _normalize_ticket_owner_batch_status(payload.status, counts),
    })
    return checkpoint


def _ticket_owner_item_key(item: dict[str, Any]) -> str:
    return str(
        item.get("itemKey")
        or item.get("caseNumber")
        or item.get("Case Number")
        or item.get("case_number")
        or "",
    ).strip()


def _packing_list_checkpoint_counts(checkpoint: dict[str, Any]) -> dict[str, int]:
    items = [item for item in (checkpoint.get("items") or []) if isinstance(item, dict)]
    result = checkpoint.get("result") if isinstance(checkpoint.get("result"), dict) else {}
    total = int(result.get("totalPackingListCount") or result.get("totalGroupCount") or checkpoint.get("totalCount") or len(items) or 0)
    completed = sum(1 for item in items if str(item.get("status") or "").lower() == "success")
    failed = sum(1 for item in items if str(item.get("status") or "").lower() in {"failed", "error"})
    if not completed and result:
        completed = int(result.get("downloadedPackingListCount") or 0)
    if not failed and result:
        failed = int(result.get("failedPackingListCount") or 0)
    pending = max(0, total - completed - failed)
    return {
        "total": total,
        "completed": completed,
        "failed": failed,
        "pending": pending,
    }


def _ticket_owner_checkpoint_counts(checkpoint: dict[str, Any]) -> dict[str, int]:
    items = [item for item in (checkpoint.get("items") or []) if isinstance(item, dict)]
    result = checkpoint.get("result") if isinstance(checkpoint.get("result"), dict) else {}
    total = int(
        checkpoint.get("totalCount")
        or result.get("totalTicketCount")
        or result.get("attemptedTicketCount")
        or result.get("rowCount")
        or len(items)
        or 0,
    )
    if items:
        completed = sum(1 for item in items if _ticket_owner_item_status(item) == "success")
        failed = sum(1 for item in items if _ticket_owner_item_status(item) in {"failed", "error"})
    else:
        completed = int(
            checkpoint.get("completedCount")
            or checkpoint.get("successCount")
            or result.get("successTicketCount")
            or 0,
        )
        failed = int(
            checkpoint.get("failedCount")
            or result.get("failedTicketCount")
            or 0,
        )
    pending = int(checkpoint.get("pendingCount") or max(0, total - completed - failed))
    return {
        "total": total,
        "completed": completed,
        "failed": failed,
        "pending": pending,
    }


def _ticket_owner_item_status(item: dict[str, Any]) -> str:
    status = str(item.get("status") or "").strip().lower()
    if status:
        return status
    return "success" if item.get("ok") is not False else "failed"


def _normalize_ticket_owner_batch_status(status: str, counts: dict[str, int]) -> str:
    raw = str(status or "").strip().lower()
    if raw in {"interrupted", "failed"}:
        return raw
    if counts["total"] > 0 and counts["completed"] >= counts["total"]:
        return "success"
    if raw in {"success", "partial"}:
        return "partial" if counts["failed"] or counts["pending"] else "success"
    if counts["failed"] and counts["pending"] == 0:
        return "partial"
    return "running"


def _stored_file_checkpoint_payload(row: dict[str, Any]) -> dict[str, Any]:
    return {
        "id": row.get("id"),
        "fileRole": row.get("file_role", ""),
        "bucket": row.get("bucket", ""),
        "objectKey": row.get("object_key", ""),
        "originalFilename": row.get("original_filename", ""),
        "downloadPath": f"/api/automation/run-files/{row['id']}/download" if row.get("id") else "",
    }


def _store_packing_list_checkpoint_file(
    batch_row: dict[str, Any],
    payload: PackingListCheckpointPayload,
    file_payload: PackingListCheckpointFilePayload,
) -> dict[str, Any]:
    try:
        content = base64.b64decode(file_payload.contentBase64, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise HTTPException(status_code=400, detail="Invalid base64 checkpoint file content.") from exc

    filename = sanitize_object_segment(file_payload.fileName or "packing-list.pdf")
    file_role = sanitize_object_segment(file_payload.fileRole or "packing_list_pdf").replace("-", "_")
    bucket = batch_row.get("bucket") or get_minio_bucket(PACKING_LIST_BUCKET_KEY)
    object_prefix = str(batch_row.get("object_prefix") or "").strip()
    if not object_prefix:
        object_prefix = f"{PACKING_LIST_OBJECT_ROOT}/{_packing_list_date_folder()}/{batch_row['batch_id']}"
    attempt_id = sanitize_object_segment(payload.attemptId or "attempt")
    timestamp = datetime.utcnow().strftime("%Y%m%d%H%M%S%f")
    object_key = f"{object_prefix}/{attempt_id}/{file_role}/{timestamp}-{filename}"
    storage_record = put_object_bytes(
        bucket=bucket,
        object_key=object_key,
        content=content,
        content_type=file_payload.contentType or "application/octet-stream",
    )
    if not payload.runId:
        return {
            "id": None,
            "run_id": "",
            "file_role": file_role,
            "bucket": bucket,
            "object_key": object_key,
            "original_filename": filename,
            "content_type": file_payload.contentType or "application/octet-stream",
            "file_size": storage_record["file_size"],
            "sha256": storage_record["sha256"],
            "created_at": datetime.utcnow(),
        }
    return insert_automation_run_file({
        "run_id": payload.runId,
        "file_role": file_role,
        "bucket": bucket,
        "object_key": object_key,
        "original_filename": filename,
        "content_type": file_payload.contentType or "application/octet-stream",
        "file_size": storage_record["file_size"],
        "sha256": storage_record["sha256"],
    })


async def _try_store_upload_file(
    *,
    run_id: str,
    automation_id: str,
    upload: UploadFile,
    file_role: str,
    bucket_key: str,
    warnings: list[str],
) -> dict[str, Any] | None:
    try:
        return await _store_upload_file(
            run_id=run_id,
            automation_id=automation_id,
            upload=upload,
            file_role=file_role,
            bucket_key=bucket_key,
        )
    except Exception:
        logger.exception("Failed to store automation run upload file.")
        warnings.append("File storage is unavailable; run record was kept without attached source file.")
        return None


def _try_store_result_json(
    run_id: str,
    automation_id: str,
    result: Any,
    warnings: list[str],
) -> dict[str, Any] | None:
    try:
        return _store_result_json(run_id, automation_id, result)
    except Exception:
        logger.exception("Failed to store automation run result JSON.")
        warnings.append("File storage is unavailable; run result was kept without attached result JSON.")
        return None


def _try_store_ticket_owner_excel(
    run_id: str,
    automation_id: str,
    result: Any,
    warnings: list[str],
) -> dict[str, Any] | None:
    if automation_id != TICKET_OWNER_AUTOMATION_ID or not isinstance(result, dict):
        return None
    rows = _extract_ticket_owner_rows(result)
    if rows is None:
        return None

    try:
        content = _build_ticket_owner_workbook_bytes(rows)
        return _store_result_file_bytes(
            run_id=run_id,
            automation_id=automation_id,
            file_role="result_excel",
            filename=TICKET_OWNER_EXCEL_NAME,
            content=content,
            content_type=EXCEL_CONTENT_TYPE,
            result=result,
        )
    except Exception:
        logger.exception("Failed to build server-side ticket ownership workbook.")
        warnings.append("File storage is unavailable; ticket ownership Excel was not generated on the server.")
        return None


def _store_ticket_owner_checkpoint_result_files(
    run_id: str,
    checkpoint: dict[str, Any],
    result: dict[str, Any] | None,
) -> list[dict[str, Any]]:
    if not run_id or not isinstance(result, dict):
        return []

    storage_result = {
        **result,
        "batchId": checkpoint.get("batchId", ""),
        "attemptId": checkpoint.get("attemptId", ""),
        "checkpoint": {
            "batchId": checkpoint.get("batchId", ""),
            "attemptId": checkpoint.get("attemptId", ""),
        },
    }
    stored_files: list[dict[str, Any]] = []
    result_json = _store_result_json(run_id, TICKET_OWNER_AUTOMATION_ID, storage_result)
    if result_json:
        stored_files.append(result_json)
    ticket_owner_excel = _try_store_ticket_owner_excel(
        run_id,
        TICKET_OWNER_AUTOMATION_ID,
        storage_result,
        [],
    )
    if ticket_owner_excel:
        stored_files.append(ticket_owner_excel)
    return stored_files


def _try_store_remote_result_file(
    run_id: str,
    automation_id: str,
    file_payload: RunFilePayload,
    warnings: list[str],
) -> dict[str, Any] | None:
    try:
        return _store_remote_result_file(run_id, automation_id, file_payload)
    except Exception:
        logger.exception("Failed to store automation run remote result file.")
        warnings.append("File storage is unavailable; run result was kept without one attached file.")
        return None


def _raise_template_storage_error(template_id: int, row: dict[str, Any], exc: Exception) -> None:
    error_code = str(getattr(exc, "code", "") or getattr(exc, "response", "") or "")
    if error_code in {"NoSuchKey", "NoSuchBucket", "NoSuchObject"}:
        logger.warning(
            "Automation template object missing: template_id=%s bucket=%s object_key=%s code=%s",
            template_id,
            row.get("bucket"),
            row.get("object_key"),
            error_code,
        )
        raise HTTPException(
            status_code=404,
            detail="模板文件不存在或已从 MinIO 删除，请联系管理员重新上传模板。",
        ) from exc

    logger.exception(
        "Failed to download automation template: template_id=%s bucket=%s object_key=%s",
        template_id,
        row.get("bucket"),
        row.get("object_key"),
    )
    raise HTTPException(
        status_code=503,
        detail="模板文件暂时无法下载，请联系管理员检查 MinIO 存储连接和模板对象。",
    ) from exc


async def _store_upload_file(
    *,
    run_id: str,
    automation_id: str,
    upload: UploadFile,
    file_role: str,
    bucket_key: str,
) -> dict[str, Any]:
    content = await upload.read()
    filename = sanitize_object_segment(upload.filename or "upload.xlsx")
    bucket = get_minio_bucket(bucket_key)
    object_key = build_object_key("runs", automation_id, run_id, file_role, filename)
    storage_record = put_object_bytes(
        bucket=bucket,
        object_key=object_key,
        content=content,
        content_type=upload.content_type or "application/octet-stream",
    )
    return insert_automation_run_file({
        "run_id": run_id,
        "file_role": file_role,
        "bucket": bucket,
        "object_key": object_key,
        "original_filename": filename,
        "content_type": upload.content_type or "application/octet-stream",
        "file_size": storage_record["file_size"],
        "sha256": storage_record["sha256"],
    })


def _store_result_json(run_id: str, automation_id: str, result: Any) -> dict[str, Any]:
    content = json.dumps(result, ensure_ascii=False, indent=2).encode("utf-8")
    if automation_id == PACKING_LIST_AUTOMATION_ID:
        bucket = get_minio_bucket(PACKING_LIST_BUCKET_KEY)
        object_prefix = _packing_list_result_object_prefix(run_id, result)
        object_key = f"{object_prefix}/{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}-result.json"
    elif automation_id == TICKET_OWNER_AUTOMATION_ID:
        bucket = get_minio_bucket(TICKET_OWNER_BUCKET_KEY)
        object_prefix = _ticket_owner_result_object_prefix(run_id, result)
        object_key = f"{object_prefix}/{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}-result.json"
    else:
        bucket = get_minio_bucket("results")
        object_key = build_object_key("results", automation_id, run_id, "result.json")
    storage_record = put_object_bytes(
        bucket=bucket,
        object_key=object_key,
        content=content,
        content_type="application/json; charset=utf-8",
    )
    return insert_automation_run_file({
        "run_id": run_id,
        "file_role": "result_json",
        "bucket": bucket,
        "object_key": object_key,
        "original_filename": "result.json",
        "content_type": "application/json; charset=utf-8",
        "file_size": storage_record["file_size"],
        "sha256": storage_record["sha256"],
    })


def _packing_list_result_object_prefix(run_id: str, result: Any) -> str:
    batch_id = ""
    if isinstance(result, dict):
        checkpoint = result.get("checkpoint")
        if isinstance(checkpoint, dict):
            batch_id = str(checkpoint.get("batchId") or "").strip()
        batch_id = batch_id or str(result.get("batchId") or "").strip()
    if batch_id:
        row = get_automation_batch(batch_id)
        if row and row.get("object_prefix"):
            return f"{row['object_prefix']}/final"
        return f"{PACKING_LIST_OBJECT_ROOT}/{_packing_list_date_folder()}/{sanitize_object_segment(batch_id)}/final"
    return f"{PACKING_LIST_OBJECT_ROOT}/{_packing_list_date_folder()}/{sanitize_object_segment(run_id or 'run')}/final"


def _ticket_owner_result_object_prefix(run_id: str, result: Any) -> str:
    batch_id = ""
    attempt_id = ""
    if isinstance(result, dict):
        checkpoint = result.get("checkpoint")
        if isinstance(checkpoint, dict):
            batch_id = str(checkpoint.get("batchId") or "").strip()
            attempt_id = str(checkpoint.get("attemptId") or "").strip()
        batch_id = batch_id or str(result.get("batchId") or "").strip()
        attempt_id = attempt_id or str(result.get("attemptId") or "").strip()
    folder_id = sanitize_object_segment(batch_id or run_id or "run")
    attempt_folder = sanitize_object_segment(attempt_id or "final")
    return f"{TICKET_OWNER_OBJECT_ROOT}/{_packing_list_date_folder()}/{folder_id}/{attempt_folder}"


def _store_result_file_bytes(
    *,
    run_id: str,
    automation_id: str,
    file_role: str,
    filename: str,
    content: bytes,
    content_type: str,
    result: Any | None = None,
) -> dict[str, Any]:
    safe_filename = sanitize_object_segment(filename)
    if automation_id == TICKET_OWNER_AUTOMATION_ID:
        bucket = get_minio_bucket(TICKET_OWNER_BUCKET_KEY)
        object_prefix = _ticket_owner_result_object_prefix(run_id, result)
        object_key = f"{object_prefix}/{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}-{sanitize_object_segment(file_role)}-{safe_filename}"
    else:
        bucket = get_minio_bucket("results")
        object_key = build_object_key("results", automation_id, run_id, file_role, safe_filename)
    storage_record = put_object_bytes(
        bucket=bucket,
        object_key=object_key,
        content=content,
        content_type=content_type,
    )
    return insert_automation_run_file({
        "run_id": run_id,
        "file_role": file_role,
        "bucket": bucket,
        "object_key": object_key,
        "original_filename": safe_filename,
        "content_type": content_type,
        "file_size": storage_record["file_size"],
        "sha256": storage_record["sha256"] or sha256_bytes(content),
    })


def _store_remote_result_file(run_id: str, automation_id: str, file_payload: RunFilePayload) -> dict[str, Any]:
    if file_payload.contentBase64:
        return _store_result_file_content(run_id, automation_id, file_payload)

    if not file_payload.url.lower().startswith(("http://127.0.0.1:", "http://localhost:")):
        raise HTTPException(status_code=400, detail="Only local executor artifact URLs can be stored.")

    content, detected_content_type = download_url_bytes(file_payload.url)
    filename = sanitize_object_segment(file_payload.fileName or file_payload.url.split("/")[-1] or "result-file")
    content_type = file_payload.contentType or detected_content_type or "application/octet-stream"
    if automation_id == TICKET_OWNER_AUTOMATION_ID:
        bucket = get_minio_bucket(TICKET_OWNER_BUCKET_KEY)
        object_prefix = _ticket_owner_result_object_prefix(run_id, None)
        object_key = f"{object_prefix}/{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}-{sanitize_object_segment(file_payload.fileRole)}-{filename}"
    else:
        bucket = get_minio_bucket("results")
        object_key = build_object_key("results", automation_id, run_id, file_payload.fileRole, filename)
    storage_record = put_object_bytes(
        bucket=bucket,
        object_key=object_key,
        content=content,
        content_type=content_type,
    )
    return insert_automation_run_file({
        "run_id": run_id,
        "file_role": file_payload.fileRole,
        "bucket": bucket,
        "object_key": object_key,
        "original_filename": filename,
        "content_type": content_type,
        "file_size": storage_record["file_size"],
        "sha256": storage_record["sha256"] or sha256_bytes(content),
    })


def _store_result_file_content(run_id: str, automation_id: str, file_payload: RunFilePayload) -> dict[str, Any]:
    try:
        content = base64.b64decode(file_payload.contentBase64, validate=True)
    except (binascii.Error, ValueError) as exc:
        raise HTTPException(status_code=400, detail="Invalid base64 result file content.") from exc

    filename = sanitize_object_segment(file_payload.fileName or file_payload.url.split("/")[-1] or "result-file")
    content_type = file_payload.contentType or "application/octet-stream"
    if automation_id == TICKET_OWNER_AUTOMATION_ID:
        bucket = get_minio_bucket(TICKET_OWNER_BUCKET_KEY)
        object_prefix = _ticket_owner_result_object_prefix(run_id, None)
        object_key = f"{object_prefix}/{datetime.utcnow().strftime('%Y%m%d%H%M%S%f')}-{sanitize_object_segment(file_payload.fileRole)}-{filename}"
    else:
        bucket = get_minio_bucket("results")
        object_key = build_object_key("results", automation_id, run_id, file_payload.fileRole, filename)
    storage_record = put_object_bytes(
        bucket=bucket,
        object_key=object_key,
        content=content,
        content_type=content_type,
    )
    return insert_automation_run_file({
        "run_id": run_id,
        "file_role": file_payload.fileRole,
        "bucket": bucket,
        "object_key": object_key,
        "original_filename": filename,
        "content_type": content_type,
        "file_size": storage_record["file_size"],
        "sha256": storage_record["sha256"] or sha256_bytes(content),
    })


def _build_ticket_owner_workbook_bytes(rows: list[Any]) -> bytes:
    workbook = Workbook()
    worksheet = workbook.active
    worksheet.title = "Ticket ownership"

    header_fill = PatternFill("solid", fgColor="FF5B9BD5")
    stripe_fill = PatternFill("solid", fgColor="FFDDEBF7")
    white_fill = PatternFill("solid", fgColor="FFFFFFFF")
    thin_side = Side(style="thin", color="FF000000")
    border = Border(left=thin_side, right=thin_side, top=thin_side, bottom=thin_side)
    header_font = Font(bold=True, color="FFFFFF")
    alignment = Alignment(horizontal="center", vertical="center")

    worksheet.append(list(TICKET_OWNER_COLUMNS))
    for cell in worksheet[1]:
        cell.fill = header_fill
        cell.font = header_font
        cell.border = border
        cell.alignment = alignment
        cell.number_format = "@"

    for row_index, source_row in enumerate(rows, start=2):
        normalized_row = _normalize_ticket_owner_row(source_row)
        worksheet.append([normalized_row[column] for column in TICKET_OWNER_COLUMNS])
        fill = stripe_fill if row_index % 2 == 0 else white_fill
        for cell in worksheet[row_index]:
            cell.fill = fill
            cell.border = border
            cell.alignment = alignment
            cell.number_format = "@"

    column_widths = (16, 30, 34, 18, 22, 16, 16)
    for column_index, width in enumerate(column_widths, start=1):
        worksheet.column_dimensions[get_column_letter(column_index)].width = width

    worksheet.freeze_panes = "A2"
    table_ref = f"A1:G{max(worksheet.max_row, 1)}"
    worksheet.auto_filter.ref = table_ref
    if worksheet.max_row >= 2:
        table = Table(displayName="TicketOwnership", ref=table_ref)
        table.tableStyleInfo = TableStyleInfo(
            name="TableStyleMedium2",
            showFirstColumn=False,
            showLastColumn=False,
            showRowStripes=True,
            showColumnStripes=False,
        )
        worksheet.add_table(table)

    stream = BytesIO()
    workbook.save(stream)
    workbook.close()
    return stream.getvalue()


def _extract_ticket_owner_rows(result: Any) -> list[Any] | None:
    if not isinstance(result, dict):
        return None
    direct_rows = result.get("rows")
    if isinstance(direct_rows, list):
        return direct_rows
    workflow_result = result.get("workflowResult")
    if isinstance(workflow_result, dict) and isinstance(workflow_result.get("rows"), list):
        return workflow_result["rows"]
    return None


def _normalize_ticket_owner_row(row: Any) -> dict[str, str]:
    if not isinstance(row, dict):
        return {column: "" for column in TICKET_OWNER_COLUMNS}
    return {
        column: _normalize_ticket_owner_cell(row.get(column))
        for column in TICKET_OWNER_COLUMNS
    }


def _normalize_ticket_owner_cell(value: Any) -> str:
    if value is None:
        return ""
    return " ".join(str(value).replace("\xa0", " ").split())


def _safe_json_loads(value: Any) -> Any:
    if not value:
        return None
    if not isinstance(value, str):
        return value
    try:
        return json.loads(value)
    except json.JSONDecodeError:
        return None


def _parse_beijing_date_range(date_from: str | None, date_to: str | None) -> tuple[datetime | None, datetime | None]:
    start_date = _parse_date_only(date_from)
    end_date = _parse_date_only(date_to)
    started_from = _beijing_date_to_utc_naive(start_date) if start_date else None
    if end_date:
        started_to = _beijing_date_to_utc_naive(end_date + timedelta(days=1))
    elif start_date:
        started_to = _beijing_date_to_utc_naive(start_date + timedelta(days=1))
    else:
        started_to = None
    return started_from, started_to


def _parse_date_only(value: str | None):
    raw = str(value or "").strip()
    if not raw:
        return None
    try:
        return datetime.strptime(raw[:10], "%Y-%m-%d").date()
    except ValueError as exc:
        raise HTTPException(status_code=400, detail="日期筛选格式应为 YYYY-MM-DD。") from exc


def _beijing_date_to_utc_naive(value) -> datetime:
    beijing_dt = datetime.combine(value, time.min, tzinfo=BEIJING_TZ)
    return beijing_dt.astimezone(timezone.utc).replace(tzinfo=None)


def _format_datetime(value: Any) -> str:
    if hasattr(value, "isoformat"):
        return value.isoformat()
    return str(value or "")


def _format_beijing_datetime(value: Any) -> str:
    if not isinstance(value, datetime):
        return str(value or "")
    source = value
    if source.tzinfo is None:
        source = source.replace(tzinfo=timezone.utc)
    return source.astimezone(BEIJING_TZ).strftime("%Y-%m-%d %H:%M:%S")


def _duration_seconds(started_at: Any, finished_at: Any) -> int | None:
    if not isinstance(started_at, datetime) or not isinstance(finished_at, datetime):
        return None
    started = started_at.replace(tzinfo=timezone.utc) if started_at.tzinfo is None else started_at.astimezone(timezone.utc)
    finished = finished_at.replace(tzinfo=timezone.utc) if finished_at.tzinfo is None else finished_at.astimezone(timezone.utc)
    seconds = int((finished - started).total_seconds())
    return seconds if seconds >= 0 else None

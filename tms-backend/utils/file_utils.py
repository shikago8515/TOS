
# -*- coding: utf-8 -*-
"""
TMS 工具 - 文件工具模块
创建时间: 2026-05-18
"""

import os
import shutil
import logging
from typing import BinaryIO, Iterable, List, Optional, Protocol, Set
import openpyxl
from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

from utils.excel_upload_backup import ExcelUploadBackupContext, store_excel_upload_backup


logger = logging.getLogger(__name__)
logger.addHandler(logging.NullHandler())


class UploadFileLike(Protocol):
    file: BinaryIO


def ensure_dir(dir_path):
    """确保目录存在，不存在则创建"""
    if not os.path.exists(dir_path):
        os.makedirs(dir_path, exist_ok=True)


def get_file_extension(file_path):
    """获取文件扩展名（小写）"""
    return os.path.splitext(file_path)[1].lower()


def validate_upload_filename(
    filename: Optional[str],
    allowed_extensions: Set[str],
    label: str,
) -> str:
    """校验上传文件名，只接受 basename，避免路径遍历。"""
    raw_name = (filename or "").strip()
    if not raw_name:
        raise ValueError(f"请选择要上传的 {label}")

    if _has_path_segment(raw_name):
        raise ValueError("文件名无效")

    safe_name = os.path.basename(raw_name)
    if not safe_name or safe_name in {".", ".."}:
        raise ValueError("文件名无效")

    extension = os.path.splitext(safe_name)[1].lower()
    normalized_extensions = {entry.lower() for entry in allowed_extensions}
    if extension not in normalized_extensions:
        allowed_text = " / ".join(sorted(normalized_extensions))
        raise ValueError(f"{label} 仅支持 {allowed_text}")

    return safe_name


def resolve_download_path(root_dir: str, filename: str) -> str:
    """把下载文件名解析到指定根目录内，禁止目录逃逸。"""
    safe_name = validate_upload_filename(
        filename,
        {".xlsx", ".xlsm", ".xls"},
        "下载文件",
    )
    root = os.path.abspath(root_dir)
    file_path = os.path.abspath(os.path.join(root, safe_name))
    if os.path.commonpath([root, file_path]) != root:
        raise ValueError("文件名无效")

    return file_path


def copy_upload_to_path(
    upload: UploadFileLike,
    target_path: str,
    backup_context: ExcelUploadBackupContext | None = None,
) -> None:
    with open(target_path, "wb") as fp:
        shutil.copyfileobj(upload.file, fp)
    if backup_context is None:
        return

    try:
        store_excel_upload_backup(target_path, backup_context)
    except Exception:
        # 备份是审计增强能力，不能阻断用户当前 Excel 处理流程。
        logger.warning(
            "Excel upload backup failed: module_id=%s request_id=%s file_role=%s",
            backup_context.module_id,
            backup_context.request_id,
            backup_context.file_role,
        )


def copy_output_to_directory(output_path: str, target_dir: str) -> str:
    output_path = os.path.abspath(output_path)
    output_filename = os.path.basename(output_path)
    if not output_filename:
        raise ValueError("输出文件名无效")

    target_path = os.path.abspath(os.path.join(target_dir, output_filename))
    if os.path.abspath(os.path.dirname(output_path)) != os.path.abspath(target_dir):
        shutil.copy2(output_path, target_path)

    return output_filename


def sanitize_output_reference(text: str, output_path: str, output_filename: str) -> str:
    """把返回给前端的输出路径替换为 basename，避免泄露本机目录。"""
    sanitized_text = text
    candidates = {
        output_path,
        output_path.replace("\\", "/"),
        output_path.replace("/", "\\"),
    }
    for candidate in sorted(candidates, key=len, reverse=True):
        if candidate:
            sanitized_text = sanitized_text.replace(candidate, output_filename)

    return sanitized_text


def sanitize_output_logs(
    logs: Iterable[object],
    output_path: str,
    output_filename: str,
) -> List[str]:
    return [
        sanitize_output_reference(str(entry), output_path, output_filename)
        for entry in logs
    ]


def _has_path_segment(filename: str) -> bool:
    return (
        "/" in filename
        or "\\" in filename
        or os.path.basename(filename) != filename
        or os.path.splitdrive(filename)[0] != ""
    )


def create_thin_border():
    """创建通用的细边框样式"""
    return Border(
        left=Side(style='thin', color='00000000'),
        right=Side(style='thin', color='00000000'),
        top=Side(style='thin', color='00000000'),
        bottom=Side(style='thin', color='00000000')
    )


def adjust_column_widths(ws, max_widths=None):
    """
    自动调整Excel列宽
    max_widths: 可选，指定列的最大宽度，{列索引: 最大宽度}
    """
    if max_widths is None:
        max_widths = {}

    for col_idx in range(1, ws.max_column + 1):
        max_length = 0
        column_letter = openpyxl.utils.get_column_letter(col_idx)
        for row in range(1, ws.max_row + 1):
            cell = ws.cell(row=row, column=col_idx)
            try:
                if cell.value:
                    length = len(str(cell.value))
                    if length > max_length:
                        max_length = length
            except:
                pass
        default_max = 25
        if col_idx in max_widths:
            adjusted_width = min(max_length + 2, max_widths[col_idx])
        else:
            adjusted_width = min(max_length + 2, default_max)
        ws.column_dimensions[column_letter].width = adjusted_width

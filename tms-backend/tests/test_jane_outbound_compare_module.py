import os
import sys
import tempfile
import unittest

import openpyxl
from openpyxl.styles import PatternFill


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.jane_outbound_compare_module import JaneOutboundCompareModule


class JaneOutboundCompareModuleTests(unittest.TestCase):
    def setUp(self):
        self.module = JaneOutboundCompareModule()

    def _fill_rgb(self, cell):
        color = cell.fill.fgColor
        if color.type == "rgb":
            return (color.rgb or "").upper()
        return ""

    def _has_fill(self, cell):
        return cell.fill.fill_type is not None

    def _column_by_header(self, ws, header):
        for cell in ws[1]:
            if cell.value == header:
                return cell.column
        self.fail(f"未找到表头：{header}")

    def _save_outbound_workbook(self, folder):
        path = os.path.join(folder, "T1 OUTBOUND.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "OUTBOUND"
        ws.append([
            "Style Number",
            "Production Lot ID",
            "Lot Manufactured Date",
            "Invoice Date/Delivery Date",
            "Outbound Quantity (in Units)",
            "Units",
            "PO Number",
            "Line Number",
            "Recording Facility ID",
            "Shipment Type",
            "Outbound Delay Reason",
            "Other Reason",
        ])
        orange_fill = PatternFill("solid", fgColor="FFFFD966")
        ws["B1"].fill = orange_fill
        ws["C1"].fill = orange_fill
        ws.append([
            "LD1803",
            "OHO055-M-LD1803",
            "2026-02-01",
            "2026-03-02",
            35,
            "Each",
            "0902235116",
            "10",
            "3LP001",
            "",
            "",
            "RC2610OW005H2H",
        ])
        ws.append([
            "LD1804",
            "OHO055-M-LD1804",
            "2026-02-01",
            "2026-03-02",
            12,
            "Each",
            "0902235117",
            "10",
            "3LP001",
            "",
            "",
            "RC2610OW005H2H",
        ])
        for row in range(2, 4):
            for column in range(1, 13):
                ws.cell(row=row, column=column).fill = orange_fill
        wb.save(path)
        return path

    def _save_tms_workbook(self, folder):
        path = os.path.join(folder, "TMS.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Result Set"
        ws.append([
            "Assigned Factory",
            "Factory",
            "PO Number",
            "PO Line Item #",
            "Working Number",
            "Article Number",
            "PODD",
            "Ordered Quantity",
        ])
        ws.append([
            "3LP001",
            "3LP001",
            "0902235116",
            "10",
            "RC2610OW005H2H",
            "LD1803",
            "2026-03-05",
            30,
        ])
        ws.append([
            "3LP001",
            "3LP001",
            "0902235118",
            "10",
            "RC2610OW005H2H",
            "LD1805",
            "2026-03-06",
            8,
        ])
        ws.append([
            "3LP001",
            "3LP001",
            "0902999999",
            "10",
            "OTHER-WORKING",
            "LD9999",
            "2026-03-06",
            8,
        ])
        wb.save(path)
        return path

    def test_marks_differences_missing_outbound_and_ignores_unrelated_tms_rows(self):
        with tempfile.TemporaryDirectory() as folder:
            outbound_path = self._save_outbound_workbook(folder)
            tms_path = self._save_tms_workbook(folder)

            result = self.module.process_reports(outbound_path, tms_path, folder)

            self.assertTrue(result["success"])
            self.assertEqual(result["checked_row_count"], 2)
            self.assertEqual(result["matched_row_count"], 1)
            self.assertEqual(result["missing_tms_row_count"], 1)
            self.assertEqual(result["missing_outbound_row_count"], 1)
            self.assertEqual(result["difference_cell_count"], 2)

            output_wb = openpyxl.load_workbook(result["output_path"])
            ws = output_wb["OUTBOUND"]
            self.assertNotIn("OUTBOUND_Check", output_wb.sheetnames)

            result_col = self._column_by_header(ws, "Check Result")
            source_col = self._column_by_header(ws, "Mismatch Source")
            detail_col = self._column_by_header(ws, "Check Detail")
            tms_row_col = self._column_by_header(ws, "Copy of TMS Source Row")

            self.assertEqual(self._fill_rgb(ws["B1"]), "FFFFD966")
            self.assertEqual(self._fill_rgb(ws["D2"]), "FFFFC7CE")
            self.assertEqual(ws["E1"].value, "Correct PODD")
            self.assertEqual(ws["E2"].value, "2026-03-05")
            self.assertFalse(self._has_fill(ws["E2"]))
            self.assertEqual(self._fill_rgb(ws["F2"]), "FFFFC7CE")
            self.assertEqual(ws["G1"].value, "Correct Ordered Quantity")
            self.assertEqual(ws["G2"].value, 30)
            self.assertFalse(self._has_fill(ws["G2"]))
            self.assertEqual(ws["I1"].value, "Correct Units")
            self.assertIsNone(ws["I2"].value)
            self.assertEqual(ws["P1"].value, "Correct Working Number")
            self.assertIsNone(ws["P2"].value)
            self.assertFalse(self._has_fill(ws["B2"]))
            self.assertFalse(self._has_fill(ws["O2"]))
            self.assertEqual(ws.cell(row=2, column=result_col).value, "需核对")
            self.assertEqual(ws.cell(row=2, column=source_col).value, "值不一致：以 Copy of TMS 为准")
            self.assertIn("PODD", ws.cell(row=2, column=detail_col).value)
            self.assertEqual(ws.cell(row=2, column=tms_row_col).value, 2)
            for cell_ref in ["A3", "J3", "K3", "L3"]:
                self.assertEqual(self._fill_rgb(ws[cell_ref]), "FFFFC7CE")
            self.assertFalse(self._has_fill(ws["B3"]))
            self.assertFalse(self._has_fill(ws["M3"]))
            self.assertEqual(ws.cell(row=3, column=source_col).value, "T1 OUTBOUND存在，Copy of TMS缺失")

            self.assertEqual(ws["A4"].value, "LD1805")
            self.assertEqual(ws["D4"].value, "2026-03-06")
            self.assertIsNone(ws["E4"].value)
            self.assertEqual(ws["F4"].value, 8)
            self.assertEqual(ws["J4"].value, "0902235118")
            self.assertEqual(ws["K4"].value, "10")
            self.assertEqual(ws["L4"].value, "3LP001")
            self.assertEqual(ws["O4"].value, "RC2610OW005H2H")
            self.assertEqual(ws.cell(row=4, column=source_col).value, "Copy of TMS存在，T1 OUTBOUND缺失")
            self.assertEqual(ws.cell(row=4, column=tms_row_col).value, 3)
            for cell_ref in ["A4", "J4", "K4", "L4", "O4"]:
                self.assertEqual(self._fill_rgb(ws[cell_ref]), "FFFFC7CE")
            self.assertFalse(self._has_fill(ws["B4"]))
            self.assertNotIn(
                "OTHER-WORKING",
                [ws.cell(row=row, column=source_col).value for row in range(2, ws.max_row + 1)],
            )


if __name__ == "__main__":
    unittest.main()

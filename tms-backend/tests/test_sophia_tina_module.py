import os
import sys
import tempfile
import unittest
import zipfile
from pathlib import Path
from xml.etree import ElementTree

import pandas as pd
from openpyxl import load_workbook


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.sophia_tina_module import SophiaTinaModule, _resolve_pivot_template_path


def _workbook_zip_entries(path: str, prefix: str) -> list[str]:
    with zipfile.ZipFile(path) as archive:
        return sorted(name for name in archive.namelist() if name.startswith(prefix))


def _pivot_cache_sources(path: str) -> list[tuple[str, str | None, str | None]]:
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    sources: list[tuple[str, str | None, str | None]] = []
    with zipfile.ZipFile(path) as archive:
        cache_entries = sorted(
            name
            for name in archive.namelist()
            if name.startswith("xl/pivotCache/pivotCacheDefinition") and name.endswith(".xml")
        )
        for entry in cache_entries:
            root = ElementTree.fromstring(archive.read(entry))
            source = root.find(".//main:worksheetSource", namespace)
            sources.append(
                (
                    entry,
                    source.attrib.get("sheet") if source is not None else None,
                    source.attrib.get("ref") if source is not None else None,
                )
            )
    return sources


def _pivot_cache_refresh_values(path: str) -> list[str | None]:
    values: list[str | None] = []
    with zipfile.ZipFile(path) as archive:
        cache_entries = sorted(
            name
            for name in archive.namelist()
            if name.startswith("xl/pivotCache/pivotCacheDefinition") and name.endswith(".xml")
        )
        for entry in cache_entries:
            root = ElementTree.fromstring(archive.read(entry))
            values.append(root.attrib.get("refreshOnLoad"))
    return values


def _pivot_cache_record_counts(path: str) -> list[str | None]:
    values: list[str | None] = []
    with zipfile.ZipFile(path) as archive:
        cache_entries = sorted(
            name
            for name in archive.namelist()
            if name.startswith("xl/pivotCache/pivotCacheDefinition") and name.endswith(".xml")
        )
        for entry in cache_entries:
            root = ElementTree.fromstring(archive.read(entry))
            values.append(root.attrib.get("recordCount"))
    return values


def _pivot_cache_records_counts(path: str) -> list[str]:
    counts: list[str] = []
    with zipfile.ZipFile(path) as archive:
        records_entries = sorted(
            name
            for name in archive.namelist()
            if name.startswith("xl/pivotCache/pivotCacheRecords") and name.endswith(".xml")
        )
        for entry in records_entries:
            root = ElementTree.fromstring(archive.read(entry))
            counts.append(root.attrib["count"])
    return counts


def _table_refs(path: str) -> list[str]:
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    refs: list[str] = []
    with zipfile.ZipFile(path) as archive:
        table_entries = sorted(
            name for name in archive.namelist() if name.startswith("xl/tables/table") and name.endswith(".xml")
        )
        for entry in table_entries:
            root = ElementTree.fromstring(archive.read(entry))
            refs.append(root.attrib["ref"])
            auto_filter = root.find("main:autoFilter", namespace)
            if auto_filter is not None:
                refs.append(auto_filter.attrib["ref"])
    return refs


def _workbook_calc_pr(path: str) -> dict[str, str]:
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read("xl/workbook.xml"))
    calc_pr = root.find("main:calcPr", namespace)
    return calc_pr.attrib if calc_pr is not None else {}


def _zip_entry_text(path: str, entry: str) -> str:
    with zipfile.ZipFile(path) as archive:
        return archive.read(entry).decode("utf-8")


def _sheet_cell_text(path: str, entry: str) -> str:
    return _zip_entry_text(path, entry)


def _pivot_data_field_names(path: str, entry: str) -> list[str | None]:
    namespace = {"main": "http://schemas.openxmlformats.org/spreadsheetml/2006/main"}
    with zipfile.ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(entry))
    return [
        field.attrib.get("name")
        for field in root.findall(".//main:dataFields/main:dataField", namespace)
    ]


def _slicer_pivot_table_names(path: str, entry: str) -> list[str | None]:
    namespace = {"x14": "http://schemas.microsoft.com/office/spreadsheetml/2009/9/main"}
    with zipfile.ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(entry))
    return [
        pivot_table.attrib.get("name")
        for pivot_table in root.findall(".//x14:pivotTable", namespace)
    ]


def _worksheet_relationship_types(path: str, entry: str) -> list[str]:
    namespace = {"rel": "http://schemas.openxmlformats.org/package/2006/relationships"}
    with zipfile.ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(entry))
    return [
        relationship.attrib.get("Type", "")
        for relationship in root.findall("rel:Relationship", namespace)
    ]


def _drawing_anchor_markers(path: str, entry: str) -> dict[str, dict[str, tuple[int, int]]]:
    namespace = {"xdr": "http://schemas.openxmlformats.org/drawingml/2006/spreadsheetDrawing"}
    with zipfile.ZipFile(path) as archive:
        root = ElementTree.fromstring(archive.read(entry))

    anchors: dict[str, dict[str, tuple[int, int]]] = {}
    for anchor in root:
        name_node = anchor.find(".//xdr:cNvPr", namespace)
        from_node = anchor.find("xdr:from", namespace)
        to_node = anchor.find("xdr:to", namespace)
        if name_node is None or from_node is None or to_node is None:
            continue

        def marker(node: ElementTree.Element) -> tuple[int, int]:
            col_node = node.find("xdr:col", namespace)
            row_node = node.find("xdr:row", namespace)
            return (
                int(col_node.text or "0") if col_node is not None else 0,
                int(row_node.text or "0") if row_node is not None else 0,
            )

        anchors[name_node.attrib.get("name", "")] = {
            "from": marker(from_node),
            "to": marker(to_node),
        }
    return anchors


class SophiaTinaTemplatePathTests(unittest.TestCase):
    def test_pivot_template_path_can_be_overridden_for_runtime_packaging(self):
        original_override = os.environ.get("TOS_SOPHIA_TINA_TEMPLATE_PATH")
        try:
            with tempfile.TemporaryDirectory() as temp_dir:
                template_path = Path(temp_dir) / "sophia_tina_pivot_template.xlsx"
                template_path.write_bytes(b"template")
                os.environ["TOS_SOPHIA_TINA_TEMPLATE_PATH"] = str(template_path)

                self.assertEqual(_resolve_pivot_template_path(), str(template_path))
        finally:
            if original_override is None:
                os.environ.pop("TOS_SOPHIA_TINA_TEMPLATE_PATH", None)
            else:
                os.environ["TOS_SOPHIA_TINA_TEMPLATE_PATH"] = original_override


class SophiaTinaModulePricePriorityTests(unittest.TestCase):
    def setUp(self):
        self.module = SophiaTinaModule()

    def _write_excel(self, directory: Path, name: str, rows: list[dict]) -> str:
        path = directory / name
        pd.DataFrame(rows).to_excel(path, index=False)
        return str(path)

    def test_country_analysis_summary_blocks_are_formula_driven_plain_tables(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            work_dir = Path(temp_dir)
            tms_rows = [
                {
                    "Factory": "FACTORY-A",
                    "PO Number": "PO-CN",
                    "Working Number": "WN-CN",
                    "Article Number": "ART-CN",
                    "Article Description": "China row",
                    "Customer Request Date (CRD)": "2026-01-01",
                    "PODD": "2026-02-01",
                    "Shipment Method": "Ocean",
                    "Gps Customer Number": "GPS-CN",
                    "Country/Region": "CHINA",
                    "Ordered Quantity": 10,
                },
                {
                    "Factory": "FACTORY-A",
                    "PO Number": "PO-US",
                    "Working Number": "WN-US",
                    "Article Number": "ART-US",
                    "Article Description": "US row",
                    "Customer Request Date (CRD)": "2026-01-02",
                    "PODD": "2026-02-02",
                    "Shipment Method": "Air",
                    "Gps Customer Number": "GPS-US",
                    "Country/Region": "UNITED STATES",
                    "Ordered Quantity": 30,
                },
                {
                    "Factory": "FACTORY-A",
                    "PO Number": "PO-OTHER",
                    "Working Number": "WN-OTHER",
                    "Article Number": "ART-OTHER",
                    "Article Description": "Other row",
                    "Customer Request Date (CRD)": "2026-01-03",
                    "PODD": "2026-02-03",
                    "Shipment Method": "Rail",
                    "Gps Customer Number": "GPS-OTHER",
                    "Country/Region": "BRAZIL",
                    "Ordered Quantity": 60,
                },
                {
                    "Factory": "FACTORY-B",
                    "PO Number": "PO-B-CN",
                    "Working Number": "WN-B-CN",
                    "Article Number": "ART-B-CN",
                    "Article Description": "Factory B China row",
                    "Customer Request Date (CRD)": "2026-03-01",
                    "PODD": "2026-04-01",
                    "Shipment Method": "Ocean",
                    "Gps Customer Number": "GPS-B-CN",
                    "Country/Region": "CHINA",
                    "Ordered Quantity": 5,
                },
                {
                    "Factory": "FACTORY-B",
                    "PO Number": "PO-B-OTHER",
                    "Working Number": "WN-B-OTHER",
                    "Article Number": "ART-B-OTHER",
                    "Article Description": "Factory B other row",
                    "Customer Request Date (CRD)": "2026-03-02",
                    "PODD": "2026-04-02",
                    "Shipment Method": "Ocean",
                    "Gps Customer Number": "GPS-B-OTHER",
                    "Country/Region": "MEXICO",
                    "Ordered Quantity": 15,
                },
            ]
            tms_price_rows = [
                {
                    "Working Number (M)": "WN-CN",
                    "Article Number (A)": "ART-CN",
                    "Season (M)": "SS26",
                    "Marketing Forecast (M)": 100,
                    "Milestone (C)": "Final",
                    "Intl. FOB (C)": 20.0,
                    "Factory Group Code (MF)": "FACTORY-A",
                },
                {
                    "Working Number (M)": "WN-US",
                    "Article Number (A)": "ART-US",
                    "Season (M)": "SS26",
                    "Marketing Forecast (M)": 100,
                    "Milestone (C)": "Final",
                    "Intl. FOB (C)": 10.0,
                    "Factory Group Code (MF)": "FACTORY-A",
                },
                {
                    "Working Number (M)": "WN-OTHER",
                    "Article Number (A)": "ART-OTHER",
                    "Season (M)": "SS26",
                    "Marketing Forecast (M)": 100,
                    "Milestone (C)": "Final",
                    "Intl. FOB (C)": 5.0,
                    "Factory Group Code (MF)": "FACTORY-A",
                },
                {
                    "Working Number (M)": "WN-B-CN",
                    "Article Number (A)": "ART-B-CN",
                    "Season (M)": "SS26",
                    "Marketing Forecast (M)": 100,
                    "Milestone (C)": "Final",
                    "Intl. FOB (C)": 8.0,
                    "Factory Group Code (MF)": "FACTORY-B",
                },
                {
                    "Working Number (M)": "WN-B-OTHER",
                    "Article Number (A)": "ART-B-OTHER",
                    "Season (M)": "SS26",
                    "Marketing Forecast (M)": 100,
                    "Milestone (C)": "Final",
                    "Intl. FOB (C)": 4.0,
                    "Factory Group Code (MF)": "FACTORY-B",
                },
            ]
            price_rows = [
                {
                    "Pack": "PACK-SS26",
                    "Season": "SS26",
                    "Working Number": row["Working Number"],
                    "Article Number": row["Article Number"],
                    "Factory": row["Factory"],
                    "Factory Price": 1.0,
                    "TMS Price": 1.0,
                }
                for row in tms_rows
            ]

            tms_path = self._write_excel(work_dir, "tms.xlsx", tms_rows)
            tms_price_path = self._write_excel(work_dir, "tms-price.xlsx", tms_price_rows)
            price_path = self._write_excel(work_dir, "price.xlsx", price_rows)

            result = self.module.process_reports(
                [tms_path],
                [tms_price_path],
                [price_path],
                output_dir=str(work_dir),
            )

            self.assertTrue(result["success"], result["message"])
            wb = load_workbook(result["output_path"], data_only=False, read_only=True)
            try:
                expected_rows = [
                    ["Season SS26", "Factory", "China Order", "USA Order", "Other Order", "Total Order"],
                    ["Qty(pcs)", "FACTORY-A", 10, 30, 60, 100],
                    ["Qty percentage", "FACTORY-A", 0.1, 0.3, 0.6, None],
                    ["Value(usd)", "FACTORY-A", 200, 300, 300, 800],
                    ["Value percentage", "FACTORY-A", 0.25, 0.375, 0.375, None],
                    ["Qty(pcs)", "FACTORY-B", 5, 0, 15, 20],
                    ["Qty percentage", "FACTORY-B", 0.25, 0, 0.75, None],
                    ["Value(usd)", "FACTORY-B", 40, 0, 60, 100],
                    ["Value percentage", "FACTORY-B", 0.4, 0, 0.6, None],
                ]
                for sheet_name, first_header in (
                    ("Country Analysis (S)", "Season SS26"),
                    ("Country Analysis (Y)", "Year 2026"),
                ):
                    ws = wb[sheet_name]
                    sheet_expected_rows = [list(row) for row in expected_rows]
                    sheet_expected_rows[0][0] = first_header
                    for offset, expected_row in enumerate(sheet_expected_rows):
                        excel_row = 1 + offset
                        actual = [
                            ws.cell(row=excel_row, column=column).value
                            for column in range(1, 7)
                        ]
                        if expected_row[0] in {"Qty(pcs)", "Value(usd)"}:
                            self.assertEqual(actual[:2], expected_row[:2])
                            self.assertTrue(str(actual[2]).startswith("=SUMIFS(Result!$"))
                            self.assertEqual(actual[5], f"=SUM(C{excel_row}:E{excel_row})")
                        elif expected_row[0] in {"Qty percentage", "Value percentage"}:
                            source_row = excel_row - 1
                            self.assertEqual(actual[:2], expected_row[:2])
                            self.assertEqual(actual[2], f"=IF($F{source_row}=0,0,C{source_row}/$F{source_row})")
                            self.assertEqual(actual[5], None)
                        else:
                            self.assertEqual(actual, expected_row)

                    self.assertEqual(ws.max_column, 6)
                    self.assertEqual(ws.max_row, len(sheet_expected_rows))

                self.assertIn(
                    'Result!$B:$B,"SS26"',
                    wb["Country Analysis (S)"].cell(row=2, column=3).value,
                )
                self.assertIn(
                    'Result!$H:$H,">="&DATE(2026,1,1)',
                    wb["Country Analysis (Y)"].cell(row=2, column=3).value,
                )
            finally:
                wb.close()

            self.assertNotIn("653448", _sheet_cell_text(result["output_path"], "xl/worksheets/sheet3.xml"))
            self.assertNotIn("3067514", _sheet_cell_text(result["output_path"], "xl/worksheets/sheet4.xml"))
            self.assertNotIn(
                "pivotTable",
                "\n".join(_worksheet_relationship_types(result["output_path"], "xl/worksheets/_rels/sheet3.xml.rels")),
            )
            self.assertNotIn(
                "pivotTable",
                "\n".join(_worksheet_relationship_types(result["output_path"], "xl/worksheets/_rels/sheet4.xml.rels")),
            )
            self.assertEqual(
                {source for _, sheet_name, source in _pivot_cache_sources(result["output_path"]) if sheet_name == "Result"},
                {"A1:Q6"},
            )
            self.assertEqual(set(_pivot_cache_refresh_values(result["output_path"])), {"1"})

    def test_country_analysis_blank_season_formula_matches_result_blank_cells(self):
        summary_rows = self.module._country_summary_block_rows(
            [
                {
                    "Season": "",
                    "Factory": "FACTORY-A",
                    "Country/Region": "CHINA",
                    "Quantity": 10,
                    "TMS Amount(USD)": 200,
                }
            ],
            "Season",
            1,
        )

        self.assertEqual(summary_rows[0][1][0], "Season (blank)")
        self.assertIn('Result!$B:$B,""', summary_rows[1][2][2])
        self.assertIn('Result!$B:$B,""', summary_rows[3][2][2])
        self.assertNotIn("(blank)", summary_rows[1][2][2])

    def test_factory_price_falls_back_to_working_and_factory_when_result_season_is_blank(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            work_dir = Path(temp_dir)
            tms_path = self._write_excel(
                work_dir,
                "tms.xlsx",
                [
                    {
                        "Factory": "FACTORY-A",
                        "PO Number": "PO-FALLBACK",
                        "Working Number": "WN-FALLBACK",
                        "Article Number": "ART-MISSING",
                        "Article Description": "Fallback price row",
                        "Customer Request Date (CRD)": "2026-01-01",
                        "PODD": "2026-02-01",
                        "Shipment Method": "Ocean",
                        "Gps Customer Number": "GPS-1",
                        "Country/Region": "CHINA",
                        "Ordered Quantity": 10,
                    },
                    {
                        "Factory": "FACTORY-B",
                        "PO Number": "PO-CONFLICT",
                        "Working Number": "WN-CONFLICT",
                        "Article Number": "ART-CONFLICT",
                        "Article Description": "Fallback conflict row",
                        "Customer Request Date (CRD)": "2026-01-02",
                        "PODD": "2026-02-02",
                        "Shipment Method": "Air",
                        "Gps Customer Number": "GPS-2",
                        "Country/Region": "UNITED STATES",
                        "Ordered Quantity": 5,
                    },
                    {
                        "Factory": "FACTORY-C",
                        "PO Number": "PO-NO-PACK",
                        "Working Number": "WN-NO-PACK",
                        "Article Number": "ART-NO-PACK",
                        "Article Description": "No pack source row",
                        "Customer Request Date (CRD)": "2026-01-03",
                        "PODD": "2026-02-03",
                        "Shipment Method": "Rail",
                        "Gps Customer Number": "GPS-3",
                        "Country/Region": "BRAZIL",
                        "Ordered Quantity": 8,
                    },
                ],
            )
            tms_price_path = self._write_excel(
                work_dir,
                "tms-price.xlsx",
                [
                    {
                        "Working Number (M)": "WN-UNRELATED",
                        "Article Number (A)": "ART-UNRELATED",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 100,
                        "Milestone (C)": "Final",
                        "Intl. FOB (C)": 20.0,
                        "Factory Group Code (MF)": "FACTORY-Z",
                    },
                ],
            )
            price_path = self._write_excel(
                work_dir,
                "price.xlsx",
                [
                    {
                        "Pack": "PACK-FALLBACK",
                        "Season": "SS26",
                        "Working Number": "WN-FALLBACK",
                        "Article Number": "ART-PRICE-A",
                        "Factory": "FACTORY-A",
                        "Factory Price": 9.0,
                        "TMS Price": 19.0,
                    },
                    {
                        "Pack": "PACK-CONFLICT-A",
                        "Season": "SS26",
                        "Working Number": "WN-CONFLICT",
                        "Article Number": "ART-PRICE-B1",
                        "Factory": "FACTORY-B",
                        "Factory Price": 5.0,
                        "TMS Price": 15.0,
                    },
                    {
                        "Pack": "PACK-CONFLICT-B",
                        "Season": "FW26",
                        "Working Number": "WN-CONFLICT",
                        "Article Number": "ART-PRICE-B2",
                        "Factory": "FACTORY-B",
                        "Factory Price": 7.0,
                        "TMS Price": 17.0,
                    },
                ],
            )

            result = self.module.process_reports(
                [tms_path],
                [tms_price_path],
                [price_path],
                output_dir=str(work_dir),
            )

            self.assertTrue(result["success"], result["message"])
            wb = load_workbook(result["output_path"], data_only=False, read_only=True)
            try:
                result_ws = wb["Result"]
                headers = [cell.value for cell in result_ws[1]]
                column_by_name = {name: index + 1 for index, name in enumerate(headers)}
                rows_by_working = {
                    result_ws.cell(row=row_index, column=column_by_name["Working Number"]).value: row_index
                    for row_index in range(2, result_ws.max_row + 1)
                }

                fallback_row = rows_by_working["WN-FALLBACK"]
                self.assertEqual(
                    result_ws.cell(fallback_row, column=column_by_name["Season"]).value,
                    None,
                )
                self.assertEqual(
                    result_ws.cell(fallback_row, column=column_by_name["Pack"]).value,
                    "PACK-FALLBACK",
                )
                self.assertEqual(
                    result_ws.cell(fallback_row, column=column_by_name["Factory Price(USD)"]).value,
                    9.0,
                )
                self.assertEqual(
                    result_ws.cell(fallback_row, column=column_by_name["TMS Price(USD)"]).value,
                    19.0,
                )

                conflict_row = rows_by_working["WN-CONFLICT"]
                self.assertEqual(
                    result_ws.cell(conflict_row, column=column_by_name["Factory Price(USD)"]).value,
                    5.0,
                )
                self.assertEqual(
                    result_ws.cell(conflict_row, column=column_by_name["Pack"]).value,
                    "PACK-CONFLICT-A",
                )
                self.assertEqual(
                    result_ws.cell(conflict_row, column=column_by_name["Factory Price(USD)"]).fill.fgColor.rgb,
                    "FFFFF2CC",
                )

                no_pack_row = rows_by_working["WN-NO-PACK"]
                self.assertEqual(
                    result_ws.cell(no_pack_row, column=column_by_name["Pack"]).value,
                    None,
                )
                self.assertEqual(
                    result_ws.cell(no_pack_row, column=column_by_name["Factory Price(USD)"]).value,
                    None,
                )

                diagnostics = wb["Diagnostics"]
                diagnostics_text = "\n".join(
                    str(diagnostics.cell(row=row, column=2).value)
                    for row in range(2, diagnostics.max_row + 1)
                )
                self.assertIn("FACTORY_PRICE_CONFLICT", diagnostics_text)
                self.assertIn("PACK_MISSING", diagnostics_text)
            finally:
                wb.close()

    def test_tms_price_uses_milestone_priority_order_instead_of_source_row_order(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            work_dir = Path(temp_dir)
            tms_path = self._write_excel(
                work_dir,
                "tms.xlsx",
                [
                    {
                        "Factory": "FACTORY-A",
                        "Working Number": "WN-FINAL",
                        "Article Number": "ART-F",
                        "Article Description": "Final article",
                        "Customer Request Date (CRD)": "2026-06-01",
                        "PODD": "2026-07-01",
                        "Gps Customer Number": "GPS-1",
                        "Country/Region": "US",
                        "Ordered Quantity": 2,
                    },
                    {
                        "Factory": "FACTORY-B",
                        "Working Number": "WN-P2",
                        "Article Number": "ART-P2",
                        "Article Description": "P2 article",
                        "Customer Request Date (CRD)": "2026-06-02",
                        "PODD": "2026-07-02",
                        "Gps Customer Number": "GPS-2",
                        "Country/Region": "CN",
                        "Ordered Quantity": 3,
                    },
                    {
                        "Factory": "FACTORY-C",
                        "Working Number": "WN-P1",
                        "Article Number": "ART-P1",
                        "Article Description": "P1 article",
                        "Customer Request Date (CRD)": "2026-06-03",
                        "PODD": "2026-07-03",
                        "Gps Customer Number": "GPS-3",
                        "Country/Region": "BR",
                        "Ordered Quantity": 4,
                    },
                    {
                        "Factory": "FACTORY-D",
                        "Working Number": "WN-PREC",
                        "Article Number": "ART-PREC",
                        "Article Description": "PREC article",
                        "Customer Request Date (CRD)": "2026-06-04",
                        "PODD": "2026-07-04",
                        "Gps Customer Number": "GPS-4",
                        "Country/Region": "ZA",
                        "Ordered Quantity": 5,
                    },
                ],
            )
            article_path = self._write_excel(
                work_dir,
                "article.xlsx",
                [
                    {
                        "Working Number (M)": "WN-FINAL",
                        "Article Number (A)": "ART-F",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 100,
                        "Milestone (C)": "P2",
                        "Intl. FOB (C)": 9.99,
                    },
                    {
                        "Working Number (M)": "WN-FINAL",
                        "Article Number (A)": "ART-F",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 100,
                        "Milestone (C)": "P1",
                        "Intl. FOB (C)": 7.5,
                    },
                    {
                        "Working Number (M)": "WN-FINAL",
                        "Article Number (A)": "ART-F",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 100,
                        "Milestone (C)": "Final",
                        "Intl. FOB (C)": 15.25,
                    },
                    {
                        "Working Number (M)": "WN-P2",
                        "Article Number (A)": "ART-P2",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 200,
                        "Milestone (C)": "P1",
                        "Intl. FOB (C)": 8.75,
                    },
                    {
                        "Working Number (M)": "WN-P2",
                        "Article Number (A)": "ART-P2",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 200,
                        "Milestone (C)": "PREC",
                        "Intl. FOB (C)": 3.0,
                    },
                    {
                        "Working Number (M)": "WN-P2",
                        "Article Number (A)": "ART-P2",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 200,
                        "Milestone (C)": "P2",
                        "Intl. FOB (C)": 12.5,
                    },
                    {
                        "Working Number (M)": "WN-P1",
                        "Article Number (A)": "ART-P1",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 300,
                        "Milestone (C)": "PREC",
                        "Intl. FOB (C)": 4.5,
                    },
                    {
                        "Working Number (M)": "WN-P1",
                        "Article Number (A)": "ART-P1",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 300,
                        "Milestone (C)": "P1",
                        "Intl. FOB (C)": 13.0,
                    },
                    {
                        "Working Number (M)": "WN-PREC",
                        "Article Number (A)": "ART-PREC",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 400,
                        "Milestone (C)": "PREC",
                        "Intl. FOB (C)": 6.75,
                    },
                ],
            )
            price_path = self._write_excel(
                work_dir,
                "price.xlsx",
                [
                    {
                        "Pack": "PACK-SS26",
                        "Season": "SS26",
                        "Working Number": "WN-FINAL",
                        "Article Number": "ART-F",
                        "Factory": "FACTORY-A",
                        "Factory Price": 11.0,
                        "TMS Price": 14.0,
                    },
                    {
                        "Pack": "PACK-SS26",
                        "Season": "SS26",
                        "Working Number": "WN-P2",
                        "Article Number": "ART-P2",
                        "Factory": "FACTORY-B",
                        "Factory Price": 7.0,
                        "TMS Price": 11.0,
                    },
                    {
                        "Pack": "PACK-SS26",
                        "Season": "SS26",
                        "Working Number": "WN-P1",
                        "Article Number": "ART-P1",
                        "Factory": "FACTORY-C",
                        "Factory Price": 8.0,
                        "TMS Price": 12.0,
                    },
                    {
                        "Pack": "PACK-SS26",
                        "Season": "SS26",
                        "Working Number": "WN-PREC",
                        "Article Number": "ART-PREC",
                        "Factory": "FACTORY-D",
                        "Factory Price": 9.0,
                        "TMS Price": 10.0,
                    },
                ],
            )

            result = self.module.process_reports(
                [tms_path],
                [article_path],
                [price_path],
                output_dir=str(work_dir),
            )

            self.assertTrue(result["success"], result["message"])
            wb = load_workbook(result["output_path"], data_only=False, read_only=True)
            ws = wb["Result"]
            try:
                headers = [cell.value for cell in ws[1]]
                column_by_name = {name: index + 1 for index, name in enumerate(headers)}
                self.assertEqual(
                    headers,
                    [
                        "Pack",
                        "Season",
                        "Factory",
                        "Working Number",
                        "Article Number",
                        "Article Name",
                        "CRD",
                        "PODD",
                        "Gps Customer Number",
                        "Country/Region",
                        "Shipment Method",
                        "Marketing Forecast(M)",
                        "Quantity",
                        "Factory Price(USD)",
                        "Factory Amount(USD)",
                        "TMS Price(USD)",
                        "TMS Amount(USD)",
                    ],
                )
                rows_by_working = {
                    ws.cell(row=row_index, column=column_by_name["Working Number"]).value: row_index
                    for row_index in range(2, ws.max_row + 1)
                }

                final_row = rows_by_working["WN-FINAL"]
                self.assertEqual(
                    ws.cell(final_row, column=column_by_name["TMS Price(USD)"]).value,
                    15.25,
                )
                self.assertEqual(
                    ws.cell(final_row, column=column_by_name["TMS Amount(USD)"]).value,
                    "=P2*M2",
                )
                self.assertEqual(ws.cell(final_row, column=column_by_name["Pack"]).value, "PACK-SS26")

                p2_row = rows_by_working["WN-P2"]
                self.assertEqual(
                    ws.cell(p2_row, column=column_by_name["TMS Price(USD)"]).value,
                    12.5,
                )
                self.assertEqual(
                    ws.cell(p2_row, column=column_by_name["TMS Amount(USD)"]).value,
                    "=P3*M3",
                )

                p1_row = rows_by_working["WN-P1"]
                self.assertEqual(
                    ws.cell(p1_row, column=column_by_name["TMS Price(USD)"]).value,
                    13.0,
                )
                self.assertEqual(
                    ws.cell(p1_row, column=column_by_name["TMS Amount(USD)"]).value,
                    "=P4*M4",
                )

                prec_row = rows_by_working["WN-PREC"]
                self.assertEqual(
                    ws.cell(prec_row, column=column_by_name["TMS Price(USD)"]).value,
                    6.75,
                )
                self.assertEqual(
                    ws.cell(prec_row, column=column_by_name["TMS Amount(USD)"]).value,
                    "=P5*M5",
                )
            finally:
                wb.close()

    def test_optional_allocation_and_shipment_files_update_result_and_summaries(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            work_dir = Path(temp_dir)
            tms_path = self._write_excel(
                work_dir,
                "tms.xlsx",
                [
                    {
                        "Factory": "FACTORY-SOURCE",
                        "PO Number": "PO-1",
                        "Working Number": "WN-1",
                        "Article Number": "ART-1",
                        "Article Description": "Article one",
                        "Customer Request Date (CRD)": "2026-01-10",
                        "PODD": "2026-02-20",
                        "Shipment Method": "Ocean",
                        "Gps Customer Number": "GPS-1",
                        "Country/Region": "CHINA",
                        "Ordered Quantity": 10,
                    },
                    {
                        "Factory": "FACTORY-SOURCE",
                        "PO Number": "PO-2",
                        "Working Number": "WN-2",
                        "Article Number": "ART-2",
                        "Article Description": "Article two",
                        "Customer Request Date (CRD)": "2026-01-11",
                        "PODD": "2026-03-31",
                        "Shipment Method": "Ocean",
                        "Gps Customer Number": "GPS-2",
                        "Country/Region": "UNITED STATES",
                        "Ordered Quantity": 12,
                    },
                    {
                        "Factory": "FACTORY-SOURCE",
                        "PO Number": "PO-2",
                        "Working Number": "WN-3",
                        "Article Number": "ART-3",
                        "Article Description": "Article three",
                        "Customer Request Date (CRD)": "2026-01-12",
                        "PODD": "2026-03-31",
                        "Shipment Method": "Ocean",
                        "Gps Customer Number": "GPS-3",
                        "Country/Region": "BRAZIL",
                        "Ordered Quantity": 8,
                    },
                ],
            )
            tms_price_path = self._write_excel(
                work_dir,
                "tms-price.xlsx",
                [
                    {
                        "Working Number (M)": "WN-1",
                        "Article Number (A)": "ART-1",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 100,
                        "Milestone (C)": "Final",
                        "Intl. FOB (C)": 15.0,
                        "Factory Group Code (MF)": "FACTORY-ALLOC",
                    },
                    {
                        "Working Number (M)": "WN-2",
                        "Article Number (A)": "ART-2",
                        "Season (M)": "FW26",
                        "Marketing Forecast (M)": 200,
                        "Milestone (C)": "Final",
                        "Intl. FOB (C)": 20.0,
                        "Factory Group Code (MF)": "FACTORY-SOURCE",
                    },
                    {
                        "Working Number (M)": "WN-3",
                        "Article Number (A)": "ART-3",
                        "Season (M)": "FW26",
                        "Marketing Forecast (M)": 300,
                        "Milestone (C)": "Final",
                        "Intl. FOB (C)": 30.0,
                        "Factory Group Code (MF)": "FACTORY-SOURCE",
                    },
                ],
            )
            price_path = self._write_excel(
                work_dir,
                "price.xlsx",
                [
                    {
                        "Pack": "PACK-SS26",
                        "Season": "SS26",
                        "Working Number": "WN-1",
                        "Article Number": "ART-1",
                        "Factory": "FACTORY-ALLOC",
                        "Factory Price": 10.0,
                        "TMS Price": 15.0,
                    },
                    {
                        "Pack": "PACK-FW26",
                        "Season": "FW26",
                        "Working Number": "WN-2",
                        "Article Number": "ART-2",
                        "Factory": "FACTORY-SOURCE",
                        "Factory Price": 12.0,
                        "TMS Price": 20.0,
                    },
                    {
                        "Pack": "PACK-FW26",
                        "Season": "FW26",
                        "Working Number": "WN-3",
                        "Article Number": "ART-3",
                        "Factory": "FACTORY-SOURCE",
                        "Factory Price": 13.0,
                        "TMS Price": 30.0,
                    },
                ],
            )
            allocation_path = self._write_excel(
                work_dir,
                "allocation.xlsx",
                [
                    {"Allocation": "FACTORY-ALLOC", "PO": "PO-1"},
                ],
            )
            shipment_path = self._write_excel(
                work_dir,
                "shipment.xlsx",
                [
                    {
                        "purchasingdocument": "PO-1",
                        "shippinginstruction_desc": "AIR",
                        "po_delivery_date": "2026-02-01",
                        "pord_order_qty": 4,
                    },
                    {
                        "purchasingdocument": "PO-1",
                        "shippinginstruction_desc": "Ocean",
                        "po_delivery_date": "2026-02-15",
                        "pord_order_qty": 6,
                    },
                    {
                        "purchasingdocument": "PO-2",
                        "shippinginstruction_desc": "Rail",
                        "po_delivery_date": "2026-04-01",
                        "pord_order_qty": 20,
                    },
                ],
            )

            result = self.module.process_reports(
                [tms_path],
                [tms_price_path],
                [price_path],
                output_dir=str(work_dir),
                allocation_paths=[allocation_path],
                shipment_method_paths=[shipment_path],
            )

            self.assertTrue(result["success"], result["message"])
            wb = load_workbook(result["output_path"], data_only=False, read_only=True)
            try:
                self.assertEqual(
                    wb.sheetnames[:9],
                    [
                        "Monthly Summary (By Fty)",
                        "Seasonal Summary (By Fty)",
                        "Country Analysis (S)",
                        "Country Analysis (Y)",
                        "Y2Y Comparison(Qtty & Value)",
                        "Fty Order Analysis",
                        "Ship Method Analysis",
                        "Development Style Qty",
                        "S2S Development Analysis",
                    ],
                )
                ws = wb["Result"]
                headers = [cell.value for cell in ws[1]]
                column_by_name = {name: index + 1 for index, name in enumerate(headers)}
                self.assertEqual(ws.max_row, 5)
                result_rows = [
                    {
                        header: ws.cell(row=row_index, column=column_by_name[header]).value
                        for header in headers
                    }
                    for row_index in range(2, ws.max_row + 1)
                ]
                po1_rows = [row for row in result_rows if row["Working Number"] == "WN-1"]
                self.assertEqual([row["Quantity"] for row in po1_rows], [4, 6])
                self.assertEqual({row["Factory"] for row in po1_rows}, {"FACTORY-ALLOC"})
                self.assertEqual([row["Shipment Method"] for row in po1_rows], ["AIR", "Ocean"])
                self.assertEqual(
                    [str(row["PODD"])[:10] for row in po1_rows],
                    ["2026-02-01", "2026-02-15"],
                )

                po2_rows = [row for row in result_rows if row["Working Number"] in {"WN-2", "WN-3"}]
                self.assertEqual([row["Quantity"] for row in po2_rows], [12, 8])
                self.assertEqual({row["Shipment Method"] for row in po2_rows}, {"Ocean"})

                self.assertEqual(ws["O2"].value, "=N2*M2")
                self.assertEqual(ws["Q2"].value, "=P2*M2")

                diagnostics = wb["Diagnostics"]
                diagnostics_text = "\n".join(
                    str(diagnostics.cell(row=row, column=2).value)
                    for row in range(2, diagnostics.max_row + 1)
                )
                self.assertIn("SHIPMENT_PO_MULTI_COMBO", diagnostics_text)
            finally:
                wb.close()

            pivot_entries = _workbook_zip_entries(result["output_path"], "xl/pivotTables/pivotTable")
            self.assertEqual(len(pivot_entries), 9)
            self.assertTrue(
                _workbook_zip_entries(result["output_path"], "xl/pivotCache/pivotCacheDefinition")
            )
            self.assertEqual(_table_refs(result["output_path"]), ["A1:Q5", "A1:Q5"])
            self.assertEqual(
                {source for _, sheet_name, source in _pivot_cache_sources(result["output_path"]) if sheet_name == "Result"},
                {"A1:Q5"},
            )
            self.assertEqual(set(_pivot_cache_refresh_values(result["output_path"])), {"1"})
            self.assertEqual(set(_pivot_cache_record_counts(result["output_path"])), {"0"})
            self.assertEqual(set(_pivot_cache_records_counts(result["output_path"])), {"0"})
            self.assertFalse(_workbook_zip_entries(result["output_path"], "xl/calcChain.xml"))
            self.assertEqual(_workbook_calc_pr(result["output_path"]).get("fullCalcOnLoad"), "1")
            self.assertEqual(_workbook_calc_pr(result["output_path"]).get("forceFullCalc"), "1")
            y2y_drawing_anchors = _drawing_anchor_markers(result["output_path"], "xl/drawings/drawing1.xml")
            self.assertEqual(y2y_drawing_anchors["Factory"], {"from": (6, 1), "to": (8, 7)})
            self.assertEqual(y2y_drawing_anchors["Chart 1"], {"from": (9, 1), "to": (17, 16)})

            content_types_xml = _zip_entry_text(result["output_path"], "[Content_Types].xml")
            self.assertIn(
                '<Types xmlns="http://schemas.openxmlformats.org/package/2006/content-types">',
                content_types_xml,
            )
            self.assertNotIn("<ns0:Types", content_types_xml)

            workbook_rels_xml = _zip_entry_text(result["output_path"], "xl/_rels/workbook.xml.rels")
            self.assertIn(
                '<Relationships xmlns="http://schemas.openxmlformats.org/package/2006/relationships">',
                workbook_rels_xml,
            )
            self.assertNotIn("<ns0:Relationships", workbook_rels_xml)

            workbook_xml = _zip_entry_text(result["output_path"], "xl/workbook.xml")
            self.assertIn('mc:Ignorable="x15"', workbook_xml)
            self.assertIn(
                'xmlns:x15="http://schemas.microsoft.com/office/spreadsheetml/2010/11/main"',
                workbook_xml,
            )

    def test_0617_report_contracts_use_grouped_price_helpers_and_updated_pivots(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            work_dir = Path(temp_dir)
            tms_path = self._write_excel(
                work_dir,
                "tms.xlsx",
                [
                    {
                        "Factory": "FACTORY-A",
                        "PO Number": "PO-PRICE",
                        "Working Number": "WN-PRICE",
                        "Article Number": "ART-RESULT",
                        "Article Description": "Price grouping row",
                        "Customer Request Date (CRD)": "2026-01-01",
                        "PODD": "2026-01-10",
                        "Shipment Method": "Ocean",
                        "Gps Customer Number": "GPS-1",
                        "Country/Region": "BRAZIL",
                        "Ordered Quantity": 10,
                    },
                    {
                        "Factory": "FACTORY-A",
                        "PO Number": "PO-US",
                        "Working Number": "WN-US",
                        "Article Number": "ART-US",
                        "Article Description": "US row",
                        "Customer Request Date (CRD)": "2026-01-02",
                        "PODD": "2026-01-11",
                        "Shipment Method": "Air",
                        "Gps Customer Number": "GPS-2",
                        "Country/Region": "UNITED STATES",
                        "Ordered Quantity": 5,
                    },
                    {
                        "Factory": "FACTORY-A",
                        "PO Number": "PO-CN",
                        "Working Number": "WN-CN",
                        "Article Number": "ART-CN",
                        "Article Description": "CN row",
                        "Customer Request Date (CRD)": "2026-01-03",
                        "PODD": "2026-01-12",
                        "Shipment Method": "Ocean",
                        "Gps Customer Number": "GPS-3",
                        "Country/Region": "CHINA",
                        "Ordered Quantity": 15,
                    },
                ],
            )
            tms_price_path = self._write_excel(
                work_dir,
                "tms-price.xlsx",
                [
                    {
                        "Working Number (M)": "WN-PRICE",
                        "Article Number (A)": "ART-RESULT",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 100,
                        "Milestone (C)": "Final",
                        "Intl. FOB (C)": 20.0,
                        "Factory Group Code (MF)": "FACTORY-A",
                    },
                    {
                        "Working Number (M)": "WN-US",
                        "Article Number (A)": "ART-US",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 50,
                        "Milestone (C)": "Final",
                        "Intl. FOB (C)": 30.0,
                        "Factory Group Code (MF)": "FACTORY-A",
                    },
                    {
                        "Working Number (M)": "WN-CN",
                        "Article Number (A)": "ART-CN",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 150,
                        "Milestone (C)": "Final",
                        "Intl. FOB (C)": 40.0,
                        "Factory Group Code (MF)": "FACTORY-A",
                    },
                    {
                        "Working Number (M)": "WN-DEV-ONLY",
                        "Article Number (A)": "ART-DEV",
                        "Season (M)": "SS26",
                        "Marketing Forecast (M)": 0,
                        "Milestone (C)": "Final",
                        "Intl. FOB (C)": 0.0,
                        "Factory Group Code (MF)": "FACTORY-A",
                    },
                ],
            )
            price_path = self._write_excel(
                work_dir,
                "price.xlsx",
                [
                    {
                        "Pack": "PACK-SS26",
                        "Season": "SS26",
                        "Working Number": "WN-PRICE",
                        "Article Number": "ART-PRICE-A",
                        "Factory": "FACTORY-A",
                        "Factory Price": 9.0,
                        "TMS Price": 19.0,
                    },
                    {
                        "Pack": "PACK-SS26",
                        "Season": "SS26",
                        "Working Number": "WN-PRICE",
                        "Article Number": "ART-PRICE-B",
                        "Factory": "FACTORY-A",
                        "Factory Price": 11.0,
                        "TMS Price": 21.0,
                    },
                    {
                        "Pack": "PACK-SS26",
                        "Season": "SS26",
                        "Working Number": "WN-US",
                        "Article Number": "ART-US",
                        "Factory": "FACTORY-A",
                        "Factory Price": 12.0,
                        "TMS Price": 30.0,
                    },
                    {
                        "Pack": "PACK-SS26",
                        "Season": "SS26",
                        "Working Number": "WN-CN",
                        "Article Number": "ART-CN",
                        "Factory": "FACTORY-A",
                        "Factory Price": 13.0,
                        "TMS Price": 40.0,
                    },
                ],
            )

            result = self.module.process_reports(
                [tms_path],
                [tms_price_path],
                [price_path],
                output_dir=str(work_dir),
            )

            self.assertTrue(result["success"], result["message"])
            wb = load_workbook(result["output_path"], data_only=False, read_only=True)
            try:
                result_ws = wb["Result"]
                headers = [cell.value for cell in result_ws[1]]
                column_by_name = {name: index + 1 for index, name in enumerate(headers)}
                rows_by_working = {
                    result_ws.cell(row=row_index, column=column_by_name["Working Number"]).value: row_index
                    for row_index in range(2, result_ws.max_row + 1)
                }
                price_row = rows_by_working["WN-PRICE"]
                self.assertEqual(
                    result_ws.cell(price_row, column=column_by_name["Pack"]).value,
                    "PACK-SS26",
                )
                self.assertEqual(
                    result_ws.cell(price_row, column=column_by_name["Factory Price(USD)"]).value,
                    9.0,
                )
                self.assertEqual(
                    result_ws.cell(price_row, column=column_by_name["Factory Price(USD)"]).fill.fgColor.rgb,
                    "FFFFF2CC",
                )

                diagnostics = wb["Diagnostics"]
                diagnostics_text = "\n".join(
                    str(diagnostics.cell(row=row, column=2).value)
                    for row in range(2, diagnostics.max_row + 1)
                )
                self.assertIn("FACTORY_PRICE_CONFLICT", diagnostics_text)

                development_ws = wb["Development Style Qty"]
                self.assertEqual(development_ws.max_row, 1)
                self.assertEqual(
                    [development_ws.cell(1, column).value for column in range(1, 4)],
                    ["Season", "Factory", "Development Style Count"],
                )

                self.assertEqual(wb["Country Analysis Source"].sheet_state, "hidden")
                country_headers = [cell.value for cell in wb["Country Analysis Source"][1]]
                country_group_column = country_headers.index("Country Group") + 1
                country_groups = {
                    wb["Country Analysis Source"].cell(row=row, column=country_group_column).value
                    for row in range(2, wb["Country Analysis Source"].max_row + 1)
                }
                self.assertEqual(country_groups, {"CHINA", "UNITED STATES", "OTHER COUNTRIES"})

                self.assertEqual(wb["Ship Method Source"].sheet_state, "hidden")
                ship_headers = [cell.value for cell in wb["Ship Method Source"][1]]
                self.assertIn("TMS Amount (USD)", ship_headers)
                self.assertNotIn("TMS Amount USD)", ship_headers)
                ship_rows = [
                    {
                        header: wb["Ship Method Source"].cell(row=row_index, column=index + 1).value
                        for index, header in enumerate(ship_headers)
                    }
                    for row_index in range(2, wb["Ship Method Source"].max_row + 1)
                ]
                ocean_row = next(row for row in ship_rows if row["Shipment Method"] == "Ocean")
                air_row = next(row for row in ship_rows if row["Shipment Method"] == "Air")
                self.assertAlmostEqual(ocean_row["Quantity (%)"], 25 / 30)
                self.assertAlmostEqual(air_row["Quantity (%)"], 5 / 30)

                self.assertEqual(wb["S2S Development Source"].sheet_state, "hidden")
                s2s_headers = [cell.value for cell in wb["S2S Development Source"][1]]
                self.assertEqual(
                    s2s_headers,
                    [
                        "Season",
                        "Factory",
                        "Development Style Count",
                        "Bulk Style Count",
                        "Bulk Qty (pcs)",
                    ],
                )
                s2s_row = {
                    header: wb["S2S Development Source"].cell(row=2, column=index + 1).value
                    for index, header in enumerate(s2s_headers)
                }
                self.assertEqual(s2s_row["Development Style Count"], 4)
                self.assertEqual(s2s_row["Bulk Style Count"], 3)
                self.assertEqual(s2s_row["Bulk Qty (pcs)"], 30)
            finally:
                wb.close()

            self.assertNotIn(
                "TMS Amount USD)",
                _zip_entry_text(result["output_path"], "xl/pivotTables/pivotTable8.xml"),
            )
            self.assertIn(
                "TMS Amount (USD)",
                _zip_entry_text(result["output_path"], "xl/pivotTables/pivotTable8.xml"),
            )
            self.assertIn(
                "Bulk Qty (pcs)",
                _zip_entry_text(result["output_path"], "xl/pivotCache/pivotCacheDefinition6.xml"),
            )
            self.assertIn(
                "Bulk Qty (pcs)",
                _zip_entry_text(result["output_path"], "xl/pivotTables/pivotTable9.xml"),
            )
            self.assertEqual(
                _pivot_data_field_names(result["output_path"], "xl/pivotTables/pivotTable9.xml")[:2],
                ["Development Style Count", "Bulk Style Count"],
            )
            self.assertEqual(
                set(_slicer_pivot_table_names(result["output_path"], "xl/slicerCaches/slicerCache1.xml")),
                {"PivotTable15", "PivotTable16"},
            )


if __name__ == "__main__":
    unittest.main()

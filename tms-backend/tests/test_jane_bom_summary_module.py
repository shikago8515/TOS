import os
import sys
import tempfile
import unittest

import openpyxl


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.jane_bom_summary_module import JaneBomSummaryModule


class JaneBomSummaryModuleTests(unittest.TestCase):
    def setUp(self):
        self.module = JaneBomSummaryModule()

    def _save_pack_workbook(self, folder, rows=None):
        path = os.path.join(folder, "Pack.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append(["Pack", "Season", "Working Number", "Merch"])
        for row in rows or [["SS26 ASOS Pack", "SS26", "RC2610OW007H2H", "Diana"]]:
            ws.append(row)
        wb.save(path)
        return path

    def _save_bom_workbook(
        self,
        folder,
        filename="BOM-CH512-1L8006-20261.xlsx",
        *,
        working="RC2610OW007H2H",
        season="SS26",
        factory="1L8006 | DANDONG SLT GARMENT INDUSTRY | B",
        left_article="LD1813",
        right_article="LD1812",
        material_descriptions=None,
    ):
        path = os.path.join(folder, filename)
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "CH512-1L8006-20261"

        ws["A1"] = "Working #"
        ws["B1"] = working
        ws["A3"] = "Season"
        ws["B3"] = season
        ws["A5"] = "Factory"
        ws["B5"] = factory
        ws["P9"] = f"DEFAULT GROUP - BLACK | {left_article} | IN RANGE"
        ws["R9"] = f"DEFAULT GROUP - CRYSTAL SKY S26 | {right_article} | IN RANGE"

        headers = [
            "Variation",
            "Part Name",
            "Part ID",
            "Part Group #",
            "Material Reference #",
            "Material Name",
            "Group Code Supplier",
            "Supplier Name",
            "Supplier Material ID",
            "Supplier Material Name",
            "Material Description",
            "Material Supplier Lifecycle",
            "Material Lifecycle Validate",
            "Part Remark",
            "Color Layer Name",
            "Color",
            "Color Treatments",
            "Color",
        ]
        for column_index, value in enumerate(headers, start=1):
            ws.cell(row=10, column=column_index, value=value)

        ws["B11"] = "MAIN COMPONENT"
        descriptions = material_descriptions or ["97%POLYESTER"]
        for index, material_description in enumerate(descriptions, start=1):
            material_ref = f"7002017{index}"
            ws.append([
                "",
                f"B main body {index}",
                "7155",
                str(9 + index),
                material_ref,
                f"{material_ref} Satin",
                "1SB001",
                "TMS (CHN)",
                f'="{material_ref}_1SB001"',
                "",
                material_description,
                "Released",
                "",
                "",
                "",
                "095A BLACK",
                "",
                "AE65 CRYSTAL SKY S26",
            ])
        trim_row = 12 + len(descriptions)
        ws.cell(row=trim_row, column=2, value="TRIM")
        ws.append([
            "",
            "B trim",
            "",
            "200",
            "80011226",
            "80011226 Tape",
            "1SB001",
            "TMS (CHN)",
            "",
            "",
            "Tape",
            "",
            "",
            "",
            "",
            "095A BLACK",
            "",
            "001A WHITE",
        ])
        wb.save(path)
        return path

    def _is_highlighted(self, cell):
        return cell.fill.patternType == "solid" and cell.fill.fgColor.rgb in {
            "00FFF2CC",
            "FFFFF2CC",
        }

    def test_generates_rows_from_main_component_with_pack_lookup(self):
        with tempfile.TemporaryDirectory() as folder:
            pack_path = self._save_pack_workbook(folder)
            bom_path = self._save_bom_workbook(folder)

            result = self.module.process_reports([bom_path], pack_path, folder)

            self.assertTrue(result["success"])
            self.assertEqual(result["row_count"], 2)

            output_wb = openpyxl.load_workbook(result["output_path"], data_only=False)
            ws = output_wb.active
            rows = list(ws.iter_rows(values_only=True))

            self.assertEqual(
                rows[0],
                (
                    "Pack",
                    "Working #",
                    "Season",
                    "Factory",
                    "Articles",
                    "Part Group #",
                    "Material Reference #",
                    "Material Name",
                    "Group Code Supplier",
                    "Supplier Name",
                    "Material Description",
                    "Color",
                ),
            )
            self.assertEqual(rows[1][0:6], (
                "SS26 ASOS Pack",
                "RC2610OW007H2H",
                "SS26",
                "1L8006 | DANDONG SLT GARMENT INDUSTRY | B",
                "LD1812",
                "10",
            ))
            self.assertEqual(rows[1][6:12], (
                "70020171",
                "70020171 Satin",
                "1SB001",
                "TMS (CHN)",
                "97%POLYESTER",
                "AE65 CRYSTAL SKY S26",
            ))
            self.assertEqual(rows[2][4], "LD1813")
            self.assertEqual(rows[2][11], "095A BLACK")

    def test_sorts_rows_and_highlights_material_description_missing_required_words(self):
        with tempfile.TemporaryDirectory() as folder:
            pack_path = self._save_pack_workbook(
                folder,
                rows=[
                    ["Pack B", "SS26", "W-B", "Diana"],
                    ["Pack A", "FW25", "W-A", "Diana"],
                ],
            )
            bom_b_path = self._save_bom_workbook(
                folder,
                "BOM-B.xlsx",
                working="W-B",
                season="SS26",
                factory="Factory B",
                left_article="ZZZ",
                right_article="AAA",
                material_descriptions=["100% COTTON"],
            )
            bom_a_path = self._save_bom_workbook(
                folder,
                "BOM-A.xlsx",
                working="W-A",
                season="FW25",
                factory="Factory A",
                left_article="ZZZ",
                right_article="AAA",
                material_descriptions=["97%POLYESTER", "contains recycled yarn", ""],
            )

            result = self.module.process_reports([bom_b_path, bom_a_path], pack_path, folder)

            self.assertTrue(result["success"])
            output_wb = openpyxl.load_workbook(result["output_path"], data_only=False)
            ws = output_wb.active
            rows = list(ws.iter_rows(values_only=True))
            header = rows[0]
            description_column = header.index("Material Description") + 1

            self.assertEqual(
                [
                    (
                        row[header.index("Articles")],
                        row[header.index("Factory")],
                        row[header.index("Working #")],
                        row[header.index("Pack")],
                        row[header.index("Season")],
                        row[header.index("Material Description")],
                    )
                    for row in rows[1:]
                ],
                [
                    ("AAA", "Factory A", "W-A", "Pack A", "FW25", "97%POLYESTER"),
                    ("AAA", "Factory A", "W-A", "Pack A", "FW25", "contains recycled yarn"),
                    ("AAA", "Factory A", "W-A", "Pack A", "FW25", None),
                    ("AAA", "Factory B", "W-B", "Pack B", "SS26", "100% COTTON"),
                    ("ZZZ", "Factory A", "W-A", "Pack A", "FW25", "97%POLYESTER"),
                    ("ZZZ", "Factory A", "W-A", "Pack A", "FW25", "contains recycled yarn"),
                    ("ZZZ", "Factory A", "W-A", "Pack A", "FW25", None),
                    ("ZZZ", "Factory B", "W-B", "Pack B", "SS26", "100% COTTON"),
                ],
            )

            highlight_by_description = {}
            for row_index in range(2, ws.max_row + 1):
                description = ws.cell(row=row_index, column=description_column).value
                highlight_by_description.setdefault(description, set()).add(
                    self._is_highlighted(ws.cell(row=row_index, column=description_column))
                )

            self.assertEqual(highlight_by_description["97%POLYESTER"], {True})
            self.assertEqual(highlight_by_description[None], {True})
            self.assertEqual(highlight_by_description["100% COTTON"], {False})
            self.assertEqual(highlight_by_description["contains recycled yarn"], {False})

    def test_rejects_ambiguous_pack_mapping(self):
        with tempfile.TemporaryDirectory() as folder:
            pack_path = self._save_pack_workbook(folder)
            wb = openpyxl.load_workbook(pack_path)
            ws = wb.active
            ws.append(["Different Pack", "SS26", "RC2610OW007H2H", "Diana"])
            wb.save(pack_path)

            bom_path = self._save_bom_workbook(folder)

            result = self.module.process_reports([bom_path], pack_path, folder)

            self.assertFalse(result["success"])
            self.assertIn("Pack 表存在重复映射", result["message"])


if __name__ == "__main__":
    unittest.main()

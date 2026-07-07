import os
import sys
import tempfile
from typing import Any, Dict, List
import unittest

import openpyxl
import pandas as pd


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.jane_module import JaneModule


class JaneModuleWorkbookTests(unittest.TestCase):
    def setUp(self):
        self.module = JaneModule()

    def _tms_row(
        self,
        working_number: str,
        article_number: str,
        country: str,
        customer_no: str,
        po_number: int,
        po_line: int,
        ordered_quantity: int,
    ) -> Dict[str, Any]:
        return {
            "Working Number": working_number,
            "PO Number": po_number,
            "Market PO Number": po_number,
            "PO Line Item #": po_line,
            "Company Code": "TMS",
            "Article Number": article_number,
            "Article Description": f"{article_number} color",
            "Customer Request Date (CRD)": "2026-07-01",
            "PODD": "2026-07-10",
            "Plant Code": "P1",
            "Customer Size Run": "",
            "Technical Notation": "",
            "Shipment Mode": "Ocean",
            "Gps Customer Number": customer_no,
            "Country/Region": country,
            "Technical Size": "36",
            "Customer Size": "36",
            "Ordered Quantity": ordered_quantity,
            "Factory": "F1",
        }

    def _country_rows(self, rows: List[Dict[str, Any]]) -> List[Dict[str, Any]]:
        return [
            {
                "CST NO": row["Gps Customer Number"],
                "DESTINATION": row["Country/Region"],
                "REGION": "EMEA" if row["Country/Region"] in {"SOUTH AFRICA", "USA", "CANADA"} else "APAC",
            }
            for row in rows
        ]

    def _process_rows(
        self,
        rows: List[Dict[str, Any]],
        working_filters: List[str] | None = None,
        country_rows: List[Dict[str, Any]] | None = None,
    ) -> openpyxl.Workbook:
        with tempfile.TemporaryDirectory() as temp_dir:
            tms_path = os.path.join(temp_dir, "Copy of TMS.xlsx")
            country_path = os.path.join(temp_dir, "country.xlsx")
            pd.DataFrame(rows).to_excel(tms_path, sheet_name="Result Set", index=False)
            pd.DataFrame(country_rows or self._country_rows(rows)).to_excel(country_path, index=False)

            result = self.module.process_reports(
                tms_path=tms_path,
                country_path=country_path,
                working_filters=working_filters,
                output_dir=temp_dir,
            )

            self.assertTrue(result["success"], result["message"])
            return openpyxl.load_workbook(result["output_path"], data_only=False)

    def _summary_working_article_rows(self, wb: openpyxl.Workbook) -> List[tuple[str, str]]:
        summary_rows: List[tuple[str, str]] = []
        current_working = None
        for row in wb["Summary"].iter_rows(min_row=3, max_col=3, values_only=True):
            if row[1]:
                current_working = row[1]
            if current_working and row[2]:
                summary_rows.append((current_working, row[2]))
        return summary_rows

    def _summary_working_order(self, wb: openpyxl.Workbook) -> List[str]:
        return [
            row[0]
            for row in wb["Summary"].iter_rows(min_row=3, min_col=2, max_col=2, values_only=True)
            if row[0] and row[0] != "TOTAL"
        ]

    def test_subtotal_po_qty_sums_po_qty_column_rows(self):
        df = pd.DataFrame(
            [
                {
                    "Working Number": "WN-001",
                    "PO Number": 823016,
                    "Market PO Number": 823016,
                    "PO Line Item #": 10,
                    "Company Code": "TMS",
                    "Article Number": "ART-1",
                    "Customer Request Date (CRD)": "2026-07-01",
                    "PODD": "2026-07-10",
                    "Plant Code": "P1",
                    "Customer Size Run": "",
                    "Technical Notation": "",
                    "Shipment Mode": "By Sea",
                    "Gps Customer Number": "848012",
                    "Technical Size": "A28",
                    "Customer Size": "A/XS",
                    "Ordered Quantity": 58,
                },
                {
                    "Working Number": "WN-001",
                    "PO Number": 823016,
                    "Market PO Number": 823016,
                    "PO Line Item #": 20,
                    "Company Code": "TMS",
                    "Article Number": "ART-1",
                    "Customer Request Date (CRD)": "2026-07-01",
                    "PODD": "2026-07-10",
                    "Plant Code": "P1",
                    "Customer Size Run": "",
                    "Technical Notation": "",
                    "Shipment Mode": "By Sea",
                    "Gps Customer Number": "848012",
                    "Technical Size": "A32",
                    "Customer Size": "A/S",
                    "Ordered Quantity": 94,
                },
            ]
        )

        wb = self.module.create_auto_workbook(df, ["WN-001"], {}, [])
        ws = wb["WN-001"]

        subtotal_row = next(
            row_idx
            for row_idx in range(1, ws.max_row + 1)
            if ws.cell(row=row_idx, column=14).value == "Sub. Total "
        )
        po_qty_col = next(
            col_idx
            for col_idx in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=col_idx).value == "PO QTY"
        )
        po_qty_letter = ws.cell(row=1, column=po_qty_col).column_letter

        self.assertEqual(
            ws.cell(row=subtotal_row, column=po_qty_col).value,
            f"=SUM({po_qty_letter}3:{po_qty_letter}{subtotal_row - 1})",
        )

    def test_korea_orders_are_combined_into_dedicated_size_section(self) -> None:
        def sized_row(
            country: str,
            customer_no: str,
            po_line: int,
            technical_size: str,
            customer_size: str,
            quantity: int,
        ) -> Dict[str, Any]:
            row = self._tms_row(
                "WN-001",
                "ART-1",
                country,
                customer_no,
                101 if country == "GERMANY" else 202,
                po_line,
                quantity,
            )
            row["Technical Size"] = technical_size
            row["Customer Size"] = customer_size
            return row

        rows = [
            sized_row("GERMANY", "1001", 10, "42", "XS", 10),
            sized_row("GERMANY", "1001", 20, "46", "S", 20),
            sized_row("GERMANY", "1001", 30, "50", "M", 30),
            sized_row("GERMANY", "1001", 40, "54", "L", 40),
            sized_row("GERMANY", "1001", 50, "58", "XL", 50),
            sized_row("KOREA", "823016", 10, "42", "XS", 7),
            sized_row("KOREA", "823016", 20, "46", "S", 62),
            sized_row("KOREA", "823016", 30, "48", "M", 114),
            sized_row("KOREA", "823016", 40, "50", "L", 136),
            sized_row("KOREA", "823016", 50, "54", "XL", 106),
            sized_row("KOREA", "823016", 60, "56", "2XL", 25),
            sized_row("KOREA", "823016", 70, "58", "3XL", 5),
        ]
        wb = self.module.create_auto_workbook(pd.DataFrame(rows), ["WN-001"], {}, [])
        ws = wb["WN-001"]
        destination_col = 14
        size_start_col = 15

        sections: List[Dict[str, Any]] = []
        row_idx = 1
        while row_idx <= ws.max_row:
            if ws.cell(row=row_idx, column=destination_col).value != "sourcing size":
                row_idx += 1
                continue

            data_start_row = row_idx + 2
            data_end_row = data_start_row - 1
            next_row = row_idx + 1
            for candidate_row in range(data_start_row, ws.max_row + 1):
                label = ws.cell(row=candidate_row, column=destination_col).value
                if label in {"Sub. Total ", "sourcing size"}:
                    data_end_row = candidate_row - 1
                    next_row = candidate_row + 1
                    break

            sizes: List[str] = []
            customer_sizes: List[str] = []
            col_idx = size_start_col
            while col_idx <= ws.max_column:
                header = ws.cell(row=row_idx, column=col_idx).value
                if header == "PO QTY":
                    break
                if header:
                    sizes.append(str(header))
                    customer_sizes.append(str(ws.cell(row=row_idx + 1, column=col_idx).value))
                col_idx += 1

            destinations = [
                str(ws.cell(row=data_row, column=destination_col).value)
                for data_row in range(data_start_row, data_end_row + 1)
                if ws.cell(row=data_row, column=destination_col).value
            ]
            sections.append(
                {
                    "sizes": sizes,
                    "customer_sizes": customer_sizes,
                    "destinations": destinations,
                }
            )
            row_idx = next_row

        korea_sections = [
            section for section in sections
            if "KOREA" in section["destinations"]
        ]
        non_korea_sections = [
            section for section in sections
            if "KOREA" not in section["destinations"]
        ]

        self.assertEqual(1, len(korea_sections), sections)
        self.assertEqual({"KOREA"}, set(korea_sections[0]["destinations"]))
        self.assertEqual(
            ["42", "46", "48", "50", "54", "56", "58"],
            korea_sections[0]["sizes"],
        )
        self.assertEqual(
            ["XS", "S", "M", "L", "XL", "2XL", "3XL"],
            korea_sections[0]["customer_sizes"],
        )
        self.assertTrue(any("GERMANY" in section["destinations"] for section in non_korea_sections))
        self.assertFalse(any("KOREA" in section["destinations"] for section in non_korea_sections))

    def test_inch_customer_sizes_with_and_without_quote_split_sections(self) -> None:
        size_pairs = [
            ("40", "28", '28"'),
            ("42", "30", '30"'),
            ("46", "32", '32"'),
            ("50", "34", '34"'),
            ("54", "36", '36"'),
            ("58", "38", '38"'),
            ("62", "40", '40"'),
        ]
        plain_rows: List[Dict[str, Any]] = []
        quoted_rows: List[Dict[str, Any]] = []
        for index, (technical_size, plain_customer_size, quoted_customer_size) in enumerate(size_pairs, 1):
            plain_row = self._tms_row(
                "WN-001",
                "ART-1",
                "BRAZIL",
                "672027",
                101,
                index * 10,
                index,
            )
            plain_row["Technical Size"] = technical_size
            plain_row["Customer Size"] = plain_customer_size
            plain_rows.append(plain_row)

            quoted_row = self._tms_row(
                "WN-001",
                "ART-1",
                "GERMANY",
                "516333",
                202,
                index * 10,
                index + 20,
            )
            quoted_row["Technical Size"] = technical_size
            quoted_row["Customer Size"] = quoted_customer_size
            quoted_rows.append(quoted_row)

        rows = quoted_rows[:-1] + plain_rows + quoted_rows[-1:]

        wb = self.module.create_auto_workbook(pd.DataFrame(rows), ["WN-001"], {}, [])
        ws = wb["WN-001"]
        destination_col = 14
        size_start_col = 15
        section_rows = [
            row_idx
            for row_idx in range(1, ws.max_row + 1)
            if ws.cell(row=row_idx, column=destination_col).value == "sourcing size"
        ]

        sections: List[Dict[str, Any]] = []
        for header_row in section_rows:
            sizes: List[str] = []
            customer_sizes: List[str] = []
            col_idx = size_start_col
            while col_idx <= ws.max_column:
                header = ws.cell(row=header_row, column=col_idx).value
                if header == "PO QTY":
                    break
                if header:
                    sizes.append(str(header))
                    customer_sizes.append(str(ws.cell(row=header_row + 1, column=col_idx).value))
                col_idx += 1

            data_start_row = header_row + 2
            data_end_row = data_start_row - 1
            for candidate_row in range(data_start_row, ws.max_row + 1):
                label = ws.cell(row=candidate_row, column=destination_col).value
                if label in {"Sub. Total ", "sourcing size"}:
                    data_end_row = candidate_row - 1
                    break

            destinations = [
                str(ws.cell(row=data_row, column=destination_col).value)
                for data_row in range(data_start_row, data_end_row + 1)
                if ws.cell(row=data_row, column=destination_col).value
            ]
            column_totals = [
                sum(
                    ws.cell(row=data_row, column=size_start_col + offset).value or 0
                    for data_row in range(data_start_row, data_end_row + 1)
                )
                for offset in range(len(sizes))
            ]
            sections.append(
                {
                    "sizes": sizes,
                    "customer_sizes": customer_sizes,
                    "destinations": destinations,
                    "column_totals": column_totals,
                }
            )

        self.assertEqual(2, len(sections), sections)
        sections_by_customer_size = {
            tuple(section["customer_sizes"]): section
            for section in sections
        }
        plain_section = sections_by_customer_size[("28", "30", "32", "34", "36", "38", "40")]
        quoted_section = sections_by_customer_size[('28"', '30"', '32"', '34"', '36"', '38"', '40"')]

        expected_sizes = ["40", "42", "46", "50", "54", "58", "62"]
        self.assertEqual(expected_sizes, plain_section["sizes"])
        self.assertEqual(expected_sizes, quoted_section["sizes"])
        self.assertEqual({"BRAZIL"}, set(plain_section["destinations"]))
        self.assertEqual({"GERMANY"}, set(quoted_section["destinations"]))
        self.assertEqual([1, 2, 3, 4, 5, 6, 7], plain_section["column_totals"])
        self.assertEqual([21, 22, 23, 24, 25, 26, 27], quoted_section["column_totals"])

    def test_south_africa_is_merged_into_europe_category(self) -> None:
        rows = [
            self._tms_row("WN-001", "ART-1", "SOUTH AFRICA", "1001", 101, 10, 400),
        ]

        wb = self._process_rows(rows)
        ws = wb["WN-001"]
        category_helper_col = next(
            col_idx
            for col_idx in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=col_idx).value == "_Jane_CATEGORY"
        )
        category_labels = [
            ws.cell(row=row_idx, column=6).value
            for row_idx in range(1, ws.max_row + 1)
        ]
        data_categories = [
            ws.cell(row=row_idx, column=category_helper_col).value
            for row_idx in range(2, ws.max_row + 1)
            if ws.cell(row=row_idx, column=category_helper_col).value
        ]

        self.assertIn("欧美单", category_labels)
        self.assertNotIn("南非单", category_labels)
        self.assertEqual(["欧美单"], data_categories)

    def test_legacy_south_africa_category_is_merged_into_europe_category(self) -> None:
        rows = [
            self._tms_row("WN-001", "ART-1", "", "1001", 101, 10, 400),
        ]
        country_rows = [
            {
                "CST NO": "1001",
                "DESTINATION": "",
                "REGION": "",
                "CATEGORY": "南非单",
            }
        ]

        wb = self._process_rows(rows, country_rows=country_rows)
        ws = wb["WN-001"]
        category_helper_col = next(
            col_idx
            for col_idx in range(1, ws.max_column + 1)
            if ws.cell(row=1, column=col_idx).value == "_Jane_CATEGORY"
        )
        data_categories = [
            ws.cell(row=row_idx, column=category_helper_col).value
            for row_idx in range(2, ws.max_row + 1)
            if ws.cell(row=row_idx, column=category_helper_col).value
        ]

        self.assertEqual(["欧美单"], data_categories)

    def test_process_reports_sorts_by_click_sort_priority(self) -> None:
        rows = [
            self._tms_row("WN-2", "ART-2", "CANADA", "1002", 206, 10, 60),
            self._tms_row("WN-1", "ART-2", "BRAZIL", "1002", 104, 10, 40),
            self._tms_row("WN-1", "ART-2", "GERMANY", "1002", 103, 10, 30),
            self._tms_row("WN-1", "ART-1", "GERMANY", "1001", 102, 10, 20),
            self._tms_row("WN-1", "ART-1", "GERMANY", "1001", 101, 10, 10),
            self._tms_row("WN-10", "ART-0", "USA", "9999", 305, 10, 50),
        ]

        sorted_rows = self.module.sort_tms_rows_for_generation(pd.DataFrame(rows))
        self.assertEqual(
            [104, 101, 102, 103, 206, 305],
            list(sorted_rows["PO Number"]),
        )

        wb = self._process_rows(rows)

        self.assertEqual(
            ["Summary", "WN-1", "WN-2", "WN-10"],
            wb.sheetnames,
        )
        self.assertEqual(
            [
                ("WN-1", "ART-2"),
                ("WN-1", "ART-1"),
                ("WN-2", "ART-2"),
                ("WN-10", "ART-0"),
            ],
            self._summary_working_article_rows(wb),
        )

    def test_detail_sort_uses_country_file_destination_when_tms_country_is_blank(self) -> None:
        rows = [
            self._tms_row("WN-001", "ART-1", "", "1002", 102, 10, 20),
            self._tms_row("WN-001", "ART-1", "", "1001", 101, 10, 10),
        ]
        country_lookup = {
            "001002": {"destination": "ITALY"},
            "001001": {"destination": "GERMANY"},
        }

        sorted_rows = self.module.sort_tms_rows_for_generation(pd.DataFrame(rows), country_lookup)

        self.assertEqual([101, 102], list(sorted_rows["PO Number"]))

    def test_summary_working_order_uses_natural_ascending_working_number(self) -> None:
        rows = [
            self._tms_row("RC2702OW006", "LK0670", "GERMANY", "1001", 1006, 10, 240),
            self._tms_row("RC2702OW005", "LK0670", "GERMANY", "1001", 1005, 10, 170),
            self._tms_row("RC2702OW001", "LK0670", "GERMANY", "1001", 1001, 10, 600),
            self._tms_row("RC2702OW000", "LK0670", "GERMANY", "1001", 1000, 10, 754),
        ]

        wb = self._process_rows(rows)

        self.assertEqual(
            ["RC2702OW000", "RC2702OW001", "RC2702OW005", "RC2702OW006"],
            self._summary_working_order(wb),
        )
        self.assertEqual(
            ["Summary", "RC2702OW000", "RC2702OW001", "RC2702OW005", "RC2702OW006"],
            wb.sheetnames,
        )

    def test_summary_working_order_places_suffix_after_base_working_number(self) -> None:
        rows = [
            self._tms_row("RC2701OW001", "LI8292", "GERMANY", "1001", 1001, 10, 100),
            self._tms_row("RC2701OW000A", "LI8290", "GERMANY", "1001", 1002, 10, 100),
            self._tms_row("RC2701OW000", "LI8286", "GERMANY", "1001", 1003, 10, 100),
        ]

        wb = self._process_rows(rows)

        self.assertEqual(
            ["RC2701OW000", "RC2701OW000A", "RC2701OW001"],
            self._summary_working_order(wb),
        )

    def test_working_filter_sorts_filtered_rows_by_click_sort_priority(self) -> None:
        rows = [
            self._tms_row("WN-1", "ART-1", "BRAZIL", "1002", 103, 10, 30),
            self._tms_row("WN-10", "ART-1", "GERMANY", "1001", 102, 10, 20),
            self._tms_row("WN-2", "ART-1", "GERMANY", "1001", 101, 10, 10),
        ]

        wb = self._process_rows(rows, working_filters=["WN-1", "WN-10", "WN-2"])

        self.assertEqual(["Summary", "WN-1", "WN-2", "WN-10"], wb.sheetnames)


if __name__ == "__main__":
    unittest.main()

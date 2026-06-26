import os
import sys
import tempfile
import unittest

from openpyxl import load_workbook


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.eric_module import EricModule, YTIC_DESTINATION_HEADERS, YTIC_SP_HEADERS


class EricModuleYticSourceTests(unittest.TestCase):
    def setUp(self):
        self.module = EricModule()

    def test_extract_sp_rows_uses_actual_confirmation_sections(self):
        rows = [
            ["*CUSTOMER  PO NUMBER"],
            ["0902792931", "Adidas Originals", "May - Oct", "2026", "CO-BA15", 106],
            [],
            ["", "PO Line Items"],
            [],
            [
                "",
                "",
                "*STYLE NUMBER",
                "PO LINE NUMBER",
                "CUSTOMER SEASON",
                "CUSTOMER SEASON YEAR",
                "PARENT STYLE",
                "TMS OFFICE",
                "PRODCUTION DIVISION",
                "DISTRIBUTOR DIVISION",
                "*CUSTOMER  DELIVERY  DATE",
                "*SHIP MODE",
                "*DESTINATION",
                "PO QTY",
                "SIZE RANGE",
                "UOM",
                "SALES CONFIRMATION NUMBER",
                "SC STATUS",
                "PURCHASE CONFIRMATION NUMBER",
                "PC STATUS",
            ],
            [
                "",
                "",
                "RC2606OW000",
                1,
                "FW26",
                "2026",
                "",
                "TM4",
                "",
                "ISC",
                46218,
                "By Sea",
                "United Arab Emirates",
                106,
                "2T,3T,4T,5T,2XS,XS,S,M,L,XL",
                "PIECES",
            ],
            [],
            ["", "", "", "PO Sub Details"],
            [],
            ["", "", "", "", "COLOR", "SIZE", "QTY", "PC NO", "ADJ QTY"],
            ["", "", "", "", "LI2854", "XS", 32, 1],
            [],
            ["", "", "", "Sales Confirmation"],
            [],
            [
                "",
                "",
                "",
                "",
                "CUSTOMER",
                "*SELLER/SUPPLIER",
                "*BILL TO",
                "*CONSIGNEE",
                "MAX ALLOW SHIP (%)",
                "MAX ALLOW SHORT SHIP (%)",
                "REQUIRED DELIVERY DATE",
                "SHIP MODE",
                "QUOTA CTRY",
                "SALES CONFIRMATION NUMBER",
                "PO QTY",
                "FOB",
            ],
            [
                "",
                "",
                "",
                "",
                "Adidas Originals",
                "TMS Fashion (H.K.) Ltd.",
                "Adidas International Trading AG",
                "Adidas Emerging Markets FZE",
                0,
                0,
                46218,
                "By Sea",
                "United Arab Emirates",
                "",
                106,
                16.5,
            ],
            [],
            ["", "", "", "Purchase Confirmation"],
            [],
            [
                "",
                "",
                "",
                "",
                "PC NUMBER",
                "FACTORY",
                "*BILL TO",
                "*CONSIGNEE",
                "SHIP TO",
                "MAX ALLOW SHIP (%)",
                "REQUIRED DELIVERY DATE",
                "SHIP MODE",
                "CTRY ORIGIN",
            ],
            [
                "",
                "",
                "",
                "",
                1,
                "Yuen Thai Industrial Company Ltd.",
                "TMS Fashion (H.K.) Ltd.",
                "Adidas Emerging Markets FZE",
                "Retail Logistics LLC, Dubai, UAE",
                0,
                46208,
                "By Sea",
                "United Arab Emirates",
            ],
        ]

        rows_out = self.module.extract_sp_rows_from_source(rows)

        self.assertEqual(
            rows_out,
            [
                [
                    "0902792931",
                    "RC2606OW000",
                    "FW26",
                    "2026",
                    "",
                    "TM4",
                    "",
                    "ISC",
                    46218,
                    "",
                    "By Sea",
                    "United Arab Emirates",
                    106,
                    "2T,3T,4T,5T,2XS,XS,S,M,L,XL",
                    "PIECES",
                ],
                [
                    "",
                    "",
                    "Adidas Originals",
                    "TMS Fashion (H.K.) Ltd.",
                    "Adidas International Trading AG",
                    "Adidas Emerging Markets FZE",
                    0,
                    0,
                    46218,
                    10,
                    "By Sea",
                    "United Arab Emirates",
                    "",
                    "",
                    "",
                ],
                [
                    "",
                    "",
                    1,
                    "Yuen Thai Industrial Company Ltd.",
                    "TMS Fashion (H.K.) Ltd.",
                    "Adidas Emerging Markets FZE",
                    "Retail Logistics LLC, Dubai, UAE",
                    0,
                    46208,
                    "",
                    "By Sea",
                    "United Arab Emirates",
                    "",
                    "",
                    "",
                ],
            ],
        )

    def test_build_size_check_rows_prefers_ytic_order(self):
        final_map = {
            ("0902773420", "KV1165", "2XS"): 25,
            ("0902792931", "LI2854", "XS"): 32,
            ("0902775685", "KV1165", "XS"): 51,
        }
        ytic_map = {
            ("0902792931", "LI2854", "XS"): 32,
            ("0902775685", "KV1165", "XS"): 51,
            ("0902773420", "KV1165", "2XS"): 25,
        }
        ytic_order = [
            ("0902792931", "LI2854", "XS"),
            ("0902775685", "KV1165", "XS"),
            ("0902773420", "KV1165", "2XS"),
        ]

        rows = self.module.build_size_check_rows(final_map, ytic_map, ytic_order)

        self.assertEqual([row[:3] for row in rows], [list(key) for key in ytic_order])

    def test_normalize_ship_mode_matches_ytic_destination_rules(self):
        self.assertEqual(self.module.normalize_ship_mode("By Sea"), "Ocean")
        self.assertEqual(self.module.normalize_ship_mode("By Courier"), "Air Express")
        self.assertEqual(self.module.normalize_ship_mode("By Courier", "Germany"), "Ocean")

    def test_reconciliation_workbook_contains_required_sheets_and_text_compare(self):
        final_data = {
            "headers": ["PO Number", "Working Number"],
            "records": [
                {"PO Number": "0902792931", "Working Number": "RC2606OW000"},
                {"PO Number": "", "Working Number": "RC2606OW000"},
                {"PO Number": "0902775685", "Working Number": "RC2606OW000"},
                {"PO Number": "0902773420", "Working Number": "RC2606OW000"},
            ],
        }
        ytic_data = {
            "size_rows": [],
            "destination_headers": list(YTIC_DESTINATION_HEADERS),
            "destination_rows": [
                [
                    "0902792931",
                    "RC2606OW000",
                    "FW26",
                    "2026",
                    "TM4",
                    "ISC",
                    46218,
                    46218,
                    0,
                    "By Sea",
                    "Ocean",
                    "",
                    "United Arab Emirates",
                    "UNITED ARAB EMIRATES",
                    1,
                    106,
                    "2T,3T,4T,5T,2XS,XS,S,M,L,XL",
                    "PIECES",
                ],
                [
                    "0902775685",
                    "RC2606OW000",
                    "FW26",
                    "2026",
                    "TM4",
                    "ISC",
                    46218,
                    46228,
                    10,
                    "By Sea",
                    "Ocean",
                    "",
                    "United Arab Emirates",
                    "UNITED ARAB EMIRATES",
                    "0",
                    106,
                    "2T,3T,4T,5T,2XS,XS,S,M,L,XL",
                    "PIECES",
                ],
                [
                    "0900000001",
                    "RC2606OW000",
                    "FW26",
                    "2026",
                    "TM4",
                    "ISC",
                    46218,
                    46218,
                    0,
                    "By Sea",
                    "Ocean",
                    "",
                    "United Arab Emirates",
                    "UNITED ARAB EMIRATES",
                    1,
                    106,
                    "2T,3T,4T,5T,2XS,XS,S,M,L,XL",
                    "PIECES",
                ],
            ],
            "sp_headers": list(YTIC_SP_HEADERS),
            "sp_rows": [
                [
                    "0902792931",
                    "RC2606OW000",
                    "FW26",
                    "2026",
                    "",
                    "TM4",
                    "",
                    "ISC",
                    46218,
                    "",
                    "By Sea",
                    "Germany",
                    106,
                    "2T,3T,4T,5T,2XS,XS,S,M,L,XL",
                    "PIECES",
                ],
                [
                    "",
                    "",
                    "Adidas Originals",
                    "TMS Fashion (H.K.) Ltd.",
                    "Adidas International Trading AG",
                    "Adidas Emerging Markets FZE",
                    0,
                    0,
                    46218,
                    10,
                    "By Sea",
                    "United kingdom",
                    "",
                    "",
                    "",
                ],
                [
                    "",
                    "",
                    1,
                    "Yuen Thai Industrial Company Ltd.",
                    "TMS Fashion (H.K.) Ltd.",
                    "Adidas Emerging Markets FZE",
                    "Retail Logistics LLC, Dubai, UAE",
                    0,
                    46208,
                    "",
                    "By Sea",
                    "United Arab Emirates",
                    "",
                    "",
                    "",
                ],
            ],
        }
        size_check_rows = [
            ["0902792931", "LI2854", "XS", 32, 32, 0, "OK"],
        ]

        with tempfile.TemporaryDirectory() as temp_dir:
            output_file = os.path.join(temp_dir, "eric_reconcile.xlsx")
            self.module.write_reconciliation_workbook(
                output_file,
                final_data,
                ytic_data,
                size_check_rows,
                [],
                [],
            )

            wb = load_workbook(output_file, read_only=True, data_only=False)
            try:
                self.assertEqual(
                    wb.sheetnames,
                    [
                        "Size_Check",
                        "Final_Data",
                        "Destination1Extract",
                        "SP_Extract",
                        "PO_Text_Compare",
                    ],
                )
                size_rows = list(wb["Size_Check"].iter_rows(values_only=True))
                self.assertEqual(
                    size_rows,
                    [
                        (
                            "PO Number",
                            "Article Number",
                            "Size",
                            "Final Data Quantity",
                            "PO Quantity",
                            "MARGIN",
                        ),
                        ("0902792931", "LI2854", "XS", 32, 32, "=D2-E2"),
                    ],
                )
                self.assertNotEqual(wb["Size_Check"]["F1"].fill.fgColor.rgb, "FFFFFF00")
                self.assertNotEqual(wb["Size_Check"]["F2"].fill.fgColor.rgb, "FFFFFF00")
                self.assertEqual(wb["Destination1Extract"]["I2"].value, "=H2-G2")
                self.assertEqual(wb["Destination1Extract"]["I3"].value, "=H3-G3")
                self.assertIs(wb["Destination1Extract"]["O2"].value, True)
                self.assertIs(wb["Destination1Extract"]["O3"].value, False)
                self.assertEqual(wb["SP_Extract"]["J3"].value, "=I3-I4")
                self.assertIsNone(wb["SP_Extract"]["J4"].value)
                sp_sheet = wb["SP_Extract"]
                highlighted_fill = "FFEBF1DE"
                for cell_address in ("A2", "L2", "O2", "A3", "L3", "O3"):
                    self.assertEqual(sp_sheet[cell_address].fill.fgColor.rgb, highlighted_fill)
                for cell_address in ("A4", "L4", "O4"):
                    self.assertNotEqual(sp_sheet[cell_address].fill.fgColor.rgb, highlighted_fill)
                self.assertNotEqual(
                    wb["Destination1Extract"]["A2"].fill.fgColor.rgb,
                    highlighted_fill,
                )
                rows = list(wb["PO_Text_Compare"].iter_rows(values_only=True))
                self.assertEqual(
                    rows,
                    [
                        (
                            "PO Text",
                            "In Final_Data PO Number",
                            "In YTIC Destination CUSTOMER PO NUMBER",
                            "Status",
                        ),
                        ("0902792931", True, True, "OK"),
                        ("0902775685", True, True, "OK"),
                        ("0900000001", False, True, "YTIC_ONLY"),
                        ("0902773420", True, False, "FINAL_ONLY"),
                    ],
                )
            finally:
                wb.close()


if __name__ == "__main__":
    unittest.main()

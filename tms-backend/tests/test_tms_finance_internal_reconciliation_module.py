import os
import sys
import tempfile
import unittest

import openpyxl
from fastapi import FastAPI
from fastapi.testclient import TestClient


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.tms_finance_internal_reconciliation_module import (  # noqa: E402
    TmsFinanceInternalReconciliationModule,
)


TARGET_HEADERS = [
    "Date1",
    "Sales invoice",
    "Date2",
    "Purchase invoice",
    "REMARK",
    "VENDOR",
    "CUSTOMER",
    "QTY",
    "Purchase amount",
    "Sales amount  with Tax 13%",
    "货描",
    "WORKING  NO.(款号)",
    "PO ORDER NO.",
    "ARTICLE NO.(货号)",
    "CUSTOMER NO.(客户编码)",
    "CUSTOMER ORDER NO.(客户订单号)",
    "交期",
    "MR",
    "Commercial Invoice",
    "工厂Promo附加费",
    "采购",
    "销售",
]


class TmsFinanceInternalReconciliationModuleTests(unittest.TestCase):
    def setUp(self) -> None:
        self.module = TmsFinanceInternalReconciliationModule()

    def _save_target_workbook(
        self,
        folder: str,
        *,
        rows_before_blank: int = 208,
        exclude_rows: str | None = None,
        exclude_columns: str | None = None,
    ) -> str:
        path = os.path.join(folder, "内销对账单.xlsx")
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "未清账"
        ws.append(TARGET_HEADERS)
        for row in range(2, rows_before_blank + 1):
            ws.append(
                [
                    f"date1-{row}",
                    f"sales-{row}",
                    f"date2-{row}",
                    f"purchase-{row}",
                    "KEEP",
                    "KEEP-VENDOR",
                    "adidas",
                    1,
                    1.11,
                    2.22,
                    "历史货描",
                    f"HISTORY{row}",
                    f"PO{row}",
                    f"ART{row}",
                    "825059",
                    "",
                    "2026-04-30",
                    "Jane",
                    f"FS{row}",
                    0.12,
                    f"#FS{row}(purchase-{row})",
                    f"#FS{row}(sales-{row})",
                ],
            )
        # 第一个全空行是 rows_before_blank + 1；下方 Total/历史行必须保留且不参与处理。
        ws.cell(rows_before_blank + 2, 1).value = "Total"
        ws.cell(rows_before_blank + 2, 8).value = "历史合计"

        archive = workbook.create_sheet("清 from 2407")
        archive.append(TARGET_HEADERS)
        archive.append(
            [
                "2024-07-05",
                "old-sales",
                "2024-07-05",
                "old-purchase",
                "Bulk",
                "新龙泰",
                "adidas",
                1,
                1,
                1,
                "归档",
                "ARCHIVE",
                "0900000001",
                "AA0001",
                "825057",
                "0300000001",
                "2024-07-05",
                "Rosa",
                "",
                "",
                "",
                "",
            ],
        )

        if exclude_rows is not None or exclude_columns is not None:
            rules = workbook.create_sheet("_Exclude_Rows_Cols")
            rules["A1"] = "排除行"
            rules["B1"] = "排除列"
            if exclude_rows is not None:
                rules["A2"] = exclude_rows
            if exclude_columns is not None:
                rules["B2"] = exclude_columns

        workbook.save(path)
        return path

    def _append_source_headers(self, ws, header_row: int) -> None:
        ws.cell(header_row, 1).value = "Pack"
        ws.cell(header_row, 2).value = "Season"
        ws.cell(header_row, 3).value = "款号"
        ws.cell(header_row, 4).value = "颜色"
        ws.cell(header_row, 5).value = "描述"
        ws.cell(header_row, 6).value = "订单号"
        ws.cell(header_row, 7).value = "客户订单号"
        ws.cell(header_row, 8).value = "客户编号"
        ws.cell(header_row, 9).value = "客户仓库"
        ws.cell(header_row, 10).value = "大货交期"
        ws.cell(header_row, 11).value = "数量"
        ws.cell(header_row, 12).value = "工厂价格"
        ws.cell(header_row + 1, 12).value = "单价"
        ws.cell(header_row + 1, 13).value = "PO税金"
        ws.cell(header_row + 1, 14).value = "PO总金额=单价*PO件数+PO税金"
        ws.cell(header_row, 15).value = "TMS价格"
        ws.cell(header_row + 1, 15).value = "单价"
        ws.cell(header_row + 1, 16).value = "TMS税金"
        ws.cell(header_row + 1, 17).value = "VAS/SHAS附加费"
        ws.cell(header_row + 1, 18).value = "PO总金额=单价*PO件数+PO税金+VAS/SHAS"
        ws.cell(header_row, 19).value = "工厂"
        ws.cell(header_row + 1, 19).value = "工厂code"
        ws.cell(header_row + 1, 20).value = "工厂"
        ws.cell(header_row, 22).value = "TMS 业务"
        ws.cell(header_row, 24).value = "TMS发票#"

    def _append_source_row(
        self,
        ws,
        row: int,
        *,
        remark: str,
        style: str,
        order: str,
        article: str,
        qty: int | float,
        purchase_amount: float,
        sales_amount: float,
        vendor_code: str = "1L8006",
        vendor_name: str = "丹东新龙太",
        customer_order: str = "(blank)",
        commercial_invoice: str | None = None,
    ) -> None:
        ws.cell(row, 1).value = "FW26"
        ws.cell(row, 2).value = "FW26"
        ws.cell(row, 3).value = style
        ws.cell(row, 4).value = article
        ws.cell(row, 5).value = "女式梭织上衣"
        ws.cell(row, 6).value = order
        ws.cell(row, 7).value = customer_order
        ws.cell(row, 8).value = "825059" if remark == "Sample" else "825057"
        ws.cell(row, 9).value = "MSO\u200b" if remark == "Sample" else "天津仓"
        ws.cell(row, 10).value = "2026-05-13" if remark == "Sample" else "2026-05-07"
        ws.cell(row, 11).value = qty
        ws.cell(row, 14).value = purchase_amount
        ws.cell(row, 17).value = 0.96 if remark == "Sample" else 969.58
        ws.cell(row, 18).value = sales_amount
        ws.cell(row, 19).value = vendor_code
        ws.cell(row, 20).value = vendor_name
        ws.cell(row, 22).value = "Bacy" if remark == "Sample" else "Caroline"
        ws.cell(row, 24).value = commercial_invoice

    def _save_sample_workbook(self, folder: str) -> str:
        path = os.path.join(folder, "合并Sample 202605.xlsx")
        workbook = openpyxl.Workbook()
        hidden = workbook.active
        hidden.title = "仅供中国单使用-4月份出货"
        hidden.sheet_state = "hidden"
        self._append_source_headers(hidden, 2)
        self._append_source_row(
            hidden,
            4,
            remark="Sample",
            style="SHOULD_IGNORE",
            order="4500000000",
            article="OLD001",
            qty=99,
            purchase_amount=99,
            sales_amount=99,
        )

        ws = workbook.create_sheet("5月")
        self._append_source_headers(ws, 2)
        for offset in range(14):
            purchase = 3.39 if offset < 13 else 376.29
            sales = 7.86 if offset < 13 else 377.37
            self._append_source_row(
                ws,
                4 + offset,
                remark="Sample",
                style=f"RC2613OW{offset:03d}",
                order=f"4501749{offset:03d}",
                article=f"LE{8200 + offset}",
                qty=3,
                purchase_amount=purchase,
                sales_amount=sales,
            )
        ws.cell(19, 3).value = "AFTER_BLANK_SAMPLE"
        ws.cell(19, 6).value = "4509999999"
        ws.cell(19, 11).value = 42
        ws.cell(19, 14).value = 420.36
        ws.cell(19, 18).value = 479.55

        dc = workbook.create_sheet("DC")
        dc.append(["客户编号", "客户仓库"])
        dc.append(["825059", "MSO"])
        workbook.save(path)
        return path

    def _save_bulk_workbook(self, folder: str) -> str:
        path = os.path.join(folder, "合并BULK 202605.xlsx")
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "仅供中国单使用-5月"
        self._append_source_headers(ws, 1)
        quantities = [10, 10, 20, 30, 50, 60, 100, 120]
        purchases = [11.3, 11.3, 22.6, 33.9, 113, 135.6, 226, 406.8]
        sales = [1118.23, 2848.08, 1047.98, 2747.43, 1739.13, 991.37, 1160.86, 2467.04]
        vendor_codes = ["ELP012", "1L8012", "1L8006", "ELP008", "ABC001", "1L8006", "1L8006", "1L8006"]
        vendor_names = ["", "", "丹东新龙太", "", "测试工厂", "新龙泰", "新龙泰", "新龙泰"]
        for offset, qty in enumerate(quantities):
            self._append_source_row(
                ws,
                3 + offset,
                remark="Bulk",
                style=f"RC2610OW{offset:03d}",
                order=f"0901888{offset:03d}",
                article=f"KX{1867 + offset}",
                qty=qty,
                purchase_amount=purchases[offset],
                sales_amount=sales[offset],
                vendor_code=vendor_codes[offset],
                vendor_name=vendor_names[offset],
                customer_order=f"03058{offset:05d}",
                commercial_invoice="CI-BULK-000" if offset == 0 else None,
            )
        ws.cell(12, 3).value = "AFTER_BLANK_BULK"
        ws.cell(12, 6).value = "0901999999"
        ws.cell(12, 11).value = 400
        ws.cell(12, 14).value = 960.5
        ws.cell(12, 18).value = 14120.12
        workbook.save(path)
        return path

    def test_updates_tail_target_rows_from_multiple_sources_without_appending(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            target_path = self._save_target_workbook(folder)
            bulk_path = self._save_bulk_workbook(folder)
            sample_path = self._save_sample_workbook(folder)

            result = self.module.process_files([bulk_path, sample_path], target_path, folder)

            self.assertTrue(result["success"])
            self.assertEqual(result["updated_count"], 22)
            self.assertEqual(result["appended_count"], 0)
            self.assertEqual(result["source_row_count"], 22)
            self.assertEqual(result["target_row_count"], 207)
            self.assertEqual(result["source_summary"]["bulk_rows"], 8)
            self.assertEqual(result["source_summary"]["sample_rows"], 14)
            self.assertEqual(result["totals"]["quantity"], 442)
            self.assertEqual(result["totals"]["purchase_amount"], 1380.86)
            self.assertEqual(result["totals"]["sales_amount_with_tax"], 14599.67)
            self.assertEqual(result["excluded_rows"], [])
            self.assertEqual(result["excluded_columns"], [])

            output_wb = openpyxl.load_workbook(result["output_path"], data_only=False)
            self.assertIn("清 from 2407", output_wb.sheetnames)
            ws = output_wb["未清账"]
            self.assertEqual(ws.max_row, 210)
            self.assertEqual(ws.cell(186, 12).value, "HISTORY186")
            self.assertEqual(ws.cell(187, 5).value, "Bulk")
            self.assertEqual(ws.cell(187, 6).value, "万代")
            self.assertEqual(ws.cell(187, 8).value, 10)
            self.assertEqual(ws.cell(187, 9).value, 11.3)
            self.assertEqual(ws.cell(187, 9).number_format, "0.00")
            self.assertEqual(ws.cell(187, 10).value, 1118.23)
            self.assertEqual(ws.cell(187, 10).number_format, "0.00")
            self.assertEqual(ws.cell(187, 19).value, "CI-BULK-000")
            self.assertEqual(ws.cell(187, 21).value, "#CI-BULK-000(purchase-187)")
            self.assertEqual(ws.cell(187, 22).value, "#CI-BULK-000(sales-187)")
            self.assertEqual(ws.cell(187, 1).value, "date1-187")
            self.assertEqual(ws.cell(187, 2).value, "sales-187")
            self.assertEqual(ws.cell(187, 3).value, "date2-187")
            self.assertEqual(ws.cell(187, 4).value, "purchase-187")
            self.assertEqual(ws.cell(188, 6).value, "万代")
            self.assertEqual(ws.cell(189, 6).value, "新龙泰")
            self.assertEqual(ws.cell(191, 6).value, "测试工厂")
            self.assertEqual(ws.cell(194, 5).value, "Bulk")
            self.assertEqual(ws.cell(194, 8).value, 120)
            self.assertEqual(ws.cell(195, 5).value, "Sample")
            self.assertEqual(ws.cell(195, 6).value, "新龙泰")
            self.assertEqual(ws.cell(208, 8).value, 3)
            self.assertEqual(ws.cell(208, 9).value, 376.29)
            self.assertTrue(all(ws.cell(209, column).value in (None, "") for column in range(1, 23)))
            self.assertEqual(ws.cell(210, 1).value, "Total")
            self.assertEqual(ws.cell(210, 8).value, "历史合计")
            self.assertNotIn(
                "SHOULD_IGNORE",
                [ws.cell(row, 12).value for row in range(187, 209)],
            )
            self.assertNotIn(
                "AFTER_BLANK_BULK",
                [ws.cell(row, 12).value for row in range(187, 209)],
            )
            self.assertEqual(output_wb["清 from 2407"].cell(2, 12).value, "ARCHIVE")

    def test_respects_exclude_helper_sheet_and_keeps_excluded_cells(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            target_path = self._save_target_workbook(
                folder,
                exclude_rows="203",
                exclude_columns="Purchase amount, V",
            )
            bulk_path = self._save_bulk_workbook(folder)

            result = self.module.process_files([bulk_path], target_path, folder)

            self.assertEqual(result["updated_count"], 8)
            self.assertEqual(result["excluded_rows"], [203])
            self.assertEqual(result["excluded_columns"], [9, 22])

            output_wb = openpyxl.load_workbook(result["output_path"], data_only=False)
            ws = output_wb["未清账"]
            self.assertEqual(ws.cell(200, 5).value, "Bulk")
            self.assertEqual(ws.cell(200, 9).value, 1.11)
            self.assertEqual(ws.cell(200, 10).value, 1118.23)
            self.assertEqual(ws.cell(200, 22).value, "#FS200(sales-200)")
            self.assertEqual(ws.cell(203, 12).value, "HISTORY203")
            self.assertEqual(ws.cell(203, 5).value, "KEEP")

    def test_rejects_invalid_exclude_helper_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            target_path = self._save_target_workbook(
                folder,
                exclude_rows="abc",
                exclude_columns="不存在的列",
            )
            bulk_path = self._save_bulk_workbook(folder)

            with self.assertRaisesRegex(ValueError, "排除配置"):
                self.module.process_files([bulk_path], target_path, folder)

    def test_api_accepts_repeated_source_files_legacy_fields_and_downloads_result(self) -> None:
        with tempfile.TemporaryDirectory() as folder:
            sample_path = self._save_sample_workbook(folder)
            bulk_path = self._save_bulk_workbook(folder)
            target_path = self._save_target_workbook(folder)

            from api import tms_finance_internal_reconciliation_api as api

            original_upload_dir = api.UPLOAD_DIR
            api.UPLOAD_DIR = folder
            app = FastAPI()
            app.include_router(api.router, prefix="/api")
            client = TestClient(app)
            try:
                with (
                    open(bulk_path, "rb") as bulk_file,
                    open(sample_path, "rb") as sample_file,
                    open(target_path, "rb") as target_file,
                ):
                    response = client.post(
                        "/api/tms-finance/internal-reconciliation/process",
                        files=[
                            (
                                "source_files",
                                (
                                    "合并BULK 202605.xlsx",
                                    bulk_file,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                ),
                            ),
                            (
                                "source_files",
                                (
                                    "合并Sample 202605.xlsx",
                                    sample_file,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                ),
                            ),
                            (
                                "target_file",
                                (
                                    "内销对账单 202605.xlsx",
                                    target_file,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                ),
                            ),
                        ],
                    )
                self.assertEqual(response.status_code, 200)
                payload = response.json()
                self.assertTrue(payload["success"])
                self.assertEqual(payload["updated_count"], 22)
                self.assertEqual(payload["source_row_count"], 22)
                self.assertTrue(payload["output_file"].endswith(".xlsx"))

                download = client.get(
                    f"/api/tms-finance/internal-reconciliation/download/{payload['output_file']}",
                )
                self.assertEqual(download.status_code, 200)
                self.assertGreater(len(download.content), 1000)

                target_path = self._save_target_workbook(folder)
                with (
                    open(sample_path, "rb") as sample_file,
                    open(bulk_path, "rb") as bulk_file,
                    open(target_path, "rb") as target_file,
                ):
                    legacy_response = client.post(
                        "/api/tms-finance/internal-reconciliation/process",
                        files={
                            "sample_file": (
                                "合并Sample 202605.xlsx",
                                sample_file,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            ),
                            "bulk_file": (
                                "合并BULK 202605.xlsx",
                                bulk_file,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            ),
                            "target_file": (
                                "内销对账单 202605.xlsx",
                                target_file,
                                "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                            ),
                        },
                    )
                self.assertEqual(legacy_response.status_code, 200)
                self.assertEqual(legacy_response.json()["updated_count"], 22)

                bad_target_path = self._save_target_workbook(
                    folder,
                    exclude_rows="bad-row",
                    exclude_columns=None,
                )
                with (
                    open(bulk_path, "rb") as bulk_file,
                    open(bad_target_path, "rb") as target_file,
                ):
                    bad_response = client.post(
                        "/api/tms-finance/internal-reconciliation/process",
                        files=[
                            (
                                "source_files",
                                (
                                    "合并BULK 202605.xlsx",
                                    bulk_file,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                ),
                            ),
                            (
                                "target_file",
                                (
                                    "内销对账单.xlsx",
                                    target_file,
                                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                                ),
                            ),
                        ],
                    )
                self.assertEqual(bad_response.status_code, 400)
            finally:
                api.UPLOAD_DIR = original_upload_dir


if __name__ == "__main__":
    unittest.main()

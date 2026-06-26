import os
import sys
import tempfile
import unittest

import openpyxl


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.draft_packing_compare_module import (  # noqa: E402
    DraftPackingCompareModule,
    ExtractedRecord,
    PackingExtractedPage,
)


DIRECT_AND_ATTACHMENT_DRAFT_TEXT = """
1 THIRTEEN (13) CARTONS OF ADIDAS BRAND GARMENT
WOMEN'S WOVEN JACKET,FROM BODY: 100% COTTON
PO# 0902172742
ARTICLE NO.: KV0625
STYLE NO.: AF26INSPW072
CUST NO.: 0306485444
QUANTITY: 150 PCS
HS Code:620432
PSR 150 PIECES
2 SIXTY ONE (61) CARTONS OF ADIDAS BRAND GARMENT
WOMEN'S WOVEN PANTS,MAIN MATERIAL: 100% COTTON
PSR 537 PIECES
(SEE ATTACHMENT)
11.Declaration by the exporter
Attachment
PO# 0902319411
ARTICLE NO.: LE8303
STYLE NO.: RC2613OW006
CUST NO.: 0306700330
QUANTITY: 537 PCS
HS Code:620462
"""


PACKING_PAGE_TEXT = """
HONG KONG Shipment Number Invoice Number
Phone: +85285221484699 ..
553714618450 10-04-26-0537
PO No PO Line
Aggregator
Working No Article
No
Article Description Model
No
Model Name Market PO
Number
Category Brand QTY Ctn
Count
Net Net Net Gross Vol Vol Wt
0902172742 1 AF26INSPW072 KV0625 ST DNM WSH JK
CRSK
P6083 ST DNM WSH
JK
0306485444 NOT SPORTS
SPECIFIC
ADIDAS 150 13 108.770 111.770 126.070 0.874 0.000
Sourcing Size/Manufacturing Size HTS Description Quota Category Quantity
A28,A32,A36,A40,A44 62043290 APP/ JACKET 150
Goods Description
WOMEN'S WOVEN ST DNM WSH JACKET,FROM BODY: 100% COTTON
PO No PO Line
Aggregator
Working No Article No Article Description Model No Model Name Market PO
Number
Category Brand QTY Ctn Count Net Net Net Gross Vol Vol Wt
0902319334 1 RC2613OW003 LE8298 SKIRT BLACK CH594 SKIRT 0306699266 ORIGINALS ADIDAS 824 33 314.540 331.020 365.670 1.848 0.000
Goods Description
WOMEN'S WOVEN SKIRT,MAIN MATERIAL: 100% COTTON
"""


PACKING_PAGE = PackingExtractedPage(
    text=PACKING_PAGE_TEXT,
    tables=[
        [
            [
                "PO No",
                "PO Line\nAggregator",
                "Working No",
                "Article\nNo",
                "Article Description",
                "Model\nNo",
                "Model Name",
                "Market PO\nNumber",
                "Category",
                "Brand",
                "QTY",
                "C t n\nCount",
            ],
            [
                "0902172742",
                "1",
                "AF26INSPW072",
                "KV0625",
                "ST DNM WSH JK\nCRSK",
                "P6083",
                "ST DNM WSH\nJK",
                "0306485444",
                "NOT SPORTS\nSPECIFIC",
                "ADIDAS",
                "150",
                "13",
            ],
        ],
        [
            [
                "Sourcing Size/Manufacturing Size",
                "HTS",
                "Description",
                "Quota Category",
                "Quantity",
            ],
            ["A28,A32,A36,A40,A44", "62043290", "APP/ JACKET", "", "150"],
        ],
        [
            [
                "PO No",
                "PO Line\nAggregator",
                "Working No",
                "Article No",
                "Article Description",
                "Model No",
                "Model Name",
                "Market PO\nNumber",
                "Category",
                "Brand",
                "QTY",
                "C t n\nCount",
            ],
            [
                "0902319334",
                "1",
                "RC2613OW003",
                "LE8298",
                "SKIRT BLACK",
                "CH594",
                "SKIRT",
                "0306699266",
                "ORIGINALS",
                "ADIDAS",
                "824",
                "33",
            ],
        ],
    ],
)


PACKING_PAGE_WITH_SUMMARY_FOOTER = PackingExtractedPage(
    text="""
PO No PO Line
0902319350 1 RC2613OW006 LE5888 PANTS 0306700331 ADIDAS 703 78
Goods Description
WOMEN'S WOVEN DENIM PANTS,MAIN MATERIAL: 100% COTTON
This document is a summary and does not contain all the terms and conditions applicable to the transaction.
The complete document may be accessed on the system. Page 6 of 8
""",
    tables=[
        [
            [
                "PO No",
                "PO Line",
                "Working No",
                "Article No",
                "Market PO Number",
                "QTY",
                "C t n Count",
            ],
            [
                "0902319350",
                "1",
                "RC2613OW006",
                "LE5888",
                "0306700331",
                "703",
                "78",
            ],
        ],
    ],
)


CHINA_PERU_FTA_TEXT = """
CERTIFICATE OF ORIGIN
Form for China-Peru FTA
6.Item 7.Number and kind of 8.HS code 9.Origin 10.Gross 11.Number and 12.Invoiced
number Packages;description of (Six digit criterion weight,quantity date of invoice value
(Max goods code) (Quantity Unit)
20) or other
measures (liters,
m³,etc.)
1 EIGHTEEN (18) CARTONS OF KID'S 620462 PSR 67.04% 252 PIECES 10-04-26-0537 USD
WOVEN DENIM PANT,MAIN G.WEIGHT MAY.05,2026 22838.65
MATERIAL:100% COTTON 161.24 KGS
PO# 0902275914
ARTICLE NO.: KU8337
STYLE NO.: OJF26250705
2 SIXTEEN (16) CARTONS OF KID'S 620462 PSR 67.04% 251 PIECES
WOVEN DENIM PANT,MAIN G.WEIGHT
MATERIAL:100% COTTON 107.56 KGS
PO# 0902275872
ARTICLE NO.: KU8317
STYLE NO.: OLKF26250706
3 THIRTY NINE (39) CARTONS OF 620432 PSR 95.21% 504 PIECES
KID'S WOVEN DENIM JACKET,MAIN G.WEIGHT
MATERIAL:100% COTTON 415.23 KGS
PO# 0902275843
ARTICLE NO.: KS1734
STYLE NO.: OJF26250703
***
"""


ACFTA_FORM_E_ATTACHMENT_TEXT = """
ASEAN-CHINA FREE TRADE AREA
FORM E
1 FIFTY SIX (56) CARTONS OF WOMEN'S WOVEN DENIM PANTS, PSR 560 PIECES 6541156264
MAIN MATERIAL: 100% COTTON MAY.06,2026
PO# 0902319353
ARTICLE NO.: LE5888
STYLE NO.: RC2613OW006
CUST ORDER NO.:0306700299
ARTICLE DESCRIPTION:DENIM BARREL LTBLUE
HS Code:620462
2 FOURTEEN (14) CARTONS OF WOMEN'S WOVEN SKIRT,MAIN PSR 371 PIECES
MATERIAL: 100% COTTON
PO# 0902319344
ARTICLE NO.: LE8298
STYLE NO.: RC2613OW003
CUST ORDER NO.:0306700300
ARTICLE DESCRIPTION:SKIRT BLACK
HS Code:620452
3 CART NO. ELEVEN (11) CARTONS OF WOMEN'S WOVEN SKIRT,MAIN PSR 285 PIECES
CUST O/N MATERIAL: 100% COTTON
PO NO PO# 0902319335
ART NO ARTICLE NO.: LE8297
SIZE ARTICLE NO. STYLE NO.: RC2613OW003
QTY STYLE NO. CUST ORDER NO.:0306700305
MADE IN CHINA ARTICLE DESCRIPTION:SKIRT OLISTR
HS Code:620452
(SEE ATTACHMENT)
ADDRESS:1008 WANDA BUILDING NO.9 JIEFANG STREET ZHONGSHAN
FAX:0411-82643422 TEL:0411-82532807
"""


CAMBODIA_FORM_D_TEXT = """
KHTH2605006456
XO TEX INDUSTRIAL CO., LTD.
KINGDOM OF CAMBODIA.
ADIDAS (THAILAND) CO.,LTD.
07-APR-2026
730 PIECE
1 CUST O/N : HS CODE: 610910 CC+SP 730 PIECE 17-03-26-07
PO NO. : WOMEN'S 93% COTTON 7% ELASTANE KNIT T-SHIRT 94
ART NO. : PO NO # ART NO # CUST. O/NO # ART DES # 30-MAR-2026
SIZE : 0901891240 KY7558 0305678582 3S TEE VIVRED
QTY : 14 CARTON
MADE IN
CAMBODIA
TOTAL (PIECE): SEVEN HUNDRED THIRTY ONLY
THIRD-COUNTRY INVOICING:
INVOICE NUMBER: 6541141861
DATE:31-MAR-2026
"""


CAMBODIA_FORM_D_MULTI_ITEM_TEXT = """
KHVN2604030487
XO TEX INDUSTRIAL CO., LTD.
KINGDOM OF CAMBODIA.
ADIDAS VIETNAM CO LTD
05-APR-2026
895 PIECE
1 CUST O/N : HS CODE: 610990 CC 515 PIECE 17-03-26-07
PO NO. : ADIDAS BRANDED : MEN'S 100% POLYESTER 96
ART NO. : (100%RECYCLED) KNITJERSEY (SHORT SLEEVE) 30-MAR-2026
SIZE : PO NO # ART NO # CUST. O/NO #
QTY : 0901944686 KX1258 0305670288
QTY:515 PCS
13 CARTON
2 MADE IN HS CODE: 610910 CC 380 PIECE
CAMBODIA ADIDAS BRANDED : MEN'S 100% COTTON KNIT T-SHIRT
N.W.: PO NO # ART NO # CUST. O/NO #
G.W.: 0901891214 KY2114 0305678561
QTY:380 PCS
13 CARTON
TOTAL (PIECE): EIGHT HUNDRED NINETY FIVE ONLY
"""


CAMBODIA_FORM_D_PACKING_PAGE = PackingExtractedPage(
    text="""
TMS FASHION (H.K.) LIMITED Invoice Number
NANYANG PLAZA 17-03-26-0794
PO No PO Line Working No Article No Article Description Model No Model Name Market PO Category Brand QTY Ctn Count Net Net Net Gross Vol Vol Wt
Aggregator Number
0901891240 1 RC2608OW004 KY7558 3S TEE VIVRED P4708 3S TEE 0305678582 ORIGINALS ADIDAS 730 14 156.950 157.090 171.830 0.885 0.000
Goods Description
WOMEN'S 93% COTTON 7% ELASTANE KNITTED T-SHIRT(HTS:6109.10.0000)
Total Quantity 730
""",
    tables=[
        [
            [
                "PO No",
                "PO Line",
                "Working No",
                "Article No",
                "Article Description",
                "Model No",
                "Model Name",
                "Market PO Number",
                "Category",
                "Brand",
                "QTY",
                "Ctn Count",
            ],
            [
                "0901891240",
                "1",
                "RC2608OW004",
                "KY7558",
                "3S TEE VIVRED",
                "P4708",
                "3S TEE",
                "0305678582",
                "ORIGINALS",
                "ADIDAS",
                "730",
                "14",
            ],
        ],
    ],
)


CAMBODIA_FORM_D_MULTI_PACKING_PAGE = PackingExtractedPage(
    text="""
TMS FASHION (H.K.) LIMITED Invoice Number
NANYANG PLAZA 17-03-26-0796
PO No PO Line Working No A r t i c l e Article Description M o d e l Model Name Market PO Category Brand QTY C t n Net Net Net Gross Vol Vol Wt
Aggregator No No Number Count
0901891214 1 RC2611OM012 KY2114 C C C A L I T E E IE071 C C C A L I 0305678561 ORIGINALS ADIDAS 380 13 91.870 92.000 102.000 0.493 0.000
WHITE TEE
Sourcing Size/Manufacturing Size HTS Description Quota Category Quantity
A42,A46,A50,A54,A38,A58,A62 61091010 NOT APPLICABLE 380
Goods Description
MEN'S 100% COTTON KNITTED T-SHIRT (HTS:6109.10.0000)
PO No PO Line Working No Article No Article Description Model No Model Name Market PO Category Brand QTY Ctn Count Net Net Net Gross Vol Vol Wt
Aggregator Number
0901944686 1 RC2610OM009 KX1258 JERSEY BLUE CA843 JERSEY 0305670288 ORIGINALS ADIDAS 515 13 120.640 120.770 133.290 0.739 0.000
Sourcing Size/Manufacturing Size HTS Description Quota Category Quantity
A38,A42,A46,A50,A54,A58 62113390 APP/ TRACKSUIT/ OTHERS 515
Goods Description
MEN'S 100% POLYESTER (100% RECYCLED) KNIT SHIRT(HTS:6105.20.0000)
Total Quantity 895
""",
    tables=[
        [
            [
                "PO No",
                "PO Line",
                "Working No",
                "Article No",
                "Article Description",
                "Model No",
                "Model Name",
                "Market PO Number",
                "Category",
                "Brand",
                "QTY",
                "C t n Count",
            ],
            [
                "0901891214",
                "1",
                "RC2611OM012",
                "KY2114",
                "C C C A L I T E E\nWHITE TEE",
                "IE071",
                "C C C A L I",
                "0305678561",
                "ORIGINALS",
                "ADIDAS",
                "380",
                "13",
            ],
        ],
        [
            [
                "Sourcing Size/Manufacturing Size",
                "HTS",
                "Description",
                "Quota Category",
                "Quantity",
            ],
            ["A42,A46,A50,A54,A38,A58,A62", "61091010", "NOT APPLICABLE", "", "380"],
        ],
        [
            [
                "PO No",
                "PO Line",
                "Working No",
                "Article No",
                "Article Description",
                "Model No",
                "Model Name",
                "Market PO Number",
                "Category",
                "Brand",
                "QTY",
                "Ctn Count",
            ],
            [
                "0901944686",
                "1",
                "RC2610OM009",
                "KX1258",
                "JERSEY BLUE",
                "CA843",
                "JERSEY",
                "0305670288",
                "ORIGINALS",
                "ADIDAS",
                "515",
                "13",
            ],
        ],
        [
            [
                "Sourcing Size/Manufacturing Size",
                "HTS",
                "Description",
                "Quota Category",
                "Quantity",
            ],
            ["A38,A42,A46,A50,A54,A58", "62113390", "APP/ TRACKSUIT/ OTHERS", "", "515"],
        ],
    ],
)


class FailingOriginCertificateOcr:
    def __init__(self) -> None:
        self.called = False

    def extract_records(self, pdf_path):
        self.called = True
        raise AssertionError("text-readable FTA PDFs should not trigger OCR fallback")


class DraftPackingCompareModuleTests(unittest.TestCase):
    def setUp(self) -> None:
        self.module = DraftPackingCompareModule()

    def _fill_rgb(self, cell) -> str:
        color = cell.fill.fgColor
        if color.type == "rgb":
            return (color.rgb or "").upper()
        return ""

    def _has_fill(self, cell) -> bool:
        return cell.fill.fill_type is not None

    def test_parse_draft_text_keeps_style_and_article_separate_and_uses_attachment_context(self):
        records = self.module.parse_draft_text(DIRECT_AND_ATTACHMENT_DRAFT_TEXT)

        self.assertEqual(len(records), 2)
        first = records[0]
        self.assertEqual(first.po_number, "0902172742")
        self.assertEqual(first.working_number, "AF26INSPW072")
        self.assertEqual(first.article_number, "KV0625")
        self.assertEqual(first.customer_number, "0306485444")
        self.assertEqual(first.quantity, 150)
        self.assertEqual(first.cartons, 13)
        self.assertEqual(first.cartons_in_words, "THIRTEEN")
        self.assertEqual(first.hs_code, "620432")
        self.assertEqual(first.goods_description, "WOMEN'S WOVEN JACKET,FROM BODY: 100% COTTON")

        attachment = records[1]
        self.assertEqual(attachment.po_number, "0902319411")
        self.assertEqual(attachment.working_number, "RC2613OW006")
        self.assertEqual(attachment.article_number, "LE8303")
        self.assertEqual(attachment.cartons, 61)
        self.assertEqual(attachment.cartons_in_words, "SIXTY ONE")
        self.assertEqual(attachment.goods_description, "WOMEN'S WOVEN PANTS,MAIN MATERIAL: 100% COTTON")

    def test_parse_origin_certificate_pdf_uses_text_fta_records_without_ocr(self):
        ocr = FailingOriginCertificateOcr()
        module = DraftPackingCompareModule(origin_certificate_ocr=ocr)
        module._extract_pdf_text = lambda _pdf_path: CHINA_PERU_FTA_TEXT

        records = module.parse_origin_certificate_pdf("text-readable-fta.pdf")

        self.assertFalse(ocr.called)
        self.assertEqual(len(records), 3)
        first = records[0]
        self.assertEqual(first.po_number, "0902275914")
        self.assertEqual(first.working_number, "OJF26250705")
        self.assertEqual(first.article_number, "KU8337")
        self.assertEqual(first.quantity, 252)
        self.assertEqual(first.cartons, 18)
        self.assertEqual(first.cartons_in_words, "EIGHTEEN")
        self.assertEqual(first.goods_description, "KID'S WOVEN DENIM PANT,MAIN MATERIAL:100% COTTON")
        self.assertEqual(first.hs_code, "620462")
        self.assertTrue(any("Cust Number" in issue for issue in first.issues))

    def test_parse_origin_certificate_pdf_uses_text_acfta_records_without_ocr(self):
        ocr = FailingOriginCertificateOcr()
        module = DraftPackingCompareModule(origin_certificate_ocr=ocr)
        module._extract_pdf_text = lambda _pdf_path: ACFTA_FORM_E_ATTACHMENT_TEXT

        records = module.parse_origin_certificate_pdf("text-readable-acfta-form-e.pdf")

        self.assertFalse(ocr.called)
        self.assertEqual(len(records), 3)
        attachment = records[2]
        self.assertEqual(attachment.po_number, "0902319335")
        self.assertEqual(attachment.working_number, "RC2613OW003")
        self.assertEqual(attachment.article_number, "LE8297")
        self.assertEqual(attachment.customer_number, "0306700305")
        self.assertEqual(attachment.quantity, 285)
        self.assertEqual(attachment.cartons, 11)
        self.assertEqual(attachment.hs_code, "620452")
        self.assertEqual(attachment.goods_description, "WOMEN'S WOVEN SKIRT,MAIN MATERIAL: 100% COTTON")
        self.assertFalse(any("PO Number" in issue for issue in attachment.issues))

    def test_form_d_records_match_packing_when_certificate_has_no_style(self):
        ocr = FailingOriginCertificateOcr()
        module = DraftPackingCompareModule(origin_certificate_ocr=ocr)
        module._extract_pdf_text = lambda _pdf_path: CAMBODIA_FORM_D_TEXT

        draft_records = module.parse_origin_certificate_pdf("text-readable-form-d.pdf")
        packing_records = module.parse_packing_pages([CAMBODIA_FORM_D_PACKING_PAGE])
        comparison = module.build_comparison_result(draft_records, packing_records)

        self.assertFalse(ocr.called)
        self.assertEqual(len(draft_records), 1)
        self.assertEqual(len(packing_records), 1)
        self.assertEqual(packing_records[0].hs_code, "6109.10.0000")
        self.assertEqual(comparison.group_count, 1)
        self.assertIsNotNone(comparison.groups[0].draft)
        self.assertIsNotNone(comparison.groups[0].packing)
        self.assertFalse(any(issue.field_name == "Record" for issue in comparison.groups[0].issues))

        with tempfile.TemporaryDirectory() as tmpdir:
            result = module.process_extracted_data(draft_records, packing_records, tmpdir)
            workbook = openpyxl.load_workbook(result["output_path"], data_only=True)
            ws = workbook[module.OUTPUT_SHEET]

            self.assertEqual(ws["A2"].value, "0901891240")
            self.assertIsNone(ws["D2"].value)
            self.assertEqual(ws["E2"].value, "KY7558")
            self.assertEqual(ws["F2"].value, "0305678582")
            self.assertEqual(ws["G2"].value, 730)
            self.assertEqual(ws["H2"].value, 14)
            self.assertEqual(ws["J2"].value, "WOMEN'S 93% COTTON 7% ELASTANE KNIT T-SHIRT")
            self.assertEqual(ws["K2"].value, "610910")
            self.assertEqual(ws["D3"].value, "RC2608OW004")
            self.assertEqual(ws["K3"].value, "6109.10.0000")
            self.assertNotIn("未找到对应记录", ws["M2"].value or "")

    def test_multi_item_form_d_records_match_packing_when_certificate_has_no_style(self):
        ocr = FailingOriginCertificateOcr()
        module = DraftPackingCompareModule(origin_certificate_ocr=ocr)
        module._extract_pdf_text = lambda _pdf_path: CAMBODIA_FORM_D_MULTI_ITEM_TEXT

        draft_records = module.parse_origin_certificate_pdf("text-readable-form-d-multi.pdf")
        packing_records = module.parse_packing_pages([CAMBODIA_FORM_D_MULTI_PACKING_PAGE])
        comparison = module.build_comparison_result(draft_records, packing_records)

        self.assertFalse(ocr.called)
        self.assertEqual(len(draft_records), 2)
        self.assertEqual(len(packing_records), 2)
        self.assertEqual(comparison.group_count, 2)
        self.assertFalse(any(issue.field_name == "Record" for issue in comparison.issues))
        self.assertEqual(
            {record.po_number: record.hs_code for record in draft_records},
            {"0901944686": "610990", "0901891214": "610910"},
        )
        self.assertEqual(
            {record.po_number: record.cartons for record in draft_records},
            {"0901944686": 13, "0901891214": 13},
        )

    def test_process_acfta_form_e_writes_populated_packing_rows(self):
        module = DraftPackingCompareModule()
        module._extract_pdf_text = lambda _pdf_path: ACFTA_FORM_E_ATTACHMENT_TEXT
        draft_records = module.parse_origin_certificate_pdf("text-readable-acfta-form-e.pdf")
        packing_records = [
            ExtractedRecord(
                source="Packing List",
                po_number="0902319353",
                invoice_number="10-04-26-0537",
                working_number="RC2613OW006",
                article_number="LE5888",
                customer_number="0306700299",
                quantity=560,
                cartons=56,
                goods_description="WOMEN'S WOVEN DENIM PANTS,MAIN MATERIAL: 100% COTTON",
            ),
            ExtractedRecord(
                source="Packing List",
                po_number="0902319344",
                invoice_number="10-04-26-0537",
                working_number="RC2613OW003",
                article_number="LE8298",
                customer_number="0306700300",
                quantity=371,
                cartons=14,
                goods_description="WOMEN'S WOVEN SKIRT,MAIN MATERIAL: 100% COTTON",
            ),
            ExtractedRecord(
                source="Packing List",
                po_number="0902319335",
                invoice_number="10-04-26-0537",
                working_number="RC2613OW003",
                article_number="LE8297",
                customer_number="0306700305",
                quantity=285,
                cartons=11,
                goods_description="WOMEN'S WOVEN SKIRT,MAIN MATERIAL: 100% COTTON",
            ),
        ]

        with tempfile.TemporaryDirectory() as tmpdir:
            result = module.process_extracted_data(draft_records, packing_records, tmpdir)
            workbook = openpyxl.load_workbook(result["output_path"], data_only=True)
            ws = workbook[module.OUTPUT_SHEET]

            packing_rows = [
                [cell.value for cell in row]
                for row in ws.iter_rows(min_row=2)
                if row[1].value == "Packing List"
            ]

        self.assertEqual(len(packing_rows), 3)
        self.assertFalse(
            [
                row
                for row in packing_rows
                if all(value in (None, "") for value in row[3:11])
            ]
        )
        self.assertTrue(
            any(
                row[:10]
                == [
                    "0902319335",
                    "Packing List",
                    "10-04-26-0537",
                    "RC2613OW003",
                    "LE8297",
                    "0306700305",
                    285,
                    11,
                    None,
                    "WOMEN'S WOVEN SKIRT,MAIN MATERIAL: 100% COTTON",
                ]
                for row in packing_rows
            )
        )

    def test_parse_packing_pages_uses_tables_and_reports_missing_hts(self):
        records = self.module.parse_packing_pages([PACKING_PAGE])

        self.assertEqual(len(records), 2)
        first = records[0]
        self.assertEqual(first.po_number, "0902172742")
        self.assertEqual(first.working_number, "AF26INSPW072")
        self.assertEqual(first.article_number, "KV0625")
        self.assertEqual(first.customer_number, "0306485444")
        self.assertEqual(first.invoice_number, "10-04-26-0537")
        self.assertEqual(first.quantity, 150)
        self.assertEqual(first.cartons, 13)
        self.assertEqual(first.hs_code, "62043290")
        self.assertEqual(
            first.goods_description,
            "WOMEN'S WOVEN ST DNM WSH JACKET,FROM BODY: 100% COTTON",
        )

        missing_hts = records[1]
        self.assertEqual(missing_hts.po_number, "0902319334")
        self.assertEqual(missing_hts.invoice_number, "10-04-26-0537")
        self.assertEqual(missing_hts.hs_code, "")
        self.assertTrue(any("HTS" in issue for issue in missing_hts.issues))

    def test_parse_packing_pages_removes_summary_footer_from_goods_description(self):
        records = self.module.parse_packing_pages([PACKING_PAGE_WITH_SUMMARY_FOOTER])

        self.assertEqual(len(records), 1)
        self.assertEqual(
            records[0].goods_description,
            "WOMEN'S WOVEN DENIM PANTS,MAIN MATERIAL: 100% COTTON",
        )
        self.assertNotIn("This document is a summary", records[0].goods_description)
        self.assertNotIn("Page 6 of 8", records[0].goods_description)

    def test_parse_packing_pages_carries_leading_description_to_previous_page_record(self):
        page_with_trailing_record = PackingExtractedPage(
            text="""
PO Details
PO No PO Line Working No Article No Article Description Model No Model Name Market PO Number Category Brand QTY Ctn Count
Aggregator Number
0902319344 1 RC2613OW003 LE8298 SKIRT BLACK CH594 SKIRT 0306700300 ORIGINALS ADIDAS 371 14
This document is a summary and does not contain all the terms and conditions applicable to the transaction.
Page 6 of 8
""",
            tables=[
                [
                    [
                        "PO No",
                        "PO Line",
                        "Working No",
                        "Article No",
                        "Article Description",
                        "Model No",
                        "Model Name",
                        "Market PO Number",
                        "Category",
                        "Brand",
                        "QTY",
                        "Ctn Count",
                    ],
                    [
                        "0902319344",
                        "1",
                        "RC2613OW003",
                        "LE8298",
                        "SKIRT BLACK",
                        "CH594",
                        "SKIRT",
                        "0306700300",
                        "ORIGINALS",
                        "ADIDAS",
                        "371",
                        "14",
                    ],
                ],
            ],
        )
        page_with_leading_description = PackingExtractedPage(
            text="""
Goods Description
WOMEN'S WOVEN SKIRT,MAIN MATERIAL: 100% COTTON
PO No PO Line Working No Article No Article Description Model No Model Name Market PO Number Category Brand QTY Ctn Count
Aggregator Number
0902319345 1 RC2613OW004B LF3779 SSTR TT CM626 SSTR TT 0306700311 ORIGINALS ADIDAS 405 23
Goods Description
WOMEN'S WOVEN DENIM TRACK TOP,MAIN MATERIAL: 99% COTTON 1% ELASTANE
""",
            tables=[
                [
                    [
                        "PO No",
                        "PO Line",
                        "Working No",
                        "Article No",
                        "Article Description",
                        "Model No",
                        "Model Name",
                        "Market PO Number",
                        "Category",
                        "Brand",
                        "QTY",
                        "Ctn Count",
                    ],
                    [
                        "0902319345",
                        "1",
                        "RC2613OW004B",
                        "LF3779",
                        "SSTR TT",
                        "CM626",
                        "SSTR TT",
                        "0306700311",
                        "ORIGINALS",
                        "ADIDAS",
                        "405",
                        "23",
                    ],
                ],
            ],
        )

        records = self.module.parse_packing_pages(
            [page_with_trailing_record, page_with_leading_description]
        )
        records_by_po = {record.po_number: record for record in records}

        self.assertEqual(
            records_by_po["0902319344"].goods_description,
            "WOMEN'S WOVEN SKIRT,MAIN MATERIAL: 100% COTTON",
        )
        self.assertEqual(
            records_by_po["0902319345"].goods_description,
            "WOMEN'S WOVEN DENIM TRACK TOP,MAIN MATERIAL: 99% COTTON 1% ELASTANE",
        )

    def test_build_comparison_result_marks_strict_hs_difference(self):
        draft_records = self.module.parse_draft_text(DIRECT_AND_ATTACHMENT_DRAFT_TEXT)
        packing_records = self.module.parse_packing_pages([PACKING_PAGE])

        result = self.module.build_comparison_result(draft_records[:1], packing_records[:1])

        self.assertEqual(result.group_count, 1)
        self.assertEqual(result.issue_count, 2)
        self.assertEqual(result.mismatch_count, 2)
        self.assertEqual(result.missing_field_count, 0)
        self.assertIn("HS Code / HTS Code", result.groups[0].issue_detail)
        self.assertIn("Goods Description", result.groups[0].issue_detail)

    def test_missing_field_status_uses_generic_feedback_label(self):
        records = self.module.parse_packing_pages([PACKING_PAGE])

        result = self.module.build_comparison_result([], records[1:2])

        self.assertEqual(result.groups[0].status, "需反馈")

    def test_process_extracted_data_writes_two_row_excel_with_issue_sheet_and_fills(self):
        draft_records = self.module.parse_draft_text(DIRECT_AND_ATTACHMENT_DRAFT_TEXT)
        packing_records = self.module.parse_packing_pages([PACKING_PAGE])

        with tempfile.TemporaryDirectory() as folder:
            result = self.module.process_extracted_data(draft_records[:1], packing_records[:1], folder)

            self.assertTrue(result["success"])
            self.assertEqual(result["group_count"], 1)
            self.assertEqual(result["issue_count"], 2)

            workbook = openpyxl.load_workbook(result["output_path"])
            ws = workbook["产地证 vs Packing"]
            issues = workbook["Issues"]

            self.assertEqual(
                [cell.value for cell in ws[1]],
                [
                    "PO Number",
                    "Source",
                    "INV number",
                    "Working Number / Style Number",
                    "Article Number",
                    "Cust Number / Market PO Number",
                    "Quantity",
                    "Cartons",
                    "Cartons In Words",
                    "Goods Description",
                    "HS Code / HTS Code",
                    "Check Status",
                    "Issue Detail",
                ],
            )
            self.assertEqual(ws["A2"].value, "0902172742")
            self.assertEqual(ws["B2"].value, "产地证")
            self.assertEqual(ws["B3"].value, "Packing List")
            self.assertEqual(ws["C2"].value, "10-04-26-0537")
            self.assertEqual(ws["C3"].value, "10-04-26-0537")
            self.assertEqual(ws["D2"].value, "AF26INSPW072")
            self.assertEqual(ws["E2"].value, "KV0625")
            self.assertEqual(ws["I2"].value, "THIRTEEN")
            self.assertIsNone(ws["I3"].value)
            self.assertEqual(ws["K2"].value, "620432")
            self.assertEqual(ws["K3"].value, "62043290")
            self.assertEqual(ws["L2"].value, "需核对")
            self.assertEqual(ws.max_row, 3)
            self.assertEqual(self._fill_rgb(ws["K1"]), "FFD9EAF7")
            self.assertFalse(self._has_fill(ws["K2"]))
            self.assertFalse(self._has_fill(ws["K3"]))
            self.assertGreaterEqual(issues.max_row, 2)
            self.assertIn("HS Code / HTS Code", [issues.cell(row=row, column=4).value for row in range(2, issues.max_row + 1)])

    def test_process_extracted_data_uses_filled_blank_separator_rows(self):
        draft_records = self.module.parse_draft_text(DIRECT_AND_ATTACHMENT_DRAFT_TEXT)
        packing_records = self.module.parse_packing_pages([PACKING_PAGE])

        with tempfile.TemporaryDirectory() as folder:
            result = self.module.process_extracted_data(draft_records, packing_records, folder)

            workbook = openpyxl.load_workbook(result["output_path"])
            ws = workbook["产地证 vs Packing"]

            self.assertEqual(ws.max_row, 1 + result["group_count"] * 2 + result["group_count"] - 1)
            separator_rows = [4 + group_index * 3 for group_index in range(result["group_count"] - 1)]

            for row_index in separator_rows:
                values = [ws.cell(row=row_index, column=column_index).value for column_index in range(1, len(self.module.HEADERS) + 1)]
                self.assertTrue(all(value in (None, "") for value in values))
                for column_index in range(1, len(self.module.HEADERS) + 1):
                    cell = ws.cell(row=row_index, column=column_index)
                    self.assertEqual(self._fill_rgb(cell), "FFEFF6FF")

            self.assertEqual(self._fill_rgb(ws["I4"]), "FFEFF6FF")
            self.assertNotEqual(self._fill_rgb(ws["I4"]), "FFFFC7CE")
            self.assertEqual(self._fill_rgb(ws["K4"]), "FFEFF6FF")

    def test_process_file_batches_writes_one_sheet_per_invoice(self):
        module = DraftPackingCompareModule()
        draft_by_path = {
            "draft-0794.pdf": [
                ExtractedRecord(
                    source=module.ORIGIN_CERTIFICATE_SOURCE,
                    po_number="0901891240",
                    article_number="KY7558",
                    customer_number="0305678582",
                    quantity=730,
                    cartons=14,
                    goods_description="WOMEN'S 93% COTTON 7% ELASTANE KNIT T-SHIRT",
                    hs_code="610910",
                )
            ],
            "draft-0796.pdf": [
                ExtractedRecord(
                    source=module.ORIGIN_CERTIFICATE_SOURCE,
                    po_number="0901944686",
                    article_number="KX1258",
                    customer_number="0305670288",
                    quantity=515,
                    cartons=13,
                    goods_description="MEN'S 100% POLYESTER KNIT SHIRT",
                    hs_code="610990",
                )
            ],
        }
        packing_by_path = {
            "packing-0794.pdf": [
                ExtractedRecord(
                    source="Packing List",
                    po_number="0901891240",
                    invoice_number="17-03-26-0794",
                    working_number="RC2608OW004",
                    article_number="KY7558",
                    customer_number="0305678582",
                    quantity=730,
                    cartons=14,
                    goods_description="WOMEN'S 93% COTTON 7% ELASTANE KNIT T-SHIRT",
                    hs_code="6109.10.0000",
                )
            ],
            "packing-0796.pdf": [
                ExtractedRecord(
                    source="Packing List",
                    po_number="0901944686",
                    invoice_number="17-03-26-0796",
                    working_number="RC2610OM009",
                    article_number="KX1258",
                    customer_number="0305670288",
                    quantity=515,
                    cartons=13,
                    goods_description="MEN'S 100% POLYESTER KNIT SHIRT",
                    hs_code="6109.90.0000",
                )
            ],
        }
        module.parse_origin_certificate_pdf = lambda path: draft_by_path[os.fspath(path)]
        module.parse_packing_pdf = lambda path: packing_by_path[os.fspath(path)]

        with tempfile.TemporaryDirectory() as folder:
            result = module.process_file_batches(
                ["draft-0794.pdf", "draft-0796.pdf"],
                ["packing-0794.pdf", "packing-0796.pdf"],
                folder,
            )
            workbook = openpyxl.load_workbook(result["output_path"], data_only=True)

        self.assertEqual(result["sheet_count"], 2)
        self.assertEqual(result["draft_file_count"], 2)
        self.assertEqual(result["packing_file_count"], 2)
        self.assertIn("17-03-26-0794", workbook.sheetnames)
        self.assertIn("17-03-26-0796", workbook.sheetnames)
        self.assertIn(module.ISSUES_SHEET, workbook.sheetnames)
        self.assertNotIn(module.OUTPUT_SHEET, workbook.sheetnames)
        ws = workbook["17-03-26-0794"]
        self.assertEqual(ws["C2"].value, "17-03-26-0794")
        self.assertEqual(ws["C3"].value, "17-03-26-0794")
        self.assertIsNone(ws["D2"].value)
        self.assertEqual(ws["D3"].value, "RC2608OW004")
        issues = workbook[module.ISSUES_SHEET]
        self.assertEqual(issues["A1"].value, "INV number")
        self.assertEqual(issues["B1"].value, "Sheet Name")
        self.assertEqual(issues["A2"].value, "17-03-26-0794")
        self.assertEqual(issues["B2"].value, "17-03-26-0794")

    def test_process_file_batches_puts_unmatched_records_on_unmatched_sheet(self):
        module = DraftPackingCompareModule()
        module.parse_origin_certificate_pdf = lambda _path: [
            ExtractedRecord(
                source=module.ORIGIN_CERTIFICATE_SOURCE,
                po_number="0900000000",
                article_number="UNKNOWN",
                quantity=1,
                cartons=1,
            )
        ]
        module.parse_packing_pdf = lambda _path: [
            ExtractedRecord(
                source="Packing List",
                po_number="0901891240",
                invoice_number="17-03-26-0794",
                working_number="RC2608OW004",
                article_number="KY7558",
                customer_number="0305678582",
                quantity=730,
                cartons=14,
            )
        ]

        with tempfile.TemporaryDirectory() as folder:
            result = module.process_file_batches(["draft.pdf"], ["packing.pdf"], folder)
            workbook = openpyxl.load_workbook(result["output_path"], data_only=True)

        self.assertIn(module.UNMATCHED_SHEET, workbook.sheetnames)
        unmatched_ws = workbook[module.UNMATCHED_SHEET]
        self.assertEqual(unmatched_ws["A2"].value, "0900000000")
        self.assertIn("无法唯一匹配 Packing 发票号", unmatched_ws["M2"].value)
        issues = workbook[module.ISSUES_SHEET]
        issue_details = [
            issues.cell(row=row, column=8).value
            for row in range(2, issues.max_row + 1)
        ]
        self.assertTrue(
            any("无法唯一匹配 Packing 发票号" in str(detail) for detail in issue_details)
        )

    def test_process_file_batches_sanitizes_invoice_sheet_names(self):
        module = DraftPackingCompareModule()
        module.parse_origin_certificate_pdf = lambda _path: []
        module.parse_packing_pdf = lambda _path: [
            ExtractedRecord(
                source="Packing List",
                po_number="0901891240",
                invoice_number="INV/2026:0734?*[]\\LONG-LONG-LONG",
                working_number="RC2608OW004",
                article_number="KY7558",
                quantity=730,
                cartons=14,
            )
        ]

        with tempfile.TemporaryDirectory() as folder:
            result = module.process_file_batches(["draft.pdf"], ["packing.pdf"], folder)
            workbook = openpyxl.load_workbook(result["output_path"], data_only=True)

        self.assertEqual(result["sheet_count"], 1)
        sheet_name = workbook.sheetnames[0]
        self.assertLessEqual(len(sheet_name), 31)
        self.assertNotRegex(sheet_name, r"[\[\]:*?/\\]")


if __name__ == "__main__":
    unittest.main()

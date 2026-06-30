import os
import sys
import tempfile
import unittest
from datetime import date
from typing import Any, Dict, List, Tuple

from openpyxl import Workbook, load_workbook
import pandas as pd


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.jessca_module import (
    InvoiceRecord,
    InvoiceSummaryRecord,
    JesscaModule,
    TcInvoiceExtractedPage,
    TcInvoiceRecord,
    TcInvoiceSummary,
    TcSummaryComparisonRow,
)


class JesscaModuleReferenceTableTests(unittest.TestCase):
    def setUp(self):
        self.module = JesscaModule()

    def _build_reference_df(self) -> pd.DataFrame:
        return pd.DataFrame(
            [
                {"Price": 12.34, "Style NO.": "STYLE-OK", "Article NO.": "ART-OK"},
                {"Price": 20.0, "Style NO.": "STYLE-PRICE", "Article NO.": "ART-PRICE"},
                {"Price": 10.0, "Style NO.": "STYLE-GOOD", "Article NO.": "ART-STYLE"},
                {"Price": 11.0, "Style NO.": "STYLE-ARTICLE", "Article NO.": "ART-ARTICLE"},
                {"Price": 7.0, "Style NO.": "STYLE-MULTI-A", "Article NO.": "ART-MULTI"},
                {"Price": 7.0, "Style NO.": "STYLE-MULTI-B", "Article NO.": "ART-MULTI"},
                {"Price": 8.0, "Style NO.": "STYLE-NOT-INVOICE", "Article NO.": "ART-NOT-INVOICE"},
            ]
        )

    def _sheet_records(self, worksheet: Any) -> List[Dict[str, Any]]:
        header_row_index = None
        headers = []
        for row_index, row in enumerate(worksheet.iter_rows(values_only=True), 1):
            row_values = list(row)
            if "核对状态" in row_values or "状态" in row_values or "Check Status" in row_values:
                header_row_index = row_index
                headers = row_values
                break

        self.assertIsNotNone(header_row_index)
        records = []
        for row in worksheet.iter_rows(min_row=header_row_index + 1, values_only=False):
            values = [cell.value for cell in row]
            if not any(value not in (None, "") for value in values):
                continue
            records.append(
                {
                    "values": dict(zip(headers, values)),
                    "cells": dict(zip(headers, row)),
                }
            )
        return records

    def test_update_reference_table_accepts_title_case_headers(self):
        ref_df = pd.DataFrame(
            [
                {"Price": 12.34, "Style NO.": "STYLE-1", "Article NO.": "ART-1"},
                {"Price": 99.0, "Style NO.": "STYLE-2", "Article NO.": "ART-2"},
            ]
        )
        invoice_data = {
            ("ART-1", "STYLE-1"): {"invoice-a.xls": 12.34},
            ("ART-2", "STYLE-2"): {"invoice-b.xls": 88.0},
        }

        result_df, matches = self.module.update_reference_table(ref_df, invoice_data)

        self.assertEqual(matches, {"一致": 1, "不一致": 1, "未找到": 0})
        self.assertEqual(result_df.loc[0, "核对状态"], "一致")
        self.assertEqual(result_df.loc[1, "核对状态"], "价格不一致")
        self.assertEqual(result_df.loc[1, "发票价格"], 88.0)

    def test_update_reference_table_classifies_two_field_diagnostics(self):
        ref_df = self._build_reference_df()
        invoice_data = {
            ("ART-OK", "STYLE-OK"): {"ok.xls": 12.34},
            ("ART-PRICE", "STYLE-PRICE"): {"price.xls": 19.5},
            ("ART-STYLE", "STYLE-WRONG"): {"style.xls": 10.0},
            ("ART-WRONG", "STYLE-ARTICLE"): {"article.xls": 11.0},
            ("ART-MULTI", "STYLE-WRONG"): {"multi.xls": 7.0},
        }

        result_df, matches = self.module.update_reference_table(ref_df, invoice_data)

        self.assertEqual(matches, {"一致": 1, "不一致": 5, "未找到": 1})
        status_by_article = dict(zip(result_df["Article NO."], result_df["核对状态"]))
        field_by_article = dict(zip(result_df["Article NO."], result_df["疑似错误字段"]))

        self.assertEqual(status_by_article["ART-OK"], "一致")
        self.assertEqual(status_by_article["ART-PRICE"], "价格不一致")
        self.assertEqual(status_by_article["ART-STYLE"], "发票款号疑似错误")
        self.assertEqual(status_by_article["ART-ARTICLE"], "发票Article疑似错误")
        self.assertEqual(status_by_article["ART-MULTI"], "字段疑似错误-候选多条")
        self.assertEqual(status_by_article["ART-NOT-INVOICE"], "未在发票中找到")

        self.assertEqual(field_by_article["ART-STYLE"], "款号")
        self.assertEqual(field_by_article["ART-ARTICLE"], "Article")
        self.assertEqual(field_by_article["ART-MULTI"], "款号")
        self.assertEqual(
            result_df.loc[result_df["Article NO."] == "ART-STYLE", "建议参考款号"].iloc[0],
            "STYLE-GOOD",
        )
        self.assertEqual(
            result_df.loc[result_df["Article NO."] == "ART-ARTICLE", "建议参考Article"].iloc[0],
            "ART-ARTICLE",
        )

    def test_summary_lookup_accepts_title_case_headers(self):
        ref_df = pd.DataFrame(
            [
                {"Price": 12.34, "Style NO.": "STYLE-1", "Article NO.": "ART-1"},
                {"Price": 99.0, "Style NO.": "STYLE-2", "Article NO.": "ART-2"},
            ]
        )
        invoice_data = {
            ("ART-1", "STYLE-1"): {"invoice-a.xls": 12.34},
            ("ART-2", "STYLE-2"): {"invoice-b.xls": 88.0},
        }
        result_df, _ = self.module.update_reference_table(ref_df, invoice_data)

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "result.xlsx")
            save_result = self.module.save_excel_with_summary(
                result_df,
                invoice_data,
                ["invoice-a.xls", "invoice-b.xls"],
                ["invoice-a.xls", "invoice-b.xls"],
                ref_df,
                output_path,
            )

        self.assertEqual(save_result["missing_count"], 0)
        self.assertEqual(save_result["data_count"], 2)

    def test_summary_sheet_outputs_diagnostics_and_status_colors(self):
        ref_df = self._build_reference_df()
        invoice_data = {
            ("ART-OK", "STYLE-OK"): {"ok.xls": 12.34},
            ("ART-STYLE", "STYLE-WRONG"): {"style.xls": 10.0},
            ("ART-WRONG", "STYLE-ARTICLE"): {"article.xls": 11.0},
            ("ART-MULTI", "STYLE-WRONG"): {"multi.xls": 7.0},
            ("ART-MISSING", "STYLE-MISSING"): {"missing.xls": 99.0},
        }
        result_df, _ = self.module.update_reference_table(ref_df, invoice_data)

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "result.xlsx")
            save_result = self.module.save_excel_with_summary(
                result_df,
                invoice_data,
                ["ok.xls", "style.xls", "article.xls", "multi.xls", "missing.xls"],
                ["ok.xls", "style.xls", "article.xls", "multi.xls", "missing.xls"],
                ref_df,
                output_path,
            )
            workbook = load_workbook(output_path)

        self.assertEqual(save_result["diagnostics"]["一致"], 1)
        self.assertEqual(save_result["diagnostics"]["发票款号疑似错误"], 1)
        self.assertEqual(save_result["diagnostics"]["发票Article疑似错误"], 1)
        self.assertEqual(save_result["diagnostics"]["字段疑似错误-候选多条"], 1)
        self.assertEqual(save_result["diagnostics"]["参考表未找到"], 1)

        summary_records = self._sheet_records(workbook["汇总表"])
        summary_by_invoice = {
            record["values"]["来源发票"]: record
            for record in summary_records
            if record["values"].get("来源发票") not in (None, "-")
        }

        style_record = summary_by_invoice["style.xls"]
        article_record = summary_by_invoice["article.xls"]
        multi_record = summary_by_invoice["multi.xls"]
        missing_record = summary_by_invoice["missing.xls"]

        self.assertEqual(style_record["values"]["状态"], "发票款号疑似错误")
        self.assertEqual(style_record["values"]["疑似错误字段"], "款号")
        self.assertEqual(style_record["values"]["建议参考款号"], "STYLE-GOOD")
        self.assertEqual(article_record["values"]["状态"], "发票Article疑似错误")
        self.assertEqual(article_record["values"]["建议参考Article"], "ART-ARTICLE")
        self.assertEqual(multi_record["values"]["状态"], "字段疑似错误-候选多条")
        self.assertEqual(missing_record["values"]["状态"], "参考表未找到")
        self.assertEqual(style_record["cells"]["状态"].fill.fgColor.rgb, "FFFFDDDD")
        self.assertEqual(missing_record["cells"]["状态"].fill.fgColor.rgb, "FFFFFFCC")

        main_records = self._sheet_records(workbook["核对结果"])
        main_by_article = {
            record["values"]["Article NO."]: record
            for record in main_records
            if record["values"].get("Article NO.")
        }
        self.assertEqual(main_by_article["ART-STYLE"]["values"]["核对状态"], "发票款号疑似错误")
        self.assertEqual(main_by_article["ART-NOT-INVOICE"]["values"]["核对状态"], "未在发票中找到")
        self.assertEqual(main_by_article["ART-STYLE"]["cells"]["核对状态"].fill.fgColor.rgb, "FFFFDDDD")
        self.assertEqual(main_by_article["ART-NOT-INVOICE"]["cells"]["核对状态"].fill.fgColor.rgb, "FFFFFFCC")

    def test_read_invoice_records_extracts_header_po_quantity_and_price(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            invoice_path = os.path.join(temp_dir, "invoice.xlsx")
            wb = Workbook()
            ws = wb.active
            ws["I4"] = "Inv#:"
            ws["J4"] = "10-06-26-0712"
            ws["I7"] = "Date:"
            ws["J7"] = date(2026, 6, 7)
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([" PO NO.", "STOCK NO.", "COLOR", "DESCRIPTIONS", None, "QTY(PC)", "UNIT PRICE(PC)", None, "FOB"])
            ws.append(["WOMEN'S KNITTED SHORT"])
            ws.append(["PO:", "0902694555", None, None, None, 200, "USD", 6.6, "USD", 1320.0])
            ws.append(["ARTICLE NO.:", "LG4321"])
            ws.append(["STYLE NO.:", "RC2620OW008"])
            wb.save(invoice_path)

            records = self.module.read_invoice_records(invoice_path)

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].invoice_file, "invoice.xlsx")
        self.assertEqual(records[0].invoice_number, "10-06-26-0712")
        self.assertEqual(records[0].invoice_date, "2026-06-07")
        self.assertEqual(records[0].po_number, "0902694555")
        self.assertEqual(records[0].article, "LG4321")
        self.assertEqual(records[0].style, "RC2620OW008")
        self.assertEqual(records[0].quantity, 200)
        self.assertEqual(records[0].price, 6.6)
        self.assertEqual(records[0].line_amount, 1320.0)
        self.assertEqual(records[0].goods_description, "WOMEN'S KNITTED SHORT")

    def test_read_invoice_summary_extracts_bottom_total_and_final_total(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            invoice_path = os.path.join(temp_dir, "invoice.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.append(["PO NO.", "STOCK NO.", "COLOR", "DESCRIPTIONS", None, "QTY(PC)", "UNIT PRICE(PC)", None, "FOB"])
            ws.append(["WOMEN'S 100% POLYESTER WOVEN JACKET"])
            ws.append(["PO:", "0901889028", None, None, None, 305, "USD", 13.0, "USD", 3965.0])
            ws.append(["ARTICLE NO.:", "KX1885"])
            ws.append(["STYLE NO.:", "RC2610OW007"])
            ws.append(["WOMEN'S 63% POLYESTER WOVEN PANTS"])
            ws.append(["PO:", "0901937666", None, None, None, 140, "USD", 16.05, "USD", 2247.0])
            ws.append(["ARTICLE NO.:", "KX1870"])
            ws.append(["STYLE NO.:", "RC2610OW001"])
            ws.append([])
            ws.append([None, None, "Total", None, None, 445, None, None, "USD", 6212.0])
            ws.append([])
            ws.append([None, None, "Freight Charge", None, None, None, None, None, "USD", 39.06])
            ws.append([None, None, "DOCUMENTATION CHARGE", None, None, None, None, None, "USD", 100.0])
            ws.append([None, None, "Agreed Discount", None, None, None, None, None, "USD", -50.0])
            ws.append([None, None, "Final Total", None, None, None, None, None, "USD", 6301.06])
            wb.save(invoice_path)

            records = self.module.read_invoice_records(invoice_path)
            summary = self.module.read_invoice_summary(invoice_path, records)

        self.assertEqual(summary.source_file, "invoice.xlsx")
        self.assertEqual(summary.total_quantity, 445)
        self.assertEqual(summary.total_amount, 6212.0)
        self.assertEqual(summary.freight_charge, 39.06)
        self.assertEqual(summary.documentation_charge, 100.0)
        self.assertEqual(summary.agreed_discount, -50.0)
        self.assertEqual(summary.final_total_amount, 6301.06)
        self.assertEqual(summary.currency, "USD")

    def test_read_invoice_summary_accepts_doc_charge_alias(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            invoice_path = os.path.join(temp_dir, "invoice.xlsx")
            wb = Workbook()
            ws = wb.active
            ws.append(["PO NO.", "STOCK NO.", "COLOR", "DESCRIPTIONS", None, "QTY(PC)", "UNIT PRICE(PC)", None, "FOB"])
            ws.append(["KID'S WOVEN DENIM JACKET,MAIN MATERIAL:100% COTTON"])
            ws.append(["PO#", "0902638287", None, None, None, 100, "USD", 26.65, "USD", 2665.0])
            ws.append(["ARTICLE NO.:", "KS1734"])
            ws.append(["STYLE NO.:", "OJF26250703"])
            ws.append([])
            ws.append([None, None, "Total", None, None, 100, None, None, "USD", 2665.0])
            ws.append([None, None, "FREIGHT CHARGE", None, None, None, None, None, "USD", 45.0])
            ws.append([None, None, "DOC CHARGE", None, None, None, None, None, "USD", 100.0])
            ws.append([None, None, "Final Total", None, None, None, None, None, "USD", 2810.0])
            wb.save(invoice_path)

            records = self.module.read_invoice_records(invoice_path)
            summary = self.module.read_invoice_summary(invoice_path, records)

        self.assertEqual(summary.documentation_charge, 100.0)

    def test_parse_optional_number_treats_parentheses_as_negative(self):
        self.assertEqual(self.module._parse_optional_number("(795.31)"), -795.31)
        self.assertEqual(
            self.module._extract_labeled_number("Agreed Discount (795.31)", "Agreed Discount"),
            -795.31,
        )

    def test_read_invoice_records_applies_trailing_article_to_multiple_po_rows(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            invoice_path = os.path.join(temp_dir, "invoice.xlsx")
            wb = Workbook()
            ws = wb.active
            ws["G4"] = "INV NO.:"
            ws["H4"] = "10-01-26-0076"
            ws["G7"] = "DATE:"
            ws["H7"] = date(2026, 2, 1)
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append(["PO NO.", "STOCK NO.", "COLOR", "DESCRIPTIONS", None, "QTY(PC)", "UNIT PRICE(PC)", None, "FOB"])
            ws.append(["WOMEN'S WOVEN DENIM SKIRT,MAIN MATERIAL: 100% COTTON"])
            ws.append(["PO#", "0901792204", None, None, None, 317, "USD", 12.67, "USD", 4016.39])
            ws.append(["PO#", "0901792050", None, None, None, 313, "USD", 12.67, "USD", 3965.71])
            ws.append(["ARTICLE NO.:", "KY8114"])
            ws.append(["STYLE NO.:", "RC2609OW007"])
            wb.save(invoice_path)

            records = self.module.read_invoice_records(invoice_path)

        self.assertEqual(len(records), 2)
        self.assertEqual([record.po_number for record in records], ["0901792204", "0901792050"])
        self.assertEqual([record.article for record in records], ["KY8114", "KY8114"])
        self.assertEqual([record.style for record in records], ["RC2609OW007", "RC2609OW007"])
        self.assertEqual([record.price for record in records], [12.67, 12.67])

    def test_update_reference_table_ignores_multi_po_trailing_article_noise(self):
        ref_df = pd.DataFrame(
            [
                {"Price": 12.67, "Style NO.": "RC2609OW007", "Article NO.": "KY8113"},
                {"Price": 12.67, "Style NO.": "RC2609OW007", "Article NO.": "KY8114"},
            ]
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            invoice_path = os.path.join(temp_dir, "0076.xlsx")
            wb = Workbook()
            ws = wb.active
            ws["G4"] = "INV NO.:"
            ws["H4"] = "10-01-26-0076"
            ws["G7"] = "DATE:"
            ws["H7"] = date(2026, 2, 1)
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append(["PO NO.", "STOCK NO.", "COLOR", "DESCRIPTIONS", None, "QTY(PC)", "UNIT PRICE(PC)", None, "FOB"])
            ws.append(["WOMEN'S WOVEN DENIM SKIRT,MAIN MATERIAL: 100% COTTON"])
            ws.append(["PO#", "0901792046", None, None, None, 7596, "USD", 12.67, "USD", 96241.32])
            ws.append(["ARTICLE NO.:", "KY8113"])
            ws.append(["STYLE NO.:", "RC2609OW007"])
            ws.append(["WOMEN'S WOVEN DENIM SKIRT,MAIN MATERIAL: 100% COTTON"])
            ws.append(["PO#", "0901792204", None, None, None, 317, "USD", 12.67, "USD", 4016.39])
            ws.append(["PO#", "0901792050", None, None, None, 313, "USD", 12.67, "USD", 3965.71])
            ws.append(["ARTICLE NO.:", "KY8114"])
            ws.append(["STYLE NO.:", "RC2609OW007"])
            wb.save(invoice_path)

            invoice_records = self.module.read_invoice_records(invoice_path)

        invoice_data: Dict[Tuple[str, str], Dict[str, float]] = {}
        for record in invoice_records:
            invoice_data.setdefault((record.article, record.style), {})[record.invoice_file] = record.price

        result_df, matches = self.module.update_reference_table(ref_df, invoice_data)

        self.assertEqual(matches, {"一致": 2, "不一致": 0, "未找到": 0})
        self.assertNotIn("字段疑似错误-候选多条", set(result_df["核对状态"]))
        self.assertEqual(list(result_df["核对状态"]), ["一致", "一致"])

    def test_read_invoice_records_extracts_inline_no_and_month_date_header(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            invoice_path = os.path.join(temp_dir, "invoice.xlsx")
            wb = Workbook()
            ws = wb.active
            ws["J4"] = "NO: 17-05-26-1190"
            ws["J7"] = "DATE: MAY 25 2026"
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([" PO NO.", "STOCK NO.", "COLOR", "DESCRIPTIONS", None, "QTY(PC)", "UNIT PRICE(PC)", None, "FOB"])
            ws.append(["WOMEN'S WOVEN TRACK TOP"])
            ws.append(["PO:", "0902590165", None, None, None, 100, "USD", 15.55, "USD", 1555.0])
            ws.append(["ARTICLE NO.:", "LD5339"])
            ws.append(["STYLE NO.:", "RC2613OW007"])
            wb.save(invoice_path)

            records = self.module.read_invoice_records(invoice_path)

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].invoice_number, "17-05-26-1190")
        self.assertEqual(records[0].invoice_date, "2026-05-25")
        self.assertEqual(records[0].po_number, "0902590165")
        self.assertEqual(records[0].article, "LD5339")
        self.assertEqual(records[0].style, "RC2613OW007")

    def test_read_invoice_records_extracts_split_inv_no_label_header(self):
        with tempfile.TemporaryDirectory() as temp_dir:
            invoice_path = os.path.join(temp_dir, "invoice.xlsx")
            wb = Workbook()
            ws = wb.active
            ws["G4"] = "INV NO.:"
            ws["H4"] = "10-04-26-0460"
            ws["G7"] = "DATE:"
            ws["H7"] = date(2026, 4, 15)
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([])
            ws.append([" PO NO.", "STOCK NO.", "COLOR", "DESCRIPTIONS", None, "QTY(PC)", "UNIT PRICE(PC)", None, "FOB"])
            ws.append(["WOMEN'S WOVEN ST DNM WSH JACKET"])
            ws.append(["PO#", "0902187227", None, None, None, 342, "USD", 21.8, "USD", 7455.6])
            ws.append(["ARTICLE NO.:", "KV0625"])
            ws.append(["STYLE NO.:", "AF26INSPW072"])
            wb.save(invoice_path)

            records = self.module.read_invoice_records(invoice_path)

        self.assertEqual(len(records), 1)
        self.assertEqual(records[0].invoice_number, "10-04-26-0460")
        self.assertEqual(records[0].invoice_date, "2026-04-15")
        self.assertEqual(records[0].po_number, "0902187227")
        self.assertEqual(records[0].article, "KV0625")
        self.assertEqual(records[0].style, "AF26INSPW072")

    def test_parse_tc_invoice_pages_extracts_rows_and_goods_descriptions(self):
        pages = [
            TcInvoiceExtractedPage(
                text=(
                    "Invoice Number\n17-04-26-0914\n"
                    "PO No PO Line Market PO Sales Order No Working No Article No Article Description Gender Category Total Qty\n"
                    "0901937666 1 0305837705 5052174707 RC2610OW001 KX1870 CLASSIC TP CRSK W ORIGINALS 140\n"
                    "Customer Size XS S M L XL Total Quantity Price Total Amount\n"
                    "QTY 28 44 36 20 12 140 17.65 2,471.00\n"
                    "Goods Description WOMEN'S 63% POLYESTER (100% RECYCLED) 34% VISCOSE,3% ELASTANE WOVEN PANTS (HTS:6204.63.0000)\n"
                    "Customer Size Sourcing Size/Manufacturing Size HTS Description Quota Category Export Quota Price Quantity\n"
                    "XS,S,M,L,XL 32,36,40,44,48 6204630000 WOMENS PANTS,SHORTS,ETC;WOVEN 140\n"
                    "PO No PO Line Market PO Sales Order No Working No Article No Article Description Gender Category Total Qty\n"
                    "0901889028 1 0305837711 5052174713 RC2610OW007 KX1885 STN FB TT LINEN/MAGBEI W ORIGINALS 305\n"
                    "Customer Size XS S M L XL Total Quantity Price Total Amount\n"
                    "QTY 55 87 83 52 28 305 14.25 4,346.25\n"
                    "Goods Description WOMEN'S 100% POLYESTER (100% RECYCLED) WOVEN JACKET(HTS:6202.40.0000)\n"
                    "Customer Size Sourcing Size/Manufacturing Size HTS Description Quota Category Export Quota Price Quantity\n"
                    "XS,S,M,L,XL 32,36,40,44,48 6202400000 WOMANS ANORAKS,WINDCHEATERS;WOVEN 305\n"
                    "Total Quantity 445\n"
                    "Total Carton 19\n"
                    "Total Gross Weight 200.740 KG\n"
                    "Total Net Weight 180.860 KG\n"
                    "Total Net Net Weight 180.670 KG\n"
                    "Total PO Net Amount 6,817.25\n"
                    "Agreed Discount (50.00)\n"
                    "Additional charge 39.06\n"
                    "Documentation charge 100.00\n"
                    "Total VAT 0.00\n"
                    "Invoice Total 6,817.25"
                ),
                tables=[
                    [
                        [
                            "PO No",
                            "PO Line\nAggregator",
                            "Market PO\nNumber",
                            "Sales Order No",
                            "Working No",
                            "Article No",
                            "Article Description",
                            "Gender",
                            "Category",
                            "Total\nQty",
                        ],
                        [
                            "0901937666",
                            "1",
                            "0305837705",
                            "5052174707",
                            "RC2610OW001",
                            "KX1870",
                            "CLASSIC TP CRSK",
                            "W",
                            "ORIGINALS",
                            "140",
                        ],
                    ],
                    [
                        ["Customer Size", "XS", "S", "M", "L", "XL", "Total Quantity", "Price", "Total Amount"],
                        ["QTY", "28", "44", "36", "20", "12", "140", "17.65", "2,471.00"],
                        ["", None, None, None, None, None, None, None, "2,471.00"],
                    ],
                    [
                        [
                            "PO No",
                            "PO Line\nAggregator",
                            "Market PO\nNumber",
                            "Sales Order No",
                            "Working No",
                            "Article No",
                            "Article Description",
                            "Gender",
                            "Category",
                            "Total\nQty",
                        ],
                        [
                            "0901889028",
                            "1",
                            "0305837711",
                            "5052174713",
                            "RC2610OW007",
                            "KX1885",
                            "STN FB TT LINEN/MAGBEI",
                            "W",
                            "ORIGINALS",
                            "305",
                        ],
                    ],
                    [
                        ["Customer Size", "XS", "S", "M", "L", "XL", "Total Quantity", "Price", "Total Amount"],
                        ["QTY", "55", "87", "83", "52", "28", "305", "14.25", "4,346.25"],
                    ],
                ],
            )
        ]

        records = self.module.parse_tc_invoice_pages(pages, "tc.pdf")
        summary = self.module.parse_tc_invoice_summary_pages(pages, "tc.pdf")

        self.assertEqual(len(records), 2)
        self.assertEqual(records[0].source_file, "tc.pdf")
        self.assertEqual(records[0].invoice_number, "17-04-26-0914")
        self.assertEqual(records[0].po_number, "0901937666")
        self.assertEqual(records[0].market_po, "0305837705")
        self.assertEqual(records[0].working_number, "RC2610OW001")
        self.assertEqual(records[0].article_number, "KX1870")
        self.assertEqual(records[0].quantity, 140)
        self.assertEqual(records[0].price, 17.65)
        self.assertEqual(records[0].total_amount, 2471.0)
        self.assertEqual(
            records[0].goods_description,
            "WOMEN'S 63% POLYESTER (100% RECYCLED) 34% VISCOSE,3% ELASTANE WOVEN PANTS",
        )
        self.assertEqual(records[1].quantity, 305)
        self.assertEqual(records[1].invoice_number, "17-04-26-0914")
        self.assertEqual(records[1].price, 14.25)
        self.assertEqual(records[1].total_amount, 4346.25)
        self.assertEqual(
            records[1].goods_description,
            "WOMEN'S 100% POLYESTER (100% RECYCLED) WOVEN JACKET",
        )
        self.assertEqual(summary.source_file, "tc.pdf")
        self.assertEqual(summary.invoice_number, "17-04-26-0914")
        self.assertEqual(summary.total_quantity, 445)
        self.assertEqual(summary.total_carton, 19)
        self.assertEqual(summary.total_gross_weight, 200.740)
        self.assertEqual(summary.total_net_weight, 180.860)
        self.assertEqual(summary.total_net_net_weight, 180.670)
        self.assertEqual(summary.total_po_net_amount, 6817.25)
        self.assertEqual(summary.agreed_discount, -50.0)
        self.assertEqual(summary.additional_charge, 39.06)
        self.assertEqual(summary.documentation_charge, 100.0)
        self.assertEqual(summary.total_vat, 0)
        self.assertEqual(summary.invoice_total, 6817.25)

    def test_parse_tc_invoice_number_skips_log_no_header_token(self):
        pages = [
            TcInvoiceExtractedPage(
                text=(
                    "TMS FASHION (H.K.) LIMITED Invoice as of Thu Jun 18 02:09:03 UTC 2026\n"
                    "UNIT 1001-1002, 10 FLOOR Invoice Number Log No\n"
                    "NANYANG PLAZA KOWLOON\n"
                    "57 HUNG TO ROAD 10-06-26-0748 565799195730\n"
                    "PO No PO Line Market PO Sales Order No Working No Article No Article Description Gender Category Total Qty\n"
                    "0902648443 1 0307154286 RC2613OW008 LE8309 CHECKS TROUSER CRLI/AUCO/ORBGRN/vermil W ORIGINALS 100\n"
                    "Customer Size S M XS L XL Total Quantity Price Total Amount\n"
                    "QTY 26 25 19 19 11 100 18.15 1,815.00\n"
                    "Total Quantity 100\n"
                    "Total PO Net Amount 2,364.51\n"
                    "Invoice Total 2,482.51"
                ),
                tables=[
                    [
                        [
                            "PO No",
                            "PO Line\nAggregator",
                            "Market PO\nNumber",
                            "Sales Order No",
                            "Working No",
                            "Article No",
                            "Article Description",
                            "Gender",
                            "Category",
                            "Total\nQty",
                        ],
                        [
                            "0902648443",
                            "1",
                            "0307154286",
                            "",
                            "RC2613OW008",
                            "LE8309",
                            "CHECKS TROUSER",
                            "W",
                            "ORIGINALS",
                            "100",
                        ],
                    ],
                    [
                        ["Customer Size", "S", "M", "XS", "L", "XL", "Total Quantity", "Price", "Total Amount"],
                        ["QTY", "26", "25", "19", "19", "11", "100", "18.15", "1,815.00"],
                    ],
                ],
            )
        ]

        records = self.module.parse_tc_invoice_pages(pages, "tc.pdf")
        summary = self.module.parse_tc_invoice_summary_pages(pages, "tc.pdf")

        self.assertEqual(records[0].invoice_number, "10-06-26-0748")
        self.assertEqual(summary.invoice_number, "10-06-26-0748")

    def test_build_tc_invoice_comparison_matches_invoice_and_flags_quantity(self):
        invoice_records = [
            InvoiceRecord(
                invoice_file="invoice.xlsx",
                invoice_number="10-06-26-0712",
                invoice_date="2026-06-07",
                po_number="0902694555",
                article="LG4321",
                style="RC2620OW008",
                quantity=200,
                price=6.6,
                line_amount=1320.0,
                goods_description="WOMEN'S 100% POLYESTER (100%RECYCLED) WOVEN JACKET",
            ),
            InvoiceRecord(
                invoice_file="invoice.xlsx",
                invoice_number="10-06-26-0712",
                invoice_date="2026-06-07",
                po_number="0902694557",
                article="LG4323",
                style="RC2620OW008",
                quantity=200,
                price=6.6,
                line_amount=1320.0,
                goods_description="WOMEN'S KNITTED SHORT MAIN MATERIAL 100% POLYESTER",
            ),
        ]
        tc_records = [
            TcInvoiceRecord(
                source_file="tc-a.pdf",
                po_number="0902694555",
                market_po="0307073961",
                working_number="RC2620OW008",
                article_number="LG4321",
                quantity=200,
                price=6.6,
                total_amount=1320.0,
                goods_description="WOMEN'S 100% POLYESTER (100% RECYCLED) WOVEN JACKET(HTS:6202.40.0000)",
            ),
            TcInvoiceRecord(
                source_file="tc-a.pdf",
                po_number="0902694557",
                market_po="0307073963",
                working_number="RC2620OW008",
                article_number="LG4323",
                quantity=180,
                price=6.6,
                total_amount=1188.0,
                goods_description="WOMEN'S KNITTED SHORT MAIN MATERIAL 100% POLYESTER",
            ),
        ]

        rows = self.module.build_tc_invoice_comparison(invoice_records, tc_records)

        self.assertEqual(len(rows), 2)
        self.assertEqual(rows[0].status, "一致")
        self.assertEqual(rows[0].issue_detail, "")
        self.assertEqual(rows[1].status, "需核对")
        self.assertIn("QTY", rows[1].issue_detail)
        self.assertEqual(rows[1].fty_quantity, 200)
        self.assertEqual(rows[1].tc_quantity, 180)

    def test_build_tc_invoice_comparison_pairs_invoice_group_remaining_rows_by_order(self):
        invoice_records = [
            InvoiceRecord(
                invoice_file="FTY INV.xls",
                invoice_number="10-06-26-0712",
                invoice_date="2026-06-07",
                po_number="0902694555",
                article="LG4321",
                style="RC2620OW008",
                quantity=200,
                price=6.6,
                line_amount=1320.0,
                goods_description="WOMEN'S WOVEN JACKET",
            )
        ]
        tc_records = [
            TcInvoiceRecord(
                source_file="TC INV.pdf",
                invoice_number="10-06-26-0712",
                po_number="0902694999",
                market_po="0307073961",
                working_number="RC2620OW999",
                article_number="LG4999",
                quantity=200,
                price=6.6,
                total_amount=1320.0,
                goods_description="WOMEN'S WOVEN JACKET",
            )
        ]

        rows = self.module.build_tc_invoice_comparison(invoice_records, tc_records)

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].status, "需核对")
        self.assertIn("PO No", rows[0].issue_detail)
        self.assertIn("Article No", rows[0].issue_detail)
        self.assertIn("Working/Style No", rows[0].issue_detail)
        self.assertEqual(rows[0].fty_po_number, "0902694555")
        self.assertEqual(rows[0].tc_po_number, "0902694999")
        self.assertEqual(rows[0].fty_article, "LG4321")
        self.assertEqual(rows[0].tc_article, "LG4999")
        self.assertEqual(rows[0].fty_style, "RC2620OW008")
        self.assertEqual(rows[0].tc_style, "RC2620OW999")

        wb = Workbook()
        ws = wb.active
        self.module._write_tc_invoice_comparison_sheet(ws, rows, [])
        sheet_rows = list(ws.iter_rows(values_only=True))
        self.assertNotIn(
            ("Invoice Number", "PO No", "Article No", "Working/Style No", "核对项", "来源", "数值", "差额", "结果"),
            [row[:9] for row in sheet_rows],
        )
        matrix_header_index = next(
            row_index for row_index, row in enumerate(sheet_rows, 1)
            if row[:12] == (
                "来源",
                "Invoice Number",
                "PO No",
                "Article No",
                "Working/Style No",
                "Quantity",
                "Total Amount / Total PO Net Amount",
                "Freight/Additional Charge",
                "Documentation Charge",
                "Agreed Discount",
                "Final Total / Invoice Total",
                "Goods Description",
            )
        )
        self.assertEqual(sheet_rows[matrix_header_index][:12], ("FTY", "10-06-26-0712", "0902694555", "LG4321", "RC2620OW008", 200, "-", "-", "-", "-", "-", "WOMEN'S WOVEN JACKET"))
        self.assertEqual(sheet_rows[matrix_header_index + 1][:12], ("TC", "10-06-26-0712", "0902694999", "LG4999", "RC2620OW999", 200, "-", "-", "-", "-", "-", "WOMEN'S WOVEN JACKET"))
        self.assertEqual(sheet_rows[matrix_header_index + 2][:12], ("结果", "一致", "不一致", "不一致", "不一致", "一致", "一致", "一致", "一致", "一致", "一致", "一致"))
        for column_index in (3, 4, 5):
            self.assertEqual(ws.cell(matrix_header_index + 1, column_index).fill.fgColor.rgb, "FFFFDDDD")
            self.assertEqual(ws.cell(matrix_header_index + 2, column_index).fill.fgColor.rgb, "FFFFDDDD")
            self.assertEqual(ws.cell(matrix_header_index + 3, column_index).fill.fgColor.rgb, "FFFFDDDD")
        self.assertNotEqual(ws.cell(matrix_header_index + 1, 1).fill.fgColor.rgb, "FFFFDDDD")
        self.assertNotEqual(ws.cell(matrix_header_index + 2, 1).fill.fgColor.rgb, "FFFFDDDD")
        self.assertEqual(ws.cell(matrix_header_index + 3, 6).fill.fgColor.rgb, "FFDDFFDD")

    def test_build_tc_invoice_comparison_ignores_detail_price_and_line_amount(self):
        invoice_records = [
            InvoiceRecord(
                invoice_file="FTY INV.xls",
                invoice_number="17-04-26-0914",
                invoice_date="2026-04-04",
                po_number="0901889028",
                article="KX1885",
                style="RC2610OW007",
                quantity=305,
                price=13.0,
                line_amount=3965.0,
                goods_description="WOMEN'S 100% POLYESTER (100% RECYCLED) WOVEN JACKET",
            )
        ]
        tc_records = [
            TcInvoiceRecord(
                source_file="TC INV.pdf",
                po_number="0901889028",
                market_po="0305837711",
                working_number="RC2610OW007",
                article_number="KX1885",
                quantity=305,
                price=14.25,
                total_amount=4346.25,
                goods_description="WOMEN'S 100% POLYESTER (100% RECYCLED) WOVEN JACKET",
            )
        ]

        rows = self.module.build_tc_invoice_comparison(invoice_records, tc_records)

        self.assertEqual(rows[0].status, "一致")
        self.assertNotIn("Unit Price", rows[0].issue_detail)
        self.assertNotIn("Line Amount", rows[0].issue_detail)
        self.assertEqual(rows[0].fty_price, 13.0)
        self.assertEqual(rows[0].tc_price, 14.25)
        self.assertEqual(rows[0].fty_line_amount, 3965.0)
        self.assertEqual(rows[0].tc_total_amount, 4346.25)

    def test_build_tc_invoice_summary_comparison_flags_total_amounts(self):
        rows = self.module.build_tc_invoice_summary_comparison(
            [
                InvoiceSummaryRecord(
                    source_file="FTY INV.xls",
                    invoice_number="17-04-26-0914",
                    total_quantity=445,
                    total_amount=6212.0,
                    freight_charge=39.06,
                    documentation_charge=100.0,
                    final_total_amount=6212.0,
                    currency="USD",
                )
            ],
            [
                TcInvoiceSummary(
                    source_file="TC INV.pdf",
                    invoice_number="17-04-26-0914",
                    total_quantity=445,
                    total_carton=19,
                    total_gross_weight=200.740,
                    total_net_weight=180.860,
                    total_net_net_weight=180.670,
                    total_po_net_amount=6817.25,
                    additional_charge=39.06,
                    documentation_charge=100.0,
                    total_vat=0,
                    invoice_total=6817.25,
                )
            ],
        )

        self.assertEqual(len(rows), 1)
        self.assertEqual(rows[0].status, "需核对")
        self.assertIn("Total Amount", rows[0].issue_detail)
        self.assertIn("Final Total", rows[0].issue_detail)
        self.assertEqual(rows[0].fty_total_quantity, 445)
        self.assertEqual(rows[0].tc_total_quantity, 445)
        self.assertEqual(rows[0].fty_total_amount, 6212.0)
        self.assertEqual(rows[0].tc_total_po_net_amount, 6817.25)
        self.assertEqual(rows[0].fty_freight_charge, 39.06)
        self.assertEqual(rows[0].tc_additional_charge, 39.06)
        self.assertEqual(rows[0].fty_documentation_charge, 100.0)
        self.assertEqual(rows[0].tc_documentation_charge, 100.0)
        self.assertNotIn("Freight/Additional Charge", rows[0].issue_detail)
        self.assertNotIn("Documentation Charge", rows[0].issue_detail)

    def test_build_tc_invoice_summary_comparison_flags_charge_mismatch(self):
        rows = self.module.build_tc_invoice_summary_comparison(
            [
                InvoiceSummaryRecord(
                    source_file="FTY INV.xls",
                    total_quantity=217,
                    total_amount=3532.0,
                    freight_charge=39.06,
                    documentation_charge=100.0,
                    final_total_amount=3671.06,
                    currency="USD",
                )
            ],
            [
                TcInvoiceSummary(
                    source_file="TC INV.pdf",
                    total_quantity=217,
                    total_po_net_amount=3532.0,
                    additional_charge=40.06,
                    documentation_charge=90.0,
                    total_vat=0,
                    invoice_total=3662.06,
                )
            ],
        )

        self.assertEqual(rows[0].status, "需核对")
        self.assertIn("Freight/Additional Charge", rows[0].issue_detail)
        self.assertIn("Documentation Charge", rows[0].issue_detail)
        self.assertEqual(rows[0].fty_freight_charge, 39.06)
        self.assertEqual(rows[0].tc_additional_charge, 40.06)
        self.assertEqual(rows[0].fty_documentation_charge, 100.0)
        self.assertEqual(rows[0].tc_documentation_charge, 90.0)

    def test_tc_invoice_comparison_accepts_matching_agreed_discount_adjustment(self):
        detail_rows = self.module.build_tc_invoice_comparison(
            [
                InvoiceRecord(
                    invoice_file="FTY INV(2).xls",
                    invoice_number="10-04-26-0513",
                    invoice_date="2026-04-30",
                    po_number="0902319229",
                    article="LE8319",
                    style="RC2613OW013",
                    quantity=751,
                    price=15.95,
                    line_amount=11978.45,
                    goods_description="WOMEN'S KNIT TRACK TOP,MAIN MATERIAL: 100% COTTON",
                )
            ],
            [
                TcInvoiceRecord(
                    source_file="TC Invoice 10-04-26-0513.pdf",
                    po_number="0902319229",
                    market_po="0306692572",
                    working_number="RC2613OW013",
                    article_number="LE8319",
                    quantity=751,
                    price=17.65,
                    total_amount=13255.15,
                    goods_description="WOMEN'S KNIT TRACK TOP,MAIN MATERIAL: 100% COTTON",
                )
            ],
            suppress_amount_differences=True,
        )
        summary_rows = self.module.build_tc_invoice_summary_comparison(
            [
                InvoiceSummaryRecord(
                    source_file="FTY INV(2).xls",
                    invoice_number="10-04-26-0513",
                    total_quantity=751,
                    total_amount=11978.45,
                    freight_charge=135.18,
                    agreed_discount=-795.31,
                    final_total_amount=11318.32,
                    currency="USD",
                )
            ],
            [
                TcInvoiceSummary(
                    source_file="TC Invoice 10-04-26-0513.pdf",
                    invoice_number="10-04-26-0513",
                    total_quantity=751,
                    total_po_net_amount=13323.27,
                    agreed_discount=-795.31,
                    additional_charge=135.18,
                    total_vat=0,
                    invoice_total=12663.14,
                )
            ],
        )

        self.assertEqual(detail_rows[0].status, "一致")
        self.assertEqual(detail_rows[0].issue_detail, "")
        self.assertEqual(summary_rows[0].status, "一致")
        self.assertEqual(summary_rows[0].fty_agreed_discount, -795.31)
        self.assertEqual(summary_rows[0].tc_agreed_discount, -795.31)
        self.assertNotIn("Total Amount", summary_rows[0].issue_detail)
        self.assertNotIn("Final Total", summary_rows[0].issue_detail)

        wb = Workbook()
        ws = wb.active
        self.module._write_tc_invoice_comparison_sheet(ws, detail_rows, summary_rows)
        rows = list(ws.iter_rows(values_only=True))
        self.assertNotIn("核对结论", [row[0] for row in rows if row and row[0]])
        self.assertNotIn("汇总差异", [row[0] for row in rows if row and row[0]])
        self.assertNotIn(
            ("Invoice Number", "PO No", "Article No", "Working/Style No", "核对项", "来源", "数值", "差额", "结果"),
            [row[:9] for row in rows],
        )
        matrix_header_index = next(
            index for index, row in enumerate(rows, 1)
            if row[:12] == (
                "来源",
                "Invoice Number",
                "PO No",
                "Article No",
                "Working/Style No",
                "Quantity",
                "Total Amount / Total PO Net Amount",
                "Freight/Additional Charge",
                "Documentation Charge",
                "Agreed Discount",
                "Final Total / Invoice Total",
                "Goods Description",
            )
        )
        self.assertEqual(
            rows[matrix_header_index][:12],
            ("FTY", "10-04-26-0513", "0902319229", "LE8319", "RC2613OW013", 751, 11978.45, 135.18, "-", -795.31, 11318.32, "WOMEN'S KNIT TRACK TOP,MAIN MATERIAL: 100% COTTON"),
        )
        self.assertEqual(
            rows[matrix_header_index + 1][:12],
            ("TC", "10-04-26-0513", "0902319229", "LE8319", "RC2613OW013", 751, 13323.27, 135.18, "-", -795.31, 12663.14, "WOMEN'S KNIT TRACK TOP,MAIN MATERIAL: 100% COTTON"),
        )
        self.assertEqual(rows[matrix_header_index + 2][:12], ("结果", "一致", "一致", "一致", "一致", "一致", "一致", "一致", "一致", "一致", "一致", "一致"))

    def test_build_tc_invoice_comparison_flags_goods_description(self):
        invoice_records = [
            InvoiceRecord(
                invoice_file="invoice.xlsx",
                invoice_number="17-04-26-0914",
                invoice_date="2026-04-04",
                po_number="0901889028",
                article="KX1885",
                style="RC2610OW007",
                quantity=305,
                price=13.0,
                line_amount=3965.0,
                goods_description="WOMEN'S 100% POLYESTER WOVEN JACKET",
            )
        ]
        tc_records = [
            TcInvoiceRecord(
                source_file="tc.pdf",
                po_number="0901889028",
                market_po="0305837711",
                working_number="RC2610OW007",
                article_number="KX1885",
                quantity=305,
                price=13.0,
                total_amount=3965.0,
                goods_description="WOMEN'S 63% POLYESTER WOVEN PANTS",
            )
        ]

        rows = self.module.build_tc_invoice_comparison(invoice_records, tc_records)

        self.assertEqual(rows[0].status, "需核对")
        self.assertIn("Goods Description", rows[0].issue_detail)

    def test_save_excel_with_summary_adds_tc_invoice_sheet_when_rows_exist(self):
        ref_df = pd.DataFrame(
            [{"Price": 6.6, "Style NO.": "RC2620OW008", "Article NO.": "LG4321"}],
        )
        invoice_data = {("LG4321", "RC2620OW008"): {"invoice.xlsx": 6.6}}
        result_df, _ = self.module.update_reference_table(ref_df, invoice_data)
        tc_rows = self.module.build_tc_invoice_comparison(
            [
                InvoiceRecord(
                    invoice_file="invoice.xlsx",
                    invoice_number="10-06-26-0712",
                    invoice_date="2026-06-07",
                    po_number="0902694555",
                    article="LG4321",
                    style="RC2620OW008",
                    quantity=200,
                    price=6.6,
                    goods_description="WOMEN'S 100% POLYESTER (100%RECYCLED) WOVEN JACKET",
                )
            ],
            [
                TcInvoiceRecord(
                    source_file="tc.pdf",
                    po_number="0902694555",
                    market_po="0307073961",
                    working_number="RC2620OW008",
                    article_number="LG4321",
                    quantity=200,
                    goods_description="WOMEN'S 100% POLYESTER (100% RECYCLED) WOVEN JACKET",
                )
            ],
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "result.xlsx")
            save_result = self.module.save_excel_with_summary(
                result_df,
                invoice_data,
                ["invoice.xlsx"],
                ["invoice.xlsx"],
                ref_df,
                output_path,
                tc_comparison_rows=tc_rows,
            )
            workbook = load_workbook(output_path)

        self.assertIn("TC INV核对", workbook.sheetnames)
        self.assertEqual(save_result["tc_count"], 1)
        self.assertEqual(save_result["tc_matched_count"], 1)
        self.assertEqual(save_result["tc_issue_count"], 0)
        tc_sheet = workbook["TC INV核对"]
        rows = list(tc_sheet.iter_rows(values_only=True))
        self.assertNotIn("核对结论", [row[0] for row in rows if row and row[0]])
        self.assertNotIn("汇总差异", [row[0] for row in rows if row and row[0]])
        self.assertNotIn(
            ("Invoice Number", "PO No", "Article No", "Working/Style No", "核对项", "来源", "数值", "差额", "结果"),
            [row[:9] for row in rows],
        )
        matrix_header_index = next(
            index for index, row in enumerate(rows, 1)
            if row[:12] == (
                "来源",
                "Invoice Number",
                "PO No",
                "Article No",
                "Working/Style No",
                "Quantity",
                "Total Amount / Total PO Net Amount",
                "Freight/Additional Charge",
                "Documentation Charge",
                "Agreed Discount",
                "Final Total / Invoice Total",
                "Goods Description",
            )
        )
        self.assertEqual(
            rows[matrix_header_index][:12],
            ("FTY", "10-06-26-0712", "0902694555", "LG4321", "RC2620OW008", 200, "-", "-", "-", "-", "-", "WOMEN'S 100% POLYESTER (100%RECYCLED) WOVEN JACKET"),
        )
        self.assertEqual(
            rows[matrix_header_index + 1][:12],
            ("TC", "10-06-26-0712", "0902694555", "LG4321", "RC2620OW008", 200, "-", "-", "-", "-", "-", "WOMEN'S 100% POLYESTER (100% RECYCLED) WOVEN JACKET"),
        )
        self.assertEqual(rows[matrix_header_index + 2][:12], ("结果", "一致", "一致", "一致", "一致", "一致", "一致", "一致", "一致", "一致", "一致", "一致"))

    def test_save_excel_with_summary_adds_tc_summary_section(self):
        ref_df = pd.DataFrame(
            [{"Price": 13.0, "Style NO.": "RC2610OW007", "Article NO.": "KX1885"}],
        )
        invoice_data = {("KX1885", "RC2610OW007"): {"FTY INV.xls": 13.0}}
        result_df, _ = self.module.update_reference_table(ref_df, invoice_data)
        tc_detail_rows = self.module.build_tc_invoice_comparison(
            [
                InvoiceRecord(
                    invoice_file="FTY INV.xls",
                    invoice_number="17-04-26-0914",
                    invoice_date="2026-04-04",
                    po_number="0901889028",
                    article="KX1885",
                    style="RC2610OW007",
                    quantity=305,
                    price=13.0,
                    line_amount=3965.0,
                    goods_description="WOMEN'S 100% POLYESTER WOVEN JACKET",
                ),
                InvoiceRecord(
                    invoice_file="FTY INV.xls",
                    invoice_number="17-04-26-0914",
                    invoice_date="2026-04-04",
                    po_number="0901937666",
                    article="KX1870",
                    style="RC2610OW001",
                    quantity=140,
                    price=16.05,
                    line_amount=2247.0,
                    goods_description="WOMEN'S 63% POLYESTER 34% VISCOSE 3% ELASTANE WOVEN PANTS",
                )
            ],
            [
                TcInvoiceRecord(
                    source_file="TC INV.pdf",
                    po_number="0901889028",
                    market_po="0305837711",
                    working_number="RC2610OW007",
                    article_number="KX1885",
                    quantity=305,
                    price=14.25,
                    total_amount=4346.25,
                    goods_description="WOMEN'S 100% POLYESTER WOVEN JACKET",
                ),
                TcInvoiceRecord(
                    source_file="TC INV.pdf",
                    po_number="0901937666",
                    market_po="0305837712",
                    working_number="RC2610OW001",
                    article_number="KX1870",
                    quantity=140,
                    price=17.65,
                    total_amount=2471.0,
                    goods_description="WOMEN'S 63% POLYESTER 34% VISCOSE 3% ELASTANE WOVEN PANTS",
                )
            ],
        )
        tc_summary_rows = self.module.build_tc_invoice_summary_comparison(
            [
                InvoiceSummaryRecord(
                    source_file="FTY INV.xls",
                    invoice_number="17-04-26-0914",
                    total_quantity=445,
                    total_amount=6212.0,
                    freight_charge=39.06,
                    documentation_charge=100.0,
                    final_total_amount=6212.0,
                    currency="USD",
                )
            ],
            [
                TcInvoiceSummary(
                    source_file="TC INV.pdf",
                    invoice_number="17-04-26-0914",
                    total_quantity=445,
                    total_carton=19,
                    total_gross_weight=200.740,
                    total_net_weight=180.860,
                    total_net_net_weight=180.670,
                    total_po_net_amount=6817.25,
                    additional_charge=39.06,
                    documentation_charge=100.0,
                    total_vat=0,
                    invoice_total=6817.25,
                )
            ],
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "result.xlsx")
            save_result = self.module.save_excel_with_summary(
                result_df,
                invoice_data,
                ["FTY INV.xls"],
                ["FTY INV.xls"],
                ref_df,
                output_path,
                tc_comparison_rows=tc_detail_rows,
                tc_summary_rows=tc_summary_rows,
            )
            workbook = load_workbook(output_path)

        tc_sheet = workbook["TC INV核对"]
        rows = list(tc_sheet.iter_rows(values_only=True))
        labels = [row[0] for row in rows if row and row[0]]
        self.assertNotIn("核对结论", labels)
        self.assertNotIn("金额差额", labels)
        self.assertNotIn("汇总差异", labels)
        issue_fill_rgb = "FFFFDDDD"
        self.assertNotIn(
            ("Invoice Number", "PO No", "Article No", "Working/Style No", "核对项", "来源", "数值", "差额", "结果"),
            [row[:9] for row in rows],
        )
        matrix_header_index = next(
            index for index, row in enumerate(rows, 1)
            if row[:12] == (
                "来源",
                "Invoice Number",
                "PO No",
                "Article No",
                "Working/Style No",
                "Quantity",
                "Total Amount / Total PO Net Amount",
                "Freight/Additional Charge",
                "Documentation Charge",
                "Agreed Discount",
                "Final Total / Invoice Total",
                "Goods Description",
            )
        )
        self.assertEqual(
            rows[matrix_header_index][:12],
            ("FTY", "17-04-26-0914", "0901889028", "KX1885", "RC2610OW007", 305, 6212.0, 39.06, 100.0, "-", 6212.0, "WOMEN'S 100% POLYESTER WOVEN JACKET"),
        )
        self.assertEqual(
            rows[matrix_header_index + 1][:12],
            ("TC", "17-04-26-0914", "0901889028", "KX1885", "RC2610OW007", 305, 6817.25, 39.06, 100.0, "-", 6817.25, "WOMEN'S 100% POLYESTER WOVEN JACKET"),
        )
        self.assertEqual(
            rows[matrix_header_index + 2][:12],
            ("结果", "一致", "一致", "一致", "一致", "一致", "不一致；差额 605.25", "一致", "一致", "一致", "不一致；差额 605.25", "一致"),
        )
        for column_index in (1, 2, 3, 4, 5, 6, 8, 9, 10, 12):
            self.assertNotEqual(tc_sheet.cell(matrix_header_index + 1, column_index).fill.fgColor.rgb, issue_fill_rgb)
            self.assertNotEqual(tc_sheet.cell(matrix_header_index + 2, column_index).fill.fgColor.rgb, issue_fill_rgb)
            self.assertNotEqual(tc_sheet.cell(matrix_header_index + 3, column_index).fill.fgColor.rgb, issue_fill_rgb)
        for column_index in (7, 11):
            self.assertEqual(tc_sheet.cell(matrix_header_index + 1, column_index).fill.fgColor.rgb, issue_fill_rgb)
            self.assertEqual(tc_sheet.cell(matrix_header_index + 2, column_index).fill.fgColor.rgb, issue_fill_rgb)
            self.assertEqual(tc_sheet.cell(matrix_header_index + 3, column_index).fill.fgColor.rgb, issue_fill_rgb)
        self.assertFalse(
            any(row[:6] == ("来源", "PO No", "Article No", "Working/Style No", "Quantity", "Goods Description") for row in rows)
        )
        self.assertEqual(rows[matrix_header_index + 3][0], None)
        next_matrix_header = rows[matrix_header_index + 4][:12]
        self.assertEqual(
            next_matrix_header,
            (
                "来源",
                "Invoice Number",
                "PO No",
                "Article No",
                "Working/Style No",
                "Quantity",
                "Total Amount / Total PO Net Amount",
                "Freight/Additional Charge",
                "Documentation Charge",
                "Agreed Discount",
                "Final Total / Invoice Total",
                "Goods Description",
            ),
        )
        all_values = [value for row in rows for value in row if value is not None]
        self.assertEqual(sum(1 for value in next_matrix_header if value == "Quantity"), 1)
        self.assertNotIn("Unit Price", all_values)
        self.assertNotIn("Line Amount", all_values)
        self.assertEqual(save_result["tc_summary_count"], 1)
        self.assertEqual(save_result["tc_summary_issue_count"], 1)
        self.assertEqual(save_result["tc_total_issue_count"], 1)

    def test_save_excel_with_summary_marks_missing_tc_rows_red(self):
        ref_df = pd.DataFrame(
            [{"Price": 6.6, "Style NO.": "RC2620OW008", "Article NO.": "LG4321"}],
        )
        invoice_data = {("LG4321", "RC2620OW008"): {"invoice.xlsx": 6.6}}
        result_df, _ = self.module.update_reference_table(ref_df, invoice_data)
        tc_rows = self.module.build_tc_invoice_comparison(
            [
                InvoiceRecord(
                    invoice_file="invoice.xlsx",
                    invoice_number="10-06-26-0712",
                    invoice_date="2026-06-07",
                    po_number="0902694555",
                    article="LG4321",
                    style="RC2620OW008",
                    quantity=200,
                    price=6.6,
                    goods_description="WOMEN'S 100% POLYESTER WOVEN JACKET",
                )
            ],
            [],
        )

        with tempfile.TemporaryDirectory() as temp_dir:
            output_path = os.path.join(temp_dir, "result.xlsx")
            self.module.save_excel_with_summary(
                result_df,
                invoice_data,
                ["invoice.xlsx"],
                ["invoice.xlsx"],
                ref_df,
                output_path,
                tc_comparison_rows=tc_rows,
            )
            workbook = load_workbook(output_path)

        tc_sheet = workbook["TC INV核对"]
        rows = list(tc_sheet.iter_rows(values_only=True))
        matrix_header_index = next(
            index for index, row in enumerate(rows, 1)
            if row[:12] == (
                "来源",
                "Invoice Number",
                "PO No",
                "Article No",
                "Working/Style No",
                "Quantity",
                "Total Amount / Total PO Net Amount",
                "Freight/Additional Charge",
                "Documentation Charge",
                "Agreed Discount",
                "Final Total / Invoice Total",
                "Goods Description",
            )
        )
        self.assertEqual(rows[matrix_header_index][:12], ("FTY", "10-06-26-0712", "0902694555", "LG4321", "RC2620OW008", 200, "-", "-", "-", "-", "-", "WOMEN'S 100% POLYESTER WOVEN JACKET"))
        self.assertEqual(rows[matrix_header_index + 1][:12], ("TC", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-", "-"))
        self.assertEqual(rows[matrix_header_index + 2][:12], ("结果", "缺失", "缺失", "缺失", "缺失", "缺失", "一致", "一致", "一致", "一致", "一致", "缺失"))
        for column_index in (2, 3, 4, 5, 6, 12):
            self.assertEqual(tc_sheet.cell(matrix_header_index + 1, column_index).fill.fgColor.rgb, "FFFFDDDD")
            self.assertEqual(tc_sheet.cell(matrix_header_index + 2, column_index).fill.fgColor.rgb, "FFFFDDDD")
            self.assertEqual(tc_sheet.cell(matrix_header_index + 3, column_index).fill.fgColor.rgb, "FFFFDDDD")
        for column_index in (1, 7, 8, 9, 10, 11):
            self.assertNotEqual(tc_sheet.cell(matrix_header_index + 1, column_index).fill.fgColor.rgb, "FFFFDDDD")
            self.assertNotEqual(tc_sheet.cell(matrix_header_index + 2, column_index).fill.fgColor.rgb, "FFFFDDDD")
            self.assertNotEqual(tc_sheet.cell(matrix_header_index + 3, column_index).fill.fgColor.rgb, "FFFFDDDD")

    def test_tc_invoice_sheet_inserts_blue_separator_between_invoice_groups(self):
        wb = Workbook()
        ws = wb.active
        self.module._write_tc_invoice_comparison_sheet(
            ws,
            [],
            [
                TcSummaryComparisonRow(
                    status="一致",
                    issue_detail="",
                    invoice_number="INV-A",
                    fty_total_quantity=10,
                    tc_total_quantity=10,
                    fty_total_amount=None,
                    tc_total_po_net_amount=None,
                    fty_freight_charge=None,
                    tc_additional_charge=None,
                    fty_documentation_charge=None,
                    tc_documentation_charge=None,
                    fty_agreed_discount=None,
                    tc_agreed_discount=None,
                    fty_final_total_amount=None,
                    tc_invoice_total=None,
                    tc_total_carton=None,
                    tc_total_gross_weight=None,
                    tc_total_net_weight=None,
                    tc_total_net_net_weight=None,
                    tc_total_vat=None,
                    source_invoice="a.xls",
                    source_tc_pdf="a.pdf",
                ),
                TcSummaryComparisonRow(
                    status="一致",
                    issue_detail="",
                    invoice_number="INV-B",
                    fty_total_quantity=20,
                    tc_total_quantity=20,
                    fty_total_amount=None,
                    tc_total_po_net_amount=None,
                    fty_freight_charge=None,
                    tc_additional_charge=None,
                    fty_documentation_charge=None,
                    tc_documentation_charge=None,
                    fty_agreed_discount=None,
                    tc_agreed_discount=None,
                    fty_final_total_amount=None,
                    tc_invoice_total=None,
                    tc_total_carton=None,
                    tc_total_gross_weight=None,
                    tc_total_net_weight=None,
                    tc_total_net_net_weight=None,
                    tc_total_vat=None,
                    source_invoice="b.xls",
                    source_tc_pdf="b.pdf",
                ),
            ],
        )

        rows = list(ws.iter_rows(values_only=True))
        separator_row_index = next(
            row_index for row_index in range(1, ws.max_row + 1)
            if ws.cell(row_index, 1).fill.fgColor.rgb == "FFD9EAF7"
        )
        self.assertLess(separator_row_index, next(row_index for row_index, row in enumerate(rows, 1) if row[0] == "Invoice Number: INV-B"))
        for column_index in range(1, 9):
            self.assertEqual(ws.cell(separator_row_index, column_index).fill.fgColor.rgb, "FFD9EAF7")

    def test_process_invoices_labels_reference_style_when_tc_confirms_invoice(self):
        class VentReferenceJesscaModule(JesscaModule):
            def read_invoice_records(self, _invoice_path: str) -> List[InvoiceRecord]:
                return [
                    InvoiceRecord(
                        invoice_file="0490.xls",
                        invoice_number="10-04-26-0490",
                        invoice_date="2026-04-23",
                        po_number="0902107846",
                        article="KQ6747",
                        style="F2625W617",
                        quantity=357,
                        price=31.82,
                        goods_description="WOMEN'S WOVEN JACKET",
                    )
                ]

            def read_tc_invoice_records(self, _tc_path: str) -> List[TcInvoiceRecord]:
                return [
                    TcInvoiceRecord(
                        source_file="tc.pdf",
                        po_number="0902107846",
                        market_po="0306326640",
                        working_number="F2625W617",
                        article_number="KQ6747",
                        quantity=357,
                        goods_description="WOMEN'S WOVEN JACKET",
                    )
                ]

        with tempfile.TemporaryDirectory() as temp_dir:
            ref_path = os.path.join(temp_dir, "reference.xlsx")
            pd.DataFrame(
                [{"Price": 31.82, "Style NO.": "F2625W167", "Article NO.": "KQ6747"}]
            ).to_excel(ref_path, index=False)

            module = VentReferenceJesscaModule()
            result = module.process_invoices(
                [os.path.join(temp_dir, "0490.xls")],
                ref_path,
                temp_dir,
                tc_invoice_paths=[os.path.join(temp_dir, "tc.pdf")],
            )
            workbook = load_workbook(result["output_path"])

        self.assertTrue(result["success"])
        summary_records = self._sheet_records(workbook["汇总表"])
        summary_record = next(
            record for record in summary_records
            if record["values"].get("来源发票") == "0490.xls"
        )
        self.assertEqual(summary_record["values"]["状态"], "参考表款号疑似错误")
        self.assertEqual(summary_record["values"]["异常文件"], "参考表")
        self.assertEqual(summary_record["values"]["建议参考款号"], "F2625W167")
        self.assertEqual(summary_record["cells"]["状态"].fill.fgColor.rgb, "FFFFDDDD")

        main_records = self._sheet_records(workbook["核对结果"])
        main_record = next(
            record for record in main_records
            if record["values"].get("Article NO.") == "KQ6747"
        )
        self.assertEqual(main_record["values"]["核对状态"], "参考表款号疑似错误")

    def test_process_invoices_keeps_invoice_style_suspect_without_tc_confirmation(self):
        class VentReferenceJesscaModule(JesscaModule):
            def read_invoice_records(self, _invoice_path: str) -> List[InvoiceRecord]:
                return [
                    InvoiceRecord(
                        invoice_file="0490.xls",
                        invoice_number="10-04-26-0490",
                        invoice_date="2026-04-23",
                        po_number="0902107846",
                        article="KQ6747",
                        style="F2625W617",
                        quantity=357,
                        price=31.82,
                        goods_description="WOMEN'S WOVEN JACKET",
                    )
                ]

        with tempfile.TemporaryDirectory() as temp_dir:
            ref_path = os.path.join(temp_dir, "reference.xlsx")
            pd.DataFrame(
                [{"Price": 31.82, "Style NO.": "F2625W167", "Article NO.": "KQ6747"}]
            ).to_excel(ref_path, index=False)

            module = VentReferenceJesscaModule()
            result = module.process_invoices(
                [os.path.join(temp_dir, "0490.xls")],
                ref_path,
                temp_dir,
            )
            workbook = load_workbook(result["output_path"])

        self.assertTrue(result["success"])
        summary_records = self._sheet_records(workbook["汇总表"])
        summary_record = next(
            record for record in summary_records
            if record["values"].get("来源发票") == "0490.xls"
        )
        self.assertEqual(summary_record["values"]["状态"], "发票款号疑似错误")
        self.assertNotEqual(summary_record["values"]["异常文件"], "参考表")

    def test_process_invoices_merges_multiple_tc_invoice_pdf_records(self):
        class MultiTcJesscaModule(JesscaModule):
            def __init__(self):
                super().__init__()
                self.read_tc_paths: List[str] = []

            def read_invoice_records(self, _invoice_path: str) -> List[InvoiceRecord]:
                return [
                    InvoiceRecord(
                        invoice_file="invoice.xlsx",
                        invoice_number="10-06-26-0712",
                        invoice_date="2026-06-07",
                        po_number="0902694555",
                        article="LG4321",
                        style="RC2620OW008",
                        quantity=200,
                        price=6.6,
                        goods_description="WOMEN'S 100% POLYESTER WOVEN JACKET",
                    ),
                    InvoiceRecord(
                        invoice_file="invoice.xlsx",
                        invoice_number="10-06-26-0712",
                        invoice_date="2026-06-07",
                        po_number="0902694557",
                        article="LG4323",
                        style="RC2620OW008",
                        quantity=180,
                        price=7.1,
                        goods_description="WOMEN'S KNITTED SHORT",
                    ),
                ]

            def read_tc_invoice_records(self, tc_path: str) -> List[TcInvoiceRecord]:
                self.read_tc_paths.append(tc_path)
                if tc_path.endswith("tc-a.pdf"):
                    return [
                        TcInvoiceRecord(
                            source_file="tc-a.pdf",
                            po_number="0902694555",
                            market_po="0307073961",
                            working_number="RC2620OW008",
                            article_number="LG4321",
                            quantity=200,
                            goods_description="WOMEN'S 100% POLYESTER WOVEN JACKET",
                        )
                    ]

                return [
                    TcInvoiceRecord(
                        source_file="tc-b.pdf",
                        po_number="0902694557",
                        market_po="0307073963",
                        working_number="RC2620OW008",
                        article_number="LG4323",
                        quantity=180,
                        goods_description="WOMEN'S KNITTED SHORT",
                    )
                ]

        with tempfile.TemporaryDirectory() as temp_dir:
            ref_path = os.path.join(temp_dir, "reference.xlsx")
            pd.DataFrame(
                [
                    {"Price": 6.6, "Style NO.": "RC2620OW008", "Article NO.": "LG4321"},
                    {"Price": 7.1, "Style NO.": "RC2620OW008", "Article NO.": "LG4323"},
                ]
            ).to_excel(ref_path, index=False)

            module = MultiTcJesscaModule()
            result = module.process_invoices(
                [os.path.join(temp_dir, "invoice.xlsx")],
                ref_path,
                temp_dir,
                tc_invoice_paths=[
                    os.path.join(temp_dir, "tc-a.pdf"),
                    os.path.join(temp_dir, "tc-b.pdf"),
                ],
            )

        self.assertTrue(result["success"])
        self.assertEqual(len(module.read_tc_paths), 2)
        self.assertEqual(result["tc_count"], 2)
        self.assertEqual(result["tc_matched_count"], 2)
        self.assertEqual(result["tc_issue_count"], 0)


if __name__ == "__main__":
    unittest.main()

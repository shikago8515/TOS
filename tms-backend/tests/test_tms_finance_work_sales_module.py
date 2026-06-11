import os
import sys
import tempfile
import unittest
from datetime import datetime

import openpyxl
from fastapi import FastAPI
from fastapi.testclient import TestClient


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.tms_finance_work_sales_module import (  # noqa: E402
    TmsFinanceWorkSalesModule,
)


BULK_SALES_HEADERS = [
    "R",
    "MONTH",
    "YEAR",
    "SALES INVOICE NUMBER",
    "SALES INVOICE DATE",
    "*SALES INVOICE AMT (GROSS)",
    "*SALES INVOICE AMT (NET)",
    "TAX (%)",
    "TAX AMT",
    "*BUYER NAME",
    "BILL TO",
    "FACTORY  PAYMENT TERM",
    "BUYER PAYMENT TERM",
    "ACTUAL BUYER PAYMENT TERM",
    "STYLE NUMBER",
    "CUSTOMER SEASON",
    "ORDER NUMBER",
    "COLOR CODE",
    "FACTORY  CURRENCY",
    "FACTORY  UNIT PX",
    "BUYER  CURRENCY",
    "BUYER  UNIT PX",
    "ESTIMATED LANDED COST UNIT PRICE (USD)",
    "ESTIMATED LANDED COST AMOUNT (USD)",
    "ACT SAILING DATE",
    "PROJECTED SAILING DATE",
    "REAL FACTORY  PRICE / USD",
    "ROE AGREED",
    "*ORDER QTY",
    "SHIP QTY",
    "L/C NUMBER",
    "INVOICE STATUS",
    "NAME OF FACTORY",
    "LCs BENEFICIARY",
    "PUR INVOICE AMOUNT",
    "SHIP MODE",
    "ACTUAL SHIP MODE",
    "HANDOVER DATE",
]

SALES_HEADERS = [
    "Style Number",
    "Unit Price(include VAT)",
    "Unit Price(exclude VAT)",
    "Ship Quantity",
    "Total price excluding VAT and VAS",
    "VAS",
    "Promo price upcharge",
    "Sales Amount after deduction",
    "VAT Amount",
    "Gross Amount(include the VAT)",
    "AR Amount",
    "Total amount in Iplex system (include VAT)",
    "Total amount in Iplex system (exclude VAT)",
    "difference",
    "",
    "MERCH",
    "HANDOVER DATE",
    "SALES INVOICE NUMBER",
]

PURCHASE_HEADERS = [
    "Style Number",
    "Unit Price(include VAT)",
    "Unit Price(exclude VAT)",
    "Ship Quantity",
    "Purchase Amount",
    "NET RE-ROUTE SURCHARGE",
    "VAT Amount",
    "RE-ROUTE SURCHARGE",
    "Gross Amount",
    "AP Amount",
    "Total amount in Iplex system (include VAT)",
    "Total amount in Iplex system (exclude VAT)",
    "difference",
    "",
    "",
    "MERCH",
    "HANDOVER DATE",
    "SALES INVOICE NUMBER",
]


def _bulk_rows() -> list[dict[str, object]]:
    return [
        {
            "invoice": "14-05-26-0063",
            "invoice_date": datetime(2026, 5, 7),
            "style": "RC2610OW001.",
            "order": "0901888496",
            "buyer_unit_price": 125.17,
            "factory_unit_price": 113.82,
            "quantity": 2977,
            "sales_amount_net": 422682.26,
            "purchase_amount": 338842.14,
            "handover_date": datetime(2026, 5, 7),
        },
        {
            "invoice": "14-05-26-0062",
            "invoice_date": datetime(2026, 5, 7),
            "style": "RC2610OW000.",
            "order": "0901888497",
            "buyer_unit_price": 127.3,
            "factory_unit_price": 115.6,
            "quantity": 1804,
            "sales_amount_net": 260496.63,
            "purchase_amount": 208542.4,
            "handover_date": datetime(2026, 5, 7),
        },
    ]


def _old_target_row() -> dict[str, object]:
    return {
        "invoice": "14-04-26-0001",
        "style": "OLDSTYLE.",
        "buyer_unit_price": 99.9,
        "factory_unit_price": 88.8,
        "quantity": 10,
        "sales_amount_net": 1128.87,
        "purchase_amount": 888.0,
        "handover_date": datetime(2026, 4, 30),
    }


class TmsFinanceWorkSalesModuleTests(unittest.TestCase):
    def setUp(self) -> None:
        self.module = TmsFinanceWorkSalesModule()

    def _create_bulk_sales_workbook(self, path: str, rows: list[dict[str, object]]) -> None:
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Sheet1"
        ws.append(BULK_SALES_HEADERS)
        for row in rows:
            values = {header: "" for header in BULK_SALES_HEADERS}
            values.update(
                {
                    "MONTH": "5",
                    "YEAR": "2026",
                    "SALES INVOICE NUMBER": row["invoice"],
                    "SALES INVOICE DATE": row.get("invoice_date"),
                    "*SALES INVOICE AMT (NET)": row["sales_amount_net"],
                    "STYLE NUMBER": row["style"],
                    "ORDER NUMBER": row["order"],
                    "FACTORY  CURRENCY": "RMB",
                    "FACTORY  UNIT PX": row["factory_unit_price"],
                    "BUYER  CURRENCY": "RMB",
                    "BUYER  UNIT PX": row["buyer_unit_price"],
                    "SHIP QTY": row["quantity"],
                    "PUR INVOICE AMOUNT": row["purchase_amount"],
                    "HANDOVER DATE": row["handover_date"],
                }
            )
            ws.append([values[header] for header in BULK_SALES_HEADERS])
        workbook.save(path)

    def _create_turnover_workbook(self, path: str, rows: list[dict[str, object]]) -> None:
        workbook = openpyxl.Workbook()
        ws = workbook.active
        ws.title = "Turnover Details"
        ws["A1"] = "Turnover Details MAY 2026"
        ws["E2"] = "SALES"
        ws["F2"] = "OTHER INCOME"
        ws["I2"] = "TAX"
        ws["J2"] = "AR"
        for column, header in enumerate(SALES_HEADERS, start=1):
            ws.cell(3, column).value = header
        for row_index, row in enumerate(rows, start=4):
            self._write_sales_template_row(ws, row_index, row)
        sales_total_row = 4 + len(rows)
        self._write_total_row(ws, sales_total_row, 4, sales_total_row - 1)
        ws.cell(sales_total_row + 1, 9).value = f"=I{sales_total_row}"

        purchase_title_row = sales_total_row + 4
        purchase_header_row = sales_total_row + 6
        ws.cell(purchase_title_row, 1).value = "Purchase Details MAY 2026"
        ws.cell(purchase_title_row + 1, 5).value = "Purchase"
        for column, header in enumerate(PURCHASE_HEADERS, start=1):
            ws.cell(purchase_header_row, column).value = header
        for row_index, row in enumerate(rows, start=purchase_header_row + 1):
            self._write_purchase_template_row(ws, row_index, row)
        purchase_total_row = purchase_header_row + 1 + len(rows)
        self._write_total_row(ws, purchase_total_row, purchase_header_row + 1, purchase_total_row - 1)
        workbook.save(path)

    def _write_sales_template_row(self, ws, row_index: int, row: dict[str, object]) -> None:
        ws.cell(row_index, 1).value = row["style"]
        ws.cell(row_index, 3).value = row["buyer_unit_price"]
        ws.cell(row_index, 4).value = row["quantity"]
        ws.cell(row_index, 5).value = f"=ROUND(C{row_index}*D{row_index},2)"
        ws.cell(row_index, 6).value = f"=ROUND(0.483581*D{row_index},2)"
        ws.cell(row_index, 8).value = f"=ROUND(E{row_index}+F{row_index}+G{row_index},2)"
        ws.cell(row_index, 9).value = f"=ROUND(H{row_index}*0.13,2)"
        ws.cell(row_index, 10).value = f"=H{row_index}+I{row_index}"
        ws.cell(row_index, 12).value = row["sales_amount_net"]
        ws.cell(row_index, 14).value = f"=J{row_index}-L{row_index}"
        ws.cell(row_index, 15).value = "Issued VAT inv."
        ws.cell(row_index, 16).value = "Caroline"
        ws.cell(row_index, 17).value = row["handover_date"]
        ws.cell(row_index, 18).value = row["invoice"]

    def _write_purchase_template_row(self, ws, row_index: int, row: dict[str, object]) -> None:
        ws.cell(row_index, 1).value = row["style"]
        ws.cell(row_index, 3).value = row["factory_unit_price"]
        ws.cell(row_index, 4).value = row["quantity"]
        ws.cell(row_index, 5).value = f"=ROUND(C{row_index}*D{row_index},2)"
        ws.cell(row_index, 7).value = f"=ROUND((E{row_index})*0.13,2)"
        ws.cell(row_index, 9).value = f"=E{row_index}+G{row_index}+F{row_index}"
        ws.cell(row_index, 12).value = row["purchase_amount"]
        ws.cell(row_index, 13).value = f"=E{row_index}-L{row_index}"
        ws.cell(row_index, 14).value = "received VAT inv."
        ws.cell(row_index, 16).value = "Caroline"
        ws.cell(row_index, 17).value = row["handover_date"]
        ws.cell(row_index, 18).value = row["invoice"]

    def _write_total_row(self, ws, total_row: int, start_row: int, end_row: int) -> None:
        for column in range(4, 15):
            letter = openpyxl.utils.get_column_letter(column)
            ws.cell(total_row, column).value = f"=SUM({letter}{start_row}:{letter}{end_row})"

    def test_appends_bulk_sales_rows_to_turnover_details_sections(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            bulk_path = os.path.join(tmpdir, "bulk sales.xlsx")
            turnover_path = os.path.join(tmpdir, "turnover.xlsx")
            self._create_bulk_sales_workbook(bulk_path, _bulk_rows())
            self._create_turnover_workbook(turnover_path, [_old_target_row()])

            result = self.module.process_files(
                bulk_sales_path=bulk_path,
                turnover_path=turnover_path,
                output_dir=tmpdir,
            )

            self.assertTrue(result["success"])
            self.assertEqual(result["source_row_count"], 2)
            self.assertEqual(result["sales_appended_count"], 2)
            self.assertEqual(result["purchase_appended_count"], 2)
            self.assertEqual(result["duplicate_count"], 0)
            output_wb = openpyxl.load_workbook(result["output_path"], data_only=False)
            try:
                ws = output_wb["Turnover Details"]
                self.assertNotIn("Work Sales Summary", output_wb.sheetnames)
                self.assertEqual(ws.cell(5, 1).value, "RC2610OW001.")
                self.assertEqual(ws.cell(5, 3).value, 125.17)
                self.assertEqual(ws.cell(5, 4).value, 2977)
                self.assertEqual(ws.cell(5, 12).value, 422682.26)
                self.assertEqual(ws.cell(5, 15).value, "Issued VAT inv.")
                self.assertEqual(ws.cell(5, 16).value, "Caroline")
                self.assertEqual(ws.cell(5, 18).value, "14-05-26-0063")
                self.assertEqual(ws.cell(5, 5).value, "=ROUND(C5*D5,2)")
                self.assertEqual(ws.cell(7, 4).value, "=SUM(D4:D6)")
                self.assertEqual(ws.cell(15, 1).value, "RC2610OW001.")
                self.assertEqual(ws.cell(15, 3).value, 113.82)
                self.assertEqual(ws.cell(15, 4).value, 2977)
                self.assertEqual(ws.cell(15, 12).value, 338842.14)
                self.assertEqual(ws.cell(15, 14).value, "received VAT inv.")
                self.assertEqual(ws.cell(15, 16).value, "Caroline")
                self.assertEqual(ws.cell(15, 18).value, "14-05-26-0063")
                self.assertEqual(ws.cell(15, 5).value, "=ROUND(C15*D15,2)")
                self.assertEqual(ws.cell(17, 4).value, "=SUM(D14:D16)")
            finally:
                output_wb.close()

    def test_skips_rows_already_present_in_sales_and_purchase_sections(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            bulk_path = os.path.join(tmpdir, "bulk sales.xlsx")
            turnover_path = os.path.join(tmpdir, "turnover.xlsx")
            self._create_bulk_sales_workbook(bulk_path, _bulk_rows())
            self._create_turnover_workbook(turnover_path, _bulk_rows())

            result = self.module.process_files(
                bulk_sales_path=bulk_path,
                turnover_path=turnover_path,
                output_dir=tmpdir,
            )

            self.assertEqual(result["source_row_count"], 2)
            self.assertEqual(result["sales_appended_count"], 0)
            self.assertEqual(result["purchase_appended_count"], 0)
            self.assertEqual(result["duplicate_count"], 2)
            output_wb = openpyxl.load_workbook(result["output_path"], data_only=False)
            try:
                ws = output_wb["Turnover Details"]
                self.assertEqual(ws.cell(6, 4).value, "=SUM(D4:D5)")
                self.assertEqual(ws.cell(15, 4).value, "=SUM(D13:D14)")
            finally:
                output_wb.close()

    def test_rejects_turnover_workbook_without_turnover_details_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as tmpdir:
            bulk_path = os.path.join(tmpdir, "bulk sales.xlsx")
            turnover_path = os.path.join(tmpdir, "turnover.xlsx")
            self._create_bulk_sales_workbook(bulk_path, _bulk_rows())
            workbook = openpyxl.Workbook()
            workbook.active.title = "Other"
            workbook.save(turnover_path)

            with self.assertRaisesRegex(ValueError, "Turnover Details"):
                self.module.process_files(
                    bulk_sales_path=bulk_path,
                    turnover_path=turnover_path,
                    output_dir=tmpdir,
                )

    def test_api_processes_bulk_sales_and_turnover_uploads(self) -> None:
        from api.tms_finance_work_sales_api import router

        app = FastAPI()
        app.include_router(router, prefix="/api")
        client = TestClient(app)

        with tempfile.TemporaryDirectory() as tmpdir:
            bulk_path = os.path.join(tmpdir, "bulk sales.xlsx")
            turnover_path = os.path.join(tmpdir, "turnover.xlsx")
            self._create_bulk_sales_workbook(bulk_path, _bulk_rows())
            self._create_turnover_workbook(turnover_path, [_old_target_row()])

            with open(bulk_path, "rb") as bulk_file, open(turnover_path, "rb") as turnover_file:
                response = client.post(
                    "/api/tms-finance/work-sales/process",
                    files={
                        "bulk_sales_file": (
                            "bulk sales.xlsx",
                            bulk_file,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        ),
                        "turnover_file": (
                            "turnover.xlsx",
                            turnover_file,
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        ),
                    },
                )

            self.assertEqual(response.status_code, 200)
            payload = response.json()
            self.assertEqual(payload["source_row_count"], 2)
            self.assertEqual(payload["sales_appended_count"], 2)
            self.assertEqual(payload["purchase_appended_count"], 2)
            self.assertIn("output_file", payload)

            download_response = client.get(
                f"/api/tms-finance/work-sales/download/{payload['output_file']}",
            )
            self.assertEqual(download_response.status_code, 200)

    def test_api_rejects_invalid_bulk_sales_extension(self) -> None:
        from api.tms_finance_work_sales_api import router

        app = FastAPI()
        app.include_router(router, prefix="/api")
        client = TestClient(app)

        response = client.post(
            "/api/tms-finance/work-sales/process",
            files={
                "bulk_sales_file": ("bulk.txt", b"bad", "text/plain"),
                "turnover_file": ("turnover.xlsx", b"bad", "application/octet-stream"),
            },
        )

        self.assertEqual(response.status_code, 400)


if __name__ == "__main__":
    unittest.main()

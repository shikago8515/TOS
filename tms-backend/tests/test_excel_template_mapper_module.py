import os
import sys
import tempfile
import unittest
from pathlib import Path

import openpyxl


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.excel_template_mapper_module import (  # noqa: E402
    ExcelTemplateFieldMapping,
    ExcelTemplateMapperConfig,
    ExcelTemplateMapperModule,
)


class ExcelTemplateMapperModuleTests(unittest.TestCase):
    def test_inspect_reads_xlsx_headers_and_sample_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "source.xlsx"
            self._write_source_workbook(source_path)

            module = ExcelTemplateMapperModule()
            inspection = module.inspect_workbook(source_path, sheet_name="Source", header_row=1)

        self.assertEqual(inspection["sheets"], [
            {"name": "Source", "max_row": 3, "max_column": 3},
        ])
        self.assertEqual(inspection["selected_sheet"]["name"], "Source")
        self.assertEqual(inspection["selected_sheet"]["data_row_count"], 2)
        self.assertEqual(
            inspection["selected_sheet"]["headers"],
            [
                {"index": 1, "letter": "A", "label": "PO Number", "sample_value": "PO-001"},
                {"index": 2, "letter": "B", "label": "Style Name", "sample_value": "Jacket"},
                {"index": 3, "letter": "C", "label": "Qty", "sample_value": "12"},
            ],
        )
        self.assertEqual(
            inspection["selected_sheet"]["sample_rows"],
            [
                {"row_number": 2, "values": ["PO-001", "Jacket", "12"]},
                {"row_number": 3, "values": ["PO-002", "Pants", "8"]},
            ],
        )

    def test_process_writes_mapped_source_values_into_template(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "source.xlsx"
            template_path = Path(temp_dir) / "template.xlsx"
            self._write_source_workbook(source_path)
            self._write_template_workbook(template_path)

            module = ExcelTemplateMapperModule()
            result = module.process_files(
                source_path=source_path,
                template_path=template_path,
                config=ExcelTemplateMapperConfig(
                    source_sheet_name="Source",
                    template_sheet_name="Template",
                    source_header_row=1,
                    source_data_start_row=2,
                    template_header_row=1,
                    template_data_start_row=2,
                    mappings=[
                        ExcelTemplateFieldMapping(target_column=1, source_column=1, target_header="Order"),
                        ExcelTemplateFieldMapping(target_column=2, source_column=2, target_header="Style"),
                        ExcelTemplateFieldMapping(target_column=3, source_column=3, target_header="Quantity"),
                    ],
                ),
                output_dir=temp_dir,
            )

            workbook = openpyxl.load_workbook(result["output_path"], data_only=False)
            sheet = workbook["Template"]

        self.assertTrue(result["output_path"].endswith(".xlsx"))
        self.assertEqual(result["source_row_count"], 2)
        self.assertEqual(result["written_row_count"], 2)
        self.assertEqual(result["mapped_field_count"], 3)
        self.assertEqual(result["unmapped_required_fields"], [])
        self.assertEqual(sheet.cell(row=1, column=1).value, "Order")
        self.assertEqual(sheet.cell(row=2, column=1).value, "PO-001")
        self.assertEqual(sheet.cell(row=2, column=2).value, "Jacket")
        self.assertEqual(sheet.cell(row=2, column=3).value, 12)
        self.assertEqual(sheet.cell(row=3, column=1).value, "PO-002")
        self.assertEqual(sheet.cell(row=3, column=2).value, "Pants")
        self.assertEqual(sheet.cell(row=3, column=3).value, 8)

    def test_process_rejects_unmapped_required_template_field(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "source.xlsx"
            template_path = Path(temp_dir) / "template.xlsx"
            self._write_source_workbook(source_path)
            self._write_template_workbook(template_path)

            module = ExcelTemplateMapperModule()
            with self.assertRaisesRegex(ValueError, "Style"):
                module.process_files(
                    source_path=source_path,
                    template_path=template_path,
                    config=ExcelTemplateMapperConfig(
                        source_sheet_name="Source",
                        template_sheet_name="Template",
                        source_header_row=1,
                        source_data_start_row=2,
                        template_header_row=1,
                        template_data_start_row=2,
                        mappings=[
                            ExcelTemplateFieldMapping(target_column=1, source_column=1, target_header="Order"),
                            ExcelTemplateFieldMapping(
                                target_column=2,
                                source_column=0,
                                target_header="Style",
                                required=True,
                            ),
                        ],
                    ),
                    output_dir=temp_dir,
                )

    def test_inspect_rejects_missing_sheet(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            source_path = Path(temp_dir) / "source.xlsx"
            self._write_source_workbook(source_path)

            module = ExcelTemplateMapperModule()
            with self.assertRaisesRegex(ValueError, "Missing"):
                module.inspect_workbook(source_path, sheet_name="Missing", header_row=1)

    def _write_source_workbook(self, path: Path) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Source"
        sheet.append(["PO Number", "Style Name", "Qty"])
        sheet.append(["PO-001", "Jacket", 12])
        sheet.append(["PO-002", "Pants", 8])
        workbook.save(path)

    def _write_template_workbook(self, path: Path) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Template"
        sheet.append(["Order", "Style", "Quantity", "Comment"])
        sheet.append(["", "", "", "keep"])
        workbook.save(path)


if __name__ == "__main__":
    unittest.main()

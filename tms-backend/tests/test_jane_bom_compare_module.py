import os
import sys
import tempfile
import unittest

import openpyxl
from openpyxl.styles import PatternFill


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.jane_bom_compare_module import JaneBomCompareModule


class JaneBomCompareModuleTests(unittest.TestCase):
    def setUp(self):
        self.module = JaneBomCompareModule()

    def _fill_rgb(self, cell):
        color = cell.fill.fgColor
        if color.type == "rgb":
            return (color.rgb or "").upper()
        return ""

    def _column_by_header(self, ws, header):
        for cell in ws[1]:
            if cell.value == header:
                return cell.column
        self.fail(f"未找到表头：{header}")

    def _rows_by_value(self, ws, column, value):
        return [
            row_index
            for row_index in range(2, ws.max_row + 1)
            if ws.cell(row=row_index, column=column).value == value
        ]

    def _append_production_row(
        self,
        ws,
        *,
        style,
        lot,
        quantity,
        units,
        production_date,
        factory,
        material,
        used_units,
    ):
        ws.append([
            style,
            lot,
            quantity,
            units,
            production_date,
            factory,
            material,
            f"LOT-{material}",
            "INV-001",
            0,
            used_units,
            "Yards",
            "OLD-SUPPLIER",
            "",
            "",
            "",
            "",
            2026,
            "Spring/Summer",
            "Submit",
            "RC2610OW005H2H",
            "Diana",
            "SS26 ASOS Pack",
        ])

    def _save_production_workbook(self, folder):
        path = os.path.join(folder, "T1 PRODUCTION.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "PRODUCTION"

        headers = [
            "Style ID",
            "Production Lot ID",
            "Production Quantity (in Units)",
            "Units",
            "Production Date",
            "Recording Facility ID",
            "Input Material UID/ID",
            "Input Material Lot ID/ TC Lot No",
            "TC Ref. Number/ Invoice Number",
            "Input Lot Quantity Used (In Kgs)",
            "Input Lot Quantity Used (In Units)",
            "Input Units",
            "Seller Facility ID",
            "Input Material Production Date",
            "Input Material Type",
            "Input Material Name",
            "Material Composition(%)",
            "Year",
            "Season",
            "Status",
            "WORKING",
            "MER",
            "PACK",
        ]
        ws.append(headers)

        orange_fill = PatternFill("solid", fgColor="FFFFD966")
        green_fill = PatternFill("solid", fgColor="FFD9EAD3")
        for index, cell in enumerate(ws[1], start=1):
            cell.fill = green_fill if 7 <= index <= 12 else orange_fill

        self._append_production_row(
            ws,
            style="LD1803",
            lot="OHO055-M-LD1803",
            quantity=100,
            units="Each",
            production_date="2026-02-01",
            factory="3LP001",
            material="70020171",
            used_units=60,
        )
        self._append_production_row(
            ws,
            style="LD1803",
            lot="OHO055-M-LD1803",
            quantity=100,
            units="Each",
            production_date="2026-02-01",
            factory="3LP001",
            material="70020171",
            used_units=40,
        )
        self._append_production_row(
            ws,
            style="LD1804",
            lot="OHO055-M-LD1804",
            quantity=200,
            units="Each",
            production_date="2026-01-01",
            factory="3LP001",
            material="70020171",
            used_units=30,
        )
        self._append_production_row(
            ws,
            style="LD1804",
            lot="OHO055-M-LD1804",
            quantity=200,
            units="Each",
            production_date="2026-02-01",
            factory="3LP001",
            material="62712089",
            used_units=20,
        )
        self._append_production_row(
            ws,
            style="LD1807",
            lot="OHO056-M-LD1807",
            quantity=120,
            units="Each",
            production_date="2026-03-01",
            factory="3LP001",
            material="70020171",
            used_units=70,
        )
        self._append_production_row(
            ws,
            style="LD1807",
            lot="OHO056-M-LD1807",
            quantity=120,
            units="Each",
            production_date="2026-03-01",
            factory="3LP001",
            material="62712089",
            used_units=30,
        )
        self._append_production_row(
            ws,
            style="LD1807",
            lot="OHO056-M-LD1807",
            quantity=120,
            units="Each",
            production_date="2026-03-01",
            factory="3LP001",
            material="70020122",
            used_units=10,
        )

        for row_index in range(2, ws.max_row + 1):
            for column in range(1, 24):
                ws.cell(row=row_index, column=column).fill = (
                    green_fill if 7 <= column <= 12 else orange_fill
                )
        wb.save(path)
        return path

    def _save_bom_summary_workbook(self, folder):
        path = os.path.join(folder, "BOM汇总.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "Sheet1"
        ws.append([
            "Pack",
            "Working #",
            "Season",
            "Factory",
            "Articles",
            "Part Group #",
            "Material Reference #",
            "Material Name",
            "Group Code Supplier",
            "Supplier Name",
            "Material Description",
            "Color",
        ])
        for article in ["LD1803", "LD1804", "LD1807"]:
            ws.append([
                "SS26 ASOS Pack",
                "RC2610OW005H2H",
                "SS26",
                "3LP001 | XO TEX INDUSTRIAL CO., LTD. | A",
                article,
                "10",
                "70020171",
                "70020171 Satin",
                "1SB001",
                "TMS (CHN)",
                "97%POLYESTER",
                "117A NIGHT INDIGO",
            ])
            ws.append([
                "SS26 ASOS Pack",
                "RC2610OW005H2H",
                "SS26",
                "3LP001 | XO TEX INDUSTRIAL CO., LTD. | A",
                article,
                "20",
                "62712089",
                "62712089 Tricot",
                "299002",
                "HUAFENG (CHINA)",
                "100%POLYESTER",
                "117A NIGHT INDIGO",
            ])
        wb.save(path)
        return path

    def _save_raw_bom_workbook(self, folder):
        path = os.path.join(folder, "BOM-LD1803-3LP001.xlsx")
        wb = openpyxl.Workbook()
        ws = wb.active
        ws.title = "LD1803-3LP001"
        ws["B1"] = "RC2610OW005H2H"
        ws["B3"] = "SS26"
        ws["B5"] = "3LP001 | XO TEX INDUSTRIAL CO., LTD. | A"
        ws["P9"] = "DEFAULT GROUP - NIGHT INDIGO | LD1803 | IN RANGE"
        headers = [
            "Variation",
            "Part Name",
            "Part ID",
            "Part Group #",
            "Material Reference #",
            "Material Name",
            "Group Code Supplier",
            "Supplier Name",
            "Supplier Material ID",
            "Supplier Material Name",
            "Material Description",
            "Material Supplier Lifecycle",
            "Material Lifecycle Validate",
            "Part Remark",
            "Color Layer Name",
            "Color",
            "Color Treatments",
        ]
        for column_index, value in enumerate(headers, start=1):
            ws.cell(row=10, column=column_index, value=value)
        ws["B11"] = "MAIN COMPONENT"
        ws.append(["", "T main body", "", "10", "70020171", "", "1SB001", "", "", "", "", "", "", "", "", "117A NIGHT INDIGO", ""])
        ws.append(["", "T pocket bag", "", "20", "62712089", "", "299002", "", "", "", "", "", "", "", "", "117A NIGHT INDIGO", ""])
        wb.save(path)
        return path

    def test_production_compare_marks_inconsistent_extra_missing_and_rate(self):
        with tempfile.TemporaryDirectory() as folder:
            production_path = self._save_production_workbook(folder)
            bom_summary_path = self._save_bom_summary_workbook(folder)

            result = self.module.process_reports(production_path, bom_summary_path, folder)

            self.assertTrue(result["success"])
            self.assertEqual(result["bom_count"], 1)
            self.assertEqual(result["bom_material_row_count"], 6)
            self.assertEqual(result["checked_row_count"], 7)
            self.assertEqual(result["inconsistent_group_count"], 1)
            self.assertEqual(result["extra_material_row_count"], 1)
            self.assertEqual(result["missing_row_count"], 1)
            self.assertEqual(result["rate_row_count"], 7)
            self.assertEqual(result["no_bom_key_count"], 0)

            output_wb = openpyxl.load_workbook(result["output_path"], data_only=False)
            ws = output_wb["PRODUCTION"]

            rate_col = self._column_by_header(ws, "料率")
            used_units_col = self._column_by_header(ws, "Input Lot Quantity Used (In Units)")
            input_units_col = self._column_by_header(ws, "Input Units")
            detail_col = self._column_by_header(ws, "Check Detail")
            result_col = self._column_by_header(ws, "Check Result")

            self.assertEqual(rate_col, used_units_col + 1)
            self.assertEqual(input_units_col, rate_col + 1)
            self.assertEqual(ws.cell(row=2, column=rate_col).value, '=IFERROR(SUMIFS($K:$K,$A:$A,A2,$B:$B,B2,$G:$G,G2)/C2,"")')
            self.assertEqual(ws.cell(row=2, column=rate_col).number_format, "0.0000")

            self.assertEqual(ws.cell(row=4, column=1).value, "LD1803")
            self.assertEqual(ws.cell(row=4, column=2).value, "OHO055-M-LD1803")
            self.assertEqual(ws.cell(row=4, column=3).value, 100)
            self.assertEqual(ws.cell(row=4, column=4).value, "Each")
            self.assertEqual(ws.cell(row=4, column=5).value, "2026-02-01")
            self.assertEqual(ws.cell(row=4, column=6).value, "3LP001")
            self.assertEqual(ws.cell(row=4, column=7).value, "62712089")
            self.assertIsNone(ws.cell(row=4, column=used_units_col).value)
            self.assertIsNone(ws.cell(row=4, column=rate_col).value)
            self.assertEqual(ws.cell(row=4, column=result_col).value, "需补入")
            self.assertIn("Production 缺少材料", ws.cell(row=4, column=detail_col).value)
            self.assertIn("62712089", ws.cell(row=4, column=detail_col).value)

            for row_index in self._rows_by_value(ws, 1, "LD1804"):
                self.assertEqual(self._fill_rgb(ws.cell(row=row_index, column=5)), "FFFFC7CE")
                self.assertIn("Production Date", ws.cell(row=row_index, column=detail_col).value)

            extra_row = self._rows_by_value(ws, 7, "70020122")[0]
            self.assertEqual(ws.cell(row=extra_row, column=result_col).value, "需删除")
            self.assertIn("Production 多出材料", ws.cell(row=extra_row, column=detail_col).value)
            for column in range(1, ws.max_column + 1):
                self.assertTrue(ws.cell(row=extra_row, column=column).font.strike)

    def test_raw_bom_input_keeps_route_contract_without_supplier_compare(self):
        with tempfile.TemporaryDirectory() as folder:
            production_path = self._save_production_workbook(folder)
            bom_path = self._save_raw_bom_workbook(folder)

            result = self.module.process_reports(production_path, [bom_path], folder)

            self.assertTrue(result["success"])
            output_wb = openpyxl.load_workbook(result["output_path"], data_only=False)
            ws = output_wb["PRODUCTION"]

            self._column_by_header(ws, "料率")
            with self.assertRaises(AssertionError):
                self._column_by_header(ws, "Correct Group Code Supplier (BOM)")

            seller_col = self._column_by_header(ws, "Seller Facility ID")
            self.assertNotEqual(self._fill_rgb(ws.cell(row=2, column=seller_col)), "FFFFC7CE")


if __name__ == "__main__":
    unittest.main()

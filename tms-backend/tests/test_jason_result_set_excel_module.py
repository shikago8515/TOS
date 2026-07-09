from __future__ import annotations

import os
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from zipfile import ZipFile

import openpyxl


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.jason_result_set_excel_module import JasonResultSetExcelModule  # noqa: E402


EXPECTED_HEADER_MERGES = {
    "A1:A2",
    "B1:B2",
    "C1:C2",
    "D1:D2",
    "E1:E2",
    "F1:F2",
    "G1:G2",
    "H1:H2",
    "I1:I2",
    "J1:J2",
    "K1:K2",
    "L1:N1",
    "O1:R1",
    "S1:T1",
    "U1:U2",
    "V1:V2",
    "W1:W2",
}


class JasonResultSetExcelModuleTests(unittest.TestCase):
    def test_process_generates_target_rows_formulas_and_totals_without_external_links(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_path = temp_path / "result-set.xlsx"
            template_path = temp_path / "template.xlsx"
            self._write_result_set_workbook(source_path, not_shipped_count=59)
            self._write_template_workbook(template_path)

            module = JasonResultSetExcelModule(template_path=template_path)
            result = module.process_result_set(
                result_set_path=source_path,
                target_month="2026-07",
                output_dir=temp_path,
            )

            workbook = openpyxl.load_workbook(result["output_path"], data_only=False, keep_links=False)
            self.assertEqual(workbook.sheetnames, ["目标表", "Sheet1", "Sheet2", "Sheet3"])
            summary_sheet = workbook["Sheet1"]
            sheet = workbook["目标表"]

            self.assertEqual(result["success"], True)
            self.assertEqual(result["written_row_count"], 60)
            self.assertEqual(result["not_shipped_row_count"], 59)
            self.assertEqual(result["partial_row_count"], 1)
            self.assertEqual(result["unknown_lookup_count"], 1)
            self.assertGreaterEqual(len(result["warnings"]), 1)

            self.assertEqual(summary_sheet.max_row, 64)
            self.assertEqual(
                [summary_sheet.cell(row=1, column=column_index).value for column_index in range(1, 12)],
                ["Order Type", "BULK", None, None, None, None, None, None, None, None, None],
            )
            self.assertEqual(
                [summary_sheet.cell(row=3, column=column_index).value for column_index in range(1, 12)],
                [
                    "Working Number",
                    "Article Number",
                    "PO Number",
                    "Market PO Number",
                    "Gps Customer Number",
                    "Assigned Factory",
                    "PODD",
                    "Price Per Unit",
                    "Total Adjustments",
                    "Sum of Ordered Quantity",
                    "Desc.",
                ],
            )
            self.assertEqual(summary_sheet.auto_filter.ref, "A3:K63")
            self.assertEqual(summary_sheet.cell(row=4, column=1).value, "WN000")
            self.assertEqual(summary_sheet.cell(row=4, column=3).value, "0900000000")
            self.assertEqual(summary_sheet.cell(row=4, column=7).value.date().isoformat(), "2026-07-15")
            self.assertEqual(summary_sheet.cell(row=4, column=8).value, 120)
            self.assertEqual(summary_sheet.cell(row=4, column=9).value, 5)
            self.assertEqual(summary_sheet.cell(row=4, column=10).value, 10)
            self.assertEqual(summary_sheet.cell(row=4, column=11).value, "BULK")
            self.assertEqual(summary_sheet.cell(row=63, column=1).value, "WN999")
            self.assertEqual(summary_sheet.cell(row=63, column=3).value, "0900009999")
            self.assertEqual(summary_sheet.cell(row=63, column=7).value.date().isoformat(), "2026-06-30")
            self.assertEqual(summary_sheet.cell(row=63, column=9).value, 6)
            self.assertEqual(summary_sheet.cell(row=63, column=10).value, 4)
            self.assertEqual(summary_sheet.cell(row=63, column=11).value, "BULK")
            self.assertEqual(summary_sheet.cell(row=64, column=1).value, "Grand Total")
            self.assertEqual(summary_sheet.cell(row=64, column=10).value, 594)

            self.assertEqual(sheet.max_row, 63)
            self.assertIsNone(sheet.cell(row=3, column=1).value)
            self.assertIsNone(sheet.cell(row=3, column=2).value)
            self.assertEqual(sheet.cell(row=3, column=3).value, "WN000")
            self.assertIsNone(sheet.cell(row=3, column=5).value)
            self.assertEqual(sheet.cell(row=3, column=6).value, "0900000000")
            self.assertEqual(sheet.cell(row=3, column=9).value, "苏州仓")
            self.assertEqual(sheet.cell(row=3, column=10).value.date().isoformat(), "2026-07-10")
            self.assertEqual(sheet.cell(row=3, column=11).value, 10)
            self.assertIsNone(sheet.cell(row=3, column=12).value)
            self.assertEqual(sheet.cell(row=3, column=13).value, "=ROUND(0.13*L3*K3,2)")
            self.assertEqual(sheet.cell(row=3, column=14).value, "=ROUND(L3*K3+M3,2)")
            self.assertEqual(sheet.cell(row=3, column=15).value, 120)
            self.assertEqual(sheet.cell(row=3, column=16).value, "=ROUND(0.13*(K3*O3+Q3),2)")
            self.assertEqual(sheet.cell(row=3, column=17).value, 5)
            self.assertEqual(sheet.cell(row=3, column=18).value, "=ROUND(O3*K3+P3+Q3,2)")
            self.assertEqual(sheet.cell(row=3, column=20).value, "丹东新龙太")
            self.assertIsNone(sheet.cell(row=3, column=22).value)
            self.assertIsNone(sheet.cell(row=3, column=23).value)

            self.assertEqual(sheet.cell(row=62, column=3).value, "WN999")
            self.assertEqual(sheet.cell(row=62, column=6).value, "0900009999")
            self.assertEqual(sheet.cell(row=62, column=10).value.date().isoformat(), "2026-06-30")
            self.assertEqual(sheet.cell(row=62, column=11).value, 4)
            self.assertIsNone(sheet.cell(row=62, column=12).value)
            self.assertEqual(sheet.cell(row=62, column=17).value, 6)
            price_fill_columns = set(range(12, 19))
            for row_index in (3, 62):
                for column_index in range(1, 24):
                    fill_type = sheet.cell(row=row_index, column=column_index).fill.fill_type
                    if column_index in price_fill_columns:
                        self.assertEqual(fill_type, "solid")
                    else:
                        self.assertIsNone(fill_type)

            self.assertEqual(sheet.cell(row=63, column=6).value, "=COUNTA(F3:F62)")
            self.assertEqual(sheet.cell(row=63, column=11).value, "=SUM(K3:K62)")
            self.assertEqual(sheet.cell(row=63, column=13).value, "=SUM(M3:M62)")
            self.assertEqual(sheet.cell(row=63, column=14).value, "=SUM(N3:N62)")
            self.assertEqual(sheet.cell(row=63, column=18).value, "=SUM(R3:R62)")
            for column_index in range(1, 24):
                self.assertIsNone(sheet.cell(row=63, column=column_index).fill.fill_type)
            self.assertEqual({str(range_ref) for range_ref in sheet.merged_cells.ranges}, EXPECTED_HEADER_MERGES)
            self.assertEqual(sheet["S1"].fill.fill_type, "solid")
            self.assertEqual(sheet["S2"].fill.fill_type, sheet["S1"].fill.fill_type)
            self.assertEqual(sheet["T2"].fill.fill_type, sheet["S1"].fill.fill_type)
            self.assertEqual(sheet["S2"].fill.fgColor.rgb, sheet["S1"].fill.fgColor.rgb)
            self.assertEqual(sheet["T2"].fill.fgColor.rgb, sheet["S1"].fill.fgColor.rgb)
            self.assertIsNone(sheet.auto_filter.ref)
            workbook.close()

            with ZipFile(result["output_path"]) as output_zip:
                output_names = output_zip.namelist()
                self.assertFalse(any(name.startswith("xl/externalLinks/") for name in output_names))
                self.assertFalse(any(name.startswith("xl/pivotCache/") for name in output_names))
                all_xml = b"".join(
                    output_zip.read(name)
                    for name in output_names
                    if name.endswith(".xml") or name.endswith(".rels")
                )
                self.assertNotIn(b"MR.xlsx", all_xml)

    def test_process_uses_optional_target_fields_from_result_set_when_present(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_path = temp_path / "result-set-with-optional-fields.xlsx"
            template_path = temp_path / "template.xlsx"
            self._write_result_set_workbook(
                source_path,
                not_shipped_count=2,
                include_optional_target_fields=True,
            )
            self._write_template_workbook(template_path)

            module = JasonResultSetExcelModule(template_path=template_path)
            result = module.process_result_set(
                result_set_path=source_path,
                target_month="2026-07",
                output_dir=temp_path,
            )

            workbook = openpyxl.load_workbook(result["output_path"], data_only=False, keep_links=False)
            try:
                sheet = workbook["目标表"]
                self.assertEqual(sheet.cell(row=3, column=1).value, "Source Pack 0")
                self.assertEqual(sheet.cell(row=3, column=2).value, "Source Season 0")
                self.assertEqual(sheet.cell(row=3, column=5).value, "Source Description 0")
                self.assertEqual(sheet.cell(row=3, column=22).value, "Source TMS 0")
                self.assertEqual(sheet.cell(row=3, column=23).value, "Source Factory 0")
            finally:
                workbook.close()

    def test_process_filters_sample_rows_and_writes_sample_label(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_path = temp_path / "result-set-sample.xlsx"
            template_path = temp_path / "template.xlsx"
            self._write_result_set_workbook(source_path, not_shipped_count=0)

            source_workbook = openpyxl.load_workbook(source_path)
            try:
                source_sheet = source_workbook["Result Set"]
                self._append_result_set_row(
                    source_sheet,
                    po="0900001000",
                    market_po="0300001000",
                    working_number="WN010",
                    article_number="ART001",
                    gps_customer_number="825066",
                    assigned_factory="1L8006",
                    podd=datetime(2026, 7, 18),
                    order_type="ZSAM",
                    shipped_status="Not Shipped",
                    price=120,
                    adjustments=5,
                    ordered_quantity=8,
                    shipped_quantity=0,
                )
                source_workbook.save(source_path)
            finally:
                source_workbook.close()

            self._write_template_workbook(template_path)
            module = JasonResultSetExcelModule(template_path=template_path)
            result = module.process_result_set(
                result_set_path=source_path,
                date_from="2026-07-01",
                date_to="2026-07-31",
                order_type_filter="sample",
                output_dir=temp_path,
            )

            workbook = openpyxl.load_workbook(result["output_path"], data_only=False, keep_links=False)
            try:
                summary_sheet = workbook["Sheet1"]
                target_sheet = workbook["目标表"]
                self.assertEqual(result["written_row_count"], 1)
                self.assertEqual(result["order_type_label"], "SAMPLE")
                self.assertEqual(summary_sheet.cell(row=1, column=2).value, "SAMPLE")
                self.assertEqual(summary_sheet.cell(row=4, column=1).value, "WN010")
                self.assertEqual(summary_sheet.cell(row=4, column=11).value, "SAMPLE")
                self.assertEqual(target_sheet.cell(row=3, column=3).value, "WN010")
                self.assertEqual(target_sheet.max_column, 23)
            finally:
                workbook.close()

    def test_process_bulk_sample_keeps_partial_rows_and_writes_combined_label(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_path = temp_path / "result-set-bulk-sample.xlsx"
            template_path = temp_path / "template.xlsx"
            self._write_result_set_workbook(source_path, not_shipped_count=0)

            source_workbook = openpyxl.load_workbook(source_path)
            try:
                source_sheet = source_workbook["Result Set"]
                self._append_result_set_row(
                    source_sheet,
                    po="0900001000",
                    market_po="0300001000",
                    working_number="WN010",
                    article_number="ART001",
                    gps_customer_number="825066",
                    assigned_factory="1L8006",
                    podd=datetime(2026, 7, 18),
                    order_type="ZSAM",
                    shipped_status="Not Shipped",
                    price=120,
                    adjustments=5,
                    ordered_quantity=8,
                    shipped_quantity=0,
                )
                source_workbook.save(source_path)
            finally:
                source_workbook.close()

            self._write_template_workbook(template_path)
            module = JasonResultSetExcelModule(template_path=template_path)
            result = module.process_result_set(
                result_set_path=source_path,
                date_from="2026-07-01",
                date_to="2026-07-31",
                order_type_filter="bulk_sample",
                output_dir=temp_path,
            )

            workbook = openpyxl.load_workbook(result["output_path"], data_only=False, keep_links=False)
            try:
                summary_sheet = workbook["Sheet1"]
                self.assertEqual(result["written_row_count"], 2)
                self.assertEqual(result["partial_row_count"], 1)
                self.assertEqual(result["order_type_label"], "BULK, SAMPLE")
                self.assertEqual(summary_sheet.cell(row=1, column=2).value, "BULK, SAMPLE")
                self.assertEqual(
                    [summary_sheet.cell(row=row_index, column=11).value for row_index in range(4, 6)],
                    ["SAMPLE", "BULK"],
                )
            finally:
                workbook.close()

    def test_process_without_date_filter_does_not_filter_podd(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_path = temp_path / "result-set-no-date-filter.xlsx"
            template_path = temp_path / "template.xlsx"
            self._write_result_set_workbook(source_path, not_shipped_count=0)

            source_workbook = openpyxl.load_workbook(source_path)
            try:
                source_sheet = source_workbook["Result Set"]
                self._append_result_set_row(
                    source_sheet,
                    po="0900000500",
                    market_po="0300000500",
                    working_number="WN005",
                    article_number="ART001",
                    gps_customer_number="825066",
                    assigned_factory="1L8006",
                    podd=datetime(2026, 5, 9),
                    order_type="ZGPS",
                    shipped_status="Not Shipped",
                    price=120,
                    adjustments=5,
                    ordered_quantity=7,
                    shipped_quantity=0,
                )
                source_workbook.save(source_path)
            finally:
                source_workbook.close()

            self._write_template_workbook(template_path)
            module = JasonResultSetExcelModule(template_path=template_path)
            result = module.process_result_set(
                result_set_path=source_path,
                date_from=None,
                date_to=None,
                order_type_filter="bulk",
                output_dir=temp_path,
            )

            workbook = openpyxl.load_workbook(result["output_path"], data_only=False, keep_links=False)
            try:
                summary_sheet = workbook["Sheet1"]
                self.assertEqual(result["written_row_count"], 2)
                self.assertEqual(summary_sheet.cell(row=4, column=1).value, "WN005")
                self.assertEqual(summary_sheet.cell(row=4, column=7).value.date().isoformat(), "2026-05-09")
            finally:
                workbook.close()

    def test_process_rejects_partial_date_range(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_path = temp_path / "result-set-partial-date.xlsx"
            template_path = temp_path / "template.xlsx"
            self._write_result_set_workbook(source_path, not_shipped_count=1)
            self._write_template_workbook(template_path)

            module = JasonResultSetExcelModule(template_path=template_path)
            with self.assertRaisesRegex(ValueError, "date_from 和 date_to 必须同时填写"):
                module.process_result_set(
                    result_set_path=source_path,
                    date_from="2026-07-01",
                    date_to=None,
                    order_type_filter="bulk",
                    output_dir=temp_path,
                )

    def test_process_sorts_rows_by_working_number_then_po_number(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_path = temp_path / "result-set-sort.xlsx"
            template_path = temp_path / "template.xlsx"
            self._write_result_set_workbook(source_path, not_shipped_count=0)

            source_workbook = openpyxl.load_workbook(source_path)
            try:
                source_sheet = source_workbook["Result Set"]
                self._append_result_set_row(
                    source_sheet,
                    po="0900000002",
                    market_po="0300000002",
                    working_number="WN100",
                    article_number="ART001",
                    gps_customer_number="825066",
                    assigned_factory="1L8006",
                    podd=datetime(2026, 7, 15),
                    order_type="ZGPS",
                    shipped_status="Not Shipped",
                    price=120,
                    adjustments=5,
                    ordered_quantity=10,
                    shipped_quantity=0,
                )
                self._append_result_set_row(
                    source_sheet,
                    po="0900009998",
                    market_po="0300009998",
                    working_number="WN050",
                    article_number="ART001",
                    gps_customer_number="825066",
                    assigned_factory="1L8006",
                    podd=datetime(2026, 7, 15),
                    order_type="ZGPS",
                    shipped_status="Not Shipped",
                    price=120,
                    adjustments=5,
                    ordered_quantity=10,
                    shipped_quantity=0,
                )
                self._append_result_set_row(
                    source_sheet,
                    po="0900000001",
                    market_po="0300000001",
                    working_number="WN100",
                    article_number="ART999",
                    gps_customer_number="825066",
                    assigned_factory="1L8006",
                    podd=datetime(2026, 7, 15),
                    order_type="ZGPS",
                    shipped_status="Not Shipped",
                    price=120,
                    adjustments=5,
                    ordered_quantity=10,
                    shipped_quantity=0,
                )
                source_workbook.save(source_path)
            finally:
                source_workbook.close()
            self._write_template_workbook(template_path)

            module = JasonResultSetExcelModule(template_path=template_path)
            result = module.process_result_set(
                result_set_path=source_path,
                target_month="2026-07",
                output_dir=temp_path,
            )

            workbook = openpyxl.load_workbook(result["output_path"], data_only=False, keep_links=False)
            try:
                sheet = workbook["目标表"]
                summary_sheet = workbook["Sheet1"]
                self.assertEqual(result["written_row_count"], 4)
                self.assertEqual(
                    [
                        (
                            sheet.cell(row=row_index, column=3).value,
                            sheet.cell(row=row_index, column=6).value,
                            sheet.cell(row=row_index, column=4).value,
                        )
                        for row_index in range(3, 7)
                    ],
                    [
                        ("WN050", "0900009998", "ART001"),
                        ("WN100", "0900000001", "ART999"),
                        ("WN100", "0900000002", "ART001"),
                        ("WN999", "0900009999", "ART999"),
                    ],
                )
                self.assertEqual(
                    [
                        (
                            summary_sheet.cell(row=row_index, column=1).value,
                            summary_sheet.cell(row=row_index, column=3).value,
                            summary_sheet.cell(row=row_index, column=2).value,
                        )
                        for row_index in range(4, 8)
                    ],
                    [
                        ("WN050", "0900009998", "ART001"),
                        ("WN100", "0900000001", "ART999"),
                        ("WN100", "0900000002", "ART001"),
                        ("WN999", "0900009999", "ART999"),
                    ],
                )
            finally:
                workbook.close()

    def test_process_rejects_missing_required_result_set_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            temp_path = Path(temp_dir)
            source_path = temp_path / "missing-columns.xlsx"
            template_path = temp_path / "template.xlsx"
            workbook = openpyxl.Workbook()
            sheet = workbook.active
            sheet.title = "Result Set"
            sheet.append(["PO Number", "Order Type"])
            sheet.append(["0900000000", "ZGPS"])
            workbook.save(source_path)
            self._write_template_workbook(template_path)

            module = JasonResultSetExcelModule(template_path=template_path)
            with self.assertRaisesRegex(ValueError, "Working Number"):
                module.process_result_set(
                    result_set_path=source_path,
                    target_month="2026-07",
                    output_dir=temp_path,
                )

    def _write_result_set_workbook(
        self,
        path: Path,
        *,
        not_shipped_count: int,
        include_optional_target_fields: bool = False,
    ) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Result Set"
        headers = [
            "For Account Of",
            "Seller",
            "Assigned Factory",
            "Factory",
            "PO Number",
            "PO Line Aggregator",
            "PO Line Item #",
            "MTFC Number",
            "Shipped Status",
            "Gps Customer Number",
            "Company Code",
            "Plant Code",
            "Customer Number",
            "Item Category",
            "Account Assignment Category",
            "Shipment Method",
            "Market PO Number",
            "Customer PO",
            "Class Code",
            "Order Type",
            "Release Status",
            "Classification",
            "Types",
            "Last Update Date",
            "Customer Request Date (CRD)",
            "Plan Date",
            "PO Batch Date",
            "Working Number",
            "Model Number",
            "Model Name",
            "Gender",
            "Article Number",
            "Article Description",
            "Business Model Attribute",
            "PSDD",
            "PODD",
            "First Production Date",
            "Last Production Date",
            "Confirmation Delay",
            "Delivery Delay",
            "Mark For Address",
            "Total Qty",
            "Price Per Unit",
            "Total Merchandise Amount",
            "Total Tax Amount",
            "Total Adjustments",
            "Total Document Amount",
            "Customer Size Run",
            "Technical Notation",
            "Customer Size",
            "Technical Size",
            "Ordered Quantity",
            "Shipped Qty",
        ]
        if include_optional_target_fields:
            headers.extend([
                "Pack",
                "Season",
                "Description",
                "TMS Merchandiser",
                "Factory Merchandiser",
            ])
        sheet.append(headers)

        for index in range(not_shipped_count):
            optional_target_fields = None
            if include_optional_target_fields:
                optional_target_fields = {
                    "Pack": f"Source Pack {index}",
                    "Season": f"Source Season {index}",
                    "Description": f"Source Description {index}",
                    "TMS Merchandiser": f"Source TMS {index}",
                    "Factory Merchandiser": f"Source Factory {index}",
                }
            self._append_result_set_row(
                sheet,
                po=f"090000{index:04d}",
                market_po=f"030000{index:04d}",
                working_number=f"WN{index:03d}",
                article_number="ART001" if index < not_shipped_count - 1 else "UNKNOWN",
                gps_customer_number="825066",
                assigned_factory="1L8006",
                podd=datetime(2026, 7, 15),
                order_type="ZGPS",
                shipped_status="Not Shipped",
                price=120,
                adjustments=5,
                ordered_quantity=10,
                shipped_quantity=0,
                optional_target_fields=optional_target_fields,
            )

        self._append_result_set_row(
            sheet,
            po="0900009999",
            market_po="0300009999",
            working_number="WN999",
            article_number="ART999",
            gps_customer_number="825066",
            assigned_factory="1L8006",
            podd=datetime(2026, 6, 30),
            order_type="ZGPS",
            shipped_status="Partially Shipped",
            price=220,
            adjustments=10,
            ordered_quantity=10,
            shipped_quantity=6,
            optional_target_fields=None,
        )
        self._append_result_set_row(
            sheet,
            po="0900008888",
            market_po="0300008888",
            working_number="WN888",
            article_number="ART001",
            gps_customer_number="825066",
            assigned_factory="1L8006",
            podd=datetime(2026, 7, 20),
            order_type="ZGPS",
            shipped_status="Fully Shipped",
            price=120,
            adjustments=5,
            ordered_quantity=10,
            shipped_quantity=10,
            optional_target_fields=None,
        )
        workbook.save(path)

    def _append_result_set_row(
        self,
        sheet: openpyxl.worksheet.worksheet.Worksheet,
        *,
        po: str,
        market_po: str,
        working_number: str,
        article_number: str,
        gps_customer_number: str,
        assigned_factory: str,
        podd: datetime,
        order_type: str,
        shipped_status: str,
        price: float,
        adjustments: float,
        ordered_quantity: int,
        shipped_quantity: int,
        optional_target_fields: dict[str, object] | None = None,
    ) -> None:
        row = [None] * 53
        row[2] = assigned_factory
        row[4] = po
        row[8] = shipped_status
        row[9] = gps_customer_number
        row[16] = market_po
        row[19] = order_type
        row[27] = working_number
        row[31] = article_number
        row[32] = f"{article_number} English"
        row[35] = podd
        row[42] = price
        row[45] = adjustments
        row[51] = ordered_quantity
        row[52] = shipped_quantity
        if optional_target_fields is not None:
            row.extend([
                optional_target_fields.get("Pack"),
                optional_target_fields.get("Season"),
                optional_target_fields.get("Description"),
                optional_target_fields.get("TMS Merchandiser"),
                optional_target_fields.get("Factory Merchandiser"),
            ])
        sheet.append(row)

    def _write_template_workbook(self, path: Path) -> None:
        workbook = openpyxl.Workbook()
        target = workbook.active
        target.title = "目标表"
        target.append([
            "Pack",
            "Season",
            "Working Number",
            "Article Number",
            "Description",
            "PO Number",
            "Market PO Number",
            "Gps Customer Number",
            "Customer Warehouse",
            "Bulk Handover Date",
            "Ordered Quantity",
            "Factory Price",
            None,
            None,
            "TMS Price",
            None,
            None,
            None,
            "Factory",
            None,
            "Remark",
            "TMS Merchandiser",
            "Factory Merchandiser",
        ])
        target.append([
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            None,
            "Price Per Unit",
            "PO Tax",
            " PO Total Price",
            "Price Per Unit",
            "PO Tax",
            "Total Adjustments",
            " PO Total Price",
            "Code",
            "Name",
            None,
            None,
            None,
        ])
        target.append([
            "Pack A",
            "SS26",
            "WN000",
            "ART001",
            "中文描述",
            "0900000000",
            "0300000000",
            "825066",
            "苏州仓",
            datetime(2026, 7, 10),
            10,
            100,
            "=ROUND(0.13*L3*K3,2)",
            "=ROUND(L3*K3+M3,2)",
            120,
            "=ROUND(0.13*(K3*O3+Q3),2)",
            5,
            "=ROUND(O3*K3+P3+Q3,2)",
            "1L8006",
            "丹东新龙太",
            None,
            "Bacy",
            "Jasmine",
        ])
        target.append([
            "Pack B",
            "FW26",
            "WN999",
            "ART999",
            "夹克",
            "0900009999",
            "0300009999",
            "825066",
            "苏州仓",
            datetime(2026, 6, 30),
            4,
            200,
            "=ROUND(0.13*L4*K4,2)",
            "=ROUND(L4*K4+M4,2)",
            220,
            "=ROUND(0.13*(K4*O4+Q4),2)",
            6,
            "=ROUND(O4*K4+P4+Q4,2)",
            "1L8006",
            "丹东新龙太",
            None,
            "Jason",
            "Judy",
        ])
        target.append([None, None, None, None, None, "=COUNTA(F3:F4)", None, None, None, None, "=SUM(K3:K4)"])
        target.auto_filter.ref = "A2:W5"
        for column_index in range(1, 24):
            target.cell(row=3, column=column_index).fill = openpyxl.styles.PatternFill(
                fill_type="solid",
                fgColor="D9EAF7",
            )
        target["S1"].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="F4C7A1")
        target["S2"].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="D9E6EF")
        target["T2"].fill = openpyxl.styles.PatternFill(fill_type="solid", fgColor="D9E6EF")

        warehouse = workbook.create_sheet("Sheet2")
        warehouse.append(["Gps Customer Number", "Customer Warehouse"])
        warehouse.append(["825066", "苏州仓"])

        factory = workbook.create_sheet("Sheet3")
        factory.append(["Assigned Factory", "Name"])
        factory.append(["1L8006", "丹东新龙太"])
        workbook.save(path)


if __name__ == "__main__":
    unittest.main()

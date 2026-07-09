from __future__ import annotations

import os
import sys
import tempfile
import unittest
from datetime import datetime
from pathlib import Path
from typing import Callable

import openpyxl
from fastapi.testclient import TestClient


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

import main  # noqa: E402
from api import jason_result_set_excel_api as api  # noqa: E402
from modules.jason_result_set_excel_module import JasonResultSetExcelModule  # noqa: E402


class JasonResultSetExcelApiTests(unittest.TestCase):
    def setUp(self) -> None:
        self.client = TestClient(main.app)
        self.temp_dir = tempfile.TemporaryDirectory()
        self.runtime_dir = Path(self.temp_dir.name)
        self.template_path = self.runtime_dir / "template.xlsx"
        self.source_path = self.runtime_dir / "source.xlsx"
        self._restore_callbacks: list[Callable[[], None]] = []
        self._write_template_workbook(self.template_path)
        self._write_source_workbook(self.source_path)
        self._patch("UPLOAD_DIR", str(self.runtime_dir))
        self._patch("jason_result_set_excel_module", JasonResultSetExcelModule(template_path=self.template_path))
        main.app.openapi_schema = None

    def tearDown(self) -> None:
        for restore in reversed(self._restore_callbacks):
            restore()
        self.temp_dir.cleanup()
        main.app.openapi_schema = None

    def test_process_route_uploads_result_set_and_returns_download_metadata(self) -> None:
        with self.source_path.open("rb") as file_obj:
            response = self.client.post(
                "/api/jason/result-set-excel/process",
                files={
                    "result_set_file": (
                        "result-set.xlsx",
                        file_obj,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={"target_month": "2026-07"},
            )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["success"], True)
        self.assertEqual(payload["written_row_count"], 1)
        self.assertEqual(payload["output_file"].split(".")[-1], "xlsx")
        self.assertEqual(payload["requestId"], payload["request_id"])
        self.assertIn("/api/jason/result-set-excel/download/", payload["resultDownloadPath"])

        download = self.client.get(payload["resultDownloadPath"])
        self.assertEqual(download.status_code, 200)
        self.assertGreater(len(download.content), 0)

    def test_process_route_accepts_date_range_and_order_type_filter(self) -> None:
        with self.source_path.open("rb") as file_obj:
            response = self.client.post(
                "/api/jason/result-set-excel/process",
                files={
                    "result_set_file": (
                        "result-set.xlsx",
                        file_obj,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={
                    "date_filter_mode": "range",
                    "date_from": "2026-07-01",
                    "date_to": "2026-07-31",
                    "order_type_filter": "bulk",
                },
            )

        self.assertEqual(response.status_code, 200)
        payload = response.json()
        self.assertEqual(payload["success"], True)
        self.assertEqual(payload["date_filter_mode"], "range")
        self.assertEqual(payload["date_from"], "2026-07-01")
        self.assertEqual(payload["date_to"], "2026-07-31")
        self.assertEqual(payload["order_type_filter"], "bulk")
        self.assertEqual(payload["order_type_label"], "BULK")
        self.assertEqual(payload["written_row_count"], 1)

    def test_process_route_rejects_partial_date_range(self) -> None:
        with self.source_path.open("rb") as file_obj:
            response = self.client.post(
                "/api/jason/result-set-excel/process",
                files={
                    "result_set_file": (
                        "result-set.xlsx",
                        file_obj,
                        "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                    )
                },
                data={
                    "date_filter_mode": "range",
                    "date_from": "2026-07-01",
                    "order_type_filter": "bulk",
                },
            )

        self.assertEqual(response.status_code, 400)
        self.assertIn("date_from 和 date_to 必须同时填写", response.json()["detail"])

    def test_process_route_rejects_bad_upload_filename(self) -> None:
        response = self.client.post(
            "/api/jason/result-set-excel/process",
            files={
                "result_set_file": (
                    "../source.xlsx",
                    b"fake",
                    "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                )
            },
            data={"target_month": "2026-07"},
        )

        self.assertEqual(response.status_code, 400)

    def test_download_route_rejects_path_traversal(self) -> None:
        response = self.client.get("/api/jason/result-set-excel/download/..%2Fsecret.xlsx")

        self.assertEqual(response.status_code, 400)

    def _patch(self, name: str, value: object) -> None:
        original = getattr(api, name)
        setattr(api, name, value)
        self._restore_callbacks.append(lambda: setattr(api, name, original))

    def _write_source_workbook(self, path: Path) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Result Set"
        headers = [
            "Assigned Factory",
            "PO Number",
            "Shipped Status",
            "Gps Customer Number",
            "Market PO Number",
            "Order Type",
            "Working Number",
            "Article Number",
            "PODD",
            "Price Per Unit",
            "Total Adjustments",
            "Ordered Quantity",
            "Shipped Qty",
        ]
        sheet.append(headers)
        sheet.append([
            "1L8006",
            "0900000001",
            "Not Shipped",
            "825066",
            "0300000001",
            "ZGPS",
            "WN001",
            "ART001",
            datetime(2026, 7, 15),
            120,
            5,
            10,
            0,
        ])
        workbook.save(path)

    def _write_template_workbook(self, path: Path) -> None:
        workbook = openpyxl.Workbook()
        target = workbook.active
        target.title = "目标表"
        target.append(["Pack", "Season", "Working Number", "Article Number", "Description"])
        target.append([None, None, None, None, None])
        target.append([
            "Pack A",
            "SS26",
            "WN001",
            "ART001",
            "中文描述",
            "0900000001",
            "0300000001",
            "825066",
            "苏州仓",
            datetime(2026, 7, 10),
            10,
            100,
            "=ROUND(0.13*L3*K3,2)",
            "=ROUND(L3*K3+M3,2)",
            120,
            "=ROUND(0.13*(K3*O3+Q3),2)",
            5,
            "=ROUND(O3*K3+P3+Q3,2)",
            "1L8006",
            "丹东新龙太",
            None,
            "Bacy",
            "Jasmine",
        ])
        warehouse = workbook.create_sheet("Sheet2")
        warehouse.append(["Gps Customer Number", "Customer Warehouse"])
        warehouse.append(["825066", "苏州仓"])
        factory = workbook.create_sheet("Sheet3")
        factory.append(["Assigned Factory", "Name"])
        factory.append(["1L8006", "丹东新龙太"])
        workbook.save(path)


if __name__ == "__main__":
    unittest.main()

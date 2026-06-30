import asyncio
import io
import os
import sys
import tempfile
import unittest

import openpyxl
from fastapi import HTTPException
from fastapi.testclient import TestClient


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from api import excel_template_mapper_api  # noqa: E402
from main import app  # noqa: E402


class ExcelTemplateMapperApiTests(unittest.TestCase):
    def test_api_inspects_processes_and_downloads_result(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            original_upload_dir = excel_template_mapper_api.UPLOAD_DIR
            excel_template_mapper_api.UPLOAD_DIR = temp_dir
            try:
                client = TestClient(app)

                inspect_response = client.post(
                    "/api/excel-template-mapper/inspect",
                    files={
                        "excel_file": (
                            "source.xlsx",
                            self._build_source_workbook(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        ),
                    },
                    data={"sheet_name": "Source", "header_row": "1"},
                )
                self.assertEqual(inspect_response.status_code, 200)
                self.assertEqual(
                    inspect_response.json()["selected_sheet"]["headers"][0]["label"],
                    "PO Number",
                )

                process_response = client.post(
                    "/api/excel-template-mapper/process",
                    files={
                        "source_file": (
                            "source.xlsx",
                            self._build_source_workbook(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        ),
                        "template_file": (
                            "template.xlsx",
                            self._build_template_workbook(),
                            "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                        ),
                    },
                    data={
                        "config_json": (
                            '{"source_sheet_name":"Source","template_sheet_name":"Template",'
                            '"source_header_row":1,"source_data_start_row":2,'
                            '"template_header_row":1,"template_data_start_row":2,'
                            '"mappings":['
                            '{"target_column":1,"source_column":1,"target_header":"Order","required":true},'
                            '{"target_column":2,"source_column":2,"target_header":"Style","required":true}'
                            "]}"
                        ),
                    },
                )
                self.assertEqual(process_response.status_code, 200)
                payload = process_response.json()
                self.assertTrue(payload["success"])
                self.assertEqual(payload["source_row_count"], 2)
                self.assertEqual(payload["written_row_count"], 2)
                self.assertTrue(payload["output_file"].endswith(".xlsx"))
                self.assertNotIn("\\", payload["output_file"])
                self.assertNotIn("/", payload["output_file"])

                download_response = client.get(
                    f"/api/excel-template-mapper/download/{payload['output_file']}",
                )
                self.assertEqual(download_response.status_code, 200)
                workbook = openpyxl.load_workbook(io.BytesIO(download_response.content))
                self.assertEqual(workbook["Template"].cell(row=2, column=1).value, "PO-001")
            finally:
                excel_template_mapper_api.UPLOAD_DIR = original_upload_dir

    def test_download_rejects_path_traversal(self) -> None:
        with self.assertRaises(HTTPException) as context:
            asyncio.run(excel_template_mapper_api.download_excel_template_mapper_result("../secret.xlsx"))

        self.assertEqual(context.exception.status_code, 400)

    def _build_source_workbook(self) -> bytes:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Source"
        sheet.append(["PO Number", "Style Name"])
        sheet.append(["PO-001", "Jacket"])
        sheet.append(["PO-002", "Pants"])
        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue()

    def _build_template_workbook(self) -> bytes:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Template"
        sheet.append(["Order", "Style"])
        sheet.append(["", ""])
        stream = io.BytesIO()
        workbook.save(stream)
        return stream.getvalue()


if __name__ == "__main__":
    unittest.main()

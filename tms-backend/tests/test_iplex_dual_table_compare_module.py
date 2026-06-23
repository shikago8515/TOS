import os
import sys
import tempfile
import unittest
from pathlib import Path
from unittest.mock import patch

import openpyxl


BACKEND_ROOT = os.path.abspath(os.path.join(os.path.dirname(__file__), ".."))
if BACKEND_ROOT not in sys.path:
    sys.path.insert(0, BACKEND_ROOT)

from modules.iplex_dual_table_compare_module import (  # noqa: E402
    IplexDualTableCompareConfig,
    IplexDualTableCompareModule,
)


class IplexDualTableCompareModuleTests(unittest.TestCase):
    def test_inspect_reads_xlsx_and_xls_headers_with_one_based_columns(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            lookup_path = Path(temp_dir) / "lookup.xls"
            self._write_main_xlsx(main_path)
            self._write_lookup_xls(lookup_path)

            module = IplexDualTableCompareModule()

            with patch("modules.iplex_dual_table_compare_module.xlrd.open_workbook", return_value=FakeXlsBook()):
                main_inspection = module.inspect_workbook(str(main_path), header_row=1)
                lookup_inspection = module.inspect_workbook(str(lookup_path), header_row=1)

        self.assertEqual(main_inspection["sheets"][0]["name"], "Main")
        self.assertEqual(main_inspection["selected_sheet"]["headers"][0]["index"], 1)
        self.assertEqual(main_inspection["selected_sheet"]["headers"][0]["label"], "PO #")
        self.assertEqual(main_inspection["selected_sheet"]["headers"][2]["label"], "PO #")
        self.assertEqual(lookup_inspection["selected_sheet"]["headers"][2]["index"], 3)
        self.assertEqual(lookup_inspection["selected_sheet"]["headers"][2]["label"], "BUYER ORDER NO.")

    def test_process_writes_formula_columns_helper_sheet_filter_and_freeze_panes(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            lookup_path = Path(temp_dir) / "lookup.xls"
            self._write_main_xlsx(main_path)
            self._write_lookup_xls(lookup_path)

            module = IplexDualTableCompareModule()
            with patch("modules.iplex_dual_table_compare_module.xlrd.open_workbook", return_value=FakeXlsBook()):
                result = module.process_files(
                    main_path=str(main_path),
                    lookup_path=str(lookup_path),
                    config=IplexDualTableCompareConfig(
                        main_sheet_name="Main",
                        lookup_sheet_name="Lookup",
                        main_header_row=1,
                        lookup_header_row=1,
                        main_key_column=1,
                        lookup_key_column=3,
                        four_digit_main_column=4,
                        four_digit_lookup_column=7,
                        two_digit_main_column=5,
                        two_digit_lookup_column=10,
                    ),
                    output_dir=temp_dir,
                )

            workbook = openpyxl.load_workbook(result["output_path"], data_only=False)
            sheet = workbook["Main"]
            helper = workbook["_iplex_lookup"]

        self.assertTrue(result["output_path"].endswith(".xlsx"))
        self.assertEqual(result["main_row_count"], 3)
        self.assertEqual(result["lookup_row_count"], 2)
        self.assertEqual(result["matched_count"], 2)
        self.assertEqual(result["unmatched_count"], 1)
        self.assertEqual(result["four_digit_mismatch_count"], 1)
        self.assertEqual(result["two_digit_mismatch_count"], 1)
        self.assertEqual(
            result["preview_rows"],
            [
                {
                    "row_number": 3,
                    "key": "0902893237",
                    "status": "不一致",
                    "four_digit": {
                        "main_value": "0.2501",
                        "lookup_value": "0.2500",
                        "difference": "0.0001",
                    },
                    "two_digit": {
                        "main_value": "25.49",
                        "lookup_value": "25.50",
                        "difference": "-0.01",
                    },
                },
                {
                    "row_number": 4,
                    "key": "0900000000",
                    "status": "未匹配",
                    "four_digit": {
                        "main_value": "1.0000",
                        "lookup_value": "#N/A",
                        "difference": "#N/A",
                    },
                    "two_digit": {
                        "main_value": "50.00",
                        "lookup_value": "#N/A",
                        "difference": "#N/A",
                    },
                },
            ],
        )

        self.assertEqual(sheet.cell(row=1, column=6).value, "4位小数差值")
        self.assertEqual(sheet.cell(row=1, column=7).value, "2位小数差值")
        self.assertEqual(
            sheet.cell(row=2, column=6).value,
            "=ROUND(D2-VLOOKUP(A2,'_iplex_lookup'!$A$2:$C$3,2,FALSE),4)",
        )
        self.assertEqual(
            sheet.cell(row=2, column=7).value,
            "=ROUND(E2-VLOOKUP(A2,'_iplex_lookup'!$A$2:$C$3,3,FALSE),2)",
        )
        self.assertEqual(sheet.cell(row=4, column=6).value, "=ROUND(D4-VLOOKUP(A4,'_iplex_lookup'!$A$2:$C$3,2,FALSE),4)")
        self.assertEqual(sheet.cell(row=2, column=6).number_format, "0.0000")
        self.assertEqual(sheet.cell(row=2, column=7).number_format, "0.00")
        self.assertNotEqual(sheet.cell(row=2, column=1).fill.fgColor.rgb, "FFFFC7CE")
        self.assertEqual(sheet.cell(row=3, column=1).fill.fgColor.rgb, "FFFFC7CE")
        self.assertEqual(sheet.cell(row=3, column=7).fill.fgColor.rgb, "FFFFC7CE")
        self.assertEqual(sheet.cell(row=4, column=1).fill.fgColor.rgb, "FFFFC7CE")
        self.assertEqual(sheet.cell(row=4, column=7).fill.fgColor.rgb, "FFFFC7CE")
        self.assertEqual(sheet.auto_filter.ref, "A1:G4")
        self.assertEqual(sheet.freeze_panes, "A2")
        self.assertEqual(helper.sheet_state, "hidden")
        self.assertEqual(helper.cell(row=2, column=1).value, "0902893225")
        self.assertEqual(helper.cell(row=2, column=2).value, 0.1)

    def test_process_allows_custom_result_headers(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            lookup_path = Path(temp_dir) / "lookup.xls"
            self._write_main_xlsx(main_path)
            self._write_lookup_xls(lookup_path)

            module = IplexDualTableCompareModule()
            with patch("modules.iplex_dual_table_compare_module.xlrd.open_workbook", return_value=FakeXlsBook()):
                result = module.process_files(
                    main_path=str(main_path),
                    lookup_path=str(lookup_path),
                    config=IplexDualTableCompareConfig(
                        main_sheet_name="Main",
                        lookup_sheet_name="Lookup",
                        main_header_row=1,
                        lookup_header_row=1,
                        main_key_column=1,
                        lookup_key_column=3,
                        four_digit_main_column=4,
                        four_digit_lookup_column=7,
                        two_digit_main_column=5,
                        two_digit_lookup_column=10,
                        four_digit_result_header="Unit Diff",
                        two_digit_result_header="Total Diff",
                    ),
                    output_dir=temp_dir,
                )
            workbook = openpyxl.load_workbook(result["output_path"], data_only=False)
            sheet = workbook["Main"]

        self.assertEqual(sheet.cell(row=1, column=6).value, "Unit Diff")
        self.assertEqual(sheet.cell(row=1, column=7).value, "Total Diff")

    def test_process_skips_blank_key_total_rows(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xlsx"
            lookup_path = Path(temp_dir) / "lookup.xls"
            self._write_main_with_total_row_xlsx(main_path)
            self._write_lookup_xls(lookup_path)

            module = IplexDualTableCompareModule()
            with patch("modules.iplex_dual_table_compare_module.xlrd.open_workbook", return_value=FakeXlsBook()):
                result = module.process_files(
                    main_path=str(main_path),
                    lookup_path=str(lookup_path),
                    config=IplexDualTableCompareConfig(
                        main_sheet_name="Main",
                        lookup_sheet_name="Lookup",
                        main_header_row=1,
                        lookup_header_row=1,
                        main_key_column=1,
                        lookup_key_column=3,
                        four_digit_main_column=4,
                        four_digit_lookup_column=7,
                        two_digit_main_column=5,
                        two_digit_lookup_column=10,
                    ),
                    output_dir=temp_dir,
                )

            workbook = openpyxl.load_workbook(result["output_path"], data_only=False)
            sheet = workbook["Main"]

        self.assertEqual(result["main_row_count"], 1)
        self.assertEqual(result["matched_count"], 1)
        self.assertEqual(result["unmatched_count"], 0)
        self.assertEqual(result["preview_rows"], [])
        self.assertEqual(sheet.cell(row=2, column=6).value, "=ROUND(D2-VLOOKUP(A2,'_iplex_lookup'!$A$2:$C$3,2,FALSE),4)")
        self.assertIsNone(sheet.cell(row=3, column=6).value)
        self.assertIsNone(sheet.cell(row=3, column=7).value)
        self.assertNotEqual(sheet.cell(row=3, column=1).fill.fgColor.rgb, "FFFFC7CE")
        self.assertNotEqual(sheet.cell(row=3, column=7).fill.fgColor.rgb, "FFFFC7CE")

    def test_process_formats_xls_date_columns_as_short_dates(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            main_path = Path(temp_dir) / "main.xls"
            lookup_path = Path(temp_dir) / "lookup.xls"
            self._write_lookup_xls(lookup_path)

            def open_fake_workbook(path: str, *args: object, **kwargs: object) -> FakeXlsBook:
                if Path(path).name == "main.xls":
                    return FakeXlsBook(FakeMainDateXlsSheet())
                return FakeXlsBook()

            module = IplexDualTableCompareModule()
            with patch("modules.iplex_dual_table_compare_module.xlrd.open_workbook", side_effect=open_fake_workbook):
                result = module.process_files(
                    main_path=str(main_path),
                    lookup_path=str(lookup_path),
                    config=IplexDualTableCompareConfig(
                        main_sheet_name="Main",
                        lookup_sheet_name="Lookup",
                        main_header_row=1,
                        lookup_header_row=1,
                        main_key_column=1,
                        lookup_key_column=3,
                        four_digit_main_column=4,
                        four_digit_lookup_column=7,
                        two_digit_main_column=5,
                        two_digit_lookup_column=10,
                    ),
                    output_dir=temp_dir,
                )

            workbook = openpyxl.load_workbook(result["output_path"], data_only=False)
            sheet = workbook["Main"]

        self.assertEqual(sheet.cell(row=1, column=6).value, "CREATE DATE")
        self.assertEqual(sheet.cell(row=2, column=6).value.date().isoformat(), "2026-06-23")
        self.assertEqual(sheet.cell(row=2, column=6).number_format, "yyyy-mm-dd")

    def _write_main_xlsx(self, path: Path) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Main"
        sheet.append(["PO #", "Article Number", "PO #", "Adjustment_per_unit", "Total Adjustment Amount"])
        sheet.append(["0902893225", "LG4295", "duplicate", 0.1000, 10.00])
        sheet.append(["0902893237", "LG4292", "duplicate", 0.2501, 25.49])
        sheet.append(["0900000000", "MISSING", "duplicate", 1.0000, 50.00])
        workbook.save(path)

    def _write_main_with_total_row_xlsx(self, path: Path) -> None:
        workbook = openpyxl.Workbook()
        sheet = workbook.active
        sheet.title = "Main"
        sheet.append(["PO #", "Article Number", "PO #", "Adjustment_per_unit", "Total Adjustment Amount"])
        sheet.append(["0902893225", "LG4295", "duplicate", 0.1000, 10.00])
        sheet.append(["", "TOTAL", "", None, 1123.32])
        workbook.save(path)

    def _write_lookup_xls(self, path: Path) -> None:
        path.write_bytes(b"legacy xls placeholder")


class FakeXlsSheet:
    name = "Lookup"
    rows = [
        [
            "WORKING NUMBER",
            "ARTICLE NUMBER",
            "BUYER ORDER NO.",
            "CURRENCY",
            "PROMO PRICE UPCHARGE (%)",
            "KIDS FIRE LABEL",
            "SHAS PRICE PER UNIT",
            "",
            "",
            "TOTAL ADJUSTMENT",
        ],
        ["RC2620OW014", "LG4295", "0902893225", "USD", 0, 0, 0.1000, "", "", 10.00],
        ["RC2620OW014", "LG4292", "0902893237", "USD", 0, 0, 0.2500, "", "", 25.50],
    ]
    nrows = len(rows)
    ncols = len(rows[0])

    def row_values(self, row_index: int) -> list[object]:
        return self.rows[row_index]

    def cell_value(self, row_index: int, column_index: int) -> object:
        return self.rows[row_index][column_index]

    def cell_type(self, row_index: int, column_index: int) -> int:
        value = self.cell_value(row_index, column_index)
        if value in (None, ""):
            return 0
        if isinstance(value, (int, float)):
            return 2
        return 1


class FakeMainDateXlsSheet:
    name = "Main"
    rows = [
        [
            "BUYER ORDER NO.",
            "ARTICLE NUMBER",
            "REMARK",
            "SHAS PRICE PER UNIT",
            "TOTAL ADJUSTMENT",
            "CREATE DATE",
        ],
        ["0902893225", "LG4295", "", 0.1000, 10.00, 46196.41319444445],
    ]
    nrows = len(rows)
    ncols = len(rows[0])

    def row_values(self, row_index: int) -> list[object]:
        return self.rows[row_index]

    def cell_value(self, row_index: int, column_index: int) -> object:
        return self.rows[row_index][column_index]

    def cell_type(self, row_index: int, column_index: int) -> int:
        if row_index == 1 and column_index == 5:
            return 3
        value = self.cell_value(row_index, column_index)
        if value in (None, ""):
            return 0
        if isinstance(value, (int, float)):
            return 2
        return 1


class FakeXlsBook:
    datemode = 0

    def __init__(self, sheet: object | None = None) -> None:
        self.sheet = sheet or FakeXlsSheet()

    def sheets(self) -> list[FakeXlsSheet]:
        return [self.sheet]

    def sheet_by_name(self, name: str) -> FakeXlsSheet:
        if name != self.sheet.name:
            raise KeyError(name)
        return self.sheet

    def release_resources(self) -> None:
        return None


if __name__ == "__main__":
    unittest.main()

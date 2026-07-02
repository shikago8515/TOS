import os
import base64
import json
import tempfile
import unittest
from datetime import datetime
from io import BytesIO
from pathlib import Path
from types import SimpleNamespace
from unittest.mock import AsyncMock, patch

from openpyxl import load_workbook

from api import automation_storage_api
from utils import credential_crypto, mysql_store
from utils.credential_crypto import decrypt_secret, encrypt_secret
from utils.mysql_store import SCHEMA_DDL
from api.automation_storage_api import (
    PackingListCheckpointFilePayload,
    PackingListCheckpointPayload,
    RunFilePayload,
    RunUpdatePayload,
    TicketOwnerAttemptPayload,
    TicketOwnerCheckpointPayload,
    TicketOwnerInterruptPayload,
    _attachment_content_disposition,
    _credential_lookup_ids,
    _normalize_account_key,
    _run_file_payload,
    _store_remote_result_file,
    _template_payload,
)
from scripts.seed_automation_templates import AUTOMATION_TEMPLATES, read_template_content


class AutomationStorageTests(unittest.TestCase):
    def test_schema_contains_new_tos_business_tables(self):
        schema_text = "\n".join(SCHEMA_DDL)

        self.assertIn("tos_people", schema_text)
        self.assertIn("tos_modules", schema_text)
        self.assertIn("tos_activity_records", schema_text)
        self.assertIn("tos_activity_files", schema_text)
        self.assertIn("tos_module_templates", schema_text)
        self.assertIn("tos_login_accounts", schema_text)
        self.assertIn("tos_release_records", schema_text)
        self.assertNotIn("CREATE TABLE IF NOT EXISTS automation_runs", schema_text)
        self.assertNotIn("CREATE TABLE IF NOT EXISTS excel_templates", schema_text)

    def test_list_automation_runs_uses_lightweight_columns(self):
        executed_sql: list[str] = []

        class CursorContext:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def execute(self, sql, _params=()):
                executed_sql.append(sql)

            def fetchall(self):
                return []

        class ConnectionContext:
            def __enter__(self):
                return SimpleNamespace(cursor=lambda: CursorContext())

            def __exit__(self, *_args):
                return False

        with patch.object(mysql_store, "ensure_schema"), \
             patch.object(mysql_store, "mysql_connection", return_value=ConnectionContext()):
            rows = mysql_store.list_automation_runs(limit=80)

        self.assertEqual(rows, [])
        self.assertEqual(len(executed_sql), 1)
        self.assertNotIn("result_json", executed_sql[0].lower())
        self.assertIn("started_at", executed_sql[0].lower())

    def test_list_automation_batches_can_use_lightweight_columns(self):
        executed_sql: list[str] = []

        class CursorContext:
            def __enter__(self):
                return self

            def __exit__(self, *_args):
                return False

            def execute(self, sql, _params=()):
                executed_sql.append(sql)

            def fetchall(self):
                return []

        class ConnectionContext:
            def __enter__(self):
                return SimpleNamespace(cursor=lambda: CursorContext())

            def __exit__(self, *_args):
                return False

        with patch.object(mysql_store, "ensure_schema"), \
             patch.object(mysql_store, "mysql_connection", return_value=ConnectionContext()):
            rows = mysql_store.list_automation_batches(
                "packing-list-auto-download",
                limit=20,
                include_payloads=False,
            )

        self.assertEqual(rows, [])
        self.assertEqual(len(executed_sql), 1)
        sql = executed_sql[0].lower()
        self.assertNotIn("checkpoint_json", sql)
        self.assertNotIn("source_file_json", sql)
        self.assertIn("total_count", sql)

    def test_latest_packing_list_batch_only_loads_full_resumable_batch(self):
        lightweight_rows = [
            {
                "batch_id": "completed-batch",
                "automation_id": "packing-list-auto-download",
                "module_id": "packing-list-auto-download",
                "run_id": "run-1",
                "batch_name": "done",
                "source_file_name": "done.xlsx",
                "source_file_sha256": "",
                "status": "success",
                "message": "",
                "total_count": 1,
                "completed_count": 1,
                "failed_count": 0,
                "pending_count": 0,
                "bucket": "tos-packing-list-auto-download",
                "object_prefix": "",
                "created_at": datetime(2026, 7, 1, 8, 0, 0),
                "updated_at": datetime(2026, 7, 1, 8, 0, 0),
            },
            {
                "batch_id": "resumable-batch",
                "automation_id": "packing-list-auto-download",
                "module_id": "packing-list-auto-download",
                "run_id": "run-2",
                "batch_name": "pending",
                "source_file_name": "pending.xlsx",
                "source_file_sha256": "",
                "status": "pending",
                "message": "",
                "total_count": 10,
                "completed_count": 3,
                "failed_count": 0,
                "pending_count": 7,
                "bucket": "tos-packing-list-auto-download",
                "object_prefix": "",
                "created_at": datetime(2026, 7, 1, 9, 0, 0),
                "updated_at": datetime(2026, 7, 1, 9, 0, 0),
            },
        ]
        full_row = packing_list_batch_row({
            **lightweight_rows[1],
            "source_file": {"downloadPath": "/download/source.xlsx"},
            "checkpoint": {"items": [{"no": "NO-001", "status": "pending"}]},
        })

        with patch.object(automation_storage_api, "list_automation_batches", return_value=lightweight_rows) as list_batches, \
             patch.object(automation_storage_api, "get_automation_batch", return_value=full_row) as get_batch:
            response = automation_storage_api.read_latest_packing_list_batch()

        list_batches.assert_called_once_with(
            "packing-list-auto-download",
            limit=20,
            include_payloads=False,
        )
        get_batch.assert_called_once_with("resumable-batch")
        self.assertEqual(response["batch"]["batchId"], "resumable-batch")
        self.assertEqual(response["batch"]["checkpoint"]["items"][0]["no"], "NO-001")

    def test_credentials_encrypt_and_decrypt_with_local_key_file(self):
        original_key_file = os.environ.get(credential_crypto.CREDENTIAL_KEY_FILE_ENV)
        original_key = os.environ.get(credential_crypto.CREDENTIAL_KEY_ENV)
        try:
            os.environ.pop(credential_crypto.CREDENTIAL_KEY_ENV, None)
            with tempfile.TemporaryDirectory() as temp_dir:
                key_file = Path(temp_dir) / "credential.key"
                os.environ[credential_crypto.CREDENTIAL_KEY_FILE_ENV] = str(key_file)

                encrypted = encrypt_secret("plain-password")

                self.assertNotEqual(encrypted, "plain-password")
                self.assertEqual(decrypt_secret(encrypted), "plain-password")
                self.assertTrue(key_file.exists())
        finally:
            restore_env(credential_crypto.CREDENTIAL_KEY_FILE_ENV, original_key_file)
            restore_env(credential_crypto.CREDENTIAL_KEY_ENV, original_key)

    def test_xinlongtai_shipping_template_is_registered_from_source_file(self):
        template = next(
            item for item in AUTOMATION_TEMPLATES
            if item["module_id"] == "xinlongtai-shipping-automation"
        )

        content = read_template_content(template)

        self.assertEqual(template["filename"], "新龙泰-shipping-自动化模板.XLS")
        self.assertGreater(len(content), 0)

    def test_wandai_shipping_template_is_registered_from_source_file(self):
        template = next(
            item for item in AUTOMATION_TEMPLATES
            if item["module_id"] == "shipping-automation"
        )

        content = read_template_content(template)

        self.assertEqual(template["filename"], "万代-shipping-自动化模板.xlsx")
        self.assertEqual(template["display_name"], "万代 Shipping 自动化 Excel 模板")
        self.assertGreater(len(content), 0)

    def test_template_download_header_supports_chinese_filename(self):
        header = _attachment_content_disposition("新龙泰-shipping-自动化模板.XLS")

        self.assertIn('filename="', header)
        self.assertIn("filename*=UTF-8''", header)
        self.assertIn("%E6%96%B0%E9%BE%99%E6%B3%B0", header)

    def test_template_download_reports_missing_minio_object(self):
        class MissingObjectError(Exception):
            code = "NoSuchKey"

        row = {
            "bucket": "tos-templates",
            "object_key": "templates/missing.xlsx",
            "original_filename": "template.xlsx",
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        with patch.object(automation_storage_api, "get_excel_template", return_value=row), \
             patch.object(automation_storage_api, "get_object_response", side_effect=MissingObjectError()):
            with self.assertRaises(automation_storage_api.HTTPException) as context:
                automation_storage_api.download_template(1)

        self.assertEqual(context.exception.status_code, 404)
        self.assertIn("模板文件不存在", context.exception.detail)

    def test_template_download_reports_storage_unavailable(self):
        row = {
            "bucket": "tos-templates",
            "object_key": "templates/template.xlsx",
            "original_filename": "template.xlsx",
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        }
        with patch.object(automation_storage_api, "get_excel_template", return_value=row), \
             patch.object(automation_storage_api, "get_object_response", side_effect=RuntimeError("storage down")):
            with self.assertRaises(automation_storage_api.HTTPException) as context:
                automation_storage_api.download_template(1)

        self.assertEqual(context.exception.status_code, 503)
        self.assertIn("MinIO", context.exception.detail)

    def test_automation_credentials_are_isolated_by_page(self):
        lookup_ids = _credential_lookup_ids("shipping-automation-demo")
        released_bulk_lookup_ids = _credential_lookup_ids("shipping-automation-2")
        packing_list_lookup_ids = _credential_lookup_ids("packing-list-auto-download")

        self.assertEqual(lookup_ids, ["shipping-automation-demo"])
        self.assertEqual(released_bulk_lookup_ids, ["shipping-automation-2"])
        self.assertEqual(packing_list_lookup_ids, ["packing-list-auto-download"])

    def test_resolve_credentials_does_not_fallback_to_another_page(self):
        rows = {
            "packing-list-auto-download": {
                "account_key": "default",
                "username": "broken-user",
                "password_ciphertext": "bad-ciphertext",
            },
            "po-auto-download": {
                "account_key": "default",
                "username": "shared-user",
                "password_ciphertext": "good-ciphertext",
            },
        }
        lookup_calls: list[tuple[str, str]] = []

        def fake_decrypt_secret(ciphertext):
            if ciphertext == "bad-ciphertext":
                raise ValueError("wrong key")
            return "plain-password"

        def fake_get_automation_credentials(automation_id, account_key):
            lookup_calls.append((automation_id, account_key))
            return rows.get(automation_id)

        with patch.object(
            automation_storage_api,
            "get_automation_credentials",
            side_effect=fake_get_automation_credentials,
        ), \
             patch.object(automation_storage_api, "decrypt_secret", side_effect=fake_decrypt_secret):
            with self.assertRaises(automation_storage_api.HTTPException) as context:
                automation_storage_api.resolve_credentials(
                    "packing-list-auto-download",
                    accountKey="default",
                )

        self.assertEqual(context.exception.status_code, 409)
        self.assertEqual(lookup_calls, [("packing-list-auto-download", "default")])

    def test_microsoft_credentials_are_isolated_by_page(self):
        lookup_ids = _credential_lookup_ids("ticket-owner-statistics")
        microsoft_lookup_ids = _credential_lookup_ids("microsoft-login-n8n")

        self.assertEqual(lookup_ids, ["ticket-owner-statistics"])
        self.assertEqual(microsoft_lookup_ids, ["microsoft-login-n8n"])

    def test_credential_account_key_is_normalized_for_saved_profiles(self):
        self.assertEqual(_normalize_account_key("  Lily  "), "Lily")
        self.assertEqual(_normalize_account_key(""), "default")
        self.assertEqual(len(_normalize_account_key("x" * 160)), 120)

    def test_create_run_continues_when_source_file_storage_is_unavailable(self):
        row = build_run_row("run-1")

        with patch.object(automation_storage_api, "create_automation_run", return_value=row), \
             patch.object(
                 automation_storage_api,
                 "_store_upload_file",
                 new=AsyncMock(side_effect=RuntimeError("storage backend detail")),
             ):
            response = run_async(automation_storage_api.create_run(
                automation_id="infornexus-auto-add",
                module_id="",
                run_name="debug",
                message="started",
                source_file=SimpleNamespace(filename="source.xlsx"),
            ))

        self.assertTrue(response["ok"])
        self.assertEqual(response["run"]["runId"], "run-1")
        self.assertEqual(response["files"], [])
        self.assertEqual(len(response["warnings"]), 1)
        self.assertNotIn("storage backend detail", response["warnings"][0])

    def test_finish_run_updates_status_when_result_file_storage_is_unavailable(self):
        existing_row = build_run_row("run-2", status="running")
        updated_row = build_run_row("run-2", status="failed", message="executor failed")
        payload = RunUpdatePayload(
            status="failed",
            message="executor failed",
            result={"ok": False, "message": "executor failed"},
            resultFiles=[],
        )

        with patch.object(automation_storage_api, "get_automation_run", return_value=existing_row), \
             patch.object(automation_storage_api, "update_automation_run", return_value=updated_row), \
             patch.object(
                 automation_storage_api,
                 "_store_result_json",
                 side_effect=RuntimeError("storage backend detail"),
             ):
            response = run_async(automation_storage_api.finish_run("run-2", payload))

        self.assertTrue(response["ok"])
        self.assertEqual(response["run"]["status"], "failed")
        self.assertEqual(response["files"], [])
        self.assertEqual(len(response["warnings"]), 1)
        self.assertNotIn("storage backend detail", response["warnings"][0])

    def test_result_file_can_be_stored_from_base64_content(self):
        payload = RunFilePayload(
            url="http://127.0.0.1:3002/artifacts/result.xlsx",
            fileName="Ticket ownership.xlsx",
            fileRole="result_excel",
            contentType="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            contentBase64=base64.b64encode(b"excel-bytes").decode("ascii"),
        )
        inserted_row = {
            "id": 12,
            "run_id": "run-12",
            "file_role": "result_excel",
            "bucket": "tos-results",
            "object_key": "results/ticket-owner-statistics/run-12/result.xlsx",
            "original_filename": "Ticket ownership.xlsx",
            "content_type": payload.contentType,
            "file_size": len(b"excel-bytes"),
            "sha256": "c" * 64,
            "created_at": datetime(2026, 6, 29, 12, 0, 0),
        }

        with patch.object(automation_storage_api, "get_minio_bucket", return_value="tos-results"), \
             patch.object(automation_storage_api, "build_object_key", return_value=inserted_row["object_key"]), \
             patch.object(automation_storage_api, "put_object_bytes", return_value={
                 "file_size": len(b"excel-bytes"),
                 "sha256": "c" * 64,
             }) as put_object, \
             patch.object(automation_storage_api, "download_url_bytes") as download_url, \
             patch.object(automation_storage_api, "insert_automation_run_file", return_value=inserted_row):
            row = _store_remote_result_file("run-12", "ticket-owner-statistics", payload)

        self.assertEqual(row["id"], 12)
        put_object.assert_called_once()
        self.assertEqual(put_object.call_args.kwargs["content"], b"excel-bytes")
        download_url.assert_not_called()

    def test_packing_list_batch_uses_dedicated_bucket_and_date_batch_prefix(self):
        stored_batch_rows = []

        class FakeUploadFile:
            filename = "source.xlsx"
            content_type = automation_storage_api.EXCEL_CONTENT_TYPE

            async def read(self):
                return b"excel-source"

        def fake_create_automation_batch(batch):
            row = packing_list_batch_row(batch)
            stored_batch_rows.append(row)
            return row

        with patch.object(automation_storage_api, "get_minio_bucket", return_value="tos-packing-list-auto-download") as get_bucket, \
             patch.object(automation_storage_api, "put_object_bytes", return_value={
                 "file_size": len(b"excel-source"),
                 "sha256": "f" * 64,
             }) as put_object, \
             patch.object(automation_storage_api, "create_automation_batch", side_effect=fake_create_automation_batch), \
             patch.object(automation_storage_api, "insert_automation_run_file", return_value={
                 "id": 31,
                 "run_id": "run-plad",
                 "file_role": "source_excel",
                 "bucket": "tos-packing-list-auto-download",
                 "object_key": "packing-list-auto-download/2026-07-01/plad/source/source.xlsx",
                 "original_filename": "source.xlsx",
                 "content_type": automation_storage_api.EXCEL_CONTENT_TYPE,
                 "file_size": len(b"excel-source"),
                 "sha256": "f" * 64,
                 "created_at": datetime(2026, 7, 1, 8, 0, 0),
             }):
            response = run_async(automation_storage_api.create_packing_list_batch(
                run_id="run-plad",
                batch_name="test batch",
                source_file=FakeUploadFile(),
            ))

        self.assertTrue(response["ok"])
        get_bucket.assert_called_once_with("packing_list_auto_download")
        put_kwargs = put_object.call_args.kwargs
        self.assertEqual(put_kwargs["bucket"], "tos-packing-list-auto-download")
        self.assertRegex(
            put_kwargs["object_key"],
            r"^packing-list-auto-download/\d{4}-\d{2}-\d{2}/plad-\d{14}-[a-f0-9]{8}/source/source\.xlsx$",
        )
        self.assertEqual(response["batch"]["bucket"], "tos-packing-list-auto-download")
        self.assertEqual(response["batch"]["sourceFile"]["bucket"], "tos-packing-list-auto-download")
        self.assertEqual(stored_batch_rows[0]["automation_id"], "packing-list-auto-download")

    def test_packing_list_checkpoint_marks_interrupted_and_stores_files_under_attempt(self):
        batch_row = packing_list_batch_row({
            "batch_id": "plad-20260701100000-abcd1234",
            "automation_id": "packing-list-auto-download",
            "module_id": "packing-list-auto-download",
            "run_id": "run-plad",
            "batch_name": "source.xlsx",
            "source_file_name": "source.xlsx",
            "source_file_sha256": "f" * 64,
            "source_file": {
                "bucket": "tos-packing-list-auto-download",
                "objectKey": "packing-list-auto-download/2026-07-01/plad-20260701100000-abcd1234/source/source.xlsx",
                "originalFilename": "source.xlsx",
            },
            "status": "running",
            "message": "running",
            "total_count": 3,
            "completed_count": 1,
            "failed_count": 0,
            "pending_count": 2,
            "checkpoint": {
                "items": [
                    {"no": "NO-001", "status": "success"},
                    {"no": "NO-002", "status": "pending"},
                    {"no": "NO-003", "status": "pending"},
                ],
                "groupResults": [{"no": "NO-001", "ok": True, "filePath": "Packing list NO-001.pdf"}],
            },
            "bucket": "tos-packing-list-auto-download",
            "object_prefix": "packing-list-auto-download/2026-07-01/plad-20260701100000-abcd1234",
        })
        updated_rows = []

        def fake_update_batch(_batch_id, update):
            row = packing_list_batch_row({
                **batch_row,
                "status": update["status"],
                "message": update["message"],
                "total_count": update["total_count"],
                "completed_count": update["completed_count"],
                "failed_count": update["failed_count"],
                "pending_count": update["pending_count"],
                "checkpoint": update["checkpoint"],
            })
            updated_rows.append(row)
            return row

        payload = PackingListCheckpointPayload(
            runId="run-plad",
            attemptId="pla-attempt-1",
            mode="continue",
            status="interrupted",
            message="browser closed",
            checkpoint={
                "items": [
                    {"no": "NO-001", "status": "success"},
                    {"no": "NO-002", "status": "pending"},
                    {"no": "NO-003", "status": "pending"},
                ],
                "groupResults": [{"no": "NO-001", "ok": True, "filePath": "Packing list NO-001.pdf"}],
            },
            groupResult={"no": "NO-001", "ok": True, "filePath": "Packing list NO-001.pdf"},
            files=[PackingListCheckpointFilePayload(
                fileName="Packing list NO-001.pdf",
                fileRole="packing_list_pdf",
                contentType="application/pdf",
                contentBase64=base64.b64encode(b"%PDF-1.4").decode("ascii"),
            )],
        )

        with patch.object(automation_storage_api, "get_automation_batch", return_value=batch_row), \
             patch.object(automation_storage_api, "get_automation_batch_attempt", return_value=None), \
             patch.object(automation_storage_api, "create_automation_batch_attempt") as create_attempt, \
             patch.object(automation_storage_api, "update_automation_batch_checkpoint", side_effect=fake_update_batch) as update_batch, \
             patch.object(automation_storage_api, "update_automation_batch_attempt") as update_attempt, \
             patch.object(automation_storage_api, "put_object_bytes", return_value={
                 "file_size": len(b"%PDF-1.4"),
                 "sha256": "a" * 64,
             }) as put_object, \
             patch.object(automation_storage_api, "insert_automation_run_file", return_value={
                 "id": 32,
                 "run_id": "run-plad",
                 "file_role": "packing_list_pdf",
                 "bucket": "tos-packing-list-auto-download",
                 "object_key": "packing-list-auto-download/2026-07-01/plad-20260701100000-abcd1234/pla-attempt-1/packing_list_pdf/file.pdf",
                 "original_filename": "Packing list NO-001.pdf",
                 "content_type": "application/pdf",
                 "file_size": len(b"%PDF-1.4"),
                 "sha256": "a" * 64,
                 "created_at": datetime(2026, 7, 1, 8, 0, 0),
             }):
            response = automation_storage_api.write_packing_list_checkpoint(
                "plad-20260701100000-abcd1234",
                payload,
            )

        self.assertTrue(response["ok"])
        create_attempt.assert_called_once()
        update_batch.assert_called_once()
        update_attempt.assert_called_once()
        self.assertEqual(response["batch"]["status"], "interrupted")
        self.assertTrue(response["batch"]["resumable"])
        self.assertEqual(response["batch"]["completedCount"], 1)
        self.assertEqual(response["batch"]["pendingCount"], 2)
        put_kwargs = put_object.call_args.kwargs
        self.assertEqual(put_kwargs["bucket"], "tos-packing-list-auto-download")
        self.assertRegex(
            put_kwargs["object_key"],
            r"^packing-list-auto-download/2026-07-01/plad-20260701100000-abcd1234/pla-attempt-1/packing_list_pdf/\d{20}-Packing-list-NO-001\.pdf$",
        )
        stored_checkpoint = _safe_json(updated_rows[0]["checkpoint_json"])
        self.assertEqual(stored_checkpoint["status"], "interrupted")
        self.assertEqual(stored_checkpoint["items"][0]["storedFiles"][0]["bucket"], "tos-packing-list-auto-download")

    def test_ticket_owner_result_excel_is_built_on_server(self):
        existing_row = build_run_row("run-ticket", status="running")
        existing_row["automation_id"] = "ticket-owner-statistics"
        existing_row["module_id"] = "ticket-owner-statistics"
        updated_row = {
            **existing_row,
            "status": "success",
            "message": "done",
            "finished_at": datetime(2026, 6, 29, 12, 0, 0),
        }
        result = {
            "ok": True,
            "rows": [{
                "Case Number": "10682971",
                "Task Type": "Provide Feedback",
                "Request": "Transport Mode Change",
                "PO Number": "0902793368",
                "Working Number": "RC2606OW001",
                "Factory": "3LP001",
                "Merch": "Rosa",
            }],
        }
        payload = RunUpdatePayload(
            status="success",
            message="done",
            result=result,
            resultFiles=[RunFilePayload(
                url="http://127.0.0.1:3002/artifacts/local.xlsx",
                fileName="Ticket ownership.xlsx",
                fileRole="result_excel",
            )],
        )
        inserted_excel_row = {
            "id": 22,
            "run_id": "run-ticket",
            "file_role": "result_excel",
            "bucket": "tos-results",
            "object_key": "results/ticket-owner-statistics/run-ticket/result_excel/Ticket ownership.xlsx",
            "original_filename": "Ticket ownership.xlsx",
            "content_type": automation_storage_api.EXCEL_CONTENT_TYPE,
            "file_size": 1,
            "sha256": "d" * 64,
            "created_at": datetime(2026, 6, 29, 12, 0, 0),
        }

        with patch.object(automation_storage_api, "get_automation_run", return_value=existing_row), \
             patch.object(automation_storage_api, "update_automation_run", return_value=updated_row), \
             patch.object(automation_storage_api, "_store_result_json", return_value={
                 "id": 21,
                 "run_id": "run-ticket",
                 "file_role": "result_json",
                 "bucket": "tos-results",
                 "object_key": "results/ticket-owner-statistics/run-ticket/result.json",
                 "original_filename": "result.json",
                 "content_type": "application/json; charset=utf-8",
                 "file_size": 100,
                 "sha256": "e" * 64,
                 "created_at": datetime(2026, 6, 29, 12, 0, 0),
             }), \
             patch.object(automation_storage_api, "get_minio_bucket", return_value="tos-results"), \
             patch.object(automation_storage_api, "build_object_key", return_value=inserted_excel_row["object_key"]), \
             patch.object(automation_storage_api, "put_object_bytes", return_value={
                 "file_size": 2048,
                 "sha256": "d" * 64,
             }) as put_object, \
             patch.object(automation_storage_api, "insert_automation_run_file", return_value=inserted_excel_row), \
             patch.object(automation_storage_api, "_store_remote_result_file") as remote_store:
            response = run_async(automation_storage_api.finish_run("run-ticket", payload))

        self.assertEqual([file["fileRole"] for file in response["files"]], ["result_json", "result_excel"])
        self.assertEqual(response["files"][1]["downloadPath"], "/api/automation/run-files/22/download")
        remote_store.assert_not_called()
        excel_content = put_object.call_args.kwargs["content"]
        workbook = load_workbook(BytesIO(excel_content))
        sheet = workbook["Ticket ownership"]
        self.assertEqual(sheet["A1"].value, "Case Number")
        self.assertEqual(sheet["D2"].value, "0902793368")
        self.assertEqual(sheet["A1"].fill.fgColor.rgb, "FF5B9BD5")
        self.assertEqual(sheet["A2"].fill.fgColor.rgb, "FFDDEBF7")
        self.assertEqual(sheet.auto_filter.ref, "A1:G2")
        workbook.close()

    def test_ticket_owner_result_excel_uses_workflow_result_rows(self):
        result = {
            "ok": True,
            "workflowResult": {
                "rowCount": 1,
                "rows": [{
                    "Case Number": "GTS82967-1",
                    "Task Type": "Review Sub-Ticket Resolution",
                    "Request": "Bulk - Additional Support on WFM",
                    "PO Number": "0901943835",
                    "Working Number": "RC2610OM005",
                    "Factory": "3LP001",
                    "Merch": "Maggie",
                }],
            },
        }

        rows = automation_storage_api._extract_ticket_owner_rows(result)

        self.assertIsNotNone(rows)
        self.assertEqual(rows[0]["Case Number"], "GTS82967-1")
        content = automation_storage_api._build_ticket_owner_workbook_bytes(rows)
        workbook = load_workbook(BytesIO(content))
        sheet = workbook["Ticket ownership"]
        self.assertEqual(sheet["A2"].value, "GTS82967-1")
        self.assertEqual(sheet["G2"].value, "Maggie")
        workbook.close()

    def test_ticket_owner_checkpoint_interruption_and_resume_flow(self):
        current_row = ticket_owner_batch_row({
            "batch_id": "tos-20260702100000-abcd1234",
            "run_id": "run-ticket-owner",
            "batch_name": "Ticket ownership",
            "status": "pending",
            "total_count": 2,
            "completed_count": 0,
            "failed_count": 0,
            "pending_count": 2,
            "checkpoint": {
                "batchId": "tos-20260702100000-abcd1234",
                "status": "pending",
                "totalCount": 2,
                "completedCount": 0,
                "failedCount": 0,
                "pendingCount": 2,
                "items": [],
            },
            "bucket": "tos-ticket-owner-statistics",
            "object_prefix": "ticket-owner-statistics/2026-07-02/tos-20260702100000-abcd1234",
        })
        attempts: dict[str, dict] = {}
        inserted_files: list[dict] = []

        def fake_get_batch(_batch_id):
            return current_row

        def fake_create_attempt(payload):
            row = {
                "attempt_id": payload["attempt_id"],
                "batch_id": payload["batch_id"],
                "run_id": payload.get("run_id", ""),
                "mode": payload.get("mode", "new"),
                "status": payload.get("status", "running"),
                "message": payload.get("message", ""),
                "started_at": payload.get("started_at") or datetime(2026, 7, 2, 10, 0, 0),
                "finished_at": None,
                "created_at": datetime(2026, 7, 2, 10, 0, 0),
                "updated_at": datetime(2026, 7, 2, 10, 0, 0),
            }
            attempts[row["attempt_id"]] = row
            return row

        def fake_get_attempt(attempt_id):
            return attempts.get(attempt_id)

        def fake_update_attempt(attempt_id, update):
            row = attempts.setdefault(attempt_id, {
                "attempt_id": attempt_id,
                "batch_id": current_row["batch_id"],
                "run_id": "",
                "mode": "new",
                "status": "running",
                "message": "",
                "started_at": datetime(2026, 7, 2, 10, 0, 0),
                "finished_at": None,
                "created_at": datetime(2026, 7, 2, 10, 0, 0),
                "updated_at": datetime(2026, 7, 2, 10, 0, 0),
            })
            row.update(update)
            return row

        def fake_update_batch(_batch_id, update):
            nonlocal current_row
            current_row = ticket_owner_batch_row({
                **current_row,
                "status": update["status"],
                "message": update["message"],
                "total_count": update["total_count"],
                "completed_count": update["completed_count"],
                "failed_count": update["failed_count"],
                "pending_count": update["pending_count"],
                "checkpoint": update["checkpoint"],
            })
            return current_row

        def fake_insert_file(row):
            file_id = len(inserted_files) + 81
            inserted = {
                "id": file_id,
                "run_id": row.get("run_id", "run-ticket-owner"),
                "file_role": row["file_role"],
                "bucket": row["bucket"],
                "object_key": row["object_key"],
                "original_filename": row["original_filename"],
                "content_type": row["content_type"],
                "file_size": row.get("file_size", 1),
                "sha256": row.get("sha256", "f" * 64),
                "created_at": datetime(2026, 7, 2, 10, 0, 0),
            }
            inserted_files.append(inserted)
            return inserted

        with patch.object(automation_storage_api, "get_automation_batch", side_effect=fake_get_batch), \
             patch.object(automation_storage_api, "get_automation_batch_attempt", side_effect=fake_get_attempt), \
             patch.object(automation_storage_api, "create_automation_batch_attempt", side_effect=fake_create_attempt), \
             patch.object(automation_storage_api, "update_automation_batch_checkpoint", side_effect=fake_update_batch), \
             patch.object(automation_storage_api, "update_automation_batch_attempt", side_effect=fake_update_attempt), \
             patch.object(automation_storage_api, "put_object_bytes", return_value={
                 "file_size": 2048,
                 "sha256": "f" * 64,
             }), \
             patch.object(automation_storage_api, "insert_automation_run_file", side_effect=fake_insert_file):
            first_checkpoint = automation_storage_api.write_ticket_owner_checkpoint(
                "tos-20260702100000-abcd1234",
                TicketOwnerCheckpointPayload(
                    runId="run-ticket-owner",
                    attemptId="tos-attempt-1",
                    mode="new",
                    status="running",
                    message="ticket 1 done",
                    checkpoint={
                        "status": "running",
                        "totalCount": 2,
                        "completedCount": 1,
                        "failedCount": 0,
                        "pendingCount": 1,
                    },
                    itemResult={
                        "itemKey": "10682971",
                        "caseNumber": "10682971",
                        "taskType": "Provide Feedback",
                        "request": "Transport Mode Change",
                        "poNumber": "0902793368",
                        "workingNumber": "RC2606OW001",
                        "ok": True,
                    },
                ),
            )

            self.assertEqual(first_checkpoint["batch"]["status"], "running")
            self.assertEqual(first_checkpoint["batch"]["completedCount"], 1)
            self.assertEqual(first_checkpoint["batch"]["pendingCount"], 1)
            self.assertEqual(attempts["tos-attempt-1"]["status"], "running")
            self.assertIsNone(attempts["tos-attempt-1"]["finished_at"])

            interrupted = automation_storage_api.interrupt_ticket_owner_batch(
                "tos-20260702100000-abcd1234",
                TicketOwnerInterruptPayload(
                    runId="run-ticket-owner",
                    attemptId="tos-attempt-1",
                    message="browser closed",
                ),
            )

            self.assertEqual(interrupted["batch"]["status"], "interrupted")
            self.assertTrue(interrupted["batch"]["resumable"])
            self.assertEqual(interrupted["batch"]["completedCount"], 1)
            self.assertEqual(interrupted["batch"]["pendingCount"], 1)

            continue_attempt = automation_storage_api.create_ticket_owner_batch_attempt(
                "tos-20260702100000-abcd1234",
                TicketOwnerAttemptPayload(runId="run-ticket-owner-resume", mode="continue"),
            )

            self.assertEqual(continue_attempt["batch"]["batchId"], "tos-20260702100000-abcd1234")
            self.assertEqual(continue_attempt["attempt"]["mode"], "continue")
            attempt2_id = continue_attempt["attempt"]["attemptId"]
            self.assertNotEqual(attempt2_id, "tos-attempt-1")

            final_checkpoint = automation_storage_api.write_ticket_owner_checkpoint(
                "tos-20260702100000-abcd1234",
                TicketOwnerCheckpointPayload(
                    runId="run-ticket-owner-resume",
                    attemptId=attempt2_id,
                    mode="continue",
                    status="success",
                    message="done",
                    checkpoint={
                        "status": "success",
                        "totalCount": 2,
                        "completedCount": 2,
                        "failedCount": 0,
                        "pendingCount": 0,
                        "items": [{
                            "itemKey": "10682971",
                            "caseNumber": "10682971",
                            "status": "success",
                        }, {
                            "itemKey": "GTS82967-1",
                            "caseNumber": "GTS82967-1",
                            "status": "success",
                        }],
                    },
                    result={
                        "ok": True,
                        "rowCount": 2,
                        "failedTicketCount": 0,
                        "attemptedTicketCount": 2,
                        "rows": [{
                            "Case Number": "10682971",
                            "Task Type": "Provide Feedback",
                            "Request": "Transport Mode Change",
                            "PO Number": "0902793368",
                            "Working Number": "RC2606OW001",
                            "Factory": "3LP001",
                            "Merch": "Rosa",
                        }, {
                            "Case Number": "GTS82967-1",
                            "Task Type": "Review Sub-Ticket Resolution",
                            "Request": "Bulk - Additional Support on WFM",
                            "PO Number": "0901943835",
                            "Working Number": "RC2610OM005",
                            "Factory": "3LP001",
                            "Merch": "Maggie",
                        }],
                    },
                ),
            )

        self.assertEqual(final_checkpoint["batch"]["status"], "success")
        self.assertFalse(final_checkpoint["batch"]["resumable"])
        self.assertEqual(final_checkpoint["batch"]["completedCount"], 2)
        self.assertEqual(final_checkpoint["batch"]["pendingCount"], 0)
        file_roles = [file["fileRole"] for file in final_checkpoint["batch"]["checkpoint"]["storedFiles"]]
        self.assertEqual(file_roles, ["result_json", "result_excel"])
        self.assertTrue(all(file["downloadPath"].startswith("/api/automation/run-files/") for file in final_checkpoint["batch"]["checkpoint"]["storedFiles"]))
        self.assertTrue(all(f"/{attempt2_id}/" in file["object_key"] for file in inserted_files))

    def test_ticket_owner_progress_checkpoint_keeps_counts_without_items(self):
        checkpoint = {
            "status": "running",
            "event": "progress",
            "totalCount": 28,
            "completedCount": 22,
            "successCount": 22,
            "failedCount": 1,
            "pendingCount": 5,
            "items": [],
        }

        counts = automation_storage_api._ticket_owner_checkpoint_counts(checkpoint)

        self.assertEqual(counts["total"], 28)
        self.assertEqual(counts["completed"], 22)
        self.assertEqual(counts["failed"], 1)
        self.assertEqual(counts["pending"], 5)

    def test_delete_ticket_owner_batch_removes_batch_record(self):
        row = ticket_owner_batch_row({
            "batch_id": "tos-ticket-delete",
            "automation_id": automation_storage_api.TICKET_OWNER_AUTOMATION_ID,
        })

        with patch.object(automation_storage_api, "get_automation_batch", return_value=row), \
             patch.object(automation_storage_api, "delete_automation_batch", return_value=True) as delete_batch:
            response = automation_storage_api.delete_ticket_owner_batch("tos-ticket-delete")

        self.assertTrue(response["ok"])
        self.assertEqual(response["batchId"], "tos-ticket-delete")
        self.assertTrue(response["deleted"])
        delete_batch.assert_called_once_with("tos-ticket-delete")

    def test_template_payload_exposes_management_fields(self):
        row = {
            "id": 7,
            "module_id": "tc-inv-automation",
            "template_key": "default",
            "display_name": "TC INV 模板",
            "original_filename": "tc-inv.xlsx",
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "file_size": 1200,
            "sha256": "a" * 64,
            "is_active": 0,
            "created_at": datetime(2026, 6, 26, 8, 0, 0),
            "updated_at": datetime(2026, 6, 26, 9, 0, 0),
        }

        payload = _template_payload(row)

        self.assertEqual(payload["id"], 7)
        self.assertFalse(payload["isActive"])
        self.assertEqual(payload["downloadPath"], "/api/automation/templates/7/download")
        self.assertIn("2026-06-26", payload["updatedAt"])

    def test_run_file_payload_exposes_download_path(self):
        row = {
            "id": 11,
            "run_id": "run-11",
            "file_role": "result_excel",
            "bucket": "tos-results",
            "object_key": "results/run-11/result.xlsx",
            "original_filename": "result.xlsx",
            "content_type": "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
            "file_size": 4096,
            "sha256": "b" * 64,
            "created_at": datetime(2026, 6, 26, 10, 0, 0),
        }

        payload = _run_file_payload(row)

        self.assertEqual(payload["id"], 11)
        self.assertEqual(payload["fileRole"], "result_excel")
        self.assertEqual(payload["downloadPath"], "/api/automation/run-files/11/download")

    def test_read_templates_can_list_all_modules_for_management(self):
        rows = [
            {
                "id": 1,
                "module_id": "shipping-automation",
                "template_key": "default",
                "display_name": "万代模板",
                "original_filename": "wandai.xlsx",
                "content_type": "",
                "file_size": 1,
                "sha256": "",
                "is_active": 1,
                "created_at": None,
                "updated_at": None,
            },
            {
                "id": 2,
                "module_id": "tc-inv-automation",
                "template_key": "default",
                "display_name": "TC INV 模板",
                "original_filename": "tc-inv.xlsx",
                "content_type": "",
                "file_size": 1,
                "sha256": "",
                "is_active": 1,
                "created_at": None,
                "updated_at": None,
            },
        ]

        with patch.object(automation_storage_api, "list_excel_templates", return_value=rows) as list_templates:
            response = automation_storage_api.read_templates(moduleId=None, includeInactive=True, limit=500)

        list_templates.assert_called_once_with(None, include_inactive=True, limit=500)
        self.assertEqual(response["moduleId"], "")
        self.assertEqual([item["moduleId"] for item in response["templates"]], [
            "shipping-automation",
            "tc-inv-automation",
        ])


def restore_env(name, value):
    if value is None:
        os.environ.pop(name, None)
    else:
        os.environ[name] = value


def build_run_row(run_id, status="running", message="started"):
    now = datetime(2026, 6, 24, 3, 30, 0)
    return {
        "run_id": run_id,
        "automation_id": "infornexus-auto-add",
        "module_id": "infornexus-auto-add",
        "run_name": "debug",
        "status": status,
        "message": message,
        "result_json": None,
        "started_at": now,
        "finished_at": now if status != "running" else None,
        "created_at": now,
        "updated_at": now,
    }


def packing_list_batch_row(batch):
    now = datetime(2026, 7, 1, 8, 0, 0)
    source_file = batch.get("source_file") or _safe_json(batch.get("source_file_json")) or {}
    checkpoint = batch.get("checkpoint") or _safe_json(batch.get("checkpoint_json")) or {}
    return {
        "batch_id": batch.get("batch_id", "plad-20260701080000-abcdef12"),
        "automation_id": batch.get("automation_id", "packing-list-auto-download"),
        "module_id": batch.get("module_id", "packing-list-auto-download"),
        "run_id": batch.get("run_id", ""),
        "batch_name": batch.get("batch_name", ""),
        "source_file_name": batch.get("source_file_name", ""),
        "source_file_sha256": batch.get("source_file_sha256", ""),
        "source_file_json": json.dumps(source_file, ensure_ascii=False),
        "status": batch.get("status", "pending"),
        "message": batch.get("message", ""),
        "total_count": batch.get("total_count", 0),
        "completed_count": batch.get("completed_count", 0),
        "failed_count": batch.get("failed_count", 0),
        "pending_count": batch.get("pending_count", 0),
        "checkpoint_json": json.dumps(checkpoint, ensure_ascii=False),
        "bucket": batch.get("bucket", "tos-packing-list-auto-download"),
        "object_prefix": batch.get("object_prefix", "packing-list-auto-download/2026-07-01/plad-20260701080000-abcdef12"),
        "created_at": batch.get("created_at", now),
        "updated_at": batch.get("updated_at", now),
    }


def ticket_owner_batch_row(batch):
    now = datetime(2026, 7, 2, 10, 0, 0)
    source_file = batch.get("source_file") or _safe_json(batch.get("source_file_json")) or {}
    checkpoint = batch.get("checkpoint") or _safe_json(batch.get("checkpoint_json")) or {}
    return {
        "batch_id": batch.get("batch_id", "tos-20260702100000-abcdef12"),
        "automation_id": batch.get("automation_id", "ticket-owner-statistics"),
        "module_id": batch.get("module_id", "ticket-owner-statistics"),
        "run_id": batch.get("run_id", ""),
        "batch_name": batch.get("batch_name", ""),
        "source_file_name": batch.get("source_file_name", ""),
        "source_file_sha256": batch.get("source_file_sha256", ""),
        "source_file_json": json.dumps(source_file, ensure_ascii=False),
        "status": batch.get("status", "pending"),
        "message": batch.get("message", ""),
        "total_count": batch.get("total_count", 0),
        "completed_count": batch.get("completed_count", 0),
        "failed_count": batch.get("failed_count", 0),
        "pending_count": batch.get("pending_count", 0),
        "checkpoint_json": json.dumps(checkpoint, ensure_ascii=False),
        "bucket": batch.get("bucket", "tos-ticket-owner-statistics"),
        "object_prefix": batch.get("object_prefix", "ticket-owner-statistics/2026-07-02/tos-20260702100000-abcdef12"),
        "created_at": batch.get("created_at", now),
        "updated_at": batch.get("updated_at", now),
    }


def _safe_json(value):
    if isinstance(value, dict):
        return value
    if not value:
        return {}
    return json.loads(value)


def run_async(awaitable):
    import asyncio

    return asyncio.run(awaitable)


if __name__ == "__main__":
    unittest.main()
